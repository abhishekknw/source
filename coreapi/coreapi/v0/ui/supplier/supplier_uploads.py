from __future__ import print_function
from __future__ import absolute_import
from openpyxl import load_workbook, Workbook
from rest_framework.views import APIView
from v0.ui.utils import (get_supplier_id, handle_response, get_content_type, save_flyer_locations, make_supplier_data,
                         get_model, get_serializer, save_supplier_data, get_region_based_query, get_supplier_image,
                         save_basic_supplier_details)
from .models import (SupplierTypeSociety, SupplierAmenitiesMap, SupplierTypeCorporate, SupplierTypeGym,
                    SupplierTypeRetailShop, CorporateParkCompanyList, CorporateBuilding, SupplierTypeBusDepot,
                    SupplierTypeCode, SupplierTypeBusShelter, CorporateCompanyDetails, RETAIL_SHOP_TYPE, SupplierMaster, AddressMaster)
from .serializers import SupplierMasterSerializer
from v0.ui.account.models import ContactDetails
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from v0.ui.finances.models import PriceMappingDefault
from v0.ui.location.models import State, City, CityArea, CitySubArea
from v0.ui.inventory.models import AdInventoryType, InventorySummary
from v0.ui.base.models import DurationType
import v0.ui.utils as ui_utils
from v0.ui.utils import get_from_dict
from v0.ui.controller import inventory_summary_insert
from django.core.exceptions import ObjectDoesNotExist
from v0.ui.difflib import get_close_matches
import os

def get_state_map():
    all_city = City.objects.all()
    state_map = {}
    for city in all_city:
        state_map[city.city_code] = {'state_name': city.state_code.state_name, 'state_code': city.state_code.state_code}
    return state_map


def create_price_mapping_default(days_count, adinventory_name, adinventory_type, new_supplier,
                                 actual_supplier_price, content_type, supplier_id):
    duration_types = DurationType.objects.filter()
    adinventory_types = AdInventoryType.objects.filter(adinventory_name=adinventory_name)
    for adinv_type in adinventory_types:
        for dur_type in duration_types:
            if dur_type.days_count == days_count and adinv_type.adinventory_name == adinventory_name and adinv_type.adinventory_type == adinventory_type:
                obj,created = PriceMappingDefault.objects.get_or_create(supplier=new_supplier.supplier_id, duration_type=dur_type,
                                              adinventory_type=adinv_type,
                                              content_type=content_type,
                                              object_id=supplier_id)

                obj.actual_supplier_price = actual_supplier_price
                obj.save()
            else:
                if (adinv_type.adinventory_name == 'FLIER' and dur_type.duration_name == 'Unit Daily') or adinv_type.adinventory_name != 'FLIER':
                    PriceMappingDefault.objects.get_or_create(supplier=new_supplier, duration_type=dur_type,
                                                          adinventory_type=adinv_type,
                                                          content_type=content_type,
                                                          object_id=supplier_id)

def get_duplicate_supplier(index, item):
    supplier_names = []
    supplier_dict = {}

    if item[1].value != "RS":
        suppliers = SupplierMaster.objects.filter(
            supplier_type=item[1].value,
            area= item[4].value,
            city= item[6].value,
            state= item[7].value,
            subarea= item[5].value,
        )
        supplier_names = [supplier.supplier_name for supplier in suppliers]
        supplier_dict = {supplier.supplier_name:supplier.supplier_id for supplier in suppliers}
    else:
        suppliers = SupplierTypeSociety.objects.filter(
            society_locality= item[4].value,
            society_city= item[6].value,
            society_state= item[7].value,
            society_subarea= item[5].value,
        )
        supplier_names = [supplier.society_name for supplier in suppliers]
        supplier_dict = {supplier.society_name:supplier.supplier_id for supplier in suppliers}

    matches = get_close_matches(item[0].value, supplier_names, n=1, cutoff=0.7)

    if matches:
        return {
            "row_no": index+1,
            "supplier_name": item[0].value,
            "supplier_type": item[1].value,
            "supplier_id": supplier_dict.get(matches[0]),
            "matched_with": matches[0]
        }

def add_contact_data(supplier_id, row):
    contact_number = row[17].value
    contact_name = row[18].value
    contact_type = row[19].value
    
    if contact_number and supplier_id:
        contact_details = ContactDetails.objects.filter(mobile=contact_number, object_id=supplier_id)
        if contact_details:
            if contact_name and contact_type:
                contact_details.update(name=contact_name, contact_type=contact_type)                
        else:
            ContactDetails(**{
                'object_id': supplier_id,
                'mobile' : contact_number, 
                'name' : contact_name, 
                'contact_type' : contact_type,
            }).save()

@method_decorator(csrf_exempt, name='dispatch')
class CorporateParkDataImport(APIView):
    """

    """

    def post(self, request):
        duplicates_list = []
        general_error = set()

        book = Workbook()
        sheet = book.active
        
        cities = City.objects.all()
        city_dict = {row.city_name:row.id for row in cities}

        areas = CityArea.objects.all()
        area_dict = {row.label:row.id for row in areas}

        subareas = CitySubArea.objects.all()
        subarea_dict = {row.subarea_name:row.id for row in subareas}

        source_file = request.data['file']
        wb = load_workbook(source_file)
        ws = wb.get_sheet_by_name(wb.get_sheet_names()[0])
        cp_data_list = []
        for index, row in enumerate(ws.iter_rows()):
            if index == 0:
                header_list = ['Supplier Id']
                for cell in row:
                    header_list.append(cell.value)
                sheet.append(header_list)

            elif index > 0 and row[0].value:

                if not city_dict.get(row[6].value) or not area_dict.get(row[4].value) or not subarea_dict.get(row[5].value):
                    error = ""
                    if not city_dict.get(row[6].value):
                        error = row[6].value+" doesn't exist."
                    if not area_dict.get(row[4].value):
                        error = row[4].value+" doesn't exist."
                    if not subarea_dict.get(row[5].value):
                        error = row[5].value+" doesn't exist."
                    
                    general_error.add(error)
                    continue
                
                supplier_data1 = {
                    'city_id': city_dict.get(row[6].value),
                    'area_id': area_dict.get(row[4].value),
                    'subarea_id': subarea_dict.get(row[5].value),
                    'supplier_type': row[1].value,
                    'supplier_code': row[1].value,
                    'supplier_name': row[0].value,    
                }
                supplier_id = get_supplier_id(supplier_data1)

                if row[20].value != 'D':
                    duplicate = get_duplicate_supplier(index, row)
                    if duplicate:
                        duplicates_list.append(duplicate)

                        cell_list = [duplicate["supplier_id"]]
                        add_contact_data(duplicate["supplier_id"], row)
                        for cell in row:
                            cell_list.append(cell.value)
                        sheet.append(cell_list)

                        continue
                    
                cell_list = [supplier_id]
                for cell in row:
                    cell_list.append(cell.value)
                sheet.append(cell_list)

                supplier_data = {
                    'supplier_id': supplier_id,
                    'supplier_name': row[0].value,
                    'supplier_type': row[1].value,
                    'unit_primary_count': row[2].value,
                    'representative': request.user.profile.organisation.organisation_id,
                    'unit_secondary_count': row[3].value,
                    'area': row[4].value,
                    'city': row[6].value,
                    'state': row[7].value,
                    'subarea': row[5].value,
                    'zipcode': row[8].value,
                    'address1': row[9].value,
                    'landmark': row[10].value,
                    'latitude': row[11].value,
                    'longitude': row[12].value,
                    'feedback': row[13].value,
                    'quality_rating': row[14].value,
                    'locality_rating': row[15].value,
                    'quantity_rating': row[16].value,
                }

                add_contact_data(supplier_id, row)

                if row[1].value != "RS":
                    serializer = SupplierMasterSerializer(data=supplier_data)
                    if serializer.is_valid():
                        serializer.save()

                    AddressMaster(**{
                        'supplier_id': supplier_id,
                        'area': row[4].value,
                        'city': row[6].value,
                        'state': row[7].value,
                        'subarea': row[5].value,
                        'zipcode': row[8].value,
                        'address1': row[9].value,
                        'nearest_landmark': row[10].value,
                        'latitude': row[11].value,
                        'longitude': row[12].value,
                    }).save()

                ui_utils.create_supplier_from_master(supplier_data, row[1].value)

        filepath = 'files/export_supplier.xlsx'
        book.save(filepath)

        return handle_response({}, data={"invalid_row_detail": duplicates_list, "general_error": list(general_error) }, success=True)
