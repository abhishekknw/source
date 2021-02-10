from __future__ import print_function
from __future__ import absolute_import
import json
import requests
import logging

logger = logging.getLogger(__name__)

from bson.objectid import ObjectId
from django.urls import reverse
from django.forms.models import model_to_dict
from django.http import HttpResponse
from django.db.models import Prefetch

from openpyxl import load_workbook, Workbook
import csv
from rest_framework.views import APIView
from rest_framework import viewsets
from rest_framework.decorators import list_route
from v0.ui.utils import (get_supplier_id, handle_response, get_content_type, save_flyer_locations, make_supplier_data,
                         get_model, get_serializer, save_supplier_data, get_region_based_query, get_supplier_image,
                         save_basic_supplier_details, get_user_organisation_id)
from v0.ui.website.utils import manipulate_object_key_values, return_price, manipulate_master_to_rs
import v0.ui.website.utils as website_utils

from v0.ui.dynamic_booking.models import BookingInventoryActivity
from v0.ui.dynamic_suppliers.models import SupplySupplier
from .models import (SupplierTypeSociety, SupplierAmenitiesMap, SupplierTypeCorporate, SupplierTypeGym, SupplierMaster,
                    SupplierTypeRetailShop, CorporateParkCompanyList, CorporateBuilding, SupplierTypeBusDepot,
                    SupplierTypeCode, SupplierTypeBusShelter, CorporateCompanyDetails, RETAIL_SHOP_TYPE, 
                    SupplierEducationalInstitute, SupplierHording, SupplierMaster, AddressMaster, SupplierRelationship)
from .serializers import (UICorporateSerializer, SupplierTypeSocietySerializer, SupplierAmenitiesMapSerializer,
                         UISalonSerializer, SupplierTypeSalon, SupplierTypeGymSerializer, RetailShopSerializer,
                         SupplierInfoSerializer, SupplierInfo, SupplierTypeCorporateSerializer,
                         CorporateBuildingGetSerializer, CorporateParkCompanyListSerializer, CorporateBuildingSerializer,
                         SupplierTypeSalonSerializer, BusDepotSerializer, CorporateParkCompanySerializer,
                         BusShelterSerializer, SupplierTypeBusShelterSerializer, SupplierEducationalInstituteSerializer, SupplierHordingSerializer, SupplierMasterSerializer, AddressMasterSerializer)
from v0.ui.inventory.serializers import ShortlistedSpacesSerializer
from v0.ui.location.serializers import CityAreaSerializer
from v0.ui.account.models import (ContactDetails, OwnershipDetails)
from v0.ui.inventory.models import AdInventoryType, PosterInventory, StandeeInventory
from v0.ui.finances.models import PriceMappingDefault, ShortlistedInventoryPricingDetails
from v0.ui.base.models import DurationType
from v0.ui.serializers import (UISocietySerializer, SocietyListSerializer)
from v0.ui.proposal.serializers import ImageMappingSerializer
from v0.ui.proposal.models import ImageMapping, ShortlistedSpaces
from v0.ui.account.serializers import (ContactDetailsSerializer, ContactDetailsGenericSerializer, OwnershipDetailsSerializer)
from v0.ui.account.models import ContactDetailsGeneric
from v0.ui.components.models import (SocietyTower, CorporateBuildingWing, CompanyFloor, FlatType, LiftDetails)
from v0.ui.components.serializers import CorporateBuildingWingSerializer
from v0.ui.proposal.models import (ProposalInfo, ProposalCenterMapping, SupplierAssignment)
from v0.ui.organisation.models import (Organisation)
import v0.constants as v0_constants
from rest_framework.response import Response
from django.contrib.contenttypes.models import ContentType
from rest_framework.pagination import PageNumberPagination
from django.db.models import Q, Sum
from rest_framework import status
from django.core.exceptions import ObjectDoesNotExist, MultipleObjectsReturned
from v0.ui.utils import get_from_dict
from v0.ui.controller import inventory_summary_insert
import v0.ui.utils as ui_utils
from v0.ui.inventory.models import InventorySummary, InventoryActivityImage
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from v0.ui.utils import get_supplier_id
import openpyxl
from v0.ui.location.models import State, City, CityArea, CitySubArea
from coreapi.settings import BASE_DIR, BASE_URL
from v0.ui.proposal.models import ProposalInfo
from v0.constants import flat_type_dict
from v0 import errors
from django.db import transaction
from v0.utils import create_cache_key, get_values
from django.conf import settings
from django.apps import apps
from django.db import IntegrityError
from .supplier_uploads import create_price_mapping_default
from django.db import connection
from v0.ui.common.pagination import paginate
from django.utils import timezone
from django.views.generic import ListView

import logging
logger = logging.getLogger(__name__)

def get_values(list_name,key):
    values = []
    for item in list_name:
        values.append(item[key])
    return values

def get_state_map():
    all_city = City.objects.all()
    state_map = {}
    for city in all_city:
        state_map[city.city_code] = {'state_name': city.state_code.state_name, 'state_code': city.state_code.state_code}
    return state_map


def get_city_map():
    all_city = City.objects.all()
    city_map = {}
    for city in all_city:
        city_map[city.city_code] = city.city_name
    return city_map


def get_city_area_map():
    all_city_area = CityArea.objects.all()
    city_area_map = {}
    for city_area in all_city_area:
        city_area_map[city_area.area_code] = city_area.label
    return city_area_map


def get_city_subarea_map():
    all_city_subarea = CitySubArea.objects.all()
    city_subarea_map = {}
    for city_subarea in all_city_subarea:
        city_subarea_map[city_subarea.subarea_code] = city_subarea.subarea_name
    return city_subarea_map


def get_flat_count_type(flat_count):
    if flat_count is None:
        return None
    if flat_count<150:
        flat_type = '1-150'
    elif flat_count<400:
        flat_type = '151-400'
    else:
        flat_type = '401+'
    return flat_type


def update_contact_and_ownership_detail(data):
    basic_contacts = data.get('contactData', None)
    object_id = data.get('supplier_id', None)
    master_supplier_data = data.get('master_data', None)
    supplier_address_data = master_supplier_data.get('address_supplier')

    if master_supplier_data:
        landmark = master_supplier_data.get('landmark', None)
        if not landmark:
            landmark = master_supplier_data.get('nearest_landmark', None)

        master_data = {
            "supplier_id": object_id,
            "supplier_name": master_supplier_data.get('supplier_name', None),
            "supplier_type": master_supplier_data.get('supplier_type', None),
            "unit_primary_count": master_supplier_data.get('unit_primary_count') if master_supplier_data.get('unit_primary_count') else 0,
            "unit_secondary_count": master_supplier_data.get('unit_secondary_count') if master_supplier_data.get('unit_secondary_count') else 0,
            "unit_tertiary_count": master_supplier_data.get('unit_tertiary_count') if master_supplier_data.get('unit_tertiary_count') else 0,
            "representative": data.get('representative', None),
            "area": supplier_address_data.get('area', None),
            "subarea": supplier_address_data.get('subarea', None),
            "city": supplier_address_data.get('city', None),
            "state": supplier_address_data.get('state', None),
            "zipcode": supplier_address_data.get('zipcode', None),
            "latitude": supplier_address_data.get('latitude', None),
            "longitude": supplier_address_data.get('longitude', None),
            "landmark": landmark,
            "feedback": data.get('feedback', None),
        }
        supplier_type = master_supplier_data.get('supplier_type', None)
        if supplier_type in ["CP", "BS", "HO", "CO", "HL", "RC", "TV", "GN", "BU"]:
            master_data["quality_rating"] = data.get('quality_rating', None)
            master_data["locality_rating"] = data.get('locality_rating', None)
            master_data["quantity_rating"] = data.get('quantity_rating', None)
        elif supplier_type in ["GY", "SA"]:
            master_data["quality_rating"] = data.get('category', None)
            master_data["locality_rating"] = data.get('locality_rating', None)
        elif supplier_type in ["RE"]:
            master_data["quality_rating"] = data.get('rating', None)
            master_data["locality_rating"] = data.get('rating', None)
            master_data["quantity_rating"] = data.get('store_size', None)

        supplier_master_data = SupplierMaster.objects.filter(supplier_id=object_id).first()
        if supplier_master_data and supplier_master_data.supplier_id:
            supplier_master_serializer = SupplierMasterSerializer(supplier_master_data, data=master_data)
        else:
            supplier_master_serializer = SupplierMasterSerializer(data=master_data)
        
        if supplier_master_serializer.is_valid():
            supplier_master_serializer.save()

        address_data = {
            "supplier": object_id,
            "address1": supplier_address_data.get('address1', None),
            "address2": supplier_address_data.get('address2', None),
            "area": supplier_address_data.get('area', None),
            "subarea": supplier_address_data.get('subarea', None),
            "city": supplier_address_data.get('city', None),
            "state": supplier_address_data.get('state', None),
            "zipcode": supplier_address_data.get('zipcode', None),
            "latitude": supplier_address_data.get('latitude', None),
            "longitude": supplier_address_data.get('longitude', None),
            "nearest_landmark": landmark,
        }
        address_master_data = AddressMaster.objects.filter(supplier_id=object_id).first()
        if address_master_data and address_master_data.supplier_id:
            address_master_serializer = AddressMasterSerializer(address_master_data, data=address_data)
        else:
            address_master_serializer = AddressMasterSerializer(data=address_data)

        if address_master_serializer.is_valid():
            address_master_serializer.save()

    not_remove_contacts = []

    if basic_contacts:
        for contact in basic_contacts:
            contact['object_id'] = object_id
            if 'id' in contact:
                contact_detail = ContactDetails.objects.filter(pk=contact['id']).first()
                contact_serializer = ContactDetailsSerializer(contact_detail, data=contact)
            else:
                contact_serializer = ContactDetailsSerializer(data=contact)
            if contact_serializer.is_valid():
                contact_serializer.save(object_id=object_id)
                not_remove_contacts.append(contact_serializer.data['id'])

    ContactDetails.objects.filter(object_id=object_id).exclude(pk__in=not_remove_contacts).delete()

    ownership = data.get('ownership_details', None)

    if ownership:
        if ownership.get('id'):
            ownership_detail = OwnershipDetails.objects.filter(pk=ownership['id']).first()
            ownership_serializer = OwnershipDetailsSerializer(ownership_detail, data=ownership)
        else:
            ownership_serializer = OwnershipDetailsSerializer(data=ownership)
        
        if ownership_serializer.is_valid():
            ownership_serializer.save()
    return True


def retrieve_contact_and_ownership_detail(pk):

    retail_shop_instance = ContactDetails.objects.filter(object_id=pk)
    contact_serializer = ContactDetailsSerializer(retail_shop_instance, many=True)
        
    contact_data = contact_serializer.data

    ownership_details_instance = OwnershipDetails.objects.filter(object_id=pk).first()
    ownership_details_serializer = OwnershipDetailsSerializer(ownership_details_instance, many=False)
    
    ownership_data = ownership_details_serializer.data
    
    if ownership_data and ownership_data.get('object_id'):
        ownership_details = ownership_data
    else:
        ownership_details = {}
    return contact_data, ownership_details


@method_decorator(csrf_exempt, name='dispatch')
class SocietyDataImport(APIView):
    """

    """

    def post(self, request):
        source_file = request.data['file']
        wb = load_workbook(source_file)
        ws = wb.get_sheet_by_name(wb.get_sheet_names()[0])
        society_data_list = []

        for index, row in enumerate(ws.iter_rows()):
            if index > 0:
                print(index)
                flat_count = int(row[18].value) if row[18].value else None
                vendor_name = row[0].value if row[0].value else None
                representative_id = None
                if vendor_name:
                    ventor_organisation_all = Organisation.objects.filter(name=vendor_name).all()
                    if len(ventor_organisation_all):
                        representative_id = ventor_organisation_all[0].organisation_id
                flat_count_type = get_flat_count_type(flat_count)
                society_data_list.append({
                    'representative_id': representative_id,
                    'society_name': row[1].value if row[1].value else None,
                    'society_city': str(row[2].value) if row[2].value else None,
                    'society_city_code': str(row[3].value) if row[3].value else None,
                    'society_locality': row[4].value if row[4].value else None,
                    'society_locality_code': row[5].value if row[5].value else None,
                    'society_subarea': row[6].value if row[6].value else None,
                    'society_subarea_code': row[7].value if row[7].value else None,
                    'society_code': row[8].value if row[8].value else None,
                    'supplier_code': row[9].value if row[9].value else None,
                    'supplier_id': row[10].value if row[10].value else None,
                    'society_zip': int(row[11].value) if row[11].value else None,
                    'society_address1' : row[12].value if row[12].value else None,
                    'landmark' : row[13].value if row[13].value else None,
                    'society_type_quality' : row[14].value if row[14].value else None,
                    'society_latitude': float(row[15].value) if row[15].value else None,
                    'society_longitude': float(row[16].value) if row[16].value else None,
                    'tower_count': int(row[17].value) if row[17].value else None,
                    'flat_count': int(row[18].value) if row[18].value else None,
                    'flat_count_type': flat_count_type,
                    'vacant_flat_count' : int(row[19].value) if row[19].value else None,
                    'bachelor_tenants_allowed': row[20].value if row[20].value in ['Y', 'y', 't', 'T', 'true','True'] else False,
                    'designation': row[21].value if row[21].value else None,
                    'salutation': row[22].value if row[22].value else None,
                    'contact_name': row[23].value if row[23].value else None,
                    'email': row[24].value if row[24].value else None,
                    'mobile': row[25].value if row[25].value else None,
                    'landline': row[26].value if row[26].value else None,
                    'name_for_payment': row[27].value if row[27].value else None,
                    'ifsc_code': row[28].value if row[28].value else None,
                    'bank_name': row[29].value if row[29].value else None,
                    'account_no': row[30].value if row[30].value else None,
                    'relationship_manager' : row[31].value if row[31].value else None,
                    'age_of_society' : row[32].value if row[32].value else None,
                    'stall_allowed': True if row[33].value in ['Y', 'y', 't', 'T', 'true', 'True'] else False,
                    'total_stall_count': row[34].value if row[34].value else None,
                    'poster_allowed_nb': True if row[35].value in ['Y', 'y', 't', 'T', 'true', 'True'] else False,
                    'nb_per_tower': int(row[36].value) if row[36].value else None,
                    'poster_allowed_lift': True if row[37].value in ['Y', 'y', 't', 'T', 'true', 'True'] else False,
                    'lift_per_tower': int(row[38].value) if row[38].value else None,
                    'flier_allowed': True if row[39].value in ['Y', 'y', 't', 'T', 'true', 'True'] else False,
                    'flier_frequency': int(row[40].value) if row[40].value else None,
                    'stall_price': float(row[41].value) if row[41].value else None,
                    'poster_price': float(row[42].value) if row[42].value else None,
                    'flier_price': float(row[43].value) if row[43].value else None,
                    'status': row[44].value,
                    'comments': row[45].value,
                })
        all_states_map = get_state_map()
        all_city_map = get_city_map()
        all_city_area_map = get_city_area_map()
        all_city_subarea_map = get_city_subarea_map()
        for society in society_data_list:
            if society['supplier_code'] is not None:
                data = {
                    'city_code': society['society_city_code'],
                    'area_code': society['society_locality_code'],
                    'subarea_code': society['society_subarea_code'],
                    'supplier_type': 'RS',
                    'supplier_code': society['supplier_code'],
                    'supplier_name': society['society_name']
                }

                supplier_id = None
                if society['supplier_id']:
                    supplier_id = society['supplier_id']
                else:
                    supplier_id = get_supplier_id(data)

                supplier_length = len(SupplierTypeSociety.objects.filter(supplier_id=supplier_id))
                if len(SupplierTypeSociety.objects.filter(society_name=society['society_name'])):

                    # instance = SupplierTypeSociety.objects.get(supplier_id=supplier_id)
                    instance = SupplierTypeSociety.objects.filter(society_name=society['society_name'])[0]
                    supplier_id = instance.supplier_id
                    instance.society_name = society['society_name']
                    instance.representative_id = society['representative_id']
                    instance.society_locality = society['society_locality']
                    instance.society_city = society['society_city']
                    instance.society_state = all_states_map[society['society_city_code']]['state_name']
                    instance.society_subarea = society['society_subarea']
                    instance.supplier_code = society['supplier_code']
                    instance.society_zip = society['society_zip']
                    instance.society_address1 = society['society_address1']
                    instance.landmark = society['landmark']
                    instance.society_type_quality = society['society_type_quality']
                    instance.society_latitude = society['society_latitude']
                    instance.society_longitude = society['society_longitude']
                    instance.tower_count = society['tower_count']
                    instance.flat_count = society['flat_count']
                    instance.vacant_flat_count = society['vacant_flat_count']
                    instance.bachelor_tenants_allowed = society['bachelor_tenants_allowed']
                    instance.name_for_payment = society['name_for_payment']
                    instance.ifsc_code = society['ifsc_code']
                    instance.bank_name = society['bank_name']
                    instance.account_no = society['account_no']
                    instance.relationship_manager = society['relationship_manager']
                    instance.age_of_society = society['age_of_society']
                    instance.stall_allowed = society['stall_allowed']
                    instance.supplier_status = society['status']
                    instance.comments = society['comments']
                    instance.save()
                    new_society = instance

                elif supplier_length:

                    instance = SupplierTypeSociety.objects.get(supplier_id=supplier_id)
                    instance.society_name = society['society_name']
                    instance.representative_id = society['representative_id']
                    instance.society_locality = society['society_locality']
                    instance.society_city = society['society_city']
                    instance.society_state = all_states_map[society['society_city_code']]['state_name']
                    instance.society_subarea = society['society_subarea']
                    instance.supplier_code = society['supplier_code']
                    instance.society_zip = society['society_zip']
                    instance.society_address1 = society['society_address1']
                    instance.landmark = society['landmark']
                    instance.society_type_quality = society['society_type_quality']
                    instance.society_latitude = society['society_latitude']
                    instance.society_longitude = society['society_longitude']
                    instance.tower_count = society['tower_count']
                    instance.flat_count = society['flat_count']
                    instance.vacant_flat_count = society['vacant_flat_count']
                    instance.bachelor_tenants_allowed = society['bachelor_tenants_allowed']
                    instance.name_for_payment = society['name_for_payment']
                    instance.ifsc_code = society['ifsc_code']
                    instance.bank_name = society['bank_name']
                    instance.account_no = society['account_no']
                    instance.relationship_manager = society['relationship_manager']
                    instance.age_of_society = society['age_of_society']
                    instance.stall_allowed = society['stall_allowed']
                    instance.supplier_status = society['status']
                    instance.comments = society['comments']
                    instance.save()
                    new_society = instance

                else:

                    new_society = SupplierTypeSociety(**{
                        'supplier_id': supplier_id,
                        'society_name': society['society_name'],
                        'representative_id': society['representative_id'],
                        'society_locality': society['society_locality'],
                        'society_city': society['society_city'],
                        'society_state': all_states_map[society['society_city_code']]['state_name'],
                        'society_subarea': society['society_subarea'],
                        'supplier_code': society['supplier_code'],
                        'society_zip': society['society_zip'],
                        'society_address1': society['society_address1'],
                        'landmark': society['landmark'],
                        'society_type_quality': society['society_type_quality'],
                        'society_latitude': society['society_latitude'],
                        'society_longitude': society['society_longitude'],
                        'tower_count': society['tower_count'],
                        'flat_count': society['flat_count'],
                        'vacant_flat_count': society['vacant_flat_count'],
                        'bachelor_tenants_allowed': society['bachelor_tenants_allowed'],
                        'name_for_payment': society['name_for_payment'],
                        'ifsc_code': society['ifsc_code'],
                        'bank_name': society['bank_name'],
                        'account_no': society['account_no'],
                        'relationship_manager': society['relationship_manager'],
                        'age_of_society': society['age_of_society'],
                        'stall_allowed': society['stall_allowed'],
                        'supplier_status': society['status'],
                        'comments': society['comments'],
                    })
                    new_society.save()

                new_contact_data = {
                    'name': society['contact_name'],
                    'email': society['email'],
                    'designation': society['designation'],
                    'salutation': society['salutation'],
                    'mobile': society['mobile'],
                    'landline': society['landline'],
                    'content_type': get_content_type('RS').data['data'],
                    'object_id': supplier_id
                }
                contacts = ContactDetails.objects.filter(mobile=society['mobile'])
                if not len(contacts):
                    ContactDetails.objects.create(**new_contact_data)

                rs_content_type = get_content_type('RS').data['data']
                print(society['society_name'])
                try:
                    create_price_mapping_default('7', "POSTER", "A4", new_society,
                                             society['poster_price'], rs_content_type, supplier_id)
                except Exception as e:
                    pass
                try:
                    create_price_mapping_default('0', "POSTER LIFT", "A4", new_society,
                                             0, rs_content_type, supplier_id)
                except Exception as e:
                    pass
                try:
                    create_price_mapping_default('0', "STANDEE", "Small", new_society,
                                             0, rs_content_type, supplier_id)
                except Exception as e:
                    pass
                try:
                    create_price_mapping_default('1', "STALL", "Small", new_society,
                                             society['stall_price'], rs_content_type, supplier_id)
                except Exception as e:
                    pass
                try:
                    create_price_mapping_default('0', "CAR DISPLAY", "A4", new_society,
                                             0, rs_content_type, supplier_id)
                except Exception as e:
                    pass
                save_flyer_locations(0, 1, new_society, society['supplier_code'])
                try:
                    create_price_mapping_default('1', "FLIER", "Door-to-Door", new_society,
                                             society['flier_price'], rs_content_type, supplier_id)
                except Exception as e:
                    pass

                inventory_obj = InventorySummary.objects.filter(object_id=supplier_id).first()
                inventory_id = inventory_obj.id if inventory_obj else None
                request_data = {
                    'id': inventory_id,
                    'd2d_allowed': True,
                    'poster_allowed_nb': True,
                    'supplier_type_code': 'RS',
                    'lift_count': society['tower_count'] * society['lift_per_tower'] if society['tower_count'] and society['lift_per_tower'] else None,
                    'stall_allowed': True,
                    'object_id': supplier_id,
                    'flier_allowed': True,
                    'nb_count': society['tower_count'] * society['nb_per_tower'] if society['tower_count'] and society['nb_per_tower'] else None,
                    'user': 1,
                    'content_type': 46,
                    'flier_frequency': society['flier_frequency'] if society['flier_frequency'] else None,
                    'total_stall_count': society['total_stall_count'] if society['total_stall_count'] else None,
                    'poster_allowed_lift': True,
                }
                class_name = self.__class__.__name__
                response = ui_utils.get_supplier_inventory(request_data.copy(), supplier_id)
                supplier_inventory_data = response.data['data']['request_data']

                if not response.data['status']:
                    return response

                supplier_inventory_data = response.data['data']['request_data']
                final_data = {
                    'id': get_from_dict(request_data, supplier_id),
                    'supplier_object': get_from_dict(response.data['data'], 'supplier_object'),
                    'inventory_object': get_from_dict(response.data['data'], 'inventory_object'),
                    'supplier_type_code': get_from_dict(request_data, 'supplier_type_code'),
                    'poster_allowed_nb': get_from_dict(request_data, 'poster_allowed_nb'),
                    'nb_count': get_from_dict(request_data, 'nb_count'),
                    'poster_campaign': get_from_dict(request_data, 'poster_campaign'),
                    'lift_count': get_from_dict(request_data, 'lift_count'),
                    'poster_allowed_lift': get_from_dict(request_data, 'poster_allowed_lift'),
                    'standee_allowed': get_from_dict(request_data, 'standee_allowed'),
                    'total_standee_count': get_from_dict(request_data, 'total_standee_count'),
                    'stall_allowed': get_from_dict(request_data, 'stall_allowed'),
                    'total_stall_count': get_from_dict(request_data, 'total_stall_count'),
                    'flier_allowed': get_from_dict(request_data, 'flier_allowed'),
                    'flier_frequency': get_from_dict(request_data, 'flier_frequency'),
                    'flier_campaign': get_from_dict(request_data, 'flier_campaign'),
                }
                try:
                    inventory_summary_insert(final_data, supplier_inventory_data)
                except ObjectDoesNotExist as e:
                    print(e)
                    # return ui_utils.handle_response(class_name, exception_object=e, request=request)
                except Exception as e:
                    print(e)
                    # return Response(data={"status": False, "error": str(e.message)}, status=status.HTTP_400_BAD_REQUEST)
        return handle_response({}, data='success', success=True)


class TransactionDataImport(APIView):
    def post(self, request):
        source_file = request.data['file']
        wb = load_workbook(source_file)
        ws = wb.get_sheet_by_name(wb.get_sheet_names()[0])
        transaction_data_list = []
        for index, row in enumerate(ws.iter_rows()):
            if index > 0:
                transaction_data_list.append({
                    'supplier_id': int(row[0].value) if row[0].value else None,
                    'poster_allowed': str(row[1].value) if row[1].value else None,
                    'poster_count': int(row[2].value) if row[2].value else None,
                    'poster_price': float(row[3].value) if row[3].value else None,
                    'standee_allowed': row[4].value if row[4].value else None,
                    'standee_count': int(row[5].value) if row[5].value else None,
                    'standee_price': float(row[6].value) if row[6].value else None,
                    'flier_allowed': row[7].value if row[7].value else None,
                    'flier_frequency': int(row[8].value) if row[8].value else None,
                    'flier_price': float(row[9].value) if row[9].value else None,
                    'stall_allowed': row[10].value if row[10].value else None,
                    'stall_duration': int(row[11].value) if row[11].value else None,
                    'arch_allowed': row[12].value if row[12].value else None,
                    'arch_price': row[13].value if row[13].value else None,
                    # 'comment': row[27].value,
                })
        return handle_response({}, data='transaction-success', success=True)


class checkSupplierCodeAPIView(APIView):
    def get(self, request, code, format=None):
        try:
            society = SupplierTypeSociety.objects.get(supplier_code=code)
            if society:
                return Response(status=200)
        except SupplierTypeSociety.DoesNotExist:
            return Response(status=404)


class GenerateSupplierIdAPIView(APIView):
    """
    Generic API that generates unique supplier id and also saves supplier data
    """

    def post(self, request):

        class_name = self.__class__.__name__
        try:
            user = request.user
            data = dict({
                'city_id': request.data['city_id'],
                'area_id': request.data['area_id'],
                'subarea_id': request.data['subarea_id'],
                'supplier_type': request.data['supplier_type'],
                'supplier_name': request.data['supplier_name'],
            })
            supplier_name = request.data['supplier_name']
            if len(supplier_name) == 0:
                return handle_response(class_name, data='Supplier Name is required', success=False)
            data['supplier_code'] = supplier_name.upper()

            if len(supplier_name) > 2:
                data['supplier_code'] = supplier_name[:3].upper()

            data['supplier_id'] = get_supplier_id(data)
            data['supplier_type_code'] = request.data['supplier_type']
            data['current_user'] = request.user
            response = make_supplier_data(data)
            all_supplier_data = response.data['data']
            return handle_response(class_name, data=save_supplier_data(user, all_supplier_data), success=True)
        except ObjectDoesNotExist as e:
            return handle_response(class_name, data='Supplier {supplier_code} does not exists'.format(supplier_code=request.data['supplier_type']))
        except Exception as e:
            return handle_response(class_name, exception_object=e)


class SupplierImageDetails(APIView):
    """
    This API gives supplier data and all images for supllier which are used to display on supplier
    Details page.
    """

    def get(self, request, id, format=None):

        class_name = self.__class__.__name__

        try:
            result = {}

            supplier_type_code = request.query_params.get('supplierTypeCode')
            if not supplier_type_code:
                return handle_response(class_name, data='No Supplier type code provided')

            content_type_response = get_content_type(supplier_type_code)
            if not content_type_response.data['status']:
                return handle_response(class_name, data='No content type found')

            content_type = content_type_response.data['data']
            supplier_object = get_model(supplier_type_code).objects.get(pk=id)

            serializer_class = get_serializer(supplier_type_code)
            serializer = serializer_class(supplier_object)
            result['supplier_data'] = serializer.data

            images = ImageMapping.objects.filter(object_id=id, content_type=content_type)
            image_serializer = ImageMappingSerializer(images, many=True)
            result['supplier_images'] = image_serializer.data

            amenities = SupplierAmenitiesMap.objects.filter(object_id=id, content_type=content_type)
            amenity_serializer = SupplierAmenitiesMapSerializer(amenities, many=True)
            result['amenities'] = amenity_serializer.data

            return handle_response(class_name, data=result, success=True)

        except Exception as e:
            return handle_response(class_name, exception_object=e, request=request)


class SocietyAPIView(APIView):
    # permission_classes = (permissions.IsAuthenticated, IsOwnerOrManager,)

    def get(self, request, id, format=None):
        try:
            response = {}
            item = SupplierTypeSociety.objects.get(pk=id)
            self.check_object_permissions(self.request, item)
            serializer = UISocietySerializer(item)

            # Start : Code changes to display images
            # todo : also write code for content_type
            # below code commented due to generic API SupplierImageDetails created to get all images
            # images = ImageMapping.objects.filter(object_id=id)
            # image_serializer = ImageMappingSerializer(images, many=True)
            # response['society_images'] = image_serializer.data
            # End : Code changes to display images
            # inventory_summary = InventorySummary.objects.get(supplier=item)
            # inventory_serializer = InventorySummarySerializer(inventory_summary)

            ###### Check if to use filter or get

            # poster = PosterInventory.objects.filter(supplier=item)
            # print "poster.adinventory_id is :",poster.adinventory_id
            # poster_serializer = PosterInventorySerializer(poster)

            # stall = StallInventory.objects.filter(supplier=item)
            # stall_serializer = StallInventorySerializer(stall)

            # doorToDoor = DoorToDoorInfo.objects.filter(**** how to filter on tower foreignkey  ***** )
            # doorToDoor_serializer = DoorToDoorInfoserializer(doorToDoor)

            # mailbox = MailboxInfo.objects.filter(**** how to filter on tower foreignkey  *****)
            # mailbox_serializer = MailboxInfoSerializer(mailbox)

            # response['society'] = serializer.data
            # response['inventory'] = inventory_serializer.data
            # response['poster'] = poster_serializer.data
            # response['doorToDoor'] = doorToDoor_serializer.data
            # response['mailbox'] = mailbox_serializer.data

            # return Response(response, status=200)

            response['society_data'] = serializer.data
            return Response(response, status=200)
        except SupplierTypeSociety.DoesNotExist:
            return Response(status=404)

    def delete(self, request, id, format=None):
        try:
            item = SupplierTypeSociety.objects.get(pk=id)
        except SupplierTypeSociety.DoesNotExist:
            return Response(status=404)
        contacts = item.get_contact_list()
        for contact in contacts:
            contact.delete()
        item.delete()
        return Response(status=204)

    def post(self, request, format=None):
        supplier_id = request.data.get('supplier_id', None)
        tower_count = request.data.get('tower_count', None)
        basic_contact_available = request.data.get('basic_contact_available', None)
        basic_reference_available = request.data.get('basic_reference_available', None)
        basic_reference_contacts = request.data.get('basic_reference_contacts', None)
        basic_contacts = request.data.get('basic_contacts', None)
        if supplier_id:
            society = SupplierTypeSociety.objects.filter(pk=supplier_id).first()
            if society:
                serializer = SupplierTypeSocietySerializer(society, data=request.data)
            else:
                serializer = SupplierTypeSocietySerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
        else:
            return Response(serializer.errors, status=400)
        society = SupplierTypeSociety.objects.filter(pk=serializer.data['supplier_id']).first()
        object_id = serializer.data.get('supplier_id')
        content_type = ContentType.objects.get_for_model(SupplierTypeSociety)

        not_remove_contacts = []
        # here we will start storing contacts
        if basic_contact_available:
            for contact in basic_contacts:
                if 'id' in contact:
                    item = ContactDetails.objects.filter(pk=contact['id']).first()
                    contact_serializer = ContactDetailsSerializer(item, data=contact)
                else:
                    contact_serializer = ContactDetailsSerializer(data=contact)
                if contact_serializer.is_valid():
                    contact_serializer.save(object_id=object_id, content_type=content_type)
                    not_remove_contacts.append(contact_serializer.data['id'])
  
        if basic_reference_available and basic_reference_contacts:
            for basic_reference_contact in basic_reference_contacts:
                if 'id' in basic_reference_contact:
                    item = ContactDetails.objects.filter(pk=basic_reference_contact['id']).first()
                    contact_serializer = ContactDetailsSerializer(item, data=basic_reference_contact)
                else:
                    contact_serializer = ContactDetailsSerializer(data=basic_reference_contact)
                if contact_serializer.is_valid():
                    contact_serializer.save(contact_type="Reference", object_id=object_id, content_type=content_type)
                    not_remove_contacts.append(contact_serializer.data['id'])
        
        ContactDetails.objects.filter(object_id=supplier_id).exclude(pk__in=not_remove_contacts).delete()

        society_tower_count = SocietyTower.objects.filter(supplier=society).count()
        difference_of_tower_count = 0
        if tower_count and tower_count > society_tower_count:
            difference_of_tower_count = tower_count - society_tower_count
        if difference_of_tower_count > 0:
            for i in range(difference_of_tower_count):
                tower = SocietyTower(supplier=society, object_id=object_id, content_type=content_type)
                tower.save()
        return Response(serializer.data, status=201)

class SocietyAPIFiltersView(APIView):

    def get(self, request, format=None):
        try:
            item = CityArea.objects.all()
            serializer = CityAreaSerializer(item, many=True)
            return Response(serializer.data)
        except CityArea.DoesNotExist:
            return Response(status=404)

class SocietyList(APIView):
    """
    API to list all societies for a given user. This is new api and hence should be preferred over other.
    """

    def get(self, request):
        class_name = self.__class__.__name__
        try:
            user = request.user
            org_id = request.user.profile.organisation.organisation_id
            state = request.GET.get("state")
            state_name = request.GET.get("state_name")
            search = request.GET.get("search")
            mobile_number = request.GET.get("mobile_number")
            society_objects = []
            if mobile_number:

                count = len(str(mobile_number))
                if count is not 10:
                    
                    return handle_response(class_name,
                     data={"error":"Please enter 10 digit number"}, 
                     request=request)
                else:
                    society_objects = SupplierTypeSociety.objects
                    contact_details = ContactDetails.objects.values_list('object_id', flat=True).filter(
                        Q(landline=mobile_number) | Q(mobile=mobile_number))
                    
                    if contact_details:
                        society_objects =  society_objects.filter(
                            supplier_id__in = contact_details)
                    else:
                        return handle_response(class_name,data={"error":"No data found"}, 
                             request=request)
            else:

                if user.is_superuser:
                    society_objects = SupplierTypeSociety.objects
                    #sort list based on state selected.
                    if state and state_name:
                        society_objects = society_objects.filter(Q(society_state=state) | Q(society_state=state_name))

                    #sort list based on search keyword.
                    if search:
                        society_objects = society_objects.filter(Q(society_name__icontains=search) | Q(society_address1__icontains=search) | Q(society_address2__icontains=search)
                                          | Q(society_city__icontains=search) | Q(supplier_id__icontains=search) | Q(supplier_code__icontains=search))

                    society_objects = society_objects.all().order_by('society_name')
                else:
                    # city_query = get_region_based_query(user, v0_constants.valid_regions['CITY'],
                    #                                              v0_constants.society)                
                    vendor_ids = Organisation.objects.filter(created_by_org=org_id).values('organisation_id')
                    society_objects = SupplierTypeSociety.objects.filter((Q(representative__in=vendor_ids) | Q(representative=org_id))
                                                                          & Q(representative__isnull=False))
                    #sort list based on state selected.
                    if state and state_name:
                        society_objects = society_objects.filter(Q(society_state=state) | Q(society_state=state_name))

                    #sort list based on search keyword.
                    if search:
                        society_objects = society_objects.filter(Q(society_name__icontains=search) | Q(society_address1__icontains=search) | Q(society_address2__icontains=search)
                                          | Q(society_city__icontains=search) | Q(supplier_id__icontains=search) | Q(supplier_code__icontains=search))

                    society_objects = society_objects.all().order_by('society_name')

            #pagination
            society_objects_paginate = paginate(society_objects,SupplierTypeSocietySerializer,request)
            suppliers = manipulate_object_key_values(society_objects_paginate["list"])
            societies_with_images = get_supplier_image(suppliers, v0_constants.society_name)

            data = {
                'count': society_objects_paginate["count"],
                'has_next':society_objects_paginate["has_next"],
                'has_previous':society_objects_paginate["has_previous"],
                'societies': societies_with_images,
            }

            return handle_response(class_name, data=data, success=True)
        except Exception as e:
            return handle_response(class_name, exception_object=e, request=request)



class SocietyListByNumber(APIView):

    def get(self, request):
        class_name = self.__class__.__name__

        try:
            search = request.GET.get("search")
            count = len(str(search))
            if count is not 10:
                
                return handle_response(class_name,
                 data={"error":"Please enter 10 digit number"}, 
                 request=request)
            else:
                society_objects = SupplierTypeSociety.objects
                contact_details = ContactDetails.objects.values_list('object_id', flat=True).filter(
                    Q(landline=search) | Q(mobile=search))
                
                if contact_details:
            
                    society_objects =  society_objects.filter(
                        supplier_id__icontains = contact_details)
                else:
                    return handle_response(class_name,data={"error":"No data found"}, 
                         request=request)
                #pagination
                society_objects_paginate = paginate(society_objects,SupplierTypeSocietySerializer,request)
                suppliers = manipulate_object_key_values(society_objects_paginate["list"])
                societies_with_images = get_supplier_image(suppliers, v0_constants.society_name)

                data = {
                    'count': society_objects_paginate["count"],
                    'has_next':society_objects_paginate["has_next"],
                    'has_previous':society_objects_paginate["has_previous"],
                    'societies': societies_with_images,
                }
                return handle_response(class_name, data=data, success=True)

        except Exception as e:
            return handle_response(class_name, exception_object=e, request=request)



class CorporateViewSet(viewsets.ViewSet):
    """
    A view set around corporates that gives corporate list with filters for state, keyword and pagination.
    """

    def list(self, request):
        class_name = self.__class__.__name__
        try:
            user = request.user
            org_id = request.user.profile.organisation.organisation_id
            corporate_objects = []
            state = request.GET.get("state")
            state_name = request.GET.get("state_name")
            search = request.GET.get("search")

            if user.is_superuser:
                corporate_objects = SupplierTypeCorporate.objects
        
                if state and state_name:
                    corporate_objects = corporate_objects.filter(Q(state=state) | Q(state=state_name))

                if search:
                    corporate_objects = corporate_objects.filter(Q(name__icontains=search) | Q(address1__icontains=search) | Q(address2__icontains=search) 
                                      | Q(city__icontains=search) | Q(supplier_id__icontains=search) | Q(supplier_code__icontains=search))

                corporate_objects = corporate_objects.all().order_by('name')    
            else:
                vendor_ids = Organisation.objects.filter(created_by_org=org_id).values('organisation_id')
                corporate_objects = SupplierTypeCorporate.objects.filter((Q(representative__in=vendor_ids) | Q(representative=org_id))
                                                                      & Q(representative__isnull=False))

                if state and state_name:
                    corporate_objects = corporate_objects.filter(Q(state=state) | Q(state=state_name))

                if search:
                    corporate_objects = corporate_objects.filter(Q(name__icontains=search) | Q(address1__icontains=search) | Q(address2__icontains=search) 
                                      | Q(city__icontains=search) | Q(supplier_id__icontains=search) | Q(supplier_code__icontains=search))

                corporate_objects = corporate_objects.all().order_by('name')

            #pagination
            corporate_objects_paginate = paginate(corporate_objects,UICorporateSerializer,request)
            corporates_with_images = get_supplier_image(corporate_objects_paginate["list"], 'Corporate')

            data = {
                'count': corporate_objects_paginate["count"],
                'has_next':corporate_objects_paginate["has_next"],
                'has_previous':corporate_objects_paginate["has_previous"],
                'corporates': corporates_with_images
            }
            return handle_response(class_name, data=data, success=True)
        except ObjectDoesNotExist as e:
            return handle_response(class_name, exception_object=e, request=request)

        except Exception as e:
            return handle_response(class_name, exception_object=e, request=request)

    def retrieve(self, request, pk=None):
        """
        Retrieve Corporate
        """
        class_name = self.__class__.__name__
        try:
            corporate_instance = SupplierTypeCorporate.objects.get(pk=pk)
            serializer = UICorporateSerializer(corporate_instance)
            return handle_response(class_name, data=serializer.data, success=True)
        except ObjectDoesNotExist as e:
            return handle_response(class_name, data=pk, exception_object=e)
        except Exception as e:
            return handle_response(class_name, exception_object=e, request=request)

    def update(self, request, pk=None):
        """
        Update a corporate
        Args:
            request: A request body
            pk: pk value

        Returns: updated one object

        """
        class_name = self.__class__.__name__
        try:
            corporate_instance = SupplierTypeCorporate.objects.get(pk=pk)
            serializer = UICorporateSerializer(corporate_instance, data=request.data)
            if serializer.is_valid():
                serializer.save()
                return handle_response(class_name, data=serializer.data, success=True)
            return handle_response(class_name, data=serializer.errors)
        except ObjectDoesNotExist as e:
            return handle_response(class_name, data=pk, exception_object=e)
        except Exception as e:
            return handle_response(class_name, exception_object=e, request=request)

    def create(self, request):
        """
        Create a corporate
        Args:
            request:  Request body

        Returns: Created Corporate
        """
        class_name = self.__class__.__name__
        try:
            serializer = UICorporateSerializer(data=request.data)
            if serializer.is_valid():
                serializer.save()
                return handle_response(class_name, data=serializer.data, success=True)
            return handle_response(class_name, data=serializer.errors)
        except Exception as e:
            return handle_response(class_name, exception_object=e, request=request)

    @list_route(methods=['GET'])
    def search(self, request):
        """
        search a corporate
        :param request:
        :return:
        """
        class_name = self.__class__.__name__
        try:
            query = request.query_params['query']

            corporates = SupplierTypeCorporate.objects.filter(
                Q(supplier_id__icontains=query) | Q(name__icontains=query) | Q(address1__icontains=query) | Q(
                    city__icontains=query) | Q(state__icontains=query)).order_by('name')
            serializer = UICorporateSerializer(corporates, many=True)

            corporates_with_images = get_supplier_image(serializer.data, 'Corporate')
            paginator = PageNumberPagination()
            result_page = paginator.paginate_queryset(corporates_with_images, request)

            paginator_response = paginator.get_paginated_response(result_page)
            data = {
                'count': len(corporates_with_images),
                'corporates': paginator_response.data
            }
            return handle_response(class_name, data=serializer.data, success=True)
        except Exception as e:
            return handle_response(class_name, exception_object=e, request=request)


class SalonViewSet(viewsets.ViewSet):
    def list(self, request):
        class_name = self.__class__.__name__
        try:
            user = request.user
            org_id = request.user.profile.organisation.organisation_id
            salon_objects = []
            state = request.GET.get("state")
            state_name = request.GET.get("state_name")
            search = request.GET.get("search")

            if user.is_superuser:
                salon_objects = SupplierTypeSalon.objects

                if state and state_name:
                    salon_objects = salon_objects.filter(Q(state=state) | Q(state=state_name))

                if search:
                    salon_objects = salon_objects.filter(Q(name__icontains=search) | Q(address1__icontains=search) | Q(address2__icontains=search) 
                                      | Q(city__icontains=search) | Q(supplier_id__icontains=search) | Q(supplier_code__icontains=search))
 
                salon_objects = salon_objects.all().order_by('name')
            else:
                vendor_ids = Organisation.objects.filter(created_by_org=org_id).values('organisation_id')
                salon_objects = SupplierTypeSalon.objects.filter((Q(representative__in=vendor_ids) | Q(representative=org_id))
                                                                      & Q(representative__isnull=False))

                if state and state_name:
                    salon_objects = salon_objects.filter(Q(state=state) | Q(state=state_name))

                if search:
                    salon_objects = salon_objects.filter(Q(name__icontains=search) | Q(address1__icontains=search) | Q(address2__icontains=search) 
                                      | Q(city__icontains=search) | Q(supplier_id__icontains=search) | Q(supplier_code__icontains=search))

                salon_objects = salon_objects.all().order_by('name')

            #pagination
            salon_objects_paginate = paginate(salon_objects,UISalonSerializer,request)
            items = get_supplier_image(salon_objects_paginate["list"], 'Salon')

            data = {
                'count': salon_objects_paginate["count"],
                'has_next':salon_objects_paginate["has_next"],
                'has_previous':salon_objects_paginate["has_previous"],
                'salons': items
            }
            return handle_response(class_name, data=data, success=True)
        except ObjectDoesNotExist as e:
            return handle_response(class_name, data='Salon does not exist', exception_object=e)
        except Exception as e:
            return handle_response(class_name, exception_object=e, request=request)

class GymViewSet(viewsets.ViewSet):

    def list(self, request):

        class_name = self.__class__.__name__
        try:
            user = request.user
            org_id = request.user.profile.organisation.organisation_id
            gym_objects = []
            state = request.GET.get("state")
            state_name = request.GET.get("state_name")
            search = request.GET.get("search")

            if user.is_superuser:
                gym_objects = SupplierTypeGym.objects
        
                if state and state_name:
                    gym_objects = gym_objects.filter(Q(state=state) | Q(state=state_name))

                if search:
                    gym_objects = gym_objects.filter(Q(name__icontains=search) | Q(address1__icontains=search) | Q(address2__icontains=search) 
                                      | Q(city__icontains=search) | Q(supplier_id__icontains=search) | Q(supplier_code__icontains=search))

                gym_objects = gym_objects.all().order_by('name')
            else:
                vendor_ids = Organisation.objects.filter(created_by_org=org_id).values('organisation_id')
                gym_objects = SupplierTypeGym.objects.filter((Q(representative__in=vendor_ids) | Q(representative=org_id))
                                                                      & Q(representative__isnull=False))

                if state and state_name:
                    gym_objects = gym_objects.filter(Q(state=state) | Q(state=state_name))

                if search:
                    gym_objects = gym_objects.filter(Q(name__icontains=search) | Q(address1__icontains=search) | Q(address2__icontains=search) 
                                      | Q(city__icontains=search) | Q(supplier_id__icontains=search) | Q(supplier_code__icontains=search))

                gym_objects = gym_objects.all().order_by('name')

            #pagination
            gym_objects_paginate = paginate(gym_objects,SupplierTypeGymSerializer,request)
            items = get_supplier_image(gym_objects_paginate['list'], 'Gym')
            
            data = {
                'count': gym_objects_paginate["count"],
                'has_next':gym_objects_paginate["has_next"],
                'has_previous':gym_objects_paginate["has_previous"],
                'gyms': items
            }
            return handle_response(class_name, data=data, success=True)
        except ObjectDoesNotExist as e:
            return handle_response(class_name, data='Gym does not exist', exception_object=e)
        except Exception as e:
            return handle_response(class_name, exception_object=e, request=request)

class SocietyAPIFiltersListView(APIView):

    # self.paginator = None
    # self.serializer = None

    # def get(self,request, format=None):
    #     return self.paginator.get_paginated_response(self.serializer.data)

    def post(self, request, format=None):
        try:
            # two list to disable search on society and flats if all the options in both the fields selected
            allflatquantity = [u'Large', u'Medium', u'Small', u'Very Large']  # in sorted order
            allsocietytype = ['High', 'Medium High', 'Standard', 'Ultra High']  # in sorted order
            cityArea = []
            societytype = []
            flatquantity = []
            inventorytype = []
            citySubArea = []
            filter_present = False
            subareas = False
            areas = False

            if 'subLocationValueModel' in request.data:
                for key in request.data['subLocationValueModel']:
                    citySubArea.append(key['subarea_name'])
                    # filter_present = True
                subareas = True if citySubArea else False

            if (not subareas) and 'locationValueModel' in request.data:
                for key in request.data['locationValueModel']:
                    cityArea.append(key['label'])
                    # filter_present = True
                areas = True if cityArea else False

            if 'typeValuemodel' in request.data:
                for key in request.data['typeValuemodel']:
                    societytype.append(key['label'])
                    # filter_present = True

            if 'checkboxes' in request.data:
                for key in request.data['checkboxes']:
                    if key['checked']:
                        flatquantity.append(key['name'])
                        # filter_present = True

            if 'types' in request.data:
                for key in request.data['types']:
                    if key['checked']:
                        inventorytype.append(key['inventoryname'])
                        # filter_present = True

            flatquantity.sort()
            societytype.sort()
            if flatquantity == allflatquantity:  # sorted comparison to avoid mismatch based on index
                flatquantity = []
            if societytype == allsocietytype:  # same as above
                societytype = []

            if subareas or areas or societytype or flatquantity or inventorytype:
                filter_present = True

            if filter_present:
                # if subareas:
                #     items = SupplierTypeSociety.objects.filter(Q(society_subarea__in = citySubArea) & Q(society_type_quality__in = societytype) & Q(society_type_quantity__in = flatquantity))
                # else :
                #     items = SupplierTypeSociety.objects.filter(Q(society_locality__in = cityArea) & Q(society_type_quality__in = societytype) & Q(society_type_quantity__in = flatquantity))
                # serializer = UISocietySerializer(items, many= True)

                # Sample Code

                if subareas:
                    if societytype and flatquantity and inventorytype:
                        items = SupplierTypeSociety.objects.filter(
                            Q(society_subarea__in=citySubArea) & Q(society_type_quality__in=societytype) & Q(
                                society_type_quantity__in=flatquantity))
                    elif societytype and flatquantity:
                        items = SupplierTypeSociety.objects.filter(
                            Q(society_subarea__in=citySubArea) & Q(society_type_quality__in=societytype) & Q(
                                society_type_quantity__in=flatquantity))
                    elif societytype and inventorytype:
                        items = SupplierTypeSociety.objects.filter(
                            Q(society_subarea__in=citySubArea) & Q(society_type_quality__in=societytype))
                    elif flatquantity and inventorytype:
                        items = SupplierTypeSociety.objects.filter(
                            Q(society_subarea__in=citySubArea) & Q(society_type_quantity__in=flatquantity))
                    elif societytype:
                        items = SupplierTypeSociety.objects.filter(
                            Q(society_subarea__in=citySubArea) & Q(society_type_quality__in=societytype))
                    elif flatquantity:
                        items = SupplierTypeSociety.objects.filter(
                            Q(society_subarea__in=citySubArea) & Q(society_type_quantity__in=flatquantity))
                    # elif inventorytype:
                    #     do something
                    else:
                        items = SupplierTypeSociety.objects.filter(society_subarea__in=citySubArea)

                elif areas:
                    if societytype and flatquantity and inventorytype:
                        items = SupplierTypeSociety.objects.filter(
                            Q(society_locality__in=cityArea) & Q(society_type_quality__in=societytype) & Q(
                                society_type_quantity__in=flatquantity))
                    elif societytype and flatquantity:
                        items = SupplierTypeSociety.objects.filter(
                            Q(society_locality__in=cityArea) & Q(society_type_quality__in=societytype) & Q(
                                society_type_quantity__in=flatquantity))
                    elif societytype and inventorytype:
                        items = SupplierTypeSociety.objects.filter(
                            Q(society_locality__in=cityArea) & Q(society_type_quality__in=societytype))
                    elif flatquantity and inventorytype:
                        items = SupplierTypeSociety.objects.filter(
                            Q(society_locality__in=cityArea) & Q(society_type_quantity__in=flatquantity))
                    elif societytype:
                        items = SupplierTypeSociety.objects.filter(
                            Q(society_locality__in=cityArea) & Q(society_type_quality__in=societytype))
                    elif flatquantity:
                        items = SupplierTypeSociety.objects.filter(
                            Q(society_locality__in=cityArea) & Q(society_type_quantity__in=flatquantity))
                    # elif inventorytype:
                    #     do something

                    else:
                        items = SupplierTypeSociety.objects.filter(society_locality__in=cityArea)

                elif societytype or flatquantity or inventorytype:
                    if societytype and flatquantity and inventorytype:
                        items = SupplierTypeSociety.objects.filter(
                            Q(society_type_quality__in=societytype) & Q(society_type_quantity__in=flatquantity))
                    elif societytype and flatquantity:
                        items = SupplierTypeSociety.objects.filter(
                            Q(society_type_quality__in=societytype) & Q(society_type_quantity__in=flatquantity))
                    elif societytype and inventorytype:
                        items = SupplierTypeSociety.objects.filter(Q(society_type_quality__in=societytype))
                    elif flatquantity and inventorytype:
                        items = SupplierTypeSociety.objects.filter(Q(society_type_quantity__in=flatquantity))
                    elif societytype:
                        items = SupplierTypeSociety.objects.filter(Q(society_type_quality__in=societytype))
                    elif flatquantity:
                        items = SupplierTypeSociety.objects.filter(Q(society_type_quantity__in=flatquantity))
                    # elif inventorytype:
                    #     do something

                else:
                    items = SupplierTypeSociety.objects.all()

                ## UISocietySerializer --> SocietyListSerializer
                # serializer = SocietyListSerializer(items, many= True)
                # Sample Code
            else:
                items = SupplierTypeSociety.objects.all()
                ## UISocietySerializer --> SocietyListSerializer

            serializer = SocietyListSerializer(items, many=True)
            # paginator = PageNumberPagination()
            # result_page = paginator.paginate_queryset(items, request)
            # serializer = SocietyListSerializer(result_page, many=True)

            # return paginator.get_paginated_response(serializer.data)

            return Response(serializer.data, status=200)
        except SupplierTypeSociety.DoesNotExist:
            return Response(status=404)

        # def set_paginator(self, paginator, serializer):
        #     self.filter_socities_paginator = paginator
        #     self.filter_socities_serializer = serializer

        # def get_paginator(self):
        #     return self.filter_socities_paginator,

class SocietyAPISortedListView(APIView):
    def post(self, request, format=None):
        order = request.query_params.get('order', None)
        society_ids = request.data

        if order == 'asc':
            societies = SupplierTypeSociety.objects.filter(supplier_id__in=society_ids).order_by('society_name')
            serializer = SocietyListSerializer(societies, many=True)
            return Response(serializer.data, status=200)
        elif order == 'desc':
            societies = SupplierTypeSociety.objects.filter(supplier_id__in=society_ids).order_by('-society_name')
            serializer = SocietyListSerializer(societies, many=True)
            return Response(serializer.data, status=200)
        else:
            return Response(status=200)

class SocietyAPISocietyIdsView(APIView):
    def get(self, request, format=None):
        supplier_type_code = request.query_params.get('supplier_type_code', None)
        if supplier_type_code == 'RS':
            supplier_ids = SupplierTypeSociety.objects.all().values_list('supplier_id', flat=True)
        else:
            supplier_ids = SupplierMaster.objects.all().values_list('supplier_id', flat=True)
        return Response({'supplier_ids': supplier_ids}, status=200)

class SupplierInfoAPIView(APIView):

    def get(self, request, id, format=None):
        try:
            item = SupplierInfo.objects.get(pk=id)
            serializer = SupplierInfoSerializer(item)
            return Response(serializer.data)
        except SupplierInfo.DoesNotExist:
            return Response(status=404)

    def put(self, request, id, format=None):
        try:
            item = SupplierInfo.objects.get(pk=id)
        except SupplierInfo.DoesNotExist:
            return Response(status=404)
        serializer = SupplierInfoSerializer(item, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=400)

    def delete(self, request, id, format=None):
        try:
            item = SupplierInfo.objects.get(pk=id)
        except SupplierInfo.DoesNotExist:
            return Response(status=404)
        item.delete()
        return Response(status=204)

class SupplierInfoAPIListView(APIView):

    def get(self, request, format=None):
        items = SupplierInfo.objects.all()
        paginator = PageNumberPagination()
        result_page = paginator.paginate_queryset(items, request)
        serializer = SupplierInfoSerializer(result_page, many=True)
        return paginator.get_paginated_response(serializer.data)

    def post(self, request, format=None):
        serializer = SupplierInfoSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=201)
        return Response(serializer.errors, status=400)

class SupplierTypeSocietyAPIListView(APIView):

    def get(self, request, format=None):
        items = SupplierTypeSociety.objects.all()
        paginator = PageNumberPagination()
        result_page = paginator.paginate_queryset(items, request)
        serializer = SupplierTypeSocietySerializer(result_page, many=True)
        return paginator.get_paginated_response(serializer.data)

    def post(self, request, format=None):
        """
        Args:
            request: request param containg society data
            format: none

        Returns: The data created by this particular user
        """
        serializer = SupplierTypeSocietySerializer(data=request.data)
        current_user = request.user
        if serializer.is_valid():
            serializer.save(created_by=current_user)
            return Response(serializer.data, status=201)
        return Response(serializer.errors, status=400)

class SupplierTypeSocietyAPIView(APIView):

    def get(self, request, id, format=None):
        try:
            item = SupplierTypeSociety.objects.get(pk=id)
            serializer = SupplierTypeSocietySerializer(item)
            return Response(serializer.data)
        except SupplierTypeSociety.DoesNotExist:
            return Response(status=404)

    def put(self, request, id, format=None):
        try:
            item = SupplierTypeSociety.objects.get(pk=id)
        except SupplierTypeSociety.DoesNotExist:
            return Response(status=404)
        serializer = SupplierTypeSocietySerializer(item, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=400)

    def delete(self, request, id, format=None):
        try:
            item = SupplierTypeSociety.objects.get(pk=id)
        except SupplierTypeSociety.DoesNotExist:
            return Response(status=404)
        item.delete()
        return Response(status=204)


class RetailShopViewSet(viewsets.ViewSet):
    """
    View Set around RetailShop
    """

    def create(self, request):

        class_name = self.__class__.__name__
        try:
            serializer = RetailShopSerializer(data=request.data)
            if serializer.is_valid():
                serializer.save()
                return handle_response(class_name, data=serializer.data, success=True)
            return handle_response(class_name, data=serializer.errors)
        except Exception as e:
            return handle_response(class_name, exception_object=e, request=request)

    def list(self, request):
        class_name = self.__class__.__name__
        try:
            # all retail_shop_objects sorted by name
            user = request.user
            org_id = request.user.profile.organisation.organisation_id
            retail_shop_objects = []
            state = request.GET.get("state")
            state_name = request.GET.get("state_name")
            search = request.GET.get("search")

            if user.is_superuser:
                retail_shop_objects = SupplierTypeRetailShop.objects

                if state and state_name:
                    retail_shop_objects = retail_shop_objects.filter(Q(state=state) | Q(state=state_name))

                if search:
                    retail_shop_objects = retail_shop_objects.filter(Q(name__icontains=search) | Q(address1__icontains=search) | Q(address2__icontains=search) 
                                      | Q(city__icontains=search) | Q(supplier_id__icontains=search) | Q(supplier_code__icontains=search))

                retail_shop_objects = retail_shop_objects.all().order_by('name')
            else:
                vendor_ids = Organisation.objects.filter(created_by_org=org_id).values('organisation_id')
                retail_shop_objects = SupplierTypeRetailShop.objects.filter((Q(representative__in=vendor_ids) | Q(representative=org_id))
                                                                      & Q(representative__isnull=False))
               
                if state and state_name:
                    retail_shop_objects = retail_shop_objects.filter(Q(state=state) | Q(state=state_name))

                if search:
                    retail_shop_objects = retail_shop_objects.filter(Q(name__icontains=search) | Q(address1__icontains=search) | Q(address2__icontains=search) 
                                      | Q(city__icontains=search) | Q(supplier_id__icontains=search) | Q(supplier_code__icontains=search))

                retail_shop_objects = retail_shop_objects.all().order_by('name')

            #pagination
            retail_shop_objects_paginate = paginate(retail_shop_objects,RetailShopSerializer,request)
            retail_shop_with_images = get_supplier_image(retail_shop_objects_paginate["list"], 'Retail Shop')

            data = {
                'count': retail_shop_objects_paginate["count"],
                'has_next':retail_shop_objects_paginate["has_next"],
                'has_previous':retail_shop_objects_paginate["has_previous"],
                'retail_shop_objects': retail_shop_with_images
            }
            return handle_response(class_name, data=data, success=True)
        except Exception as e:
            return handle_response(class_name, data=str(e), request=request)

    def update(self, request, pk):
        class_name = self.__class__.__name__
        try:

            basic_contacts = request.data.get('basic_contacts', None)
            object_id = request.data.get('supplier_id', None)
            if basic_contacts is not None:
                for contact in basic_contacts:
                    if 'id' in contact:
                        item = ContactDetails.objects.filter(pk=contact['id']).first()
                        contact_serializer = ContactDetailsSerializer(item, data=contact)
                    else:
                        contact_serializer = ContactDetailsSerializer(data=contact)
                    if contact_serializer.is_valid():
                        contact_serializer.save()

            supplier_type_code = request.data.get('supplier_type_code', None)

            if not supplier_type_code:
                return ui_utils.handle_response({}, data='supplier_type_code is Mandatory')
            ownership = request.data.get('ownership_details', None)

            model_name = get_model(supplier_type_code)
            serializer_name = get_serializer(supplier_type_code)


            instance = model_name.objects.get(pk=pk)
            serializer = serializer_name(instance=instance, data=request.data)
            
            if serializer.is_valid():
                serializer.save()
                contact_and_ownership = update_contact_and_ownership_detail(request.data)
                return handle_response(class_name, data=serializer.data, success=True)
            
            return handle_response(class_name, data=serializer.data, success=True)
        
        except Exception as e:
            return handle_response(class_name, exception_object=e, request=request)

    def retrieve(self, request, pk):
        class_name = self.__class__.__name__
        try:
            retail_shop_instance = SupplierTypeRetailShop.objects.get(pk=pk)
            serializer = RetailShopSerializer(instance=retail_shop_instance)

            shopData = serializer.data

            shopData['contactData'],shopData['ownership_details'] = retrieve_contact_and_ownership_detail(pk)

            return handle_response(class_name, data=shopData, success=True)
        except Exception as e:
            return handle_response(class_name, exception_object=e, request=request)


class HordingViewSet(viewsets.ViewSet):
    
    def list(self, request):
        class_name = self.__class__.__name__
        try:
            # all retail_shop_objects sorted by name
            user = request.user
            org_id = request.user.profile.organisation.organisation_id
            hording_objects = []
            state = request.GET.get("state")
            state_name = request.GET.get("state_name")
            search = request.GET.get("search")

            if user.is_superuser:
                hording_objects = SupplierHording.objects

                if state and state_name:
                    hording_objects = hording_objects.filter(Q(state=state) | Q(state=state_name))

                if search:
                    hording_objects = hording_objects.filter(Q(name__icontains=search) | Q(address1__icontains=search) | Q(address2__icontains=search) 
                                      | Q(city__icontains=search) | Q(supplier_id__icontains=search) | Q(supplier_code__icontains=search))

                hording_objects = hording_objects.all().order_by('name')
            else:

                hording_objects = SupplierHording.objects
                
                if state and state_name:
                    hording_objects = hording_objects.filter(Q(state=state) | Q(state=state_name))

                if search:
                    hording_objects = hording_objects.filter(Q(name__icontains=search) | Q(address1__icontains=search) | Q(address2__icontains=search) 
                                      | Q(city__icontains=search) | Q(supplier_id__icontains=search) | Q(supplier_code__icontains=search))

                hording_objects = hording_objects.all().order_by('name')

            #pagination
            hording_objects_paginate = paginate(hording_objects, SupplierHordingSerializer, request)
            hording_with_images = get_supplier_image(hording_objects_paginate["list"], 'Retail Shop')

            data = {
                'count': hording_objects_paginate["count"],
                'has_next':hording_objects_paginate["has_next"],
                'has_previous':hording_objects_paginate["has_previous"],
                'hording_objects': hording_with_images
            }
            return handle_response(class_name, data=data, success=True)
        except Exception as e:
            return handle_response(class_name, data=str(e), request=request)

    def retrieve(self, request, pk):
        class_name = self.__class__.__name__
        try:
            hording_instance = SupplierHording.objects.get(pk=pk)
            serializer = SupplierHordingSerializer(instance=hording_instance)

            result = serializer.data

            result['contactData'],result['ownership_details'] = retrieve_contact_and_ownership_detail(pk)

            return handle_response(class_name, data=result, success=True)
        except Exception as e:
            return handle_response(class_name, exception_object=e, request=request)

    def update(self, request, pk):
        class_name = self.__class__.__name__
        try:
            contact_and_ownership = update_contact_and_ownership_detail(request.data)
            hording_instance = SupplierHording.objects.get(pk=pk)
            serializer = SupplierHordingSerializer(instance=hording_instance, data=request.data)
            if serializer.is_valid():
                serializer.save()
                return handle_response(class_name, data=serializer.data, success=True)
            return handle_response(class_name, data=serializer.data, success=True)
        except Exception as e:
            return handle_response(class_name, exception_object=e, request=request)


class EducationalInstituteViewSet(viewsets.ViewSet):
    """
    View Set around EducationalInstitute
    """
    def list(self, request):
        class_name = self.__class__.__name__
        try:
           
            user = request.user
            org_id = request.user.profile.organisation.organisation_id
            educational_institute_objects = []
            state = request.GET.get("state")
            state_name = request.GET.get("state_name")
            search = request.GET.get("search")

            if user.is_superuser:

                educational_institute_objects = SupplierEducationalInstitute.objects

                if state and state_name:
                    educational_institute_objects = educational_institute_objects.filter(Q(state=state) | Q(state=state_name))

                if search:
                    educational_institute_objects = educational_institute_objects.filter(Q(name__icontains=search) | Q(address1__icontains=search) | Q(address2__icontains=search) 
                                      | Q(city__icontains=search) | Q(supplier_id__icontains=search) | Q(supplier_code__icontains=search))

                educational_institute_objects = educational_institute_objects.all().order_by('name')
            else:
                #vendor_ids = Organisation.objects.filter(created_by_org=org_id).values('organisation_id')
                #educational_institute_objects = SupplierEducationalInstitute.objects.filter((Q(representative__in=vendor_ids) | Q(representative=org_id)) & Q(representative__isnull=False))

                educational_institute_objects = SupplierEducationalInstitute.objects.filter()
               
                if state and state_name:
                    educational_institute_objects = educational_institute_objects.filter(Q(state=state) | Q(state=state_name))

                if search:
                    educational_institute_objects = educational_institute_objects.filter(Q(name__icontains=search) | Q(address1__icontains=search) | Q(address2__icontains=search) 
                                      | Q(city__icontains=search) | Q(supplier_id__icontains=search) | Q(supplier_code__icontains=search))
               
                educational_institute_objects = educational_institute_objects.all().order_by('name')
          # #pagination
            educational_institute_objects_paginate = paginate(educational_institute_objects,SupplierEducationalInstituteSerializer,request)
            educational_institute_with_images = get_supplier_image(educational_institute_objects_paginate["list"], 'Educational Institute')

            data = {
                'count': educational_institute_objects_paginate["count"],
                'has_next':educational_institute_objects_paginate["has_next"],
                'has_previous':educational_institute_objects_paginate["has_previous"],
                'educational_institute_objects': educational_institute_with_images
            }
            return handle_response(class_name, data=data, success=True)
        except Exception as e:
            return handle_response(class_name, data=str(e), request=request)
    

class SaveBasicCorporateDetailsAPIView(APIView):
    def post(self, request, id, format=None):
        class_name = self.__class__.__name__
        try:
            companies = []
            error = {}
            # Round 1 Saving basic data
            if 'supplier_id' in request.data:
                corporate = SupplierTypeCorporate.objects.filter(pk=request.data['supplier_id']).first()
                if corporate:
                    corporate_serializer = SupplierTypeCorporateSerializer(corporate, data=request.data)
                else:
                    corporate_serializer = SupplierTypeCorporateSerializer(data=request.data)
                if corporate_serializer.is_valid():
                    corporate_serializer.save()
                else:
                    error['message'] = 'Invalid Corporate Info data'
                    error = json.dumps(error)
                    return Response(corporate_serializer.errors, status=406)
            else:
                return Response({"status": False, "error": "No supplier id in request.data"},
                                status=status.HTTP_400_BAD_REQUEST)

            # Round 2 Saving List of companies

            corporate_id = request.data['supplier_id']
            companies_name = request.data.get('list1')
            company_ids = list(
                CorporateParkCompanyList.objects.filter(supplier_id=corporate_id).values_list('id', flat=True))

            for company_name in companies_name:
                if 'id' is company_name.get('name'):
                    company = CorporateParkCompanyList.objects.get(id=id)
                    company.name = company_name.get('name')
                    company.employeeCount = company_name.get('employeeCount')
                    company_ids.remove(company.id)
                    companies.append(company)
                else:
                    company = CorporateParkCompanyList(supplier_id_id=corporate_id, name=company_name.get('name'), employeeCount=company_name.get('employeeCount'))
                    companies.append(company)

            CorporateParkCompanyList.objects.bulk_create(companies)
            CorporateParkCompanyList.objects.filter(id__in=company_ids).delete()
            # Round 3 - Saving contacts
            try:
                instance = SupplierTypeCorporate.objects.get(supplier_id=id)
            except SupplierTypeCorporate.DoesNotExist:
                return Response({'message': 'This corporate park does not exist'}, status=406)

            content_type = ContentType.objects.get_for_model(SupplierTypeCorporate)
            object_id = instance.supplier_id
            # in order to save get calls in a loop, prefetch all the contacts for this supplier beforehand
            # making a dict which key is object id and contains another
            contact_detail_objects = {contact.id: contact for contact in
                                      ContactDetails.objects.filter(content_type=content_type, object_id=object_id)}

            # get all contact id's in a set. Required for contact deletion
            contact_detail_ids = set([contact_id for contact_id in contact_detail_objects.keys()])

            for contact in request.data.get('contacts'):
                # make the data you want to save in contacts
                contact_data = {
                    'name': contact.get('name'),
                    'country_code': contact.get('country_code'),
                    'std_code': contact.get('std_code'),
                    'mobile': contact.get('mobile'),
                    'contact_type': contact.get('contact_type'),
                    'object_id': object_id,
                    'content_type': content_type.id,
                    'email': contact.get('email'),
                    'salutation': contact.get('salutation'),
                    'landline': contact.get('landline'),
                }

                # get the contact instance to be updated if id was present else create a brand new instance
                contact_instance = contact_detail_objects[contact['id']] if 'id' in contact else None

                # if contact instance was there this means this is an update request. we need to remove this id
                # from the set of id's we are maintaining because we do not want this instance to be deleted.

                if contact_instance:
                    contact_detail_ids.remove(contact['id'])
                # save the data
                serializer = ContactDetailsSerializer(contact_instance, data=contact_data)

                if serializer.is_valid():
                    serializer.save()
                else:
                    return handle_response(class_name, data=serializer.errors)

            # in the end we need to delete the crossed out contacts. all contacts now in the list of ids are
            # being deleted by the user hence we delete it from here too.
            ContactDetails.objects.filter(id__in=contact_detail_ids).delete()

            # todo: to be changed later
            '''

            content_type = ContentType.objects.get_for_model(SupplierTypeCorporate)

            contacts_ids = ContactDetailsGeneric.objects.filter(content_type=content_type, object_id=instance.supplier_id).values_list('id',flat=True)
            contacts_ids = list(contacts_ids)


            for contact in request.data['contacts']:
                if 'id' in contact:
                    contact_instance = ContactDetailsGeneric.objects.get(id=contact['id'])
                    contacts_ids.remove(contact_instance.id)
                    serializer = ContactDetailsGenericSerializer(contact_instance, data=contact)
                    if serializer.is_valid():
                        serializer.save()
                    else:
                        return Response(status=404)

                else:
                    contact['object_id'] = instance.supplier_id
                    serializer = ContactDetailsGenericSerializer(data=contact)
                    if serializer.is_valid():
                        serializer.save(content_type=content_type)
                    else:
                        return Response(status=404)

            ContactDetailsGeneric.objects.filter(id__in=contacts_ids).delete()

            '''
            # Round 4 - Creating number of fields in front end in Building Model.
            buildingcount = CorporateBuilding.objects.filter(corporatepark_id=request.data['supplier_id']).count()
            diff_count = 0
            new_list = []
            building_count_recieved = int(request.data['building_count'])
            if building_count_recieved > buildingcount:
                diff_count = building_count_recieved - buildingcount
            if 'building_count' in request.data:
                for i in range(diff_count):
                    instance = CorporateBuilding(corporatepark_id_id=request.data['supplier_id'])
                    new_list.append(instance)

            CorporateBuilding.objects.bulk_create(new_list)

        except KeyError as e:
            return Response({"status": False, "error": str(e.message)}, status=status.HTTP_400_BAD_REQUEST)
        except ObjectDoesNotExist as e:
            return Response({"status": False, "error": str(e.message)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({"status": False, "error": str(e.message)}, status=status.HTTP_400_BAD_REQUEST)

        return Response(status=200)

    def get(self, request, id, format=None):
        try:
            supplier = SupplierTypeCorporate.objects.get(supplier_id=id)
            serializer = SupplierTypeCorporateSerializer(supplier)
            companies = CorporateParkCompanyList.objects.filter(supplier_id=id)
            corporate_serializer = CorporateParkCompanyListSerializer(companies, many=True)
            content_type = ContentType.objects.get_for_model(model=SupplierTypeCorporate)
            contacts = ContactDetails.objects.filter(content_type=content_type, object_id=id)
            contacts_serializer = ContactDetailsSerializer(contacts, many=True)
            result = {'basicData': serializer.data, 'companyList': corporate_serializer.data,
                      'contactData': contacts_serializer.data}
            return Response(result)

        except ObjectDoesNotExist as e:
            return Response({"status": False, "error": str(e.message)}, status=status.HTTP_400_BAD_REQUEST)
        except MultipleObjectsReturned as e:
            return Response({"status": False, "error": str(e.message)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({"status": False, "error": str(e.message)}, status=status.HTTP_400_BAD_REQUEST)


# This API is for saving the buildings and wings details of a corporate space

class SaveBuildingDetailsAPIView(APIView):

    def get(self, request, id, format=None):
        try:
            corporate = SupplierTypeCorporate.objects.get(supplier_id=id)
        except SupplierTypeCorporate.DoesNotExist:
            return Response({'message': 'Invalid Corporate ID'}, status=status.HTTP_400_BAD_REQUEST)

        buildings = corporate.get_buildings()
        building_serializer = CorporateBuildingGetSerializer(buildings, many=True)
        return Response(building_serializer.data, status=status.HTTP_200_OK)

    def post(self, request, id, format=None):
        try:
            corporate_object = SupplierTypeCorporate.objects.get(supplier_id=id)
        except SupplierTypeCorporate.DoesNotExist:
            return Response({'message': 'Invalid Corporate ID'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            building_count = len(request.data)
            if corporate_object.building_count != building_count:
                corporate_object.building_count = building_count
                corporate_object.save()

            buildings_ids = set(
                CorporateBuilding.objects.filter(corporatepark_id=corporate_object).values_list('id', flat=True))
            wing_ids_superset = set()

            # Logic for delete - Put all currently present fields in the database into a list.
            # Check the received data from front-end, if ID of any field matches with any in the list, remove that field from list.
            # In the end, delete all the fields left in the list.

            for building in request.data:
                if 'id' in building:
                    basic_data_instance = CorporateBuilding.objects.get(id=building['id'])
                    buildings_ids.remove(building['id'])
                    building_serializer = CorporateBuildingSerializer(basic_data_instance, data=building)
                else:
                    building_serializer = CorporateBuildingSerializer(data=building)

                if building_serializer.is_valid():
                    building_object = building_serializer.save()
                else:
                    return Response(status=406)

                wings_ids = set(
                    CorporateBuildingWing.objects.filter(building_id=building_object).values_list('id', flat=True))

                for wing in building['wingInfo']:

                    if 'id' in wing:
                        wing_object = CorporateBuildingWing.objects.get(id=wing['id'])
                        wings_ids.remove(wing_object.id)
                        wing_serializer = CorporateBuildingWingSerializer(wing_object, data=wing)
                    else:
                        wing_serializer = CorporateBuildingWingSerializer(data=wing)

                    if wing_serializer.is_valid():
                        wing_serializer.save(building_id=building_object)
                    else:
                        return Response({'message': 'Invalid Wing Data', \
                                         'errors': wing_serializer.errors}, status=406)

                if wings_ids:
                    wing_ids_superset = wing_ids_superset.union(wings_ids)

            if wing_ids_superset:
                CorporateBuildingWing.objects.filter(id__in=wing_ids_superset).delete()
            if buildings_ids:
                CorporateBuilding.objects.filter(id__in=buildings_ids).delete()

            return Response({"status": True, "data": ""}, status=status.HTTP_200_OK)

        except KeyError as e:
            return Response({"status": False, "error": str(e.message)}, status=status.HTTP_400_BAD_REQUEST)
        except ObjectDoesNotExist as e:
            return Response({"status": False, "error": str(e.message)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({"status": False, "error": str(e.message)}, status=status.HTTP_400_BAD_REQUEST)


# This API is for fetching the companies and buildings of a specific corporate space.
class CompanyDetailsAPIView(APIView):
    def get(self, request, id, format=True):
        company = SupplierTypeCorporate.objects.get(supplier_id=id)
        company_list = CorporateParkCompanyList.objects.filter(supplier_id_id=company)
        serializer = CorporateParkCompanyListSerializer(company_list, many=True)
        building = SupplierTypeCorporate.objects.get(supplier_id=id)
        building_list = CorporateBuilding.objects.filter(corporatepark_id_id=building)
        serializer1 = CorporateBuildingGetSerializer(building_list, many=True)
        response_dict = {'companies': serializer.data, 'buildings': serializer1.data}
        return Response(response_dict, status=200)


class CorporateCompanyDetailsAPIView(APIView):
    '''
    This API is for saving details of all companies belonging to a specific corporate space.
    '''

    def get(self, request, id=None, format=None):
        try:
            corporate = SupplierTypeCorporate.objects.get(supplier_id=id)
        except SupplierTypeCorporate.DoesNotExist:
            return Response({'message': 'No such Corporate'}, status=406)

        companies = corporate.get_corporate_companies()
        companies_serializer = CorporateParkCompanySerializer(companies, many=True)

        return Response(companies_serializer.data, status=200)

    def post(self, request, id, format=True):
        try:
            corporate = SupplierTypeCorporate.objects.get(supplier_id=id)
        except SupplierTypeCorporate.DoesNotExist:
            return Response({'message': "Corporate Doesn't Exist"}, status=404)

        company_detail_list = []
        floor_list = []
        company_detail_ids = set(
            CorporateCompanyDetails.objects.filter(company_id__supplier_id=corporate).values_list('id', flat=True))
        floor_superset = set()
        company_detail_set = set()
        flag = False

        for company in request.data:
            for company_detail in company['companyDetailList']:
                try:
                    if not company_detail['building_name']:
                        continue
                    # get the right CorporateCompanyDetails object if 'id' exist in the company_detail object.
                    company_detail_instance = CorporateCompanyDetails.objects.get(id=company_detail['id'])

                    # set the id of the company for which this object belongs to
                    company_detail['company_id'] = company['id']

                    # remove the id of CorporateCompanyDetails object from the set of id's.
                    company_detail_ids.remove(company_detail['id'])

                    # update this instance with new data
                    company_detail_instance.__dict__.update(company_detail)

                    # save
                    company_detail_instance.save()

                    flag = True
                except KeyError:
                    # create the CorporateCompanyDetails object if 'id' is not in company_detail object
                    company_detail_instance = CorporateCompanyDetails(company_id_id=company['id'],
                                                                      building_name=company_detail['building_name'],
                                                                      wing_name=company_detail['wing_name'])
                    # company_detail_list.append(company_detail_instance)
                    # uncomment this line if error occurs
                    company_detail_instance.save()
                    flag = False
                # if the CorporateCompanyDetails object was updated
                if flag:
                    # for a given building, find all the floors
                    floor_ids = set(
                        CompanyFloor.objects.filter(company_details_id_id=company_detail_instance.id).values_list('id',
                                                                                                                  flat=True))

                    # for each floor object in company_details['listOfFloors']
                    for floor in company_detail['listOfFloors']:
                        try:
                            # get the floor object
                            floor_instance = CompanyFloor.objects.get(id=floor['id'])

                            # remove the floor id from set of floor id's
                            floor_ids.remove(floor_instance.id)

                            # if the floor number for this building is same as new floor number no point in updating
                            if floor_instance.floor_number == floor['floor_number']:
                                continue

                            # else update the new floor number and also save for which building ( CorporateCompanyDetails ) this floor was updated
                            floor_instance.company_detail_id, floor_instance.floor_number = company_detail_instance, \
                                                                                            floor['floor_number']

                            # floor_instance.__dict__.update(floor)
                            floor_instance.save()
                        except KeyError:
                            # if we do not find the floor object , create a new Floor object for this building instance
                            floor_list_instance = CompanyFloor(company_details_id=company_detail_instance,
                                                               floor_number=floor['floor_number'])

                            # append the floor instance to floor_list
                            floor_list.append(floor_list_instance)
                    floor_superset = floor_superset.union(floor_ids)

                # if CorporateCompanydetails was not updated , it was created fresh new, then create new floor objects
                else:
                    # for each floor
                    for floor in company_detail['listOfFloors']:
                        # create a new floor object
                        floor_list_instance = CompanyFloor(company_details_id=company_detail_instance,
                                                           floor_number=floor['floor_number'])
                        floor_list.append(floor_list_instance)

        # floor_list = []
        #         for floor in company_detail['listOfFloors']:
        #             floor_list_instance = CompanyFloor(company_details_id=company_detail_instance,floor_number=floor)
        #             floor_list.append(floor_list_instance)
        #         CompanyFloor.objects.bulk_create(floor_list)

        # CorporateCompanyDetails.objects.bulk_create(company_detail_list)
        CompanyFloor.objects.filter(id__in=floor_superset).delete()
        CorporateCompanyDetails.objects.filter(id__in=company_detail_ids).delete()
        # create all the floor objects at once
        CompanyFloor.objects.bulk_create(floor_list)
        return Response(status=200)


# Saving and fetching basic data of a salon.
class saveBasicSalonDetailsAPIView(APIView):
    def get(self, request, id, format=None):
        try:
            data1 = SupplierTypeSalon.objects.get(supplier_id=id)
            serializer = SupplierTypeSalonSerializer(data1)
            data2 = ContactDetailsGeneric.objects.filter(object_id=id)
            serializer1 = ContactDetailsGenericSerializer(data2, many=True)
            result = {'basicData': serializer.data, 'contactData': serializer1.data}
            return Response(result)
        except SupplierTypeSalon.DoesNotExist:
            return Response(status=404)
        except SupplierTypeSalon.MultipleObjectsReturned:
            return Response(status=406)

    def post(self, request, id, format=None):
        error = {}
        if 'supplier_id' in request.data:
            salon = SupplierTypeSalon.objects.filter(pk=request.data['supplier_id']).first()
            if salon:
                salon_serializer = SupplierTypeSalonSerializer(salon, data=request.data)
            else:
                salon_serializer = SupplierTypeSalonSerializer(data=request.data)
        if salon_serializer.is_valid():
            salon_serializer.save()
        else:
            error['message'] = 'Invalid Salon Info data'
            error = json.dumps(error)
            return Response(error, status=406)

        # Now saving contacts
        try:
            instance = SupplierTypeSalon.objects.get(supplier_id=id)
        except SupplierTypeSalon.DoesNotExist:
            return Response({'message': 'This salon does not exist'}, status=406)

        content_type = ContentType.objects.get_for_model(SupplierTypeSalon)

        contacts_ids = ContactDetailsGeneric.objects.filter(content_type=content_type,
                                                            object_id=instance.supplier_id).values_list('id', flat=True)
        contacts_ids = list(contacts_ids)

        for contact in request.data['contacts']:
            if 'id' in contact:
                contact_instance = ContactDetailsGeneric.objects.get(id=contact['id'])
                contacts_ids.remove(contact_instance.id)
                serializer = ContactDetailsGenericSerializer(contact_instance, data=contact)
                if serializer.is_valid():
                    serializer.save()
                else:
                    return Response(status=404)
            else:
                contact['object_id'] = instance.supplier_id
                serializer = ContactDetailsGenericSerializer(data=contact)
                if serializer.is_valid():
                    serializer.save(content_type=content_type)
                else:
                    return Response(status=404)

        ContactDetailsGeneric.objects.filter(id__in=contacts_ids).delete()
        return Response(status=200)
        # End of contact saving

# Saving and fetching basic data of a gym.
class saveBasicGymDetailsAPIView(APIView):

    def get(self, request, id, format=None):
        try:
            data1 = SupplierTypeGym.objects.get(supplier_id=id)
            serializer = SupplierTypeGymSerializer(data1)
            data2 = ContactDetailsGeneric.objects.filter(object_id=id)
            serializer1 = ContactDetailsGenericSerializer(data2, many=True)

            ownership_details_instance = OwnershipDetails.objects.filter(object_id=id).first()
            ownership_details_serializer = OwnershipDetailsSerializer(ownership_details_instance, many=False)
                
            result = {'basicData': serializer.data, 'contactData': serializer1.data, 'ownership_details':ownership_details_serializer.data}
            return Response(result)
        except SupplierTypeGym.DoesNotExist:
            return Response(status=404)
        except SupplierTypeGym.MultipleObjectsReturned:
            return Response(status=406)

    def post(self, request, id, format=None):
        error = {}
        class_name = self.__class__.__name__
        if 'supplier_id' in request.data:
            gym = SupplierTypeGym.objects.filter(pk=request.data['supplier_id']).first()
            if gym:
                gym_serializer = SupplierTypeGymSerializer(gym, data=request.data)
            else:
                gym_serializer = SupplierTypeGymSerializer(data=request.data)
        if gym_serializer.is_valid():
            gym_serializer.save()
        else:
            return handle_response(class_name, data='Invalid Gym Info data')

        # Now saving contacts
        try:
            instance = SupplierTypeGym.objects.get(supplier_id=id)
        except SupplierTypeGym.DoesNotExist:
            return Response({'message': 'This gym does not exist'}, status=406)

        content_type = ContentType.objects.get_for_model(SupplierTypeGym)

        contacts_ids = ContactDetailsGeneric.objects.filter(content_type=content_type,
                                                            object_id=instance.supplier_id).values_list('id', flat=True)
        contacts_ids = list(contacts_ids)

        for contact in request.data['contacts']:
            if 'id' in contact:
                contact_instance = ContactDetailsGeneric.objects.get(id=contact['id'])
                contacts_ids.remove(contact_instance.id)
                serializer = ContactDetailsGenericSerializer(contact_instance, data=contact)
                if serializer.is_valid():
                    serializer.save()
                else:
                    return Response(status=404)

            else:
                contact['object_id'] = instance.supplier_id
                serializer = ContactDetailsGenericSerializer(data=contact)
                if serializer.is_valid():
                    serializer.save(content_type=content_type)
                else:
                    return Response(status=404)

        ContactDetailsGeneric.objects.filter(id__in=contacts_ids).delete()

        ownership = request.data.get('ownership_details', None)

        if ownership:
            if ownership.get('id'):
                item1 = OwnershipDetails.objects.filter(pk=ownership['id']).first()
                ownership_serializer = OwnershipDetailsSerializer(item1, data=ownership)
            else:
                ownership_serializer = OwnershipDetailsSerializer(data=ownership)
            
            if ownership_serializer.is_valid():
                ownership_serializer.save()
        return Response(status=200)
        # End of contact Saving

class BusShelter(APIView):
    """
    The class provides api's for fetching and saving Bus Shelter details
    """

    def post(self, request, id):

        """
        API saves Bus Shelter details
        ---
        parameters:
        - name: supplier_type_code
          required: true
        - name: lit_status
        - name: halt_buses_count
        - name: name
        """

        # save the name of the class you are in to be useful for logging purposes
        class_name = self.__class__.__name__

        # check for supplier_type_code
        supplier_type_code = request.data.get('supplier_type_code')
        if not supplier_type_code:
            return handle_response(class_name, data='provide supplier type code', success=False)
        data = request.data.copy()
        data['supplier_id'] = id
        basic_details_response = save_basic_supplier_details(supplier_type_code, data)
        if not basic_details_response.data['status']:
            return basic_details_response

        basic_contacts = request.data.get('basic_contacts', None)

        if basic_contacts:
            for contact in basic_contacts:
                if 'id' in contact:
                    item = ContactDetails.objects.filter(pk=contact['id']).first()
                    contact_serializer = ContactDetailsSerializer(item, data=contact)
                else:
                    contact_serializer = ContactDetailsSerializer(data=contact)
                if contact_serializer.is_valid():
                    contact_serializer.save()

        ownership = request.data.get('ownership_details', None)

        if ownership:
            if ownership.get('id'):
                item1 = OwnershipDetails.objects.filter(pk=ownership['id']).first()
                ownership_serializer = OwnershipDetailsSerializer(item1, data=ownership)
            else:
                ownership_serializer = OwnershipDetailsSerializer(data=ownership)
            
            if ownership_serializer.is_valid():
                ownership_serializer.save()            
        
        return handle_response(class_name, data=basic_details_response.data['data'], success=True)

    def get(self, request):
        # fetch all and list Bus Shelters

        class_name = self.__class__.__name__

        try:
            user = request.user
            org_id = request.user.profile.organisation.organisation_id
            state = request.GET.get("state")
            state_name = request.GET.get("state_name")
            search = request.GET.get("search")
            bus_objects = []

            if user.is_superuser:
                bus_objects = SupplierTypeBusShelter.objects
                #sort list based on state selected.
                if state and state_name:
                    bus_objects = bus_objects.filter(Q(state=state) | Q(state=state_name))

                #sort list based on search keyword.
                if search:
                    bus_objects = bus_objects.filter(Q(name__icontains=search) | Q(address1__icontains=search) | Q(address2__icontains=search)
                                      | Q(city__icontains=search) | Q(supplier_id__icontains=search) | Q(supplier_code__icontains=search))

                bus_objects = bus_objects.all().order_by('name')
            else:
                vendor_ids = Organisation.objects.filter(created_by_org=org_id).values('organisation_id')
                bus_objects = SupplierTypeBusShelter.objects.filter((Q(representative__in=vendor_ids) | Q(representative=org_id))
                                                                      & Q(representative__isnull=False))
                #sort list based on state selected.
                if state and state_name:
                    bus_objects = bus_objects.filter(Q(state=state) | Q(state=state_name))

                #sort list based on search keyword.
                if search:
                    bus_objects = bus_objects.filter(Q(name__icontains=search) | Q(address1__icontains=search) | Q(address2__icontains=search)
                                      | Q(city__icontains=search) | Q(supplier_id__icontains=search) | Q(supplier_code__icontains=search))

                bus_objects = bus_objects.all().order_by('name')
                
            #pagination
            bus_objects_paginate = paginate(bus_objects,BusShelterSerializer,request)
            items = get_supplier_image(bus_objects_paginate["list"], 'Bus Shelter')
            # disabling pagination because search cannot be performed in whole data set.
            # paginator = PageNumberPagination()
            # result_page = paginator.paginate_queryset(items, request)
            # paginator_response = paginator.get_paginated_response(result_page)
            data = {
                'count': bus_objects_paginate["count"],
                'has_next':bus_objects_paginate["has_next"],
                'has_previous':bus_objects_paginate["has_previous"],
                'busshelters': items
            }
            return handle_response(class_name, data=data, success=True)
        except SupplierTypeBusShelter.DoesNotExist as e:
            return handle_response(class_name, data='Bus Shelter object does not exist', exception_object=e)
        except Exception as e:
            return handle_response(class_name, exception_object=e, request=request)
    

class getBusShelter(APIView):
    """
    The class provides api's for fetching and saving Bus Shelter details
    """

    def get(self, request, id):
        class_name = self.__class__.__name__
        try:
            retail_shop_instance = SupplierTypeBusShelter.objects.get(supplier_id=id)
            serializer = BusShelterSerializer(instance=retail_shop_instance)

            result = serializer.data

            retail_shop_instance = ContactDetails.objects.filter(object_id=id)
            contact_serializer = ContactDetailsSerializer(retail_shop_instance, many=True)
                
            result['contacData'] = contact_serializer.data

            ownership_details_instance = OwnershipDetails.objects.filter(object_id=id).first()
            ownership_details_serializer = OwnershipDetailsSerializer(ownership_details_instance, many=False)

            result['ownership_details'] = ownership_details_serializer.data

            return handle_response(class_name, data=result, success=True)
            
        except Exception as e:
            return handle_response(class_name, exception_object=e, request=request)


class BusShelterSearchView(APIView):
    """
    Searches particular bus shelters on the basis of search query
    """

    def get(self, request, format=None):
        """
        GET api fetches all bus shelters on search query. later it's implentation will be changed.
        """

        class_name = self.__class__.__name__
        try:
            user = request.user
            search_txt = request.query_params.get('search', None)
            if not search_txt:
                return handle_response(class_name, data='Search Text is not provided')
            items = SupplierTypeBusShelter.objects.filter(
                Q(supplier_id__icontains=search_txt) | Q(name__icontains=search_txt) | Q(
                    address1__icontains=search_txt) | Q(city__icontains=search_txt) | Q(
                    state__icontains=search_txt)).order_by('name')
            paginator = PageNumberPagination()
            result_page = paginator.paginate_queryset(items, request)
            serializer = SupplierTypeBusShelterSerializer(result_page, many=True)
            return paginator.get_paginated_response(serializer.data)
        except SupplierTypeBusShelter.DoesNotExist as e:
            return handle_response(class_name, data='Bus Shelter object does not exist', exception_object=e)
        except Exception as e:
            return handle_response(class_name, exception_object=e, request=request)


class SocietyAPIListView(APIView):
    def get(self, request):
        class_name = self.__class__.__name__
        try:
            user = request.user

            search_txt = request.query_params.get('search', None)
            if search_txt:
                society_objects = SupplierTypeSociety.objects.filter(
                    Q(supplier_id__icontains=search_txt) | Q(society_name__icontains=search_txt) | Q(
                        society_address1__icontains=search_txt) | Q(society_city__icontains=search_txt) | Q(
                        society_state__icontains=search_txt)).order_by('society_name')
            else:
                if user.is_superuser:
                    society_objects = SupplierTypeSociety.objects.all().order_by('society_name')
                else:
                    city_query = get_region_based_query(user, v0_constants.valid_regions['CITY'],
                                                        v0_constants.society)
                    society_objects = SupplierTypeSociety.objects.filter(city_query)

            # modify items to have society images data
            society_objects = get_supplier_image(society_objects, 'Society')
            paginator = PageNumberPagination()
            result_page = paginator.paginate_queryset(society_objects, request)
            paginator_response = paginator.get_paginated_response(result_page)
            data = {
                'count': len(society_objects),
                'societies': paginator_response.data
            }
            return handle_response(class_name, data=data, success=True)
        except SupplierTypeSociety.DoesNotExist:
            return Response(status=404)


class SuppliersMeta(APIView):
    """
    Fetches meta information about suppliers. How many are there in the system, count of each one of them.
    """

    def get(self, request):
        class_name = self.__class__.__name__
        try:
            valid_supplier_type_code_instances = SupplierTypeCode.objects.all()
            org_id= request.user.profile.organisation.organisation_id
            data = {}

            for instance in valid_supplier_type_code_instances:
                supplier_type_code = instance.supplier_type_code
                error = False
                try:
                    if request.user.is_superuser:
                        if instance.supplier_type_code == 'RS':
                            count = SupplierTypeSociety.objects.all().count()
                        else :
                            count = SupplierMaster.objects.filter(supplier_type = instance.supplier_type_code).count()
                    else:
                        vendor_ids = Organisation.objects.filter(created_by_org=org_id).values('organisation_id')
                        if instance.supplier_type_code == 'RS':
                            count = SupplierTypeSociety.objects.all().filter((Q(representative__in=vendor_ids) | Q(representative=org_id))
                                                                    & Q(representative__isnull=False)).count()
                        else :
                            count = SupplierMaster.objects.filter(Q(supplier_type = instance.supplier_type_code) & (Q(representative__in=vendor_ids) | Q(representative=org_id))
                                                                & Q(representative__isnull=False)).count()
                except Exception:
                    count = 0
                    error = True
               
                data[supplier_type_code] = {
                    'count': count,
                    'name': instance.supplier_type_name,
                    'error': error
                }
            return handle_response(class_name, data=data, success=True)

        except Exception :
            return handle_response(class_name, data='Something went wrong please try again later', success=False)


class SuppliersMetaData(APIView):
    """
    Fetches meta information about suppliers. Gives Count on the basis of selected state or state name and supplier type.
    """

    def get(self, request):
        """

        :param request:
        :return:
        """
        class_name = self.__class__.__name__
        try:
            data = {}
            supplier_type_code = request.query_params.get('supplier_code', None)

            if not supplier_type_code:
                return ui_utils.handle_response({}, data='supplier_type_code is Mandatory')
            
            model_name = get_model(supplier_type_code)

            state = request.GET.get("state")
            state_name = request.GET.get("state_name")
            user = request.user
            org_id = request.user.profile.organisation.organisation_id

            if user.is_superuser :
                if supplier_type_code == 'RS':
                    if state and state_name:
                        count = SupplierTypeSociety.objects.filter(Q(society_state=state) | Q(society_state=state_name)).count()
                else:
                    address_supplier = Prefetch('address_supplier',queryset=AddressMaster.objects.all())

                    if state_name:
                        count = SupplierMaster.objects.prefetch_related(address_supplier).filter(supplier_type=supplier_type_code,address_supplier__state__icontains=state_name).count()
                    else:
                        count = SupplierMaster.objects.prefetch_related(address_supplier).filter(supplier_type=supplier_type_code).count()
            else:
                if supplier_type_code == 'RS':
                    vendor_ids = Organisation.objects.filter(created_by_org=org_id).values('organisation_id')
                    society_objects = SupplierTypeSociety.objects.filter((Q(representative__in=vendor_ids) | Q(representative=org_id))
                                                                      & Q(representative__isnull=False))
                
                    if state and state_name:
                        count = society_objects.filter(Q(society_state=state) | Q(society_state=state_name)).count()
                
                else:
                    vendor_ids = Organisation.objects.filter(created_by_org=org_id).values('organisation_id')
                    supplier_objects = SupplierMaster.objects.filter((Q(representative__in=vendor_ids) | Q(representative=org_id)) & Q(representative__isnull=False))

                    address_supplier = Prefetch('address_supplier',queryset=AddressMaster.objects.all())

                    if state_name:
                        count = supplier_objects.prefetch_related(address_supplier).filter(supplier_type=supplier_type_code,address_supplier__state__icontains=state_name).count()
                    else:
                        count = supplier_objects.prefetch_related(address_supplier).filter(supplier_type=supplier_type_code).count()
            
            data[supplier_type_code] = {'count': count}

            return handle_response(class_name, data=data, success=True)   

        except Exception as e:
            return handle_response(class_name, exception_object=e, request=request)


class BusDepotViewSet(viewsets.ViewSet):
    """
    ViewSet around Bus depot
    """

    def create(self, request):
        class_name = self.__class__.__name__
        try:
            data = request.data
            serializer = BusDepotSerializer(data=data)
            if serializer.is_valid():
                serializer.save()
                return handle_response(class_name, data=serializer.data, success=True)
            return handle_response(class_name, data=serializer.errors)
        except Exception as e:
            return handle_response(class_name, exception_object=e, request=request)

    def list(self, request):
        class_name = self.__class__.__name__
        try:
            bus_depots = SupplierTypeBusDepot.objects.all()
            serializer = BusDepotSerializer(bus_depots, many=True)
            return handle_response(class_name, data=serializer.data, success=True)
        except Exception as e:
            return handle_response(class_name, exception_object=e, request=request)

    def update(self, request, pk=None):
        class_name = self.__class__.__name__
        try:
            instance = SupplierTypeBusDepot.objects.get(pk=pk)
            serializer = BusDepotSerializer(instance=instance, data=request.data)
            if serializer.is_valid():
                serializer.save()
                return handle_response(class_name, data=serializer.data, success=True)
            return handle_response(class_name, data=serializer.errors)
        except Exception as e:
            return handle_response(class_name, exception_object=e, request=request)

    def retrieve(self, request, pk=None):
        class_name = self.__class__.__name__
        try:
            instance = SupplierTypeBusDepot.objects.get(pk=pk)
            serializer = BusDepotSerializer(instance=instance)
            return handle_response(class_name, data=serializer.data, success=True)
        except Exception as e:
            return handle_response(class_name, exception_object=e, request=request)


class FilteredSuppliers(APIView):
    """
    This API gives suppliers based on different filters from mapView and gridView Page.
    Very expensive API. Use carefully.
    """

    def post(self, request):
        """
        The request looks like :
        {
          'supplier_type_code': 'CP',
          'common_filters': { 'latitude': 12, 'longitude': 11, 'radius': 2, 'quality': [ 'UH', 'H' ],'quantity': ['VL'] }
          'inventory_filters': ['PO', 'ST'],
          'amenities': [ ... ]
          'specific_filters': { 'real_estate_allowed': True, 'employees_count': {min: 10, max: 100},}
          'center_id': '23',
          'proposal_id': 'abc'
        }
        and the response looks like.:
        {
            "status": true,
            "data": {
                "suppliers": {
                    "RS": []
                },
                "suppliers_meta": {
                    "RS": {
                        "count": 0,
                        "inventory_count": {
                            "posters": null,
                            "stalls": null,
                            "standees": null,
                            "fliers": null
                        }
                    },
                }
            }
        }
        caching is done for three types of filters: for common filters, for inventory filters and for pi index filters.
        we are using hashlib.md5() or sha1() to get a hex digest to use it as a key. this is because the key in cache doest not
        allow spaces or underscores chars.
        """
        class_name = self.__class__.__name__
        try:
            # if not supplier type code, return
            supplier_type_code = request.data.get('supplier_type_code')
            if not supplier_type_code:
                return ui_utils.handle_response(class_name, data='provide supplier type code')

            # common filters are necessary

            common_filters = request.data['common_filters']  # maps to BaseSupplier Model or a few other models.
            inventory_filters = request.data.get('inventory_filters')  # maps to InventorySummary model
            priority_index_filters = request.data.get('priority_index_filters')  # maps to specific supplier table and are used in calculation of priority index
            proposal_id = request.data['proposal_id']
            center_id = request.data.get('center_id')

            # cannot be handled  under specific society filters because it involves operation of two columns in database which cannot be generalized to a query.
            # To get business name
            proposal = ProposalInfo.objects.get(proposal_id=proposal_id)
            organisation_name = proposal.account.organisation.name

            # get the right model and content_type
            supplier_model = ui_utils.get_model(supplier_type_code)
            response = ui_utils.get_content_type(supplier_type_code)
            if not response:
                return response
            content_type = response.data.get('data')

            # first fetch common query which is applicable to all suppliers. The results of this query will form
            # our master supplier list.
            response = website_utils.handle_common_filters(common_filters, supplier_type_code, proposal)
            if not response.data['status']:
                return response
            common_filters_query = response.data['data']

            # container to store inventory filters type of suppliers
            inventory_type_query_suppliers = []

            # this is the main list. if no filter is selected this is what is returned by default
            if supplier_type_code == 'RS':
                master_suppliers_list = set(list(supplier_model.objects.filter(common_filters_query).values_list('supplier_id', flat=True)))
            else :
                master_suppliers_list = set(list(SupplierMaster.objects.filter(common_filters_query).values_list('supplier_id', flat=True)))
            
            # now fetch all inventory_related suppliers
            # handle inventory related filters. it involves quite an involved logic hence it is in another function.
            response = website_utils.handle_inventory_filters(inventory_filters)
            if not response.data['status']:
                return response
            inventory_type_query = response.data['data']

            if inventory_type_query.__len__():
                inventory_type_query &= Q(content_type=content_type)
                inventory_type_query_suppliers = set(list(InventorySummary.objects.filter(inventory_type_query).values_list('object_id', flat=True)))

            # if inventory query was non zero in length, set final_suppliers_id_list to inventory_type_query_suppliers.
            if inventory_type_query.__len__():
                final_suppliers_id_list = inventory_type_query_suppliers
            else:
                final_suppliers_id_list = master_suppliers_list

            # when the final supplier list is prepared. we need to take intersection with master list.
            final_suppliers_id_list = final_suppliers_id_list.intersection(master_suppliers_list)

            
            result = {}

            if supplier_type_code == 'RS':
                filtered_suppliers = supplier_model.objects.filter(supplier_id__in=final_suppliers_id_list, representative=proposal.principal_vendor)
                supplier_serializer = ui_utils.get_serializer(supplier_type_code)
            else: 
                filtered_suppliers = SupplierMaster.objects.filter(supplier_type=supplier_type_code, supplier_id__in=final_suppliers_id_list)
                supplier_serializer = SupplierMasterSerializer
           
            serializer = supplier_serializer(filtered_suppliers, many=True)

            

            # to include only those suppliers that lie within radius, we need to send coordinates
            coordinates = {
                'radius': common_filters['radius'],
                'latitude': common_filters['latitude'],
                'longitude': common_filters['longitude']
            }
            
            # set initial value of total_suppliers. if we do not find suppliers which were saved initially, this is the final list
            initial_suppliers = website_utils.get_suppliers_within_circle(serializer.data, coordinates, supplier_type_code)

            
            # adding earlier saved shortlisted suppliers in the results for this center only
            shortlisted_suppliers = website_utils.get_shortlisted_suppliers_map(proposal_id, content_type, center_id)
            total_suppliers = website_utils.union_suppliers(initial_suppliers, shortlisted_suppliers.values())

            # because some suppliers can be outside the given radius, we need to recalculate list of
            # supplier_id's.
            final_suppliers_id_list = list(total_suppliers.keys())
            
            # We are applying ranking on combined list of previous saved suppliers plus the new suppliers if any. now the suppliers are filtered. we have to rank these suppliers. Get the ranking by calling this function.
            pi_index_map = {}
            supplier_id_to_pi_map = {}

            if priority_index_filters:
                # if you have provided pi filters only then a pi index map is calculated for each supplier
                pi_index_map = website_utils.handle_priority_index_filters(supplier_type_code, priority_index_filters, final_suppliers_id_list)
                supplier_id_to_pi_map = {supplier_id: detail['total_priority_index'] for supplier_id, detail in pi_index_map.items()}

            # the following function sets the pricing as before and it's temprorary.
            total_suppliers, suppliers_inventory_count = website_utils.set_pricing_temproray(list(total_suppliers.values()), final_suppliers_id_list, supplier_type_code, coordinates, supplier_id_to_pi_map)

            # before returning final result. change some keys of society to common keys we have defined.
            total_suppliers = website_utils.manipulate_object_key_values(total_suppliers, supplier_type_code=supplier_type_code)

            # construct the response and return
            # set the business name
            result['organisation_name'] = organisation_name

            # use this to show what kind of pricing we are using to fetch from pmd table for each kind of inventory
            result['inventory_pricing_meta'] = v0_constants.inventory_type_duration_dict_list

            # set total suppliers
            result['suppliers'] = {}
            result['suppliers'][supplier_type_code] = total_suppliers

            # set meta info about suppliers
            result['suppliers_meta'] = {}
            result['suppliers_meta'][supplier_type_code] = {}
            result['suppliers_meta'][supplier_type_code]['count'] = 0
            result['suppliers_meta'][supplier_type_code]['inventory_count'] = suppliers_inventory_count

            # send explanation separately
            result['suppliers_meta'][supplier_type_code]['pi_index_explanation'] = pi_index_map

            return ui_utils.handle_response(class_name, data=result, success=True)

        except Exception as e:
            logger.exception(e)
            return ui_utils.handle_response(class_name, exception_object=e, request=request)


class FilteredSuppliersAPIView(APIView):
    """
    This API gives suppliers based on different filters from mapView and gridView Page
    Currently implemented filters are locality and location (Standard, Medium High etc.)
    flat_count (100 - 250 etc.) flat_type(1BHK, 2BHK etc.)
    """

    def get(self, request, format=None):
        '''
        Response is in form
        {
            "status": true,
            "data": {
                "suppliers": {
                    "RS": []
                },
                "suppliers_meta": {
                    "RS": {
                        "count": 0,
                        "inventory_count": {
                            "posters": null,
                            "stalls": null,
                            "standees": null,
                            "fliers": null
                        }
                    }
                }
            }
        }

        lat -- latitude
        lng -- longitude
        r -- radius
        loc -- location params
        qlt -- supplier quality params example UH, HH, MH
        qnt -- supplier quantity params example LA, MD
        flc -- flat count
        inv -- inventory params example PO, ST, SL
        supplier_type_code -- RS, CP etc
        '''
        class_name = self.__class__.__name__
        try:

            response = {}

            latitude = request.query_params.get('lat', None)
            longitude = request.query_params.get('lng', None)
            radius = request.query_params.get('r', None)  # radius change
            location_params = request.query_params.get('loc', None)
            supplier_quality_params = request.query_params.get('qlt', None)
            supplier_quantity_params = request.query_params.get('qnt', None)
            flat_count = request.query_params.get('flc', None)
            flat_type_params = request.query_params.get('flt', None)
            inventory_params = request.query_params.get('inv', None)
            supplier_code = request.query_params.get('supplier_type_code')
            filter_query = Q()

            supplier_code = 'RS'

            if not supplier_code:
                return Response({'status': False, 'error': 'Provide supplier type code'},
                                status=status.HTTP_400_BAD_REQUEST)

            if not latitude or not longitude or not radius:
                return Response({'message': 'Please Provide longitude and latitude and radius as well'}, status=406)

            latitude = float(latitude)
            longitude = float(longitude)
            radius = float(radius)


            if flat_type_params:
                flat_types = []
                flat_type_params = flat_type_params.split()
                for param in flat_type_params:
                    try:
                        flat_types.append(flat_type_dict[param])
                    except KeyError:
                        pass
                if flat_types:
                    ''' We can improve performance here  by appending .distinct('supplier_id') when using postgresql '''
                    supplier_ids = set(FlatType.objects.filter(flat_type__in=flat_types).values_list('object_id', flat=True))
                    filter_query &= Q(supplier_id__in=supplier_ids)
                    # here to include those suppliers which don't have this info nothing can be done
                    # It is simply all the suppliers -> filter becomes useless
            delta_dict = website_utils.get_delta_latitude_longitude(radius, latitude)

            # add data to get min,max lat-long
            delta_dict['latitude'] = float(latitude)
            delta_dict['longitude'] = float(longitude)
            delta_dict['supplier_type_code'] = supplier_code

            # call the function to get min,max, lat,long data
            min_max_lat_long_response = website_utils.get_min_max_lat_long(delta_dict)
            if not min_max_lat_long_response.data['status']:
                return min_max_lat_long_response

            # make right params to call make_filter_query function
            filter_query_params = min_max_lat_long_response.data['data']
            filter_query_params['location_params'] = location_params
            filter_query_params['supplier_quality_params'] = supplier_quality_params
            filter_query_params['supplier_quantity_params'] = supplier_quantity_params
            filter_query_params['flat_count'] = flat_count
            filter_query_params['inventory_params'] = inventory_params
            filter_query_params['supplier_type_code'] = supplier_code

            # call function to make filter
            #todo: make filter function specific to supplier
            filter_query_response = website_utils.make_filter_query(filter_query_params)
            if not filter_query_response.data['status']:
                return filter_query_response
            filter_query &= filter_query_response.data['data']

            # get all suppliers  with all columns for the filter we constructed.
            suppliers = ui_utils.get_model(supplier_code).objects.filter(filter_query)

            serializer = ui_utils.get_serializer(supplier_code)(suppliers, many=True)

            suppliers = serializer.data
            supplier_ids = []
            suppliers_count = 0
            suppliers_data = []

            # iterate over all suppliers to  generate the final response
            for supplier in suppliers:

                supplier['supplier_latitude'] = supplier['society_latitude'] if supplier['society_latitude'] else supplier['latitude']
                supplier['supplier_longitude'] = supplier['society_longitude'] if supplier['society_longitude'] else supplier['longitude']

                if website_utils.space_on_circle(latitude, longitude, radius, supplier['supplier_latitude'],supplier['supplier_longitude']):
                    supplier_inventory_obj = InventorySummary.objects.get_supplier_type_specific_object(request.data.copy(),
                                                                                                        supplier['supplier_id'])
                    if supplier_inventory_obj:

                        supplier['shortlisted'] = True
                        supplier['buffer_status'] = False
                        adinventory_type_dict = ui_utils.adinventory_func()
                        duration_type_dict = ui_utils.duration_type_func()

                        if supplier_inventory_obj.poster_allowed_nb or supplier_inventory_obj.poster_allowed_lift:
                            supplier['total_poster_count'] = supplier_inventory_obj.total_poster_count
                            supplier['poster_price'] = return_price(adinventory_type_dict, duration_type_dict, 'poster_a4',
                                                                    'campaign_weekly')

                        if supplier_inventory_obj.standee_allowed:
                            supplier['total_standee_count'] = supplier_inventory_obj.total_standee_count
                            supplier['standee_price'] = return_price(adinventory_type_dict, duration_type_dict,
                                                                     'standee_small', 'campaign_weekly')

                        if supplier_inventory_obj.stall_allowed:
                            supplier['total_stall_count'] = supplier_inventory_obj.total_stall_count
                            supplier['stall_price'] = return_price(adinventory_type_dict, duration_type_dict, 'stall_small',
                                                                   'unit_daily')
                            supplier['car_display_price'] = return_price(adinventory_type_dict, duration_type_dict,
                                                                         'car_display_standard', 'unit_daily')

                        if supplier_inventory_obj.flier_allowed:
                            supplier['flier_frequency'] = supplier_inventory_obj.flier_frequency
                            supplier['filer_price'] = return_price(adinventory_type_dict, duration_type_dict,
                                                                   'flier_door_to_door', 'unit_daily')

                        # ADDNEW -->
                supplier_ids.append(supplier['supplier_id'])
                suppliers_data.append(supplier)
                suppliers_count += 1

                for society_key, actual_key in v0_constants.society_common_keys.items():
                    if society_key in supplier.keys():
                        value = supplier[society_key]
                        del supplier[society_key]
                        supplier[actual_key] = value

            suppliers_inventory_count = InventorySummary.objects.get_supplier_type_specific_objects({'supplier_type_code': supplier_code},
                                                                                                    supplier_ids).aggregate(posters=Sum('total_poster_count'), \
                                                                                                        standees=Sum('total_standee_count'),
                                                                                                        stalls=Sum('total_stall_count'),
                                                                                                        fliers=Sum('flier_frequency'))

            result = {}

            result['suppliers'] = {}
            result['suppliers_meta'] = {}

            result['suppliers']['RS'] = suppliers_data

            result['suppliers_meta']['RS'] = {}

            result['suppliers_meta']['RS']['count'] = suppliers_count
            result['suppliers_meta']['RS']['inventory_count'] = suppliers_inventory_count

            return ui_utils.handle_response(class_name, data=result, success=True)

            # response['suppliers'] = suppliers_data
            # response['supplier_inventory_count'] = suppliers_inventory_count
            # response['supplier_count'] = suppliers_count
            # response['supplier_type_code'] = supplier_code
            # return Response(response, status=200)
        except Exception as e:
            return Response({'status': False, 'error': e.message}, status=status.HTTP_400_BAD_REQUEST)


class ImportSocietyData(APIView):
    """
    This API reads a csv file and  makes supplier id's for each row. then it adds the data along with
    supplier id in the  supplier_society table. it also populates society_tower table.

    """
    def get(self, request):
        """
        :param request: request object
        :return: success response in case it succeeds else failure message.
        """
        class_name = self.__class__.__name__
        try:
            source_file = open(BASE_DIR + '/files/modified_new_tab.csv', 'rb')
            response = ui_utils.get_content_type(v0_constants.society)
            if not response.data['status']:
                return response
            content_type = response.data['data']

            with transaction.atomic():
                reader = csv.reader(source_file)
                for num, row in enumerate(reader):
                    data = {}
                    if num == 0:
                        continue
                    else:
                        # todo: city is not being saved in society.
                        if len(row) != len(v0_constants.supplier_keys):
                            return ui_utils.handle_response(class_name, data=errors.LENGTH_MISMATCH_ERROR.format(len(row), len(v0_constants.supplier_keys)))

                        for index, key in enumerate(v0_constants.supplier_keys):
                            if row[index] == '':
                                data[key] = None
                            else:
                                data[key] = row[index]

                        state_name = v0_constants.state_name
                        state_code = v0_constants.state_code
                        state_object = State.objects.get(state_name=state_name, state_code=state_code)
                        city_object = City.objects.get(city_code=data['city_code'], state_code=state_object)
                        area_object = CityArea.objects.get(area_code=data['area_code'], city_code=city_object)
                        subarea_object = CitySubArea.objects.get(subarea_code=data['subarea_code'],area_code=area_object)
                        # make the data needed to make supplier_id
                        supplier_id_data = {
                            'city_code': data['city_code'],
                            'area_code': data['area_code'],
                            'subarea_code': data['subarea_code'],
                            'supplier_type': data['supplier_type'],
                            'supplier_code': data['supplier_code']
                        }

                        data['supplier_id'] = get_supplier_id(supplier_id_data)
                        (society_object, value) = SupplierTypeSociety.objects.get_or_create(supplier_id=data['supplier_id'])
                        data['society_location_type'] = subarea_object.locality_rating
                        #data['society_state'] = 'Maharashtra'Uttar Pradesh
                        data['society_state'] = 'Haryana'
                        supplier_id = data['supplier_id']
                        society_object.__dict__.update(data)
                        society_object.save()

                        # make entry into PMD here.
                        ui_utils.set_default_pricing(supplier_id, data['supplier_type'])
                        towercount = SocietyTower.objects.filter(supplier=society_object).count()

                        # what to do if tower are less
                        tower_count_given = int(data['tower_count'])
                        if tower_count_given > towercount:
                            abc = tower_count_given - towercount
                            for i in range(abc):
                                tower = SocietyTower(supplier=society_object, object_id=supplier_id, content_type=content_type)
                                tower.save()
                        print("{0} done \n".format(data['supplier_id']))
            source_file.close()
            return Response(data="success", status=status.HTTP_200_OK)
        except ObjectDoesNotExist as e:
            return ui_utils.handle_response(class_name, data=e.args, exception_object=e)
        except KeyError as e:
            return ui_utils.handle_response(class_name, data=e.args, exception_object=e)
        except Exception as e:
            return ui_utils.handle_response(class_name, exception_object=e, request=request)


class ImportSupplierDataFromSheet(APIView):
    """
    The API tries to read a given sheet in a fixed format and updates the database with data in the sheet. This is used
    to enter data into the system from sheet. Currently Society data is supported, but the API is made generic.
    Please ensure following things before using this API:
       1. Events are defined in v0/constants.py. The event name should match with event names in headers.
       2. Amenities are already defined in database and their names must match with amenity names used in the headers.
       3. Check once the Flats constants defined and the headers. Both should match.
    NOTE:
        1.The API will not consider the item if it's marked 'NO' or 'N'. it will delete the appropriate item from the system if
        it's already present.
        2. The suppliers if not present in the system are created. if already created, they are updated.
        3. In response of the API, you can get to know how many suppliers were possible to be processed, how many of them
        got successfully created, how many were updated, how many were invalid because of issue in the row etc.
        4. An idea of positive or negative objects is introduced. Any object which has 'No' or 'N' in the sheet is qualified
        as negative. Only delete operation, if applicable is applied on such objects. Update and create operations are applied
        on positive objects.

    """
    def post(self, request):
        """
        Args:
            request:

        Returns:

            collect the basic society data in 'basic_data' key
            collect flat data in 'flats' key
            collect amenity data in 'amenities' key
            collect events data in  'events' key
            collect inventories and there pricing  data in 'inventories' key

        """
        class_name = self.__class__.__name__
        try:
            source_file = request.data['file']
            state_code = request.data['state_code']
            wb = openpyxl.load_workbook(source_file)
            ws = wb.get_sheet_by_name('supplier_data')
            supplier_type_code = request.data['supplier_type_code']
            data_import_type = request.data['data_import_type']

            # collects invalid row indexes
            invalid_rows_detail = {}
            possible_suppliers = 0
            new_cities_created = []
            new_areas_created = []
            new_subareas_created = []

            # will store all the info of society
            result = {}
            supplier_id_per_row = {}
            state_instance = State.objects.get(state_code=state_code)

            # put debugging information here
            invalid_rows_detail['detail'] = {}

            base_headers = ['city', 'city_code', 'area', 'area_code', 'sub_area', 'sub_area_code',  'supplier_code', 'supplier_name']
            # iterate through all rows and populate result array
            for index, row in enumerate(ws.iter_rows()):
                if index == 0:
                    website_utils.validate_supplier_headers(supplier_type_code, row, data_import_type)
                    continue
                possible_suppliers += 1
                supplier_id_per_row[index] = ''

                row_response = website_utils.get_mapped_row(ws, row)
                if not row_response.data['status']:
                    return row_response
                row_dict = row_response.data['data']

                # check for basic headers presence
                check = False
                for valid_header in base_headers:
                    if not row_dict[valid_header]:
                        invalid_rows_detail['detail'][index + 1] = '{0} not present in this row'.format(valid_header)
                        check = True

                # we do not proceed further if headers not valid
                if check:
                    continue

                city_code = row_dict['city_code'].strip()
                city_instance, is_created = City.objects.get_or_create(city_code=city_code, state_code=state_instance)
                city_instance.city_name = row_dict['city']
                city_instance.save()
                if is_created:
                    new_cities_created.append((city_code, city_instance.city_name))

                area_code = row_dict['area_code'].strip()
                area_instance,  is_created = CityArea.objects.get_or_create(area_code=area_code, city_code=city_instance)
                area_instance.label = row_dict['area']
                area_instance.save()
                if is_created:
                    new_areas_created.append((area_code, area_instance.label))

                subarea_code = str(row_dict['sub_area_code']).strip()
                sub_area_instance, is_created = CitySubArea.objects.get_or_create(subarea_code=subarea_code, area_code=area_instance)
                sub_area_instance.subarea_name = row_dict['sub_area']
                sub_area_instance.save()
                if is_created:
                    new_subareas_created.append((subarea_code, sub_area_instance.subarea_name))

                supplier_id = row_dict['city_code'].strip() + row_dict['area_code'].strip() + str(row_dict['sub_area_code']).strip() + supplier_type_code.strip() + row_dict['supplier_code'].strip()
                supplier_id_per_row[index] = supplier_id

                if supplier_id in result.keys():
                    invalid_rows_detail['detail'][index + 1] = '{0} supplier_id is duplicated in the sheet'.format(supplier_id)
                    continue

                result[supplier_id] = {
                    'common_data': {'state_name': state_code},
                    'flats': {'positive': {}, 'negative': {}},
                    'amenities': {'positive': {}, 'negative': {}},
                    'events': {'positive': {}, 'negative': {}},
                    'inventories': {'positive': {}, 'negative': {}}
                }
                result = website_utils.collect_supplier_common_data(result, supplier_type_code, supplier_id, row_dict, data_import_type)
                # result = website_utils.collect_amenity_data(result, supplier_id, row_dict)
                # result = website_utils.collect_events_data(result, supplier_id, row_dict)
                # result = website_utils.collect_flat_data(result, supplier_id, row_dict)

            input_supplier_ids = set(result.keys())
            if not input_supplier_ids:
                return ui_utils.handle_response(class_name, data=invalid_rows_detail)

            model = ui_utils.get_model(supplier_type_code)
            content_type = ui_utils.fetch_content_type(supplier_type_code)
            already_existing_supplier_ids = set(model.objects.filter(supplier_id__in=input_supplier_ids).values_list('supplier_id', flat=True))
            already_existing_supplier_count = len(already_existing_supplier_ids)

            new_supplier_ids = input_supplier_ids.difference(already_existing_supplier_ids)

            # create the supplier first here
            model.objects.bulk_create(model(**{'supplier_id': supplier_id}) for supplier_id in new_supplier_ids)

            supplier_instance_map = model.objects.in_bulk(input_supplier_ids)
            summary = website_utils.handle_supplier_data_from_sheet(result, supplier_instance_map, content_type, supplier_type_code)

            data = {
                'total_new_suppliers_made': len(new_supplier_ids),
                'total_valid_suppliers': len(input_supplier_ids),
                'total_invalid_suppliers': len(invalid_rows_detail['detail'].keys()),
                'total_existing_suppliers_out_of_valid_suppliers': already_existing_supplier_count,
                'total_possible_suppliers': possible_suppliers,
                'invalid_row_detail': invalid_rows_detail,
                'new_suppliers_created': new_supplier_ids,
                'already_existing_suppliers_from_this_sheet': already_existing_supplier_ids,
                'summary': summary,
                'supplier_id_per_row': supplier_id_per_row,
                'total_suppliers_in_sheet': len(supplier_id_per_row.keys()),
                'new_cities_created': new_cities_created,
                'new_areas_created': new_areas_created,
                'new_subareas_created': new_subareas_created

            }

            return ui_utils.handle_response(class_name, data=data, success=True)
        except Exception as e:
            print(e.message or e.args)
            return ui_utils.handle_response(class_name, exception_object=e, request=request)


class ImportContactDetails(APIView):
    """
    Saves contact details in db for each supplier.
    The API expects source file to be named as contacts.csv and should be placed in a folder called 'file'.
    """

    def get(self, request):
        class_name = self.__class__.__name__
        with transaction.atomic():

            file = open(BASE_DIR + '/files/contacts.csv', 'rb')
            try:
                reader = csv.reader(file)
                file.seek(0)
                for num, row in enumerate(reader):
                    if num == 0:
                        continue
                    else:
                        data = {}

                        length_of_row = len(row)
                        length_of_predefined_keys = len(v0_constants.contact_keys)
                        if length_of_row != length_of_predefined_keys:
                            return ui_utils.handle_response(class_name, data=errors.LENGTH_MISMATCH_ERROR.format(length_of_row, length_of_predefined_keys))

                        # make the data
                        for index, key in enumerate(v0_constants.contact_keys):
                            if row[index] == '':
                                data[key] = None
                            else:
                                data[key] = row[index]

                        if data.get('landline'):
                            landline_number = data['landline'].split('-')
                            data['landline'] = landline_number[1]
                            data['std_code'] = landline_number[0]

                        data['country_code'] = v0_constants.COUNTRY_CODE

                        try:
                            data['supplier_id'] = get_supplier_id(data)
                            society_object = SupplierTypeSociety.objects.get(supplier_id=data['supplier_id'])
                            data['spoc'] = False
                            data['supplier'] = society_object
                            contact_object = ContactDetails()
                            contact_object.__dict__.update(data)
                            contact_object.save()

                            # print it for universe satisfaction that something is going on !
                            print('{0} supplier contact done'.format(data['supplier_id']))
                        except ObjectDoesNotExist as e:
                            return ui_utils.handle_response(class_name, exception_object=e, request=request)
                        except Exception as e:
                            return ui_utils.handle_response(class_name, exception_object=e, request=request)

            finally:
                file.close()
            return ui_utils.handle_response(class_name, data='success', success=True)


class ImportSupplierData(APIView):
    """
    This API basically takes an excel sheet , process the data and saves it in the database.

    The request by which the sheet is made  is in form:
    [
         {
              center : { id : 1 , center_name: c1, ...   } ,
              suppliers: { 'RS' : [ { 'supplier_type_code': 'RS', 'status': 'R', 'supplier_id' : '1'}, {...}, {...} }
              suppliers_meta: {
                                 'RS': { 'inventory_type_selected' : [ 'PO', 'POST', 'ST' ]  },
                                 'CP': { 'inventory_type_selected':  ['ST']
              }
         }
    ]
    """

    def post(self, request, proposal_id=None):
        """
        Args:
            request: request param
            proposal_id: proposal_id

        Returns: Saves the  data in db
        """
        class_name = self.__class__.__name__
        try:

            if not request.FILES:
                return ui_utils.handle_response(class_name, data='No File Found')
            my_file = request.FILES['file']

            wb = openpyxl.load_workbook(my_file)

            # fetch all sheets
            all_sheets = wb.get_sheet_names()

            result = {}

            # iterate over multiple sheets
            for sheet in all_sheets:

                # fetch supplier_type_code from sheet name
                supplier_type_code = v0_constants.sheet_names_to_codes.get(sheet)
                if not supplier_type_code:
                    continue

                # fetch the worksheet object to work with
                ws = wb.get_sheet_by_name(sheet)

                # fetch all the center id's

                center_id_list_response = website_utils.get_center_id_list(ws, v0_constants.index_of_center_id)

                if not center_id_list_response.data['status']:
                    return center_id_list_response

                center_id_list = center_id_list_response.data['data']

                for index, center_id in enumerate(center_id_list):
                    if not result.get(center_id):
                        result[center_id] = {}

                # iterate through all rows and populate result array
                for index, row in enumerate(ws.iter_rows()):

                    if index == 0 or website_utils.is_empty_row(row):
                        continue

                    """
                      Number of rows in 'result' is number of distinct centers. each center has a center_id which is mapped
                      to an index in the 'result' list. each element of result list is a center_object.
                      Goal is to populate the right center_object with data of it's 3 keys:
                     'societies_inventory', 'societies', 'center'

                    """
                    # in order to proceed further we need a dict in which keys are header names with spaces
                    # removed and values are value of the row which we are processing

                    row_response = website_utils.get_mapped_row(ws, row)
                    if not row_response.data['status']:
                        return row_response
                    row = row_response.data['data']

                    # # get the center index mapped for this center_id
                    # center_index = center_id_to_index_mapping[int(row['center_id'])]

                    # # get the actual center_object from result list to process further
                    # center_object = result[center_index]

                    center_id = int(row['center_id'])

                    center_object = result[center_id]

                    # initialize the center_object  with necessary keys if not already
                    center_object = website_utils.initialize_keys(center_object, supplier_type_code)

                    # add 1 supplier that represents this row to the list of suppliers this object has already
                    center_object = website_utils.make_suppliers(center_object, row, supplier_type_code, proposal_id,
                                                                 center_id)

                    # add the 'center' data  in center_object
                    center_object = website_utils.make_center(center_object, row)

                    # update the center dict in result with modified center_object
                    result[center_id] = center_object

            # during an import, do not delete filters, but save whatever ShortlistedSpaces data. Don't touch filter data, just save spaces.
            website_utils.setup_create_final_proposal_post(result.values(), proposal_id,
                                                           delete_and_save_filter_data=False,
                                                           delete_and_save_spaces=True, exclude_shortlisted_space=True)

            # # data for this supplier is made. populate the shortlisted_inventory_details table before hitting the urls
            message = website_utils.populate_shortlisted_inventory_pricing_details(result, proposal_id, request.user)

            website_utils.import_proposal_cost_data(my_file,proposal_id)

            # if response.status_code != status.HTTP_200_OK:
            #     return Response({'status': False, 'error in import-metric-data api ': response.text},
            #                     status=status.HTTP_400_BAD_REQUEST)

            # prepare a new name for this file and save it in the required table
            file_name = website_utils.get_file_name(request.user, proposal_id, is_exported=False)

            # upload the file here to s3
            my_file.seek(0)  # rewind the file to start
            website_utils.upload_to_amazon(file_name, my_file)

            # fetch proposal instance and change it's status to 'finalized'.
            proposal = ProposalInfo.objects.get(proposal_id=proposal_id)
            proposal.campaign_state = v0_constants.proposal_finalized
            proposal.invoice_number = proposal.name.replace(" ", "")[:9].upper()
            proposal.save()

            return Response({'status': True, 'data': file_name}, status=status.HTTP_200_OK)
        except Exception as e:
            return ui_utils.handle_response(class_name, exception_object=e, request=request)


class SupplierSearch(APIView):
    """
    Generic API to perform simple search on predefined fields. The API will be slow because search uses
    basic icontains query. It can be made faster by making  FULLTEXT indexes on the db columns
    todo: attach user level permissions later
    """

    def get(self, request):

        class_name = self.__class__.__name__
       
        try:
            
            proposal_id = request.query_params.get('proposal_id')
            all_shortlisted_spaces = []
            if proposal_id:
                all_shortlisted_spaces = ShortlistedSpaces.objects.filter(proposal_id=proposal_id).values_list('object_id', flat=True)
                

            search_txt = request.query_params.get('search', None)
            supplier_type_code = request.query_params.get('supplier_type_code', None)
            vendor = request.query_params.get('vendor', None)


            if not supplier_type_code :
                return ui_utils.handle_response(class_name, data='provide supplier type code')

            
            if supplier_type_code == "RS" :

                model = ui_utils.get_model(supplier_type_code)
                search_query = Q()
                search_query &= Q(supplier_code__isnull=False)

                if search_txt :
                    search_text_query = Q()
                    for search_field in v0_constants.search_fields[supplier_type_code]:
                        if search_text_query:
                            search_text_query |= Q(**{search_field: search_txt})
                        else:
                            search_text_query = Q(**{search_field: search_txt})
                    search_query &= search_text_query

                if request.query_params.get("supplier_center"):
                    if search_query:
                        search_query &= Q(society_city__icontains=request.query_params.get("supplier_center"))
                    else:
                        search_query = Q(society_city__icontains=request.query_params.get("supplier_center"))
                
                if request.query_params.get("supplier_area"):
                    if search_query:
                        search_query &= Q(society_locality__icontains=request.query_params.get("supplier_area"))
                    else:
                        search_query = Q(society_locality__icontains=request.query_params.get("supplier_area"))
                
                if request.query_params.get("supplier_area_subarea"):
                    
                    if search_query:
                        search_query &= Q(society_subarea__contains = request.query_params.get("supplier_area_subarea"))
                    else:
                        search_query = Q(society_subarea__contains = request.query_params.get("supplier_area_subarea"))
                
                if vendor:
                    suppliers = model.objects.filter(search_query, representative=vendor)
                else:
                    suppliers = model.objects.filter(search_query)

                if len(all_shortlisted_spaces) > 0:
                    suppliers = suppliers.exclude(supplier_id__in=all_shortlisted_spaces)
                
                
            
                serializer_class = ui_utils.get_serializer(supplier_type_code)
                serializer = serializer_class(suppliers, many=True)
                suppliers = website_utils.manipulate_object_key_values_generic(serializer.data, supplier_type_code=supplier_type_code, **{'status': v0_constants.status})

                
                return ui_utils.handle_response(class_name, data=suppliers, success=True)
            else : 

                search_query = Q()
                search_query &= Q(supplier_type=supplier_type_code)

                if search_txt:
                    search_query &= (
                        Q(address_supplier__state__icontains=search_txt)
                        | Q(supplier_name__icontains=search_txt) 
                        | Q(address_supplier__address1__icontains=search_txt) 
                        | Q(address_supplier__address2__icontains=search_txt) 
                        | Q(address_supplier__city__icontains=search_txt) 
                    )
                
                if request.query_params.get("supplier_center"):
                    search_query &= Q(address_supplier__city__icontains = request.query_params.get("supplier_center"))
                    
                if request.query_params.get("supplier_area"):
                    search_query &= Q(address_supplier__area__icontains = request.query_params.get("supplier_area"))
                
                if request.query_params.get("supplier_area_subarea"):
                    search_query &= Q(address_supplier__subarea__icontains = request.query_params.get("supplier_area_subarea"))

                address_supplier = Prefetch('address_supplier',queryset=AddressMaster.objects.all())
                master_supplier_objects = SupplierMaster.objects.prefetch_related(address_supplier).filter(search_query).order_by('supplier_name')

                if len(all_shortlisted_spaces) > 0:
                    master_supplier_objects = master_supplier_objects.exclude(supplier_id__in=all_shortlisted_spaces)

                supplier_master_serializer = SupplierMasterSerializer(master_supplier_objects, many=True)


                suppliers = website_utils.manipulate_object_key_values_generic(supplier_master_serializer.data, supplier_type_code=supplier_type_code, **{'status': v0_constants.status})
                

                return ui_utils.handle_response(class_name, data=suppliers, success=True)
            
        except:
            return ui_utils.handle_response(class_name, data='something went wrong please try again later.')
        

class SupplierDetails(APIView):
    """
    Detail of individual supplier
    """
    def get(self, request):
        """
        Args:
            self:
            request:
        Returns: matching supplier object from supplier type model

        """
        class_name = self.__class__.__name__

        try:
            supplier_id = request.query_params['supplier_id']
            supplier_type_code = request.query_params['supplier_type_code']
            content_type = ui_utils.fetch_content_type(supplier_type_code)

            supplier_model = ContentType.objects.get(pk=content_type.id).model
            model = apps.get_model(settings.APP_NAME,supplier_model)

            supplier_object = model.objects.filter(supplier_id=supplier_id)
            data = {}
            if len(supplier_object) > 0:
                supplier_object = model.objects.get(supplier_id=supplier_id)
            else:
                try:
                    supplier_object = SupplierTypeRetailShop.objects.get(supplier_id=supplier_id)
                except Exception as e:
                    supplier_object = SupplySupplier.objects.raw({'_id': ObjectId(supplier_id)})
                    if supplier_object.count() > 0:
                        supplier = supplier_object[0]
                        data = {
                            "name": supplier.name,
                            "supplier_id": str(supplier._id)
                        }
                        for attr in supplier.supplier_attributes:
                            data[attr['name'].lower()] = attr['value'] if 'value' in attr else ''
            if not data:
                data = model_to_dict(supplier_object)
                data = website_utils.manipulate_object_key_values([data])[0]

            return ui_utils.handle_response(class_name, data=data, success=True)

        except Exception as e:
            return ui_utils.handle_response(class_name, exception_object=e, request=request)

    def put(self, request):
        """
        updates a particular supplier
        Args:
            request:
        Returns: returns updated supplier object

        """
        class_name = self.__class__.__name__
        try:
            supplier_id = request.data['supplier_id']
            supplier_type_code = request.data['supplier_type_code']

            data = request.data.copy()

            data.pop('supplier_id')
            data.pop('supplier_type_code')

            response = ui_utils.get_content_type(supplier_type_code)
            if not response.data['status']:
                return response
            content_type = response.data['data']
            supplier_model = content_type.model

            model = apps.get_model(settings.APP_NAME, supplier_model)

            model.objects.filter(supplier_id=supplier_id).update(**data)

            supplier_object = model.objects.get(pk=supplier_id)
            return ui_utils.handle_response(class_name, data=model_to_dict(supplier_object), success=True)

        except Exception as e:
            return ui_utils.handle_response(class_name, exception_object=e, request=request)


class MultiSupplierDetails(APIView):
    """
    Detail of multi supplier
    """
    def post(self, request):
        """
        get a particular supplier
        Args:
            request:
        Returns: returns supplier details

        """
        class_name = self.__class__.__name__
        try:
            organisation_id = get_user_organisation_id(request.user)
            # Visible only for machadalo users
            if organisation_id != 'MAC1421':
                return Response(data={"status": False, "error": "Permission Error"},
                                status=status.HTTP_400_BAD_REQUEST)
            supplier_ids = request.data.get('supplier_ids', None)
            supplier_type_code = request.data.get('supplier_type_code', 'RS')
            is_multiple_contact_number = request.data.get('is_multiple_contact_number', False)
            is_multiple_contact_name = request.data.get('is_multiple_contact_name', False)
            is_contact_name = request.data.get('is_contact_name', False)
            is_contact_number = request.data.get('is_contact_number', False)
            is_society = False

            if not supplier_ids or not supplier_type_code:
                return Response(data={'status': False, 'error': 'Missing supplier_ids or supplier_type_code'},
                                status=status.HTTP_400_BAD_REQUEST)
            try:
                response = ui_utils.get_content_type(supplier_type_code)
            except Exception as e:
                return Response(data={'status': False, 'error': "Please provide valid supplier type code"},
                                status=status.HTTP_400_BAD_REQUEST)
            content_type = response.data['data']
            supplier_model = content_type.model
            model = apps.get_model(settings.APP_NAME, supplier_model)

            if supplier_type_code == 'RS':
                is_society = True
                suppliers = model.objects.filter(pk__in=supplier_ids).values('society_name', 'society_locality',
                                                                             'society_city', 'society_subarea', 'society_state',
                                                                             'society_latitude','society_longitude','society_address1',
                                                                             'supplier_id', 'flat_count', 'society_type_quality')
            else:
                suppliers = model.objects.filter(pk__in=supplier_ids).values('name', 'area','city', 'subarea','state','latitude', 'longitude',
                                                                             'address1','supplier_id')
            # Get contact name & number
            contact_details = ContactDetails.objects.filter(object_id__in=supplier_ids).values('object_id', 'name', 'mobile', 'contact_type')
            multiple_supplier_details_with_contact = []
            supplier_list = []
            for supplier in suppliers:
                index = 0
                supplier_object = {
                    'supplier_type_code': supplier_type_code,
                    'supplier_id': supplier['supplier_id'],
                    'name': supplier['society_name'] if is_society else supplier['name'],
                    'area': supplier['society_locality'] if is_society else supplier['area'],
                    'city': supplier['society_city'] if is_society else supplier['city'],
                    'subarea': supplier['society_subarea'] if is_society else supplier['subarea'],
                    'latitude': supplier['society_latitude'] if is_society else supplier['latitude'],
                    'longitude': supplier['society_longitude'] if is_society else supplier['longitude'],
                    'state': supplier['society_state'] if is_society else supplier['state'],
                    'address': supplier['society_address1'] if is_society else supplier['address1'],
                    'flat_count': supplier['flat_count'] if is_society else None,
                    'society_type': supplier['society_type_quality'] if is_society else None,
                    'is_society': is_society if is_society else False
                }
                if contact_details:
                    for contact_detail in contact_details:
                        # For duplicate contact numbers
                        if not is_multiple_contact_name and is_multiple_contact_number and contact_detail['mobile'] is None:
                            continue
                        if not is_multiple_contact_number and is_multiple_contact_name and contact_detail['name'] is None:
                            continue
                        if contact_detail['object_id'] == supplier['supplier_id']:
                            supplier_object['contact_name'] = contact_detail['name'],
                            supplier_object['contact_number'] = contact_detail['mobile'],
                            supplier_object['contact_type'] = contact_detail['contact_type']
                            multiple_supplier_details_with_contact.append({
                                'id': index,
                                'supplier_type_code': supplier_type_code,
                                'supplier_id': supplier['supplier_id'],
                                'name': supplier['society_name'] if is_society else supplier['name'],
                                'area': supplier['society_locality'] if is_society else supplier['area'],
                                'city': supplier['society_city'] if is_society else supplier['city'],
                                'subarea': supplier['society_subarea'] if is_society else supplier['subarea'],
                                'latitude': supplier['society_latitude'] if is_society else supplier['latitude'],
                                'longitude': supplier['society_longitude'] if is_society else supplier['longitude'],
                                'state': supplier['society_state'] if is_society else supplier['state'],
                                'address': supplier['society_address1'] if is_society else supplier['address1'],
                                'contact_name': contact_detail['name'],
                                'contact_number': contact_detail['mobile'],
                                'contact_type': contact_detail['contact_type'],
                                'flat_count': supplier['flat_count'] if is_society else None,
                                'society_type': supplier['society_type_quality'] if is_society else None,
                                'is_society': is_society if is_society else False
                            })
                    index += 1
                if not is_contact_number and not is_contact_name:
                    supplier_list.append(supplier_object)
                if not is_contact_number and is_contact_name and (supplier_object['contact_name'] is not None or not supplier_object['contact_name']):
                    supplier_list.append(supplier_object)
                if not is_contact_name and is_contact_number and (supplier_object['contact_number'] is not None or not supplier_object['contact_number']):
                    supplier_list.append(supplier_object)
            if is_multiple_contact_name or is_multiple_contact_number:
                return Response(data={'status': True, 'data': multiple_supplier_details_with_contact}, status=status.HTTP_200_OK)
            else:
                return Response(data={'status': True, 'data': supplier_list}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response(data={'status': False, 'error': str(e)},
                            status=status.HTTP_400_BAD_REQUEST)


class addSupplierDirectToCampaign(APIView):
    """

    """
    def post(self, request):
        """

        :param request:
        :return:
        """
        class_name = self.__class__.__name__
        try:
            campaign_data = request.data
            campaign_data['is_import_sheet'] = False
            center = ProposalCenterMapping.objects.filter(proposal=campaign_data['campaign_id'])[0]
            campaign = ProposalInfo.objects.get(pk=campaign_data['campaign_id'])

            for supplier_code in campaign_data['center_data']:
                response = website_utils.save_shortlisted_suppliers_data(center, supplier_code, campaign_data, campaign)
                if not response.data['status']:
                    return response
                response = website_utils.save_shortlisted_inventory_pricing_details_data(center, supplier_code,
                                                            campaign_data, campaign, create_inv_act_data=True)

            response = website_utils.prepare_shortlisted_spaces_and_inventories(campaign_data['campaign_id'], 1, '', 0, '', '', '', '', '', '')                                                            
            return ui_utils.handle_response(class_name, data=response.data['data']['shortlisted_suppliers'], success=True)
        except Exception as e:
            return ui_utils.handle_response(class_name, exception_object=e, request=request)


class deleteSuppliers(APIView):
    # delete supplier if it is not shortlisted
    @staticmethod
    def delete(request):
        try:
            supplier_id_array = request.data
            remaining_suppliers = get_values(SupplierTypeSociety.objects.filter(supplier_id__in=supplier_id_array).\
                values('supplier_id'),'supplier_id')
            shortlisted_query = ShortlistedSpaces.objects.filter(object_id__in=remaining_suppliers).\
                values('object_id').distinct()
            unremoved_suppliers = []
            for supplier in shortlisted_query:
                unremoved_suppliers.append(supplier['object_id'])
                remaining_suppliers.remove(supplier['object_id'])

            for supplier_id in remaining_suppliers:
                tower_ids = get_values(SocietyTower.objects.filter(object_id=supplier_id).values('tower_id'),'tower_id')
                StandeeInventory.objects.filter(tower_id__in=tower_ids).delete()
                LiftDetails.objects.filter(tower_id__in=tower_ids).delete()
                SocietyTower.objects.filter(object_id=supplier_id).delete()
                PriceMappingDefault.objects.filter(object_id=supplier_id).delete()
                PosterInventory.objects.filter(object_id=supplier_id).delete()
                InventorySummary.objects.filter(object_id=supplier_id).delete()
                FlatType.objects.filter(object_id=supplier_id).delete()
                SupplierTypeSociety.objects.filter(supplier_id=supplier_id).delete()

            print("Suppliers in shortlisted spaces (not deleted)", unremoved_suppliers)

            return ui_utils.handle_response({}, data='success', success=True)
        except Exception as e:
            return ui_utils.handle_response({}, exception_object=e, request=request)


class deleteShortlistedSpaces(APIView):
    # delete shortlisted_spaces (with pricing) if there is no assocaited image
    @staticmethod
    def post(request):
        try:
            shortlisted_spaces_id_array = request.data
            remaining_spaces = get_values(ShortlistedSpaces.objects.filter(id__in=shortlisted_spaces_id_array).\
                values('id'),'id')
            spaces_containing_images = []
            for space_id in remaining_spaces:
                inventory_activity_images = InventoryActivityImage.objects.filter(inventory_activity_assignment__inventory_activity__shortlisted_inventory_details_id=space_id).all()
                if len(inventory_activity_images) > 0:
                    spaces_containing_images.append(space_id)
                    remaining_spaces.remove(space_id)
            for shortlisted_space_id in remaining_spaces:
                ShortlistedInventoryPricingDetails.objects.filter(shortlisted_spaces_id=shortlisted_space_id).delete()
                ShortlistedSpaces.objects.filter(id=shortlisted_space_id).delete()
            print("ShortlistedSpaces containing images (not deleted)", spaces_containing_images)
            return ui_utils.handle_response({}, data='success', success=True)
        except Exception as e:
            return ui_utils.handle_response({}, exception_object=e, request=request)


class insertFlatCountType(APIView):
    # inserts flat count type to society on the basis of number of flats it has
    @staticmethod
    def post(request):
        existing_data = SupplierTypeSociety.objects.all()
        n = 0
        for curr_data in existing_data:
            flat_count = curr_data.flat_count
            if flat_count is None:
                continue
            if flat_count<150:
                flat_type = '1-150'
            elif flat_count<400:
                flat_type = '151-400'
            else:
                flat_type = '401+'
            curr_data.flat_count_type = flat_type
            print(curr_data)
            curr_data.save()
            n=n+1
            print("saved: ", n)
        return ui_utils.handle_response({}, data='success', success=True)


class listCampaignSuppliers(APIView):
    # inserts flat count type to society on the basis of number of flats it has
    @staticmethod
    def get(request, campaign_id):

        all_shortlisted_supplier = ShortlistedSpaces.objects.filter(proposal_id=campaign_id)
        if request.user.profile.name == "Intern":
            assigned_suppliers = SupplierAssignment.objects.filter(campaign=campaign_id, assigned_to=request.user).values_list("supplier_id", flat=True)
            all_shortlisted_supplier = all_shortlisted_supplier.filter(object_id__in=assigned_suppliers)
        all_shortlisted_supplier = all_shortlisted_supplier.values('proposal_id', 'object_id', 'phase_no_id', 'is_completed', 'proposal__name',
                'proposal__tentative_start_date',
                'proposal__tentative_end_date', 'proposal__campaign_state','supplier_code')
        all_supplier_ids = [supplier['object_id'] for supplier in all_shortlisted_supplier if supplier['supplier_code'] == "RS"]
        all_campaign_societies = SupplierTypeSociety.objects.filter(supplier_id__in=all_supplier_ids).all()
        serializer = SupplierTypeSocietySerializer(all_campaign_societies, many=True)

        all_supplier_master_ids = [supplier['object_id'] for supplier in all_shortlisted_supplier if supplier['supplier_code'] != "RS"]
        all_campaign_master_societies = SupplierMaster.objects.filter(supplier_id__in=all_supplier_master_ids).all()
        master_serializer = SupplierMasterSerializer(all_campaign_master_societies, many=True)

        all_societies = manipulate_object_key_values(serializer.data)
        master_suppliers = manipulate_master_to_rs(master_serializer.data)
        all_societies.extend(master_suppliers)

        booking_inv_activities = BookingInventoryActivity.objects.raw({"campaign_id": campaign_id})
        supplier_ids = [ObjectId(supplier.supplier_id) for supplier in booking_inv_activities]
        suppliers = SupplySupplier.objects.raw({'_id': {'$in': (supplier_ids)}})
        dynamic_suppliers = []
        if suppliers.count() > 0:
            for supplier in suppliers:
                data = {
                    "name": supplier.name,
                    "supplier_id": str(supplier._id)
                }
                for attr in supplier.supplier_attributes:
                    data[attr['name']] = attr['value'] if 'value' in attr else None
                dynamic_suppliers.append(data)
        all_societies = all_societies + dynamic_suppliers
        all_societies = [dict(society) for society in all_societies]

        return ui_utils.handle_response({}, data=all_societies, success=True)

class ListCampaignSuppliers(APIView):
    # inserts flat count type to society on the basis of number of flats it has
    @staticmethod
    def get(request, campaign_id):
        all_shortlisted_supplier = ShortlistedSpaces.objects.filter(proposal_id=campaign_id). \
            values('proposal_id', 'object_id', 'phase_no_id', 'is_completed', 'proposal__name',
                   'proposal__tentative_start_date',
                   'proposal__tentative_end_date', 'proposal__campaign_state')
        all_supplier_ids = [supplier['object_id'] for supplier in all_shortlisted_supplier]
        all_campaign_societies = SupplierMaster.objects.filter(supplier_id__in=all_supplier_ids).all()
        serializer = SupplierMasterSerializer(all_campaign_societies, many=True)
        all_societies = manipulate_object_key_values(serializer.data)
        booking_inv_activities = BookingInventoryActivity.objects.raw({"campaign_id": campaign_id})
        supplier_ids = [ObjectId(supplier.supplier_id) for supplier in booking_inv_activities]
        suppliers = SupplySupplier.objects.raw({'_id': {'$in': (supplier_ids)}})
        dynamic_suppliers = []
        if suppliers.count() > 0:
            for supplier in suppliers:
                data = {
                    "name": supplier.name,
                    "supplier_id": str(supplier._id)
                }
                for attr in supplier.supplier_attributes:
                    data[attr['name']] = attr['value'] if 'value' in attr else None
                dynamic_suppliers.append(data)
        all_societies = all_societies + dynamic_suppliers
        all_societies = [dict(society) for society in all_societies]
        return ui_utils.handle_response({}, data=all_societies, success=True)

class CreateSupplierPriceMappingObjects(APIView):
    def get(self, request):
        suppliers = SupplierTypeSociety.objects.all()
        pmd_object = PriceMappingDefault.objects.all()
        suppliers_no_pmd = []
        for supplier in suppliers:
            if len(pmd_object.filter(object_id=supplier.supplier_id)) == 0:
                suppliers_no_pmd.append(supplier)
        rs_content_type = get_content_type('RS').data['data']
        print(len(suppliers_no_pmd))
        index = 0
        for supplier in suppliers_no_pmd:
            index = index + 1
            print(supplier.society_name, index)
            new_society = supplier
            supplier_id = supplier.supplier_id

            try:
                create_price_mapping_default('7', "POSTER", "A4", new_society,
                                             0, rs_content_type, supplier_id)
            except Exception as e:
                pass
            try:
                create_price_mapping_default('0', "POSTER LIFT", "A4", new_society,
                                             0, rs_content_type, supplier_id)
            except Exception as e:
                pass
            try:
                create_price_mapping_default('0', "STANDEE", "Small", new_society,
                                             0, rs_content_type, supplier_id)
            except Exception as e:
                pass
            try:
                create_price_mapping_default('1', "STALL", "Small", new_society,
                                             0, rs_content_type, supplier_id)
            except Exception as e:
                pass
            try:
                create_price_mapping_default('0', "CAR DISPLAY", "A4", new_society,
                                             0, rs_content_type, supplier_id)
            except Exception as e:
                pass
            try:
                save_flyer_locations(0, 1, new_society, 'RS')
            except Exception as e:
                pass
            try:
                create_price_mapping_default('1', "FLIER", "Door-to-Door", new_society,
                                             0, rs_content_type, supplier_id)
            except Exception as e:
                pass
        return ui_utils.handle_response({}, data={}, success=True)

class GetLocationDataInSheet(APIView):
    @staticmethod
    def get(request):
        suppliers = SupplierTypeSociety.objects.all()
        serializer = SupplierTypeSocietySerializer(suppliers, many=True)
        data = serializer.data
        header_list = ['Index', 'Society Name', 'Supplier Id', 'State', 'City', 'Area', 'Subarea']
        for supplier in data:
            book = Workbook()
            sheet = book.active
            sheet.append(header_list)
            index = 0
            for supplier in data:
                index = index + 1
                supplier_data = []

                supplier_data.append(index)
                supplier_data.append(supplier['society_name'])
                supplier_data.append(supplier['supplier_id'])
                supplier_data.append(supplier['society_state'])
                supplier_data.append(supplier['society_city'])
                supplier_data.append(supplier['society_locality'])
                supplier_data.append(supplier['society_subarea'])
                sheet.append(supplier_data)
            resp = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            resp['Content-Disposition'] = 'attachment; filename=mydata.xlsx'
            book.save(resp)
            return resp
        return ()


class SupplierGenericViewSet(viewsets.ViewSet):
    """
    View Set around All Suppliers
    """ 

    def list(self, request):
        class_name = self.__class__.__name__
        try:  
            user = request.user
            org_id = request.user.profile.organisation.organisation_id
            retail_shop_objects = []
            supplier_type_code = request.GET.get("supplier_type_code", None)
            state = request.GET.get("state")
            state_name = request.GET.get("state_name")
            search = request.GET.get("search")

            if not supplier_type_code:
                return ui_utils.handle_response({}, data='supplier_type_code is Mandatory')

            search_query = Q()
            search_query &= Q(supplier_type=supplier_type_code)

            if search:
                search_query &= (
                    Q(address_supplier__state__icontains=search)
                    | Q(address_supplier__address1__icontains=search) 
                        | Q(address_supplier__address1__icontains=search) 
                    | Q(address_supplier__address1__icontains=search) 
                    | Q(address_supplier__address2__icontains=search) 
                        | Q(address_supplier__address2__icontains=search) 
                    | Q(address_supplier__address2__icontains=search) 
                    | Q(address_supplier__city__icontains=search) 
                        | Q(address_supplier__city__icontains=search) 
                    | Q(address_supplier__city__icontains=search) 
                    | Q(supplier_id=search)
                    | Q(supplier_name__icontains=search)
                )
            master_supplier_objects = SupplierMaster.objects
            if not user.is_superuser:
                vendor_ids = Organisation.objects.filter(created_by_org=org_id).values('organisation_id')
                search_query &= ((Q(representative__in=vendor_ids) | Q(representative=org_id)) & Q(representative__isnull=False))

            address_supplier = Prefetch('address_supplier',queryset=AddressMaster.objects.all())
            master_supplier_objects = master_supplier_objects.prefetch_related(address_supplier).filter(search_query).order_by('supplier_name')

            if state_name:
                master_supplier_objects = master_supplier_objects.filter(supplier_type=supplier_type_code,address_supplier__state__icontains=state_name)
            
            master_supplier_objects = master_supplier_objects.all().order_by('supplier_name')
        
            supplier_objects_paginate = paginate(master_supplier_objects, SupplierMasterSerializer, request)
            supplier_with_images = get_supplier_image(supplier_objects_paginate["list"], 'Supplier')

            data = {
                'count': supplier_objects_paginate["count"],
                'has_next': supplier_objects_paginate["has_next"],
                'has_previous': supplier_objects_paginate["has_previous"],
                'supplier_objects': supplier_with_images
            }
            return handle_response(class_name, data=data, success=True)
        except Exception as e:
            return handle_response(class_name, data="Something went wrong please try again later.", request=request)

    def retrieve(self, request, pk):
        class_name = self.__class__.__name__

        try:
            supplier_type_code = request.GET.get("supplier_type_code", None)
            
            if not supplier_type_code:
                return ui_utils.handle_response({}, data='supplier_type_code is Mandatory')

            master_supplier_objects = SupplierMaster.objects.filter(pk=pk).first()
            supplier_master_serializer = SupplierMasterSerializer(master_supplier_objects, many=False)

            model_name = get_model(supplier_type_code)
            supplier_instance = model_name.objects.filter(pk=pk).first()

            if supplier_instance:
                serializer_name = get_serializer(supplier_type_code)
                serializer = serializer_name(instance=supplier_instance)
                shopData = serializer.data
                shopData['master_data'] = supplier_master_serializer.data
                shopData['contactData'],shopData['ownership_details'] = retrieve_contact_and_ownership_detail(pk)
                return handle_response(class_name, data=shopData, success=True)
            else:
                return handle_response(class_name, data="Supplier Id does not exist.", success=True)
        except Exception as e:
            return handle_response(class_name, data="Something went wrong please try again later.", request=request)

    def update(self, request, pk):
        class_name = self.__class__.__name__
        try:
            supplier_type_code = request.data.get('supplier_type_code', None)

            if not supplier_type_code:
                return handle_response({}, data='supplier_type_code is Mandatory')

            model_name = get_model(supplier_type_code)
            serializer_name = get_serializer(supplier_type_code)
            instance = model_name.objects.get(pk=pk)
            
            serializer = serializer_name(instance=instance, data=request.data)
            if serializer.is_valid():
                serializer.save()
                update_contact_and_ownership_detail(request.data)
                return handle_response(class_name, data=serializer.data, success=True)
            return handle_response(class_name, data=serializer.data, success=True)
        except Exception as e:
            return handle_response(class_name, exception_object=e, request=request)
class SocietySupplierRelationship(APIView):
    @staticmethod
    def post(request):
        try:
            supplier_type = request.query_params.get('type', 'RE')
            society_id = request.query_params.get('society_id', None)
            suppliers = request.data.get('suppliers', None)
            if not supplier_type:
                return ui_utils.handle_response({}, data={'message': 'Missing supplier type'}, success=False)
            if not suppliers:
                return ui_utils.handle_response({}, data={'message': 'Please provide data'}, success=False)
            for supplier in suppliers:
                SupplierRelationship(society_id=society_id, supplier_id=supplier['supplier_id'], supplier_type=supplier_type, type=supplier['type']).save()
            return ui_utils.handle_response({}, data='Supplier added successfully', success=True)
        except Exception as e:
            logger.exception(e)
            return ui_utils.handle_response({}, exception_object=e)

    @staticmethod
    def get(request):
        try:
            society_id = request.query_params.get('society_id', None)
            supplier_type = request.query_params.get('supplier_type', 'RE')
            type = request.query_params.get('type', None)
            if not society_id and not type:
                return Response(data={"status": False, "error": "Missing supplier_id and type"},
                                status=status.HTTP_400_BAD_REQUEST)
            supplier_list = SupplierRelationship.objects.filter(society_id=society_id, type=type)
            supplier_model = get_model(supplier_type)
            supplier_ids = [supplier.supplier_id for supplier in supplier_list]
            if supplier_type == 'RS':
                supplier_details = supplier_model.objects.filter(supplier_id__in=supplier_ids).values('society_name', 'supplier_id')
            else:
                supplier_details = supplier_model.objects.filter(supplier_id__in=supplier_ids).values('name','supplier_id')
            return ui_utils.handle_response({}, data=supplier_details, success=True)
        except Exception as e:
            logger.exception(e)
            return ui_utils.handle_response({}, exception_object=e)

class ListSuppliers(APIView):
    @staticmethod
    def get(request):
        try:
            type = request.query_params.get('type', 'RE')
            supplier_id = request.query_params.get('supplier_id', None)
            supplier_type = request.query_params.get('supplier_type', 'RS')
            city = None
            area = None
            if not supplier_id:
                return Response(data={"status": False, "error": "Missing supplier_id"}, status=status.HTTP_400_BAD_REQUEST)
            # Get area, subarea using supplier id
            if supplier_type == 'RS':
                society_details = SupplierTypeSociety.objects.filter(supplier_id=supplier_id).values('society_city','society_locality')
                if len(society_details) > 0:
                    city = society_details[0]['society_city']
                    area = society_details[0]['society_locality']
            else:
                supplier_model = get_model(supplier_type)
                supplier_details = supplier_model.objects.filter(supplier_id=supplier_id).values('area', 'city')
                if len(supplier_details) > 0:
                    city = supplier_details[0]['city']
                    area = supplier_details[0]['area']
            if not city and not area:
                return Response(data={"status": False, "error": "Missing paramaters city, area"}, status=status.HTTP_400_BAD_REQUEST)
            model = get_model(type)
            if type == 'RS':
                supplier_list = model.objects.filter(society_city__icontains=city, society_locality__icontains=area).values('society_name','society_locality','society_subarea',
                                                                                          'society_city', 'supplier_id')
            else:
                supplier_list = model.objects.filter(city__icontains=city, area__icontains=area).values('name','supplier_id')
            return ui_utils.handle_response({}, data=supplier_list, success=True)
        except Exception as e:
            logger.exception(e)
            return ui_utils.handle_response({}, exception_object=e)

class AssignSupplierUsers(APIView):
    def put(selt, request):
        quality_rating = request.data.get("quality_rating")
        assigned_to_ids = request.data.get("assigned_to_ids")
        campaign_id = request.data.get("campaign_id")
        supplier_ids = ShortlistedSpaces.objects.filter(proposal_id=campaign_id).values_list("object_id", flat=True)
        supplier_new_ids = SupplierMaster.objects.filter(supplier_id__in=supplier_ids, quality_rating=quality_rating).values_list("supplier_id", flat=True)
        if not supplier_new_ids:
            supplier_new_ids = SupplierTypeSociety.objects.filter(supplier_id__in=supplier_ids, society_type_quality=quality_rating).values_list("supplier_id", flat=True)
        
        user_old = SupplierAssignment.objects.filter(campaign=campaign_id, supplier_id__in=supplier_new_ids).values("supplier_id","assigned_to")
        user_old_obj = {}
        for row in user_old:
            if not user_old_obj.get(row["supplier_id"]):
                user_old_obj[row["supplier_id"]] = []
            user_old_obj[row["supplier_id"]].append(row["assigned_to"])

        data = []
        now_time = timezone.now()
        for supplier_id in supplier_new_ids:
            for user_id in assigned_to_ids:
                if not user_old_obj.get(supplier_id) or not user_id in user_old_obj[supplier_id]:
                    data.append(SupplierAssignment(**{
                        "campaign_id": campaign_id,
                        "supplier_id": supplier_id,
                        "assigned_by": request.user,
                        "assigned_to_id": user_id,
                        "created_at": now_time,
                        "updated_at": now_time
                    }))

        SupplierAssignment.objects.bulk_create(data)
        return ui_utils.handle_response({}, data={}, success=True)

import os, sys

class UpdateSupplierDataImport(APIView):
    def post(self, request):
        try:
            data = request.data
            updated_contact_file = open(os.path.join(sys.path[0], "updated_supplier.txt"), "a")
            new_contact_file = open(os.path.join(sys.path[0], "new_contact.txt"), "a")
            for supplier in data:
                supplier_id = supplier['supplier_id']
                contact_number = supplier['contact_number']
                contact_name = supplier['contact_name'].title()
                name = supplier['name']	
                flat_count = supplier['flat_count']	
                society_type = supplier['society_type']	
                latitude = supplier['latitude']	
                longitude = supplier['longitude']	
                area = supplier['area']	
                subarea = supplier['subarea']
                city = supplier['city']	
                address = supplier['address']	
                contact_type = supplier['contact_type']
                if contact_type is None:
                    contact_type = 'Committee Member'
                try:
                    if supplier_id:
                        supplier_data = SupplierTypeSociety.objects.filter(supplier_id=supplier_id)
                        if supplier_data:
                            supplier_data.update(society_name=name, society_city=city, society_longitude=longitude,
                            society_latitude=latitude, flat_count=flat_count, society_type_quality=society_type,
                            society_locality=area, society_subarea=subarea, society_address1=address )
                    
                    if contact_number and supplier_id:
                        # Check contact number in contact details
                        contact_details = ContactDetails.objects.filter(mobile=contact_number, object_id=supplier_id)
                        if contact_details:
                            if contact_name and contact_type:
                                contact_details.update(name=contact_name, contact_type=contact_type)
                                updated_contact_file.write(
                                    "{mobile},{name},{id}\n".format(mobile=contact_number, name=contact_name, id=supplier_id))
                        else:
                            contact_details = ContactDetails(object_id=supplier_id, name=contact_name, contact_type=contact_type, mobile=contact_number)
                            contact_details.save()
                            new_contact_file.write(
                                "{mobile},{name},{id}\n".format(mobile=contact_number, name=contact_name, id=supplier_id))
                except Exception as e:
                    print(e)
        except Exception as e:
            print(e)
        return handle_response({}, data='Data updated Successfully', success=True)


class DownloadExcel(ListView):
    permission_classes = ()
    authentication_classes = ()

    def get(self, request):
        wb = load_workbook('files/export_supplier.xlsx')
        resp = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = 'attachment; filename=files/export_supplier.xlsx'
        wb.save(resp)
        return resp