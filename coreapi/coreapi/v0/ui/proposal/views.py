from __future__ import print_function
from __future__ import absolute_import
from rest_framework.views import APIView
from rest_framework import viewsets, status
from django.db.models import Count, Sum, F
from django.contrib.contenttypes.models import ContentType
from django.core.exceptions import ObjectDoesNotExist
from django.utils import timezone
from pygeocoder import Geocoder, GeocoderError
from openpyxl import load_workbook
from requests.exceptions import ConnectionError
from celery.result import GroupResult
import shutil
import hashlib
import dateutil.parser
import os
import pytz
import v0.permissions as v0_permissions
from django.conf import settings
import datetime
import time
import random
import string
from .models import ProposalInfo, ProposalCenterMapping, ProposalCenterSuppliers
from .serializers import (ProposalInfoSerializer, ProposalCenterMappingSerializer, ProposalCenterMappingSpaceSerializer,
                         ProposalCenterMappingVersionSpaceSerializer)
from rest_framework.decorators import detail_route, list_route
from v0.ui.account.models import AccountInfo, ContactDetails
import v0.ui.website.utils as website_utils
import v0.ui.utils as ui_utils
from v0.ui.organisation.models import Organisation
from django.db import transaction
from v0.ui.location.models import City, CityArea, CitySubArea
from v0.ui.campaign.models import GenericExportFileName
from v0.ui.website.views import GenericExportFileSerializerReadOnly
from rest_framework.response import Response
from v0.ui.proposal.models import HashTagImages, SupplierAssignment
from v0.ui.proposal.serializers import (ProposalInfoSerializer, ProposalCenterMappingSerializer,
                                        ProposalCenterMappingVersionSerializer, ProposalInfoVersionSerializer,
                                        SpaceMappingSerializer, ProposalCenterMappingSpaceSerializer,
                                        ProposalCenterMappingVersionSpaceSerializer, SpaceMappingVersionSerializer,
                                        ProposalSocietySerializer, ProposalCorporateSerializer, HashtagImagesSerializer,
                                        SupplierAssignmentSerializer, ShortlistedSpacesVersionSerializer, BookingStatusSerializer)
from v0.ui.inventory.models import (SupplierTypeSociety, AdInventoryType, InventorySummary)
from .models import (ProposalInfo, ProposalCenterMapping, ProposalCenterMappingVersion, SpaceMappingVersion,
                    SpaceMapping, ShortlistedSpacesVersion, ShortlistedSpaces, SupplierPhase, BookingStatus, BookingSubstatus, TypeOfEndCustomer)
from .serializers import (SupplierPhaseSerializer)
from v0.ui.inventory.models import (AdInventoryType,InventoryActivityAssignment,InventorySummary,InventoryTypeVersion,
                                    InventoryType,InventoryActivity)
from v0.ui.inventory.serializers import (InventoryTypeSerializer, InventoryTypeVersionSerializer)
from v0.ui.finances.models import (ShortlistedInventoryPricingDetails, PriceMappingDefault, getPriceDict)
from v0.ui.campaign.models import Campaign
from v0.ui.website.utils import return_price
import v0.constants as v0_constants
import v0.ui.website.tasks as tasks
from v0.ui.email.views import send_email
from v0.ui.supplier.models import SupplierTypeCorporate, SupplierTypeRetailShop, SupplierMaster
from v0.ui.supplier.serializers import SupplierTypeSocietySerializer, SupplierMasterSerializer
from v0.ui.base.models import DurationType
from v0 import errors
from rest_framework import viewsets
from .models import SupplierPhase, ProposalInfo
from .serializers import SupplierPhaseSerializer
from v0.ui.utils import handle_response
from v0.ui.common.models import BaseUser
from v0.ui.campaign.models import CampaignComments
from v0.ui.common.models import mongo_client, mongo_test
from django.db.models import Prefetch, Max

from v0.ui.campaign.models import CampaignAssignment
from django.db.models.functions import Trim

import logging
logger = logging.getLogger(__name__)

def convert_date_format(date):
    if isinstance(date, datetime.datetime):
        return date
    try:
        date = datetime.datetime.strptime(str(date), '%d/%m/%Y')
        return date
    except Exception as ex:
        print(ex)
    try:
        date = datetime.datetime.strptime(str(date), '%d/%m/%y')
        return date
    except Exception as ex:
        print(ex)
    try:
        date = datetime.datetime.strptime(str(date), '%d-%m-%Y')
        return date
    except Exception as ex:
        print(ex)
    try:
        date = datetime.datetime.strptime(str(date), '%d-%m-%y')
        return date
    except Exception as ex:
        print(ex)
    try:
        date = datetime.datetime.strptime(str(date), '%m-%d-%Y')
        return date
    except Exception as ex:
        print(ex)
    try:
        date = datetime.datetime.strptime(str(date), '%d %b %Y')
        return date
    except Exception as ex:
        print(ex)
    try:
        date = datetime.datetime.strptime(str(date), '%Y-%d-%m %H:%M:%S')
        print(date)
        return date
    except Exception as ex:
        print(ex)
    try:
        date = datetime.datetime.strptime(str(date), '%m-%d-%y')
        return date
    except Exception as ex:
        print(ex)
        return date


def get_Date_Values(values):
    function_name = get_Date_Values.__name__
    try:
        result = []
        if values:
            values_list = [x for x in str(values).split(',')]
            if len(values_list) > 1:
                for value in values_list:
                    result.append(convert_date_format(value))
            else:
                result = values_list
        return result

    except Exception as e:
        return ui_utils.handle_response(function_name, data="error in Date conversion")


def create_proposal_object(organisation_id, account_id, user, tentative_cost, name):
    account = AccountInfo.objects.get_permission(user=user, account_id=account_id)
    return {
        'organisation_id': organisation_id,
        'account_id': account_id,
        'proposal_id': website_utils.get_generic_id(
            [Organisation.objects.get(pk=organisation_id).name, AccountInfo.objects.get(pk=account_id).name]),
        'account': account.account_id,
        'user': user.id,
        'created_by': user.first_name,
        'tentative_cost': tentative_cost,
        'name': name,
        'updated_by':user.first_name,
        'campaign_state': 'PTC',
        'principal_vendor_id': user.profile.organisation_id
    }

def genrate_supplier_data(data,user):
    function_name = genrate_supplier_data.__name__
    try:
        society_data_list = []
        contact_data_list = []
        total_society_id_list = []
        total_society_id_list_obj = {}
        source_file = data['file']
        wb = load_workbook(source_file)
        ws = wb.get_sheet_by_name(wb.get_sheet_names()[0])
        for index, row in enumerate(ws.iter_rows()):
            if index > 0:
                print("row is " + str(index),row[1].value.strip())
                
                try:
                    if not row[1].value:
                        continue
                    city = City.objects.filter(city_name__icontains=row[1].value.strip()).first()
                    city_code = city.city_code
                except ObjectDoesNotExist as e:
                    error = 'No City Found at - ' + str(index) + ',' + str(row[1].value)
                    return ui_utils.handle_response(function_name, data=error)
                try:
                    if not row[2].value:
                        continue
                    area = CityArea.objects.filter(label=row[2].value.strip()).first()
                    area_code = area.area_code
                except ObjectDoesNotExist as e:
                    error = 'No Area Found at - ' + str(index) + ',' + str(row[2].value)
                    return ui_utils.handle_response(function_name, data=error)
                try:
                    if not row[3].value:
                        continue
                    subarea = CitySubArea.objects.filter(subarea_name=row[3].value.strip()).first()
                    subarea_code = subarea.subarea_code
                except ObjectDoesNotExist as e:
                    error = 'No SubArea Found at - ' + str(index) + ',' + str(row[3].value)
                    return ui_utils.handle_response(function_name, data=error)
                try:
                    error = 'No Supplier Code - ' + str(index) + ',' + str(row[3].value)
                    if row[4].value:
                        supplier_type_code = row[4].value.strip()
                    else:
                        return ui_utils.handle_response(function_name, data=error)
                except ObjectDoesNotExist as e:
                    error = 'supplier code error - ' + str(index) + ',' + str(row[3].value)
                    return ui_utils.handle_response(function_name, data=error)
                
                supplier_code = row[0].value.strip().upper()

                if len(supplier_code) > 2:
                    supplier_code = supplier_code[:3]
                    
                supplier_id_dict = {
                    'city_code': city_code,
                    'area_code': area_code,
                    'subarea_code': subarea_code,
                    'supplier_type': supplier_type_code,
                    'supplier_code': supplier_code,
                }
                supplier_id = ui_utils.get_supplier_id(supplier_id_dict)
                
                content_type = ui_utils.get_content_type(supplier_type_code).data['data']

                supplier = SupplierTypeSociety.objects.filter(supplier_id=supplier_id).first()
                
                if not supplier:
                    submit_data = {
                        'city_id': city.id,
                        'area_id': area.id,
                        'subarea_id': subarea.id,
                        'latitude': float(row[6].value) if row[6].value else None,
                        'longitude': float(row[7].value) if row[7].value else None,
                        'landmark' : row[13].value if row[13].value else None,
                        'zipcode': int(row[5].value) if row[5].value else None,
                        'address1': row[12].value.strip() if row[12].value else None,
                        'supplier_code': supplier_code,
                        'supplier_type_code': supplier_type_code,
                        'supplier_name': row[0].value.strip(),
                        'supplier_id': supplier_id,
                        'current_user': user,
                        'unit_primary_count': int(row[8].value) if row[8].value else None, #flat_count
                        'unit_secondary_count': int(row[9].value) if row[9].value else None, #tower_count
                    }

                    response = ui_utils.make_supplier_data(submit_data)
                    all_supplier_data = response.data['data']
                    ui_utils.save_supplier_data(user, all_supplier_data)
                    
                    contact_data_list.append(ContactDetails(**{
                        'name': row[10].value.strip() if row[10].value else None,
                        'designation': 'Manager',
                        'salutation': 'Mr',
                        'mobile': row[11].value if row[11].value else None,
                        'object_id': supplier_id,
                        'content_type': content_type
                    }))

                temp_data =  {
                    'id': supplier_id,
                    'status' : 'F',
                    'index' : 0,
                    'rl_count' : 0,
                    'cl_count' : 0,
                }
                # total_society_id_list.append(temp_data)
                if not total_society_id_list_obj.get(supplier_type_code):
                    total_society_id_list_obj[supplier_type_code] = {
                        'supplier_data' : [],
                        'filter_codes' : [
                            { 'id' : 'PO'},
                            {'id': 'SL'},
                            {'id': 'ST'},
                            {'id': 'FL'}
                        ]
                    }
                
                total_society_id_list_obj[supplier_type_code]["supplier_data"].append(temp_data)
                

        try:
            ContactDetails.objects.bulk_create(contact_data_list)
        except Exception as e:
            return ui_utils.handle_response(function_name, data="error in bulk create contact")
        try:
            result = {
                'center_data' : total_society_id_list_obj,
                'proposal_id' : data['proposal_id'],
                'center_id' : data['center_id'],
                'invoice_number' : data['invoice_number'],
                'tentative_start_date' : data['tentative_start_date'],
                'tentative_end_date' : data['tentative_end_date'],
                'is_import_sheet' : True,
                'assigned_by' : data['assigned_by'],
                'assigned_to' : data['assigned_to']
            }

        except Exception as e:
            return ui_utils.handle_response(function_name, data="error in data creation")


        return ui_utils.handle_response(function_name, data=result, success=True)
    except Exception as e:
        return Exception(function_name, ui_utils.get_system_error(e))

def get_value_from_list_by_key(list1, key):
    try:
        return list1[key].value
    except:
        pass

def genrate_supplier_data2(data,user):

    source_file = data['file']
    wb = load_workbook(source_file)
    ws = wb.get_sheet_by_name(wb.get_sheet_names()[0])
    headers = {}
    for index, row in enumerate(ws.iter_rows()):
        if index == 0:
            i = 0
            for key in row:
                key1 = key.value.lower()
                key1 = key1.strip()
                headers[key1] = i
                i+=1
        else:
            supplier_id = get_value_from_list_by_key(row, headers.get('supplier id'))
            booking_status = get_value_from_list_by_key(row, headers.get('booking status'))
            booking_status1 = None
            if booking_status:
                booking_status1 = BookingStatus.objects.filter(name__iexact=booking_status).first()
            internal_comments = get_value_from_list_by_key(row, headers.get('internal comments'))
            external_comments = get_value_from_list_by_key(row, headers.get('external comments'))
            next_action_date = get_value_from_list_by_key(row, headers.get('next action date'))
            completion_status = get_value_from_list_by_key(row, headers.get('completion status'))
            phases_no = get_value_from_list_by_key(row, headers.get('phases'))
            assign_to_user = get_value_from_list_by_key(row, headers.get('assign user'))
            phase = None
            if phases_no:
                phase = SupplierPhase.objects.filter(campaign_id=data["proposal_id"], phase_no=phases_no).first()
            
            requirement_given_text = get_value_from_list_by_key(row, headers.get('requirement given'))
            requirement_given = "no"
            if requirement_given_text:
                requirement_given = requirement_given_text.lower()
            if supplier_id:
                shortlisted_spaces = ShortlistedSpaces.objects.filter(proposal_id=data["proposal_id"], object_id=supplier_id).first()
                if not shortlisted_spaces:
                    supplier_code = "RS"
                    supplier = SupplierMaster.objects.filter(supplier_id=supplier_id).first()
                    if supplier:
                        supplier_code = supplier.supplier_type

                    content_type = ui_utils.fetch_content_type(supplier_code)

                    shortlisted_spaces = ShortlistedSpaces(
                        proposal_id=data["proposal_id"],
                        supplier_code=supplier_code,
                        content_type=content_type,
                        object_id=supplier_id,
                        center_id=data['center_id'],
                        status='F'
                    )
                    
                if booking_status1:
                    shortlisted_spaces.booking_status = booking_status1.code
                if next_action_date:
                    shortlisted_spaces.next_action_date = next_action_date
                if completion_status and (completion_status.lower() == "yes" or completion_status.lower() == "y"):
                    shortlisted_spaces.is_completed = 1
                if phase:
                    shortlisted_spaces.phase = phase.phase_no
                    shortlisted_spaces.phase_no = phase                
                if not shortlisted_spaces.requirement_given == requirement_given:
                    shortlisted_spaces.requirement_given = requirement_given
                    shortlisted_spaces.requirement_given_date = datetime.datetime.now()

                shortlisted_spaces.save()

                now_time = timezone.now()
                if assign_to_user:
                    organisation = user.profile.organisation
                    assign_to = BaseUser.objects.filter(profile__organisation = organisation, username = assign_to_user).first()
                    if assign_to:
                        SupplierAssignment(
                            campaign_id=data["proposal_id"],
                            supplier_id=supplier_id,
                            assigned_by= user,
                            assigned_to= assign_to,
                            created_at= now_time,
                            updated_at= now_time
                        ).save()

                if internal_comments:
                    CampaignComments(
                        comment=internal_comments,
                        campaign_id=data["proposal_id"],
                        shortlisted_spaces=shortlisted_spaces,
                        user=user,
                        related_to='INTERNAL'
                    ).save()
                if external_comments:
                    CampaignComments(
                        comment=external_comments,
                        campaign_id=data["proposal_id"],
                        shortlisted_spaces=shortlisted_spaces,
                        user=user,
                        related_to='EXTERNAL'
                    ).save()
    return

def assign_inv_dates(data):
    function_name = assign_inv_dates.__name__
    try:
        inv_data = InventoryActivity.objects.select_related('shortlisted_inventory_details',
                                                 'shortlisted_inventory_details__shortlisted_spaces').filter(
            shortlisted_inventory_details__shortlisted_spaces__proposal_id=data['proposal_id']).\
            annotate(space_id=F('shortlisted_inventory_details__shortlisted_spaces'),inv_name=F('shortlisted_inventory_details__ad_inventory_type__adinventory_name'),
                     supplier_id=F('shortlisted_inventory_details__shortlisted_spaces__object_id')).\
            values('id','space_id', 'inv_name','activity_type','supplier_id')
        supplier_ids_mapping = {supplier['id'] : supplier for supplier in data['center_data']['RS']['supplier_data']}
        format_str = '%d/%m/%Y'  # The format
        assigned_by_user = BaseUser.objects.get(id=data['assigned_by'])
        assigned_to_user = BaseUser.objects.get(id=data['assigned_to'])
        inv_act_assignement_list = []
        index = 0
        for inv in inv_data:
            index += 1
            assigned_by = None
            assigned_to = None
            if inv['inv_name'] == 'POSTER' or inv['inv_name'] == 'FLIER' or inv['inv_name'] == 'STANDEE':
                if inv['supplier_id'] in supplier_ids_mapping:
                    date = supplier_ids_mapping[inv['supplier_id']]['inv_code'][inv['inv_name']]
                    if not date:
                        continue
                    if date:
                        assigned_by = assigned_by_user
                        assigned_to = assigned_to_user

                    if inv['activity_type'] == 'RELEASE':
                        temp_data = InventoryActivityAssignment(**{
                            'inventory_activity' : InventoryActivity.objects.get(id=inv['id']),
                            'activity_date' : convert_date_format(date),
                            'assigned_by' : assigned_by,
                            'assigned_to' : assigned_to
                        })
                        inv_act_assignement_list.append(temp_data)

                    # elif inv['activity_type'] == 'CLOSURE' and inv['inv_name'] == 'POSTER':
                    #     temp_data = InventoryActivityAssignment(**{
                    #         'inventory_activity': InventoryActivity.objects.get(id=inv['id']),
                    #         'activity_date': convert_date_format(date + datetime.timedelta(days=3)),
                    #         'assigned_by': assigned_by,
                    #         'assigned_to': assigned_to
                    #     })
                    #     inv_act_assignement_list.append(temp_data)

            if inv['inv_name'] == 'STALL':
                date = None

                if inv['activity_type'] == 'RELEASE':
                    try:
                        if inv['supplier_id'] in supplier_ids_mapping:
                            if supplier_ids_mapping[inv['supplier_id']]['inv_code'][inv['inv_name']]:
                                if supplier_ids_mapping[inv['supplier_id']]['rl_count'] >= supplier_ids_mapping[inv['supplier_id']]['index']:
                                    supplier_ids_mapping[inv['supplier_id']]['rl_count'] = supplier_ids_mapping[inv['supplier_id']]['index'] -1
                                date = supplier_ids_mapping[inv['supplier_id']]['inv_code'][inv['inv_name']][supplier_ids_mapping[inv['supplier_id']]['rl_count']]
                                supplier_ids_mapping[inv['supplier_id']]['rl_count'] += 1
                                assigned_to = assigned_to_user
                                assigned_by = assigned_by_user
                            temp_data = InventoryActivityAssignment(**{
                                'inventory_activity': InventoryActivity.objects.get(id=inv['id']),
                                'activity_date': convert_date_format(date),
                                'assigned_by': assigned_by,
                                'assigned_to': assigned_to
                            })
                            inv_act_assignement_list.append(temp_data)

                    except KeyError:
                        error = "Stall count and date not matching" + str(inv['supplier_id'])
                        return ui_utils.handle_response(function_name, error)

                elif inv['activity_type'] == 'CLOSURE':
                    try:
                        if inv['supplier_id'] in supplier_ids_mapping:
                            if supplier_ids_mapping[inv['supplier_id']]['inv_code'][inv['inv_name']]:
                                if supplier_ids_mapping[inv['supplier_id']]['cl_count'] >= supplier_ids_mapping[inv['supplier_id']]['index']:
                                    supplier_ids_mapping[inv['supplier_id']]['cl_count'] = supplier_ids_mapping[inv['supplier_id']]['index'] -1
                                date = supplier_ids_mapping[inv['supplier_id']]['inv_code'][inv['inv_name']][supplier_ids_mapping[inv['supplier_id']]['cl_count']]

                                assigned_to = assigned_to_user
                                assigned_by = assigned_by_user
                            temp_data = InventoryActivityAssignment(**{
                                'inventory_activity': InventoryActivity.objects.get(id=inv['id']),
                                'activity_date': convert_date_format(date),
                                'assigned_by': assigned_by,
                                'assigned_to': assigned_to
                            })
                            inv_act_assignement_list.append(temp_data)

                    except KeyError:
                        error = "Stall count and date not matching" + str(inv['supplier_id'])
                        return ui_utils.handle_response(function_name, error)



        # InventoryActivityAssignment.objects.filter(inventory_activity__shortlisted_inventory_details__shortlisted_spaces__proposal_id=data['proposal_id']).delete()
        InventoryActivityAssignment.objects.bulk_create(inv_act_assignement_list)
        return ui_utils.handle_response(function_name, data={}, success=True)
    except Exception as e:

        return Exception(function_name, ui_utils.get_system_error(e))

class CreateInitialProposal(APIView):
    """
    This is first step in creating proposal. Basic data get's stored here.
    because we have reduced number of models in which Proposal
    data is stored hence we have created new classes CreateInitialProposal and CreateFinalProposal API.
    author: nikhil
    """

    # permission_classes = (v0_permissions.IsGeneralBdUser,)

    def post(self, request, account_id):
        """
        Args:
            request: request param
            proposal_id:  proposal_id for which data is to be saved

        Returns: success or failure depending an initial proposal is created or not.

        """
        class_name = self.__class__.__name__
        try:
            with transaction.atomic():
                proposal_data = request.data
                user = request.user
                organisation_id = proposal_data['organisation_id']
                

                # create a unique proposal id
                proposal_data['proposal_id'] = website_utils.get_generic_id(
                    [Organisation.objects.get(pk=organisation_id).name, AccountInfo.objects.get(pk=account_id).name])

                # get the account object. required for creating the proposal
                account = AccountInfo.objects.get_permission(user=user, account_id=account_id)
                proposal_data['account'] = account.account_id
                proposal_data['user'] = user.id

                # query for parent. if available set it. if it's available, then this is an EDIT request.
                parent = request.data.get('parent')
                parent = parent if parent != '0' else None
                proposal_data['parent'] = parent
                # set parent if available
                if parent:
                    proposal_data['parent'] = ProposalInfo.objects.get_permission(user=user,
                                                                                  proposal_id=parent).proposal_id
                
                proposal_data['is_mix'] = False
                if request.data.get('centers') and request.data["centers"][0].get("suppliers"):
                    supplier_type_count = len([row for row in request.data["centers"][0]["suppliers"] if row["selected"] == True or row["selected"] == "True"])
                    proposal_data['is_mix'] = True if supplier_type_count > 1 else False
                
                # call the function that saves basic proposal information
                proposal_data['created_by'] = user.username
                proposal_data['updated_by'] = user.username
                response = website_utils.create_basic_proposal(proposal_data)
                if not response.data['status']:
                    return response

                # time to save all the centers data
                response = website_utils.save_center_data(proposal_data, user)
                if not response.data['status']:
                    return response

                # return the proposal_id of the new proposal created
                proposal_id = proposal_data['proposal_id']
                return ui_utils.handle_response(class_name, data=proposal_id, success=True)
        except Exception as e:
            logger.exception("Something bad happened in CreateInitialProposal")
            return ui_utils.handle_response(class_name, exception_object=e, request=request)

def calculate_hotness_level(multi_level_is_hot):
    hotness_level = 0
    for is_hot_level in multi_level_is_hot:
        if multi_level_is_hot[is_hot_level]:
            if int(is_hot_level[-1]) > hotness_level:
                hotness_level = int(is_hot_level[-1])
    return hotness_level


def calculate_is_hot(curr_lead, global_hot_lead_criteria):
    # checking 'or' global_hot_lead_criteria
    any_is_hot = False
    multi_level_is_hot = {is_hot_level: False for is_hot_level in global_hot_lead_criteria}
    curr_lead_data_dict = {str(item['item_id']): item for item in curr_lead['data']}
    for is_hot_level in global_hot_lead_criteria:
        for item_id in global_hot_lead_criteria[is_hot_level]['or']:
            if item_id in curr_lead_data_dict and curr_lead_data_dict[item_id]['value'] is not None:
                if str(curr_lead_data_dict[item_id]['value']) in global_hot_lead_criteria[is_hot_level]['or'][item_id]:
                    multi_level_is_hot[is_hot_level] = True
                    any_is_hot = True
                if "AnyValue" in global_hot_lead_criteria[is_hot_level]['or'][item_id] and str(
                        curr_lead_data_dict[item_id]['value']).lower() != 'na':
                    multi_level_is_hot[is_hot_level] = True
                    any_is_hot = True
    hotness_level = calculate_hotness_level(multi_level_is_hot)
    return any_is_hot, multi_level_is_hot, hotness_level


def create_lead_hash(lead_dict):
    lead_hash_string = ''
    lead_hash_string += str(lead_dict['leads_form_id'])

    for item in lead_dict['data']:
        if item['value']:
            if isinstance(item["value"], (str,bytes)):
                lead_hash_string += str(item['value'].strip())
            else:
                lead_hash_string += str(item['value'])
    return hashlib.sha256(lead_hash_string.encode('utf-8')).hexdigest()

def get_supplier_data_by_type(name, area_name, city_name):
    suppliers = SupplierTypeSociety.objects.filter(society_name__icontains=name, 
                                                    society_city__icontains=city_name, 
                                                    society_locality__icontains=area_name).values('supplier_id',
                                                                                     'society_name').all()
    if len(suppliers) > 0:
        return suppliers
    else:
        suppliers = SupplierTypeRetailShop.objects.filter(name__contains=name).values('supplier_id', 'name').all()
        if len(suppliers) > 0:
            return suppliers
    return []


class CreateDummyProposal(APIView):
    """docstring for CreateDummyProposal"""
    def post(self, request):
        class_name = self.__class__.__name__

        source_file = request.data['file']
        campaign_id = request.data['campaign_id']
        user = request.user

        wb = load_workbook(source_file)
        ws = wb.get_sheet_by_name(wb.get_sheet_names()[0])

        not_present_in_shortlisted_societies = []
        more_than_ones_same_shortlisted_society = []
        matched_societies = []
        unmatched_societies = []
        found_supplier_id = []

        leads_dict = []
        campaign_lead_form = list(mongo_client.leads_forms.find({"campaign_id": str(campaign_id)}))
        for leads_id in campaign_lead_form:
            leads_form_id = leads_id['leads_form_id']

        lead_form = mongo_client.leads_forms.find_one({"leads_form_id": int(leads_form_id)})
        fields = len(lead_form['data'])
        global_hot_lead_criteria = lead_form['global_hot_lead_criteria']
        entry_id = lead_form['last_entry_id'] + 1 if 'last_entry_id' in lead_form else 1

        all_sha256 = list(mongo_client.leads.find({"leads_form_id": int(leads_form_id)},{"lead_sha_256": 1, "_id": 0}))
        all_sha256_list = [str(element['lead_sha_256']) for element in all_sha256]
        apartment_index = None
        club_name_index = None
        area_index = None
        city_index = None
        for index, row in enumerate(ws.iter_rows()):
            if index == 0:
                for idx, i in enumerate(row):
                    if i.value and 'apartment' in i.value.lower():
                        apartment_index = idx

                    if i.value and 'area' in i.value.lower():
                        area_index = idx

                    if i.value and 'city' in i.value.lower():
                        city_index = idx

                if not apartment_index:
                    for idx, i in enumerate(row):
                        if i.value and 'club name' in i.value.strip().lower():
                            club_name_index = idx
                            break

            if apartment_index is None:
                return handle_response('', data="Apartment column missing in the sheet", success=False)
            
            if area_index is None:
                return handle_response('', data="Could not find Area column in sheet", success=False)

            if city_index is None:
                return handle_response('', data="Could not find City column in sheet", success=False)

            if index > 0:
                if not (row[apartment_index].value is None):
                    society_name = row[apartment_index].value

                if not (row[area_index].value is None):
                    area_name = row[area_index].value

                if not (row[city_index].value is None):
                    city_name = row[city_index].value

                suppliers = get_supplier_data_by_type(society_name, area_name, city_name)

                if len(suppliers) == 0:
                    if society_name not in unmatched_societies and society_name is not None:
                        unmatched_societies.append(society_name)
                    continue
                else:
                    if len(suppliers) == 1:
                        found_supplier_id = suppliers[0]['supplier_id']
                        matched_societies.append(society_name)

                        supplier_code = "RS"
                        content_type = ui_utils.fetch_content_type(supplier_code)  

                        data = {
                            'object_id': found_supplier_id,
                            'proposal_id': campaign_id,
                            'content_type': content_type,
                        }
                        obj, is_created = ShortlistedSpaces.objects.get_or_create(**data)
                        obj.save()

                        shortlisted_spaces = ShortlistedSpaces.objects.filter(object_id=found_supplier_id).filter(
                            proposal_id=campaign_id).all()
                        if len(shortlisted_spaces) == 0:
                            not_present_in_shortlisted_societies.append(society_name)
                            continue
                        current_date = datetime.datetime.now().strftime('%Y-%m-%dT%H:%M:%S.%f%z')
                        current_date = dateutil.parser.parse(current_date)
                        lead_dict = {"data": [], "is_hot": False, "created_at": current_date, "supplier_id": found_supplier_id,
                                     "campaign_id": campaign_id, "leads_form_id": int(leads_form_id), "entry_id": entry_id}
                        for item_id in range(0, fields):
                            curr_item_id = item_id + 1
                            curr_form_item_dict = lead_form['data'][str(curr_item_id)]
                            key_name = curr_form_item_dict['key_name']
                            value = row[item_id].value if row[item_id].value else None
                            if isinstance(value, datetime.datetime) or isinstance(value, datetime.time):
                                value = str(value)
                            item_dict = {
                                'key_name': key_name,
                                'value': value,
                                'item_id': curr_item_id
                            }
                            lead_dict["data"].append(item_dict)
                        lead_sha_256 = create_lead_hash(lead_dict)
                        lead_dict["lead_sha_256"] = lead_sha_256
                        lead_dict["is_hot"], lead_dict["multi_level_is_hot"], lead_dict["hotness_level"] = calculate_is_hot(lead_dict, global_hot_lead_criteria)
                        lead_already_exist = True if lead_sha_256 in all_sha256_list else False
                        if not lead_already_exist:
                            mongo_client.leads.insert_one(lead_dict)
                            entry_id = entry_id + 1  # will be saved in the end

                        mongo_client.leads_forms.update_one({"leads_form_id": leads_form_id}, {"$set": {"last_entry_id": entry_id}})
                        unmatched_societies.sort()

                    else:
                        supplier_ids = []
                        for s in suppliers:
                            supplier_ids.append(s['supplier_id'])
                        shortlisted_spaces = ShortlistedSpaces.objects.filter(proposal_id=campaign_id,
                                                                              object_id__in=supplier_ids).values(
                            'object_id', 'id').all()
                        if len(shortlisted_spaces) > 1:
                            more_than_ones_same_shortlisted_society.append(society_name)
                            continue
                        if len(shortlisted_spaces) == 0:
                            not_present_in_shortlisted_societies.append(society_name)
                            continue
                        else:
                            found_supplier_id = shortlisted_spaces[0]['object_id']

        lead_count_matched=dict(zip(list(matched_societies),[list(matched_societies).count(i) for i in list(matched_societies)]))
        lead_count_unmatched = dict(zip(list(unmatched_societies),[list(unmatched_societies).count(i) for i in list(unmatched_societies)]))
        missing_dict = {
            "lead_count_matched":lead_count_matched,
            "lead_count_unmatched":lead_count_unmatched,
            "matched_societies": list(set(matched_societies)),
            "unmatched_societies": list(set(unmatched_societies)),
            "more_than_ones_same_shortlisted_society": list(set(more_than_ones_same_shortlisted_society)),
            "not_present_in_shortlisted_societies": list(set(not_present_in_shortlisted_societies))
        }
        return handle_response({}, data=missing_dict, success=True)
        

class CreateInitialProposalBulkBasic(APIView):
    def post(self, request):
        class_name = self.__class__.__name__
        source_file = request.data['file']
        wb = load_workbook(source_file)
        ws = wb.get_sheet_by_name(wb.get_sheet_names()[0])
        proposal_list = []
        campaign_list = []
        proposal_data = {}
        organisation_id = request.data['organisation_id']
        account_id = request.data['account_id']
        user = request.user
        tentative_cost = request.data['tentative_cost']
        name = request.data['name']
        proposal_data = create_proposal_object(organisation_id, account_id, user, tentative_cost, name)
        response = website_utils.create_basic_proposal(proposal_data)
        if not response.data['status']:
            return response
        center_name = request.data['center_name']
        city = request.data['city']
        area = request.data['area']
        subarea = request.data['subarea']
        pincode = request.data['pincode']
        radius = request.data['radius']
        address = request.data['address']
        supplier_codes = request.data['codes'].split(',')
        supplier_codes = [x.strip(' ') for x in supplier_codes]
        city_id = City.objects.get(city_name=city).id
        area_id = CityArea.objects.get(label=area).id
        all_suppliers = {'RS': 'Societies', 'CP': 'Corporate Parks', 'BS': 'Bus Shelter', 'GY': 'Gym', 'SA': 'Saloon',
                         'RE': 'Retail Shop'}
        suppliers = []
        for supplier in all_suppliers:
            selected_status = False
            if supplier in supplier_codes:
                selected_status = True
            suppliers.append({
                'selected': selected_status,
                'code': supplier,
                'name': all_suppliers[supplier]
            })
        centers = [{
            'isEditProposal': False,
            'city': city_id,
            'area': area_id,
            'suppliers': suppliers,
            'center': {
                'city': city,
                'area': area,
                'codes': supplier_codes,
                'center_name': center_name,
                'subarea': subarea,
                'pincode': pincode,
                'radius': radius,
                'address': address
            }}]
        proposal_data['centers'] = centers
        center_response = website_utils.save_center_data(proposal_data, user)
        center_id = center_response.data['data']['center_id']
        if not center_response.data['status']:
            return response
        for index, row in enumerate(ws.iter_rows()):
            if index > 0:
                start_date = row[0].value if row[0].value else None
                end_date = row[1].value if row[1].value else None
                invoice_no = int(row[2].value) if row[2].value else None
                poster_filter = True if int(row[3].value) == 1 else False
                standee_filter = True if int(row[4].value) == 1 else False
                flier_filter = True if int(row[5].value) == 1 else False
                stall_filter = True if int(row[6].value) == 1 else False
                gatewayarch_filter = True if int(row[7].value) == 1 else False
                supplier_id = row[8].value if row[8].value else None
                total_negotiated_price = int(row[9].value) if row[9].value else None
                start_date = start_date.replace(tzinfo=pytz.UTC) if start_date else None
                end_date = end_date.replace(tzinfo=pytz.UTC) if end_date else None
                proposal_data['tentative_start_date'] = start_date
                proposal_data['tentative_end_date'] = end_date
                proposal_data['center_id'] = center_id
                proposal_data['invoice_number'] = invoice_no
                proposal_data['total_negotiated_price'] = total_negotiated_price
                center = ProposalCenterMapping.objects.get(pk=center_id)
                filter_codes = []
                if poster_filter:
                    filter_codes.append({'id': 'PO'})
                if stall_filter:
                    filter_codes.append({'id': 'SL'})
                if standee_filter:
                    filter_codes.append({'id': 'ST'})
                if flier_filter:
                    filter_codes.append({'id': 'FL'})
                if gatewayarch_filter:
                    filter_codes.append({'id': 'GA'})
                supplier_data = [{
                    'status': 'F',
                    'id': supplier_id,
                    'total_negotiated_price': total_negotiated_price
                }]
                center_data = {
                    'RS': {
                        'filter_codes': filter_codes,
                        'supplier_data': supplier_data
                    }
                }
                proposal_data['center_data'] = center_data
                proposal = ProposalInfo.objects.get(pk=proposal_data['proposal_id'])
                for supplier_code in proposal_data['center_data']:
                    response = website_utils.save_filters(center, supplier_code, proposal_data, proposal)
                    if not response.data['status']:
                        return response
                    response = website_utils.save_shortlisted_suppliers_data(center, supplier_code, proposal_data,
                                                                             proposal)
                    if not response.data['status']:
                        return response
                    response = website_utils.save_shortlisted_inventory_pricing_details_data(center, supplier_code,
                                                                                             proposal_data, proposal)
                    if not response.data['status']:
                        return response
                response = website_utils.update_proposal_invoice_and_state(proposal_data, proposal)
                if not response.data['status']:
                    return response
                response = website_utils.create_generic_export_file_data(proposal)
                if not response.data['status']:
                    return response
        return ui_utils.handle_response({}, data='success', success=True)


class ChildProposals(APIView):
    """
    Fetches child proposals of a given proposal. if the given proposal is string '0', all the parent proposals are fetched.
    """

    def get(self, request, proposal_id=None):

        """
        Fetches all proposals who have parent = pk
        Args:
            request: request param
            pk: parent pk value. if pk == '0',this means we need to fetch all proposals whose parent is NULL.
            proposal_id: Proposal id
        Returns:
        """
        class_name = self.__class__.__name__
        try:
            account_id = request.query_params.get('account_id')
            account_id = account_id if account_id != '0' else None
            data = {
                'parent': proposal_id if proposal_id != '0' else None,
                'user': request.user,
                'account_id': account_id
            }
            return ui_utils.handle_response(class_name, data=website_utils.child_proposals(data), success=True)
        except Exception as e:
            return ui_utils.handle_response(class_name, exception_object=e, request=request)


class ProposalViewSet(viewsets.ViewSet):
    """
    A ViewSet handling various operations related to ProposalModel.
    This viewset was made instead of creating separate ApiView's because all the api's in this viewset
    are related to Proposal domain. so keeping them at one place makes sense.
    """

    # parser_classes = (JSONParser, FormParser)
    # permission_classes = (v0_permissions.IsProposalAuthenticated, )

    def retrieve(self, request, pk=None):
        """
        Fetches one Proposal object
        Args:
            request: request parameter
            pk: primary key of proposal

        Returns: one proposal object

        """
        class_name = self.__class__.__name__
        try:

            proposal = ProposalInfo.objects.get(proposal_id=pk)
            serializer = ProposalInfoSerializer(proposal)
            return ui_utils.handle_response(class_name, data=serializer.data, success=True)
        except Exception as e:
            return ui_utils.handle_response(class_name, exception_object=e, request=request)

    def list(self, request):
        class_name = self.__class__.__name__
        try:
            proposal = ProposalInfo.objects.all()
            serializer = ProposalInfoSerializer(proposal, many=True)
            return ui_utils.handle_response(class_name, data=serializer.data, success=True)
        except Exception as e:
            return ui_utils.handle_response(class_name, exception_object=e, request=request)

    def update(self, request, pk=None):
        """
        Args:
            request: The request body
            pk: primary key

        Returns: Updated proposal object
        """
        class_name = self.__class__.__name__
        try:
            # prepare the data to be updated
            data = request.data.copy()
            data['proposal_id'] = pk

            proposal = ProposalInfo.objects.get(proposal_id=pk)
            serializer = ProposalInfoSerializer(proposal, data=data)
            if serializer.is_valid():
                serializer.save()
                return ui_utils.handle_response(class_name, data=serializer.data, success=True)
            return ui_utils.handle_response(class_name, data=serializer.errors)
        except Exception as e:
            return ui_utils.handle_response(class_name, exception_object=e, request=request)

    @detail_route(methods=['GET'])
    def proposal_centers(self, request, pk=None):
        """
        Fetches all centers associated with this proposal
        Args:
            request: request parameter
            pk: primary key of proposal

        Returns: Fetches all centers associated with this proposal. in each center object we have 'codes' array
        which contains which suppliers are allowed. Response is like this
        {
            "status": true,
            "data": [
                {
                  "id": 44,
                  "created_at": "2016-12-01T00:00:00Z",
                  "updated_at": "2016-12-01T00:00:00Z",
                  "center_name": "powai",
                  "address": "powai",
                  "latitude": 19.1153798,
                  "longitude": 72.9091436,
                  "radius": 1,
                  "subarea": "Hiranandani Gardens",
                  "area": "Powai",
                  "city": "Mumbai",
                  "pincode": 400076,
                  "user": 1,
                  "proposal": "BVIDBHBH157a5",
                  "codes": [
                    "RS"
                  ]
                }
              ]
        }
        """
        class_name = self.__class__.__name__
        try:
            response = website_utils.construct_proposal_response(pk)
            if not response.data['status']:
                return response
            return ui_utils.handle_response(class_name, data=response.data['data'], success=True)
        except Exception as e:
            return ui_utils.handle_response(class_name, exception_object=e, request=request)

    @list_route()
    def invoice_proposals(self, request):
        """
        Args:
            request: request body
        Returns: All the proposal features  which have invoice_number not null
        """
        class_name = self.__class__.__name__
        try:
            # can't use distinct() to return only unique proposal_id's because .distinct('proposal') is not supported
            # for MySql

            organisation_id = request.user.profile.organisation.organisation_id
            username_list = BaseUser.objects.filter(profile__organisation=organisation_id).values_list('username')
            file_objects = GenericExportFileName.objects.select_related('proposal', 'user').filter(
                proposal__invoice_number__isnull=False, is_exported=False,
                proposal__created_by__in=username_list).order_by('-proposal__created_on')
            
            # we need to make a unique list where proposal_id do not repeat.
            seen = set()
            final_file_objects = []
            for file_object in file_objects:
                proposal_id = file_object.proposal_id
                if proposal_id not in seen:
                    seen.add(proposal_id)
                    final_file_objects.append(file_object)
            file_serializer = GenericExportFileSerializerReadOnly(final_file_objects, many=True).data

            seen = list(seen)
            comments_list = CampaignComments.objects.values('campaign_id').annotate(latest_id=Max('id'), comment_max=Trim('comment')).filter(campaign_id__in=seen, related_to='campaign')
            comments_list_dict = {row["campaign_id"]: row["comment_max"] for row in comments_list}
            for row in file_serializer:
                row["latest_comment"] = comments_list_dict.get(row["proposal"]["proposal_id"])

            return ui_utils.handle_response(class_name, data=file_serializer, success=True)
        except Exception as e:
            return ui_utils.handle_response(class_name, exception_object=e, request=request)

    @detail_route(methods=['POST'])
    def get_spaces(self, request, pk=None):
        """
        The API  fetches all the data required to display on the grid view page.
        response looks like :
        {
           'status': true,
           'data' : {
              "business_name": '',
              "suppliers":
                    [
                        {
                           suppliers: { RS: [], CP: [] } ,
                           center: { ...   codes: [] }
                           suppliers_meta: {
                                             'RS': { 'inventory_type_selected' : [ 'PO', 'POST', 'ST' ]  },
                                             'CP': { 'inventory_type_selected':  ['ST']
                            }
                        }
                        ,
                        { }
                    ]
              }

        }
        Args:
            request:  request param
            pk: proposal_id
        Returns: collects data for all shortlisted suppliers and filters and send them.
        ---
        parameters:
        - name: center_id
          description:  center_id
        - name: radius
          description: radius
        - name: latitude
        - name: longitude
        """
        class_name = self.__class__.__name__
        try:
            center_id = request.data.get('center_id')
            radius = request.data.get('radius')
            latitude = request.data.get('latitude')
            longitude = request.data.get('longitude')

            data = {
                'proposal_id': pk,
                'center_id': center_id,
                'radius': radius,
                'latitude': latitude,
                'longitude': longitude,
            }

            # update center's radius here. This is being updated so that when next time get_spaces is called, radius is reflected in front end
            if center_id:
                instance = ProposalCenterMapping.objects.get(id=int(center_id))
                instance.radius = float(radius)
                instance.save()

            response = website_utils.suppliers_within_radius(data)
            if not response.data['status']:
                return response
            response = website_utils.add_shortlisted_suppliers_get_spaces(pk, request.user, response.data['data'])
            if not response.data['status']:
                return response
            result = response.data['data']
            result["suppliers"] = list(result["suppliers"])
            return ui_utils.handle_response(class_name, data=result, success=True)
        except Exception as e:
            return ui_utils.handle_response(class_name, exception_object=e, request=request)

    @detail_route(methods=['GET'])
    def shortlisted_suppliers(self, request, pk=None):
        """
        Fetches all shortlisted suppliers for this proposal.
        Response looks like :
        {
           'status': true,
           'data' : [
                {
                   suppliers: { RS: [], CP: [] } ,
                   center: { ...   codes: [] }
                }
           ]
        }

        Args:
            request: request
            pk: pk

        Returns:

        """
        class_name = self.__class__.__name__
        try:
            data = {
                'proposal_id': pk,
                'user': request.user
            }
            response = website_utils.proposal_shortlisted_spaces(data)
            if not response.data['status']:
                return response
            
            return ui_utils.handle_response(class_name, data=response.data['data'], success=True)
        except Exception as e:
            return ui_utils.handle_response(class_name, exception_object=e, request=request)

    @detail_route(methods=['PUT'])
    def shortlisted_suppliers_status(self, request, pk=None):
        """
            Update shortlisted suppliers based on their status value.
        Response looks like :def list
        {
           'status': true,
           'data' : [
                {
                   suppliers: { RS: [], CP: [] } ,
                   center: { ...   codes: [] }
                }
           ]
        }

        Args:
            request: request
            pk: pk

        Returns:

        """
        class_name = self.__class__.__name__
        try:
            center_id = request.data['center']['id']
            proposal = request.data['proposal']
            shortlisted_suppliers = []

            fixed_data = {
                'center': center_id,
                'proposal': proposal,
            }
            unique_supplier_codes = list(request.data['suppliers'].keys())
            for code in unique_supplier_codes:
                # get the right model and content_type
                response = ui_utils.get_content_type(code)
                if not response:
                    return response
                content_type = response.data.get('data')
                fixed_data['content_type'] = content_type
                fixed_data['supplier_code'] = code
                shortlisted_suppliers.append(
                    website_utils.save_shortlisted_suppliers(request.data['suppliers'][code], fixed_data))

            return ui_utils.handle_response(class_name, data=shortlisted_suppliers, success=True)
        except Exception as e:
            return ui_utils.handle_response(class_name, exception_object=e, request=request)


class InitialProposalAPIView(APIView):
    '''This API creates initial proposal when the user enters the center(address, name etc.) and basic proposal
    fields are stored in the database
    ProposalInfo and ProposalCenterMapping models are used only'''

    def post(self, request, account_id=None, format=None):
        '''In this centers contain format like
        centers : [
            center : [
                space_mapping : []
            ]
            society_inventory : []  // these will be made if in center[space_mapping][society_allowed] is true
            corporate_inventory : []
        ]
        This is done to be in sync with the format on map view page as serializers.data dont allow to append
        any new (key,value) pair to its existing data
        '''

        supplier_codes = {
            'society': 'RS', 'corporate': 'CP',
            'gym': 'GY', 'salon': 'SA'
        }

        with transaction.atomic():
            proposal_data = request.data
            proposal_data['proposal_id'] = self.create_proposal_id()
            try:
                account = AccountInfo.objects.get(account_id=account_id)
            except AccountInfo.DoesNotExist:
                return Response({'message': 'Invalid Account ID'}, status=406)
            proposal_data['account'] = account.account_id
            try:
                proposal_object = ProposalInfo.objects.get(proposal_id=proposal_data['proposal_id'])
                # id already exists --> Do something
                return Response(status=404)
            except ProposalInfo.DoesNotExist:
                proposal_serializer = ProposalInfoSerializer(data=proposal_data)

                if proposal_serializer.is_valid():
                    proposal_object = proposal_serializer.save()
                else:
                    return Response({'message': 'Invalid Proposal Info', 'errors': \
                        proposal_serializer.errors}, status=406)

                for center_info in proposal_data['centers']:
                    center = center_info['center']
                    space_mapping = center['space_mapping']
                    center['proposal'] = proposal_object.proposal_id
                    address = center['address'] + "," + center['subarea'] + ',' + center['area'] + ',' + center[
                        'city'] + ' ' + center['pincode']
                    geocoder = Geocoder(api_key='AIzaSyCy_uR_SVnzgxCQTw1TS6CYbBTQEbf6jOY')
                    try:
                        geo_object = geocoder.geocode(address)
                    except GeocoderError:
                        ProposalInfo.objects.get(proposal_id=proposal_object.proposal_id).delete()
                        return Response({'message': 'Latitude Longitude Not found for address : ' + address},
                                        status=406)
                    except ConnectionError:
                        ProposalInfo.objects.get(proposal_id=proposal_object.proposal_id).delete()
                        return Response({'message': 'Unable to connect to google Maps'}, status=406)

                    center['latitude'] = geo_object.latitude
                    center['longitude'] = geo_object.longitude

                    center_serializer = ProposalCenterMappingSerializer(data=center)

                    if center_serializer.is_valid():
                        center_object = center_serializer.save()
                    else:
                        ProposalInfo.objects.get(proposal_id=proposal_object.proposal_id).delete()
                        return Response({'message': 'Invalid Center Data', 'errors': center_serializer.errors}, \
                                        status=406)

                    space_mapping['center'] = center_object.id
                    space_mapping['proposal'] = proposal_object.proposal_id
                    space_mapping_serializer = SpaceMappingSerializer(data=space_mapping)
                    if space_mapping_serializer.is_valid():
                        space_mapping_object = space_mapping_serializer.save()
                    else:
                        ProposalInfo.objects.get(proposal_id=proposal_object.proposal_id).delete()
                        return Response({
                            'message': 'Invalid Space Mapping Data',
                            'errors': space_mapping_serializer.errors
                        }, status=406)

                    # ADDNEW --> extend the list in for loop when new spaces added. Keep the variables names accordingly
                    for space in ['society', 'corporate', 'gym', 'salon']:
                        ''' This loops checks if the space is allowed and if it is allowed save the
                        inventory types chosen by the user in the inventory_type table '''
                        try:
                            space_allowed = space + '_allowed'
                            if space_mapping[space_allowed]:
                                space_inventory = space + '_inventory'
                                center_info[space_inventory]['supplier_code'] = supplier_codes[space]
                                center_info[space_inventory]['space_mapping'] = space_mapping_object.id
                                inventory_type_serializer = InventoryTypeSerializer(data=center_info[space_inventory])
                                if inventory_type_serializer.is_valid():
                                    inventory_type_serializer.save()
                                else:
                                    ProposalInfo.objects.get(proposal_id=proposal_object.proposal_id).delete()
                                    return Response({
                                        'message': 'Invalid Inventory Type Info',
                                        'errors': inventory_type_serializer.errors
                                    })
                        except KeyError:
                            pass

        return Response(proposal_object.proposal_id, status=200)

    def create_proposal_id(self):
        import random, string
        return ''.join(random.choice(string.ascii_letters) for _ in range(8))


class GetAccountProposalsAPIView(APIView):
    """
    fetches proposals for a given account_id
    """

    def get(self, request, account_id, format=None):

        try:
            account = AccountInfo.objects.get(account_id=account_id)

            proposals = ProposalInfo.objects.filter(account=account)
            proposal_serializer = ProposalInfoSerializer(proposals, many=True)

            return Response(proposal_serializer.data, status=status.HTTP_200_OK)
        except ObjectDoesNotExist as e:
            return Response({'status': False, 'error': e.message}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({'status': False, 'error': e.message}, status=status.HTTP_400_BAD_REQUEST)


class CurrentProposalAPIView(APIView):

    def get(self, request, proposal_id, format=None):
        ''' This returns the proposal info of the proposal id
        '''

        try:
            proposal_object = ProposalInfo.objects.get(proposal_id=proposal_id)
        except ProposalInfo.DoesNotExist:
            return Response({'message': 'Invalid Proposal ID'}, status=406)

        proposal_serializer = ProposalInfoSerializer(proposal_object)
        centers = proposal_object.get_centers()
        centers_list = []

        for center in centers:
            space_info_dict = {}
            space_mapping_object = center.get_space_mappings()
            center_serializer = ProposalCenterMappingSpaceSerializer(center)
            space_info_dict['center'] = center_serializer.data

            if space_mapping_object.society_allowed:
                societies_shortlisted_ids = ShortlistedSpaces.objects.filter(space_mapping=space_mapping_object,
                                                                             supplier_code='RS',
                                                                             buffer_status=False).values_list(
                    'object_id', flat=True)

                societies_shortlisted_temp = SupplierTypeSociety.objects.filter(
                    supplier_id__in=societies_shortlisted_ids).values('supplier_id', 'society_latitude',
                                                                      'society_longitude', 'society_name',
                                                                      'society_address1', 'society_subarea',
                                                                      'society_location_type')

                societies_shortlisted = []
                societies_shortlisted_count = 0

                for society in societies_shortlisted_temp:
                    society_inventory_obj = InventorySummary.objects.get_supplier_type_specific_object(
                        request.data.copy(),
                        society['supplier_id'])
                    # society_inventory_obj = InventorySummary.objects.get(supplier_id=society['supplier_id'])
                    society['shortlisted'] = True
                    society['buffer_status'] = False
                    # obj = InventorySummaryAPIView()
                    adinventory_type_dict = ui_utils.adinventory_func()
                    duration_type_dict = ui_utils.duration_type_func()

                    if society_inventory_obj.poster_allowed_nb or society_inventory_obj.poster_allowed_lift:
                        society['total_poster_count'] = society_inventory_obj.poster_count_per_tower
                        society['poster_price'] = return_price(adinventory_type_dict, duration_type_dict, 'poster_a4',
                                                               'campaign_weekly')

                    if society_inventory_obj.standee_allowed:
                        society['total_standee_count'] = society_inventory_obj.total_standee_count
                        society['standee_price'] = return_price(adinventory_type_dict, duration_type_dict,
                                                                'standee_small', 'campaign_weekly')

                    if society_inventory_obj.stall_allowed:
                        society['total_stall_count'] = society_inventory_obj.total_stall_count
                        society['stall_price'] = return_price(adinventory_type_dict, duration_type_dict, 'stall_small',
                                                              'unit_daily')
                        society['car_display_price'] = return_price(adinventory_type_dict, duration_type_dict,
                                                                    'car_display_standard', 'unit_daily')

                    if society_inventory_obj.flier_allowed:
                        society['flier_frequency'] = society_inventory_obj.flier_frequency
                        society['filer_price'] = return_price(adinventory_type_dict, duration_type_dict,
                                                              'flier_door_to_door', 'unit_daily')

                    societies_shortlisted.append(society)
                    societies_shortlisted_count += 1

                space_info_dict['societies_shortlisted'] = societies_shortlisted
                space_info_dict['societies_shortlisted_count'] = societies_shortlisted_count

                societies_buffered_ids = ShortlistedSpaces.objects.filter(space_mapping=space_mapping_object,
                                                                          supplier_code='RS', \
                                                                          buffer_status=True).values_list('object_id',
                                                                                                          flat=True)
                societies_buffered_temp = SupplierTypeSociety.objects.filter(
                    supplier_id__in=societies_buffered_ids).values('supplier_id', 'society_latitude',
                                                                   'society_longitude', 'society_name',
                                                                   'society_address1', 'society_subarea',
                                                                   'society_location_type')
                societies_buffered = []
                societies_buffered_count = 0
                for society in societies_buffered_temp:
                    society_inventory_obj = InventorySummary.objects.get_supplier_type_specific_object(
                        request.data.copy(),
                        society['supplier_id'])

                    # society_inventory_obj = InventorySummary.objects.get(supplier_id=society['supplier_id'])
                    society['shortlisted'] = True
                    society['buffer_status'] = False
                    # obj = InventorySummaryAPIView()
                    adinventory_type_dict = ui_utils.adinventory_func()
                    duration_type_dict = ui_utils.duration_type_func()

                    if society_inventory_obj.poster_allowed_nb or society_inventory_obj.poster_allowed_lift:
                        society['total_poster_count'] = society_inventory_obj.poster_count_per_tower
                        society['poster_price'] = return_price(adinventory_type_dict, duration_type_dict, 'poster_a4',
                                                               'campaign_weekly')

                    if society_inventory_obj.standee_allowed:
                        society['total_standee_count'] = society_inventory_obj.total_standee_count
                        society['standee_price'] = return_price(adinventory_type_dict, duration_type_dict,
                                                                'standee_small', 'campaign_weekly')

                    if society_inventory_obj.stall_allowed:
                        society['total_stall_count'] = society_inventory_obj.total_stall_count
                        society['stall_price'] = return_price(adinventory_type_dict, duration_type_dict, 'stall_small',
                                                              'unit_daily')
                        society['car_display_price'] = return_price(adinventory_type_dict, duration_type_dict,
                                                                    'car_display_standard', 'unit_daily')

                    if society_inventory_obj.flier_allowed:
                        society['flier_frequency'] = society_inventory_obj.flier_frequency
                        society['filer_price'] = return_price(adinventory_type_dict, duration_type_dict,
                                                              'flier_door_to_door', 'unit_daily')

                    societies_buffered.append(society)
                    societies_buffered_count += 1

                space_info_dict['societies_buffered'] = societies_buffered
                space_info_dict['societies_buffered_count'] = societies_buffered_count

                societies_inventory = InventoryType.objects.get(supplier_code='RS', space_mapping=space_mapping_object)
                societies_inventory_serializer = InventoryTypeSerializer(societies_inventory)
                # inventory count only for shortlisted ones
                # to add buffered societies as well uncomment following line
                # societies_shortlisted_ids.extend(societies_buffered_ids)
                societies_inventory_count = InventorySummary.objects.filter(
                    supplier_id__in=societies_shortlisted_ids).aggregate(posters=Sum('poster_count_per_tower'), \
                                                                         standees=Sum('total_standee_count'),
                                                                         stalls=Sum('total_stall_count'),
                                                                         fliers=Sum('flier_frequency'))

                # Count only for society_shortlisted
                space_info_dict['societies_inventory_count'] = societies_inventory_count
                space_info_dict['societies_inventory'] = societies_inventory_serializer.data

            if space_mapping_object.corporate_allowed:
                # ADDNEW -->
                pass

            if space_mapping_object.salon_allowed:
                # ADDNEW -->
                pass

            if space_mapping_object.gym_allowed:
                # ADDNEW -->
                pass

            centers_list.append(space_info_dict)

        response = {
            'proposal': proposal_serializer.data,
            'centers': centers_list,
        }

        return Response(response, status=200)

    def post(self, request, proposal_id, format=None):
        ''' Updates the buffer and shortlisted spaces. This API allows user to delete
        move buffer to shortlisted and vice versa. No addition allowed using this API
        '''
        try:
            proposal_object = ProposalInfo.objects.get(proposal_id=proposal_id)
        except ProposalInfo.DoesNotExist:
            return Response({'message': 'Invalid Proposal ID'}, status=406)

        # version save
        # proposal_version_object = ProposalInfoVersion(proposal=proposal_object, name=proposal_object.name, payment_status=proposal_object.payment_status,\
        #     created_on=proposal_object.created_on, created_by=proposal_object.created_by, tentative_cost=proposal_object.tentative_cost,\
        #     tentative_start_date=proposal_object.tentative_start_date, tentative_end_date=proposal_object.tentative_end_date)
        # proposal_version_object.save()

        shortlisted_space_list = []
        centers = request.data
        for center_info in centers:
            space_mappings = center_info['center']['space_mappings']
            space_mapping_id = space_mappings['id']

            ShortlistedSpaces.objects.filter(space_mapping_id=space_mapping_id).delete()

            if space_mappings['society_allowed']:
                supplier_code = 'RS'
                content_type = ContentType.objects.get_for_model(SupplierTypeSociety)
                for society in center_info['societies_shortlisted']:
                    if society['shortlisted']:
                        object_id = society['supplier_id']
                        shortlisted_society = ShortlistedSpaces(space_mapping_id=space_mapping_id,
                                                                content_type=content_type, \
                                                                supplier_code=supplier_code, object_id=object_id,
                                                                buffer_status=society['buffer_status'])

                        shortlisted_space_list.append(shortlisted_society)

                for society in center_info['societies_buffered']:
                    if society['shortlisted']:
                        object_id = society['supplier_id']
                        shortlisted_society = ShortlistedSpaces(space_mapping_id=space_mapping_id,
                                                                content_type=content_type,
                                                                supplier_code=supplier_code, object_id=object_id,
                                                                buffer_status=society['buffer_status'])

                        shortlisted_space_list.append(shortlisted_society)

            if space_mappings['corporate_allowed']:
                # ADDNEW -->
                pass

            if space_mappings['salon_allowed']:
                # ADDNEW -->
                pass

            if space_mappings['gym_allowed']:
                # ADDNEW -->
                pass

        ShortlistedSpaces.objects.bulk_create(shortlisted_space_list)
        return Response(status=200)


class ProposalHistoryAPIView(APIView):
    def get(self, request, proposal_id, format=None):
        ''' Sends the proposal versions for the particular proposal id
        Currently if no proposal_id  set to a default one
        sends socities shortlisted and buffered differently
        '''

        try:
            proposal_object = ProposalInfo.objects.get(proposal_id=proposal_id)
        except ProposalInfo.DoesNotExist:
            return Response({'message': 'Invalid Proposal ID'}, status=406)

        proposal_versions = proposal_object.get_proposal_versions()
        proposal_versions_list = []

        try:
            for proposal_version_object in proposal_versions:
                proposal_info_dict = {}
                proposal_version_serializer = ProposalInfoVersionSerializer(proposal_version_object)
                proposal_info_dict['proposal'] = proposal_version_serializer.data
                proposal_center_versions = ProposalCenterMappingVersion.objects.filter(
                    proposal_version=proposal_version_object)
                center_versions_list = []
                for center_version_object in proposal_center_versions:
                    space_info_dict = {}
                    center_version_serailizer = ProposalCenterMappingVersionSpaceSerializer(center_version_object)
                    space_info_dict['center'] = center_version_serailizer.data

                    space_mapping_version_object = SpaceMappingVersion.objects.get(center_version=center_version_object)

                    if space_mapping_version_object.society_allowed:
                        societies_shortlisted_ids = ShortlistedSpacesVersion.objects.filter(
                            space_mapping_version=space_mapping_version_object,
                            supplier_code='RS', buffer_status=False).values_list('object_id', flat=True)
                        societies_shortlisted_temp = SupplierTypeSociety.objects.filter(
                            supplier_id__in=societies_shortlisted_ids).values('supplier_id', 'society_latitude',
                                                                              'society_longitude', 'society_name',
                                                                              'society_address1', 'society_subarea',
                                                                              'society_location_type')
                        societies_shortlisted = []
                        societies_shortlisted_count = 0
                        # societies_shortlisted_serializer = ProposalSocietySerializer(societies_shortlisted, many=True)
                        for society in societies_shortlisted_temp:
                            society_inventory_obj = InventorySummary.objects.get_supplier_type_specific_object(
                                request.data.copy(),
                                society['supplier_id'])

                            # society_inventory_obj = InventorySummary.objects.get(supplier_id=society['supplier_id'])
                            society['shortlisted'] = True
                            society['buffer_status'] = False
                            # obj = InventorySummaryAPIView()
                            adinventory_type_dict = ui_utils.adinventory_func()
                            duration_type_dict = ui_utils.duration_type_func()

                            if society_inventory_obj.poster_allowed_nb or society_inventory_obj.poster_allowed_lift:
                                society['total_poster_count'] = society_inventory_obj.total_poster_count
                                society['poster_price'] = return_price(adinventory_type_dict, duration_type_dict,
                                                                       'poster_a4', 'campaign_weekly')

                            if society_inventory_obj.standee_allowed:
                                society['total_standee_count'] = society_inventory_obj.total_standee_count
                                society['standee_price'] = return_price(adinventory_type_dict, duration_type_dict,
                                                                        'standee_small', 'campaign_weekly')

                            if society_inventory_obj.stall_allowed:
                                society['total_stall_count'] = society_inventory_obj.total_stall_count
                                society['stall_price'] = return_price(adinventory_type_dict, duration_type_dict,
                                                                      'stall_small', 'unit_daily')
                                society['car_display_price'] = return_price(adinventory_type_dict, duration_type_dict,
                                                                            'car_display_standard', 'unit_daily')

                            if society_inventory_obj.flier_allowed:
                                society['flier_frequency'] = society_inventory_obj.flier_frequency
                                society['filer_price'] = return_price(adinventory_type_dict, duration_type_dict,
                                                                      'flier_door_to_door', 'unit_daily')

                            societies_shortlisted.append(society)
                            societies_shortlisted_count += 1

                        space_info_dict['societies_shortlisted'] = societies_shortlisted
                        space_info_dict['societies_shortlisted_count'] = societies_shortlisted_count

                        societies_buffered_ids = ShortlistedSpacesVersion.objects.filter(
                            space_mapping_version=space_mapping_version_object, \
                            supplier_code='RS', buffer_status=True).values_list('object_id', flat=True)
                        societies_buffered_temp = SupplierTypeSociety.objects.filter(
                            supplier_id__in=societies_buffered_ids).values('supplier_id', 'society_latitude',
                                                                           'society_longitude', 'society_name',
                                                                           'society_address1', 'society_subarea',
                                                                           'society_location_type')
                        # societies_buffered_serializer = ProposalSocietySerializer(societies_buffered, many=True)
                        societies_buffered = []
                        societies_buffered_count = 0
                        for society in societies_buffered_temp:
                            society_inventory_obj = InventorySummary.objects.get_supplier_type_specific_object(
                                request.data.copy(),
                                society['supplier_id'])

                            # society_inventory_obj = InventorySummary.objects.get(supplier_id=society['supplier_id'])
                            society['shortlisted'] = True
                            society['buffer_status'] = False
                            # obj = InventorySummaryAPIView()
                            adinventory_type_dict = ui_utils.adinventory_func()
                            duration_type_dict = ui_utils.duration_type_func()

                            if society_inventory_obj.poster_allowed_nb or society_inventory_obj.poster_allowed_lift:
                                society['total_poster_count'] = society_inventory_obj.total_poster_count
                                society['poster_price'] = return_price(adinventory_type_dict, duration_type_dict,
                                                                       'poster_a4', 'campaign_weekly')

                            if society_inventory_obj.standee_allowed:
                                society['total_standee_count'] = society_inventory_obj.total_standee_count
                                society['standee_price'] = return_price(adinventory_type_dict, duration_type_dict,
                                                                        'standee_small', 'campaign_weekly')

                            if society_inventory_obj.stall_allowed:
                                society['total_stall_count'] = society_inventory_obj.total_stall_count
                                society['stall_price'] = return_price(adinventory_type_dict, duration_type_dict,
                                                                      'stall_small', 'unit_daily')
                                society['car_display_price'] = return_price(adinventory_type_dict, duration_type_dict,
                                                                            'car_display_standard', 'unit_daily')

                            if society_inventory_obj.flier_allowed:
                                society['flier_freqency'] = society_inventory_obj.flier_frequency
                                society['filer_price'] = return_price(adinventory_type_dict, duration_type_dict,
                                                                      'flier_door_to_door', 'unit_daily')

                            societies_buffered.append(society)
                            societies_buffered_count += 1

                        space_info_dict['societies_buffered'] = societies_buffered
                        space_info_dict['societies_buffered_count'] = societies_buffered_count

                        societies_inventory = InventoryTypeVersion.objects.get(supplier_code='RS', \
                                                                               space_mapping_version=space_mapping_version_object)
                        societies_inventory_serializer = InventoryTypeVersionSerializer(societies_inventory)
                        societies_inventory_count = InventorySummary.objects.filter(
                            supplier_id__in=societies_shortlisted_ids).aggregate(posters=Sum('total_poster_count'), \
                                                                                 standees=Sum('total_standee_count'),
                                                                                 stalls=Sum('total_stall_count'),
                                                                                 fliers=Sum('flier_frequency'))

                        # Count only for society_shortlisted
                        space_info_dict['societies_inventory_count'] = societies_inventory_count
                        space_info_dict['societies_inventory'] = societies_inventory_serializer.data

                    if space_mapping_version_object.corporate_allowed:
                        # ADDNEW -->
                        pass

                    if space_mapping_version_object.gym_allowed:
                        # ADDNEW -->
                        pass

                    if space_mapping_version_object.salon_allowed:
                        # ADDNEW -->
                        pass

                    center_versions_list.append(space_info_dict)

                proposal_info_dict['centers'] = center_versions_list
                proposal_versions_list.append(proposal_info_dict)

            return Response(proposal_versions_list, status=200)
        except ObjectDoesNotExist as e:
            Response({'status': False, 'error': e.message}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            Response({'status': False, 'error': e.message}, status=status.HTTP_400_BAD_REQUEST)


class HashtagImagesViewSet(viewsets.ViewSet):
    """
    This class is arround hashtagged images by audit app
    """

    def create(self, request):
        """

        :param request:
        :return:
        """
        class_name = self.__class__.__name__
        try:
            data = request.data.copy()
            supplier_type_code = request.data.get('supplierTypeCode')
            response = ui_utils.get_content_type(supplier_type_code)
            if not response:
                return response
            content_type = response.data.get('data')
            data['content_type'] = content_type.id
            serializer = HashtagImagesSerializer(data=data)
            if serializer.is_valid():
                serializer.save()
                return ui_utils.handle_response(class_name, data=serializer.data, success=True)
            return ui_utils.handle_response(class_name, data=serializer.errors)
        except Exception as e:
            return ui_utils.handle_response(class_name, exception_object=e, request=request)

    def list(self, request):
        class_name = self.__class__.__name__
        try:
            campaign_id = request.query_params.get('campaign_id')
            supplier_id = request.query_params.get('supplier_id')

            try:
                images = HashTagImages.objects.filter(campaign=campaign_id,object_id=supplier_id).values()
            except ObjectDoesNotExist:
                return ui_utils.handle_response(class_name, data={}, success=True)
            return ui_utils.handle_response(class_name, data=images, success=True)

        except Exception as e:
            logger.exception
            return ui_utils.handle_response(class_name, exception_object=e, request=request)

    @detail_route(methods=['POST'])
    def upload_permission_box_image(self, request, pk):
        class_name = self.__class__.__name__
        try:
            file = request.data['file']
            campaign_id = pk
            extension = file.name.split('.')[-1]
            campaign_name = request.data['campaign_name'].replace(' ', '_')
            supplier_name = request.data['supplier_name'].replace(' ', '_')
            response = ui_utils.get_content_type(request.data['supplier_type_code'])
            if not response:
                return response
            content_type = response.data.get('data')

            file_name = campaign_name + '_' + supplier_name + '_' + 'permission_box_' + str(
                time.time()).replace('.', '_')+ "_" + ''.join(random.SystemRandom().choice(string.ascii_uppercase + string.digits) for _ in range(6)) +  '.' + extension
            website_utils.upload_to_amazon(file_name, file_content=file, bucket_name=settings.ANDROID_BUCKET_NAME)
            data = HashTagImages(**{
                "campaign_id" : campaign_id,
                "hashtag" : request.data['hashtag'],
                "object_id" : request.data['object_id'],
                "comment" : request.data['comment'],
                "content_type" : content_type,
                "image_path" : file_name
            })
            data.save()
            image = data.image_path
            return ui_utils.handle_response(class_name, data={"image_path" : image}, success=True)
        except Exception as e:
            return ui_utils.handle_response(class_name, exception_object=e, request=request)

    @list_route(methods=['GET'])
    def get_permission_box_images(self, request):
        class_name = self.__class__.__name__
        try:
            campaign_id = request.query_params.get("campaign_id")
            supplier_id = request.query_params.get("supplier_id")
            images = HashTagImages.objects.filter(campaign_id=campaign_id,object_id=supplier_id,hashtag='Permission Box')
            serializer = HashtagImagesSerializer(images, many=True)
            return ui_utils.handle_response(class_name, data=serializer.data, success=True)
        except Exception as e:
            return ui_utils.handle_response(class_name, exception_object=e, request=request)


    @detail_route(methods=['POST'])
    def upload_receipt_image(self, request, pk):
        class_name = self.__class__.__name__
        try:
            file = request.data['file']
            campaign_id = pk
            extension = file.name.split('.')[-1]
            campaign_name = request.data['campaign_name'].replace(' ', '_')
            supplier_name = request.data['supplier_name'].replace(' ', '_')
            response = ui_utils.get_content_type(request.data['supplier_type_code'])
            if not response:
                return response
            content_type = response.data.get('data')

            file_name = campaign_name + '_' + supplier_name + '_' + 'receipt_' + str(
                time.time()).replace('.', '_') + "_" + ''.join(
                random.SystemRandom().choice(string.ascii_uppercase + string.digits) for _ in range(6)) + '.' + extension
            website_utils.upload_to_amazon(file_name, file_content=file, bucket_name=settings.ANDROID_BUCKET_NAME)
            data = HashTagImages(**{
                "campaign_id": campaign_id,
                "hashtag": request.data['hashtag'],
                "object_id": request.data['object_id'],
                "comment": request.data['comment'],
                "content_type": content_type,
                "image_path": file_name
            })
            data.save()
            image = data.image_path
            return ui_utils.handle_response(class_name, data={"image_path" : image}, success=True)
        except Exception as e:
            return ui_utils.handle_response(class_name, exception_object=e, request=request)


    @list_route(methods=['GET'])
    def get_receipt_images(self, request):
        class_name = self.__class__.__name__
        try:
            campaign_id = request.query_params.get("campaign_id")
            supplier_id = request.query_params.get("supplier_id")
            images = HashTagImages.objects.filter(campaign_id=campaign_id, object_id=supplier_id, hashtag='RECEIPT')
            serializer = HashtagImagesSerializer(images, many=True)
            return ui_utils.handle_response(class_name, data=serializer.data, success=True)
        except Exception as e:
            return ui_utils.handle_response(class_name, exception_object=e, request=request)

    @list_route(methods=['GET'])
    def get_hashtag_images(self, request):
        class_name = self.__class__.__name__
        try:
            campaign_id = request.query_params.get("campaign_id")

            if not campaign_id:
                return ui_utils.handle_response(class_name, data='Please pass campaign Id', success=False)
            images = HashTagImages.objects.filter(campaign_id=campaign_id, hashtag__in=['Permission Box','RECEIPT']).order_by('-updated_at')
            if not images:
                return ui_utils.handle_response(class_name, data='No images found', success=True) 
            result_obj = {}
            for image in images:
                image.hashtag = image.hashtag.lower()
                if image.hashtag == 'permission box':
                    image.hashtag = 'permission_box'
                if image.object_id not in result_obj:
                    result_obj[image.object_id] = {}
                if image.hashtag not in result_obj[image.object_id]:
                    result_obj[image.object_id][image.hashtag] = {}
                result_obj[image.object_id][image.hashtag]["image_path"] = image.image_path
                result_obj[image.object_id][image.hashtag]["object_id"] = image.object_id
                result_obj[image.object_id][image.hashtag]["hashtag"] = image.hashtag
                result_obj[image.object_id][image.hashtag]["updated_at"] = image.updated_at
                result_list = [result_obj[result] for result in result_obj]
                
            return ui_utils.handle_response(class_name, data=result_list, success=True)
        except Exception as e:
            return ui_utils.handle_response(class_name, exception_object=e, request=request)

def upload_hashtag_images(data):
    function_name = upload_hashtag_images.__name__
    try:
        file = data['file']
        data = data
        data_dict = {}
        data_dict['content_type'] = 46
        data_dict['object_id'] = data['supplier_id']
        data_dict['hashtag'] = data['hashtag'].upper()
        data_dict['comment'] = data['comment'] if data['comment'] else None
        data_dict['latitude'] = data['lat'] if 'lat' in data else None
        data_dict['longitude'] = data['long'] if 'long' in data else None
        data_dict['campaign'] = data['campaign_id']
        campaign_name = data['campaign_name'].replace(" ", "_")
        supplier_name = data['supplier_name'].replace(" ", "_")
        extension = file.name.split('.')[-1]
        image_path = campaign_name + '_' + supplier_name + '_' + data_dict['hashtag'] + str(
            time.time()).replace('.', '_') + "_" + ''.join(
            random.SystemRandom().choice(string.ascii_uppercase + string.digits) for _ in
            range(6)) + '.' + extension
        website_utils.upload_to_amazon(image_path, file_content=file, bucket_name=settings.ANDROID_BUCKET_NAME)
        data_dict['image_path'] = image_path
        serializer = HashtagImagesSerializer(data=data_dict)
        if serializer.is_valid():
            serializer.save()
            return ui_utils.handle_response(function_name, data=serializer.data, success=True)
        return ui_utils.handle_response(function_name, data=serializer.errors)
    except Exception as e:
        return ui_utils.handle_response(function_name, exception_object=e, request=data)


class HashtagImagesNewViewSet(APIView):
    """
    This class is arround hashtagged images by audit app
    """

    def post(self, request):
        """

        :param request:
        :return:
        """
        class_name = self.__class__.__name__

        response = upload_hashtag_images(request.data)
        if response.data["status"]:
            return ui_utils.handle_response(class_name, data=response.data, success=True)
        return ui_utils.handle_response(class_name, data=response.data)



class CreateFinalProposal(APIView):
    """
    The request is in form:
        [
             {
                  center : { id : 1 , center_name: c1, ...   } ,
                  suppliers:  { 'RS' : [ { 'supplier_type_code': 'RS', 'status': 'R', 'supplier_id' : '1'}, {...}, {...}  }
                  suppliers_meta: {
                                     'RS': { 'inventory_type_selected' : [ 'PO', 'POST', 'ST' ]  },
                                     'CP': { 'inventory_type_selected':  ['ST']
                  }
             }
        ]
    This is second step for creating proposal.
    The proposal_id in the request is always a brand new proposal_id wether you hit this API from an EDIT form of
    the proposal or you are creating an entirely new proposal.

    structure of request.data is  a list. item of the list is the one center information. inside center
    information we have all the suppliers shortlisted, all the Filters and all.

    """

    permission_classes = (v0_permissions.IsProposalAuthenticated,)

    def post(self, request, proposal_id):
        """
        Args:
            request: request data
            proposal_id: proposal_id to be updated

        Returns: success if data is saved successfully.
        """
        class_name = self.__class__.__name__
        try:
            # when save button is clicked, we have already saved space status before hand on individual API call. Only filters are saved here.
            website_utils.setup_create_final_proposal_post(request.data, proposal_id, delete_and_save_filter_data=True,
                                                           delete_and_save_spaces=False)
            return ui_utils.handle_response(class_name, data='success', success=True)
        except Exception as e:
            return ui_utils.handle_response(class_name, exception_object=e, request=request)

    def put(self, request, proposal_id):
        """
        Args:
            request: The request object
            proposal_id: The proposal id

        Returns: updates ShortlistedSpaces table with new data or creates the shortlisted space if not already there. PUT does the job of creating data already.
        """
        class_name = self.__class__.__name__
        try:
            supplier_type_code = request.data['supplier_type_code']

            content_type_response = ui_utils.get_content_type(supplier_type_code)
            if not content_type_response.data['status']:
                return content_type_response
            content_type = content_type_response.data['data']

            data = {
                'center_id': request.data['center_id'],
                'proposal_id': proposal_id,
                'object_id': request.data['supplier_id'],
                'content_type': content_type,
                'supplier_code': supplier_type_code
            }
            status = request.data['status']
            obj, is_created = ShortlistedSpaces.objects.get_or_create(**data)
            obj.status = status
            obj.save()
            return ui_utils.handle_response(class_name, data='success', success=True)
        except Exception as e:
            return ui_utils.handle_response(class_name, exception_object=e, request=request)


class ProposalVersion(APIView):
    """
    The API does following tasks:
    1. creates a new proposal_id and saves all the proposal info in it if user comes from edit proposal.
    2. set's parent to proposal_id received if user comes from edit proposal
    3. saves shortlisted suppliers against the new proposal_id if user comes from edit proposal
    4. saves filter data against the new proposal_id if user comes from edit proposal
    5. sends mail to logged in user
    6. sends mail to BD head.
    7. Generates excel sheet and send's it as attachment in the mail to BD head.
    """

    # permission_classes = (v0_permissions.IsProposalAuthenticated, )

    def post(self, request, proposal_id):
        """
        Args:
            request: The request object
            proposal_id: The proposal_id. ( This is proposal_id for which new version is to be created if is_proposal_version_created is True

        Returns: success if everything succeeds.

        The request is in form:
        [
             {
                  center : { id : 1 , center_name: c1, ...   } ,
                  suppliers: { 'RS' : [ { 'supplier_type_code': 'RS', 'status': 'R', 'supplier_id' : '1'}, {...}, {...} }
                  suppliers_meta: {
                                     'RS': { 'inventory_type_selected' : [ 'PO', 'POST', 'ST' ]  },
                                     'CP': { 'inventory_type_selected':  ['ST']
                  }
             }
        ]

        """
        class_name = self.__class__.__name__
        try:
            user = request.user
            proposal = ProposalInfo.objects.get(proposal_id=proposal_id)
            account = AccountInfo.objects.get(account_id=proposal.account.account_id)
            business = Organisation.objects.get(organisation_id=account.organisation.organisation_id)

            # if you don't provide this value, No proposal version is created.
            is_proposal_version_created = request.data['is_proposal_version_created'] if request.data.get(
                'is_proposal_version_created') else False
            data = request.data['centers']

            # if this variable is true, we will have to create a new proposal version.
            if is_proposal_version_created:
                # create a unique proposal id
                new_proposal_id = website_utils.create_proposal_id(business.organisation_id, account.account_id)

                # create new ProposalInfo object for this new proposal_id
                proposal.pk = new_proposal_id
                proposal.save()

                # change the parent and save again
                proposal.parent = proposal
                proposal.save()
                # change the proposal_id variable here
                proposal_id = new_proposal_id

            # call create Final Proposal first. We need to delete and save filter data. We don't save spaces here. The single status of each space has already been done before on Grid View.
            website_utils.setup_create_final_proposal_post(data, proposal_id, delete_and_save_filter_data=True,
                                                           delete_and_save_spaces=False)

            result = website_utils.setup_generic_export(data, request.user, proposal_id)
            file_name = result['name']
            stats = result['stats']
            # todo : either we can check stats and if it's not empty we can throw error here and not allow sheet formation
            # or we can allow sheet formation and let the user know that such errors have occurred. Fix when it's clear

            bd_body = {
                'user_name': request.user.first_name,
                'business': business.name,
                'account': account.name,
                'proposal_id': proposal_id,
                'file_name': file_name
            }

            bd_body = website_utils.process_template(v0_constants.bodys['bd_head'], bd_body)

            email_data = {
                'subject': v0_constants.subjects['bd_head'],
                'body': bd_body,
                'to': ['dev@machadalo.com', 'anmol.prabhu@machadalo.com', 'anupam@machadalo.com', 'abhishek.chandel@machadalo.com', 'sarvesh.patel@synergytop.com', 'akshay.sahu@synergytop.com', 'bheema.gowda@machadalo.com', 'shyamlee.khanna@machadalo.com']
            }

            #  email_data = {
            #     'subject': v0_constants.subjects['bd_head'],
            #     'body': bd_body,
            #     'to': [v0_constants.emails['developer'], ]
            # }

            attachment = {
                'filepath': file_name,
                'mime_type': v0_constants.mime['xlsx']
            }

            # upload this shit to amazon
            upload_to_amazon_aync_id = tasks.upload_to_amazon.delay(file_name).id

            # send mail to Bd Head with attachment
            bd_head_async_id = send_email.delay(email_data, attachment=attachment).id

            # send mail to logged in user without attachment
            email_data = {
                'subject': v0_constants.subjects['agency'],
                'body': v0_constants.bodys['agency'],
                'to': [user.email]
            }

            logged_in_user_async_id = send_email.delay(email_data).id

            # prepare to send back async ids
            data = {
                'logged_in_user_async_id': logged_in_user_async_id,
                'bd_head_async_id': bd_head_async_id,
                'upload_to_amazon_async_id': upload_to_amazon_aync_id,
                'file_name': file_name,
                'stats': stats
            }
            # change campaign state
            proposal.campaign_state = v0_constants.proposal_requested
            proposal.save()

            # change the status of the proposal to 'requested' once everything is okay.
            data['stats']['inventory_summary_no_instance_error'] = list(data['stats']['inventory_summary_no_instance_error'])
            return ui_utils.handle_response(class_name, data=data, success=True)
        except Exception as e:
            return ui_utils.handle_response(class_name, exception_object=e, request=request)


class ProposalToCampaign(APIView):
    """
    sets the campaign state to right state that
    marks this proposal a campaign.
    """

    def post(self, request, proposal_id):
        """
        Args:
            request:
            proposal_id:

        Returns:

        """
        class_name = self.__class__.__name__
        try:

            proposal = ProposalInfo.objects.get(proposal_id=proposal_id)

            if not proposal.invoice_number:
                return ui_utils.handle_response(class_name, data=errors.CAMPAIGN_NO_INVOICE_ERROR, request=request)

            proposal_start_date = proposal.tentative_start_date
            proposal_end_date = proposal.tentative_end_date

            if not proposal_start_date or not proposal_end_date:
                return ui_utils.handle_response(class_name, data=errors.NO_DATES_ERROR.format(proposal_id),
                                                request=request)

            # todo: disabling this check. Now user can press accept as many times as required.
            #
            # response = website_utils.is_campaign(proposal)
            # if response.data['status']:
            #     return ui_utils.handle_response(class_name, data=errors.ALREADY_A_CAMPAIGN_ERROR.format(proposal.proposal_id), request=request)

            # these are the current inventories assigned. These are inventories assigned to this proposal when sheet was imported.
            # current_assigned_inventories = models.ShortlistedInventoryPricingDetails.objects.select_related('shortlisted_spaces').filter(shortlisted_spaces__proposal_id=proposal_id)

            # assign default dates when you have some inventories assigned
            # if current_assigned_inventories:
            #     current_assigned_inventories_map = {}
            #
            #     for inv in current_assigned_inventories:
            #         inv_tuple = (inv.inventory_content_type, inv.inventory_id)
            #         current_assigned_inventories_map[inv_tuple] = (proposal_start_date, proposal_end_date, inv)
            #
            #     # currently set the R.D and C.D of all inventories to proposal's start and end date.
            #     inventory_release_closure_list = [(inv, proposal_start_date, proposal_end_date) for inv in current_assigned_inventories]
            #     response = website_utils.insert_release_closure_dates(inventory_release_closure_list)
            #     if not response.data['status']:
            #         return response

            # convert to campaign and return
            proposal.campaign_state = v0_constants.proposal_converted_to_campaign
            proposal.save()
            return ui_utils.handle_response(class_name, data=errors.PROPOSAL_CONVERTED_TO_CAMPAIGN.format(proposal_id),
                                            success=True)

            # todo: uncomment this code and modify when date based booking of inventory comes into picture
            # # get all the proposals which are campaign and which overlap with the current campaign
            # response = website_utils.get_overlapping_campaigns(proposal)
            # if not response.data['status']:
            #     return response
            # overlapping_campaigns = response.data['data']
            #
            # if not overlapping_campaigns:
            #     # currently we have no choice but to book all inventories the same proposal_start and end date
            #     # this can be made smarter when we know for how many days a particular inventory  is allowed in a
            #     # supplier this will help in automatically determining R.D and C.D.
            #     inventory_release_closure_list = [(inv, proposal_start_date, proposal_end_date) for inv in current_assigned_inventories]
            #     response = website_utils.insert_release_closure_dates(inventory_release_closure_list)
            #     if not response.data['status']:
            #         return response
            #     proposal.campaign_state = v0_constants.proposal_converted_to_campaign
            #     proposal.save()
            #     return ui_utils.handle_response(class_name,data=errors.PROPOSAL_CONVERTED_TO_CAMPAIGN.format(proposal_id), success=True)
            #
            # already_booked_inventories = models.ShortlistedInventoryPricingDetails.objects.filter(shortlisted_spaces__proposal__in=overlapping_campaigns)
            # already_booked_inventories_map = {}
            #
            # for inv in already_booked_inventories:
            #     inv_tuple = (inv.inventory_content_type, inv.inventory_id)
            #     target_tuple = (inv.release_date, inv.closure_date, inv.shortlisted_spaces.proposal_id)
            #     try:
            #         reference = already_booked_inventories_map[inv_tuple]
            #     except KeyError:
            #         already_booked_inventories_map[inv_tuple] = []
            #         reference = already_booked_inventories_map[inv_tuple]
            #     reference.append(target_tuple)
            #
            # response = website_utils.book_inventories(current_assigned_inventories_map, already_booked_inventories_map)
            # if not response.data['status']:
            #     return response
            # booked_inventories, inventory_release_closure_list, inv_error_list = response.data['data']
            #
            # # if there is something in error list then one or more inventories overlapped with already running campaigns
            # # we do not convert a proposal into campaign in this case
            # if inv_error_list:
            #     return ui_utils.handle_response(class_name, data=errors.CANNOT_CONVERT_TO_CAMPAIGN_ERROR.format(proposal_id, inv_error_list))
            #
            # # insert the RD and CD dates for each inventory
            # response = website_utils.insert_release_closure_dates(inventory_release_closure_list)
            # if not response.data['status']:
            #     return response
            #
            # bulk_update(booked_inventories)
            # proposal.campaign_state = v0_constants.proposal_converted_to_campaign
            # proposal.save()
            #
            # return ui_utils.handle_response(class_name, data=errors.PROPOSAL_CONVERTED_TO_CAMPAIGN.format(proposal_id), success=True)

        except Exception as e:
            return ui_utils.handle_response(class_name, exception_object=e, request=request)


class FinalProposalAPIView(APIView):

    def get(self, request, proposal_id=None, format=None):
        ''' This API sends the data to frontend based on the centers of that proposal and applying
        space filter(society_allowed, corporate_allowed) and there inventories as well
        e.g. (Society --> poster, standee ) will give societies that have both poster and standee allowed
        '''
        try:
            # if proposal_id is None:
            #     proposal_id = 'AlntOlJi';
            proposal_object = ProposalInfo.objects.get(proposal_id=proposal_id)
        except ProposalInfo.DoesNotExist:
            return Response({'message': 'Invalid Proposal ID sent'}, status=406)

        space_dict, supplier_code_dict = self.get_space_code_dict()
        # ADDNEW -->
        space_model_dict = {
            'society': SupplierTypeSociety, 'corporate': SupplierTypeCorporate,
            # 'gym' : SupplierTypeGym,          'salon' : SupplierTypesalon
        }

        space_serializer_dict = {
            'society': ProposalSocietySerializer, 'corporate': ProposalCorporateSerializer,
            # 'gym' : ProposalGymSerializer,        'salon' : ProposalsalonSerializer
        }

        centers_objects = proposal_object.get_centers()
        centers_list = []

        for center_object in centers_objects:
            space_info_dict = {}
            center_serializer = ProposalCenterMappingSpaceSerializer(center_object)
            space_info_dict['center'] = center_serializer.data

            space_mapping_object = center_object.get_space_mappings()

            # for space_name in ['society','corporate','gym','salon']:
            #     if space_mapping_object.__dict__[space_name + '_allowed']:
            #         space_inventory_object = InventoryTypes.objects.get(space_mapping=space_mapping_object, supplier_code=supplier_code_dict[space_name])
            #         space_info_dict[space_dict[space_name] + '_inventory'] = InventoryTypeSerializer(space_inventory_object)

            #         space_ids = ShortlistedSpaces.objects.filter(space_mapping=space_mapping_object, supplier_code=supplier_code_dict[space_name]).values_list('object_id',flat=True)
            #         spaces = space_model_dict[space_name].objects.filter(supplier_id__in=space_ids)
            #         spaces_serializer = space_serializer_dict[space_name](spaces)
            #         space_info_dict[space_dict[space_name]] = spaces_serializer.data

            #         # still have to put inventory count and then done

            #         centers_list.append(space_info_dict)

            if space_mapping_object.society_allowed:
                societies_inventory = space_mapping_object.get_society_inventories()
                societies_inventory_serializer = InventoryTypeSerializer(societies_inventory)

                society_ids = ShortlistedSpaces.objects.filter(space_mapping=space_mapping_object,
                                                               supplier_code='RS').values_list('object_id', flat=True)
                societies_temp = SupplierTypeSociety.objects.filter(supplier_id__in=society_ids).values('supplier_id',
                                                                                                        'society_latitude',
                                                                                                        'society_longitude',
                                                                                                        'society_name',
                                                                                                        'society_address1',
                                                                                                        'society_subarea',
                                                                                                        'society_location_type')
                societies = []
                society_ids = []
                societies_count = 0
                for society in societies_temp:
                    if website_utils.space_on_circle(center_object.latitude, center_object.longitude,
                                                     center_object.radius, \
                                                     society['society_latitude'], society['society_longitude']):
                        society_inventory_obj = InventorySummary.objects.get_supplier_type_specific_object(
                            request.data.copy(),
                            society['supplier_id'])
                        # society_inventory_obj = InventorySummary.objects.get(supplier_id=society['supplier_id'])
                        society['shortlisted'] = True
                        society['buffer_status'] = False
                        # obj = InventorySummaryAPIView()
                        adinventory_type_dict = ui_utils.adinventory_func()
                        duration_type_dict = ui_utils.duration_type_func()
                        if society_inventory_obj.poster_allowed_nb or society_inventory_obj.poster_allowed_lift:
                            society['total_poster_count'] = society_inventory_obj.total_poster_count
                            society['poster_price'] = return_price(adinventory_type_dict, duration_type_dict,
                                                                   'poster_a4', 'campaign_weekly')

                        if society_inventory_obj.standee_allowed:
                            society['total_standee_count'] = society_inventory_obj.total_standee_count
                            society['standee_price'] = return_price(adinventory_type_dict, duration_type_dict,
                                                                    'standee_small', 'campaign_weekly')

                        if society_inventory_obj.stall_allowed:
                            society['total_stall_count'] = society_inventory_obj.total_stall_count
                            society['stall_price'] = return_price(adinventory_type_dict, duration_type_dict,
                                                                  'stall_small', 'unit_daily')
                            society['car_display_price'] = return_price(adinventory_type_dict, duration_type_dict,
                                                                        'car_display_standard', 'unit_daily')

                        if society_inventory_obj.flier_allowed:
                            society['flier_frequency'] = society_inventory_obj.flier_frequency
                            society['filer_price'] = return_price(adinventory_type_dict, duration_type_dict,
                                                                  'flier_door_to_door', 'unit_daily')

                        # ADDNEW --> Banner etc.
                        society_ids.append(society['supplier_id'])
                        societies.append(society)
                        societies_count += 1

                # societies_serializer =  ProposalSocietySerializer(societies, many=True)

                # following query find sum of all the variables specified in a dictionary
                # this finds sum of all inventories, if you don't need some of some inventory make it 0 in front end
                societies_inventory_count = InventorySummary.objects.filter(supplier_id__in=society_ids).aggregate(
                    posters=Sum('total_poster_count'), \
                    standees=Sum('total_standee_count'), stalls=Sum('total_stall_count'), fliers=Sum('flier_frequency'))

                space_info_dict['societies'] = societies
                space_info_dict['societies_inventory_count'] = societies_inventory_count
                space_info_dict['societies_inventory'] = societies_inventory_serializer.data
                space_info_dict['societies_count'] = societies_count

            if space_mapping_object.corporate_allowed:
                # ADDNEW -->
                pass

            if space_mapping_object.gym_allowed:
                # ADDNEW -->
                pass

            centers_list.append(space_info_dict)

        proposal_serializer = ProposalInfoSerializer(proposal_object)
        response = {
            'proposal': proposal_serializer.data,
            'centers': centers_list,
        }

        return Response(response, status=200)

    def post(self, request, proposal_id=None, format=None):
        ''' Saving the proposal from the map view. Every time mapview page is loaded and grid view is submitted from there
        This makes a new version of proposal. And also updates all the required table as well
        This expects id of both center and space mapping from the frontend as they are saved on basic proposal page (InitialProposalAPIView)
        '''

        centers = request.data
        space_dict, supplier_code_dict = self.get_space_code_dict()

        with transaction.atomic():
            try:
                shortlisted_space_list = []
                shortlisted_space_version_list = []
                for center_info in centers:
                    center = center_info['center']
                    center_id = center['id']

                    proposal_version_object_response = website_utils.save_proposal_version(center_id)
                    if not proposal_version_object_response.data['status']:
                        return proposal_version_object_response
                    proposal_version_object = proposal_version_object_response.data['data']

                    center_object = ProposalCenterMapping.objects.select_related('proposal').get(id=center_id)
                    proposal_object = center_object.proposal

                    center_serializer = ProposalCenterMappingSerializer(center_object, data=center)
                    if center_serializer.is_valid():
                        center_object = center_serializer.save()
                    else:
                        return Response({'message': 'Invalid Center Data', 'errors': center_serializer.errors},
                                        status=406)

                    # version save
                    center['proposal_version'] = proposal_version_object.id
                    del center['id']
                    center_version_serailizer = ProposalCenterMappingVersionSerializer(data=center)

                    if center_version_serailizer.is_valid():
                        center_version_object = center_version_serailizer.save()
                    else:
                        return Response(
                            {'message': 'Invalid Center Version Data', 'errors': center_version_serailizer.errors}, \
                            status=406)

                    space_mappings = center['space_mappings']
                    space_mapping_id = space_mappings['id']
                    space_mapping_object = SpaceMapping.objects.get(id=space_mapping_id)

                    space_mapping_serializer = SpaceMappingSerializer(space_mapping_object, data=space_mappings)
                    if space_mapping_serializer.is_valid():
                        space_mapping_object = space_mapping_serializer.save()
                    else:
                        return Response(
                            {'message': 'Invalid Space Mapping Data', 'errors': space_mapping_serializer.errors}, \
                            status=406)

                    # version save
                    space_mappings['center_version'] = center_version_object.id
                    space_mappings['proposal_version'] = proposal_version_object.id
                    del space_mappings['id']
                    space_mapping_version_serializer = SpaceMappingVersionSerializer(data=space_mappings)
                    if space_mapping_version_serializer.is_valid():
                        space_mapping_version_object = space_mapping_version_serializer.save()
                    else:
                        return Response({'message': 'Invalid Space Mapping Version Data',
                                         'errors': space_mapping_version_serializer.errors}, \
                                        status=406)

                    for space_name in ['society', 'corporate', 'gym', 'salon']:

                        if space_mapping_object.__dict__[space_name + "_allowed"]:
                            content_type = ContentType.objects.get(model='SupplierType' + space_name.title())
                            supplier_code = supplier_code_dict[space_name]

                            try:
                                space_inventory_type = center_info[space_dict[space_name] + '_inventory']
                            except KeyError:
                                # Just ignoring because for corporate inventory is not made
                                continue

                            try:
                                inventory_type_object = InventoryType.objects.get(id=space_inventory_type['id'])
                                inventory_type_serializer = InventoryTypeSerializer(inventory_type_object,
                                                                                    data=space_inventory_type)
                                del space_inventory_type['id']
                            except KeyError:
                                space_inventory_type['space_mapping'] = space_mapping_object.id
                                inventory_type_serializer = InventoryTypeSerializer(data=space_inventory_type)

                            if inventory_type_serializer.is_valid():
                                inventory_type_serializer.save()
                            else:
                                return Response({'message': 'Invalid Inventory Details for ' + space_name, 'errors': \
                                    inventory_type_serializer.errors}, status=406)

                            # version save
                            space_inventory_type['space_mapping_version'] = space_mapping_version_object.id
                            inventory_type_version_serializer = InventoryTypeVersionSerializer(
                                data=space_inventory_type)
                            if inventory_type_version_serializer.is_valid():
                                inventory_type_version_serializer.save()
                            else:
                                return Response(
                                    {'message': 'Invalid Inventory Details Version for ' + space_name, ' errors': \
                                        inventory_type_version_serializer.errors}, status=406)

                            space_mapping_object.get_all_spaces().delete()

                            for space in center_info[space_dict[space_name]]:
                                if space.get('shortlisted'):
                                    object_id = space['supplier_id']
                                    shortlisted_space = ShortlistedSpaces(space_mapping=space_mapping_object,
                                                                          content_type=content_type,
                                                                          supplier_code=supplier_code,
                                                                          object_id=object_id,
                                                                          buffer_status=space['buffer_status'])

                                    shortlisted_space_list.append(shortlisted_space)

                                    # version save
                                    shortlisted_version_space = ShortlistedSpacesVersion(
                                        space_mapping_version=space_mapping_version_object,
                                        content_type=content_type, supplier_code=supplier_code, object_id=object_id,
                                        buffer_status=space['buffer_status'])
                                    shortlisted_space_version_list.append(shortlisted_version_space)

                ShortlistedSpaces.objects.bulk_create(shortlisted_space_list)
                # version save
                ShortlistedSpacesVersion.objects.bulk_create(shortlisted_space_version_list)
            except Exception as e:
                return Response({'status': False, 'error': e.message}, status=status.HTTP_400_BAD_REQUEST)

        return Response({'message': 'Successfully Saved'}, status=200)

    # def put(self, request, format=None):
    #     centers = request.data
    #     for center in centers:

    def get_space_code_dict(self):
        space_dict = {  # singular space as key and its plural as value
            'society': 'societies', 'corporate': 'corporates',
            'gym': 'gyms', 'salon': 'salons',
        }

        supplier_code_dict = {
            'society': 'RS', 'corporate': 'CP',
            'gym': 'GY', 'salon': 'SA',
        }

        return space_dict, supplier_code_dict


class FinalProposalSubmit(APIView):
    """
    Saves final proposal
    """

    def post(self, request):
        for center in request.data:
            # create an entry of this proposal in  proposal_info_version table
            pass


class CreateProposalAPIView(APIView):
    def get(self, request, id, format=None):
        try:
            campaign = Campaign.objects.get(pk=id)
            items = campaign.societies.filter(booking_status='Shortlisted')
            inv_types = campaign.types.all()

            response = []
            allowed = '_allowed'
            total = 'total_'
            count = '_count'
            price_dict = getPriceDict()
            for item in items:
                society_detail = {}  # List of society-related details required to be displayed
                society_id = item.society.supplier_id
                society_name = item.society.society_name
                society_detail['id'] = society_id
                society_detail['society_name'] = society_name
                society_detail['flat_count'] = item.society.flat_count
                society_detail['tower_count'] = item.society.tower_count
                society_detail['inventory'] = []
                inv_details = InventorySummary.objects.get_supplier_type_specific_object(request.data.copy(), id)
                if not inv_details:
                    return Response({'status': False, 'error': 'Inventory object not found for {0} id'.format(id)},
                                    status=status.HTTP_400_BAD_REQUEST)

                # inv_details = InventorySummary.objects.get(supplier_id=society_id)
                for inv in inv_types:
                    inv_name = inv.type
                    inv_size = inv.sub_type
                    if (hasattr(inv_details, inv_name.lower() + allowed) and getattr(inv_details,
                                                                                     inv_name.lower() + allowed)):
                        if (inv_name == 'Flier'):
                            inv_count = 1
                        else:
                            inv_count = getattr(inv_details, total + inv_name.lower() + count)
                        inv_info = {}  # List of inventory and its count details

                        # adinventory_type = models.ForeignKey('AdInventoryType', db_column='ADINVENTORY_TYPE_ID', blank=True, null=True, on_delete=models.CASCADE)
                        # supplier = models.ForeignKey('SupplierTypeSociety', db_column='SUPPLIER_ID', related_name='default_prices', blank=True, null=True, on_delete=models.CASCADE)
                        # duration_type = models.ForeignKey('DurationType', db_column='DURATION_ID', blank=True, null=True, on_delete=models.CASCADE)
                        duration_type = DurationType.objects.get(id=int(price_dict[inv_name]['duration']))
                        adinventory_type = AdInventoryType.objects.get(id=int(price_dict[inv_name]['types'][inv_size]))
                        price_obj = PriceMappingDefault.objects.get(supplier=item.society, duration_type=duration_type,
                                                                    adinventory_type=adinventory_type)
                        inv_price = price_obj.business_price
                        inv_info['count'] = str(inv_count)
                        inv_info['price'] = str(inv_price)
                        inv_info['type'] = inv_name
                        society_detail['inventory'].append(inv_info)

                response.append(society_detail)

            # response = (response)
            return Response(response, status=200)

        except:
            return Response(status=404)


class ProposalImagesPath(APIView):
    """
    returns a path to downloaded images in zip format. The zipped file is stored in Media directory
    This API  converts the folder present in 'files' directory into zip file and returns a path to download the file.
    """

    def get(self, request):
        class_name = self.__class__.__name__
        if True:
            proposal_id = request.query_params['proposal_id']
            task_id = request.query_params['task_id']
            result = GroupResult.restore(task_id)
            if not result.successful():
                return ui_utils.handle_response(class_name, data=errors.TASK_NOT_DONE_YET_ERROR.format(task_id))

            path_to_master_dir = settings.BASE_DIR + '/files/downloaded_images/' + proposal_id
            output_path = os.path.join(settings.MEDIA_ROOT, proposal_id)

            if not os.path.exists(path_to_master_dir):
                return ui_utils.handle_response(class_name, data=errors.PATH_DOES_NOT_EXIST.format(path_to_master_dir))

            shutil.make_archive(output_path, 'zip', path_to_master_dir)
            file_url = settings.BASE_URL + proposal_id + '.zip'
            # we should remove the original folder as it will consume space.
            # os.chmod(path_to_master_dir, 777)
            # shutil.rmtree(path_to_master_dir)
            os.system("rm -rf "+path_to_master_dir)
            return ui_utils.handle_response(class_name, data=file_url, success=True)
        # except Exception as e:
        #     return ui_utils.handle_response(class_name, exception_object=e, request=request)


class CampaignToProposal(APIView):
    """
    Releases the inventories booked under a campaign and the sets the campaign state back to proposal state.
    """

    def post(self, request, campaign_id):
        """
        Args:
            request:
            campaign_id: The campaign_id

        Returns:
        """
        class_name = self.__class__.__name__
        try:
            proposal = ProposalInfo.objects.get(proposal_id=campaign_id)

            response = website_utils.is_campaign(proposal)
            if not response.data['status']:
                return response

            proposal.campaign_state = v0_constants.proposal_not_converted_to_campaign
            proposal.save()

            current_assigned_inventories = ShortlistedInventoryPricingDetails.objects.select_related(
                'shortlisted_spaces').filter(shortlisted_spaces__proposal_id=campaign_id)
            InventoryActivityAssignment.objects.filter(
                inventory_activity__shortlisted_inventory_details__in=current_assigned_inventories).delete()

            return ui_utils.handle_response(class_name, data=errors.REVERT_CAMPAIGN_TO_PROPOSAL.format(campaign_id,
                                                                                                       v0_constants.proposal_not_converted_to_campaign),
                                            success=True)
        except Exception as e:
            return ui_utils.handle_response(class_name, exception_object=e, request=request)


class convertDirectProposalToCampaign(APIView):
    """
    This will convert proposal to campaign where supplier id's should be provided in sheet
    """

    def post(self, request):
        """

        :param request:
        :return:
        """
        class_name = self.__class__.__name__
        try:
            data = request.data.copy()

            is_import_sheet = data['is_import_sheet']

            if is_import_sheet:
                response = genrate_supplier_data(data,request.user)
                if not response.data['status']:
                    return response
                proposal_data = response.data['data']
            else:
                proposal_data = data
            center_id = proposal_data['center_id']
            proposal = ProposalInfo.objects.get(pk=proposal_data['proposal_id'])
            center = ProposalCenterMapping.objects.get(pk=center_id)
            for supplier_code in proposal_data['center_data']:
                response = website_utils.save_filters(center, supplier_code, proposal_data, proposal)
                if not response.data['status']:
                    return response

                response = website_utils.save_shortlisted_suppliers_data(center, supplier_code, proposal_data, proposal)
                if not response.data['status']:
                    return response
                if is_import_sheet:
                    create_inv_act_data = True
                    response = website_utils.save_shortlisted_inventory_pricing_details_data(center, supplier_code,
                                                                 proposal_data, proposal,create_inv_act_data)

                    response = assign_inv_dates(proposal_data)
                else:
                    response = website_utils.save_shortlisted_inventory_pricing_details_data(center, supplier_code,
                                                                                         proposal_data, proposal)
                if not response.data['status']:
                    return response
            response = website_utils.update_proposal_invoice_and_state(proposal_data, proposal)
            if not response.data['status']:
                return response
            response = website_utils.create_generic_export_file_data(proposal)
            if not response.data['status']:
                return response

            return ui_utils.handle_response(class_name, data={}, success=True)
        except Exception as e:
            return ui_utils.handle_response(class_name, exception_object=e, request=request)


def strToList(str):
    elements = str.split(',')
    converted_list = [x.strip(' ') for x in elements]
    return converted_list


def convertDirectProposalToCampaignExcel(proposal_list):
    for proposal_data in proposal_list:
        center_id = proposal_data['center_id']
        proposal = ProposalInfo.objects.get(pk=proposal_data['proposal_id'])
        center = ProposalCenterMapping.objects.get(pk=center_id)

        # data conversion
        center_data = {}


class SupplierPhaseViewSet(viewsets.ViewSet):

    def list(self, request):
        class_name = self.__class__.__name__
        try:
            campaign_id = request.query_params.get('campaign_id')

            phases = SupplierPhase.objects.filter(campaign=campaign_id)
            current_date = datetime.datetime.now().date()
            result_obj = {}
            for phase in phases:
                if not (phase.end_date and phase.start_date):
                    continue
                if phase.id not in result_obj:
                    result_obj[phase.id] = {}
                result_obj[phase.id]["start_date"] = phase.start_date
                result_obj[phase.id]["end_date"] = phase.end_date
                result_obj[phase.id]["phase_no"] = phase.phase_no
                result_obj[phase.id]["created_at"] = phase.created_at
                result_obj[phase.id]["id"] = phase.id
                result_obj[phase.id]["comments"] = phase.comments
                result_obj[phase.id]["campaign"] = campaign_id

                if not phase.end_date or not phase.start_date:
                    result_obj[phase.id]["status"] = "Phase not defined"
                else:
                    if current_date > phase.end_date.date():
                        result_obj[phase.id]["status"] = "completed"
                    elif phase.start_date.date() > current_date:
                        result_obj[phase.id]["status"] = "upcoming"
                    elif phase.start_date.date() < current_date < phase.end_date.date():
                        result_obj[phase.id]["status"] = "ongoing"

            result_list = [result_obj[result] for result in result_obj]

            return ui_utils.handle_response(class_name, data=result_list, success=True)

        except Exception as e:
            return handle_response(class_name, exception_object=e, request=request)

    def create(self, request):
        class_name = self.__class__.__name__
        try:
            phases = request.data
            campaign_id = request.query_params.get('campaign_id')
            if not campaign_id or not phases:
                return handle_response(class_name, data='Please provide campaign id or phases', success=False)

            for phase in phases:
                if 'id' in phase:
                    phase_already_exists = SupplierPhase.objects.filter(campaign=campaign_id, phase_no=phase['phase_no']).exclude(id=phase['id'])
                    if phase_already_exists:
                        return handle_response(class_name, data='Phase No. already exists', success=False)

                    item = SupplierPhase.objects.get(pk=phase['id'],campaign=campaign_id)
                    phase_serializer = SupplierPhaseSerializer(item, data=phase)
                else:
                    phase_already_exists = SupplierPhase.objects.filter(campaign=campaign_id, phase_no=phase['phase_no'])
                    if phase_already_exists:
                        return handle_response(class_name, data='Phase No. already exists', success=False)
                    
                    phase['campaign'] = campaign_id
                    phase_serializer = SupplierPhaseSerializer(data=phase)
                if phase_serializer.is_valid():
                    phase_serializer.save()
            data = SupplierPhase.objects.filter(campaign=campaign_id)
            serializer = SupplierPhaseSerializer(data, many=True)
            return handle_response(class_name, data=serializer.data, success=True)
        except Exception as e:
            return handle_response(class_name, exception_object=e, request=request)

    def destroy(self, request, pk):
        """

        :param request:
        :param pk:
        :return:
        """
        class_name = self.__class__.__name__
        try:
            SupplierPhase.objects.get(pk=pk).delete()
            return ui_utils.handle_response(class_name, data=True, success=True)
        except ValueError:
            return ui_utils.handle_response(class_name, data='Phase is required getting undefined')
        except Exception as e:
            return ui_utils.handle_response(class_name, exception_object=e, request=request)


def flatten_list(list_of_lists):
    return [y for x in list_of_lists for y in x]


def get_supplier_list_by_status_ctrl(campaign_id):
    shortlisted_spaces_list = ShortlistedSpaces.objects.filter(proposal_id=campaign_id)
    shortlisted_spaces_by_phase_dict = {}

    all_phases = SupplierPhase.objects.filter(campaign_id=campaign_id).all()
    all_ss_comments = CampaignComments.objects.filter(campaign_id=campaign_id, related_to='EXTERNAL').all()
    all_ss_comments_dict = {}

    for single_ss_comment in all_ss_comments:
        if single_ss_comment.shortlisted_spaces_id not in all_ss_comments_dict:
            all_ss_comments_dict[single_ss_comment.shortlisted_spaces_id] = []
        if not single_ss_comment.inventory_type:
            all_ss_comments_dict[single_ss_comment.shortlisted_spaces_id].append(single_ss_comment.comment)
    all_phase_by_id = {}

    current_date = datetime.datetime.now().date()
    for phase in all_phases:
        all_phase_by_id[phase.id] = {'start_date': phase.start_date,
                                     'end_date': phase.end_date,
                                     'phase_no': phase.phase_no,
                                     'comments': phase.comments
                                     }

    overall_inventory_count_dict = {}

    no_phase_suppliers = []
    no_status_suppliers = []
    all_supplier_ids = list(set([space.object_id for space in shortlisted_spaces_list]))

    all_supplier_objects = SupplierMaster.objects.filter(supplier_id__in=all_supplier_ids)
    all_supplier_data_serializer = SupplierMasterSerializer(all_supplier_objects,many=True).data
    all_supplier_data = website_utils.manipulate_master_to_rs(all_supplier_data_serializer)
    if len(all_supplier_data) < 1:
        all_supplier_objects = SupplierTypeSociety.objects.filter(supplier_id__in=all_supplier_ids)
        all_supplier_data_serializer = SupplierTypeSocietySerializer(all_supplier_objects,many=True).data
        all_supplier_data = website_utils.manipulate_object_key_values_generic(all_supplier_data_serializer)

    all_supplier_dict = {supplier['supplier_id']:supplier for supplier in all_supplier_data}
    for space in shortlisted_spaces_list:
        try:
            supplier_society_serialized = all_supplier_dict[space.object_id]
        except KeyError:
            pass
        supplier_inventories = ShortlistedInventoryPricingDetails.objects.filter(shortlisted_spaces_id=space.id)
        inventory_activity_assignment = InventoryActivityAssignment.objects.filter(
            inventory_activity__shortlisted_inventory_details__shortlisted_spaces_id=space.id)\
            .values('activity_date',
              'inventory_activity__activity_type',
              'inventory_activity__shortlisted_inventory_details__ad_inventory_type__adinventory_name'
            )
        inventory_dates_dict = {
                    "POSTER": [],
                    "POSTER LIFT": [],
                    "STALL": [],
                    "STANDEE": [],
                    "FLIER": [],
                    "BANNER": [],
                    "GATEWAY ARCH": [],
                    "SUNBOARD" : [],
                    "BILLING" : [],
                    "HOARDING" : [],
                    "GANTRY" : [],
                    "BUS SHELTER" : [],
                    "BUS BACK" : [],
                    "BUS RIGHT" : [],
                    "BUS LEFT" : [],
                    "BUS WRAP" : [],
                    "FLOOR" : [],
                    "CEILING" : [],
                    "COUNTER DISPLAY" : [],
                    "TENT CARD" : [],
                    "TABLE" : [],
                    "HOARDING LIT" : [],
                    "BUS SHELTER LIT" : [],
                    "GANTRY LIT" : [],
                    "WALL" : [],
                    "CAR DISPLAY" : [],
                    "WHATSAPP INDIVIDUAL": [],
                    "WHATSAPP GROUP": []
                }
        inventory_days_dict = {
            "POSTER": None,
            "POSTER LIFT": None,
            "STALL": None,
            "STANDEE": None,
            "FLIER": None,
            "BANNER": None,
            "GATEWAY ARCH": None,
            "SUNBOARD" : None,
            "BILLING" : None,
            "HOARDING" : None,
            "GANTRY" : None,
            "BUS SHELTER" : None,
            "BUS BACK" : None,
            "BUS RIGHT" : None,
            "BUS LEFT" : None,
            "BUS WRAP" : None,
            "FLOOR" : None,
            "CEILING" : None,
            "COUNTER DISPLAY" : None,
            "TENT CARD" : None,
            "TABLE" : None,
            "HOARDING LIT" : None,
            "BUS SHELTER LIT" : None,
            "GANTRY LIT" : None,
            "WALL" : None,
            "CAR DISPLAY" : None,
            "WHATSAPP INDIVIDUAL": None,
            "WHATSAPP GROUP": None
        }
        for inventory_activity in inventory_activity_assignment:
            inventoy_name = inventory_activity.get('inventory_activity__shortlisted_inventory_details__ad_inventory_type__adinventory_name')
            activity_date = inventory_activity.get('activity_date')
            if activity_date and inventoy_name:
                activity_date = activity_date.strftime('%d %b %Y')
                if activity_date not in inventory_dates_dict[inventoy_name]:
                    inventory_dates_dict[inventoy_name].append(activity_date)
        inventory_count_dict = {}

        supplier_tower_count = supplier_society_serialized["tower_count"] if supplier_society_serialized.get("tower_count") else 0
        supplier_flat_count = supplier_society_serialized["flat_count"] if supplier_society_serialized.get("flat_count") else 0
        for inventory in supplier_inventories:
            if inventory.ad_inventory_type.adinventory_name not in inventory_count_dict:
                inventory_count_dict[inventory.ad_inventory_type.adinventory_name] = 0
            if inventory.ad_inventory_type.adinventory_name not in overall_inventory_count_dict:
                overall_inventory_count_dict[inventory.ad_inventory_type.adinventory_name] = 0

            inventory_count_dict[inventory.ad_inventory_type.adinventory_name] += 1
            overall_inventory_count_dict[inventory.ad_inventory_type.adinventory_name] += 1
            if inventory.inventory_number_of_days:
                inventory_days_dict[inventory.ad_inventory_type.adinventory_name] = inventory.inventory_number_of_days
            inventory_count_dict['FLIER'] = supplier_society_serialized["flat_count"] if supplier_society_serialized.get("flat_count") else 0
            overall_inventory_count_dict['FLIER'] = supplier_society_serialized["flat_count"] if supplier_society_serialized.get("flat_count") else 0

        supplier_society_serialized['booking_status'] = space.booking_status
        supplier_society_serialized['next_action_date'] = space.next_action_date.strftime('%Y/%m/%d') if space.next_action_date else None
        supplier_society_serialized['booking_sub_status'] = space.booking_sub_status
        supplier_society_serialized['freebies'] = space.freebies.split(",") if space.freebies else None
        supplier_society_serialized['stall_locations'] = space.stall_locations.split(",") if space.stall_locations else None
        supplier_society_serialized['comments'] = all_ss_comments_dict[space.id] if space.id  in all_ss_comments_dict else None
        supplier_society_serialized['space_id'] = space.id
        supplier_society_serialized['inventory_counts'] = inventory_count_dict
        supplier_society_serialized['inventory_dates'] = inventory_dates_dict
        supplier_society_serialized['inventory_days'] = inventory_days_dict

        if not space.phase_no_id:
            if space.booking_status:
                no_phase_suppliers.append(supplier_society_serialized)
            else:
                no_status_suppliers.append(supplier_society_serialized)
        elif space.phase_no_id:
            if space.phase_no_id not in shortlisted_spaces_by_phase_dict:
                shortlisted_spaces_by_phase_dict[space.phase_no_id] = {'BK': [], 'NB': [], 'PB': [], 'VB': [], 'SR': [], 'NI':[], 'MF':[], 'RERA':[], 'MWS':[], 'MWC':[], 'MWT':[], 'MWO':[],
                                                                       'SE': [], 'VR': [], 'CR': [], 'DPCR':[], 'DPNR':[], 'NE':[], 'UN':[], 'MWA':[], 'UPNI':[], 'UCPI':[], 'TB':[],
                                                                       'DP': [], 'TB': [], 'MC':[], 'UN':[], 'DPRR': [], 'RLC':[], 'PR': [], 'NVOS':[], 'RE': [], 'RERR':[], 'RERA':[], 'BSR': [], 'BDP': [],
                                                                       'MWA': [], 'MWS': [], 'MWC': [], 'MWT': [], 'MWO': [], 'DPNR': [], 'DPNA': [], 'DPP': [], 'DPSOO': [], 'DPOS': [], 'DPVR': [], 'DPCR': [],
                                                                       'RLO': [], 'RLC': [], 'RUB': [], 'RVE': [], 'RCR': [], 'RRS': [], 'ROS': [],'BRLO': [], 'BRLC': [], 'BRUB': [], 'BRVE': [], 'BUPNI': [], 'BUCNI':[],
                                                                       'BRCR': [], 'BRRS': [],'BROS': [],'BDPNR': [], 'BDPNA': [], 'BDPP': [], 'BDPSOO': [], 'BDPOS': [], 'BDPVR': [], 'BDPCR': [], 'BUN': [], 'BUCPI':[],
                                                                       'BNE':[], 'BNVW':[], 'BNVG':[], 'BNVA':[], 'BNVMB':[], 'BNVFT':[], 'BNVOS':[], 'UPNI':[], 'UCPI':[], 'NVW':[], 'NVG':[], 'NVA':[], 
                                                                       'NVMB':[], 'NVFT':[], 'NVOS':[], 'BNI':[],                                            
                                                                       'OEL': [], 'ESVF': [], 'ESMD': [], 'ESGR': [], 'ESOE': [], 'ESNR': [],'OCL': [], 'CLVF': [],'CLMD': [], 'CLGR': [], 'CLOE': [], 'CLNR': [],
                                                                       'OPBL': [], 'PBVF': [], 'PBMD': [], 'PBGR': [], 'PBOE': [], 'PBNR': [], 'OPFL': [], 'PFVF': [], 'PFMD': [], 'PFGR': [], 'PFOE': [], 'PFNR': [],
                                                                       'OPHL': [], 'PHVF': [], 'PHMD': [], 'PHGR': [], 'PHOE': [], 'PHNR': [], 'OP': [], 'OVF': [], 'OMD': [], 'OGR': [], 'OOE': [], 'ONR': [] }                                                                    
            if space.booking_status:                
                # if space.booking_sub_status:
                #     shortlisted_spaces_by_phase_dict[space.phase_no_id][space.booking_sub_status].append(
                #         supplier_society_serialized)
                # else:
                #     shortlisted_spaces_by_phase_dict[space.phase_no_id][space.booking_status].append(
                #     supplier_society_serialized)
                # else
                if space.booking_status:
                    if not shortlisted_spaces_by_phase_dict[space.phase_no_id].get(space.booking_status):
                        shortlisted_spaces_by_phase_dict[space.phase_no_id][space.booking_status] = []
                    
                    shortlisted_spaces_by_phase_dict[space.phase_no_id][space.booking_status].append(
                    supplier_society_serialized)

    shortlisted_spaces_by_phase_list = []
    ongoing_phase = None
    upcoming_phases = []
    completed_phases = []
    confirmed_booked_status = ['BK']
    unknown = ['BUN', 'BUPNI', 'BUCPI', 'UN', 'UPNI', 'UCPI']
    new_entity = ['NE', 'NVW', 'NVG', 'NVA', 'NVMB', 'NVFT', 'NVOS', 'BNE', 'BNVW', 'BNVG', 'BNVA', 'BNVMB', 'BNVFT', 'BNVOS']
    # not_initiated = ['NI']
    meeting_fixed = ['MF', 'MWA', 'MWS', 'MWC', 'MWT', 'MWO']
    meeting_converted = ['MC']
    verbally_booked_status = ['TB', 'VB', 'PB', 'RE', 'RERR', 'RERA']
    followup_req_status = ['DP', 'DPNR', 'DPNA', 'DPP', 'DPSOO', 'DPOS', 'SE', 'DPVR', 'DPCR']
    not_initiated_status = ['NB', 'NI', 'BNI']
    btoc_rejected_status = ['SR', 'RLO', 'RLC', 'RUB', 'RVE', 'RCR', 'RRS', 'ROS']
    btob_rejected_status = ['BSR', 'BRLO', 'BRLC', 'BRUB', 'BRVE', 'BRCR', 'BRRS', 'BROS']
    decision_pending_status = ['BDP', 'BDPNR', 'BDPNA', 'BDPP', 'BDPSOO', 'BDPOS', 'BDPVR', 'BDPCR']
    recce_required = ['DPRR']
    emergency_situation_status = ['OEL', 'ESVF', 'ESMD', 'ESGR', 'ESOE', 'ESNR']
    complete_lockdown_status = ['OCL', 'CLVF','CLMD', 'CLGR', 'CLOE', 'CLNR']
    part_building_lock_status = ['OPBL', 'PBVF', 'PBMD', 'PBGR', 'PBOE', 'PBNR']
    part_floor_lock_status =['OPFL', 'PFVF', 'PFMD', 'PFGR', 'PFOE', 'PFNR']
    part_house_lock_status = ['OPHL', 'PHVF', 'PHMD', 'PHGR', 'PHOE', 'PHNR']
    open_status = ['OP', 'OVF', 'OMD', 'OGR', 'OOE', 'ONR']
    pipeline_status = ['TB', 'DP']
    all_not_initiated_supplier = []
    all_recce_supplier = []

    proposal = ProposalInfo.objects.get(proposal_id=campaign_id)
    end_customer = proposal.type_of_end_customer.name

    booking_startdate = timezone.datetime.today().date()
    booking_enddate = booking_startdate + datetime.timedelta(days=3)
    pipeline_enddate = booking_startdate + datetime.timedelta(days=7)

    for phase_id in shortlisted_spaces_by_phase_dict:
        end_date = all_phase_by_id[phase_id]['end_date'] if phase_id in all_phase_by_id else None
        start_date = all_phase_by_id[phase_id]['start_date'] if phase_id in all_phase_by_id else None
        total_pipeline_suppliers_count = 0
        total_pipeline_flats = 0
        total_booked_suppliers_count = 0
        total_booked_flats = 0
        verbally_booked_suppliers_count = 0
        verbally_booked_flats = 0
        followup_req_booked_suppliers_count = 0
        followup_req_booked_flats = 0
        confirmed_booked_suppliers_count = 0
        confirmed_booked_flats = 0
        not_initiated_supplier_count = 0
        not_initiated_flats = 0
        rejected_supplier_count = 0
        btob_rejected_supplier_count = 0
        btob_rejected_flats = 0
        decision_pending_supplier_count = 0
        decision_pending_flats = 0
        recce_required_supplier_count = 0
        not_initiated_supplier_count = 0
        meeting_fixed_supplier_count =0
        meeting_fixed_flats =0
        meeting_converted_supplier_count = 0
        meeting_converted_flats = 0
        emergency_situation_supplier_count = 0
        emergency_situation_flats = 0
        complete_lockdown_supplier_count = 0
        complete_lockdown_flats = 0
        part_building_lock_supplier_count = 0
        part_building_lock_flats = 0
        part_floor_lock_supplier_count = 0
        part_floor_lock_flats = 0
        part_house_lock_supplier_count = 0
        part_house_lock_flats = 0
        open_supplier_count = 0
        open_flats = 0
        recce_flats = 0
        rejected_flats = 0
        pipeline_supplier_count = 0
        pipeline_flats = 0
        for status in shortlisted_spaces_by_phase_dict[phase_id]:
            phase_booked_suppliers = len(shortlisted_spaces_by_phase_dict[phase_id][status])
            phase_booked_flats = sum(supplier['flat_count'] for supplier in shortlisted_spaces_by_phase_dict[phase_id][status] if supplier['flat_count'])
            
            if end_customer == 'B to C':
                if status in verbally_booked_status:
                    total_booked_suppliers_count += phase_booked_suppliers
                    total_booked_flats += phase_booked_flats
                    verbally_booked_suppliers_count += phase_booked_suppliers
                    verbally_booked_flats += phase_booked_flats
                if status in followup_req_status:
                    total_booked_suppliers_count += phase_booked_suppliers
                    total_booked_flats += phase_booked_flats
                    followup_req_booked_suppliers_count += phase_booked_suppliers
                    followup_req_booked_flats += phase_booked_flats
                if status in confirmed_booked_status:
                    total_booked_suppliers_count += phase_booked_suppliers
                    total_booked_flats += phase_booked_flats
                    # confirmed_booked_suppliers_count += phase_booked_suppliers
                    # confirmed_booked_flats += phase_booked_flats
                if status in not_initiated_status:
                    total_booked_suppliers_count += phase_booked_suppliers
                    not_initiated_flats += phase_booked_flats
                    total_booked_flats += phase_booked_flats
                    not_initiated_supplier_count += phase_booked_suppliers
                if status in btoc_rejected_status:
                    total_booked_suppliers_count += phase_booked_suppliers
                    rejected_flats += phase_booked_flats
                    total_booked_flats += phase_booked_flats
                    rejected_supplier_count += phase_booked_suppliers
                if status in recce_required:
                    total_booked_suppliers_count += phase_booked_suppliers
                    recce_flats += phase_booked_flats
                    total_booked_flats += phase_booked_flats
                    recce_required_supplier_count += phase_booked_suppliers

            elif end_customer == 'B to B':
                if status in meeting_fixed:
                    total_booked_suppliers_count += phase_booked_suppliers
                    total_booked_flats += phase_booked_flats
                    meeting_fixed_flats += phase_booked_flats
                    meeting_fixed_supplier_count += phase_booked_suppliers
                if status in meeting_converted:
                    total_booked_suppliers_count += phase_booked_suppliers
                    total_booked_flats += phase_booked_flats
                    meeting_converted_flats += phase_booked_flats
                    meeting_converted_supplier_count += phase_booked_suppliers
                if status in btob_rejected_status:
                    total_booked_suppliers_count += phase_booked_suppliers
                    total_booked_flats += phase_booked_flats
                    btob_rejected_flats += phase_booked_flats
                    btob_rejected_supplier_count += phase_booked_suppliers
                if status in decision_pending_status:
                    total_booked_suppliers_count += phase_booked_suppliers
                    total_booked_flats += phase_booked_flats
                    decision_pending_flats += phase_booked_flats
                    decision_pending_supplier_count += phase_booked_suppliers
                if status in not_initiated_status:
                    total_booked_suppliers_count += phase_booked_suppliers
                    not_initiated_flats += phase_booked_flats
                    total_booked_flats += phase_booked_flats
                    not_initiated_supplier_count += phase_booked_suppliers

            else: 
                if status in emergency_situation_status:
                    total_booked_suppliers_count += phase_booked_suppliers
                    total_booked_flats += phase_booked_flats
                    emergency_situation_flats += phase_booked_flats
                    emergency_situation_supplier_count += phase_booked_suppliers
                if status in complete_lockdown_status:
                    total_booked_suppliers_count += phase_booked_suppliers
                    total_booked_flats += phase_booked_flats
                    complete_lockdown_flats += phase_booked_flats
                    complete_lockdown_supplier_count += phase_booked_suppliers
                if status in part_building_lock_status:
                    total_booked_suppliers_count += phase_booked_suppliers
                    total_booked_flats += phase_booked_flats
                    part_building_lock_flats += phase_booked_flats
                    part_building_lock_supplier_count += phase_booked_suppliers
                if status in part_floor_lock_status:
                    total_booked_suppliers_count += phase_booked_suppliers
                    total_booked_flats += phase_booked_flats
                    part_floor_lock_flats += phase_booked_flats
                    part_floor_lock_supplier_count += phase_booked_suppliers
                if status in part_house_lock_status:
                    total_booked_suppliers_count += phase_booked_suppliers
                    total_booked_flats += phase_booked_flats
                    part_house_lock_flats += phase_booked_flats
                    part_house_lock_supplier_count += phase_booked_suppliers
                if status in open_status:
                    total_booked_suppliers_count += phase_booked_suppliers
                    total_booked_flats += phase_booked_flats
                    open_flats += phase_booked_flats
                    open_supplier_count += phase_booked_suppliers

        total_booked_suppliers_list = flatten_list(
            [supplier for status, supplier in shortlisted_spaces_by_phase_dict[phase_id].items()])

        confirmed_booked_suppliers_list = []
        pipeline_suppliers_list = []
        total_pipeline_suppliers_list = []
        for status, supplier in shortlisted_spaces_by_phase_dict[phase_id].items():
            if status in confirmed_booked_status:
                for row in supplier:
                    if row["next_action_date"]:
                        next_action_date = datetime.datetime.strptime(row["next_action_date"], '%Y/%m/%d').date()
                        if row["next_action_date"] and next_action_date >= booking_startdate and next_action_date <= booking_enddate:
                            confirmed_booked_suppliers_list.append(row)
                            confirmed_booked_suppliers_count += 1
                            confirmed_booked_flats += row["flat_count"]

            elif status in pipeline_status:
                for row in supplier:
                    if row["next_action_date"]:
                        next_action_date = datetime.datetime.strptime(row["next_action_date"], '%Y/%m/%d').date()
                        if next_action_date and next_action_date >= booking_startdate and next_action_date <= pipeline_enddate:
                            pipeline_suppliers_list.append(row)
                            pipeline_supplier_count += 1
                            pipeline_flats += row["flat_count"]

            if status in verbally_booked_status or status in confirmed_booked_status or status in followup_req_status or status in meeting_fixed or status in meeting_converted or status in decision_pending_status or status in complete_lockdown_status or status in emergency_situation_status or status in part_building_lock_status or status in part_floor_lock_status or status in part_house_lock_status or status in open_status:
                for row in supplier:
                    total_pipeline_suppliers_list.append(row)
                    total_pipeline_suppliers_count += 1
                    total_pipeline_flats += row["flat_count"]              

        verbally_booked_suppliers_list = flatten_list(
            [supplier for status, supplier in shortlisted_spaces_by_phase_dict[phase_id].items() if
             status in verbally_booked_status])
        followup_required_suppliers_list = flatten_list(
            [supplier for status, supplier in shortlisted_spaces_by_phase_dict[phase_id].items() if
             status in followup_req_status])
        btoc_rejected_suppliers_list = flatten_list(
            [supplier for status, supplier in shortlisted_spaces_by_phase_dict[phase_id].items() if
             status in btoc_rejected_status])
        not_initiated_suppliers_list = flatten_list(
            [supplier for status, supplier in shortlisted_spaces_by_phase_dict[phase_id].items() if
             status in not_initiated_status])
        recce_required_suppliers_list = flatten_list(
            [supplier for status, supplier in shortlisted_spaces_by_phase_dict[phase_id].items() if
             status in recce_required])
        meeting_fixed_suppliers_list = flatten_list(
            [supplier for status, supplier in shortlisted_spaces_by_phase_dict[phase_id].items() if
             status in meeting_fixed])
        meeting_converted_suppliers_list = flatten_list(
            [supplier for status, supplier in shortlisted_spaces_by_phase_dict[phase_id].items() if
             status in meeting_converted])
        btob_rejected_suppliers_list = flatten_list(
            [supplier for status, supplier in shortlisted_spaces_by_phase_dict[phase_id].items() if
             status in btob_rejected_status])
        decision_pending_suppliers_list = flatten_list(
            [supplier for status, supplier in shortlisted_spaces_by_phase_dict[phase_id].items() if
             status in decision_pending_status])
        emergency_situation_suppliers_list = flatten_list(
            [supplier for status, supplier in shortlisted_spaces_by_phase_dict[phase_id].items() if
             status in emergency_situation_status])
        complete_lockdown_suppliers_list = flatten_list(
            [supplier for status, supplier in shortlisted_spaces_by_phase_dict[phase_id].items() if
             status in complete_lockdown_status])
        part_building_lock_suppliers_list = flatten_list(
            [supplier for status, supplier in shortlisted_spaces_by_phase_dict[phase_id].items() if
             status in part_building_lock_status])
        part_floor_lock_suppliers_list = flatten_list(
            [supplier for status, supplier in shortlisted_spaces_by_phase_dict[phase_id].items() if
             status in part_floor_lock_status])
        part_house_lock_suppliers_list = flatten_list(
            [supplier for status, supplier in shortlisted_spaces_by_phase_dict[phase_id].items() if
             status in part_house_lock_status])
        open_suppliers_list = flatten_list(
            [supplier for status, supplier in shortlisted_spaces_by_phase_dict[phase_id].items() if
             status in open_status])

        phase_dict = {
            'phase_no': all_phase_by_id[phase_id]['phase_no'],
            'start_date': datetime.datetime.strptime(str(all_phase_by_id[phase_id]['start_date'])[:10], '%Y-%m-%d').strftime('%d %b %Y'),
            'end_date': datetime.datetime.strptime(str(all_phase_by_id[phase_id]['end_date'])[:10], '%Y-%m-%d').strftime('%d %b %Y'),
            'comments': all_phase_by_id[phase_id]['comments'],
            'supplier_data': shortlisted_spaces_by_phase_dict[phase_id],
            'overall_inventory_counts': overall_inventory_count_dict,
            'total_booked': {
                'supplier_count': total_booked_suppliers_count,
                'flat_count': total_booked_flats,
                'supplier_data': total_booked_suppliers_list
            },
            'total_pipeline': {
                'supplier_count': total_pipeline_suppliers_count,
                'flat_count': total_pipeline_flats,
                'supplier_data': total_pipeline_suppliers_list
            },
            'confirmed_booked': {
                'supplier_count': confirmed_booked_suppliers_count,
                'flat_count': confirmed_booked_flats,
                'supplier_data': confirmed_booked_suppliers_list
            },
            'verbally_booked': {
                'supplier_count': verbally_booked_suppliers_count,
                'flat_count': verbally_booked_flats,
                'supplier_data': verbally_booked_suppliers_list
            },
            'followup_required': {
                'supplier_count': followup_req_booked_suppliers_count,
                'flat_count': followup_req_booked_flats,
                'supplier_data': followup_required_suppliers_list
            },
            'rejected': {
                'supplier_count': rejected_supplier_count,
                'flat_count': rejected_flats,
                'supplier_data': btoc_rejected_suppliers_list
            },
            'not_initiated': {
                'supplier_count': not_initiated_supplier_count,
                'flat_count': not_initiated_flats,
                'supplier_data': not_initiated_suppliers_list
            },
            'recce_required': {
                'supplier_count': recce_required_supplier_count,
                'flat_count': recce_flats,
                'supplier_data': recce_required_suppliers_list
            },
            'meeting_fixed': {
                'supplier_count': meeting_fixed_supplier_count,
                'flat_count': meeting_fixed_flats,
                'supplier_data': meeting_fixed_suppliers_list
            },
            'meeting_converted': {
                'supplier_count': meeting_converted_supplier_count,
                'flat_count': meeting_converted_flats,
                'supplier_data': meeting_converted_suppliers_list
            },
            'btob_rejected': {
                'supplier_count': btob_rejected_supplier_count,
                'flat_count': btob_rejected_flats,
                'supplier_data': btob_rejected_suppliers_list
            },
            'decision_pending': {
                'supplier_count': decision_pending_supplier_count,
                'flat_count': decision_pending_flats,
                'supplier_data': decision_pending_suppliers_list
            },
            'emergency_situation': {
                'supplier_count': emergency_situation_supplier_count,
                'flat_count': emergency_situation_flats,
                'supplier_data': emergency_situation_suppliers_list
            },
            'complete_lockdown': {
                'supplier_count': complete_lockdown_supplier_count,
                'flat_count': complete_lockdown_flats,
                'supplier_data': complete_lockdown_suppliers_list
            },
            'partial_building_lockdown': {
                'supplier_count': part_building_lock_supplier_count,
                'flat_count': part_building_lock_flats,
                'supplier_data': part_building_lock_suppliers_list
            },
            'partial_floor_lockdown': {
                'supplier_count': part_floor_lock_supplier_count,
                'flat_count': part_floor_lock_flats,
                'supplier_data': part_floor_lock_suppliers_list
            },
            'partial_house_lockdown': {
                'supplier_count': part_house_lock_supplier_count,
                'flat_count': part_house_lock_flats,
                'supplier_data': part_house_lock_suppliers_list
            },
            'open': {
                'supplier_count': open_supplier_count,
                'flat_count': open_flats,
                'supplier_data': open_suppliers_list
            },
            'pipeline': {
                'supplier_count': pipeline_supplier_count,
                'flat_count': pipeline_flats,
                'supplier_data': pipeline_suppliers_list
            }
        }
        if end_date is not None and end_date.date() >= current_date:
            shortlisted_spaces_by_phase_list.append(phase_dict)
            if start_date and start_date.date() <= current_date:
                status = 'ongoing'
                phase_dict['status'] = status
                ongoing_phase = phase_dict
            elif start_date and start_date.date() > current_date:
                status = 'upcoming'
                phase_dict['status'] = status
                upcoming_phases.append(phase_dict)
        elif end_date and end_date.date() < current_date:
            status = 'completed'
            phase_dict['status'] = status
            completed_phases.append(phase_dict)
    last_completed_phase = None
    sorted_upcoming_phases = sorted(upcoming_phases, key=lambda k: k['phase_no'])
    upcoming_beyond_three = sorted_upcoming_phases[3:]
    pipeline = {'followup_required': {'supplier_count': 0, 'flat_count': 0, 'supplier_data':[]},
                'confirmed_booked': {'supplier_count': 0, 'flat_count': 0, 'supplier_data': []},
                'verbally_booked': {'supplier_count': 0, 'flat_count': 0, 'supplier_data': []},
                'rejected': {'supplier_count': 0, 'flat_count': 0, 'supplier_data': []},
                'btob_rejected': {'supplier_count': 0, 'flat_count': 0, 'supplier_data': []},
                'not_initiated': {'supplier_count': 0, 'flat_count': 0, 'supplier_data': []},
                'total_booked': {'supplier_count': 0, 'flat_count': 0, 'supplier_data': []},
                'recce_required': {'supplier_count': 0, 'flat_count': 0, 'supplier_data': []},
                'meeting_fixed': {'supplier_count': 0, 'flat_count': 0, 'supplier_data': []},
                'meeting_converted': {'supplier_count': 0, 'flat_count': 0, 'supplier_data': []},
                'decision_pending': {'supplier_count': 0, 'flat_count': 0, 'supplier_data': []},
                'emergency_situation': {'supplier_count': 0, 'flat_count': 0, 'supplier_data': []},
                'complete_lockdown': {'supplier_count': 0, 'flat_count': 0, 'supplier_data': []},
                'partial_building_lockdown': {'supplier_count': 0, 'flat_count': 0, 'supplier_data': []},
                'partial_floor_lockdown': {'supplier_count': 0, 'flat_count': 0, 'supplier_data': []},
                'partial_house_lockdown': {'supplier_count': 0, 'flat_count': 0, 'supplier_data': []},
                'open': {'supplier_count': 0, 'flat_count': 0, 'supplier_data': []},
                'pipeline': {'supplier_count': 0, 'flat_count': 0, 'supplier_data': []},
                }

    for phase in upcoming_beyond_three:
        for status_type in ['followup_required', 'confirmed_booked', 'verbally_booked', 'rejected', 'btob_rejected', 'total_booked', 'not_initiated', 'recce_required', 'meeting_fixed', 'meeting_converted', 'decision_pending', 
                        'emergency_situation', 'complete_lockdown', 'partial_building_lockdown', 'partial_floor_lockdown', 'partial_house_lockdown', 'open']:
            pipeline[status_type]['supplier_count'] += phase[status_type]['supplier_count']
            pipeline[status_type]['flat_count'] += phase[status_type]['flat_count']
            pipeline[status_type]['supplier_data'] += phase[status_type]['supplier_data']

    for supplier in (no_status_suppliers + no_phase_suppliers):
        if not supplier['booking_status']:
            pipeline['not_initiated']['supplier_data'].append(supplier)
            pipeline['not_initiated']['flat_count'] += supplier['flat_count'] if supplier['flat_count'] else 0
            pipeline['not_initiated']['supplier_count'] += 1
        if supplier['booking_status'] in followup_req_status:
            pipeline['followup_required']['supplier_data'].append(supplier)
            pipeline['followup_required']['flat_count'] += supplier['flat_count'] if supplier['flat_count'] else 0
            pipeline['followup_required']['supplier_count'] += 1
            # pipeline['total_booked']['supplier_data'].append(supplier)
            # pipeline['total_booked']['flat_count'] += supplier['flat_count'] if supplier['flat_count'] else 0
            # pipeline['total_booked']['supplier_count'] += 1
        if supplier['booking_status'] in verbally_booked_status:
            pipeline['verbally_booked']['supplier_data'].append(supplier)
            pipeline['verbally_booked']['flat_count'] += supplier['flat_count'] if supplier['flat_count'] else 0
            pipeline['verbally_booked']['supplier_count'] += 1
        if supplier['booking_status'] in btoc_rejected_status:
            pipeline['rejected']['supplier_data'].append(supplier)
            pipeline['rejected']['flat_count'] += supplier['flat_count'] if supplier['flat_count'] else 0
            pipeline['rejected']['supplier_count'] += 1
        if supplier['booking_status'] in btob_rejected_status:
            pipeline['btob_rejected']['supplier_data'].append(supplier)
            pipeline['btob_rejected']['flat_count'] += supplier['flat_count'] if supplier['flat_count'] else 0
            pipeline['btob_rejected']['supplier_count'] += 1
        if supplier['booking_status'] in recce_required:
            pipeline['recce_required']['supplier_data'].append(supplier)
            pipeline['recce_required']['flat_count'] += supplier['flat_count'] if supplier['flat_count'] else 0
            pipeline['recce_required']['supplier_count'] += 1
            # pipeline['total_booked']['supplier_data'].append(supplier)
            # pipeline['total_booked']['flat_count'] += supplier['flat_count'] if supplier['flat_count'] else 0
            # pipeline['total_booked']['supplier_count'] += 1
        if supplier['booking_status'] in meeting_fixed:
            pipeline['meeting_fixed']['supplier_data'].append(supplier)
            pipeline['meeting_fixed']['flat_count'] += supplier['flat_count'] if supplier['flat_count'] else 0
            pipeline['meeting_fixed']['supplier_count'] += 1
        if supplier['booking_status'] in meeting_converted:
            pipeline['meeting_converted']['supplier_data'].append(supplier)
            pipeline['meeting_converted']['flat_count'] += supplier['flat_count'] if supplier['flat_count'] else 0
            pipeline['meeting_converted']['supplier_count'] += 1
        if supplier['booking_status'] in decision_pending_status:
            pipeline['decision_pending']['supplier_data'].append(supplier)
            pipeline['decision_pending']['flat_count'] += supplier['flat_count'] if supplier['flat_count'] else 0
            pipeline['decision_pending']['supplier_count'] += 1
        if supplier['booking_status'] in emergency_situation_status:
            pipeline['emergency_situation']['supplier_data'].append(supplier)
            pipeline['emergency_situation']['flat_count'] += supplier['flat_count'] if supplier['flat_count'] else 0
            pipeline['emergency_situation']['supplier_count'] += 1
        if supplier['booking_status'] in complete_lockdown_status:
            pipeline['complete_lockdown']['supplier_data'].append(supplier)
            pipeline['complete_lockdown']['flat_count'] += supplier['flat_count'] if supplier['flat_count'] else 0
            pipeline['complete_lockdown']['supplier_count'] += 1
        if supplier['booking_status'] in part_building_lock_status:
            pipeline['partial_building_lockdown']['supplier_data'].append(supplier)
            pipeline['partial_building_lockdown']['flat_count'] += supplier['flat_count'] if supplier['flat_count'] else 0
            pipeline['partial_building_lockdown']['supplier_count'] += 1
        if supplier['booking_status'] in part_floor_lock_status:
            pipeline['partial_floor_lockdown']['supplier_data'].append(supplier)
            pipeline['partial_floor_lockdown']['flat_count'] += supplier['flat_count'] if supplier['flat_count'] else 0
            pipeline['partial_floor_lockdown']['supplier_count'] += 1
        if supplier['booking_status'] in part_house_lock_status:
            pipeline['partial_house_lockdown']['supplier_data'].append(supplier)
            pipeline['partial_house_lockdown']['flat_count'] += supplier['flat_count'] if supplier['flat_count'] else 0
            pipeline['partial_house_lockdown']['supplier_count'] += 1
        if supplier['booking_status'] in open_status:
            pipeline['open']['supplier_data'].append(supplier)
            pipeline['open']['flat_count'] += supplier['flat_count'] if supplier['flat_count'] else 0
            pipeline['open']['supplier_count'] += 1
        if supplier['booking_status'] in pipeline_status and supplier["next_action_date"]:
            next_action_date = datetime.datetime.strptime(supplier["next_action_date"], '%Y/%m/%d').date()
            if next_action_date >= booking_startdate and next_action_date <= pipeline_enddate:
                pipeline['pipeline']['supplier_data'].append(supplier)
                pipeline['pipeline']['flat_count'] += supplier['flat_count'] if supplier['flat_count'] else 0
                pipeline['pipeline']['supplier_count'] += 1

        pipeline['total_booked']['supplier_data'].append(supplier)
        pipeline['total_booked']['flat_count'] += supplier['flat_count'] if supplier['flat_count'] else 0
        pipeline['total_booked']['supplier_count'] += 1
    
    breakup_header = v0_constants.breakup_header
    summary_header = v0_constants.summary_header

    if len(completed_phases) > 0:
        last_completed_phase = sorted(completed_phases, key=lambda k: int(k['phase_no']))[-1]
    if last_completed_phase:
        if ongoing_phase:
            upcoming_phase = sorted_upcoming_phases[:1]
        else:
            upcoming_phase = sorted_upcoming_phases[:2]
    else:
        if ongoing_phase:
            upcoming_phase = sorted_upcoming_phases[:2]
        else:
            upcoming_phase = sorted_upcoming_phases[:3]
    shortlisted_spaces_by_phase_dict = {
        'all_phases': shortlisted_spaces_by_phase_list,
        'last_completed_phase': last_completed_phase,
        'upcoming_phases': upcoming_phase,
        'ongoing_phase': ongoing_phase,
        'pipeline_suppliers': pipeline,
        'completed_phases': completed_phases,
        'breakup_header' : breakup_header,
        'summary_header' : summary_header,
        'type_of_end_customer' : end_customer,
    }
    return shortlisted_spaces_by_phase_dict


class getSupplierListByStatus(APIView):
    @staticmethod
    def get(request, campaign_id):
        try:
            center_data = ProposalCenterSuppliers.objects.filter(proposal_id=campaign_id).values('supplier_type_code').annotate(supplier_type=Count('supplier_type_code'))
            shortlisted_spaces_by_phase_list = get_supplier_list_by_status_ctrl(campaign_id)
            return Response({'status': True, 'data': shortlisted_spaces_by_phase_list, 'supplier_type_code':center_data})
        except Exception as e:
            logger.exception(e)
            return ui_utils.handle_response("", exception_object=e, request=request)

class ImportSheetInExistingCampaign(APIView):
    """
    This will convert proposal to campaign where supplier id's should be provided in sheet
    """

    def post(self, request):
        """

        :param request:
        :return:
        """
        class_name = self.__class__.__name__
        try:
            data = request.data.copy()

            genrate_supplier_data2(data,request.user)

            return ui_utils.handle_response(class_name, data={}, success=True)
        except Exception as e:
            logger.exception(e)
            return ui_utils.handle_response(class_name, exception_object=e, request=request)

class GetOngoingSuppliersOfCampaign(APIView):

    def get(self, request,campaign_id):
        class_name = self.__class__.__name__
        try:
            date = request.query_params.get("date",None)
            if date :
                phases = SupplierPhase.objects.filter(campaign=campaign_id,start_date__lte=date,end_date__gte=date).values_list('id')
                supplier_ids = ShortlistedSpaces.objects.filter(phase_no__in=phases).values_list('object_id')
                suppliers = SupplierTypeSociety.objects.filter(supplier_id__in=supplier_ids)
                serializer = SupplierTypeSocietySerializer(suppliers, many=True)
                form_details = mongo_client.leads_forms.find({"campaign_id":campaign_id})
                data = {}
                data["suppliers" ] = serializer.data
                if form_details.count() > 0:
                    form_id = form_details[0]['leads_form_id']
                    data['form_id'] = form_id
                    extra_leads = mongo_client.leads_extras.find({"campaign_id": campaign_id, "leads_form_id":form_id})
                    extra_leads_supplier_map = {lead["supplier_id"].replace('"','') : lead for lead in extra_leads}
                    for supplier in data["suppliers"]:
                        if supplier["supplier_id"] in extra_leads_supplier_map:
                            if 'leads' not in extra_leads_supplier_map:
                                supplier["leads"] = []

                            temp_data = {
                                'hot_leads' : extra_leads_supplier_map[supplier["supplier_id"]]["extra_hot_leads"],
                                'total_leads' : extra_leads_supplier_map[supplier["supplier_id"]]["extra_leads"],
                                'timestamp' : extra_leads_supplier_map[supplier["supplier_id"]]["created_at"]
                                            if 'created_at' in extra_leads_supplier_map[supplier["supplier_id"]] else None,
                            }
                            supplier["leads"].append(temp_data)
            return ui_utils.handle_response(class_name, data=data, success=True)
        except Exception as e:
            return ui_utils.handle_response(class_name, exception_object=e, request=request)

class GetExtraLead(APIView):
    def get(self, request, campaign_id, form_id, supplier_id):
        class_name = self.__class__.__name__
        try:
            extra_leads = mongo_client.leads_extras.find({"campaign_id": campaign_id, "leads_form_id": int(form_id), "supplier_id": supplier_id})
            data = []
            for lead in extra_leads:
                temp_data = {
                    "id": str(lead['_id']),
                    "supplier_id": lead['supplier_id'],
                    "campaign_id": lead['campaign_id'],
                    "leads_form_id":lead['leads_form_id'],
                    "extra_hot_leads": lead['extra_hot_leads'],
                    "extra_leads": lead['extra_leads'],
                    "created_at": lead['created_at']
                }
                data.append(temp_data)
            return ui_utils.handle_response(class_name, data=data, success=True)
        except Exception as e:
            return ui_utils.handle_response(class_name, exception_object=e, request=request)

class BookingAssignmentViewSet(APIView):

    def post(self, request):
        class_name = self.__class__.__name__
        try:
            print('handle_response')
            data = {}
            campaign_id = request.data["campaign_id"]
            user_ids = request.data["assigned_to_ids"]
            supplier_ids = ShortlistedSpaces.objects.filter(proposal_id=campaign_id).values('object_id')

            for supplier in supplier_ids:
                user_ids_old = SupplierAssignment.objects.filter(campaign=campaign_id, supplier_id=supplier['object_id'],
                                                             assigned_to__in=user_ids).values()
                user_ids_old = [user_id['assigned_to_id'] for user_id in user_ids_old]
                assigned_by_user = BaseUser.objects.get(id=request.data["assigned_by"])
                user_ids_new = list(set(user_ids) - set(user_ids_old))
                user_ids_new_objects = BaseUser.objects.filter(id__in=user_ids_new)
                data = []
                now_time = timezone.now()
                for user_id in user_ids_new_objects:
                    data.append(SupplierAssignment(**{
                        "campaign_id": campaign_id,
                        "supplier_id": supplier['object_id'],
                        "assigned_by": assigned_by_user,
                        "assigned_to": user_id,
                        "created_at": now_time,
                        "updated_at": now_time
                    }))

            SupplierAssignment.objects.bulk_create(data)
                
            return ui_utils.handle_response(class_name, data={}, success=True)
        except Exception as e:
            return ui_utils.handle_response(class_name, exception_object=e, request=request)


class SupplierAssignmentViewSet(viewsets.ViewSet):


    def create(self, request):
        class_name = self.__class__.__name__
        try:
            data = {}
            campaign_id = request.data["campaign_id"]
            supplier_id = request.data["supplier_id"]
            user_ids = request.data["assigned_to_ids"]
            user_ids_old = SupplierAssignment.objects.filter(campaign=campaign_id, supplier_id=supplier_id,
                                                             assigned_to__in=user_ids).values()
            user_ids_old = [user_id['assigned_to_id'] for user_id in user_ids_old]
            assigned_by_user = BaseUser.objects.get(id=request.data["assigned_by"])
            user_ids_new = list(set(user_ids) - set(user_ids_old))
            user_ids_new_objects = BaseUser.objects.filter(id__in=user_ids_new)
            data = []
            now_time = timezone.now()
            for user_id in user_ids_new_objects:
                data.append(SupplierAssignment(**{
                    "campaign_id": campaign_id,
                    "supplier_id": supplier_id,
                    "assigned_by": assigned_by_user,
                    "assigned_to": user_id,
                    "created_at": now_time,
                    "updated_at": now_time
                }))

            SupplierAssignment.objects.bulk_create(data)
                
            return ui_utils.handle_response(class_name, data={}, success=True)
        except Exception as e:
            return ui_utils.handle_response(class_name, exception_object=e, request=request)

    def list(self, request):
        class_name = self.__class__.__name__
        user = request.user.id
        campaign_id = request.query_params.get('campaign_id')

        if not campaign_id:
            return ui_utils.handle_response(class_name, data='Please pass campaign Id', success=False)
        suppliers = SupplierAssignment.objects.filter(campaign=campaign_id, assigned_by=user)
        result_obj = {}
        for supplier_obj in suppliers:
            if supplier_obj.supplier_id not in result_obj:
                result_obj[supplier_obj.supplier_id] = {}
            result_obj[supplier_obj.supplier_id]["assigned_to"] = supplier_obj.assigned_to.id
            result_obj[supplier_obj.supplier_id]["updated_at"] = supplier_obj.updated_at
            result_obj[supplier_obj.supplier_id]["supplier_id"] = supplier_obj.supplier_id
        result_list = [result_obj[supplier] for supplier in result_obj]

        return ui_utils.handle_response(class_name, data=result_list, success=True)


class BrandAssignmentViewSet(viewsets.ViewSet):

    def create(self, request):
        class_name = self.__class__.__name__
        try:

            if request.data["id"]:
                item = ShortlistedSpaces.objects.filter(id=request.data["id"]).first()
                if item:
                    item.brand_organisation_id = request.data["brand_organisation_id"]
                    item.save()
               
            return ui_utils.handle_response(class_name, data={}, success=True)
        except Exception as e:
            return ui_utils.handle_response(class_name, exception_object=e, request=request)


class ConvertProposalToCampaign(APIView):
    def post(self, request, proposal_id):
        try:
            class_name = self.__class__.__name__
            proposal_details = ProposalInfo.objects.filter(proposal_id=proposal_id)
            status = 'PTC'
            start_date = proposal_details[0].created_at
            end_date = start_date + datetime.timedelta(weeks=16)
            invoice_number = proposal_id
            proposal_details.update(tentative_start_date=start_date, tentative_end_date=end_date, campaign_state=status, invoice_number=invoice_number)

            generic_file = GenericExportFileName.objects.filter(proposal_id=proposal_id)
            generic_file.update(is_exported=0)

            return ui_utils.handle_response(class_name, data='Proposal conversion successfull', success=True)
        except Exception as e:
            return ui_utils.handle_response(class_name, data='Error converting proposal', success=False)

class BookingStatusAPI(APIView):
    def get(self, request, proposal_id):
        try:
            class_name = self.__class__.__name__
            end_customer = ProposalInfo.objects.filter(proposal_id=proposal_id).values('type_of_end_customer')
            bk_status = BookingStatus.objects.prefetch_related('booking_substatus').filter(type_of_end_customer__in = end_customer)
            serializer = BookingStatusSerializer(bk_status, many=True)

            return ui_utils.handle_response(class_name, data=serializer.data, success=True)

        except Exception as e:
            return ui_utils.handle_response(class_name, exception_object=e, success=False)

class EndCustomerType(APIView):
    def get(self, request):
        try:
            class_name = self.__class__.__name__
            end_customer = TypeOfEndCustomer.objects.all().values('id', 'name')
            return ui_utils.handle_response(class_name, data=end_customer, success=True)

        except Exception as e:
            return ui_utils.handle_response(class_name, exception_object=e, success=False)