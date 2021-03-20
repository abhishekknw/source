from __future__ import print_function
from __future__ import absolute_import
from rest_framework.response import Response
from rest_framework.views import APIView
from django.http import HttpResponse
from .models import (Gupshup,PaymentDetails,LicenseDetails,MachadaloRelationshipManager,Requirement, SuspenseLead, BrowsedLead, CampaignLeads, OrganizationLeads, PreRequirement)
from .serializers import NotificationTemplateSerializer,PaymentDetailsSerializer,LicenseDetailsSerializer,RequirementSerializer, PreRequirementSerializer, RelationshipManagerSerializer
import v0.ui.utils as ui_utils
from openpyxl import load_workbook, Workbook
from v0.ui.account.models import ContactDetails, BusinessTypes, BusinessSubTypes
from v0.ui.account.serializers import ContactDetailsSerializer
from v0.ui.supplier.models import SupplierTypeSociety, SupplierMaster, AddressMaster
from v0.ui.location.models import City,CityArea
from v0.ui.proposal.models import ProposalInfo, ShortlistedSpaces, ProposalCenterMapping
from v0.ui.proposal.serializers import ProposalInfoSerializer
from v0.ui.organisation.models import Organisation
from v0.ui.organisation.serializers import OrganisationSerializer
from django.db.models import Q
import datetime
from v0.ui.common.models import mongo_client,BaseUser
import hashlib
from v0.ui.common.pagination import paginateMongo
from bson.objectid import ObjectId
from v0.ui.common.serializers import BaseUserSerializer
from v0.ui.supplier.serializers import SupplierMasterSerializer, SupplierTypeSocietySerializer
import v0.constants as v0_constants
from v0.constants import (campaign_status, proposal_on_hold)
from v0.ui.website.utils import manipulate_object_key_values, manipulate_master_to_rs
import v0.ui.b2b.utils as b2b_utils
from django.db.models import F
from v0.ui.campaign.models import CampaignComments
from datetime import timedelta
from django.utils.timezone import make_aware
from v0.constants import supplier_code_to_names
from v0.ui.location.models import CityArea
from v0.ui.supplier.views import update_contact_and_ownership_detail

def get_value_from_list_by_key(list1, key):
    text = ""
    try:
        text = list1[key].value
        text = text.strip()
    except:
        pass
    return text

class ImportLead(APIView):

    def post(self, request, campaign_id):
        data = request.data.copy()
        source_file = data['file']
        wb = load_workbook(source_file)
        ws = wb.get_sheet_by_name(wb.get_sheet_names()[0])
        headers = {}
        skip_row_list = []
        lead_status = "Lead"
        
        for index, row in enumerate(ws.iter_rows()):
            if index == 0:
                i = 0
                for key in row:
                    key1 = key.value.lower()
                    key1 = key1.strip()
                    headers[key1] = i
                    i+=1
            else:
                # Parsing excel columns
                phone_number = get_value_from_list_by_key(row, headers.get('phone number'))
                supplier_name = get_value_from_list_by_key(row, headers.get('supplier name'))
                city = get_value_from_list_by_key(row, headers.get('city'))
                area = get_value_from_list_by_key(row, headers.get('area'))
                sector_name = get_value_from_list_by_key(row, headers.get('sector'))
                sub_sector_name = get_value_from_list_by_key(row, headers.get('sub sector'))
                impl_timeline = get_value_from_list_by_key(row, headers.get('implementation timeline'))
                meating_timeline = get_value_from_list_by_key(row, headers.get('meating timeline'))
                comment = get_value_from_list_by_key(row, headers.get('comment'))
                current_patner = get_value_from_list_by_key(row, headers.get('current partner'))
                current_patner_feedback = get_value_from_list_by_key(row, headers.get('current patner feedback'))
                current_patner_feedback_reason = get_value_from_list_by_key(row, headers.get('current patner feedback reason'))
                prefered_patners = get_value_from_list_by_key(row, headers.get('prefered partners'))
                l1_answers = get_value_from_list_by_key(row, headers.get('l1 answers'))
                l1_answer_2 = get_value_from_list_by_key(row, headers.get('l1 answer 2'))
                l2_answers = get_value_from_list_by_key(row, headers.get('l2 answers'))
                l2_answer_2 = get_value_from_list_by_key(row, headers.get('l2 answer 2'))
                call_back_preference = get_value_from_list_by_key(row, headers.get('call back preference'))

                if phone_number is None:
                    skip_row_list.append(index)

                if impl_timeline:
                    impl_timeline = impl_timeline.lower()
                else:
                    impl_timeline = "not given"

                if meating_timeline:
                    meating_timeline = meating_timeline.lower()
                else:
                    meating_timeline = "not given"

                if current_patner_feedback:
                    current_patner_feedback = current_patner_feedback
                else:
                    current_patner_feedback = "NA"

                change_current_patner = "no"
                if current_patner_feedback == "Dissatisfied" or current_patner_feedback == "Extremely Dissatisfied":
                    change_current_patner = "yes"

                submitted = "no"
                if meating_timeline is not "not given" and meating_timeline is not None:
                    submitted = "yes"

                contact_details = None
                if phone_number:
                    contact_details = ContactDetails.objects.filter(
                        Q(mobile=phone_number)|Q(landline=phone_number)).first()

                prefered_patners_array = []
                if prefered_patners:
                    prefered_patners_split = prefered_patners.split(",")
                    prefered_patners_array = [row.strip() for row in prefered_patners_split]

                prefered_patners_list = []
                prefered_patners_id_list = []
                preferred_company_other = None
                if prefered_patners_array:
                    prefered_patners_list = Organisation.objects.filter(name__in=prefered_patners_array).all()
                    prefered_patners_id_list = [row.organisation_id for row in prefered_patners_list]
                    prefered_patners_name_list = [row.name for row in prefered_patners_list]

                    if len(prefered_patners_name_list) < len(prefered_patners_array):
                        for prefered_patners_name in prefered_patners_array:
                            if prefered_patners_name not in prefered_patners_name_list:
                                preferred_company_other = prefered_patners_name
                                break

                supplier_id = ""
                supplier_type = "RS"
                sector = None
                if sector_name is not None:
                    sector = BusinessTypes.objects.filter(
                        business_type=sector_name.lower()).first()

                sub_sector = None
                if sub_sector_name is not None:
                    sub_sector = BusinessSubTypes.objects.filter(
                        business_sub_type=sub_sector_name.lower()).first()


                supplier = None
                city = None
                area = None
                supplier_name = None
                if contact_details:
                    
                    supplier_id = contact_details.object_id
                    supplier = SupplierTypeSociety.objects.filter(
                        supplier_id=supplier_id).first()
                    if supplier:
                        supplier_name = supplier.society_name
                        city = supplier.society_city
                        area = supplier.society_locality

                    else:
                        supplier = SupplierMaster.objects.filter(
                            supplier_id=supplier_id).first()

                        if supplier:
                            supplier_type = supplier.supplier_type
                            city = supplier.city
                            area = supplier.area
                            supplier_name = supplier.supplier_name

                campaign = None
                if supplier:
                    campaign = ProposalInfo.objects.filter(
                        Q(type_of_end_customer__formatted_name="b_to_b_r_g") 
                        & Q(name=area) | Q(name=city)).first()

                lead_status = b2b_utils.get_lead_status(
                    impl_timeline = impl_timeline,
                    meating_timeline = meating_timeline,
                    company=None,
                    prefered_patners=prefered_patners_list,
                    change_current_patner=change_current_patner.lower()
                    )

                if supplier and campaign:
            
                    campaign_id = campaign.proposal_id

                    shortlisted_spaces = ShortlistedSpaces.objects.filter(
                        proposal_id=campaign_id, object_id=supplier_id).first()

                    if not shortlisted_spaces:
                        content_type = ui_utils.get_content_type(supplier_type)
                        center = ProposalCenterMapping.objects.filter(
                            proposal_id=campaign_id).first()

                        shortlisted_spaces = ShortlistedSpaces(
                            proposal_id=campaign_id,
                            center=center,
                            supplier_code=supplier_type,
                            object_id=supplier_id,
                            content_type=content_type.data['data'],
                            status='F',
                            user=request.user,
                            requirement_given='yes',
                            requirement_given_date=datetime.datetime.now()
                        )
                        shortlisted_spaces.save()

                    shortlisted_spaces.requirement_given = 'yes'
                    shortlisted_spaces.requirement_given_date=datetime.datetime.now()
                    shortlisted_spaces.save()

                    current_patner_obj = None
                    current_company_other = None
                    if current_patner:
                        current_patner_obj = Organisation.objects.filter(
                            name=current_patner).first()
                        if not current_patner_obj:
                            current_company_other = current_patner

                    if submitted == "yes":
                        shortlisted_spaces.color_code = 1
                        shortlisted_spaces.save()
                        
                        pre_requirement = PreRequirement(
                            campaign_id=campaign_id,
                            shortlisted_spaces=shortlisted_spaces,
                            # company = company,
                            current_company = current_patner_obj,
                            current_company_other = current_company_other,
                            # is_current_patner = "yes" if current_patner_obj == company else "no",
                            current_patner_feedback = current_patner_feedback,
                            current_patner_feedback_reason = current_patner_feedback_reason,
                            preferred_company_other = preferred_company_other,
                            sector = sector,
                            sub_sector = sub_sector,
                            lead_by = contact_details,
                            impl_timeline = impl_timeline,
                            meating_timeline = meating_timeline,
                            lead_status = lead_status,
                            comment = comment,
                            varified_ops = 'no',
                            varified_bd = 'no',
                            lead_date = datetime.datetime.now(),
                            l1_answers = l1_answers,
                            l1_answer_2 = l1_answer_2,
                            l2_answers = l2_answers,
                            l2_answer_2 = l2_answer_2,
                            change_current_patner = change_current_patner.lower(),
                            call_back_preference = call_back_preference.lower(),
                        )
                        pre_requirement.save()

                        if prefered_patners_list:
                            pre_requirement.preferred_company.set(prefered_patners_list)

                    else:
                        if shortlisted_spaces.color_code != 1:
                            shortlisted_spaces.color_code = 2
                            shortlisted_spaces.save()

                        BrowsedLead(
                            supplier_id=supplier_id,
                            shortlisted_spaces_id=shortlisted_spaces.id,
                            campaign_id=campaign_id,
                            phone_number = phone_number,
                            supplier_name = supplier_name,
                            city = city,
                            area = area,
                            sector_id = sector.id if sector else None,
                            sub_sector_id = sub_sector.id if sub_sector else None,
                            implementation_timeline = impl_timeline,
                            meating_timeline = meating_timeline,
                            lead_status = lead_status,
                            comment = comment,
                            current_patner_id = current_patner_obj.organisation_id if current_patner_obj else None,
                            current_patner_other = current_company_other,
                            current_patner_feedback = current_patner_feedback,
                            current_patner_feedback_reason = current_patner_feedback_reason,
                            prefered_patners = prefered_patners_id_list,
                            prefered_patner_other = preferred_company_other,
                            status="closed",
                            created_at = datetime.datetime.now(),
                            updated_at = datetime.datetime.now(),
                            l1_answers = l1_answers,
                            l1_answer_2 = l1_answer_2,
                            l2_answers = l2_answers,
                            l2_answer_2 = l2_answer_2,
                            call_back_preference = call_back_preference.lower()
                        ).save()
                else:

                    SuspenseLead(
                        phone_number = phone_number,
                        supplier_name = supplier_name,
                        city = city,
                        area = area,
                        sector_name = sector_name,
                        sub_sector_name = sub_sector_name,
                        implementation_timeline = impl_timeline,
                        meating_timeline = meating_timeline,
                        lead_status = lead_status,
                        comment = comment,
                        current_patner = current_patner,
                        current_patner_feedback = current_patner_feedback,
                        current_patner_feedback_reason = current_patner_feedback_reason,
                        prefered_patners = prefered_patners_array,
                        created_at = datetime.datetime.now(),
                        updated_at = datetime.datetime.now(),
                        l1_answers = l1_answers,
                        l1_answer_2 = l1_answer_2,
                        l2_answers = l2_answers,
                        l2_answer_2 = l2_answer_2,
                        call_back_preference = call_back_preference.lower()
                    ).save()

        skip_row_list_split = ""
        if skip_row_list:
            skip_row_list_split = skip_row_list.split(",")

        return ui_utils.handle_response({}, data={"skip rows":skip_row_list_split}, success=True)
    
class RequirementClass(APIView):
    
    def get(self, request):
        shortlisted_spaces_id = request.query_params.get("shortlisted_spaces_id")
        requirements = PreRequirement.objects.filter(shortlisted_spaces_id=shortlisted_spaces_id)
        sectors = []
        verified_ops_user = {}
        for row in requirements:
            sectors.append(row.sector)

            if row.varified_ops_by and not verified_ops_user.get(row.varified_ops_by.id):
                verified_ops_user[row.varified_ops_by.id] = BaseUserSerializer(row.varified_ops_by,many=False).data
        requirement_obj = {}
        requirement_data = PreRequirementSerializer(requirements, many=True).data
        added_requirement = {}
        for row in requirement_data:
            key = str(row["lead_by"])+str(row["sub_sector"])+str(row["sector"])
            # if added_requirement.get(key):
            #     continue

            if not requirement_obj.get(row["sector"]):
                requirement_obj[row["sector"]] = dict(row)
                requirement_obj[row["sector"]]["requirements"] = []

            row["verified_ops_by_obj"] = verified_ops_user.get(row["varified_ops_by"])

            requirement_obj[row["sector"]]["requirements"].append(row)

            added_requirement[key] = True

        companies = Organisation.objects.filter(business_type__in=sectors)
        companies_data = OrganisationSerializer(companies, many=True).data
        
        return ui_utils.handle_response({}, data={"requirements": requirement_obj, "companies": companies_data}, success=True)

    def put(self, request):
        requirements = request.data.get("requirements")
        for req in requirements:

            update_req = {
                "current_company": req["current_company"],
                "current_company_other": req["current_company_other"],
                "preferred_company": req["preferred_company"],
                "preferred_company_other": req["preferred_company_other"],
                "impl_timeline": req["impl_timeline"],
                "meating_timeline": req["meating_timeline"],
                "comment": req["comment"],
                "lead_price": req["lead_price"],
            }

            call_back_preference = "NA"
            if req["call_back_preference"]:
                call_back_preference = req["call_back_preference"]
            update_req['call_back_preference'] = call_back_preference

            prefered_patners_list = Organisation.objects.filter(
                organisation_id__in=req["preferred_company"]).all()

            row = PreRequirement.objects.filter(id = req["id"]).first()

            lead_data = []

            change_current_patner = "no"
            if row.current_patner_feedback and (row.current_patner_feedback == "Dissatisfied" or row.current_patner_feedback == "Extremely Dissatisfied"):
                change_current_patner = "yes"
            
            lead_status = b2b_utils.get_lead_status(
                impl_timeline = req["impl_timeline"],
                meating_timeline = req["meating_timeline"],
                company=None,
                prefered_patners=prefered_patners_list,
                change_current_patner=change_current_patner
                )
            update_req['lead_status'] = lead_status
            requirement_data = PreRequirementSerializer(row, data=update_req)
            data = {}
            
            if requirement_data.is_valid():
                requirement_data.save()
                context = requirement_data.data
                if context['id'] == req["id"]:
                    lead_data.append({"lead_status":context['lead_status'],"id":context['id']})
            else:
                return ui_utils.handle_response({}, data={"errors":requirement_data.errors}, success=False)

        return ui_utils.handle_response({}, data=lead_data, success=True)


def remove_suspense_lead_cron():

    # Remove suspense leads
    all_suspense_lead = SuspenseLead.objects.all()
    
    for row in all_suspense_lead:
        contact_details = ContactDetails.objects.filter(Q(mobile=row.phone_number)|Q(landline=row.phone_number)).first()
        
        if contact_details:
            row.delete()
    
    # Close browsed leads
    prev_24h = datetime.datetime.now() - datetime.timedelta(days=1)

    BrowsedLead.objects.raw({'status': 'open', 'created_at': {"$lte": prev_24h}}).update({"$set":{"status":"closed"}})

    return HttpResponse(status=201)


class SuspenseLeadClass(APIView):
    permission_classes = ()
    authentication_classes = ()

    def get(self, request):

        header_list = ['Phone Number', 'Supplier Type', 'Supplier Name', 'POC Name', 'Designation', 'City', 'Area', 
            'Pin Code', 'Sector', 'Sub Sector', 'Current Partner', 'Current Patner Feedback',
            'Current Patner Feedback Reason', 'Preferred Partner','Implementation Timeline',
            'Meeting Time', 'L1 Answers','L1 Answer 2', 'L2 Answers','L2 Answer 2', 'Lead Status', 'Comment', 
            'Submitted', 'Call Back Preference']

        book = Workbook()
        sheet = book.active
        sheet.append(header_list)

        start_date = make_aware(datetime.datetime.strptime(request.GET.get("start_date"), '%Y-%m-%d'))
        end_date = make_aware(datetime.datetime.strptime(request.GET.get("end_date"), '%Y-%m-%d')) + datetime.timedelta(days=1)

        all_suspense_lead = list(mongo_client.suspense_lead.find({'created_at': {"$gte": start_date, "$lte": end_date}}))

        data = []
        for row in all_suspense_lead:
            row1 = dict(row)

            change_current_patner = "no"
            if row1['current_patner_feedback'] == "Dissatisfied" or row1['current_patner_feedback'] == "Extremely Dissatisfied":
                change_current_patner = "yes"

            try:
                l1_answer_2 = row1['l1_answer_2']
            except Exception as e:
                l1_answer_2 = None

            try:
                l2_answer_2 = row1['l2_answer_2']
            except Exception as e:
                l2_answer_2 = None

            try:
                call_back_preference = row1['call_back_preference']
            except Exception as e:
                call_back_preference = None

            try:
                lead_status = row1['lead_status']
            except Exception as e:
                lead_status = None

            try:
                poc_name = row1['poc_name']
            except Exception as e:
                poc_name = None
            
            try:
                designation = row1['designation']
            except Exception as e:
                designation = None

            try:
                supplier_type = row1['supplier_type']
                supplier_type_name = supplier_code_to_names[supplier_type]
            except Exception as e:
                supplier_type_name = None

            try:
                pin_code = row1['pin_code']
            except Exception as e:
                pin_code = None

            row2 = [
                row1['phone_number'],
                supplier_type_name,
                row1['supplier_name'],
                poc_name,
                designation,
                row1['city'],
                row1['area'],
                pin_code,
                row1['sector_name'],
                row1['sub_sector_name'],
                row1['current_patner'],
                row1['current_patner_feedback'],
                row1['current_patner_feedback_reason'],
                ", ".join(row1['prefered_patners']),
                row1['implementation_timeline'],
                row1['meating_timeline'],
                row1['l1_answers'] if row1['l1_answers'] else None,
                l1_answer_2,
                row1['l2_answers'] if row1['l2_answers'] else None,
                l2_answer_2,
                lead_status,
                row1['comment'],
                "no",
                call_back_preference,
            ]
            sheet.append(row2)

        resp = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = 'attachment; filename=mydata.xlsx'
        book.save(resp)

        return resp

class BrowsedLeadClass(APIView):

    def get(self, request):
        shortlisted_spaces_id = request.query_params.get("shortlisted_spaces_id")
        browsed_leads = BrowsedLead.objects.raw({"shortlisted_spaces_id":shortlisted_spaces_id, "status":"closed"}).values()
        phone_numers = [row.get("phone_number") for row in browsed_leads]

        contact_details = ContactDetails.objects.filter(Q(mobile__in=phone_numers)|Q(landline__in=phone_numers))
        contact_details_dict_mobile = {str(row.mobile):row.name for row in contact_details}
        contact_details_dict_landline = {str(row.landline):row.name for row in contact_details}
        sectors = []
        list1 = []
        for row in browsed_leads:
            row1 = dict(row)
            row1["_id"] = str(row1["_id"])

            row1["lead_by_name"] = contact_details_dict_mobile.get(row1.get("phone_number"))
            created_at = datetime.datetime.strptime(str(row1['created_at']), '%Y-%m-%d %H:%M:%S.%f').strftime("%Y-%m-%dT%H:%M:%S.%fZ")
            row1['created_at'] = created_at
            if not row1["lead_by_name"]:
                row1["lead_by_name"] = contact_details_dict_landline.get(row.get("phone_number"))

            list1.append(row1)
            sectors.append(row["sector_id"])
        
        companies = Organisation.objects.filter(business_type__in=sectors)
        companies_data = OrganisationSerializer(companies, many=True).data

        return ui_utils.handle_response({}, data={"browsed": list1, "companies": companies_data}, success=True)

class BrowsedLeadDelete(APIView):

    def post(self, request):
        browsed_ids = request.data.get("browsed")
        list_color_code = None
        for browsed_id in browsed_ids:
            browsed = mongo_client.browsed_lead.update({"_id": ObjectId(browsed_id["_id"])}, {"$set":{"status":"deleted"}})

            shortlisted_spaces = browsed_id['shortlisted_spaces_id']

        if shortlisted_spaces:
            list_color_code = b2b_utils.get_color_code(shortlisted_spaces)
            
        return ui_utils.handle_response({}, data={"message":"Browsed lead deleted successfully","list_color_code":list_color_code}, success=True)

class DeleteRequirement(APIView):

    def post(self, request):
        requirement_ids = request.data.get('requirement_ids')
        list_color_code = None

        for req in requirement_ids:
            requirement = PreRequirement.objects.filter(id=req).first()
            requirement.is_deleted="yes"
            requirement.save()

        if requirement.shortlisted_spaces:
            list_color_code = b2b_utils.get_color_code(requirement.shortlisted_spaces.id)

        return ui_utils.handle_response({}, data={"message":"Requirement deleted","list_color_code":list_color_code}, success=True)


class RestoreRequirement(APIView):
    # Requirement restore api (deletion revert api)

    def post(self, request):
        requirement_ids = request.data.get('requirement_ids')
        list_color_code = None

        for req in requirement_ids:
            requirement= PreRequirement.objects.filter(id=req).first()
            requirement.is_deleted = "no"
            requirement.save()
        
        if requirement.shortlisted_spaces:
            list_color_code = b2b_utils.get_color_code(requirement.shortlisted_spaces.id)
            
        return ui_utils.handle_response({}, data={"message":"Requirement restored","list_color_code":1}, success=True)
        
class LeadOpsVerification(APIView):

    def post(self, request):
        requirement_ids = request.data.get("requirement_ids")
        requirements = PreRequirement.objects.filter(id__in=requirement_ids)

        for requirement in requirements:
            companies = Organisation.objects.filter(business_type=requirement.sector)
            verified = 0
            if companies:
                if requirement.varified_ops == "no":
                    for company in companies:

                        company_campaign = ProposalInfo.objects.filter(type_of_end_customer__formatted_name="b_to_b_l_d",
                            account__organisation=company).first()
                        if company_campaign:
                            requirement.varified_ops = "yes"
                            requirement.varified_ops_date = datetime.datetime.now()
                            requirement.varified_ops_by = request.user

                            company_shortlisted_spaces = ShortlistedSpaces.objects.filter(object_id=requirement.shortlisted_spaces.object_id,
                                proposal=company_campaign.proposal_id).first()

                            if not company_shortlisted_spaces:
                                center = ProposalCenterMapping.objects.filter(proposal=company_campaign).first()
                                content_type = ui_utils.get_content_type(requirement.shortlisted_spaces.supplier_code)

                                company_shortlisted_spaces = ShortlistedSpaces(
                                    proposal=company_campaign,
                                    center=center,
                                    object_id=requirement.shortlisted_spaces.object_id,
                                    supplier_code=requirement.shortlisted_spaces.supplier_code,
                                    content_type=content_type.data['data'],
                                    status='F',
                                    user=request.user,
                                    requirement_given='yes',
                                    requirement_given_date=datetime.datetime.now(),
                                    color_code = 1
                                )
                                company_shortlisted_spaces.save()

                            requirement.company_campaign = company_campaign
                            requirement.company_shortlisted_spaces = company_shortlisted_spaces

                            shortlisted_spac = ShortlistedSpaces.objects.filter(
                                id=company_shortlisted_spaces.id).first()
                            if shortlisted_spac:
                                shortlisted_spac.color_code = 1
                                shortlisted_spac.save()
                            
                            preferred_partners_list = requirement.preferred_company.all()

                            is_preferred_company = "no"
                            if company_campaign.account.organisation:
                
                                if company_campaign.account.organisation in preferred_partners_list:
                                    is_preferred_company = "yes"

                            new_requirement = Requirement(
                            campaign_id=requirement.campaign_id,
                            shortlisted_spaces=requirement.shortlisted_spaces,
                            company = company,
                            current_company = requirement.current_company,
                            current_company_other = requirement.current_company_other,
                            is_current_patner = "yes" if requirement.current_company == company else "no",
                            current_patner_feedback = requirement.current_patner_feedback,
                            current_patner_feedback_reason = requirement.current_patner_feedback_reason,
                            preferred_company_other = requirement.preferred_company_other,
                            sector = requirement.sector,
                            sub_sector = requirement.sub_sector,
                            lead_by = requirement.lead_by,
                            impl_timeline = requirement.impl_timeline,
                            meating_timeline = requirement.meating_timeline,
                            lead_status = requirement.lead_status,
                            comment = requirement.comment,
                            varified_ops = 'yes',
                            varified_bd = 'no',
                            lead_date = requirement.lead_date,
                            lead_price  = requirement.lead_price,
                            l1_answers = requirement.l1_answers,
                            l1_answer_2 = requirement.l1_answer_2,
                            l2_answers = requirement.l2_answers,
                            l2_answer_2 = requirement.l2_answer_2,
                            change_current_patner = requirement.change_current_patner.lower(),
                            company_campaign=company_campaign,
                            company_shortlisted_spaces=company_shortlisted_spaces,
                            varified_ops_by = request.user,
                            varified_ops_date = datetime.datetime.now(),
                            call_back_preference = requirement.call_back_preference,
                            is_preferred_company = is_preferred_company
                            )
                            new_requirement.save()
                            verified += 1
                            
                            if preferred_partners_list:
                                new_requirement.preferred_company.set(preferred_partners_list)
                    requirement.save()
            else:
                return ui_utils.handle_response({}, data={"error":"No companies for the service found"}, success=False)
        color_code = None
        list_color_code = None
        if requirement.shortlisted_spaces:
            list_color_code = b2b_utils.get_color_code(requirement.shortlisted_spaces.id)
            
        if verified == 0:
            return ui_utils.handle_response({}, data={"error":"Ops verify failed as there are 0 client campaigns","color_code":color_code,"verified_ops_by":request.user.first_name + request.user.last_name,"list_color_code":list_color_code}, success=False)
        else:
            return ui_utils.handle_response({}, data={"message":"Ops Verified and distributed to "+str(verified)+" campaigns","color_code":color_code,"verified_ops_by":request.user.first_name + request.user.last_name,"list_color_code":list_color_code}, success=True)

class BrowsedToRequirement(APIView):

    def post(self, request):
        browsed_ids = request.data.get("browsed_ids")
        
        shortlisted_spaces_id = None

        for browsed_id in browsed_ids:
            browsed = dict(mongo_client.browsed_lead.find_one({"_id": ObjectId(browsed_id["_id"])}))
            if browsed:

                if browsed["meating_timeline"] == "" or browsed["meating_timeline"] == "not given" or browsed["meating_timeline"] == None:
                    return ui_utils.handle_response({}, data={
                        "error":"meeting time not given"}, success=True)

                contact_details = None
                if browsed["phone_number"]:
                    contact_details = ContactDetails.objects.filter(Q(mobile=browsed["phone_number"])|Q(landline=browsed["phone_number"])).first()
                
                prefered_patners_list = []
                if browsed["prefered_patners"]:
                    prefered_patners_list = Organisation.objects.filter(organisation_id__in=browsed["prefered_patners"]).all()
                
                shortlisted_spaces_id = browsed["shortlisted_spaces_id"]

                change_current_patner = "no"
                if browsed["current_patner_feedback"] == "Dissatisfied" or browsed["current_patner_feedback"] == "Extremely Dissatisfied":
                    change_current_patner = "yes"

                lead_status = b2b_utils.get_lead_status(
                    impl_timeline = browsed["implementation_timeline"],
                    meating_timeline = browsed["meating_timeline"],
                    company=None,
                    prefered_patners=prefered_patners_list,
                    change_current_patner=change_current_patner.lower()
                    )

                try:
                    call_back_preference = browsed["call_back_preference"]
                except Exception as e:
                    call_back_preference = "NA"

                if browsed["current_patner_id"] == "":
                    current_patner_id = None
                else:
                    current_patner_id = browsed["current_patner_id"]

                requirement = PreRequirement(
                    campaign_id=browsed["campaign_id"],
                    shortlisted_spaces_id=browsed["shortlisted_spaces_id"],
                    current_company_id = current_patner_id,
                    current_patner_feedback = browsed["current_patner_feedback"],
                    current_patner_feedback_reason = browsed["current_patner_feedback_reason"],
                    sector_id = browsed["sector_id"],
                    lead_by = contact_details,
                    impl_timeline = browsed["implementation_timeline"],
                    meating_timeline = browsed["meating_timeline"],
                    comment = browsed["comment"],
                    varified_ops = 'no',
                    varified_bd = 'no',
                    lead_status = lead_status,
                    lead_date = datetime.datetime.now(),
                    preferred_company_other = browsed["prefered_patner_other"],
                    current_company_other = browsed["current_patner_other"],
                    l1_answers = browsed["l1_answers"],
                    l1_answer_2 = browsed["l1_answer_2"],
                    l2_answers = browsed["l2_answers"],
                    l2_answer_2 = browsed["l2_answer_2"],
                    sub_sector_id = browsed["sub_sector_id"],
                    call_back_preference = call_back_preference
                )
                requirement.save()

                if prefered_patners_list:
                    requirement.preferred_company.set(prefered_patners_list)

                mongo_client.browsed_lead.update({"_id": ObjectId(browsed_id["_id"])}, {"$set":{"status":"converted"}})

        if shortlisted_spaces_id:
            ShortlistedSpaces.objects.filter(id=shortlisted_spaces_id).update(color_code=1, requirement_given='yes', requirement_given_date=datetime.datetime.now())

        return ui_utils.handle_response({}, data={"message":"Lead submitted successfully","list_color_code":1}, success=True)


class UpdateBrowsedLead(APIView):

    def post(self, request):

        browsed_leads = request.data.get("browsed_leads")

        for browsed in browsed_leads:

            prefered_patners_id_list = browsed["prefered_patners_id"]
            
            try:
                call_back_preference = browsed["call_back_preference"]
            except Exception as e:
                call_back_preference = "NA"

            impl_timeline = "not given"
            if browsed["implementation_timeline"] is not "":
                impl_timeline = browsed["implementation_timeline"]

            meating_timeline = "not given"
            if browsed["meating_timeline"] is not "":
                meating_timeline = browsed["meating_timeline"]

            if browsed["current_patner_id"] == "":
                current_patner_id = None
            else:
                current_patner_id = browsed["current_patner_id"]

            update_values = {"$set":{
                "current_patner_id":current_patner_id,
                "prefered_patners":prefered_patners_id_list,
                "implementation_timeline":impl_timeline,
                "meating_timeline":meating_timeline,
                "call_back_preference":call_back_preference,
                "comment":browsed["comment"],
                "current_patner_other":browsed["current_company_other"],
                "prefered_patner_other":browsed["preferred_company_other"]
                }}

            mongo_client.browsed_lead.update({"_id": ObjectId(browsed["_id"])},update_values)

        return ui_utils.handle_response({}, data={"message":"Browsed lead updated successfully"}, success=True)


class BdVerification(APIView):
    """docstring for BdVerification"""
    def post(self, request):
        requirement_ids = request.data.get("requirement_ids")

        requirements = Requirement.objects.filter(id__in=requirement_ids)

        now = datetime.datetime.utcnow() + datetime.timedelta(hours=5, minutes=30)

        for requirement in requirements:

            if requirement.varified_bd == "no":
                
                if requirement.company_campaign:
                
                    lead_form = mongo_client.leads_forms.find_one({"campaign_id": requirement.company_campaign.proposal_id})
                    if lead_form:

                        requirement.varified_bd = "yes"
                        requirement.varified_bd_by = request.user
                        requirement.varified_bd_date = now
                        requirement.save()

                        self.insert_lead_data(lead_form, requirement, requirement.campaign)
                        if requirement.lead_by:
                            phone = "91" + str(requirement.lead_by.mobile)
                            b2b_utils.send_whatsapp_notification(None,None,
                                phone)
                    else:
                        return ui_utils.handle_response({}, data="Please add lead form for this campaign to BD verify",
                         success=True)

        color_code = None
        if requirement.company_shortlisted_spaces:
        
            list_color_code = requirement.company_shortlisted_spaces.color_code
            requirement_exist = Requirement.objects.filter(company_shortlisted_spaces=requirement.company_shortlisted_spaces,
             varified_bd = "no")
            if not requirement_exist:
                shortlisted_spac = ShortlistedSpaces.objects.filter(
                    id=requirement.company_shortlisted_spaces.id).first()
                shortlisted_spac.color_code = 3
                shortlisted_spac.save()
                color_code = 3
                list_color_code = color_code

        return ui_utils.handle_response({}, data={"list_color_code":list_color_code,"color_code":color_code,"varified_bd_by":request.user.first_name + request.user.last_name}, success=True)


    def insert_lead_data(self, lead_form, requirement, campaign):
        entry_id = lead_form['last_entry_id'] + 1 if 'last_entry_id' in lead_form else 1
        lead_data = []
        
        supplier_name = ""
        supplier_city = ""
        supplier_area = ""
        supplier_subarea = ""
        supplier_primary_count = ""

        supplier_state = ""
        supplier_pin_code = ""
        supplier_contact_person_name = ""
        supplier_designation = ""
        supplier_moblile = ""
        
        if requirement.shortlisted_spaces.supplier_code == 'RS':
            supplier = SupplierTypeSociety.objects.filter(supplier_id = requirement.shortlisted_spaces.object_id).first()

            if supplier:
                supplier_name = supplier.society_name
                supplier_city = supplier.society_city
                supplier_area = supplier.society_locality
                supplier_subarea = supplier.society_subarea
                address = str(supplier.society_address1) + " " + str(supplier.society_address2)
                lat_long = str(supplier.society_latitude) + "/" + str(supplier.society_longitude)
                supplier_primary_count = supplier.flat_count

                supplier_state = supplier.society_state
                supplier_pin_code = supplier.society_zip

                supplier_contact_person_name = requirement.lead_by.name
                supplier_designation = requirement.lead_by.contact_type
                supplier_moblile = requirement.lead_by.mobile
        else:
            supplier = SupplierMaster.objects.filter(supplier_id = requirement.shortlisted_spaces.object_id).first()
            if supplier:
                supplier_name = supplier.supplier_name
                supplier_city = supplier.city
                supplier_area = supplier.area
                supplier_subarea = supplier.subarea
                address = str(supplier.address_supplier.address1) + " " + str(supplier.address_supplier.address2)
                lat_long = str(supplier.latitude) + "/" + str(supplier.longitude)
                supplier_primary_count = supplier.unit_primary_count

                supplier_state = supplier.state
                supplier_pin_code = supplier.zipcode

                supplier_contact_person_name = requirement.lead_by.name
                supplier_designation = requirement.lead_by.contact_type
                supplier_moblile = requirement.lead_by.mobile
        
        prefered_patner = "no"
        for row in requirement.preferred_company.all():
            if requirement.company == row:
                prefered_patner = "yes"
    
        
        lead_status = requirement.lead_status
        lead_form_key = None
        for key, lead_form_keys in lead_form["data"].items():
            if lead_form_keys["key_name"].lower() == lead_status.lower():
                lead_form_key = lead_form_keys["item_id"]
        
        lead_form_key_2 = None
        if lead_form_key:
            for key, value in lead_form["global_hot_lead_criteria"].items():
                if value.get("or"):
                    for key1, value1 in value["or"].items():
                        
                        if str(key1) == str(lead_form_key):
                            lead_form_key_2 = key
            if lead_form_key_2 and lead_form["hotness_mapping"].get(lead_form_key_2):
                lead_status = lead_form["hotness_mapping"].get(lead_form_key_2)

        if requirement.hotness_of_lead:
            h2 = None
            h3 = None 
            h4 = None
            h5 = None
            h6 = None
            level = requirement.hotness_of_lead
            if level == 'H2':
                h2 = 'Y'
            if level == 'H3':
                h2 = 'Y'
                h3 = 'Y'
            if level == 'H4':
                h2 = 'Y'
                h3 = 'Y'
                h4 = 'Y'
            if level == 'H5':
                h2 = 'Y'
                h3 = 'Y'
                h4 = 'Y'
                h5 = 'Y'
            if level == 'H6':
                h2 = 'Y'
                h3 = 'Y'
                h4 = 'Y'
                h5 = 'Y'
                h6 = 'Y'

        lead_data_dict = {
            "Supplier Name": supplier_name,
            "Supplier Type": requirement.shortlisted_spaces.supplier_code,
            "Area": supplier_area,
            "Address": address,
            "Lat/Long": lat_long,
            "City": supplier_city,
            "State": supplier_state,
            "Country": "India",
            "Pin Code": supplier_pin_code,
            "Primary Count": supplier_primary_count,
            "Service" : requirement.sector.business_type if requirement.sector else None ,
            "Sub service" : requirement.sub_sector.business_sub_type if requirement.sub_sector else None ,
            "L1.1 Answer" : requirement.l1_answers,
            "L1.2 Answer": requirement.l1_answer_2,
            "L2.1 Answer": requirement.l2_answers,
            "L2.2 Answer": requirement.l2_answer_2,
            "L3.1 Answer": None,
            "L3.2 Answer": None,
            "Prefered Patner": prefered_patner,
            "Implementation Time": requirement.impl_timeline,
            "Meeting Time": requirement.meating_timeline,
            "Call back time" : requirement.call_back_preference,
            "Comments" : requirement.comment,
            "Time Stamp" : datetime.datetime.strftime(requirement.varified_bd_date, '%Y-%m-%d %H:%M:%S'),
            "Lead Status": lead_status,
            "H2" : h2, 
            "H3" : h3,
            "H4" : h4,
            "H5" : h5,
            "H6" : h6,
            "Not Applicable" : None,
            "Name": supplier_contact_person_name,
            "Number": supplier_moblile,
            "Designation": supplier_designation,
            "Current Partner": requirement.is_current_patner,
            "Satisfaction Level" : requirement.current_patner_feedback,
            "Reasons for Dissatisfaction" : requirement.current_patner_feedback_reason,
            "Price": requirement.lead_price,
            "Client Status":requirement.client_status,
        }

        lead_data = []
        i = 1
        for key, value in lead_data_dict.items():
            row = {
                    'key_name': key,
                    'value': value,
                    'item_id': i,
                    'key_type': 'STRING'
                }
            lead_data.append(row)

            i+=1

        lead_dict = {"data": lead_data, "is_hot": False, "created_at": datetime.datetime.now(),
            "supplier_id": requirement.shortlisted_spaces.object_id, "campaign_id": requirement.campaign_id,
            "leads_form_id": lead_form['leads_form_id'], "entry_id": entry_id, "status": "active",
            "lead_status": requirement.lead_status, "lead_purchased": "no", "lead_existing_client": "no",
            "company_campaign_id": requirement.company_campaign_id, "requrement_id":requirement.id,
            "company_lead_status":lead_status, "is_current_company":requirement.is_current_patner,
            "current_patner_feedback":requirement.current_patner_feedback,
            "current_patner_feedback_reason":requirement.current_patner_feedback_reason,
            "company_id":requirement.company.organisation_id,"meating_timeline":requirement.meating_timeline,
            "impl_timeline":requirement.impl_timeline,"lead_date":requirement.varified_bd_date,
            "preferred_patner":prefered_patner,"lead_price":requirement.lead_price,
            "supplier_primary_count":supplier_primary_count,"supplier_city": supplier_city,
            "supplier_area": supplier_area,"supplier_sub_area": supplier_subarea,
            "purchased_date": requirement.purchased_date,"client_status":requirement.client_status,
            "supplier_type":requirement.shortlisted_spaces.supplier_code}

        lead_for_hash = {
            "data": lead_data,
            "leads_form_id": lead_form['leads_form_id']
        }
        
        lead_sha_256 = self.create_lead_hash(lead_for_hash)
        lead_dict["lead_sha_256"] = lead_sha_256

        mongo_client.leads.insert_one(lead_dict)

        campaign_lead = mongo_client.CampaignLeads.find_one({
            "company_campaign_id":requirement.company_campaign_id})
        if campaign_lead:
            mongo_client.CampaignLeads.update_one({"company_campaign_id": 
                requirement.company_campaign_id},{"$set": {
                        "lead_count": campaign_lead['lead_count'] + 1,
                        "updated_at": datetime.datetime.now()
                    }})
        else:
            campaign_leads_dict = {
                "updated_at": datetime.datetime.now(),
                "created_at": datetime.datetime.now(),
                "lead_count": 1,
                "company_campaign_id": requirement.company_campaign_id
            }
            mongo_client.CampaignLeads.insert_one(campaign_leads_dict)

        company_lead = mongo_client.OrganizationLeads.find_one({
            "company_id":requirement.company.organisation_id})
        if company_lead:
            mongo_client.OrganizationLeads.update_one({"company_id": 
                requirement.company.organisation_id},{"$set": {
                        "lead_count": company_lead['lead_count'] + 1,
                        "updated_at": datetime.datetime.now()
                    }})
        else:
            company_leads_dict = {
                "updated_at": datetime.datetime.now(),
                "created_at": datetime.datetime.now(),
                "lead_count": 1,
                "company_id": requirement.company.organisation_id
            }
            mongo_client.OrganizationLeads.insert_one(company_leads_dict)

        return True

    def create_lead_hash(self, lead_dict):
        lead_hash_string = ''
        lead_hash_string += str(lead_dict['leads_form_id'])

        for item in lead_dict['data']:
            if item['value']:
                if isinstance(item["value"], (str,bytes)):
                    lead_hash_string += str(item['value'].strip())
                else:
                    lead_hash_string += str(item['value'])
        return hashlib.sha256(lead_hash_string.encode('utf-8')).hexdigest()


class BdRequirement(APIView):
    
    def get(self, request):
        company_shortlisted_spaces_id = request.query_params.get("company_shortlisted_spaces_id")

        requirements = Requirement.objects.filter(company_shortlisted_spaces_id=company_shortlisted_spaces_id)
        
        sectors = []
        verified_ops_user = {}
        verified_bd_user = {}
        for row in requirements:
            sectors.append(row.sector)

            if row.varified_ops_by and not verified_ops_user.get(row.varified_ops_by.id):
                verified_ops_user[row.varified_ops_by.id] = BaseUserSerializer(row.varified_ops_by,many=False).data

            if row.varified_bd_by and not verified_bd_user.get(row.varified_bd_by.id):
                verified_bd_user[row.varified_bd_by.id] = BaseUserSerializer(row.varified_bd_by,many=False).data

        requirement_obj = {}
        requirement_data = RequirementSerializer(requirements, many=True).data
        added_requirement = {}
        for row in requirement_data:
            key = str(row["lead_by"])+str(row["sub_sector"])+str(row["sector"])

            if not requirement_obj.get(row["sector"]):
                requirement_obj[row["sector"]] = dict(row)
                requirement_obj[row["sector"]]["requirements"] = []

            # preferred_company_list = row["preferred_company"]

            # preferred_organisation = ProposalInfo.objects.filter(
            #     account__organisation__in=preferred_company_list)

            # if preferred_organisation:
            #     row["is_preferred_company"] = "yes"
            # else:
            #     row["is_preferred_company"] = "no"

            row["verified_ops_by_obj"] = verified_ops_user.get(row["varified_ops_by"])
            row["verified_bd_by_obj"] = verified_bd_user.get(row["varified_bd_by"])

            requirement_obj[row["sector"]]["requirements"].append(row)

            added_requirement[key] = True

        companies = Organisation.objects.filter(business_type__in=sectors)
        companies_data = OrganisationSerializer(companies, many=True).data
        
        return ui_utils.handle_response({}, data={"requirements": requirement_obj, "companies": companies_data}, success=True)


class GetLeadsByCampaignId(APIView):

    def get(self, request):

        where = {"lead_purchased": request.query_params.get('is_purchased'), "client_status":"Accepted"}

        if request.query_params.get("campaign_id"):
            where["company_campaign_id"] = request.query_params.get("campaign_id")
        else:
            where["company_id"] = request.user.profile.organisation.organisation_id

        leads_data = mongo_client.leads.find(where)
        data = {}
        if leads_data is not None:

            leads_data_list = list(leads_data)
            suppliers_list = []
            for lead_data in leads_data_list:
                suppliers_list.append(lead_data['supplier_id'])
            suppliers_list = list(set(suppliers_list))
            
            master_societies = SupplierMaster.objects.filter(
                supplier_id__in=suppliers_list)
            master_serializer = SupplierMasterSerializer(master_societies, many=True)
            
            supplire_societies = SupplierTypeSociety.objects.filter(
                supplier_id__in=suppliers_list)
            supplire_serializer = SupplierTypeSocietySerializer(supplire_societies, many=True)
            
            all_societies = manipulate_object_key_values(supplire_serializer.data)
            master_suppliers = manipulate_master_to_rs(master_serializer.data)
            
            supplier_data = {}
            all_societies.extend(master_suppliers)
            for supplr in all_societies:
                supplier_data[supplr['supplier_id']] = supplr
        
            data = []
            led = {}
            supplier = []
            for lead in leads_data_list:
                led = dict(lead)
                led['_id'] = str(led['_id'])
                led['supplier_data'] = supplier_data.get(lead['supplier_id'])
                data.append(led)
                
        return ui_utils.handle_response({}, data=data, success=True)

class GetLeadsForDonutChart(APIView):

    def get(self, request):
       
        where = {"client_status":"Accepted"}
        
        if request.query_params.get("campaign_id"):
            where["company_campaign_id"] = request.query_params.get("campaign_id")
        else:
            where["company_id"] = request.user.profile.organisation.organisation_id

        total_leads = mongo_client.leads.find(where).count()
        data = {}
        if total_leads:
            where["lead_purchased"] = "yes"
            leads_purchased = mongo_client.leads.find(where).count()
            leads_remain = total_leads-leads_purchased

            data = {
                "total_leads": total_leads,
                "leads_purchased_per": (leads_purchased*100)/total_leads,
                "leads_remain": leads_remain,
                "leads_remain_per": (leads_remain*100)/total_leads,
                "total_leads_purchased": leads_purchased,
            }
        return ui_utils.handle_response({}, data=data, success=True)
        

class GetLeadsSummeryForDonutChart(APIView):

    def get(self, request):
       
        where = {"is_current_company": "yes", "client_status":"Accepted"}
        if request.query_params.get("campaign_id"):
            where["company_campaign_id"] = request.query_params.get("campaign_id")
        else:
            where["company_id"] = request.user.profile.organisation.organisation_id

        total_leads = mongo_client.leads.find(where).count()
        data = {}
        if total_leads:
            where["current_patner_feedback"] = "Satisfied"
            total_satisfied = mongo_client.leads.find(where).count()

            where["lead_purchased"] = "yes"
            where["current_patner_feedback"] = { "$ne": "Satisfied" }
            
            dissatisfied_purchased = mongo_client.leads.find(where).count()

            dissatisfied_purchased_per = (dissatisfied_purchased*100)/total_leads

            where["lead_purchased"] = "no"
            dissatisfied_not_purchased = mongo_client.leads.find(where).count()

            dissatisfied_not_purchased_per = (dissatisfied_not_purchased*100)/total_leads

            satisfied_per = (total_satisfied*100)/total_leads

            data = {
                "total_leads": total_leads,
                "dissatisfied_purchased": dissatisfied_purchased,
                "dissatisfied_purchased_per":round(dissatisfied_purchased_per, 2),
                "dissatisfied_not_purchased": dissatisfied_not_purchased,
                "dissatisfied_not_purchased_per":round(dissatisfied_not_purchased_per, 2),
                "total_satisfied": total_satisfied,
                "satisfied_per": round(satisfied_per, 2),
            }
        return ui_utils.handle_response({}, data=data, success=True)
        

class GetLeadsForCurrentCompanyDonut(APIView):

    def get(self, request):
        
        where = {"is_current_company": "yes","lead_purchased":request.query_params.get("is_purchased"), "client_status":"Accepted"}
        if request.query_params.get("campaign_id"):
            where["company_campaign_id"] = request.query_params.get("campaign_id")
        else:
            where["company_id"] = request.user.profile.organisation.organisation_id

        if request.query_params.get("is_satisfied") == "yes":
            where["current_patner_feedback"] = "Satisfied"
        else:
            where["current_patner_feedback"] = { "$ne": "Satisfied" }

        lead_data = mongo_client.leads.find(where)
        leads_list = list(lead_data)
        total_leads = len(leads_list)
        context = {}
        if total_leads:
            suppliers_list = []
            for lead_data in leads_list:
                suppliers_list.append(lead_data['supplier_id'])
            suppliers_list = list(set(suppliers_list))
        
            master_societies = SupplierMaster.objects.filter(
                supplier_id__in=suppliers_list).exclude(supplier_type="RS")
            master_serializer = SupplierMasterSerializer(master_societies, many=True)
            
            supplire_societies = SupplierTypeSociety.objects.filter(
                supplier_id__in=suppliers_list,supplier_code="RS")
            supplire_serializer = SupplierTypeSocietySerializer(supplire_societies, many=True)
            
            all_societies = manipulate_object_key_values(supplire_serializer.data)
            master_suppliers = manipulate_master_to_rs(master_serializer.data)
            all_societies.extend(master_suppliers)

            supplier_data = {}
            for supplr in all_societies:
                supplier_data[supplr['supplier_id']] = supplr

            data = []
            s_led_obj = {}
            for lead in leads_list:
                lead_obj = dict(lead)
                lead_obj['_id'] = str(lead_obj['_id'])
                lead_obj['supplier_data'] = supplier_data.get(lead_obj['supplier_id'])
                data.append(lead_obj)

            context = {
                "lead_data":data,
                "total_leads":total_leads
            }

        return ui_utils.handle_response({}, data=context, success=True)

class AddLeadPrice(APIView):
    # Update requirement price and comment api

    def post(self, request):
        data = request.data.get('data')
        for row in data:
            requirement = Requirement.objects.filter(id=row['requirement_id']).first()
            requirement.lead_price = row['lead_price']
            requirement.comment = row['comment']
            requirement.client_status = row['client_status']
            requirement.hotness_of_lead = row['hotness_of_lead'].upper()
            requirement.save()
        return ui_utils.handle_response({}, data="Price and comment added", success=True)

class GetLeadsByDate(APIView):

    def get(self, request):

        date = request.query_params.get('date')
        date_time_obj = datetime.datetime.strptime(date, '%Y-%m-%d %H:%M:%S.%f')
        start_date = date_time_obj.replace(hour=0, minute=0, second=0)
        end_date = date_time_obj.replace(hour=23, minute=59, second=59)
        organisation_id = request.user.profile.organisation.organisation_id
        lead_count = mongo_client.leads.find({"$and": [{"created_at":{"$gte": start_date, "$lte": end_date}}, {"company_id": organisation_id}, {"client_status": "Accepted"}]}).count()
        existing_client_count = mongo_client.leads.find({"$and": [{"created_at":{"$gte": start_date, "$lte": end_date}}, {"company_id": organisation_id}, {"is_current_company":"yes"}, {"current_patner_feedback": { "$in": ["Dissatisfied", "Extremely Dissatisfied"]}}, {"client_status": "Accepted"}]}).count()
            
        lead_dict = {
            'lead_count' : lead_count,
            'existing_client_count' : existing_client_count,
        }
        return ui_utils.handle_response({}, data=lead_dict, success=True)

class GetLeadsCampaignByDate(APIView):

    def get(self, request):

        date = request.query_params.get('date')
        date_time_obj = datetime.datetime.strptime(date, '%Y-%m-%d %H:%M:%S.%f')
        start_date = date_time_obj.replace(hour=0, minute=0, second=0)
        end_date = date_time_obj.replace(hour=23, minute=59, second=59)
        organisation_id = request.user.profile.organisation.organisation_id

        leads = mongo_client.leads.find({"$and": [{"created_at":{"$gte": start_date, "$lte": end_date}}, {"company_id": organisation_id}, {"client_status": "Accepted"}]})
        campaign_ids = set()
        lead_count_purchased_map = {}
        lead_count_not_purchased_map = {}

        for row in leads:
            campaign_ids.add(row["company_campaign_id"])

            if not lead_count_purchased_map.get(row["company_campaign_id"]):
                lead_count_purchased_map[row["company_campaign_id"]] = 0
            
            if row["lead_purchased"] == "yes":
                lead_count_purchased_map[row["company_campaign_id"]] += 1
            
            if not lead_count_not_purchased_map.get(row["company_campaign_id"]):
                lead_count_not_purchased_map[row["company_campaign_id"]] = 0

            if row["lead_purchased"] == "no":
                lead_count_not_purchased_map[row["company_campaign_id"]] += 1
            
        campaign_ids = list(campaign_ids)

        campaigns = ProposalInfo.objects.filter(proposal_id__in=campaign_ids)
        campaign_data = ProposalInfoSerializer(campaigns, many=True).data

        campaign = {}
        for row in campaign_data:
            row["purchased_count"] = lead_count_purchased_map.get(row["proposal_id"])
            row["not_purchased_count"] = lead_count_not_purchased_map.get(row["proposal_id"])
            
        lead_dict = {
            'campaigns' : campaign_data,
        }
        return ui_utils.handle_response({}, data=lead_dict, success=True)


class GetFeedbackCount(APIView):

    def get(self, request):

        date = request.query_params.get('date')
        date_time_obj = datetime.datetime.strptime(date, '%Y-%m-%d %H:%M:%S.%f')
        start_date = date_time_obj.replace(hour=0, minute=0, second=0)
        end_date = date_time_obj.replace(hour=23, minute=59, second=59)
        organisation_id = request.user.profile.organisation.organisation_id

        satisfied_count = 0
        dissatisfied_count = 0
        extremely_dissatisfied_count = 0

        client_count = mongo_client.leads.find({"$and": [{"created_at":{"$gte": start_date, "$lte": end_date}}, {"company_id": organisation_id}, {"is_current_company":"yes"}, {"current_patner_feedback": { "$in": ["Dissatisfied", "Extremely Dissatisfied"]}}, {"client_status": "Accepted"}]})

        for row in client_count:
            feedback = row["current_patner_feedback"]
            if feedback=="Satisfied":
                satisfied_count +=1
            elif feedback=="Dissatisfied":
                dissatisfied_count +=1
            elif feedback=="Extremely Dissatisfied":
                extremely_dissatisfied_count +=1

        feedback_count = {
            'satisfied_count' : satisfied_count,
            'dissatisfied_count' : dissatisfied_count,
            'extremely_dissatisfied_count' : extremely_dissatisfied_count,
        }
                
        return ui_utils.handle_response({}, data=feedback_count, success=True)


class GetCampaignList(APIView):

    def get(self, request):
        
        organisation_id = request.user.profile.organisation.organisation_id

        campaign_data = {
            'ongoing_campaigns': [],
            'upcoming_campaigns': [],
            'completed_campaigns': [],
            'onhold_campaigns': []
        }
        current_date = datetime.datetime.now()

        campaign_data['completed_campaigns'] = ProposalInfo.objects.filter(type_of_end_customer__formatted_name="b_to_b_l_d", tentative_end_date__lt=current_date, account__organisation=organisation_id, 
            campaign_state='PTC').values('proposal_id', 'name')

        campaign_data['upcoming_campaigns'] = ProposalInfo.objects.filter(type_of_end_customer__formatted_name="b_to_b_l_d", tentative_start_date__gt=current_date, account__organisation=organisation_id,
            campaign_state='PTC').values('proposal_id', 'name')

        campaign_data['ongoing_campaigns'] = ProposalInfo.objects.filter(tentative_start_date__lte=current_date, tentative_end_date__gte=current_date, account__organisation=organisation_id, 
            campaign_state='PTC', type_of_end_customer__formatted_name="b_to_b_l_d").values('proposal_id', 'name')

        campaign_data['onhold_campaigns'] = ProposalInfo.objects.filter(type_of_end_customer__formatted_name="b_to_b_l_d", account__organisation=organisation_id,
            campaign_state='POH').values('proposal_id', 'name')

        return ui_utils.handle_response({}, data=campaign_data, success=True)

       
class GetSupplierByCampaign(APIView):

    def get(self, request):

        campaign_id = request.query_params.get('campaign_id')

        verified_supplier_ids = Requirement.objects.filter(company_campaign_id=campaign_id,
            varified_bd="yes", client_status="Accepted").values('company_shortlisted_spaces_id__object_id')

        supplier_society_data = SupplierTypeSociety.objects.filter(supplier_id__in=verified_supplier_ids).values('supplier_id').annotate(
            supplier_name = F('society_name'), unit_primary_count=F('flat_count'), city=F('society_city'), area=F('society_locality'))
        
        supplier_master_data = SupplierMaster.objects.filter(supplier_id__in=verified_supplier_ids).values('supplier_id', 'supplier_name', 'unit_primary_count', 'city', 'area')
        supplier_data = list(supplier_society_data) + list(supplier_master_data)

        return ui_utils.handle_response({}, data=supplier_data, success=True)

class FlatSummaryDetails(APIView):

    def get(self, request):

        campaign_id = request.query_params.get('campaign_id')
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')

        if start_date and end_date:
            start_date = datetime.datetime.strptime(start_date, '%Y-%m-%d')
            end_date = datetime.datetime.strptime(end_date, '%Y-%m-%d')

            where = {"company_campaign_id": campaign_id, "created_at":{"$gte": start_date, "$lte": end_date},"lead_purchased":"yes", "client_status": "Accepted"}
        else:
            where = {"company_campaign_id": campaign_id,"lead_purchased":"yes", "client_status": "Accepted"}

        lead_data = list(mongo_client.leads.find(where))
        total_leads = len(lead_data)

        lead_obj = {}
        data = []
        for lead in lead_data:
            lead_obj = dict(lead)
            lead_obj['_id'] = str(lead_obj['_id'])
            data.append(lead_obj)

        return ui_utils.handle_response({}, data=data, success=True)

class SummaryReportAndGraph(APIView):

    def get(self, request):

        final_data = {}
        campaign_id = request.query_params.get('campaign_id')

        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')

        if start_date and end_date:
            start_date = datetime.datetime.strptime(start_date, '%Y-%m-%d')
            end_date = datetime.datetime.strptime(end_date, '%Y-%m-%d')
            date = start_date
        else:
            date = datetime.datetime.now()
 
        diff_date = date - timedelta(days=7)
        final_data['last_week'] = self.get_count_data(campaign_id, start_date, end_date, diff_date)
        diff_date = date - timedelta(days=14)
        final_data['last_two_weeks'] = self.get_count_data(campaign_id, start_date, end_date, diff_date)
        diff_date = date - timedelta(days=21)
        final_data['last_three_weeks'] = self.get_count_data(campaign_id, start_date, end_date, diff_date)

        where = {"company_campaign_id": campaign_id,"lead_purchased":"yes", "client_status": "Accepted"}
        total_lead_data = list(mongo_client.leads.find(where))
        total_lead = len(total_lead_data)

        total_primary_count = 0
        for lead in total_lead_data:
            try:
                primary_count = lead['supplier_primary_count']
            except Exception as e:
                primary_count = 0
            total_primary_count = total_primary_count + primary_count

        where = {"company_campaign_id": campaign_id,"lead_status": "Hot Lead","lead_purchased":"yes", "client_status": "Accepted"}
        total_hot_lead_data = list(mongo_client.leads.find(where))
        total_hot_lead_count = len(total_hot_lead_data)

        company_hot_lead_status = None
        if total_hot_lead_data:
            company_hot_lead_status = total_hot_lead_data[0]['company_lead_status']

        where = {"company_campaign_id": campaign_id,"lead_status": "Deep Lead","lead_purchased":"yes", "client_status": "Accepted"}
        total_deep_lead_data = list(mongo_client.leads.find(where))
        total_deep_lead_count = len(total_deep_lead_data)

        where = {"company_campaign_id": campaign_id,"lead_purchased":"yes", "client_status": "Accepted"}
        total_purchased_lead = mongo_client.leads.find(where).count()

        company_deep_lead_status = None
        if total_deep_lead_data:
            company_deep_lead_status = total_deep_lead_data[-1]['company_lead_status']

        final_data['overall_data'] = {
            "total_lead":total_lead,
            "company_hot_lead_status": company_hot_lead_status,
            "company_deep_lead_status": company_deep_lead_status,
            "total_deep_count":total_deep_lead_count,
            "hot_lead_count":total_hot_lead_count,
            "primary_count":total_primary_count
            }

        return ui_utils.handle_response({}, data=final_data, success=True)

    def get_count_data(self, campaign_id, start_date, end_date, diff_date):

        if start_date and end_date:
            where = {"company_campaign_id": campaign_id,"lead_purchased":"yes", "client_status": "Accepted",
                "$and": [{"created_at":{"$gte": diff_date},
                    "created_at":{"$gte": start_date, "$lte": end_date}}]}

        else:
            where = {"company_campaign_id": campaign_id, "client_status": "Accepted", 
                    "created_at":{"$gte": diff_date},"lead_purchased":"yes"}
        
        total_lead_data = list(mongo_client.leads.find(where))
        total_lead = len(total_lead_data)
        total_primary_count = 0
        for lead in total_lead_data:
            try:
                primary_count = lead['supplier_primary_count']
            except Exception as e:
                primary_count = 0
            total_primary_count = total_primary_count + primary_count

        where.update({"lead_status": "Hot Lead"})
        total_hot_lead = mongo_client.leads.find(where).count()

        where.update({"lead_status": "Deep Lead"})
        total_deep_lead = mongo_client.leads.find(where).count()

        context = {
            "total_lead": total_lead,
            "hot_lead_count": total_hot_lead,
            "total_deep_count": total_deep_lead,
            "primary_count":total_primary_count
        }

        return context

class GetLeadDistributionCampaign(APIView):

    def get(self, request):
        lead_type = request.query_params.get("lead_type")
        organisation_id = request.user.profile.organisation.organisation_id
        if lead_type == "Survey":
            lead = list(mongo_client.leads.find({"$and": [{"company_id": organisation_id}, {"is_current_company":"yes"}, {"current_patner_feedback": { "$in": ["Dissatisfied", "Extremely Dissatisfied"]}}, {"client_status":"Accepted"}]}))
        else:
            lead = list(mongo_client.leads.find({"company_id": organisation_id, "client_status":"Accepted"}))
        
        campaign_list = []
        for row in lead:
            campaign = row["company_campaign_id"]       
            campaign_list.append(campaign)

        # if request.query_params.get('supplier_code') == "mix":
        campaign_list = ProposalInfo.objects.filter(proposal_id__in=campaign_list).values_list('proposal_id', flat=True).distinct()
        # if request.query_params.get('supplier_code') and request.query_params.get('supplier_code') != "mix" and request.query_params.get('supplier_code') != "all":
        #     campaign_list = ShortlistedSpaces.objects.filter(proposal_id__in=campaign_list, supplier_code=request.query_params.get('supplier_code')).values_list('proposal_id', flat=True).distinct()
        campaign_list = [campaign_id for campaign_id in campaign_list]

        all_shortlisted_supplier = ShortlistedSpaces.objects.filter(proposal_id__in=campaign_list).\
            values('proposal_id', 'object_id', 'is_completed', 'proposal__name', 'proposal__tentative_start_date',
                'proposal__tentative_end_date', 'proposal__campaign_state', 'supplier_code')

        all_campaign_dict = {}
        all_shortlisted_supplier_id = [supplier['object_id'] for supplier in all_shortlisted_supplier if supplier['supplier_code'] == 'RS']
        all_supplier_society = SupplierTypeSociety.objects.filter(supplier_id__in=all_shortlisted_supplier_id).values('supplier_id', 'flat_count')

        all_supplier_id = [supplier['object_id'] for supplier in all_shortlisted_supplier if supplier['supplier_code'] != 'RS']
        all_supplier_master = SupplierMaster.objects.filter(supplier_id__in=all_supplier_id).values('supplier_id', 'unit_primary_count')

        all_supplier_society_dict = {}
        current_date = datetime.datetime.now().date()
        for supplier in all_supplier_society:
            all_supplier_society_dict[supplier['supplier_id']] = {'flat_count': supplier['flat_count']}

        for supplier in all_supplier_master:
            all_supplier_society_dict[supplier['supplier_id']] = {'flat_count': supplier['unit_primary_count']}

        for shortlisted_supplier in all_shortlisted_supplier:
            if shortlisted_supplier['proposal_id'] not in all_campaign_dict:
                all_campaign_dict[shortlisted_supplier['proposal_id']] = {
                'all_supplier_ids': [], 'total_flat_counts': 0, 'purchased_survey':0}
            if shortlisted_supplier['object_id'] not in all_campaign_dict[shortlisted_supplier['proposal_id']]['all_supplier_ids']:
                all_campaign_dict[shortlisted_supplier['proposal_id']]['all_supplier_ids'].append(shortlisted_supplier['object_id'])
                if shortlisted_supplier['object_id'] in all_supplier_society_dict and all_supplier_society_dict[shortlisted_supplier['object_id']]['flat_count']:
                    all_campaign_dict[shortlisted_supplier['proposal_id']]['total_flat_counts'] += all_supplier_society_dict[shortlisted_supplier['object_id']]['flat_count']

            all_campaign_dict[shortlisted_supplier['proposal_id']]['name'] = shortlisted_supplier['proposal__name']
            all_campaign_dict[shortlisted_supplier['proposal_id']]['start_date'] = shortlisted_supplier['proposal__tentative_start_date']
            all_campaign_dict[shortlisted_supplier['proposal_id']]['end_date'] = shortlisted_supplier['proposal__tentative_end_date']
            all_campaign_dict[shortlisted_supplier['proposal_id']]['campaign_status'] = shortlisted_supplier['proposal__campaign_state']

        all_purchased_survey = mongo_client.leads.find({"company_campaign_id":campaign_list})
        for row in all_purchased_survey:
            if row["lead_purchased"] == "yes":
                all_campaign_dict[campaign_id]['purchased_survey'] += 1

        all_leads_summary = []
        for campaign_id in all_campaign_dict:
            this_campaign_status = None
            if not all_campaign_dict[campaign_id]['campaign_status'] == proposal_on_hold:
                if all_campaign_dict[campaign_id]['start_date'].date() > current_date:
                    this_campaign_status = campaign_status['upcoming_campaigns']
                elif all_campaign_dict[campaign_id]['end_date'].date() >= current_date:
                    this_campaign_status = campaign_status['ongoing_campaigns']
                elif all_campaign_dict[campaign_id]['end_date'].date() < current_date:
                    this_campaign_status = campaign_status['completed_campaigns']
            else:
                this_campaign_status = "on_hold"
            all_leads_summary.append({
                "campaign_id": campaign_id,
                "name": all_campaign_dict[campaign_id]['name'],
                "start_date": all_campaign_dict[campaign_id]['start_date'],
                "end_date": all_campaign_dict[campaign_id]['end_date'],
                "supplier_count": len(all_campaign_dict[campaign_id]['all_supplier_ids']),
                "flat_count": all_campaign_dict[campaign_id]['total_flat_counts'],
                "purchased_survey": all_campaign_dict[campaign_id]['purchased_survey'],
                "campaign_status": this_campaign_status
            })
        return ui_utils.handle_response({}, data=all_leads_summary, success=True)

class GetPurchasedLeadsData(APIView):

    def get(self, request):

        campaign_id = request.query_params.get("campaign_id")
        leads = mongo_client.leads.find({"company_campaign_id": campaign_id})
        data = []
        for row in leads:
            purchased_leads = {}           
            if row["lead_purchased"] == "yes":
                row1 = row
                if row1.get("_id"):
                    del row1["_id"]
                if row1.get("data"):
                    del row1["data"]

                purchased_leads = row1
                supplier_list = row["supplier_id"]
                supplier_society_data = SupplierTypeSociety.objects.filter(supplier_id=supplier_list).values('supplier_id').annotate(
                    supplier_name = F('society_name'), unit_primary_count=F('flat_count'), city=F('society_city'), area=F('society_locality'), subarea=F('society_subarea'))
                supplier_master_data = SupplierMaster.objects.filter(supplier_id=supplier_list).values(
                    'supplier_id', 'supplier_name', 'unit_primary_count', 'city', 'area', 'subarea')
                leads_data = list(supplier_society_data) + list(supplier_master_data)

                purchased_leads["leads_data"] = leads_data

                purchased_leads["internal_comments"] = CampaignComments.objects.filter(campaign_id=campaign_id, shortlisted_spaces__object_id=supplier_list, related_to='INTERNAL').values("comment")
                purchased_leads["external_comments"] = CampaignComments.objects.filter(campaign_id=campaign_id, shortlisted_spaces__object_id=supplier_list, related_to='EXTERNAL').values("comment")

                contact_data = ContactDetails.objects.filter(object_id=supplier_list).values('mobile', 'name', 'contact_type')
                purchased_leads["contact_data"] = contact_data

                data.append(purchased_leads)

        return ui_utils.handle_response({}, data=data, success=True)

class GetNotPurchasedLeadsData(APIView):

    def get(self, request):

        campaign_id = request.query_params.get("campaign_id")
        leads = mongo_client.leads.find({"company_campaign_id": campaign_id})
        data = []
        for row in leads:
            row1 = row
            if row1.get("_id"):
                del row1["_id"]
            if row1.get("data"):
                del row1["data"]

            not_purchased_leads = row1
            if row["lead_purchased"] == "no":
                supplier_list = []
                supplier_list = row["supplier_id"]
                supplier_society_data = SupplierTypeSociety.objects.filter(supplier_id=supplier_list).values('supplier_id').annotate(
                    supplier_name = F('society_name'), unit_primary_count=F('flat_count'), city=F('society_city'), area=F('society_locality'))
        
                supplier_master_data = SupplierMaster.objects.filter(supplier_id=supplier_list).values(
                    'supplier_id', 'supplier_name', 'unit_primary_count', 'city', 'area')
                leads_data = list(supplier_society_data) + list(supplier_master_data)

                not_purchased_leads["leads_data"] = leads_data
                not_purchased_leads["price"] = 0
                if row.get("lead_price"):
                    not_purchased_leads["price"] = row["lead_price"]

                not_purchased_leads["internal_comments"] = CampaignComments.objects.filter(campaign_id=campaign_id, shortlisted_spaces__object_id=supplier_list, related_to='INTERNAL').values("comment")
                not_purchased_leads["external_comments"] = CampaignComments.objects.filter(campaign_id=campaign_id, shortlisted_spaces__object_id=supplier_list, related_to='EXTERNAL').values("comment")

                data.append(not_purchased_leads)

        return ui_utils.handle_response({}, data=data, success=True)


class BuyLead(APIView):

    def post(self, request):
        requirement_ids = request.data.get('requirement_ids')

        for req in requirement_ids:

            requirement = Requirement.objects.filter(id=req).first()
            requirement.lead_purchased = "yes"
            requirement.purchased_date = datetime.datetime.now()
            requirement.save()

            if requirement:
                
                mongo_client.leads.update_one({"requrement_id": 
                    requirement.id,},{"$set": {
                            "lead_purchased": "yes",
                            "purchased_date": requirement.purchased_date
                        }})

                company_lead = mongo_client.OrganizationLeads.find_one({"company_id":requirement.company_id})
                if company_lead:
                    mongo_client.OrganizationLeads.update_one({"company_id":requirement.company_id},
                            {"$set": {
                                "purchased_count": company_lead['purchased_count'] + 1,
                                "updated_at": datetime.datetime.now()
                            }})
                else:
                    company_leads_dict = {
                        "updated_at": datetime.datetime.now(),
                        "created_at": datetime.datetime.now(),
                        "purchased_count": 1,
                        "company_id": requirement.company_id
                    }
                    mongo_client.OrganizationLeads.insert_one(company_leads_dict)

                campaign_lead = mongo_client.CampaignLeads.find_one({
                    "company_campaign_id":requirement.company_campaign_id})

                if campaign_lead:
                    mongo_client.CampaignLeads.update_one({"company_campaign_id":requirement.company_campaign_id},
                            {"$set": {
                                "purchased_count": campaign_lead['purchased_count'] + 1,
                                "updated_at": datetime.datetime.now()
                            }})
                else:
                    campaign_leads_dict = {
                        "updated_at": datetime.datetime.now(),
                        "created_at": datetime.datetime.now(),
                        "purchased_count": 1,
                        "company_campaign_id": requirement.company_campaign_id
                    }
                    mongo_client.CampaignLeads.insert_one(campaign_leads_dict)

        return ui_utils.handle_response({}, data="Lead buy successfully", success=True)

class GetDynamicLeadFormHeaders(APIView):

    def get(self, request):

        campaign_id = request.query_params.get("campaign_id")
        supplier_type = request.query_params.get("supplier_type")
        lead_form = mongo_client.leads_forms.find({"campaign_id": campaign_id})
        lead_type = request.query_params.get("lead_type")

        data = {}
        context = {}
        lead_form_key = None

        for key, lead_form_keys in lead_form[0]["data"].items():
            lead_form_key = lead_form_keys["item_id"]
            header_keys = lead_form_keys["key_name"]        
        
            lead_form_key_2 = None
            if lead_form_key:
                for key, value in lead_form[0]["global_hot_lead_criteria"].items():
                    if value.get("or"):
                        for key1, value1 in value["or"].items():
                            if str(key1) == str(lead_form_key):
                                lead_form_key_2 = key
                if lead_form_key_2 and lead_form[0]["hotness_mapping"].get(lead_form_key_2):
                    header_values = lead_form[0]["hotness_mapping"].get(lead_form_key_2)
                else:
                    header_values = header_keys
            
            context[header_keys] = header_values

            if supplier_type == "all":
                
                if lead_type == "Survey":
                    leads = list(mongo_client.leads.find({"$and": [{"company_campaign_id": campaign_id}, {"is_current_company":"yes"}, {"current_patner_feedback": { "$in": ["Dissatisfied", "Extremely Dissatisfied"]}}, {"client_status":"Accepted"}]}))
                else:
                    leads = list(mongo_client.leads.find({"company_campaign_id": campaign_id, "client_status":"Accepted"}))
            else:

                if lead_type == "Survey":
                    leads = list(mongo_client.leads.find({"$and": [{"company_campaign_id": campaign_id}, {"is_current_company":"yes"}, {"current_patner_feedback": { "$in": ["Dissatisfied", "Extremely Dissatisfied"]}}, {"client_status":"Accepted"}, {"supplier_type":supplier_type}]}))
                else:
                    leads = list(mongo_client.leads.find({"company_campaign_id": campaign_id, "client_status":"Accepted", "supplier_type":supplier_type}))

            values = []
            for entry in leads:
                lead = entry['data']
                lead.insert(0, {'lead_purchased' : entry['lead_purchased']})
                values.append(lead)

            data['header'] = context
            data['values'] = values

        return ui_utils.handle_response({}, data=data, success=True)


class SuspenseLeadCount(APIView):
    permission_classes = ()
    authentication_classes = ()

    def get(self, request):

        start_date = make_aware(datetime.datetime.strptime(request.GET.get("start_date"), '%Y-%m-%d'))
        end_date = make_aware(datetime.datetime.strptime(request.GET.get("end_date"), '%Y-%m-%d')) + datetime.timedelta(days=1)

        count_suspanse_lead = mongo_client.suspense_lead.find({'created_at': 
            {"$gte": start_date, "$lte": end_date}}).count()

        return ui_utils.handle_response({}, data={"count":count_suspanse_lead}, success=True)


class LicenceDetails(APIView):
    # permission_classes = ()
    # authentication_classes = ()

    def get(self, request):

        current_user = request.user
        mydetail = None
        companydetail = None
        managerdetail = None
        rltionshipmanager = None
        base_user = BaseUser.objects.filter(username=current_user).first()
        mydetail = BaseUserSerializer(base_user, many=False).data

        if base_user and base_user.profile:

            licensedetail = LicenseDetails.objects.filter(company=
                base_user.profile.organisation_id).first()
            companydetail = LicenseDetailsSerializer(licensedetail, many=False).data

            managerdetail = MachadaloRelationshipManager.objects.filter(company=
                base_user.profile.organisation_id).first()
            rltionshipmanager = RelationshipManagerSerializer(managerdetail, many=False).data

        return ui_utils.handle_response({}, data={"mydetail":mydetail,
            "companydetail":companydetail,"relationship_manager":rltionshipmanager}, success=True)

    def put(self,request):

        licensedetail = LicenseDetails.objects.filter(company=
                request.data['company']).first()
        if licensedetail:

            license = LicenseDetailsSerializer(licensedetail, data=request.data,
                partial=True)

            if license.is_valid():
                license.save()
                return ui_utils.handle_response({}, data={"data":license.data}, success=True)
            else:
                return ui_utils.handle_response({}, data={"errors":license.errors}, success=False)
        else:
            return ui_utils.handle_response({}, data={"data":"data not found"}, success=True)


class PaymentDetailsView(APIView):
    # permission_classes = ()
    # authentication_classes = ()

    def get(self, request):

        paymentdetail = None
        current_user = request.user
        base_user = BaseUser.objects.filter(username=current_user).first()

        if base_user and base_user.profile:

            paymnt = PaymentDetails.objects.filter(company=
                base_user.profile.organisation_id).first()

            paymentdetail = PaymentDetailsSerializer(paymnt, many=True).data

        return ui_utils.handle_response({}, data={"paymentdetail":paymentdetail}, success=True)


class LeadsDecisionPanding(APIView):
    # permission_classes = ()
    # authentication_classes = ()

    def get(self, request):

        data = []
        context = {}
        type_of_entity = request.query_params.get("type_of_entity")
        organisation_id = request.user.profile.organisation.organisation_id

        if organisation_id:

            if type_of_entity == "all":
                where = {"$and": [{"company_id": organisation_id}, {"client_status":"Decision Pending"}]}
            else:
                where = {"$and": [{"company_id": organisation_id}, {"client_status":"Decision Pending"},{"supplier_type":type_of_entity}]}

            leads = list(mongo_client.leads.find(where)) 
            
            for entry in leads:
            
                context['_id'] = str(ObjectId(entry['_id']))
                context['requirement_id'] = entry['requrement_id']
                context['entity_name'] = entry['data'][0]['value']
                context['entity_type'] = entry['data'][1]['value']
                context['primary_count'] = entry['data'][9]['value']
                context['area'] = entry['data'][2]['value']
                context['city'] = entry['data'][5]['value']
                context['client_status'] = entry['client_status']
                data.append(context.copy())

        return ui_utils.handle_response({}, data={"lead":data}, success=True)

class UpdateClientStatus(APIView):

    def post(self, request):

        data = request.data.get("data")

        for clnt_status in data:

            mongo_client.leads.update_one({"requrement_id": 
                int(clnt_status['requirement_id'])},{"$set": {
                        "client_status": clnt_status['client_status'],
                    }})

            req = Requirement.objects.filter(id=clnt_status["requirement_id"]).first()
            req.client_status = clnt_status["client_status"]
            req.save()
        
        return ui_utils.handle_response({}, data="Record updated successfully", success=True)

class DownloadB2BLeads(APIView):

    def get(self, request):

        campaign_id = request.query_params.get("campaign_id")

        requirement = PreRequirement.objects.filter(campaign_id=campaign_id)

        browsed_leads = list(BrowsedLead.objects.raw({"campaign_id":str(campaign_id),
         "status":"closed"}).values())

        excel_book = b2b_utils.download_b2b_leads(requirement,browsed_leads)

        campaign = ProposalInfo.objects.filter(proposal_id=campaign_id).first()

        name = str(campaign.name) if campaign else "mydata"
        filename = str(name.replace(" ", ""))

        resp = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = 'attachment; filename='+filename+'.xlsx'
        excel_book.save(resp)

        return resp


class AddNotificationTemplate(APIView):

    def post(self, request):

        serializer = NotificationTemplateSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            res = serializer.data
            status = True
        else:
            res = serializer.errors
            status = False
        
        return ui_utils.handle_response({}, data=res, success=status)

class GetAllSuspenseLead(APIView):

    def get(self, request):

        list1 = []
        suspense_lead = mongo_client.suspense_lead.find({"status":"closed"}).sort("created_at",-1)
        companies = Organisation.objects.all()
        org_dict_id = {str((row3.name).lower()):row3.organisation_id for row3 in companies}

        for row in suspense_lead:
            row1 = dict(row)
            row1["_id"] = str(row1["_id"])
            prefered_patners_list = []

            row1['current_patner_other'] = row1.get("current_patner_other")
            row1['status'] = row1.get("status")
            row1['prefered_patner_other'] = row1.get("prefered_patner_other")
            row1['poc_name'] = row1.get("poc_name")
            row1['designation'] = row1.get("designation")
            row1['pin_code'] = row1.get("pin_code")
            row1['address'] = row1.get("address")
            row1['supplier_type'] = row1.get("supplier_type")
            row1['area_id'] = None
            row1['city_id'] = None

            if row1.get("area"):

                area_detail = CityArea.objects.filter(label=
                    str(row1.get("area"))).first()
                if area_detail:
                    row1['area_id'] = area_detail.id

            if row1.get("city"):

                city_detail = City.objects.filter(city_name=
                    str(row1.get("city"))).first()
                if city_detail:
                    row1['city_id'] = city_detail.id

            if row1.get("prefered_patners"):
                for prf_prnr in row1.get("prefered_patners"):
                    if prf_prnr == "other":
                        prefered_patners_list.append("other")
                    else:
                        if prf_prnr:
                            prefered_patners_list.append(org_dict_id.get(prf_prnr.lower()))
            else:
                prefered_patners_list.append("other")

            row1['prefered_patners'] = prefered_patners_list


            if row1.get("current_patner"):
                row1["current_patner"] = org_dict_id.get(row1.get("current_patner").lower())
            
            list1.append(row1)

        companies_data = OrganisationSerializer(companies, many=True).data
        return ui_utils.handle_response({}, data={"suspense_lead": list1,"companies":companies_data}, success=True)


class UpdateSuspenseLead(APIView):

    def post(self, request):

        suspense_leads = request.data.get("suspense_leads")
        companies = Organisation.objects.all()
        org_dict_id = {str((row3.organisation_id).lower()):row3.name for row3 in companies}

        for suspense in suspense_leads:
            prefered_patners = []

            if "other" in suspense.get("prefered_patners_id") and \
                len(suspense["prefered_patners_id"]) == 1:
                prefered_patners.append("other")

            else:
                for prtnrs in suspense["prefered_patners_id"]:
                    prtnrs = prtnrs.lower()
                    if not prtnrs == "other":
                        company = org_dict_id.get(prtnrs)
                        prefered_patners.append(company)

            if "other" in suspense["prefered_patners_id"] and \
                len(suspense["prefered_patners_id"]) > 1:
                prefered_patners.append("other")

            current_patner = None
            if suspense.get("current_patner_id"):
                current_patner = suspense.get("current_patner_id").lower()
                if current_patner:
                    current_patner = org_dict_id.get(current_patner)

            
            update_values = {"$set":{
                "implementation_timeline":suspense.get("implementation_timeline"),
                "meating_timeline":suspense.get("meating_timeline"),
                "comment":suspense.get("comment"),
                "current_patner":current_patner,
                "current_patner_other":suspense.get("current_patner_other"),
                "prefered_patners":prefered_patners,
                "prefered_patner_other":suspense.get("prefered_patner_other"),
                "call_back_preference":suspense.get("call_back_preference"),
                "status":"closed",
                "updated_at":datetime.datetime.now(),
                }}

            mongo_client.suspense_lead.update({"_id": ObjectId(suspense["_id"])},update_values)

        return ui_utils.handle_response({}, data={"message":"Suspense lead updated successfully"}, success=True)


class SuspenseLeadDelete(APIView):

    def post(self, request):
        suspense_ids = request.data.get("suspense_ids")

        for suspense_id in suspense_ids:
            mongo_client.suspense_lead.update({"_id": ObjectId(suspense_id)}, {"$set":{"status":"deleted"}})

        return ui_utils.handle_response({}, data={"message":"Suspense lead removed successfully"}, success=True)


class AddSuspenseToSupplier(APIView):

    def get(self, request):

        city = request.query_params.get("city")
        area = request.query_params.get("area")
        supplier_type = request.query_params.get("supplier_type")
        suspense_id = request.query_params.get("suspense_id")
        supplier = {}

        supplier_objects = SupplierMaster.objects
        society_objects = SupplierTypeSociety.objects

        if supplier_type == 'RS':
            supplier_list = society_objects.filter(society_city=city, society_locality=area).values('society_name', 'supplier_id')
        else:
            supplier_list = supplier_objects.filter(city=city, area=area).values('supplier_name', 'supplier_id')

        if suspense_id:
            suspense_lead = mongo_client.suspense_lead.find_one({"_id": ObjectId(suspense_id)})
            supplier_id = suspense_lead['supplier_id']

            if suspense_lead['supplier_type'] == "RS":
                supplier_data = society_objects.filter(supplier_id=supplier_id).values('supplier_id').annotate(
                    supplier_name = F('society_name'), city=F('society_city'), area=F('society_locality'))
            else:
                supplier_data = supplier_objects.filter(supplier_id=supplier_id).values('supplier_name','city', 'area')

            supplier["supplier"] = supplier_data
            contact_detail = ContactDetails.objects.filter(object_id=supplier_id).values("name", "mobile", "contact_type")
            supplier["contact_detail"] = contact_detail

        return ui_utils.handle_response({}, data={"supplier_list": supplier_list, "supplier_data": supplier}, success=True)


    def post(self, request):
        
        suspense_id = request.data.get("suspense_id")
        supplier_type = request.data.get("supplier_type")
        phone_number = request.data.get("phone_number")
        supplier_name = request.data.get("supplier_name")
        poc_name = request.data.get("poc_name")
        designation = request.data.get("designation")
        city_id = request.data.get("city_id")
        city = request.data.get("city")
        area_id = request.data.get("area_id")
        area = request.data.get("area")
        supplier_id = request.data.get("supplier_id")

        #If area does not exist, add new area
        if city_id:
            area_detail = CityArea.objects.filter(id=area_id, city_code=city_id).values('id')
            if area_detail and len(area_detail) > 0:
                area_id = area_detail[0]['id']
            else:
                area_code = ui_utils.getRandomString()
                areaInserted = CityArea.objects.create(label=area.title(), area_code=area_code.upper(), city_code_id=city_id)
                if areaInserted:
                    area_id = areaInserted.id
                else:
                    print('Failed to add Area :', area)

        if supplier_type == "RS":
            supplier = SupplierTypeSociety.objects.filter(supplier_id=supplier_id)
        else:
            supplier = SupplierMaster.objects.filter(supplier_id=supplier_id)
        
        if not supplier:
            suspense_data = {
                        'city_id': city_id,
                        'area_id': area_id,
                        'subarea_id': 1,
                        'supplier_type': supplier_type,
                        'supplier_code': supplier_type,
                        'supplier_name': supplier_name,    
                    }
            supplier_id = ui_utils.get_supplier_id(suspense_data)

            #updating supplier data in suspense lead table
            update_values = {"$set":{
                "supplier_id":supplier_id,
                "supplier_type":supplier_type,
                "supplier_name":supplier_name,
                "poc_name":poc_name,
                "designation":designation,
                "city":city,
                "area":area,
                "is_updated":"True"
                }}
            mongo_client.suspense_lead.update({"_id": ObjectId(suspense_id)},update_values)

            if supplier_type != "RS":
                supplier_data = {
                'supplier_id': supplier_id,
                'supplier_name': supplier_name,
                'supplier_type': supplier_type,
                'area': area.title(),
                'city': city,
                }
                serializer = SupplierMasterSerializer(data=supplier_data)
                if serializer.is_valid():
                    serializer.save()

                AddressMaster(**{
                    'supplier_id': supplier_id,
                    'area': area.title(),
                    'city': city,
                }).save()
            else:
                supplier_data = {
                'supplier_id': supplier_id,
                'society_name': supplier_name,
                'supplier_code': supplier_type,
                'society_locality': area.title(),
                'society_city': city,
                }
                serializer = SupplierTypeSocietySerializer(data=supplier_data)
                if serializer.is_valid():
                    serializer.save()

        else:
            update_values = {"$set":{
                "supplier_id":supplier_id,
                "supplier_type":supplier_type,
                "supplier_name":supplier_name,
                "poc_name":poc_name,
                "designation":designation,
                "city":city,
                "area":area,
                "is_updated":"True"
                }}
            mongo_client.suspense_lead.update({"_id": ObjectId(suspense_id)},update_values)
            if supplier_type == "RS":
                supplier.update(society_name=supplier_name)
            else:
                supplier.update(supplier_name=supplier_name)

        if phone_number and supplier_id:
            contact_details = ContactDetails.objects.filter(mobile=phone_number, object_id=supplier_id)
            if contact_details:
                if poc_name and designation:
                    contact_details.update(name=poc_name, contact_type=designation)                
            else:
                ContactDetails(**{
                    'object_id': supplier_id,
                    'mobile' : phone_number, 
                    'name' : poc_name, 
                    'contact_type' : designation,
                }).save()

        return ui_utils.handle_response({}, data={"message":"Supplier added successfully"}, success=True)

class AddPocDetails(APIView):

    def get(self, request):

        suspense_id = request.query_params.get("suspense_id")
        suspense_lead = mongo_client.suspense_lead.find_one({"_id": ObjectId(suspense_id)})
        object_id = suspense_lead['supplier_id']

        contact_detail = ContactDetails.objects.filter(object_id=object_id).values("name", "mobile", "contact_type")

        return ui_utils.handle_response({}, data={"contact_detail": contact_detail}, success=True)

    def post(self, request):
        
        suspense_id = request.data.get("suspense_id")
        contactData = request.data.get("contactData")
        suspense_lead = mongo_client.suspense_lead.find_one({"_id": ObjectId(suspense_id)})
        object_id = suspense_lead['supplier_id']
        not_remove_contacts = []
        if contactData:
            for contact in contactData:
                contact['object_id'] = object_id
                if 'id' in contact:
                    contact_detail = ContactDetails.objects.filter(pk=contact['id']).first()
                    contact_serializer = ContactDetailsSerializer(contact_detail, data=contact)
                else:
                    contact_serializer = ContactDetailsSerializer(data=contact)
                if contact_serializer.is_valid():
                    contact_serializer.save(object_id=object_id)
                    not_remove_contacts.append(contact_serializer.data['id'])

        ContactDetails.objects.filter(object_id=object_id).exclude(pk__in=not_remove_contacts).delete()

        return ui_utils.handle_response({}, data={"message":"POC added successfully"}, success=True)


class UpdateMongoDbNotExistKey(APIView):

    def get(self, request):
        key_name = request.query_params.get("key_name")
        value = request.query_params.get("value")

        result = mongo_client.suspense_lead.update({str(key_name): {"$exists" : False}}, {"$set": {str(key_name): value}})

        return ui_utils.handle_response({}, data={"result":result}, success=True)


class SuspenseLeadOpsVerification(APIView):

    def get(self, request):

        suspense_id = request.query_params.get("_id")
        suspense = dict(mongo_client.suspense_lead.find_one({"_id": ObjectId(suspense_id)}))

        campaign = ProposalInfo.objects.filter(Q(type_of_end_customer__formatted_name="b_to_b_r_g") 
            and Q(name=str(suspense.get('city')) or Q(name=str(suspense.get('area'))))).first()

        shortlisted_spaces = ShortlistedSpaces.objects.filter(
            proposal_id=campaign.proposal_id, 
            object_id=suspense.get('supplier_id')).first()

        if not shortlisted_spaces:

            content_type = ui_utils.get_content_type(suspense.get('supplier_type'))
            center = ProposalCenterMapping.objects.filter(
                proposal_id=campaign.proposal_id).first()

            shortlisted_spaces = ShortlistedSpaces(
                proposal_id=campaign.proposal_id,
                center=center,
                supplier_code=suspense.get('supplier_type'),
                object_id=suspense.get('supplier_id'),
                content_type=content_type.data['data'],
                status='F',
                user=request.user,
                requirement_given='yes',
                requirement_given_date=datetime.datetime.now()
            )
            shortlisted_spaces.save()

        if shortlisted_spaces:
            shortlisted_spaces.color_code = 1
            shortlisted_spaces.save()

        if suspense.get('sector_name'):
            sector = BusinessTypes.objects.filter(
                business_type=suspense.get('sector_name').lower()).first()

        if suspense.get('sub_sector_name'):
            sub_sector = BusinessSubTypes.objects.filter(
                business_sub_type=suspense.get('sub_sector_name').lower()).first()

        change_current_patner = "no"
        if suspense.get('current_patner_feedback') == "Dissatisfied" or \
            suspense.get('current_patner_feedback') == "Extremely Dissatisfied":
            change_current_patner = "yes"

        prefered_patners_list = []
        if suspense.get("prefered_patners"):
            prefered_patners_list = Organisation.objects.filter(
                organisation_id__in=suspense.get("prefered_patners")).all()

        current_patner_obj = None
        if suspense.get('current_patner'):
            current_patner_obj = Organisation.objects.filter(
                name=suspense.get('current_patner')).first()

        lead_status = b2b_utils.get_lead_status(impl_timeline = suspense.get('implementation_timeline'),
            meating_timeline = suspense.get('meating_timeline'),company=None,prefered_patners=prefered_patners_list,
            change_current_patner=change_current_patner.lower())

        pre_requirement = PreRequirement(
            campaign_id=campaign.proposal_id,
            shortlisted_spaces=shortlisted_spaces,
            current_company = current_patner_obj,
            current_company_other = suspense.get('current_patner_other'),
            current_patner_feedback = suspense.get('current_patner_feedback'),
            current_patner_feedback_reason = suspense.get('current_patner_feedback_reason'),
            preferred_company_other = suspense.get('prefered_patner_other'),
            sector = sector,
            sub_sector = sub_sector,
            lead_by = None,
            impl_timeline = suspense.get('implementation_timeline'),
            meating_timeline = suspense.get('meating_timeline'),
            lead_status = lead_status,
            comment = suspense.get('comment'),
            varified_ops = 'yes',
            varified_bd = 'no',
            lead_date = datetime.datetime.now(),
            l1_answers = suspense.get('l1_answers'),
            l1_answer_2 = suspense.get('l1_answer_2'),
            l2_answers = suspense.get('l2_answers'),
            l2_answer_2 = suspense.get('l2_answer_2'),
            change_current_patner = change_current_patner.lower(),
            call_back_preference = suspense.get('call_back_preference').lower(),
            varified_ops_by_id = request.user.id,
        )
        pre_requirement.save()

        if prefered_patners_list:
            pre_requirement.preferred_company.set(prefered_patners_list)

        if pre_requirement:
            companies = Organisation.objects.filter(business_type=
                pre_requirement.sector)

            verified = 0
            if companies:
                for company in companies:

                    company_campaign = ProposalInfo.objects.filter(
                        type_of_end_customer__formatted_name="b_to_b_l_d",
                        account__organisation=company).first()

                    if company_campaign:
                        company_shortlisted_spaces = ShortlistedSpaces.objects.filter(
                            object_id=pre_requirement.shortlisted_spaces.object_id,
                            proposal=company_campaign.proposal_id).first()

                    if not company_shortlisted_spaces:
                        center_ = ProposalCenterMapping.objects.filter(proposal=company_campaign).first()
                        content_type_ = ui_utils.get_content_type(pre_requirement.shortlisted_spaces.supplier_code)

                        company_shortlisted_spaces = ShortlistedSpaces(
                            proposal=company_campaign,
                            center=center_,
                            object_id=pre_requirement.shortlisted_spaces.object_id,
                            supplier_code=pre_requirement.shortlisted_spaces.supplier_code,
                            content_type=content_type_.data['data'],
                            status='F',
                            user=request.user,
                            requirement_given='yes',
                            requirement_given_date=datetime.datetime.now(),
                            color_code = 1
                        )
                        company_shortlisted_spaces.save()

                    if company_shortlisted_spaces:
                        color_code = b2b_utils.get_color_code(company_shortlisted_spaces.id)
                        company_shortlisted_spaces.color_code = color_code
                        company_shortlisted_spaces.save()

                    o_preferred_partners_list = pre_requirement.preferred_company.all()

                    is_preferred_company = "no"
                    if company_campaign:
                        if company_campaign.account.organisation in o_preferred_partners_list:
                            is_preferred_company = "yes"

                    new_requirement = Requirement(
                        campaign_id=pre_requirement.campaign_id,
                        shortlisted_spaces=pre_requirement.shortlisted_spaces,
                        company = company,
                        current_company = pre_requirement.current_company,
                        current_company_other = pre_requirement.current_company_other,
                        is_current_patner = "yes" if pre_requirement.current_company == company else "no",
                        current_patner_feedback = pre_requirement.current_patner_feedback,
                        current_patner_feedback_reason = pre_requirement.current_patner_feedback_reason,
                        preferred_company_other = pre_requirement.preferred_company_other,
                        sector = pre_requirement.sector,
                        sub_sector = pre_requirement.sub_sector,
                        lead_by = pre_requirement.lead_by,
                        impl_timeline = pre_requirement.impl_timeline,
                        meating_timeline = pre_requirement.meating_timeline,
                        lead_status = pre_requirement.lead_status,
                        comment = pre_requirement.comment,
                        varified_ops = 'yes',
                        varified_bd = 'no',
                        lead_date = pre_requirement.lead_date,
                        lead_price  = pre_requirement.lead_price,
                        l1_answers = pre_requirement.l1_answers,
                        l1_answer_2 = pre_requirement.l1_answer_2,
                        l2_answers = pre_requirement.l2_answers,
                        l2_answer_2 = pre_requirement.l2_answer_2,
                        change_current_patner = pre_requirement.change_current_patner.lower(),
                        company_campaign=company_campaign,
                        company_shortlisted_spaces=company_shortlisted_spaces,
                        varified_ops_by = request.user,
                        varified_ops_date = datetime.datetime.now(),
                        call_back_preference = pre_requirement.call_back_preference,
                        is_preferred_company = is_preferred_company
                    )
                    new_requirement.save()
                    verified += 1

                    if o_preferred_partners_list:
                        new_requirement.preferred_company.set(o_preferred_partners_list)
            
            return ui_utils.handle_response({}, data=
                {"error":"No companies for the service found"}, success=False)

        update_suspense = mongo_client.suspense_lead.update(
                    {"_id":ObjectId(suspense_id)},{"status":"converted"})

        if verified == 0:
            return ui_utils.handle_response({}, data={"error":"Ops verify failed as there are 0 client campaigns","verified_ops_by":request.user.first_name + request.user.last_name}, success=False)
        else:
            return ui_utils.handle_response({}, data={"message":"Ops Verified and distributed to "+str(verified)+" campaigns","verified_ops_by":request.user.first_name + request.user.last_name}, success=True)
