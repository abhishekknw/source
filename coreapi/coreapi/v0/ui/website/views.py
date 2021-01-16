from __future__ import print_function
from __future__ import absolute_import
import csv
import json
import os
import requests
from openpyxl import load_workbook
from celery.result import GroupResult, AsyncResult
from django.conf import settings
from django.contrib.contenttypes.models import ContentType
from django.core.exceptions import ObjectDoesNotExist
from django.urls import reverse
from django.db import transaction
from django.db.models import Q, F, Sum, Max
from django.forms.models import model_to_dict
from django.utils.dateparse import parse_datetime
from rest_framework import status
from rest_framework.decorators import list_route
from rest_framework import viewsets
from rest_framework.response import Response
from rest_framework.views import APIView
from . import tasks
from v0.ui.components.models import SocietyTower, FlatType, Amenity
from v0.ui.components.serializers import AmenitySerializer
from v0.ui.account.serializers import (BusinessTypesSerializer, BusinessSubTypesSerializer, ProfileSimpleSerializer,
                                       UIBusinessInfoSerializer, UIAccountInfoSerializer, ProfileNestedSerializer)
from v0.ui.campaign.models import Campaign, CampaignSocietyMapping, CampaignAssignment, GenericExportFileName
from v0.ui.campaign.serializers import CampaignSerializer, CampaignSocietyMappingSerializer, CampaignListSerializer, \
    CampaignAssignmentSerializerReadOnly
from v0.ui.account.serializers import (BusinessInfoSerializer, BusinessAccountContactSerializer, AccountInfoSerializer,
                                       ContactDetailsSerializer, AccountSerializer)
from v0.ui.account.models import BusinessAccountContact, ContactDetails, AccountInfo, Profile
from v0.ui.inventory.models import (AdInventoryType, SocietyInventoryBooking,InventoryActivityAssignment,
                                    InventoryActivityImage, InventorySummary,InventoryTypeVersion, InventoryType,
                                    InventoryActivity, INVENTORY_ACTIVITY_TYPES)
from v0.ui.inventory.serializers import SocietyInventoryBookingSerializer, InventoryTypeSerializer, \
    InventoryActivityAssignmentSerializerReadOnly, InventoryActivityAssignmentSerializer, \
    InventoryTypeVersionSerializer, CampaignInventorySerializer

from v0.ui.organisation.models import Organisation, OrganisationMap, ORGANIZATION_CATEGORY
from v0.ui.organisation.serializers import OrganisationMapNestedSerializer, OrganisationSerializer
from v0.ui.location.models import State, City, CityArea, CitySubArea

from v0.ui.common.models import BaseUser
from v0.ui.user.serializers import UserProfileSerializer, UserSerializer, BaseUserSerializer, BaseUserUpdateSerializer, \
    GenericExportFileSerializerReadOnly
from v0.ui.proposal.models import ProposalInfo, ProposalCenterMapping,ProposalCenterMappingVersion, \
    SpaceMappingVersion, SpaceMapping, ShortlistedSpacesVersion, ShortlistedSpaces
from v0.ui.proposal.serializers import (ProposalInfoSerializer, ProposalCenterMappingSerializer,
    ProposalCenterMappingVersionSerializer, ProposalInfoVersionSerializer, SpaceMappingSerializer,
    ProposalCenterMappingSpaceSerializer, ProposalMasterCostSerializer, ProposalMetricsSerializer,
    ProposalCenterMappingVersionSpaceSerializer, SpaceMappingVersionSerializer, ProposalSocietySerializer,
                                        ProposalCorporateSerializer, HashtagImagesSerializer)
from v0.ui.supplier.models import SupplierAmenitiesMap, SupplierTypeCorporate, SupplierTypeSociety, SupplierTypeBusShelter, SupplierMaster
from v0.ui.supplier.serializers import (SupplierAmenitiesMapSerializer, SupplierTypeCorporateSerializer,
                                        SupplierTypeSocietySerializer, SupplierTypeBusShelterSerializer, SupplierMasterSerializer)
from v0.ui.finances.models import ShortlistedInventoryPricingDetails, PriceMappingDefault, getPriceDict
from v0.ui.permissions.models import ObjectLevelPermission, GeneralUserPermission, Role, RoleHierarchy
from v0.ui.base.serializers import ContentTypeSerializer

from . import utils as website_utils
import v0.ui.utils as ui_utils
from coreapi.settings import BASE_URL, BASE_DIR
from v0 import errors
import v0.constants as v0_constants
from v0.ui.campaign.models import CampaignComments

from django.db.models.functions import Trim


# codes for supplier Types  Society -> RS   Corporate -> CP  Gym -> GY   salon -> SA


class ShortlistSocietyAPIView(APIView):

    # discuss use of this get function no id in url
    def get(self, request, id, format=None):
        try:
            campaign = Campaign.objects.get(pk=id)
            items = campaign.societies.all().filter(booking_status='Requested')
            serializer = FinalizeInventorySerializer(items, many=True)
            return Response(serializer.data, status=200)
        except :
            return Response(status=404)

    def post(self, request, format=None):

        if 'campaign_id' in request.data:
            try:
                campaign = Campaign.objects.get(pk=request.data['campaign_id'])
            except Campaign.DoesNotExist:
                return Response(status=404)
        else:
            return Response(status=400)

        if 'society_id' in request.data:
            try:
                society_id = request.data['society_id']
                society = SupplierTypeSociety.objects.get(pk=request.data['society_id'])
            except SupplierTypeSociety.DoesNotExist:
                return Response(status=404)
        else:
            return Response(status=400)

        try:
            campaign_society = CampaignSocietyMapping.objects.get(campaign=campaign, society=society)
            total_societies = CampaignSocietyMapping.objects.filter(campaign=campaign).count()
            error = {"message" : "Already Shortlisted", 'count':total_societies}
            return Response(error, status=200)
        except CampaignSocietyMapping.DoesNotExist:
            campaign_society = CampaignSocietyMapping(campaign=campaign, society=society, booking_status='Shortlisted')
            campaign_society.save()
        # except CampaignSocietyMapping.MultipleObjectsReturned:
        #     total_societies = CampaignSocietyMapping.objects.filter(campaign=campaign).count()
        #     error = {"message" : "Already Shortlisted", 'count':total_societies}
        #     return Response(error, status=200)

        for key in campaign.get_types():
            inventory = SocietyInventoryBooking(campaign=campaign, society=society, adinventory_type=key)
            inventory.save()

        total_societies = CampaignSocietyMapping.objects.filter(campaign=campaign).count()
        return Response({"message": "Society Shortlisted", "id": society_id, 'count':total_societies}, status=200)


# Beta API below this point. Be Careful. Danger Below

def return_price(adinventory_type_dict, duration_type_dict, inv_type, dur_type):
    price_mapping = PriceMappingDefault.objects.filter(adinventory_type=adinventory_type_dict[inv_type], duration_type=duration_type_dict[dur_type])
    if price_mapping:
        return price_mapping[0].business_price
    return 0


class SpacesOnCenterAPIView(APIView):
    def get(self,request,proposal_id=None, format=None):
        ''' This function filters all the spaces(Societies, Corporates etc.) based on the center and
        radius provided currently considering radius
        This API is called before map view page is loaded
        center -- center id for which to filter societies and corporates.

        '''

        ''' !IMPORTANT --> you have to manually add all the type of spaces that are being added apart from
        Corporate and Society '''
        response = {}
        center_id = request.query_params.get('center',None)
        try:
            proposal = ProposalInfo.objects.get(proposal_id=proposal_id)
            response['business_name'] = proposal.account.business.name
        except ProposalInfo.DoesNotExist:
            return Response({'message' : 'Invalid Proposal ID sent'}, status=406)

        # if center comes in get request then just return the result for that center
        # this is to implement the reset center functionality
        if center_id:
            try:
                center_id = int(center_id)
            except ValueError:
                return Response({'message' : 'Invalid Center ID provided'}, status=406)
            # fetch center object
            proposal_centers = ProposalCenterMapping.objects.filter(id=center_id)
            if not proposal_centers:
                return Response({'message' : 'Invalid Center ID provided'}, status=406)
        else :
            # fetch all centers for this proposal id
            proposal_centers = ProposalCenterMapping.objects.filter(proposal=proposal)

        centers_data_list = []
        # iterate all centers one by one
        for proposal_center in proposal_centers:
            try:
                space_mapping_object = SpaceMapping.objects.get(center=proposal_center)
            except SpaceMapping.DoesNotExist:
                return Response({'message' : 'Space Mapping Does Not Exist'}, status=406)

            space_info_dict = {}

            # calculate the max, min lat long
            delta_dict = website_utils.get_delta_latitude_longitude(float(proposal_center.radius), float(proposal_center.latitude))

            delta_latitude = delta_dict['delta_latitude']
            min_latitude = proposal_center.latitude - delta_latitude
            max_latitude = proposal_center.latitude + delta_latitude

            delta_longitude = delta_dict['delta_longitude']
            min_longitude = proposal_center.longitude - delta_longitude
            max_longitude = proposal_center.longitude + delta_longitude
            # for society
            if space_mapping_object.society_allowed:
                q = Q(society_latitude__lt=max_latitude) & Q(society_latitude__gt=min_latitude) & Q(society_longitude__lt=max_longitude) & Q(society_longitude__gt=min_longitude)

                societies_inventory = space_mapping_object.get_society_inventories()
                societies_inventory_serializer = InventoryTypeSerializer(societies_inventory)
                # applying filter on basis of inventory
                for param in ['poster_allowed', 'standee_allowed', 'flier_allowed', 'stall_allowed', 'banner_allowed']:
                    try:
                        if param == 'poster_allowed' and societies_inventory.__dict__[param]:
                            q &= (Q(poster_allowed_nb=True) | Q(poster_allowed_lift=True))
                        elif societies_inventory.__dict__[param]:
                            q |= Q(**{param : True})

                    except KeyError:
                        pass

                # get all societies based on the query q
                societies_temp = SupplierTypeSociety.objects.filter(q).values('supplier_id','society_latitude','society_longitude','society_name','society_address1', 'society_address2', 'society_subarea', 'society_locality', 'society_location_type', 'flat_count', 'average_rent', 'machadalo_index', 'society_type_quality','tower_count','flat_count')

                # to maintain list of shortlisted societies
                societies = []

                # to maintain list of society_ids
                society_ids = []

                # to maintain total count of societies
                societies_count = 0

                # for each society within defined lat long, process the society
                for society in societies_temp:
                    if website_utils.space_on_circle(proposal_center.latitude, proposal_center.longitude, proposal_center.radius, \
                        society['society_latitude'], society['society_longitude']):
                        society_inventory_obj = InventorySummary.objects.get_supplier_type_specific_object(request.data.copy(),
                                                                                                           society['supplier_id'])
                        adinventory_type_dict = ui_utils.adinventory_func()
                        duration_type_dict = ui_utils.duration_type_func()

                        if society_inventory_obj:
                            society['shortlisted'] = True
                            society['buffer_status'] = False

                            # not required
                            if society_inventory_obj.poster_allowed_nb or society_inventory_obj.poster_allowed_lift:
                                society['total_poster_count'] = society_inventory_obj.total_poster_count
                                society['poster_price'] = return_price(adinventory_type_dict, duration_type_dict, 'poster_a4', 'campaign_weekly')
                            # not required
                            if society_inventory_obj.standee_allowed:
                                society['total_standee_count'] = society_inventory_obj.total_standee_count
                                society['standee_price'] = return_price(adinventory_type_dict, duration_type_dict, 'standee_small', 'campaign_weekly')

                            if society_inventory_obj.stall_allowed:
                                society['total_stall_count'] = society_inventory_obj.total_stall_count
                                society['stall_price'] = return_price(adinventory_type_dict, duration_type_dict, 'stall_small', 'unit_daily')
                                society['car_display_price'] = return_price(adinventory_type_dict, duration_type_dict, 'car_display_standard', 'unit_daily')

                            if society_inventory_obj.flier_allowed:
                                society['flier_frequency'] = society_inventory_obj.flier_frequency
                                society['filer_price'] = return_price(adinventory_type_dict, duration_type_dict, 'flier_door_to_door', 'unit_daily')

                        # append the society id
                        society_ids.append(society['supplier_id'])

                        # append the society object to list
                        societies.append(society)

                        # increment the societies count
                        societies_count += 1
                # aggregate information over counts of standee, flier, poster etc is required.
                societies_inventory_count =  InventorySummary.objects.filter(supplier_id__in=society_ids).aggregate(posters=Sum('total_poster_count'),\
                    standees=Sum('total_standee_count'), stalls=Sum('total_stall_count'), fliers=Sum('flier_frequency'))

                # add all societies shortlisted
                space_info_dict['societies'] = societies

                space_info_dict['societies_inventory_count'] = societies_inventory_count
                space_info_dict['societies_inventory'] = societies_inventory_serializer.data
                space_info_dict['societies_count'] = societies_count


            if space_mapping_object.corporate_allowed:
                q = Q(latitude__lt=max_latitude) & Q(latitude__gt=min_latitude) & Q(longitude__lt=max_longitude) & Q(longitude__gt=min_longitude)

                # ADDNEW --> uncomment this line when corporate inventory implemented
                corporates_inventory = space_mapping_object.get_corporate_inventories()
                corporates_inventory_serializer = InventoryTypeSerializer(corporates_inventory)
                # then run for loop almost same as above for applying filter on inventory_allowed
                # make a query for different inventory count (e.g. poster_count )

                corporates_temp = SupplierTypeCorporate.objects.filter(q).values('supplier_id','latitude','longitude')
                corporates = []
                corporate_ids = []
                corporates_count = 0
                for corporate in corporates_temp:
                    if website_utils.space_on_circle(proposal_center.latitude, proposal_center.longitude, proposal_center.radius, \
                        corporate['latitude'], corporate['longitude']):
                        corporate_inventory_obj = InventorySummary.objects.get_supplier_type_specific_object(request.data.copy(), corporate['supplier_id'])
                        adinventory_type_dict = ui_utils.adinventory_func()
                        duration_type_dict = ui_utils.duration_type_func()
                        if corporate_inventory_obj:
                            corporate['shortlisted'] = True
                            corporate['buffer_status'] = False

                            if corporate_inventory_obj.poster_allowed_nb or corporate_inventory_obj.poster_allowed_lift:
                                corporate['total_poster_count'] = corporate_inventory_obj.total_poster_count
                                corporate['poster_price'] = return_price(adinventory_type_dict, duration_type_dict, 'poster_a4', 'campaign_weekly')
                            if corporate_inventory_obj.standee_allowed:
                                corporate['total_standee_count'] = corporate_inventory_obj.total_standee_count
                                corporate['standee_price'] = return_price(adinventory_type_dict, duration_type_dict, 'standee_small', 'campaign_weekly')
                            if corporate_inventory_obj.stall_allowed:
                                corporate['total_stall_count'] = corporate_inventory_obj.total_stall_count
                                corporate['stall_price'] = return_price(adinventory_type_dict, duration_type_dict, 'stall_small', 'unit_daily')
                                corporate['car_display_price'] = return_price(adinventory_type_dict, duration_type_dict, 'car_display_standard', 'unit_daily')
                            if corporate_inventory_obj.flier_allowed:
                                corporate['flier_frequency'] = corporate_inventory_obj.flier_frequency
                                corporate['filer_price'] = return_price(adinventory_type_dict, duration_type_dict, 'flier_door_to_door', 'unit_daily')



                        corporate_ids.append(corporate['supplier_id'])
                        corporates.append(corporate)
                        corporates_count += 1

                # corporates_serializer = ProposalCorporateSerializer(corporates, many=True)

                carporates_inventory_count =  InventorySummary.objects.filter(supplier_id__in=corporate_ids).aggregate(posters=Sum('total_poster_count'),\
                    standees=Sum('total_standee_count'), stalls=Sum('total_stall_count'), fliers=Sum('flier_frequency'))

                space_info_dict['corporates'] = corporates
                space_info_dict['corporates_count'] = corporates_count
                space_info_dict['corporates_inventory_count'] = carporates_inventory_count  #implement this first
                space_info_dict['corporates_inventory'] = corporates_inventory_serializer.data

                # add inventory count information
                space_info_dict['societies_inventory_count'] = societies_inventory_count

                # add InventoryType details
                space_info_dict['societies_inventory'] = societies_inventory_serializer.data

                # add the total count of shortlisted societies
                space_info_dict['societies_count'] = societies_count

            # add center information
            proposal_center_serializer = ProposalCenterMappingSpaceSerializer(proposal_center)
            space_info_dict['center'] = proposal_center_serializer.data

            # append the structure obtained into the list for centers
            centers_data_list.append(space_info_dict)

        # return the result
        response['centers'] = centers_data_list

        return Response(response, status=200)



    def post(self, request,proposal_id=None, format=None):
        '''This API returns the spaces info when center or radius is changed
        API ONLY PRODUCE RESULTS FOR SOCIETIES ONLY
        Code to be written for other spaces
        ---


        '''
        response = {}
        center_info = request.data
        try:
            center = center_info['center']
            space_mappings = center['space_mappings']
        except KeyError:
            return Response({'message':'Improper center or space_mappings data'}, status=406)

        response['center'] = center
        latitude = float(center['latitude'])
        longitude = float(center['longitude'])
        radius = float(center['radius'])
        # area= "Andheri(E)"
        delta_dict = website_utils.get_delta_latitude_longitude(radius, latitude)

        delta_latitude = delta_dict['delta_latitude']
        min_latitude = center['latitude'] - delta_latitude
        max_latitude = center['latitude'] + delta_latitude

        delta_longitude = delta_dict['delta_longitude']
        min_longitude = center['longitude'] - delta_longitude
        max_longitude = center['longitude'] + delta_longitude
        q = Q()
        if space_mappings['society_allowed']:
            q &= Q(society_latitude__lt=max_latitude) & Q(society_latitude__gt=min_latitude) & Q(society_longitude__lt=max_longitude) & Q(society_longitude__gt=min_longitude)
            # p = Q(society_locality=area)
            try:
                societies_inventory = center_info['societies_inventory']
            except KeyError:
                return Response({'message' : 'Provide society_inventory when society selected'}, status=406)

            for param in ['poster_allowed', 'standee_allowed', 'flier_allowed', 'stall_allowed', 'banner_allowed']:
                try:
                    if param == 'poster_allowed' and societies_inventory[param]:
                        q &= (Q(poster_allowed_nb=True) | Q(poster_allowed_lift=True))
                    elif societies_inventory[param]:
                        q &= Q(**{param : True})
                except KeyError:
                    pass
            # societies_temp1 = SupplierTypeSociety.objects.filter(p).values('supplier_id','society_latitude','society_longitude','society_zip')
            societies_temp = SupplierTypeSociety.objects.filter(q).values('supplier_id','society_latitude','society_longitude','society_name','society_address1','society_subarea','society_location_type','tower_count','flat_count','society_type_quality')
            societies = []
            society_ids = []
            societies_count = 0
            for society in societies_temp:
                if website_utils.space_on_circle(latitude, longitude, radius, society['society_latitude'], society['society_longitude']):
                    society_inventory_obj = InventorySummary.objects.get_supplier_type_specific_object(request.data.copy(),
                                                                                                       society['supplier_id'])
                    if society_inventory_obj:
                    #society_inventory_obj = InventorySummary.objects.get(supplier_id=society['supplier_id'])
                        society['shortlisted'] = True
                        society['buffer_status'] = False
                        # obj = InventorySummaryAPIView()
                        adinventory_type_dict = ui_utils.adinventory_func()
                        duration_type_dict = ui_utils.duration_type_func()
                        if society_inventory_obj.poster_allowed_nb or society_inventory_obj.poster_allowed_lift:
                            society['total_poster_count'] = society_inventory_obj.total_poster_count
                            society['poster_price'] = return_price(adinventory_type_dict, duration_type_dict, 'poster_a4', 'campaign_weekly')

                        if society_inventory_obj.standee_allowed:
                            society['total_standee_count'] = society_inventory_obj.total_standee_count
                            society['standee_price'] = return_price(adinventory_type_dict, duration_type_dict, 'standee_small', 'campaign_weekly')

                        if society_inventory_obj.stall_allowed:
                            society['total_stall_count'] = society_inventory_obj.total_stall_count
                            society['stall_price'] = return_price(adinventory_type_dict, duration_type_dict, 'stall_small', 'unit_daily')
                            society['car_display_price'] = return_price(adinventory_type_dict, duration_type_dict, 'car_display_standard', 'unit_daily')

                        if society_inventory_obj.flier_allowed:
                            society['flier_frequency'] = society_inventory_obj.flier_frequency
                            society['filer_price'] = return_price(adinventory_type_dict, duration_type_dict, 'flier_door_to_door', 'unit_daily')

                    # ADDNEW -->
                    society_ids.append(society['supplier_id'])
                    societies.append(society)
                    societies_count += 1

            # societies_serializer =  ProposalSocietySerializer(societies, many=True)

            # following query find sum of all the variables specified in a dictionary
            # this finds sum of all inventories, if you don't need some of some inventory make it 0 in front end
            societies_inventory_count =  InventorySummary.objects.filter(supplier_id__in=society_ids).aggregate(posters=Sum('total_poster_count'),\
                standees=Sum('total_standee_count'), stalls=Sum('total_stall_count'), fliers=Sum('flier_frequency'))

            response['suppliers'] = societies
            response['supplier_inventory_count'] = societies_inventory_count
            response['supplier_inventory'] = societies_inventory
            response['supplier_count'] = societies_count
            # response['area_societies'] = societies_temp1

        if space_mappings['corporate_allowed']:
            pass
            # q = Q(latitude__lt=max_latitude) & Q(latitude__gt=min_latitude) & Q(longitude__lt=max_longitude) & Q(longitude__gt=min_longitude)

            # ADDNEW --> uncomment when implemented properly
            # try:
            #     corporates_inventory = center_info['corporate_inventory']
            # except KeyError:
            #     return Response({'message' : 'Provide corporates_inventory when corporates selected'}, status=406)

            # filter on basis of corporates_inventory (poster_allowed etc.) like in society_filter


            # corporates_temp = SupplierTypeCorporate.objects.filter(q)
            # corporates = []
            # corporates_count = 0
            # for corporate in corporates_temp:
            #     if space_on_circle(center['latitude'], center['longitude'], center['radius'], \
            #         corporate.latitude, corporate.longitude):
            #         corporates.append(corporate)
            #         corporates_count += 1

            # corporates_serializer = ProposalCorporateSerializer(corporates, many=True)

            # response['corporates'] = corporates_serializer.data
            # response['corporates_count'] = corporates_count
            # response['corporates_inventory'] = corporates_inventory

        if space_mappings['gym_allowed']:
            # ADDNEW --> write code for gym
            pass

        if space_mappings['salon_allowed']:
            # ADDNEW --> write code for salon
            pass

        return Response(response, status=200)


class GenericExportData(APIView):
    """
        The request is in form:
        [
             {
                  center : { id : 1 , center_name: c1, ...   } ,
                  suppliers: { 'RS' : [ { 'supplier_type_code': 'RS', 'status': 'R', 'supplier_id' : '1'}, {...}, {...} }
                  suppliers_meta: {
                                     'RS': { 'inventory_type_selected' : [ 'PO', 'POST', 'ST' ]  },
                                     'CP': { 'inventory_type_selected':  ['ST']
                  }
             }
        ]
        Exports supplier data on grid view page.
        The API is divided into two phases :
        1. extension of Headers and DATA keys. This is because one or more inventory type selected map to more HEADER
        and hence more DATA keys
        2. Making of individual rows. Number of rows in the sheet is equal to total number of societies in all centers combined
        """
    # renderer_classes = (website_renderers.XlsxRenderer, )
    # permission_classes = (v0_permissions.IsGeneralBdUser, )

    def post(self, request, proposal_id=None):
        class_name = self.__class__.__name__
        try:
            data = website_utils.setup_generic_export(request.data, request.user, proposal_id)
            response = ui_utils.handle_response(class_name, data=data, success=True)
            # attach some custom headers
            # response['Content-Disposition'] = 'attachment; filename=%s' % data['name']
            return response
        except Exception as e:
            return ui_utils.handle_response(class_name, exception_object=e, request=request)


class Business(APIView):
    """
    Test api. will be deleted later on.
    """
    def get(self, request):
        class_name = self.__class__.__name__
        try:
            master_user = BaseUser.objects.get(id=8)
            result = AccountInfo.objects.get_permission(user=master_user)
            # result = AccountInfo.objects.filter_permission(user=master_user)
            serializer = AccountSerializer(result)
            return ui_utils.handle_response(class_name, data=serializer.data, success=True)
        except Exception as e:
            return ui_utils.handle_response(class_name, exception_object=e, request=request)


class AssignCampaign(APIView):
    """
    This API assigns a campaign to a user by a user
    """
    def post(self, request):
        """
        Args:
            request: The request object
            campaign_id: Assigns campaign_id to a user by a user.
        Returns: Success in case campaign is successfully assigned to the user.
        """
        class_name = self.__class__.__name__
        try:
            assigned_by = request.user
            assigned_to = request.data['to']

            campaign_id = request.data['campaign_id']
            if assigned_by.is_anonymous:
                return ui_utils.handle_response(class_name, data='A campaign cannot be assigned by an Anonymous user')

            if not assigned_to:
                return ui_utils.handle_response(class_name, data='You must provide a user to which this campaign will be assigned')

            # fetch ProposalInfo object.
            proposal = ProposalInfo.objects.get(proposal_id=campaign_id)

            response = website_utils.is_campaign(proposal)
            if not response.data['status']:
                return response

            # todo: check for dates also. you should not assign a past campaign to any user. left for later
            # fetch BaseUser object.
            for assigned_id in assigned_to:
                assigned_to_user = BaseUser.objects.get(id=assigned_id['id'])
                if assigned_to_user:
                    # create the object.
                    pre_assignment = CampaignAssignment.objects.filter(campaign_id=campaign_id, assigned_to=assigned_to_user)
                    if len(pre_assignment) <= 0:
                        # return ui_utils.handle_response(class_name, data='Campaign Already assigned to this user')
                        instance = CampaignAssignment(
                            **{'campaign_id': campaign_id, 'assigned_by': assigned_by, 'assigned_to': assigned_to_user})
                        instance.save()

            return ui_utils.handle_response(class_name, data='success', success=True)

        except ObjectDoesNotExist as e:
            return ui_utils.handle_response(class_name, exception_object=e, request=request)
        except Exception as e:
            return ui_utils.handle_response(class_name, exception_object=e, request=request)

    def get(self, request):
        """
        Args:
            request: Request object

        Returns: AssignedCampaign objects.
        {
          "status": true,
          "data": [
            {
              "id": 1,
              "assigned_by": {
                "id": 1,
                "first_name": "Nikhil",
                "last_name": "Kumar",
              },
              "assigned_to": {
                "id": 9,
                "first_name": "kamlesh",
                "last_name": "sabziwale"
              },
              "campaign": {
                "proposal_id": "zUkqLPYR",
                "created_at": "2016-12-01T00:00:00Z",
                "updated_at": "2016-12-15T10:18:04.514008Z",
              },
            }
          ]
        }
        Handles four  cases :
        Fetches campaigns assigned by a user only
        Fetches campaigns assigned to a particular user only
        Fetches campaigns assigned by a user to a particular user.
        Fetches all campaigns in the database if the requesting user is superuser
        """
        class_name = self.__class__.__name__

        try:
            users = BaseUser.objects.all().values('id', 'username')
            user_obj = {}
            if users:
                for user in users:
                    row = {
                        "id": user.get('id', None),
                        "username": user.get('username', None)
                    }
                    if not user_obj.get(user['id']):
                        user_obj[user['id']] = row
            user = request.user

            if user.is_superuser:
                assigned_objects = CampaignAssignment.objects.all()
            elif user.profile.name == "Intern":
                assigned_objects = CampaignAssignment.objects.filter(assigned_to_id=user)
            else:
                username_list = BaseUser.objects.filter(profile__organisation=user.profile.organisation.organisation_id). \
                            values_list('username')
                assigned_objects = CampaignAssignment.objects.filter(campaign__created_by__in=username_list)
            campaigns = []
            all_proposal_ids = []
            # check each one of them weather they are campaign or not
            for assign_object in assigned_objects:
                if assign_object.campaign.is_disabled:
                    continue
                response = website_utils.is_campaign(assign_object.campaign)
                # if it is a campaign.
                if response.data['status']:
                    campaigns.append(assign_object)
            serializer = CampaignAssignmentSerializerReadOnly(campaigns, many=True)

            for data in serializer.data:
                all_proposal_ids.append(data['campaign']['proposal_id'])
            
            comments_list = CampaignComments.objects.values('campaign_id').annotate(latest_id=Max('id'), comment_max=Trim('comment')).filter(campaign_id__in=all_proposal_ids, related_to='campaign')
            comments_list_dict = {row["campaign_id"]: row["comment_max"] for row in comments_list}

            campaign_obj = {}

            for data in serializer.data:

                accountResult = AccountInfo.objects.filter(pk=data['campaign']['account']).first()
                data['campaign']['accountName'] = accountResult.name
                data['campaign']['accountOrganisationName'] = accountResult.organisation.name


                organisationResult = Organisation.objects.filter(organisation_id=data['campaign']['principal_vendor']).first()
                if organisationResult:
                    data['campaign']['organisationName'] = organisationResult.name
                else:
                    data['campaign']['organisationName'] = ''

                proposal_start_date = parse_datetime(data['campaign']['tentative_start_date'])
                proposal_end_date = parse_datetime(data['campaign']['tentative_end_date'])
                response = website_utils.get_campaign_status(proposal_start_date, proposal_end_date)
                if not response.data['status']:
                    return response
                data['campaign']['status'] = response.data['data']
                if not campaign_obj.get(data['campaign']['proposal_id']):
                    campaign_obj[data['campaign']['proposal_id']] = data
                    campaign_obj[data['campaign']['proposal_id']]["assigned"] = []
                row = {
                    "assigned_by": user_obj[data["assigned_by"]]['username'],
                    "assigned_to": user_obj[data["assigned_to"]]['username'],
                }

                campaign_obj[data['campaign']['proposal_id']]["assigned"].append(row)
                campaign_obj[data['campaign']['proposal_id']]["comment"] = comments_list_dict.get(data['campaign']['proposal_id'])

            campaign_list = [value for key,value in campaign_obj.items()]

            return ui_utils.handle_response(class_name, data=campaign_list, success=True)
        except ObjectDoesNotExist as e:
            return ui_utils.handle_response(class_name, exception_object=e, request=request)
        except KeyError as e:
            return ui_utils.handle_response(class_name, data='key Error', exception_object=e, request=request)
        except Exception as e:
            return ui_utils.handle_response(class_name, exception_object=e, request=request)


class ImportCorporateData(APIView):
    """
    This API reads a csv file and  makes supplier id's for each row. then it adds the data along with
    supplier id in the  supplier_society table. it also populates society_tower table.
    """

    def get(self, request):
        """
        :param request: request object
        :return: success response in case it succeeds else failure message.
        """
        class_name = self.__class__.__name__
        try:
            source_file = open(BASE_DIR + '/files/corporate.csv', 'rb')
            with transaction.atomic():
                reader = csv.reader(source_file)
                for num, row in enumerate(reader):
                    data = {}
                    if num == 0:
                        continue
                    else:
                        if len(row) != len(v0_constants.corporate_keys):
                            return ui_utils.handle_response(class_name, data=errors.LENGTH_MISMATCH_ERROR.format(len(row), len(v0_constants.corporate_keys)))

                        for index, key in enumerate(v0_constants.corporate_keys):
                            if row[index] == '':
                                data[key] = None
                            else:
                                data[key] = row[index]
                        state_name = v0_constants.state_name
                        state_code = v0_constants.state_code
                        state_object = State.objects.get(state_name=state_name, state_code=state_code)
                        city_object = City.objects.get(city_code=data['city_code'], state_code=state_object)
                        area_object = CityArea.objects.get(area_code=data['area_code'], city_code=city_object)
                        subarea_object = CitySubArea.objects.get(subarea_code=data['subarea_code'],area_code=area_object)
                        # make the data needed to make supplier_id
                        supplier_id_data = {
                            'city_code': data['city_code'],
                            'area_code': data['area_code'],
                            'subarea_code': data['subarea_code'],
                            'supplier_type': data['supplier_type'],
                            'supplier_code': data['supplier_code']
                        }

                        data['supplier_id'] =   get_supplier_id(supplier_id_data)

                        (corporate_object, value) = SupplierTypeCorporate.objects.get_or_create(supplier_id=data['supplier_id'])

                        data['society_state'] = 'Maharashtra'
                        corporate_object.__dict__.update(data)
                        corporate_object.save()

                        # make entry into PMD here.
                        ui_utils.set_default_pricing(data['supplier_id'], data['supplier_type'])

                        url = reverse('inventory-summary', kwargs={'id': data['supplier_id']})
                        url = BASE_URL + url[1:]
                        headers = {
                            'Content-Type': 'application/json',
                            'Authorization': request.META.get('HTTP_AUTHORIZATION', '')
                        }
                        response = requests.post(url, json.dumps(data), headers=headers)
                        print("{0} done \n".format(data['supplier_id']))

            source_file.close()
            return Response(data="success", status=status.HTTP_200_OK)
        except ObjectDoesNotExist as e:
            return ui_utils.handle_response(class_name, data=e.args, exception_object=e, request=request)
        except KeyError as e:
            return ui_utils.handle_response(class_name, data=e.args, exception_object=e, request=request)
        except Exception as e:
            return ui_utils.handle_response(class_name, exception_object=e, request=request)


class AmenityAPIView(APIView):
    """
    API to create an Amenity
    """
    def post(self, request):
        """
        create a single amenity
        Args:
            request:

        Returns:

        """
        class_name = self.__class__.__name__
        try:
            name = request.data['name']
            code = request.data['code']

            if code not in v0_constants.valid_amenities.keys():
                return ui_utils.handle_response(class_name, data=errors.INVALID_AMENITY_CODE_ERROR.format(code))

            amenity, is_created = Amenity.objects.get_or_create(code=code)
            amenity.name = name
            amenity.save()

            return ui_utils.handle_response(class_name, data=model_to_dict(amenity), success=True)
        except Exception as e:
            return ui_utils.handle_response(class_name, exception_object=e, request=request)


class GetAllAmenities(APIView):
    """
    Fetches all amenities
    """
    def get(self, request):
        """
        fetches all amenities
        Args:
            request:

        Returns:

        """
        class_name = self.__class__.__name__
        try:

            type = request.query_params.get('type', None)

            if type:
                amenities = Amenity.objects.filter(type=type)
            else:
                amenities = Amenity.objects.all()
           
            serializer = AmenitySerializer(amenities, many=True)
            return ui_utils.handle_response(class_name, data=serializer.data, success=True)
        except Exception as e:
            return ui_utils.handle_response(class_name, exception_object=e, request=request)

    def post(self, request):
        
        class_name = self.__class__.__name__
        try:

            id = request.data.get('id', None)

            if id:
                item = Amenity.objects.filter(pk=id).first()
                serializer = AmenitySerializer(item, data=request.data)
            else:
                serializer = AmenitySerializer(data=request.data)

            if serializer.is_valid():
                serializer.save()

            return ui_utils.handle_response(class_name, data='success', success=True)
        except Exception as e:
            return ui_utils.handle_response(class_name, exception_object=e, request=request)



class SupplierAmenity(APIView):
    """
    Returns all amenities per supplier
    """
    def get(self, request):
        """
        Args:
            request:

        Returns:
        """
        class_name = self.__class__.__name__
        try:
            supplier_type_code = request.query_params['supplier_type_code']

            response = ui_utils.get_content_type(supplier_type_code)
            if not response.data['status']:
                return response
            content_type = response.data['data']

            amenities = SupplierAmenitiesMap.objects.filter(object_id=request.query_params['supplier_id'], content_type=content_type)
            serializer = SupplierAmenitiesMapSerializer(amenities, many=True)
            return ui_utils.handle_response(class_name, data=serializer.data, success=True)
        except Exception as e:
            return ui_utils.handle_response(class_name, exception_object=e, request=request)

    def post(self, request):
        """
        Args:
            request:
        Returns:
        """
        class_name = self.__class__.__name__
        try:
            supplier_type_code = request.data['supplier_type_code']
            supplier_id = request.data['supplier_id']
            response = website_utils.save_amenities_for_supplier(supplier_type_code, supplier_id, request.data['amenities'])

            if not response.data['status']:
                return response
            return ui_utils.handle_response(class_name, data='success', success=True)

        except Exception as e:
            return ui_utils.handle_response(class_name, exception_object=e, request=request)


class UserMinimalList(APIView):
    """
        returns all users
    """

    def get(self, request):
        """
        Args:
            request:
        Returns: returns all users
        """
        class_name = self.__class__.__name__
        try:
            organisation_id = request.query_params.get('org_id',None)

            page_slug = "get-users-minimal-list-"+str(organisation_id)
            return_data = ui_utils.get_api_cache(page_slug)

            if not return_data:
                if organisation_id:
                    users = BaseUser.objects.filter(profile__organisation=organisation_id)
                else:
                    users = BaseUser.objects.all()
                user_list = []
                for user in users:
                    user_list.append({
                        "id": user.id,
                        "first_name": user.first_name,
                        "last_name": user.last_name,
                        "email": user.email,
                        "username": user.username,
                        "profile_id": user.profile_id,
                        "role_id": user.role_id,
                    })
                
                return_data = user_list
                ui_utils.create_api_cache(page_slug, 'user-list', return_data)

            return ui_utils.handle_response(class_name, data=return_data, success=True)
        except Exception as e:
            return ui_utils.handle_response(class_name, exception_object=e, request=request)

class UserList(APIView):
    """
        returns all users
    """

    def get(self, request):
        """
        Args:
            request:
        Returns: returns all users
        """
        class_name = self.__class__.__name__
        try:
            organisation_id = request.query_params.get('org_id',None)
            if organisation_id:
                users = BaseUser.objects.filter(profile__organisation=organisation_id)
            else:
                users = BaseUser.objects.all()
            user_serializer = BaseUserSerializer(users, many=True)
            return ui_utils.handle_response(class_name, data=user_serializer.data, success=True)
        except Exception as e:
            return ui_utils.handle_response(class_name, exception_object=e, request=request)


class BulkDownloadImagesAmazon(APIView):
    """
    downloads images from Amazon and saves in file directory.
    """
    def get(self, request):
        """
        Args:
            request:

        Returns: The idea is store list of file names ( path ) per supplier. Call a util function to download the files from
        Amazon for each supplier and store it in 'files' directory under folder 'downloaded_images/<proposal_id>'
        """
        class_name = self.__class__.__name__
        if True:
            proposal_id = request.query_params['proposal_id']
            proposal = ProposalInfo.objects.get(proposal_id=proposal_id)
            response = website_utils.is_campaign(proposal)
            if not response.data['status']:
                return response
            # fetch all images for this proposal_id
            inventory_images = InventoryActivityImage.objects.filter(inventory_activity_assignment__inventory_activity__shortlisted_inventory_details__shortlisted_spaces__proposal_id=proposal_id).values(
                'image_path', 'inventory_activity_assignment__inventory_activity__shortlisted_inventory_details__shortlisted_spaces__object_id',
                'inventory_activity_assignment__inventory_activity__shortlisted_inventory_details__shortlisted_spaces__content_type'
            )
            if not inventory_images:
                return ui_utils.handle_response(class_name, data=errors.NO_IMAGES_FOR_THIS_PROPOSAL_MESSAGE.format(proposal_id))

            # store images per supplier_id, content_type
            image_map = {}
            for detail in inventory_images:
                supplier_id = detail['inventory_activity_assignment__inventory_activity__shortlisted_inventory_details__shortlisted_spaces__object_id']
                content_type = detail['inventory_activity_assignment__inventory_activity__shortlisted_inventory_details__shortlisted_spaces__content_type']
                image_name = detail['image_path']
                try:
                    reference = image_map[supplier_id + '_' + str(content_type)]
                    reference.append(image_name)
                except KeyError:
                    image_map[supplier_id + '_' + str(content_type)] = [image_name]
            # initiate the task and return the task id.
            result = website_utils.start_download_from_amazon(proposal_id, json.dumps(image_map))
            return ui_utils.handle_response(class_name, data=result, success=True)
        # except Exception as e:
        #     return ui_utils.handle_response(class_name, exception_object=e, request=request)


class IsGroupTaskSuccessFull(APIView):
    """
    checks the status of the task
    """

    def get(self, request, task_id):
        class_name = self.__class__.__name__
        try:
            result = GroupResult.restore(task_id)
            # 'ready' means have all subtask completed ? 'status' means all substask were successful ?
            return ui_utils.handle_response(class_name, data={'ready': result.ready(), 'status': result.successful()}, success=True)
        except Exception as e:
            return ui_utils.handle_response(class_name, exception_object=e, request=request)


class IsIndividualTaskSuccessFull(APIView):
    """
    checks individual tasks weather they are successfull or not
    """
    def get(self, request, task_id):
        class_name = self.__class__.__name__
        try:
            result = AsyncResult(task_id)
            return ui_utils.handle_response(class_name, data={'ready': result.ready(), 'status': result.successful()}, success=True)
        except Exception as e:
            return ui_utils.handle_response(class_name, exception_object=e, request=request)


class DeleteFileFromSystem(APIView):
    """
    deletes a given file from system
    """

    def post(self, request):
        """
        Args:
            request:

        Returns:
        """
        class_name = self.__class__.__name__
        file_name = ''
        try:
            file_name = request.data['file_name']
            file_extension = file_name.split('.')[1]

            if file_extension not in v0_constants.valid_extensions:
                return ui_utils.handle_response(class_name, data=errors.DELETION_NOT_PERMITTED.format(file_extension), success=False)
                #raise Exception(class_name, errors.DELETION_NOT_PERMITTED.format(file_extension))

            file_path = os.path.join(settings.BASE_DIR, file_name) 

            if os.path.exists(file_path):
               os.remove(os.path.join(settings.BASE_DIR, file_name))
            
            return ui_utils.handle_response(class_name, data='success', success=True)
        except Exception as e:
            return ui_utils.handle_response(class_name, data=str(e), success=False)


class ExportAllSupplierData(APIView):
    """
    Export's all supplier data into sheet and mail it to bd head
    """
    def get(self, request):
        class_name = self.__class__.__name__
        try:
            area_codes = set()
            subarea_codes = set()
            city_codes = set()
            supplier_ids = set()
            result = []
            pricing_dict = {}

            supplier_type_code = request.query_params['supplier_type_code']
            model_class = ui_utils.get_model(supplier_type_code)
            if supplier_type_code == v0_constants.society:
                model_instances = model_class.objects.all().values('supplier_id', 'society_name')
            else:
                model_instances = model_class.objects.all().values('supplier_id', 'name')

            for instance_dict in model_instances:

                supplier_id = instance_dict['supplier_id']
                supplier_id_breakup = website_utils.expand_supplier_id(supplier_id)

                area_codes.add(supplier_id_breakup['area_code'])
                city_codes.add(supplier_id_breakup['city_code'])
                subarea_codes.add(supplier_id_breakup['subarea_code'])
                supplier_ids.add(supplier_id)
                instance_dict['supplier_id_breakup'] = supplier_id_breakup

            city_instances = City.objects.filter(city_code__in=city_codes)
            area_instances = CityArea.objects.filter(area_code__in=area_codes)
            subarea_instances = CitySubArea.objects.filter(subarea_code__in=subarea_codes)

            city_instance_map = {city.city_code: city for city in city_instances}
            area_instance_map = {area.area_code: area for area in area_instances}
            subarea_instance_map = {subarea.subarea_code: subarea for subarea in subarea_instances}

            price_mapping_instance_map = website_utils.get_price_mapping_map(supplier_ids)

            inv_summary_map = website_utils.get_inventory_summary_map(supplier_ids)

            # make pricing dict here
            for supplier_id in supplier_ids:

                pricing_dict[supplier_id] = {}
                pricing_dict[supplier_id]['error'] = 'No Error'

                try:
                    inv_summary_instance = inv_summary_map[supplier_id]
                except KeyError:
                    pricing_dict[supplier_id]['error'] = 'No inventory summary instance for this supplier. Cannot determine which inventories are allowed or not allowed.'
                    continue

                try:
                    price_instances = price_mapping_instance_map[supplier_id]
                except KeyError:
                    pricing_dict[supplier_id]['error'] = 'No pricing instance for this supplier'
                    continue

                # set all possible inventory_allowed fields to 0 first
                for code, name in v0_constants.inventory_code_to_name.items():
                    inv_name_key = website_utils.join_with_underscores(name).lower()
                    pricing_dict[supplier_id][inv_name_key + '_' + 'allowed'] = 0

                # for all inventory codes allowed.
                for inv_code in website_utils.get_inventories_allowed(inv_summary_instance):
                    inventory_name = v0_constants.inventory_code_to_name[inv_code]
                    inv_name_key = website_utils.join_with_underscores(inventory_name).lower()
                    pricing_dict[supplier_id][inv_name_key + '_' + 'allowed'] = 1
                    # for all pricing instances for this supplier
                    for price_instance in price_instances:
                        price_inventory_name = price_instance.adinventory_type.adinventory_name.upper()
                        # proceed only if you find the inventory name in the instance.
                        if inventory_name == price_inventory_name:
                            inventory_meta_type = price_instance.adinventory_type.adinventory_type
                            inventory_duration = price_instance.duration_type.duration_name
                            str_key = inv_name_key + ' ' + inventory_meta_type.lower() + ' ' + inventory_duration.lower()
                            key = website_utils.join_with_underscores(str_key)
                            pricing_dict[supplier_id][key] = price_instance.actual_supplier_price

            for instance_dict in model_instances:

                supplier_dict_breakup = instance_dict['supplier_id_breakup']
                supplier_id = instance_dict['supplier_id']
                supplier_name = instance_dict['society_name'] if instance_dict.get('society_name') else instance_dict['name']
                supplier_code = supplier_dict_breakup['supplier_code']

                city_instance = city_instance_map.get(supplier_dict_breakup['city_code'])
                area_instance = area_instance_map.get(supplier_dict_breakup['area_code'])
                subarea_instance = subarea_instance_map.get(supplier_dict_breakup['subarea_code'])

                basic_data_dict = {

                    'city_name': city_instance.city_name if city_instance else v0_constants.not_in_db_special_code,
                    'city_code': city_instance.city_code if city_instance else v0_constants.not_in_db_special_code,
                    'area_name': area_instance.label if area_instance else v0_constants.not_in_db_special_code,
                    'area_code': area_instance.area_code if area_instance else v0_constants.not_in_db_special_code,
                    'subarea_name': subarea_instance.subarea_name if subarea_instance else v0_constants.not_in_db_special_code,
                    'subarea_code': subarea_instance.subarea_code if subarea_instance else v0_constants.not_in_db_special_code ,
                    'supplier_id': supplier_id,
                    'supplier_name': supplier_name,
                    'supplier_code': supplier_code,
                    'supplier_type_code': supplier_type_code
                }
                # set key, value from pricing_dict to basic_data_dict
                for key, value in pricing_dict[supplier_id].items():
                    basic_data_dict[key] = value

                result.append(basic_data_dict)
            # add pricing headers to current headers.
            headers = v0_constants.basic_supplier_export_headers
            data_keys = v0_constants.basic_supplier_data_keys
            for inventory_name, header_list in v0_constants.price_mapping_default_headers.items():
                for header_tuple in header_list:

                    inv_name_key = website_utils.join_with_underscores(inventory_name).lower()
                    str_key = inv_name_key + ' '+ (' '.join(header_tuple)).lower()
                    key = website_utils.join_with_underscores(str_key)
                    headers.append(key)
                    data_keys.append(key)

            data = {
                'sheet_name': v0_constants.code_to_sheet_names[supplier_type_code],
                'headers': headers,
                'data_keys': data_keys,
                'suppliers': result
            }
            file_name = website_utils.generate_supplier_basic_sheet_mail(data)

            email_data = {
                'subject': 'The all society data of test server in  proper format',
                'body': 'PFA data of all suppliers in the system with supplier_type_code {0}'.format(supplier_type_code),
                'to': [v0_constants.emails['developer']]
            }

            attachment = {
                'file_name': file_name,
                'mime_type': v0_constants.mime['xlsx']
            }

            # send mail to Bd Head with attachment
            bd_head_async_id = tasks.send_email.delay(email_data, attachment=attachment).id
            return ui_utils.handle_response(class_name, data=bd_head_async_id, success=True)

        except Exception as e:
            return ui_utils.handle_response(class_name, exception_object=e, request=request)


class ContactViewSet(viewsets.ViewSet):
    """
    ViewSet around contacts
    """

    def list(self, request):
        """
        list all the contacts for a given object_id or all if none
        :param request:
        :return:
        """
        class_name = self.__class__.__name__
        try:
            object_id = request.query_params.get('object_id')
            if object_id:
                instances = ContactDetails.objects.filter(object_id=object_id)
            else:
                instances = ContactDetails.objects.all()
            serializer = ContactDetailsSerializer(instances, many=True)
            return ui_utils.handle_response(class_name, data=serializer.data, success=True)
        except Exception as e:
            return ui_utils.handle_response(class_name, exception_object=e, request=request)

    def create(self, request):
        """

        :param request:
        :return:
        """
        class_name = self.__class__.__name__
        try:
            serializer = ContactDetailsSerializer(data=request.data)
            if serializer.is_valid():
                serializer.save()
                return ui_utils.handle_response(class_name, data=serializer.data, success=True)
            return ui_utils.handle_response(class_name, data=serializer.errors)
        except Exception as e:
            return ui_utils.handle_response(class_name, exception_object=e, request=request)

    def update(self, request, pk):
        """

        :param request:
        :param pk:
        :return:
        """
        class_name = self.__class__.__name__
        try:
            instance = ContactDetails.objects.get(pk=pk)
            serializer = ContactDetailsSerializer(data=request.data, instance=instance)
            if serializer.is_valid():
                serializer.save()
                return ui_utils.handle_response(class_name, data=serializer.data, success=True)
            return ui_utils.handle_response(class_name, data=serializer.errors)
        except Exception as e:
            return ui_utils.handle_response(class_name, exception_object=e, request=request)


class ProfileViewSet(viewsets.ViewSet):
    """
    Profile View Set
    """
    def list(self, request):
        """
        returns a profile object along with general user permissions and object level permissions instances.
        works with query params which is organisation_id and it's optional
        :param request:
        :return:
        """
        class_name = self.__class__.__name__
        try:
            organisation_id = request.user.profile.organisation.organisation_id
            if request.user.is_superuser:
                instances = Profile.objects.all()
            else:
                org = Organisation.objects.get(organisation_id=organisation_id)
                instances = Profile.objects.filter(organisation=org)
            serializer = ProfileNestedSerializer(instances, many=True)
            return ui_utils.handle_response(class_name, data=serializer.data, success=True)
        except Exception as e:
            return ui_utils.handle_response(class_name, exception_object=e, request=request)

    def create(self, request):
        """
        creates a profile object. Also create associated object level and general user permissions objects.
        :param request:
        :return:
        """
        class_name = self.__class__.__name__
        try:
            serializer = ProfileSimpleSerializer(data=request.data)
            if serializer.is_valid():
                serializer.save()
                for row in request.data.get("general_user_permission"):
                    GeneralUserPermission(
                        codename=row["codename"],
                        description=row["description"],
                        is_allowed=False,
                        name=row["name"],
                        profile_id=serializer.data["id"],
                    ).save()

                instance = Profile.objects.get(pk=serializer.data["id"])
                serializer1 = ProfileNestedSerializer(instance)
                return ui_utils.handle_response(class_name, data=serializer1.data, success=True)
            return ui_utils.handle_response(class_name, data=serializer.errors, success=True)
        except Exception as e:
            return ui_utils.handle_response(class_name, exception_object=e, request=request)

    def retrieve(self, request, pk):
        """

        :param request:
        :param pk:
        :return:
        """
        class_name = self.__class__.__name__
        try:
            instance = Profile.objects.get(pk=pk)
            serializer = ProfileNestedSerializer(instance)
            return ui_utils.handle_response(class_name, data=serializer.data, success=True)
        except Exception as e:
            return ui_utils.handle_response(class_name, exception_object=e, request=request)

    def destroy(self, request, pk):
        """

        :param request:
        :param pk:
        :return:
        """
        class_name = self.__class__.__name__
        try:
            Profile.objects.get(pk=pk).delete()
            return ui_utils.handle_response(class_name, data=True, success=True)
        except Exception as e:
            return ui_utils.handle_response(class_name, exception_object=e, request=request)

    def update(self, request, pk):
        """
        :param request:
        :param pk:
        :return:
        """
        class_name = self.__class__.__name__
        try:
            instance = Profile.objects.get(pk=pk)
            serializer = ProfileSimpleSerializer(data=request.data,instance=instance)
            if serializer.is_valid():
                serializer.save()
                return ui_utils.handle_response(class_name, data=serializer.data, success=True)
        except Exception as e:
            return ui_utils.handle_response(class_name, exception_object=e, request=request)

    @list_route(methods=['GET'])
    def standard_profiles(self, request):
        """
        fetches standard profiles from the system. Any profile which is marked is_standard=True and whose Organisation Cateogry is 'MACHADALO',
        is qualified as standard profile. This list is used in creating profiles for other organisations quickly.
        Any other organisation will only clone profile, not create one.
        :param request:
        :return:
        """
        class_name = self.__class__.__name__
        try:
            instances = Profile.objects.filter(is_standard=True)
            serializer = ProfileSimpleSerializer(instances, many=True)
            return ui_utils.handle_response(class_name, data=serializer.data, success=True)
        except Exception as e:
            return ui_utils.handle_response(class_name, exception_object=e, request=request)

class OrganisationViewSet(viewsets.ViewSet):
    """
    ViewSets around organisations.
    """
    def list(self, request):
        """
        list all organisations for provided category
        :param request
        :param
        :return
        """

        class_name = self.__class__.__name__
        try:
            category = request.query_params.get('category')
            organisation_id = request.user.profile.organisation.organisation_id

            page_slug = "organisation-"+str(organisation_id)+"-"+str(category)
            return_data = ui_utils.get_api_cache(page_slug)

            if not return_data:
                if request.user.is_superuser:
                    if category:
                        instances = Organisation.objects.filter(category__in=[category,"MACHADALO"])
                    else:
                        instances = Organisation.objects.all()
                elif category:
                    instances = Organisation.objects.filter_permission(user=request.user, category__in=[category,"MACHADALO"], created_by_org=organisation_id)
                else:
                    instances = Organisation.objects.filter_permission(user=request.user, created_by_org=organisation_id)
                serializer = OrganisationSerializer(instances, many=True)

                return_data = serializer.data
                ui_utils.create_api_cache(page_slug, "organisation", return_data)

            return ui_utils.handle_response(class_name, data=return_data, success=True)
        except Exception as e:
            return ui_utils.handle_response(class_name, exception_object=e, request=request)

    def retrieve(self, request, pk):
        """
        :param request
        :return
        """
        class_name = self.__class__.__name__
        try:
            instance = Organisation.objects.get_permission(user=request.user, pk=pk)
            serializer = OrganisationSerializer(instance)
            return ui_utils.handle_response(class_name, data=serializer.data, success=True)
        except Exception as e:
            return ui_utils.handle_response(class_name,exception_object=e, request=request)

    @list_route(methods=['GET'])
    def get_organisations_for_assignment(self, request):
        class_name = self.__class__.__name__
        try:
            organisation_id = request.user.profile.organisation.organisation_id
            if organisation_id == v0_constants.MACHADALO_ORG_ID:
                data = Organisation.objects.all()
            else:
                instance = Organisation.objects.filter(organisation_id=organisation_id)
                created_by_org = Organisation.objects.filter(organisation_id=instance[0].created_by_org.organisation_id)
                data = instance.union(created_by_org)
            serializer = OrganisationSerializer(data, many=True)
            return ui_utils.handle_response(class_name, data=serializer.data, success=True)
        except Exception as e:
            return ui_utils.handle_response(class_name, exception_object=e, request=request)


    def create(self, request):
        """
        :param request:
        :return:
        """

        class_name = self.__class__.__name__
        try:
            data = request.data.copy()
            data['user'] = request.user.pk
            data['organisation_id'] = website_utils.get_generic_id([data['category'], data['name']])
            data['created_by_org'] = request.user.profile.organisation.organisation_id
            serializer = OrganisationSerializer(data=data)
            if serializer.is_valid():
                serializer.save()
                return ui_utils.handle_response(class_name, data=serializer.data, success=True)
            return ui_utils.handle_response(class_name, data=serializer.errors)
        except Exception as e:
            return ui_utils.handle_response(class_name, exception_object=e, request=request)

    def update(self, request, pk):
        """
        :param request:
        :param pk:
        :return:
        """

        class_name = self.__class__.__name__
        try:
            instance = Organisation.objects.get_permission(user=request.user, check_update_permissions=True, pk=pk)
            serializer = OrganisationSerializer(data=request.data, instance=instance)
            if serializer.is_valid():
                serializer.save()
                return ui_utils.handle_response(class_name, data=serializer.data, success=True)
            return ui_utils.handle_response(class_name, data=serializer.errors)
        except Exception as e:
            return ui_utils.handle_response(class_name, exception_object=e, request=request)


class ContentTypeViewSet(viewsets.ViewSet):
    """
    fetches all content types in the system
    """

    def list(self, request):
        """

        :param request:
        :return:
        """
        class_name = self.__class__.__name__
        try:
            valid_models = [Organisation, BaseUser, Profile, AccountInfo, ProposalInfo]
            serializer = ContentTypeSerializer(ContentType.objects.get_for_models(*valid_models).values(), many=True)
            return ui_utils.handle_response(class_name, data=serializer.data, success=True)
        except Exception as e:
            return ui_utils.handle_response(class_name, exception_object=e, request=request)

    def retrieve(self, request, pk):
        """

        :param request:
        :param pk:
        :return:
        """
        class_name = self.__class__.__name__
        try:
            instance = ContentType.objects.get_for_id(id=pk)
            serializer = ContentTypeSerializer(instance)
            return ui_utils.handle_response(class_name, data=serializer.data, success=True)
        except Exception as e:
            return ui_utils.handle_response(class_name, exception_object=e, request=request)


class CloneProfile(APIView):
    """
    Clone a profile.
    """
    def post(self, request):
        """
        clones a given profile to a new profile with a new name and new organisation_id. copies all object
        level and general user permission.
        """
        class_name = self.__class__.__name__
        try:
            profile_pk = request.data['clone_from_profile_id']
            profile_instance = Profile.objects.get(pk=profile_pk)
            new_organisation_id = request.data['new_organisation_id']
            new_profile_name = request.data['new_name']

            data = {
                'name': new_profile_name,
                'organisation': new_organisation_id,
                'is_standard': False
            }
            serializer = ProfileSimpleSerializer(data=data)

            if serializer.is_valid():
                new_profile_instance = serializer.save()
                # copy all object level permissions to new profile. Running .save() call in a loop is okay here
                # as there won't be much objects per profile in our system
                for object_level_permission in ObjectLevelPermission.objects.filter(profile=profile_instance):
                    object_level_permission.pk = None
                    object_level_permission.profile = new_profile_instance
                    object_level_permission.save()
                # copy all general user permissions to new profile. Running .save() call in a loop is okay here
                # as there won't be much objects per profile in our system here.
                for general_user_permission in GeneralUserPermission.objects.filter(profile=profile_instance):
                    general_user_permission.pk = None
                    general_user_permission.profile = new_profile_instance
                    general_user_permission.save()
                return ui_utils.handle_response(class_name, data=serializer.data, success=True)
            else:
                return ui_utils.handle_response(class_name, data=serializer.errors, success=True)
        except Exception as e:
            return ui_utils.handle_response(class_name, exception_object=e, request=request)


class OrganisationMapViewSet(viewsets.ViewSet):
    """
    viewset that around OrganisationMap model
    """

    def list(self, request):
        """
        returns a dict {
           'details' : { k1: {}, k2: {} },
           'mapping' : { k1: [ k2, k3], k2: [ k3 ], k3: [ k1 ] }
        }
        :param request:
        :return:
        """
        class_name = self.__class__.__name__
        try:
            source_organisation_id = request.query_params.get('source_organisation_id')
            instance = Organisation.objects.get(pk=source_organisation_id)
            if source_organisation_id:
                instances = OrganisationMap.objects.filter(Q(first_organisation=instance) or Q(second_organisation=instance))
            else:
                instances = OrganisationMap.objects.all()
            serializer = OrganisationMapNestedSerializer(instances, many=True)
            return ui_utils.handle_response(class_name, data=serializer.data, success=True)
        except Exception as e:
            return ui_utils.handle_response(class_name, exception_object=e, request=request)

    def create(self, request):
        """
        creates a map
        :param request:
        :return:
        """
        class_name = self.__class__.__name__
        try:
            first_organisation = Organisation.objects.get(organisation_id=request.data['first_organisation_id'])
            second_organisation = Organisation.objects.get(organisation_id=request.data['second_organisation_id'])
            instance, is_created = OrganisationMap.objects.get_or_create(first_organisation=first_organisation,second_organisation=second_organisation)
            serializer = OrganisationMapNestedSerializer(instance=instance)
            return ui_utils.handle_response(class_name, data=serializer.data, success=True)
        except Exception as e:
            return ui_utils.handle_response(class_name, exception_object=e, request=request)


class GetAssignedIdImagesListApiView(APIView):
    def get(self, request, organisation_id):

        class_name = self.__class__.__name__

        try:
            user = request.user
            proposal_query = Q(inventory_activity_assignment__inventory_activity__shortlisted_inventory_details__shortlisted_spaces__proposal__campaign_state='PTC')
            proposal_query_assignment = Q(inventory_activity__shortlisted_inventory_details__shortlisted_spaces__proposal__campaign_state='PTC')

            activity_type = request.query_params.get('type', None)
            activity_type_query = Q(inventory_activity_assignment__inventory_activity__activity_type=activity_type)
            activity_type_query_assignment = Q(inventory_activity__activity_type=activity_type)

            activity_date = request.query_params.get('date', None)
            activity_date_query = Q(inventory_activity_assignment__activity_date=activity_date)
            activity_date_query_assignment = Q(activity_date=activity_date)

            inventory = request.query_params.get('inventory', None)
            inv_query = Q(inventory_activity_assignment__inventory_activity__shortlisted_inventory_details__ad_inventory_type__adinventory_name = inventory)
            inv_query_assignment = Q(inventory_activity__shortlisted_inventory_details__ad_inventory_type__adinventory_name=inventory)
            all_users = BaseUser.objects.all().values('id', 'username')
            user_map = {detail['id']: detail['username'] for detail in all_users}

            query = Q()
            query_assignment = Q()
            if not request.user.is_superuser:
                category = request.query_params.get('category', None)
                # if category.upper() == v0_constants.category['business']:
                #     query = Q(inventory_activity_assignment__inventory_activity__shortlisted_inventory_details__shortlisted_spaces__proposal__account__organisation__organisation_id=organisation_id)
                # if category.upper() == v0_constants.category['business_agency']:
                #     query = Q(inventory_activity_assignment__inventory_activity__shortlisted_inventory_details__shortlisted_spaces__proposal__user=user)
                #
                # if category.upper() == v0_constants.category['supplier_agency'] or category.upper() == v0_constants.category['machadalo']:
                query = Q(inventory_activity_assignment__inventory_activity__shortlisted_inventory_details__shortlisted_spaces__proposal__campaignassignment__assigned_to=user)
                query_assignment = Q(inventory_activity__shortlisted_inventory_details__shortlisted_spaces__proposal__campaignassignment__assigned_to=user)

            proposal_alias_name ='inventory_activity_assignment__inventory_activity__shortlisted_inventory_details__shortlisted_spaces__proposal__name'
            proposal_alias_name_assignment = 'inventory_activity__shortlisted_inventory_details__shortlisted_spaces__proposal__name'

            proposal_alias_id = 'inventory_activity_assignment__inventory_activity__shortlisted_inventory_details__shortlisted_spaces__proposal__proposal_id'
            proposal_alias_id_assignment = 'inventory_activity__shortlisted_inventory_details__shortlisted_spaces__proposal__proposal_id'
            shortlisted_inv_alias = 'inventory_activity_assignment__inventory_activity__shortlisted_inventory_details'
            shortlisted_inv_alias_assignment = 'inventory_activity__shortlisted_inventory_details'

            supplier_id = 'inventory_activity_assignment__inventory_activity__shortlisted_inventory_details__shortlisted_spaces__object_id'
            supplier_id_assignment = 'inventory_activity__shortlisted_inventory_details__shortlisted_spaces__object_id'

            inv_act_assignment_objects = InventoryActivityAssignment.objects.select_related('inventory_activity',
                                                                                  'inventory_activity__shortlisted_inventory_details',
                                                                                  'inventory_activity__shortlisted_inventory_details__shortlisted_spaces'). \
                filter(proposal_query_assignment, query_assignment, activity_type_query_assignment, activity_date_query_assignment, inv_query_assignment). \
                annotate(name=F(proposal_alias_name_assignment), inv_id=F(shortlisted_inv_alias_assignment), object_id=F(supplier_id_assignment),
                         proposal_id=F(proposal_alias_id_assignment)). \
                values('name', 'inv_id', 'object_id', 'proposal_id')

            inv_act_image_objects = InventoryActivityImage.objects.select_related('inventory_activity_assignment',
                                                                     'inventory_activity_assignment__inventory_activity',
                                                                     'inventory_activity_assignment__inventory_activity__shortlisted_inventory_details',
                                                                     'inventory_activity_assignment__inventory_activity__shortlisted_inventory_details__shortlisted_spaces'). \
                filter(proposal_query, query, activity_type_query, activity_date_query, inv_query). \
                annotate(name=F(proposal_alias_name),inv_id=F(shortlisted_inv_alias),object_id=F(supplier_id),proposal_id=F(proposal_alias_id),
                         supplier_code=F('inventory_activity_assignment__inventory_activity__shortlisted_inventory_details__shortlisted_spaces__supplier_code')). \
                values('name','inv_id','object_id','latitude','longitude','updated_at','created_at','actual_activity_date','proposal_id','image_path','comment','supplier_code')

            supplier_id_list = []
            for supplier in inv_act_image_objects:
                supplier_id_list.append(supplier["object_id"])

            supplier_objects = SupplierTypeSociety.objects.filter(supplier_id__in=supplier_id_list)
            society_serializer = SupplierTypeSocietySerializer(supplier_objects, many=True)

            master_supplier = SupplierMaster.objects.filter(supplier_id__in=supplier_id_list)
            master_serializer = SupplierMasterSerializer(master_supplier, many=True)

            all_societies = website_utils.manipulate_object_key_values(society_serializer.data)
            master_suppliers = website_utils.manipulate_master_to_rs(master_serializer.data)
            all_societies.extend(master_suppliers)

            inv_act_image_objects_with_distance = website_utils.calculate_location_difference_between_inventory_and_supplier(
                            inv_act_image_objects, all_societies)
            
            inv_act_image_objects_with_distance_map = {}
            if inv_act_image_objects_with_distance:
                inv_act_image_objects_with_distance_map = {inv['inv_id'] : inv for inv in inv_act_image_objects_with_distance}

            result = {}
            for object in inv_act_assignment_objects:
                if object['name'] not in result:
                    result[object['name']] = {}
                    result[object['name']]['assigned'] = {}
                    result[object['name']]['completed'] = {}
                if object['inv_id'] not in result[object['name']]['assigned']:
                    result[object['name']]['assigned'][object['inv_id']] = []
                    # result[object['name']]['completed'][object['inv_id']] = []
                result[object['name']]['assigned'][object['inv_id']].append(object)
                if object['inv_id'] in inv_act_image_objects_with_distance_map:
                    if object['inv_id'] not in result[object['name']]['completed']:
                        result[object['name']]['completed'][object['inv_id']] = []
                    result[object['name']]['completed'][object['inv_id']].append(inv_act_image_objects_with_distance_map[object['inv_id']])
            return ui_utils.handle_response(class_name, data=result, success=True)
        except Exception as e:
            return ui_utils.handle_response(class_name, exception_object=e, request=request)


class proposalCenterMappingViewSet(viewsets.ViewSet):
    """
    viewset around centers created by each proposal
    """
    def list(self,request):
        """

        :param request:
        :return:
        """
        class_name = self.__class__.__name__
        try:
            proposal_id = request.query_params['proposal_id']
            proposal_centers = ProposalCenterMapping.objects.filter(proposal=proposal_id)
            serializer = ProposalCenterMappingSerializer(proposal_centers, many=True)
            return ui_utils.handle_response(class_name, data=serializer.data, success=True)
        except Exception as e:
            return ui_utils.handle_response(class_name, exception_object=e, request=request)


class GetRelationshipAndPastCampaignsData(APIView):
    def get(self, request):
        """
        This api will return supplier relationship data as wll as past campaign related data
        :param request:
        :return:
        """
        class_name = self.__class__.__name__
        try:
            supplier_type_code = request.query_params.get('supplier_code',None)
            supplier_id = request.query_params.get('supplier_id',None)
            campaign_id = request.query_params.get('campaign_id', None)
            
            if supplier_type_code == 'RS':
                supplier_data = SupplierTypeSociety.objects.filter(supplier_id=supplier_id).values('feedback','representative__name')
            else:
                supplier_data = SupplierMaster.objects.filter(supplier_id=supplier_id).values('feedback','representative__name')
            campaign_data = website_utils.get_past_campaigns_data(supplier_id,campaign_id)
            result = {
                'campaign_data': campaign_data,
                'supplier_data': supplier_data,
                'contacts': {}
            }
            contact_details = ContactDetails.objects.filter(object_id=supplier_id).first()
            if contact_details:
                contact_serializer = ContactDetailsSerializer(contact_details, many=False)
                result['contacts'] = contact_serializer.data
           
            return ui_utils.handle_response(class_name, data=result, success=True)
        except Exception as e:
            return ui_utils.handle_response(class_name, exception_object=e, request=request)


class CheckExstingSuppliers(APIView):
    def post(self, request):
        try:
            # source_file = request.data['file']
            supplier_data = SupplierTypeSociety.objects.all()
            source_file = request.data['file']
            wb = load_workbook(source_file)
            ws = wb.get_sheet_by_name(wb.get_sheet_names()[0])
            count = 0
            rcount = 0
            for index, row in enumerate(ws.iter_rows()):
                if index > 0:
                    name = row[0].value
                    id = row[1].value
                    if SupplierTypeSociety.objects.filter(society_name=name) or SupplierTypeSociety.objects.filter(supplier_id=id):
                        count = count + 1
                        s = str(count) + " " + name
                    else:
                        rcount = rcount + 1
            data = {
                'on_platform': count,
                'not_on_platform': rcount
            }
            return ui_utils.handle_response({}, data=data, success=True)
        except Exception as e:
            return ui_utils.handle_response({}, exception_object=e, request=request)

