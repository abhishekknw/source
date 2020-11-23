from __future__ import print_function
from __future__ import absolute_import
import math
import re
import datetime
from io import StringIO
from types import *
import os
import uuid
from string import Template
import json
import shutil
import string
import random
from PIL import Image, ImageDraw, ImageFont
from django.core.files.storage import default_storage
from geopy.geocoders import GoogleV3

from django.db import transaction
from django.db.models import Q, F, ExpressionWrapper, FloatField
from django.apps import apps
from django.contrib.contenttypes.models import ContentType
from django.core.exceptions import ObjectDoesNotExist, MultipleObjectsReturned
from django.shortcuts import get_object_or_404
from django.http import Http404
from django.utils import timezone
from django.conf import settings
from django.db.models import Count, Sum
from django.forms.models import model_to_dict
from django.db import connection
from django.core.cache import cache
from django.utils.dateparse import parse_datetime
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger


from rest_framework.response import Response
from rest_framework import status
from pygeocoder import Geocoder, GeocoderError
import openpyxl
import geocoder
from openpyxl import Workbook
import boto
import boto.s3
from boto.s3.key import Key
from bulk_update.helper import bulk_update
from celery import group
from celery.task.sets import TaskSet, subtask
from collections import namedtuple
import gpxpy.geo

#import v0.models as models
import v0.ui.inventory.models as inventory_models
from v0.ui.account.models import ContactDetails, AccountInfo
from v0.ui.components.models import FlatType, SocietyTower, Amenity
from v0.ui.inventory.models import (AdInventoryType, InventoryActivityAssignment,
                                    InventoryActivityImage, InventorySummary, InventoryActivity, INVENTORY_ACTIVITY_TYPES, PosterInventory)
from v0.ui.inventory.serializers import (InventoryActivityAssignmentSerializer)
from v0.ui.location.models import State, City, CityArea, CitySubArea
from v0.ui.proposal.models import (ProposalInfo, ProposalCenterMapping, ProposalCenterSuppliers, ProposalMetrics,
    ProposalInfoVersion, ProposalMasterCost, ShortlistedSpaces, SupplierPhase, SupplierAssignment)
from v0.ui.proposal.serializers import ProposalInfoSerializer, ProposalCenterMappingSerializer
from v0.ui.campaign.models import CampaignAssignment, GenericExportFileName
import v0.ui.utils as ui_utils
from v0 import errors
from . import tasks
import v0.constants as v0_constants
from v0.ui.finances.models import (RatioDetails, PrintingCost, LogisticOperationsCost, IdeationDesignCost,
                                   SpaceBookingCost, EventStaffingCost, DataSciencesCost,
                                   ShortlistedInventoryPricingDetails, PriceMappingDefault)
from v0.ui.finances.serializers import ShortlistedSpacesSerializerReadOnly
from v0.ui.supplier.models import SupplierAmenitiesMap, SupplierTypeSociety, AddressMaster, SupplierMaster
from v0.ui.supplier.serializers import SupplierTypeSocietySerializer, SupplierMasterSerializer
from v0.ui.permissions.models import Role
from v0.ui.permissions.serializers import RoleHierarchySerializer
from v0.ui.events.models import Events
from v0.ui.base.models import DurationType
from v0.ui.account.serializers import ContactDetailsSerializer
from v0.ui.campaign.serializers import GenericExportFileSerializer
from v0.ui.inventory.models import Filters
from v0.ui.inventory.serializers import FiltersSerializer
from v0.ui.dynamic_suppliers.utils import (get_dynamic_suppliers_by_campaign)
from django.db.models import Prefetch

from v0.ui.organisation.models import Organisation
from v0.ui.organisation.serializers import OrganisationSerializer

from django.db import transaction
import openpyxl

import logging

fonts_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'fonts')

logger = logging.getLogger(__name__)

def get_union_keys_inventory_code(key_type, unique_inventory_codes):
    """
    :param key_type: can take value 'HEADER' or 'DATA'
    :param unique inventory_code: can take value, 'SL', 'CD', 'ST', 'PO'
    :return: a list containing HEADERS/DATA union of all inventory types hidden in inventory_code.
    """

    assert key_type is not None, 'key type should not be None'
    assert unique_inventory_codes is not None, 'inventory code  should not be None'
    assert type(v0_constants.inventorylist) is dict, 'Inventory list is not dict type'
    function = get_union_keys_inventory_code.__name__
    try:
        # for each code in individual_codes union the data and return
        return [item for code in unique_inventory_codes for item in v0_constants.inventorylist.get(code)[key_type]]
    except Exception as e:
        raise Exception(function, ui_utils.get_system_error(e))


def getList(data, society_keys):
    return [data.get(key, '') for key in society_keys]


def remove_duplicates_preserver_order(sequence):
    """
    Args:
        sequence: a list  containing duplicatees
    Returns: a list which does not contains duplicates whilst preserving order of elements in orginal lis
    """
    assert type(sequence) is list, 'Sequence must be list type'
    seen = set()
    seen_add = seen.add
    return [x for x in sequence if not (x in seen or seen_add(x))]


def get_unique_inventory_codes(inventory_array):
    """
    Args:
        inventory_array: array of inventory codes like PL, SL, PLSL.
    Returns: an array containing unique codes. Pl, PLSL would just give [ PL, SL]
    """
    function = get_unique_inventory_codes.__name__
    try:
        # our codes are two letters long (individual)
        step = 2
        # generate individual codes from inventory_array
        individual_codes = [inventory_code[i:i + step] for inventory_code in inventory_array for i in
                            range(0, len(inventory_code), step)]
        return list(set(individual_codes))
    except Exception as e:
        raise Exception(function, ui_utils.get_system_error(e))


def get_union_inventory_price_per_flat(data, unique_inventory_codes, index):
    """
    This function adds an extra key to the dict data based on price and flat count keys
    :param data: it's a dict containing 1 society data
    :param unique_inventory_code: array of inventory codes like 'FL', 'CD', 'ST' etc
    :return: calculates inventory price per flat by dividing two keys of the the dict and stores the result  in the dict itself
    """

    # assert type(data) is dict, 'Data variable should be a dict'
    # assert type(inventory_code) is StringType, 'inventory_code should be a String {0}.'.format(inventory_code)
    function = get_union_inventory_price_per_flat.__name__
    try:
        # iterate over individual codes and calculate for each code and  return
        for code in unique_inventory_codes:
            if data.get('flat_count'):
                inventory_price = data.get(v0_constants.price_per_flat[code][1], 0)
                if not inventory_price:
                    inventory_price = 0.0
                data[v0_constants.price_per_flat[code][0]] = inventory_price / (float(data['flat_count']))
        return data
    except Exception as e:
        raise Exception(function, ui_utils.get_system_error(e))


def insert_supplier_sheet(workbook, result):
    """
    Args:
        workbook: a worksheet object
        result
    Returns: a worksheet in which the right rows are inserted and returned.
    """
    function_name = insert_supplier_sheet.__name__
    try:
        for code, supplier_data in result.items():

            # create a new sheet for each supplier type
            ws = workbook.create_sheet(index=0, title=supplier_data['sheet_name'])

            # set the heading
            ws.append(supplier_data['header_keys'])

            for supplier_object in supplier_data['objects']:
                ws.append([supplier_object[key] for key in supplier_data['data_keys']])

        # we also need to add metric sheet as part of export
        return add_metric_sheet(workbook)
        # return a workbook object
    except Exception as e:
        raise Exception(function_name, ui_utils.get_system_error(e))


def get_related_dict():
    ''' This dictionary is simply a mapping from get params to their actual values '''
    # quality_dict if for both society_quality and its location_quality and similarly for other spaces
    quality_dict = {
        'UH': 'Ultra High', 'HH': 'High',
        'MH': 'Medium High', 'ST': 'Standard'
    }

    inventory_dict = {
        'PO': 'poster_allowed_nb', 'ST': 'standee_allowed',
        'SL': 'stall_allowed', 'FL': 'flier_allowed',
        'BA': 'banner_allowed', 'CD': 'car_display_allowed',
    }

    quantity_dict = {
        'LA': 'Large', 'MD': 'Medium',
        'VL': 'Very Large', 'SM': 'Small',
    }

    return quality_dict, inventory_dict, quantity_dict


def make_filter_query(data):
    """
    This function constructs the filter that later on will be used to query the supplier table.
    """
    try:
        quality_dict, inventory_dict, supplier_quantity_dict = get_related_dict()
        supplier_type_code = data.get('supplier_type_code')

        if not data['max_latitude'] or not data['min_latitude'] or not data['max_longitude'] or not data[
            'min_longitude'] or not data['supplier_type_code']:
            return None

        elif data['supplier_type_code'] == 'RS':
            filter_query = Q(society_latitude__lt=data['max_latitude']) & Q(
                society_latitude__gt=data['min_latitude']) & Q(
                society_longitude__lt=data['max_longitude']) & Q(society_longitude__gt=data['min_longitude'])
        else:
            filter_query = Q(latitude__lt=data['max_latitude']) & Q(latitude__gt=data['min_latitude']) & Q(
                longitude__lt=data['max_longitude']) & Q(longitude__gt=data['min_longitude'])

        if data.get('location_params'):
            location_ratings = []
            location_params = data.get('location_params').split()
            for param in location_params:
                try:
                    location_ratings.append(quality_dict[param])
                except KeyError:
                    pass
            if location_ratings:
                if supplier_type_code == 'RS':  # todo: change this when clarity
                    filter_query &= Q(society_location_type__in=location_ratings)
                else:
                    filter_query &= Q(location_type__in=location_ratings)

                # if required to include suppliers with null value of this parameter uncomment following line
                # will not work if there is default value is something then replace __isnull=True --> = defaultValue
                # q &= Q(supplier_location_type__isnull=True)

        if data.get('supplier_quality_params'):
            supplier_quality_ratings = []
            supplier_quality_params = data.get('supplier_quality_params').split()
            for param in supplier_quality_params:
                try:
                    supplier_quality_ratings.append(quality_dict[param])
                except KeyError:
                    pass

            if supplier_quality_ratings:
                if supplier_type_code == 'RS':
                    filter_query &= Q(society_type_quality__in=supplier_quality_ratings)
                else:
                    filter_query &= Q(quality_rating__in=supplier_quality_ratings)

                # if required to include suppliers with null value of this parameter uncomment following line
                # will not work if there is default value is something then replace __isnull=True --> = defaultValue
                # q &= Q(supplier_type_quality__isnull=True)

        if data.get('supplier_quantity_params'):
            supplier_quantity_ratings = []
            supplier_quantity_params = data.get('supplier_quantity_params').split()
            for param in supplier_quantity_params:
                try:
                    supplier_quantity_ratings.append(supplier_quantity_dict[param])
                except KeyError:
                    pass

            if supplier_quantity_ratings:
                if supplier_type_code == 'RS':
                    filter_query &= Q(society_type_quantity__in=supplier_quantity_ratings)
                else:
                    filter_query &= Q(quantity_rating__in=supplier_quantity_ratings)

                # if required to include suppliers with null value of this parameter uncomment following line
                # will not work if there is default value is something then replace __isnull=True --> = defaultValue
                # q &= Q(supplier_type_quantity__isnull=True)

        if data.get('flat_count'):
            flat_count = data.get('flat_count').split()
            if len(flat_count) == 2:
                flat_min = flat_count[0]
                flat_max = flat_count[1]

                filter_query &= Q(flat_count__gte=flat_min) & Q(flat_count__lte=flat_max)
                # if required to include suppliers with null value of this parameter uncomment following line
                # will not work if there is default value is something then replace __isnull=True --> = defaultValue
                # q &= Q(flat_count__isnull=True)
        p = None
        if data.get('inventory_params'):
            inventory_params = data.get('inventory_params').split()
            temp = None  # temporary variable
            temp = None  # temporary variable
            for param in inventory_params:
                try:

                    # | 'STFL' | 'CDFL' | 'PSLF' | 'STSLFL' | 'POCDFL' | 'STCDFL'
                    if (param == 'POFL') | (param == 'STFL') | (param == 'SLFL') | (param == 'CDFL') | (
                            param == 'POSLFL') | (param == 'STSLFL') | (param == 'POCDFL') | (param == 'STCDFL'):

                        if param == 'POFL':
                            temp_q = (Q(poster_allowed_nb=True) & Q(flier_allowed=True))

                        if param == 'SLFL':
                            temp_q = (Q(stall_allowed=True) & Q(flier_allowed=True))

                        if param == 'STFL':
                            temp_q = (Q(standee_allowed=True) & Q(flier_allowed=True))

                        if param == 'CDFL':
                            temp_q = (Q(car_display_allowed=True) & Q(flier_allowed=True))

                        if param == 'POSLFL':
                            temp_q = (Q(poster_allowed_nb=True) & Q(stall_allowed=True) & Q(flier_allowed=True))

                        if param == 'STSLFL':
                            temp_q = (Q(stall_allowed=True) & Q(standee_allowed=True) & Q(flier_allowed=True))

                        if param == 'POCDFL':
                            temp_q = (Q(poster_allowed_nb=True) & Q(car_display_allowed=True) & Q(flier_allowed=True))

                        if param == 'STCDFL':
                            temp_q = (Q(standee_allowed=True) & Q(car_display_allowed=True) & Q(flier_allowed=True))

                    else:
                        temp_q = Q(**{"%s" % inventory_dict[param]: 'True'})

                    if temp:

                        p = (p | temp_q)
                    else:
                        p = temp_q
                        temp = 1
                except KeyError:
                    pass

        # code changes & added for 'OR' filters of inventory and 'AND' filters with inventory and others
        # code start
        if p:
            filter_query &= p
        return Response({'status': True, 'data': filter_query}, status=status.HTTP_200_OK)

    except Exception as e:
        return Response({'status': False, 'error': e.message}, status=status.HTTP_400_BAD_REQUEST)


def return_price(adinventory_type_dict, duration_type_dict, inv_type, dur_type):
    price_mapping = PriceMappingDefault.objects.filter(adinventory_type=adinventory_type_dict[inv_type],
                                                       duration_type=duration_type_dict[dur_type])
    if price_mapping:
        return price_mapping[0].actual_supplier_price
    return 0


def get_min_max_lat_long(delta_dict):
    """
    :param data:
    :return:
    """
    # calculating for latitude and longitude
    # this is done to ensure that the suppliers are sent according to the current radius and center in frontend
    # user can change center and radius and it will not be stored in the database until saved
    if not delta_dict.get('latitude') or not delta_dict.get('longitude') or not delta_dict.get('supplier_type_code'):
        return Response({'status': False, 'error': 'provide lat/long/supplier_type_code'},
                        status=status.HTTP_400_BAD_REQUEST)

    latitude = delta_dict.get('latitude')
    longitude = delta_dict.get('longitude')
    supplier_code = delta_dict.get('supplier_type_code')

    try:
        latitude = float(latitude)
        longitude = float(longitude)

        data = {}

        data['max_latitude'] = latitude + delta_dict['delta_latitude']
        data['min_latitude'] = latitude - delta_dict['delta_latitude']

        delta_longitude = delta_dict['delta_longitude']
        data['max_longitude'] = longitude + delta_dict['delta_longitude']
        data['min_longitude'] = longitude - delta_dict['delta_longitude']
        data['supplier_type_code'] = supplier_code
        return Response({'status': True, 'data': data}, status=status.HTTP_200_OK)
    except Exception as e:
        return Response({'status': False, 'error': e.message}, status=status.HTTP_400_BAD_REQUEST)


def get_delta_latitude_longitude(radius, latitude):
    delta_longitude = radius / (111.320 * math.cos(math.radians(latitude)))
    delta_latitude = radius / 110.574

    return {'delta_latitude': delta_latitude,
            'delta_longitude': delta_longitude}


def space_on_circle(latitude, longitude, radius, space_lat, space_lng):
    """
    Args:
        latitude: latitude of the center
        longitude: longitude of the center
        radius: radius of center
        space_lat: supplier latitude
        space_lng: supplier longitude

    Returns: weather the supplier coordinates actually lie within a circle drawn with  given radius  ? The center
    coordinates being latitude, longitude in the params. The function uses simple pythagoras theorem to test this.
    """
    # if any of the param is False, we return False.
    if not latitude or (not longitude) or (not radius) or (not space_lat) or (not space_lng):
        return False
    return (space_lat - latitude) ** 2 + (space_lng - longitude) ** 2 <= (radius / 110.574) ** 2


def initialize_keys(center_object, supplier_type_code):
    """
    Args:
        center_object: a dict representing one center object
    goal is to make dict request.data type
    Returns: intializes this dict  with all valid keys and defaults as values

    """
    function = initialize_keys.__name__
    try:
        if not center_object:
            # set center key
            center_object['center'] = {}
            # set suppliers dict
            center_object['suppliers'] = {supplier_type_code: []}

            center_object['shortlisted_inventory_details'] = []
        else:
            # if center_object does exist, it can happen that it does not contain entry for this
            # supplier_type_code
            if not center_object['suppliers'].get(supplier_type_code):
                center_object['suppliers'][supplier_type_code] = []
            # if center_object does exist, it can happen that shortlisted_inventory_details does not
            # exist
            if not center_object.get('shortlisted_inventory_details'):
                center_object['shortlisted_inventory_details'] = []
        return center_object
    except Exception as e:
        raise Exception(function, ui_utils.get_system_error(e))


def make_societies_inventory(center_object, row):
    """
    Args:
        center_object: a center object
        row: 1 row of sheet
    Returns: adds data for 'societies_inventory'

    """
    try:
        center_object['societies_inventory']['supplier_code'] = 'RS'  # hardcoded for society
        center_object['societies_inventory']['id'] = row['inventory_type_id']
        center_object['societies_inventory']['space_mapping'] = row['space_mapping_id']

        return Response({'status': True, 'data': center_object}, status=status.HTTP_200_OK)
    except Exception as e:
        return Response({'status': False, 'error': e.message}, status=status.HTTP_400_BAD_REQUEST)


def make_shortlisted_inventory_list(row, supplier_type_code, proposal_id, center_id):
    """
    Args:
        row: a single row of sheet.  it's dict

    Returns: a list of shortlisted inventory types for this supplier

    """
    function = make_shortlisted_inventory_list.__name__
    try:
        shortlisted_inventory_list = []
        # use the filter table to pull out those inventories which were selected.
        filter_codes = Filters.objects.filter(proposal_id=proposal_id, center_id=center_id,
                                              filter_name='inventory_type_selected',
                                              supplier_type_code=supplier_type_code).values_list('filter_code',
                                                                                                 flat=True)
        unique_inventory_codes = get_unique_inventory_codes(filter_codes)
        # check for predefined keys in the row. if available, we have that inventory !
        for code in unique_inventory_codes:
            inv_name = v0_constants.inventory_code_to_name[code]
            # get inventory base name
            base_name = join_with_underscores(v0_constants.inventory_code_to_name[code].lower())

            # get type and duration.  it's fixed for now
            inv_type = v0_constants.inventory_type_duration_dict_list[code][1]
            inv_duration = v0_constants.inventory_type_duration_dict_list[code][2]

            # inv_type = v0_constants.inventory_type_duration_dict[code]['type_duration'][0]['type']
            # inv_type = v0_constants.type_dict[inv_type]
            #
            # # get duration. it's fixed for now
            # inv_duration = v0_constants.inventory_type_duration_dict[code]['type_duration'][0]['duration']
            # inv_duration = v0_constants.duration_dict[inv_duration]

            # the inventory ids are added later. but why not here depending upon the count ?
            shortlisted_inventory_details = {
                'proposal_id': proposal_id,
                'center_id': center_id,
                'supplier_id': row['supplier_id'],
                'supplier_type_code': supplier_type_code,
                'inventory_price': row[base_name + '_business_price'],
                'inventory_count': get_default_inventory_count(inv_name, row['tower_count']),
                'type': inv_type,
                'duration': inv_duration,
                'inventory_name': v0_constants.inventory_code_to_name[code].upper(),
                'factor': row[base_name + '_price_factor']
            }

            # add it to the list
            shortlisted_inventory_list.append(shortlisted_inventory_details)
        return shortlisted_inventory_list
    except KeyError as e:
        raise KeyError(function, ui_utils.get_system_error(e))
    except Exception as e:
        raise Exception(function, ui_utils.get_system_error(e))


def get_default_inventory_count(inventory_name, tower_count):
    """

    :param inventory_name:
    :param tower_count:
    :return:
    """

    function = get_default_inventory_count.__name__
    try:

        if inventory_name.lower() == v0_constants.poster.lower() or inventory_name.lower() == v0_constants.standee.lower():
            return tower_count
        else:
            return 1
    except Exception as e:
        raise Exception(function, ui_utils.get_system_error(e))


def make_suppliers(center_object, row, supplier_type_code, proposal_id, center_id):
    """
    Args:
        center_object: a center_object
        row: 1 row of data in sheet

    Returns: adds the society that is present in this row to center_object and returns

    """
    function = make_suppliers.__name__
    try:
        # collect society data in a dict and append it to the list of societies of center_object
        supplier_header_keys = ['_'.join(header.split(' ')) for header in
                                v0_constants.inventorylist[supplier_type_code]['HEADER']]
        supplier_header_keys = [header.lower() for header in supplier_header_keys]
        supplier_data_keys = v0_constants.inventorylist[supplier_type_code]['DATA']
        supplier = {data_key: row[header] for header, data_key in zip(supplier_header_keys, supplier_data_keys)}

        # get the list of shortlisted inventory details
        shortlisted_inventory_list = make_shortlisted_inventory_list(row, supplier_type_code, proposal_id, center_id)

        # add it to the list of center_object
        center_object['suppliers'][supplier_type_code].append(supplier)

        # add shortlisted_inventory_list to the list already initialized
        center_object['shortlisted_inventory_details'].extend(shortlisted_inventory_list)

        # return the result after we are done scanning
        return center_object
    except Exception as e:
        raise Exception(function, ui_utils.get_system_error(e))


def populate_shortlisted_inventory_pricing_details(result, proposal_id, user):
    """
    Args:
        result: it's a list containing final data
        proposal_id: The proposal_id

    Returns: success if it's able to create objects and map inventory_ids in the list else failure
    """
    function = populate_shortlisted_inventory_pricing_details.__name__
    try:

        center_ids = list(result.keys())
        # this creates a mapping like { 1: 'center_object_1', 2: 'center_object_2' } etc
        center_objects = ProposalCenterMapping.objects.in_bulk(center_ids)
        proposal_object = ProposalInfo.objects.get_permission(user=user, proposal_id=proposal_id)

        # set to hold all durations
        duration_list = set()
        # set to hold inventory names
        inventory_names = set()
        # set to hold all inventory_types
        inventory_types = set()

        for center_id, center in result.items():
            for index, shortlisted_inventory_detail in enumerate(center['shortlisted_inventory_details']):
                duration_list.add(shortlisted_inventory_detail['duration'])
                inventory_names.add(shortlisted_inventory_detail['inventory_name'])
                inventory_types.add(shortlisted_inventory_detail['type'])

        # fetch all ad_inventory_type objects
        ad_inventory_type_objects = AdInventoryType.objects.filter(adinventory_name__in=inventory_names).filter(
            adinventory_type__in=inventory_types)
        # fetch all duration objects
        durations_objects = DurationType.objects.filter(duration_name__in=duration_list)

        # return error if atleast one of them is False
        if not ad_inventory_type_objects or not durations_objects:
            return ui_utils.handle_response(function, data='No ad_inventory_objects or duration objects.', success=True)

        # create a mapping like {'POSTER':{ 'A3' : ad_inv_object },'STANDEE': {'small':ad_inv_object  } }
        ad_inventory_type_objects_mapping = {}
        for ad_inventory_type_object in ad_inventory_type_objects:
            # example, 'POSTER', 'STANDEE', 'CAR DISPLAY' becomes 'CAR_DISPLAY'
            inv_name = ad_inventory_type_object.adinventory_name
            # example, 'A3', 'Small', 'High' etc
            inv_type = ad_inventory_type_object.adinventory_type

            if not ad_inventory_type_objects_mapping.get(inv_name):
                ad_inventory_type_objects_mapping[inv_name] = {}
            # key with inv_name will always be there at this  line
            if not ad_inventory_type_objects_mapping[inv_name].get(inv_type):
                ad_inventory_type_objects_mapping[inv_name][inv_type] = ad_inventory_type_object

        # create a mapping like { 'weekly' : duration_object,  } by this loop
        duration_mapping = {duration_object.duration_name: duration_object for duration_object in durations_objects}

        # this object holds the data that is to be added in the shortlisted_inventory_detail table
        # list to store ShortlistedInventoryPricingDetails objects
        output = []

        supplier_to_details_map = {}

        supplier_type_codes = set()

        for center_id, center in result.items():

            # map the supplier_related detail like status first .
            for supplier_type_code, supplier_detail in center['suppliers'].items():
                for supplier in supplier_detail:
                    supplier_id = supplier['supplier_id']
                    supplier_status = supplier['status']
                    supplier_to_details_map[supplier_type_code, supplier_id] = supplier_status

            # now map the inventory details
            for index, shortlisted_inventory_detail in enumerate(center['shortlisted_inventory_details']):

                # this dict is filled up with data and collected into output array which is sent for assignment
                shortlisted_inventory_detail_object = {
                    'user': user,
                    'center': center_id
                }

                # copy supplier_id, inventory_price, inventory_count as it is from the current object
                for key in v0_constants.shortlisted_inventory_pricing_keys:
                    shortlisted_inventory_detail_object[key] = shortlisted_inventory_detail[key]

                supplier_type_code = shortlisted_inventory_detail['supplier_type_code']
                supplier_id = shortlisted_inventory_detail['supplier_id']

                # fetch the inventory_name, type, duration. It will be used to fetch correct ad_inventory_type
                # objects from the mapping.
                inventory_name = shortlisted_inventory_detail['inventory_name']
                ad_inventory_type = shortlisted_inventory_detail['type']
                duration = shortlisted_inventory_detail['duration']
                # fetch the right ad_inventory_type object from the mapping created earlier
                shortlisted_inventory_detail_object['ad_inventory_type'] = \
                ad_inventory_type_objects_mapping[inventory_name][ad_inventory_type]
                # fetch the right duration type object created earlier
                shortlisted_inventory_detail_object['ad_inventory_duration'] = duration_mapping[duration]
                shortlisted_inventory_detail_object['inventory_name'] = inventory_name
                shortlisted_inventory_detail_object['status'] = supplier_to_details_map[supplier_type_code, supplier_id]

                supplier_type_codes.add(supplier_type_code)
                output.append(shortlisted_inventory_detail_object)

        # make the inventories assigned to suppliers
        message = make_inventory_assignments(proposal_id, output, supplier_type_codes)

        # update the data in sipd instances created for this proposal.
        update_shortlisted_inventory_pricing_data(proposal_id, output)

        return message
    except Exception as e:
        return ui_utils.handle_response(function, exception_object=e)


def update_shortlisted_inventory_pricing_data(proposal_id, output):
    """

    Args:
        proposal_id:
        output:

    Returns: after the inventories have been assigned, next job is to save the  things like business_price, factor of each inventory.

    """
    function = update_shortlisted_inventory_pricing_data.__name__
    try:
        shortlisted_inventory_pricing_instances = ShortlistedInventoryPricingDetails.objects.filter(
            shortlisted_spaces__proposal_id=proposal_id)
        # get the distinct inventoy names
        inventory_names = set([detail['inventory_name'] for detail in output])

        # get the content types
        response = ui_utils.get_content_types(inventory_names)
        if not response.data['status']:
            raise Exception(function, response.data['data'])
        inventory_content_type_map = response.data['data']

        inventory_content_type_name_map = {}

        # reverse map content_type ---> name. Required to make the sipd_instances_per_inventory_map dict
        for code, inv_content_type in inventory_content_type_map.items():
            inventory_content_type_name_map[inv_content_type] = code

        # create a map of (supplier_id, inventory_name) ---> [ list of sipd objects ]. All such objects will have same factor and business price as of now.
        sipd_instances_per_inventory_map = {}
        for instance in shortlisted_inventory_pricing_instances:
            inventory_name = inventory_content_type_name_map[instance.inventory_content_type]
            key = (instance.shortlisted_spaces.object_id, v0_constants.inventory_code_to_name[inventory_name])
            try:
                reference = sipd_instances_per_inventory_map[key]
                reference.append(instance)
            except KeyError:
                sipd_instances_per_inventory_map[key] = []
                sipd_instances_per_inventory_map[key].append(instance)

        # create a (supplier_id, inventory_name) -->  { factor, buisness_price } map for output list. This combination will
        # yeild a dict. because currently inventory type, duration is fixed, inventory name implies a default type and duration.
        output_map = {}
        for detail in output:
            key = (detail['supplier_id'], detail['inventory_name'])
            try:
                reference = output_map[key]
                if reference:
                    raise Exception(
                        'Duplicate supplier present in sheet. Kindly fix that and re upload the sheet. The duplicate supplier id is {0}'.format(
                            detail['supplier_id']))
            except KeyError:
                output_map[key] = {
                    'factor': detail['factor'],
                    'inventory_price': detail['inventory_price']
                }
        # container for holding sipd instances
        sipd_instances_updated = []
        for key, sipd_instances in sipd_instances_per_inventory_map.items():
            try:
                data = output_map[key]
                for instance in sipd_instances:
                    instance.factor = data['factor']
                    instance.inventory_price = data['inventory_price']
                    sipd_instances_updated.append(instance)
            except KeyError:
                # this means that this supplier was not assigned this inventory. It can be the case if the supplier does not allow that kind of inventory.
                pass
        bulk_update(sipd_instances_updated)
    except Exception as e:
        raise Exception(function, ui_utils.get_system_error(e))


def get_center_id_list(ws, index_of_center_id):
    """
    Args:
        ws: instance of worksheet
        index_of_center_id: index of center_id column in the sheet
    Returns: a list containing unique center_id's

    """
    function = get_center_id_list.__name__
    try:
        center_id_set = set()
        for index, row in enumerate(ws.iter_rows()):
            # skip the headers
            if index == 0:
                continue
            if row[index_of_center_id].value:
                center_id_set.add(int(row[index_of_center_id].value))
        return ui_utils.handle_response(function, data=list(center_id_set), success=True)
    except Exception as e:
        return ui_utils.handle_response(function, exception_object=e)


def get_headers(ws):
    """
    Args:
        ws: worksheet
    Returns: a list of headers

    """
    try:
        headers = []
        for index, row in enumerate(ws.rows[0]):
            headers[index] = row[index]
        return Response({'status': True, 'data': headers}, status=status.HTTP_200_OK)
    except Exception as e:
        return Response({'status': False, 'error': e.message}, status=status.HTTP_400_BAD_REQUEST)


def get_mapped_row(ws, row):
    """
    This makes a dict having headings with keys and corresponding cell values as values.
    Args:
        ws: worksheet instance
        row_index: the row index  for which the modified row with underscore keys needs to be made

    Returns: a dict in which keys are keys which are mapped to headers and value is the corresponding value from the row

    """
    try:
        row_dict = {}
        for index, cell in enumerate(row):
            heading_cell = ws.cell(row=1, column=index + 1)
            if heading_cell.value:
                heading_with_spaces = heading_cell.value.lower()
                heading_with_underscore = '_'.join(heading_with_spaces.split(' '))
                row_dict[heading_with_underscore] = cell.value
        return Response({'status': True, 'data': row_dict}, status=status.HTTP_200_OK)
    except Exception as e:
        return Response({'status': False, 'error': e.message}, status=status.HTTP_400_BAD_REQUEST)


def make_center(center_object, row):
    """
    Args:
        center_object: center_object
        row: a dict containg data of one row

    Returns: populates the center_object with data for it's 'center' key

    """
    # get the fields which also needs to sent for the other api to work upon because they are required for serializing .
    # city, area, pincode, sub_area, latitude, longitude, radius = ProposalCenterMapping.objects.values_list('city', 'area', 'pincode', 'subarea', 'latitude', 'longitude', 'radius').get(id=row['center_id'])
    function = make_center.__name__
    try:
        center_object['center']['center_name'] = row.get('center_name')
        center_object['center']['proposal'] = row['proposal']
        center_object['center']['id'] = row['center_id']
        # center_object['center']['city'] = city
        # center_object['center']['area'] = area
        # center_object['center']['subarea'] = sub_area
        # center_object['center']['pincode'] = pincode
        # center_object['center']['latitude'] = latitude
        # center_object['center']['longitude'] = longitude
        # center_object['center']['radius'] = radius
        # space_mapping_response = make_space_mappings(row)
        # if not space_mapping_response.data['status']:
        #     return space_mapping_response
        # center_object['center']['space_mappings'] = space_mapping_response.data['data']
        return center_object
    except Exception as e:
        raise Exception(function, ui_utils.get_system_error(e))


def make_space_mappings(row):
    """
    Args:
        row: data of one row

    Returns: a dict of space mappings

    """
    try:
        space_mapping = {}
        space_mapping['center'] = row['center_id']
        space_mapping['proposal'] = row['center_proposal_id']
        space_mapping['id'] = row['space_mapping_id']
        return Response({'status': True, 'data': space_mapping}, status=status.HTTP_200_OK)
    except Exception as e:
        return Response({'status': False, 'error': e.message}, status=status.HTTP_400_BAD_REQUEST)


def save_proposal_version(center_id):
    """
    Args:
        center_id: center_id for which proposal version object is to be saved
         used in FinalProposal post request

    Returns:

    """

    try:
        center_object = ProposalCenterMapping.objects.select_related('proposal').get(id=center_id)
        proposal_object = center_object.proposal

        # version save
        proposal_version_object = ProposalInfoVersion(proposal=proposal_object, name=proposal_object.name,
                                                      payment_status=proposal_object.payment_status, \
                                                      created_on=proposal_object.created_on,
                                                      created_by=proposal_object.created_by,
                                                      tentative_cost=proposal_object.tentative_cost, \
                                                      tentative_start_date=proposal_object.tentative_start_date,
                                                      tentative_end_date=proposal_object.tentative_end_date)
        proposal_version_object.save()
        return Response({'status': True, 'data': proposal_version_object}, status=status.HTTP_200_OK)
    except Exception as e:
        return Response({'status': False, 'error': e.message}, status=status.HTTP_400_BAD_REQUEST)


def get_model_name(model_name):
    """
    Args:
        model_name: a model name in lower case joined with underscores

    Returns: removes underscores and joins on Intials with a capital letter

    """
    return ' '.join(model_name.split('_')).title()


def initialize_master_data(master_data):
    """
    Args:
        master_data: empty dict for import proposal cost data api

    Returns: initialized according to needs

    """
    # only one object exist per file hence a dict
    master_data['ideation_design_cost'] = {}

    # only one object exist per file hence a dict
    master_data['logistic_operations_cost'] = {}

    #  one object exist per supplier or row  hence a list
    master_data['space_booking_cost'] = []

    # only one object exist per file hence a dict
    master_data['event_staffing_cost'] = {}

    # only one object exist per file hence a dict
    master_data['data_sciences_cost'] = {}

    # only one object exist per file hence a dict
    master_data['printing_cost'] = {}

    # only one object exist per file hence a dict
    master_data['proposal_master_cost'] = {}

    # one object exist per supplier or row hence a list because there are multiple suppliers
    master_data['proposal_metrics'] = []

    return master_data


def handle_offline_pricing_row(row, master_data):
    """
    Args:
        row: a row of offline pricing sheet

    Returns: saves the data in the right model after processing the row

    """
    try:
        # this string is the base string against we perform the matching and determine to which model the
        #  row data belongs. example  'Use of Data Sciences to Identify Targeted Spaces'
        target_string = row[0].value

        # for all models and their contents declared in constants
        for model, model_content in v0_constants.offline_pricing_data.items():
            # because model names are joined by _, join them without underscore with first letter as capital
            # this is required for apps.get_model() to work. model_content is a list. each item of this list
            # will map to a row in db

            # for all types of row in model_content
            for model_row in model_content:

                # find the pattern which is predefined for the this model_row
                pattern = model_row['match_term']

                # match it
                match = re.search(pattern, target_string, re.I)

                # no point in going further if there is no match
                if not match:
                    continue
                # get the column name
                column_name = model_row['col_name']

                # get the value for that column
                value = row[v0_constants.value_index].value

                # set data to previously saved data dict if it's a dict because here it will be updated
                # else set to an empty dict because it will be appended in the list
                data = master_data[model] if type(master_data[model]) == dict else {}

                # add the content type information to data if 'specific' is not none
                if model_row.get('specific'):
                    supplier_content_type_response = ui_utils.get_content_type(model_row.get('specific')['code'])
                    if not supplier_content_type_response.data['status']:
                        return supplier_content_type_response
                    supplier_content_type = supplier_content_type_response.data['data']
                    data['supplier_type'] = supplier_content_type.id

                # add comment information to data if comment is not none. index for comment col is defined in constants
                if row[v0_constants.comment_index].value:
                    data['comment'] = row[v0_constants.comment_index].value

                # add the specific column and value
                data[column_name] = value

                # add metric_name and value of the metric in case it's a metric model
                if model == v0_constants.metric_model:
                    data['metric_name'] = model_row.get('match_term')
                    data['value'] = value
                # we will not save here. we will collect the data in master_data
                # dict and save later. This is because multiple rows represent one object in some cases and  many objects
                # in other cases so saving on row by row basis cannot be done.
                insert_master_data_response = insert_into_master_data(data, master_data, model)
                if not insert_master_data_response.data['status']:
                    return insert_master_data_response
                master_data = insert_master_data_response.data['data']

        return ui_utils.handle_response(handle_offline_pricing_row.__name__, data=master_data, success=True)
    except KeyError as e:
        return ui_utils.handle_response(handle_offline_pricing_row.__name__, data='Key error', exception_object=e)
    except Exception as e:
        return ui_utils.handle_response(handle_offline_pricing_row.__name__, exception_object=e)


def insert_into_master_data(data, master_data, model_name):
    """
    Args:
        data: data is dict containing data of one row
        master_data: master data into which data can be assigned or appended depending upon what model_name is
        This function is made because 'data' can either be assigned or appended in master_data depending upon
        for which model you are collecting.

    Returns: master_data with modified content

    """
    try:
        # append the data to list
        if type(master_data[model_name]) == list:
            master_data[model_name].append(data)
        else:
            # update the existing dict with new data
            master_data[model_name] = data
        return ui_utils.handle_response(insert_into_master_data.__name__, data=master_data, success=True)
    except Exception as e:
        return ui_utils.handle_response(handle_offline_pricing_row.__name__, exception_object=e)


def save_master_data(master_data):
    """

    Args:
        master_data: the data to be saved in the models looks like
        {
           'data_sciences_cost':{
              'total_cost':40
           },
           'ideation_design_cost':{
              'total_cost':40
           },
           'logistic_operations_cost':{
              'total_cost':40
           },
           'space_booking_cost':[
              {
                 'total_cost':40
                 'supplier_type':46
              },
              {
                 'total_cost':40
                 'supplier_type':45
              }
           ],
           'event_staffing_cost':{
              'total_cost':40      L
           },
           'proposal_master_cost':{
              'discount':40 ,
              'total_cost':40 ,
              'tax':40 ,
              'total_impressions':50,
              'average_cost_per_impression':50 ,
              'agency_cost':40,
              'proposal':u'jogXxEUu'
           },
           'printing_cost':{
              'total_cost':40
           },
           'proposal_metrics':[
              {
                 'supplier_type':45,
                 'metric_name':'Average Cost per Society',
                 'value':40
              },
              {
                 'supplier_type':48,
                 'metric_name':'Average Cost per Salon',
                 'value':40
              },
            ]
   }
    Returns: success if models are saved successfully

    """
    function = save_master_data.__name__
    try:
        with transaction.atomic():
            # save into proposal master cost table first because it's id is required later
            serializer = ui_utils.get_serializer('proposal_master_cost')(data=master_data['proposal_master_cost'])
            if serializer.is_valid():
                serializer.save()
                proposal_master_cost_id = serializer.data['id']
            else:
                return ui_utils.handle_response(function, data=serializer.errors)

            # pull the model names to process except proposal_master_cost which has been done
            model_name_list = list(v0_constants.offline_pricing_data.keys())
            model_name_list.remove('proposal_master_cost')

            # get the right serializer and save the data
            for model in model_name_list:
                # differentiate if we have list to save or a dict to save
                if model in v0_constants.one_obect_models:
                    # need to set id of proposal master cost as according to model structure
                    master_data[model]['proposal_master_cost'] = proposal_master_cost_id
                    serializer = ui_utils.get_serializer(model)(data=master_data[model])
                else:
                    # need to set id of proposal master cost as according to model structure
                    for data in master_data[model]:
                        data['proposal_master_cost'] = proposal_master_cost_id
                    serializer = ui_utils.get_serializer(model)(data=master_data[model], many=True)

                if serializer.is_valid():
                    serializer.save()
                else:
                    return ui_utils.handle_response(save_master_data.__name__, data=serializer.errors)
        return ui_utils.handle_response(save_master_data.__name__, data='success', success=True)
    except Exception as e:
        return ui_utils.handle_response(save_master_data.__name__, exception_object=e)


def is_empty_row(row):
    """
    Args:
        row: row to be checked if empty or not
          row is row read by openpyxl lib

    Returns: true if empty else False

    """
    for x in row:
        # if x has a value, row is not empty
        if x.value:
            return False
    # return True, row is empty
    return True


def delete_proposal_cost_data(proposal_id):
    """
    Deletes the tables data associated with proposal cost each time the api is hit because we don't want
    duplicates.
    """
    function = delete_proposal_cost_data.__name__
    try:
        proposal_master_cost_object = get_object_or_404(ProposalMasterCost, proposal__proposal_id=proposal_id)
        ProposalMetrics.objects.filter(proposal_master_cost=proposal_master_cost_object).delete()
        PrintingCost.objects.filter(proposal_master_cost=proposal_master_cost_object).delete()
        DataSciencesCost.objects.filter(proposal_master_cost=proposal_master_cost_object).delete()
        EventStaffingCost.objects.filter(proposal_master_cost=proposal_master_cost_object).delete()
        SpaceBookingCost.objects.filter(proposal_master_cost=proposal_master_cost_object).delete()
        IdeationDesignCost.objects.filter(proposal_master_cost=proposal_master_cost_object).delete()
        LogisticOperationsCost.objects.filter(proposal_master_cost=proposal_master_cost_object).delete()
        proposal_master_cost_object.delete()
        return ui_utils.handle_response(function, data='success', success=True)
    except Http404:
        # first time call of this function will not fetch object, but must be a success return.
        return ui_utils.handle_response(function, data='', success=True)
    except Exception as e:
        return ui_utils.handle_response(function, exception_object=e)


def create_basic_proposal(data):
    """

    Args:
        data: data for ProposalInfo

    Returns: success if proposalInfo  is created

    """
    function_name = create_basic_proposal.__name__
    try:
        proposal_serializer = ProposalInfoSerializer(data=data)
        if proposal_serializer.is_valid():
            proposal_serializer.save()
        else:
            return ui_utils.handle_response(function_name, data=proposal_serializer.errors)
        return ui_utils.handle_response(function_name, data=proposal_serializer.data, success=True)
    except Exception as e:
        return ui_utils.handle_response(function_name, exception_object=e)


def get_geo_object_lat_long(address):
    """

    Args:
        address: given an address in string, return it's lat, long in a tuple form.
    Returns: Function tries three variations of the address and returns wherever it finds a valid lat, long.

    """
    function_name = get_geo_object_lat_long.__name__
    try:
        geocoder = Geocoder(api_key='AIzaSyCy_uR_SVnzgxCQTw1TS6CYbBTQEbf6jOY')
        # geo_object = geocoder.geocode(address)
        # split the address on comma
        address_parts = address.split(',')
        # get the length
        length = len(address_parts)
        # define upto what indexes you want to calculate addressess.
        indexes = [length, length - 1, length - 2]
        geo_object = None
        for index in indexes:
            # get this address
            address = ','.join(part for part in address_parts[:index])
            # try to get geo_object
            geo_object = geocoder.geocode(address)
            if not geo_object.data[0]['geometry']:
                continue
            elif not geo_object.data[0]['geometry']['location']:
                continue
            elif not geo_object.data[0]['geometry']['location']['lat']:
                continue
            else:
                break
        # if found, return lat, long
        if geo_object:
            latitude = geo_object.data[0]['geometry']['location']['lat'],
            longitude = geo_object.data[0]['geometry']['location']['lng'],
            latitude = latitude[0] if len(latitude) > 0 else latitude
            longitude = longitude[0] if len(longitude) > 0 else longitude
            return ui_utils.handle_response(function_name, data=(latitude, longitude), success=True)
        else:
            # return right error.
            return ui_utils.handle_response(function_name,
                                            data='no geo_object found even after three variations of the address')
    except GeocoderError as e:
        return ui_utils.handle_response(function_name, exception_object=e)
    except Exception as e:
        return ui_utils.handle_response(function_name, exception_object=e)


def save_suppliers_allowed(center_info, proposal_id, center_id, user):
    """
    Args:
        center_info: One center data
        proposal_id: The proposal_id
        center_id: The newly created center_id
        user: The User instance

    Returns: saves what suppliers are allowed in each center. and returns on success.
    """
    function_name = save_suppliers_allowed.__name__

    try:
        # fetch all the supplier codes. 'supplier_codes'
        suppliers_codes = center_info['center']['codes']
        # for all the codes
        for code in suppliers_codes:
            content_type_response = ui_utils.get_content_type(code)
            if not content_type_response.data['status']:
                return content_type_response
            content_type = content_type_response.data['data']
            # prepare the data
            data = {
                'proposal_id': proposal_id,
                'center_id': center_id,
                'supplier_content_type': content_type,
                'supplier_type_code': code,
                'user': user
            }
            ProposalCenterSuppliers.objects.get_or_create(**data)
        # return success if done
        return ui_utils.handle_response(function_name, data='success', success=True)
    except KeyError as e:
        return ui_utils.handle_response(function_name, exception_object=e)
    except Exception as e:
        return ui_utils.handle_response(function_name, exception_object=e)


def calculate_address(center):
    """
    Args:
        center: makes address for this center
    Returns: address for this center

    """
    function_name = calculate_address.__name__
    try:
        address = center['address'] + "," + center['subarea'] + ',' + center['area'] + ',' + center['city'] + ',' + \
                  center['pincode']
        return ui_utils.handle_response(function_name, data=address, success=True)

    except Exception as e:
        return ui_utils.handle_response(function_name, exception_object=e)


def save_center_data(proposal_data, user):
    """

    Args:
        proposal_data: a proposal_data with proposal_id as key must.
        user: User instance

    Returns: center data

    """
    function_name = save_center_data.__name__
    try:
        # get the proposal_id
        proposal_id = proposal_data['proposal_id']
        with transaction.atomic():
            # for all centers
            for center_info in proposal_data['centers']:
                center = center_info['center']
                # prepare center info
                center['proposal'] = proposal_id
                # get address for this center. because address can contain a complicated logic in future, it's in separate
                # function
                address_response = calculate_address(center)
                if not address_response.data['data']:
                    return address_response
                address = address_response.data['data']

                # add lat long to center's data based on address calculated
                geo_response = get_geo_object_lat_long(address)
                if not geo_response.data['status']:
                    center['latitude'], center['longitude'] = [0, 0]
                else:
                    center['latitude'], center['longitude'] = geo_response.data['data']
                center['user'] = user.id
                # set pincode to zero if there isn't any
                if not center['pincode']:
                    center['pincode'] = 0


                if 'id' in center_info:
                    # means an existing center was updated
                    center_instance = ProposalCenterMapping.objects.get_permission(user=user, id=center_info['id'])
                    center_serializer = ProposalCenterMappingSerializer(center_instance, data=center)
                else:
                    # means we need to create new center
                    center_serializer = ProposalCenterMappingSerializer(data=center)

                # save center info
                if center_serializer.is_valid():
                    center_serializer.save()
                    # now save all the suppliers associated with this center
                    response = save_suppliers_allowed(center_info, proposal_id, center_serializer.data['id'], user)
                    if not response.data['status']:
                        return response
                else:
                    return ui_utils.handle_response(function_name, data=center_serializer.errors)
            return ui_utils.handle_response(function_name,
                                            data={'status': 'success', 'center_id': center_serializer.data['id']},
                                            success=True)
    except Exception as e:
        return ui_utils.handle_response(function_name, exception_object=e)



def save_shortlisted_suppliers(suppliers, fixed_data):
    """
    Args:
        suppliers: a list of suppliers
        fixed_data: The data that will not change.
        user: User instance

    Returns: saves the list of suppliers into shortlistedspaces model

    """
    function_name = save_shortlisted_suppliers.__name__
    try:
        center = fixed_data.get('center')
        proposal = fixed_data.get('proposal')
        code = fixed_data.get('supplier_code')
        content_type = fixed_data.get('content_type')

        shortlisted_suppliers = []

        count = 0
        for supplier in suppliers:
            # by default every supplier is shortlisted, hence we save everything.
            # if supplier['status'] == v0_constants.status:
            #     continue

            # make entry for campaign_status and phase for each supplier here itself.
            campaign_status = supplier['campaign_status'] if supplier.get('campaign_status') else supplier['status']
            phase = supplier['phase'] if supplier.get('phase') else ''

            # make the data to be saved in ShortListedSpaces
            data = {
                'content_type': content_type,
                'object_id': supplier['supplier_id'],
                'center': center,
                'proposal': proposal,
                'supplier_code': code,
                'status': supplier['status'],
                'campaign_status': campaign_status,
                'phase': phase
            }
            shortlisted_suppliers.append(ShortlistedSpaces(**data))
        return shortlisted_suppliers
    except Exception as e:
        raise Exception(function_name, ui_utils.get_system_error(e))


def fetch_final_proposal_data(proposal_data, unique_supplier_codes):
    """
    The request is in form:
        [
             {
                  center : { id : 1 , center_name: c1, ...   } ,
                  suppliers: {  'RS' : [ { 'supplier_type_code': 'RS', 'status': 'R', 'supplier_id' : '1'}, {...}, {...} ],  }
                  suppliers_meta: {
                                     'RS': { 'inventory_type_selected' : [ 'PO', 'POST', 'ST' ]  },
                                     'CP': { 'inventory_type_selected':  ['ST']
                  }
             }
        ]
    Args:
        proposal_data: data of the proposal
        proposal_id: proposal_id

    Returns: collects all suppliers in societies array and inserts them into the db

    """
    function_name = fetch_final_proposal_data.__name__
    try:

        # get the proposal_id
        proposal_id = proposal_data['proposal_id']

        # get the center id
        center_id = proposal_data['center']['id']

        # get the proposal object
        proposal = ProposalInfo.objects.get(proposal_id=proposal_id)

        # get the center object
        center = ProposalCenterMapping.objects.get(id=center_id)

        fixed_data = {
            'center': center,
            'proposal': proposal,
        }

        # list to store all filter objects
        filter_data = []

        total_shortlisted_suppliers_list = []

        for code in unique_supplier_codes:

            # get the list of suppliers
            suppliers = proposal_data['suppliers'][code] if proposal_data['suppliers'].get(code) else []

            # get the content type for this supplier
            content_type = ui_utils.fetch_content_type(code)

            fixed_data['supplier_code'] = code
            fixed_data['content_type'] = content_type

            # save data of shortlisted suppliers
            total_shortlisted_suppliers_list.extend(save_shortlisted_suppliers(suppliers, fixed_data))

            # fetch suppliers_meta dict if present
            suppliers_meta = proposal_data.get('suppliers_meta')
            # check if any filters available for this partcular supplier type
            if suppliers_meta and suppliers_meta.get(code):
                filter_data.extend(save_filter_data(suppliers_meta, fixed_data))

        return total_shortlisted_suppliers_list, filter_data
    except KeyError as e:
        raise KeyError(function_name, ui_utils.get_system_error(e))
    except ObjectDoesNotExist as e:
        raise ObjectDoesNotExist(function_name, ui_utils.get_system_error(e))
    except Exception as e:
        raise Exception(function_name, ui_utils.get_system_error(e))


def save_filter_data(suppliers_meta, fixed_data):
    """
    Args:
        inventories_selected: the list of inventories which are selected
        fixed_data: The data that is fixed data and will not change
        user: The User instance
    Returns:

    """
    function_name = save_filter_data.__name__
    try:
        center = fixed_data.get('center')
        proposal = fixed_data.get('proposal')
        code = fixed_data.get('supplier_code')
        content_type = fixed_data.get('content_type')

        # creating filter objects for each filter value selected
        selected_filters_list = []
        for filter_name in v0_constants.filter_type[code]:
            if suppliers_meta.get(code) and filter_name in list(suppliers_meta[code].keys()):
                for inventory_code in suppliers_meta[code][filter_name]:
                    # TO store employee_count by codes so easy to fetch
                    if filter_name == 'employee_count':
                        key = int(inventory_code['min'])
                        inventory_code = v0_constants.employee_count_codes[key]
                    data = {
                        'center': center,
                        'proposal': proposal,
                        'supplier_type': content_type,
                        'supplier_type_code': code,
                        'filter_name': filter_name,
                        'filter_code': inventory_code,
                        'is_checked': True,
                    }
                    filter_object = Filters(**data)
                    selected_filters_list.append(filter_object)
        return selected_filters_list
    except KeyError as e:
        raise KeyError(function_name, ui_utils.get_system_error(e))
    except ObjectDoesNotExist as e:
        raise ObjectDoesNotExist(function_name, ui_utils.get_system_error(e))
    except Exception as e:
        raise Exception(function_name, ui_utils.get_system_error(e))


def create_proposal_id(organisation_id, account_id):
    """
    Args:
        organisation_id: The organisation_id
        account_id:  The account id

    Returns: A unique proposal id

    """
    function = create_proposal_id.__name__
    try:
        if not organisation_id or not account_id:
            return ui_utils.handle_response(function, data='provide business and account ids')
        # get number of business letters to append
        business_letters = v0_constants.business_letters
        # get number of account letters to append
        account_letters = v0_constants.account_letters
        # make the proposal id.
        proposal_id = organisation_id[:business_letters].upper() + account_id[:account_letters].upper() + (
        str(uuid.uuid4())[-v0_constants.proposal_id_limit:])
        return proposal_id
    except Exception as e:
        raise Exception(function, ui_utils.get_system_error(e))


def get_coordinates(radius, latitude, longitude):
    """

    Args:
        delta_dict: radius, latitude

    Returns: min lat, max lat, min long, max long

    """
    function_name = get_coordinates.__name__
    try:
        radius = float(radius)
        latitude = float(latitude)
        longitude = float(longitude)

        delta_dict = get_delta_latitude_longitude(float(radius), float(latitude))

        delta_latitude = delta_dict['delta_latitude']
        min_latitude = latitude - delta_latitude
        max_latitude = latitude + delta_latitude

        delta_longitude = delta_dict['delta_longitude']
        min_longitude = longitude - delta_longitude
        max_longitude = longitude + delta_longitude
        data = {
            'min_lat': float(min_latitude),
            'min_long': float(min_longitude),
            'max_lat': float(max_latitude),
            'max_long': float(max_longitude)
        }
        return ui_utils.handle_response(function_name, data=data, success=True)
    except Exception as e:
        return ui_utils.handle_response(function_name, exception_object=e)


def build_query(min_max_data, supplier_type_code, proposal):
    """

    Args:
        min_max_data: data of min, max lat longs.
        supplier_type_code: based on what supplier type code is, it constructs a query which shall be used
        to filter objects.

    Returns: a Q object.
    """
    function_name = build_query.__name__

    try:
        if supplier_type_code == 'RS':
            q = Q(society_latitude__lt=min_max_data['max_lat']) & Q(society_latitude__gt=min_max_data['min_lat']) & Q(
                society_longitude__lt=min_max_data['max_long']) & Q(society_longitude__gt=min_max_data['min_long']) & Q(
                representative=proposal.principal_vendor)
        else:
            q = Q(latitude__lt=min_max_data['max_lat']) & Q(latitude__gt=min_max_data['min_lat']) & Q(
                longitude__lt=min_max_data['max_long']) & Q(longitude__gt=min_max_data['min_long']) & Q(
                representative=proposal.principal_vendor)

        return ui_utils.handle_response(function_name, data=q, success=True)
    except Exception as e:
        return ui_utils.handle_response(function_name, exception_object=e)


def get_suppliers(query, supplier_type_code, coordinates):
    """
    Args:
        query: The Q object
        supplier_type_code : 'RS', 'CP'
        coordinates: a dict having radius, latitude, longitude as keys

    Returns: all the suppliers.

    """
    function_name = get_suppliers.__name__
    try:
        radius = coordinates.get('radius', 0)
        latitude = coordinates.get('latitude', 0)
        longitude = coordinates.get('longitude', 0)

        # get the suppliers data within that radius

        if supplier_type_code == 'RS':
            supplier_objects = SupplierTypeSociety.objects.filter(query)
            serializer = SupplierTypeSocietySerializer(supplier_objects, many=True)

        else:
            search_query = Q()
            search_query &= Q(supplier_type=supplier_type_code)

            address_supplier = Prefetch('address_supplier',queryset=AddressMaster.objects.all())
            supplier_objects = SupplierMaster.objects.prefetch_related(address_supplier).filter(search_query).order_by('supplier_name')
            serializer = SupplierMasterSerializer(supplier_objects, many=True)

        # result to store final suppliers
        result = []
        for supplier in serializer.data:
            # replace all society specific keys with common supplier keys
            if supplier_type_code != "RS":
                address_supplier = supplier.get('address_supplier')
                if address_supplier:
                    address_supplier = supplier.get('address_supplier')
                    
                    supplier['address1'] = address_supplier['address1'] if address_supplier['address1'] else ''
                    supplier['address2'] = address_supplier['address2']
                    supplier['area'] = address_supplier['area']
                    supplier['subarea'] = address_supplier['subarea']
                    supplier['city'] = address_supplier['city']
                    supplier['state'] = address_supplier['state']
                    supplier['zipcode'] = address_supplier['zipcode']
                    supplier['latitude'] = address_supplier['latitude']
                    supplier['longitude'] = address_supplier['longitude']
                    supplier['name'] = supplier['supplier_name']

                    

            for society_key, actual_key in v0_constants.society_common_keys.items():
                if society_key in list(supplier.keys()):
                    value = supplier[society_key]
                    del supplier[society_key]
                    supplier[actual_key] = value

            if not coordinates:
                result.append(supplier)
                continue

            if space_on_circle(latitude, longitude, radius, supplier['latitude'], supplier['longitude']):
                supplier['shortlisted'] = True
                # set status= 'S' as suppliers are shortlisted initially.
                supplier['status'] = v0_constants.status
                result.append(dict(supplier))
        return ui_utils.handle_response(function_name, data=result, success=True)

    except Exception as e:
        return ui_utils.handle_response(function_name, exception_object=e)
        proposal_id = data['proposal_id']
        center_id = data['center_id']
        supplier_content_type = data['content_type']

        # get the filter's data
        filter_objects = Filters.objects.filter(proposal_id=proposal_id, center_id=center_id,
                                                supplier_type=supplier_content_type)
        filter_serializer = FiltersSerializer(filter_objects, many=True)
        return ui_utils.handle_response(function_name, data=filter_serializer.data, success=True)
    except Exception as e:
        return ui_utils.handle_response(function_name, exception_object=e)


def handle_single_center(center, result, proposal):
    """
    Args:
        center: One center data.
        result : a dict having an entry against each center_id
        user: User instance
    Returns:

    """
    function_name = handle_single_center.__name__
    try:
        center_data = {}
        center_data['center'] = center

        radius = float(center['radius'])
        latitude = float(center['latitude'])
        longitude = float(center['longitude'])

        response = get_coordinates(radius, latitude, longitude)
        if not response.data['status']:
            return response

        min_max_data = response.data['data']

        # make room for storing supplier data
        center_data['suppliers'] = {}

        # store data for each type of supplier
        for supplier_type_code in center['codes']:
            # make an entry against the code for storing the results
            center_data['suppliers'][supplier_type_code] = {}

            # make a query
            query_response = build_query(min_max_data, supplier_type_code, proposal)
            if not query_response.data['status']:
                return query_response
            query = query_response.data['data']

            # prepare coordinates
            coordinates = {
                'radius': radius,
                'latitude': latitude,
                'longitude': longitude
            }

            # get the suppliers data
            response = get_suppliers(query, supplier_type_code, coordinates)
            if not response.data['status']:
                return response
            suppliers_data = response.data['data']
            # set in the result
            center_data['suppliers'][supplier_type_code] = suppliers_data

        result[center['id']] = center_data
        return ui_utils.handle_response(function_name, data=result, success=True)
    except KeyError as e:
        return ui_utils.handle_response(function_name, data='Key Error occurred', exception_object=e)
    except Exception as e:
        return ui_utils.handle_response(function_name, exception_object=e)


def merge_two_dicts(x, y):
    '''Given two dicts, merge them into a new dict as a shallow copy.'''
    function = merge_two_dicts.__name__
    try:
        # update x with keys which are not in x.
        x = x.copy()
        x_keys = list(x.keys())
        for key, value in y.items():
            if key not in x_keys:
                x[key] = value
        return x
    except Exception as e:
        raise Exception(function, ui_utils.get_system_error(e))


def suppliers_within_radius(data):
    """
    Args:
        data: a dict containing proposal_id, center_id, radius, supplier_codes.
        Two kinds of data is prepared and sent as of now:
        suppliers within radius and filters data

    Returns: all the suppliers withing the radius defined.

    """
    function_name = suppliers_within_radius.__name__
    try:
        proposal_id = data['proposal_id']
        center_id = data['center_id']
        proposal = ProposalInfo.objects.get(proposal_id=proposal_id)
        organisation_name = proposal.account.organisation.name

        master_result = {
            # set the business_name
            'organisation_name': organisation_name,
            # space to store the suppliers
            'suppliers': []
        }

        # todo: think of better way of separating this logic. looks ugly right now
        if center_id:
            center_id = int(center_id)
            # the queries will change if center_id is provided because we want to process
            # for this center only.
            if not data['radius'] or not data['latitude'] or not data['longitude']:
                return ui_utils.handle_response(function_name, data='if giving center_id, give radius, lat, long too!')

            centers = ProposalCenterMapping.objects.filter(id=center_id)
            serializer = ProposalCenterMappingSerializer(centers, many=True)

            serializer.data[0]['radius'] = data['radius']
            serializer.data[0]['latitude'] = data['latitude']
            serializer.data[0]['longitude'] = data['longitude']

            proposal_center_objects = ProposalCenterSuppliers.objects.filter(center_id=center_id)
            supplier_type_codes_list = proposal_center_objects.select_related('center').values('center',
                                                                                               'supplier_type_code')

            # define center_id_list to be used later
            center_id_list = [center_id]
        else:
            proposal_center_objects = ProposalCenterSuppliers.objects.filter(proposal_id=proposal_id)
            supplier_type_codes_list = proposal_center_objects.select_related('center').values('center',
                                                                                               'supplier_type_code')

            # fetch the mapped centers. This centers    were saved when CreateInitialproposal was hit.
            center_id_list = [data['center'] for data in supplier_type_codes_list]

            # query the center objects
            centers = ProposalCenterMapping.objects.filter(proposal_id=proposal_id, id__in=center_id_list)
            # centers = ProposalCenterMapping.objects.filter(proposal_id=proposal_id, id__in=center_id_list)
            serializer = ProposalCenterMappingSerializer(centers, many=True)

        # if not center_id, then fetch all the centers. centers can be a list
        # we add an extra attribute for each center object we get. Thats called codes. codes contain a list
        # of supplier_type_codees  like RS, CP.

        supplier_codes_dict = {center['id']: set() for center in serializer.data}
        if not supplier_codes_dict:
            return ui_utils.handle_response(function_name,
                                            data='Not found any centers in database against {0}'.format(proposal_id))

        for data in supplier_type_codes_list:
            center_id = int(data['center'])
            code = data['supplier_type_code']
            supplier_codes_dict[center_id].add(code)

        for center in serializer.data:
            center['codes'] = list(supplier_codes_dict[center['id']])
        # collect suppliers_meta information
        response = add_filters(proposal_id, center_id_list)
        if not response.data['status']:
            return response
        filters_data = response.data['data']

        # prepare result dict
        result = {center_id: {} for center_id in center_id_list}
        for center in serializer.data:
            center = dict(center)
            response = handle_single_center(center, result, proposal)
            if not response.data['status']:
                return response
            result = response.data['data']
            # get filter data per center from previous result
            filter_data_per_center = filters_data[center_id]['suppliers_meta']
            # assign it back to right center information
            result[center_id]['suppliers_meta'] = filter_data_per_center

        master_result['suppliers'] = list(result.values())
        return ui_utils.handle_response(function_name, data=master_result, success=True)
    except KeyError as e:
        return ui_utils.handle_response(function_name, data='Key Error occurred', exception_object=e)
    except Exception as e:
        return ui_utils.handle_response(function_name, exception_object=e)


def child_proposals(data):
    """
    Args:
        data: a dict containing parent_id for which we have to fetch all children proposals data
    Returns: all the proposalInfo data that is children of the given proposal_id

    """
    function_name = child_proposals.__name__
    try:
        # fetch all children of proposal_id and return.
        parent = data['parent']
        user = data['user']
        account_id = data['account_id']
        proposal_children = ProposalInfo.objects.filter_permission(user=user)
        if account_id:
            proposal_children = proposal_children.filter(account_id=account_id)
        proposal_children = proposal_children.filter(parent=parent).order_by('-created_on')
        serializer = ProposalInfoSerializer(proposal_children, many=True)
        return serializer.data
    except Exception as e:
        raise Exception(function_name, ui_utils.get_system_error(e))


def construct_proposal_response(proposal_id):
    """
    Args:
        proposal_id: proposal_id for which response structure is built
        user: The request.user object
    appends a list called codes in each center object.

    Returns: constructs the data in required form to be sent back to API response
    """
    function_name = construct_proposal_response.__name__
    try:
        supplier_type_codes_list = ProposalCenterSuppliers.objects.filter(proposal_id=proposal_id).select_related(
            'center').values('center', 'supplier_type_code')
        # fetch the mapped centers. This centers were saved when CreateInitialProposal was hit.
        center_id_list = [data['center'] for data in supplier_type_codes_list]

        # query the center objects
        centers = ProposalCenterMapping.objects.filter(proposal_id=proposal_id, id__in=center_id_list)
        serializer = ProposalCenterMappingSerializer(centers, many=True)
        supplier_codes_dict = {center['id']: [] for center in serializer.data}
        for data in supplier_type_codes_list:
            center_id = data['center']
            code = data['supplier_type_code']
            supplier_codes_dict[center_id].append(code)

        for center in serializer.data:
            center['codes'] = supplier_codes_dict[center['id']]

        return ui_utils.handle_response(function_name, data=serializer.data, success=True)

    except Exception as e:
        return ui_utils.handle_response(function_name, exception_object=e)


def set_inventory_pricing(supplier_ids, supplier_type_code, inventory_summary_map, stats):
    """
    :param supplier_ids:
    :param supplier_type_code:
    :param inventory_summary_map:
    :return: returns a dict in which each supplier has inventory status of allowed and the price if allowed.
    """

    function = set_inventory_pricing.__name__
    try:
        # by default all inventory codes.
        chosen_inventory_codes = ['PO', 'ST', 'CD', 'SL', 'FL']
        content_type = ui_utils.fetch_content_type(supplier_type_code)
        price_mapping_default_map, ad_inventory_map, duration_map = make_handy_price_mapping_default_duration_adinventory_type(
            supplier_ids, content_type)
        suppliers_per_supplier_type_code = {}
        for supplier_id in supplier_ids:
            if not suppliers_per_supplier_type_code.get(supplier_id):
                suppliers_per_supplier_type_code[supplier_id] = {}
                for inv_code in chosen_inventory_codes:
                    suppliers_per_supplier_type_code[supplier_id][inv_code] = {'allowed': False, 'price': None}
            try:
                inventory_codes_allowed = get_inventories_allowed(inventory_summary_map[supplier_id])
            except KeyError:
                # record this error and return in stats
                stats['inventory_summary_no_instance_error'].add(supplier_id)
                continue

            for inv_code in inventory_codes_allowed:
                suppliers_per_supplier_type_code[supplier_id][inv_code]['allowed'] = True
                try:
                    inventory_default_meta = v0_constants.inventory_type_duration_dict_list[inv_code]
                except KeyError:
                    raise Exception('The supplier {0} does not have an entry into constants.'.format(supplier_id))
                try:
                    ad_inventory_instance = ad_inventory_map[inventory_default_meta[0], inventory_default_meta[1]]
                    inventory_name = ad_inventory_instance.adinventory_name
                    inventory_type = ad_inventory_instance.adinventory_type
                    ad_inventory_id = ad_inventory_instance.id
                except KeyError:
                    raise Exception(
                        'The supplier {0} does not have ad_inventory type instance for {1}, {2}.'.format(supplier_id,
                                                                                                         inventory_default_meta[
                                                                                                             0],
                                                                                                         inventory_default_meta[
                                                                                                             1]))
                try:
                    duration_instance = duration_map[inventory_default_meta[2]]
                    duration_name = duration_instance.duration_name
                    duration_id = duration_instance.id
                except KeyError:
                    raise Exception('The supplier {0} does not have duration instance for {1}. '.format(supplier_id,
                                                                                                        inventory_default_meta[
                                                                                                            2]))
                try:
                    key = (ad_inventory_id, duration_id, supplier_id, content_type.id)
                    price = price_mapping_default_map[key] if key in price_mapping_default_map else price_mapping_default_map[list(price_mapping_default_map)[0]]
                except KeyError:
                    error_key = (inventory_name, inventory_type, duration_name, supplier_id)
                    raise Exception(
                        'The price mapping default instance does not exist for this supplier. key {0} is not in the pmd map.  detail is {1}.   \n valid keys are {2}. '.format(
                            key, error_key, list(price_mapping_default_map.keys())))
                suppliers_per_supplier_type_code[supplier_id][inv_code]['price'] = price

        return suppliers_per_supplier_type_code
    except Exception as e:
        raise Exception(function, ui_utils.get_system_error(e))


def add_inventory_summary_details(supplier_list, inventory_summary_objects_mapping, supplier_type_code,
                                  shortlisted=True, status=True):
    """
    This function adds details from inventory summary table for all the suppliers in
    supplier_list.

    Args:
        supplier_list:  [{supplier_id: 123, ..}, { }, { }  ] type structure in which
        each item is a dict containing details of only one supplier.
        supplier_type_code: a code to identify a supplier
        inventory_summary_objects_mapping: {
              supplier_id_1: inv_summ_object_1, supplier_id_2: inv_summ_object_2
            } type structure in which the right inv_object is put against supplier_id
        shortlisted: True. by default we assume suppliers are marked shortlisted = True
        status : 'X' by default we assume status of each supplier is 'X'.

    Returns: adds information like price or count from inv_summary table to each supplier
    dict .
    """
    function = add_inventory_summary_details.__name__
    try:
        statistics = {}
        content_type = ui_utils.fetch_content_type(supplier_type_code)
        supplier_ids = [supplier_dict['supplier_id'] for supplier_dict in supplier_list]
        price_mapping_default_map, ad_inventory_map, duration_map = make_handy_price_mapping_default_duration_adinventory_type(
            supplier_ids, content_type)
        inventory_count_map = get_inventory_count(supplier_ids, content_type)

        for supplier in supplier_list:
            supplier_inventory_obj = inventory_summary_objects_mapping.get(supplier['supplier_id'])
            supplier['shortlisted'] = shortlisted
            supplier['buffer_status'] = False
            # status is set to True.
            if not supplier.get('status'):
                supplier['status'] = v0_constants.status

            allowed_inventory_codes = get_inventories_allowed(supplier_inventory_obj)
            for inventory_code in allowed_inventory_codes:
                inventory_name = v0_constants.inventory_code_to_name[inventory_code].lower()
                if inventory_name == v0_constants.flier.lower():
                    supplier[
                        'flier_frequency'] = supplier_inventory_obj.flier_frequency if supplier_inventory_obj else v0_constants.default_inventory_count
                else:
                    db_key = 'total_' + inventory_name + '_count'
                    # set count from inventory summary if available. if not set count calculated previously from actual inventory tables
                    if supplier_inventory_obj:
                        try:
                            total_inventory_count = supplier_inventory_obj.__dict__[db_key]
                            # if there is no count, we try to set count from individual table of the inventory
                            if not total_inventory_count:
                                try:
                                    supplier[db_key] = inventory_count_map[supplier['supplier_id']][inventory_code]
                                except KeyError:
                                    supplier[db_key] = v0_constants.default_inventory_count
                            # else, set the default value from inventory summary
                            else:
                                supplier[db_key] = total_inventory_count
                        # if the 'total_inventory_count' isn't present as a key, set to default count
                        except KeyError:
                            supplier[db_key] = v0_constants.default_inventory_count
                    # if no inventory summary we try to set by individual inventory table
                    else:
                        try:
                            supplier[db_key] = inventory_count_map[supplier['supplier_id']][inventory_code]
                        except KeyError:
                            supplier[db_key] = v0_constants.default_inventory_count

                ad_inventory_instance = ad_inventory_map[
                    v0_constants.inventory_type_duration_dict_list[inventory_code][0],
                    v0_constants.inventory_type_duration_dict_list[inventory_code][1]]
                duration_instance = duration_map[v0_constants.inventory_type_duration_dict_list[inventory_code][2]]
                try:
                    supplier[inventory_name + '_price'] = price_mapping_default_map[
                        ad_inventory_instance.id, duration_instance.id, supplier['supplier_id'], content_type.id][
                        'actual_supplier_price']
                except KeyError:
                    supplier[inventory_name + '_price'] = v0_constants.default_inventory_price
                try:
                    supplier[inventory_name + '_duration'] = duration_instance.duration_name
                except KeyError:
                    supplier[inventory_name + '_duration'] = v0_constants.default_inventory_duration

        return supplier_list
    except Exception as e:
        raise Exception(function, ui_utils.get_system_error(e))


def get_inventories_allowed(inventory_summary_instance):
    """

    Args:
        inventory_summary_instance:
    checks for  poster, stall, standee, flier, car display.
    Returns:

    """
    function = get_inventories_allowed.__name__

    try:
        if not inventory_summary_instance:
            return []
        inventory_code_list = []
        if inventory_summary_instance.poster_allowed_nb or inventory_summary_instance.poster_allowed_lift:
            inventory_code_list.append(v0_constants.inventory_name_to_code['POSTER'])
        if inventory_summary_instance.stall_allowed:
            inventory_code_list.append(v0_constants.inventory_name_to_code['STALL'])
        if inventory_summary_instance.standee_allowed:
            inventory_code_list.append(v0_constants.inventory_name_to_code['STANDEE'])
        if inventory_summary_instance.flier_allowed:
            inventory_code_list.append(v0_constants.inventory_name_to_code['FLIER'])
        if inventory_summary_instance.car_display_allowed:
            inventory_code_list.append(v0_constants.inventory_name_to_code['CAR DISPLAY'])
        return inventory_code_list
    except Exception as e:
        raise Exception(function, ui_utils.get_system_error(e))


def make_handy_price_mapping_default_duration_adinventory_type(supplier_ids, content_type):
    """

    Returns:

    """
    function = make_handy_price_mapping_default_duration_adinventory_type.__name__
    try:
        price_mapping_instances_map = {}
        ad_inventory_instance_map = {}
        duration_instance_map = {}
        pmd_objects = PriceMappingDefault.objects.filter(object_id__in=supplier_ids, content_type=content_type).values(
            'adinventory_type', 'duration_type', 'object_id', 'content_type', 'suggested_supplier_price',
            'actual_supplier_price')
        for instance in pmd_objects:
            try:
                reference = price_mapping_instances_map[
                    instance['adinventory_type'], instance['duration_type'], instance['object_id'], instance[
                        'content_type']]
            except KeyError:
                price_mapping_instances_map[
                    instance['adinventory_type'], instance['duration_type'], instance['object_id'], instance[
                        'content_type']] = instance

        for ad_inventory_instance in AdInventoryType.objects.all():
            try:
                reference = ad_inventory_instance_map[
                    ad_inventory_instance.adinventory_name, ad_inventory_instance.adinventory_type]
            except KeyError:
                ad_inventory_instance_map[
                    ad_inventory_instance.adinventory_name, ad_inventory_instance.adinventory_type] = ad_inventory_instance

        for duration_instance in DurationType.objects.all():
            try:
                reference = duration_instance_map[duration_instance.duration_name]
            except KeyError:
                duration_instance_map[duration_instance.duration_name] = duration_instance

        return price_mapping_instances_map, ad_inventory_instance_map, duration_instance_map
    except Exception as e:
        raise Exception(function, ui_utils.get_system_error(e))


def add_shortlisted_suppliers(supplier_type_code_list, shortlisted_suppliers, inventory_summary_objects_mapping=None):
    """

    Args:
        supplier_type_code_list: ['RS', 'CP', 'GY']
        shortlisted_suppliers:  a list of object id's
        inventory_summary_objects_mapping: a map of object_id --> inv sum instance

    Returns: { RS: [], CP: [], GY: [] } a dict containing each type of suppliers in the list

    """
    function = add_shortlisted_suppliers.__name__
    try:
        # from the codes, fetch the right supplier model, supplier serializer, and put in the dict against that code
        result = {}

        supplier_ids = []

        # supplier_id : filter object mapping so that we can add relevant info from
        # filter table to each supplier dict
        supplier_to_filter_object_mapping = {}

        # shortlisted_suppliers array can be empty !
        shortlisted_suppliers = shortlisted_suppliers if shortlisted_suppliers else []

        for supplier in shortlisted_suppliers:
            supplier_id = supplier['object_id']
            supplier_ids.append(supplier_id)
            supplier_to_filter_object_mapping[supplier_id] = supplier

        for code in supplier_type_code_list:
            if code == 'RS':
                supplier_objects = SupplierTypeSociety.objects.filter(supplier_id__in=supplier_ids)
                serializer = SupplierTypeSocietySerializer(supplier_objects, many=True)
            else:
                supplier_objects = SupplierMaster.objects.filter(supplier_id__in=supplier_ids, supplier_type=code)
                serializer = SupplierMasterSerializer(supplier_objects, many=True)

            # adding status information to each supplier which is stored in shorlisted_spaces table
            for supplier in serializer.data:
                supplier_id = supplier['supplier_id']
                supplier['status'] = supplier_to_filter_object_mapping[supplier_id]['status']

            result[code] = serializer.data
            # proceed only when shortlisted_suppliers is non empty and inv_summ_object_map exist !
            if shortlisted_suppliers and inventory_summary_objects_mapping:
                result[code] = add_inventory_summary_details(serializer.data, inventory_summary_objects_mapping, code,
                                                             False, False)
            # convert society_keys to common supplier keys to access easily at frontEnd
            result[code] = manipulate_object_key_values(result[code], supplier_type_code=code)

        return result
    except Exception as e:
        raise Exception(function, ui_utils.get_system_error(e))


def proposal_shortlisted_spaces(data):
    """
    Args:
        data: a dict containing proposal_id for which we have to fetch all shortlisted spaces

    Returns: all shortlisted spaces
    """
    function = proposal_shortlisted_spaces.__name__
    try:
        proposal_id = data['proposal_id']

        # fetch all shortlisted suppliers object id's for this proposal
        shortlisted_suppliers = ShortlistedSpaces.objects.filter(proposal_id=proposal_id).select_related(
            'content_object').values()

        shortlisted_suppliers = manipulate_object_key_values(shortlisted_suppliers)


        # collect all supplier_id's
        supplier_ids = [supplier['object_id'] for supplier in shortlisted_suppliers]

        # fetch all inventory_summary objects related to each one of suppliers
        inventory_summary_objects = InventorySummary.objects.filter(object_id__in=supplier_ids)

        # generate a mapping from object_id to inv_summ_object in a dict so that right object can be fetched up
        inventory_summary_objects_mapping = {inv_sum_object.object_id: inv_sum_object for inv_sum_object in
                                             inventory_summary_objects}

        shortlisted_suppliers_centerwise = {}

        

        # populate the dict with object_id's now
        for supplier in shortlisted_suppliers:

            center_id = supplier['center_id']
            if not shortlisted_suppliers_centerwise.get(center_id):
                shortlisted_suppliers_centerwise[center_id] = []

            shortlisted_suppliers_centerwise[center_id].append(supplier)

        
       
        # construction of proposal response is isolated
        response = construct_proposal_response(proposal_id)
        if not response.data['status']:
            return response

        # all connected centers data
        centers = response.data['data']

        # collect all center_codes against each center_id
        center_id_list = [center['id'] for center in centers]
        response = add_filters(proposal_id, center_id_list)
        if not response.data['status']:
            return response

        # suppliers meta information  is available against each center_id
        filter_data = response.data['data']

        # final result dict
        result = {}

        # add extra information in each center object
        for center in centers:
            # dict to store intermediate result
            center_result = {
                # add shortlisted suppliers for codes available for this center
                'suppliers': add_shortlisted_suppliers(center['codes'],
                                                       shortlisted_suppliers_centerwise.get(center['id']),
                                                       inventory_summary_objects_mapping),
                'center': center,
                'suppliers_meta': filter_data[center['id']]['suppliers_meta'],
            }
            result[center['id']] = center_result

        dynamic_suppliers = get_dynamic_suppliers_by_campaign(proposal_id)
        if len(dynamic_suppliers) > 0:
            result['dynamic_suppliers'] = dynamic_suppliers
        return ui_utils.handle_response(function, data=result, success=True)
    except Exception as e:
        return ui_utils.handle_response(function, exception_object=e)


def add_filters(proposal_id, center_id_list):
    """
    The function is used to return all filters relatated to proposa_id, center_id,
    and supplier_type_code.
    """
    function = add_filters.__name__
    try:

        filter_objects = Filters.objects.values().filter(proposal_id=proposal_id)

        # the container to hold all filter objects per center
        filter_objects_per_center = {}

        for filter_object in filter_objects:
            center_id = filter_object['center_id']

            # if not given space for this center, give it !
            if not filter_objects_per_center.get(center_id):
                filter_objects_per_center[center_id] = []
            # collect all filter objects for this center here
            filter_objects_per_center[center_id].append(filter_object)

        # output result. The structure ouf the result is defined here
        result = {center_id: {'suppliers_meta': {}} for center_id in center_id_list}

        # iterate for valid centers
        for center_id in center_id_list:

            filter_objects = filter_objects_per_center[center_id] if filter_objects_per_center.get(center_id) else []

            # iterate for filter objects for this center
            for filter_object in filter_objects:

                supplier_type_code = filter_object['supplier_type_code']
                filter_name = filter_object.get('filter_name') or 'inventory_type_selected'
                filter_code = filter_object['filter_code']

                # give memory
                if not result[center_id]['suppliers_meta'].get(supplier_type_code):
                    result[center_id]['suppliers_meta'][supplier_type_code] = {}

                if not result[center_id]['suppliers_meta'][supplier_type_code].get(filter_name):
                    result[center_id]['suppliers_meta'][supplier_type_code][filter_name] = []

                # append the filter_code
                result[center_id]['suppliers_meta'][supplier_type_code][filter_name].append(filter_code)

        return ui_utils.handle_response(function, data=result, success=True)

    except KeyError as e:
        return ui_utils.handle_response(function, data='key Error', exception_object=e)
    except Exception as e:
        return ui_utils.handle_response(function, exception_object=e)


def unique_supplier_type_codes(data):
    """
    because the information that which  supplier types are shortlisted is not in the request,we need to derive it
    from data itself.
    Args:
        data: data
    Returns: unique supplier codes available
    """
    function_name = unique_supplier_type_codes.__name__
    try:
        supplier_type_codes = []
        for center in data:
            codes = list(center['suppliers'].keys())
            supplier_type_codes.extend(codes)
        return list(set(supplier_type_codes))
    except Exception as e:
        raise Exception(function_name, ui_utils.get_system_error(e))


def extra_header_database_keys(supplier_type_codes, data, result):
    """
    because depending upon inventory types selected for each supplier, the headers can be extended and hence the database
    keys shall be extended hence this function returns a dict for each of the suppliers the extra headers and database keys
    Args:
        data: data
    Returns: a dict containing header and database keys
    """
    function = extra_header_database_keys.__name__
    try:
        for code in supplier_type_codes:
            for center in data:
                # check that suppliers_meta is indeed present. Only then the extra headers
                # are added.
                if center.get('suppliers_meta') and center.get('suppliers_meta').get(code):
                    try:
                        inventory_codes = center['suppliers_meta'][code]['inventory_type_selected']
                    except KeyError:
                        inventory_codes = []
                else:
                    inventory_codes = []

                unique_inv_codes = get_unique_inventory_codes(inventory_codes)

                # extend the header keys with header for supplier type codes
                result[code]['header_keys'].extend(get_union_keys_inventory_code('HEADER', unique_inv_codes))

                # extend the data keys with header for supplier type codes
                result[code]['data_keys'].extend(get_union_keys_inventory_code('DATA', unique_inv_codes))

            # remove duplicates for this supplier code. we hell can't use sets because that thing will destroy the order
            # of the keys which is important and the order must match
            result[code]['header_keys'] = remove_duplicates_preserver_order(result[code]['header_keys'])
            result[code]['data_keys'] = remove_duplicates_preserver_order(result[code]['data_keys'])

            # set the counts for validation.
            result[code]['header_keys_count'] = len(result[code]['header_keys'])
            result[code]['data_keys_count'] = len(result[code]['data_keys'])
        return result
    except Exception as e:
        raise Exception(function, ui_utils.get_system_error(e))


def initialize_export_final_response(supplier_type_codes, result):
    """
    because there can be multiple type of supplier_type_codes, this function prepares a dict and sets it to
    keys and values required  for further processing for example with sheet names etc.
    Args:
        supplier_type_codes: ['RS', 'CP']

    Returns: a initialized dict

    """
    function = initialize_export_final_response.__name__
    try:
        for code in supplier_type_codes:
            result[code] = {}
            sheet_name = v0_constants.sheet_names[code]
            result[code]['sheet_name'] = sheet_name

            # set fixed headers for center
            result[code]['header_keys'] = get_union_keys_inventory_code('HEADER', ['CENTER'])

            # set fixed data keys for center
            result[code]['data_keys'] = get_union_keys_inventory_code('DATA', ['CENTER'])

            # set fixed header keys for supplier
            result[code]['header_keys'].extend(get_union_keys_inventory_code('HEADER', [code]))

            # set fixed data keys for supplier
            result[code]['data_keys'].extend(get_union_keys_inventory_code('DATA', [code]))

            result[code]['objects'] = []
        return result
    except Exception as e:
        raise Exception(function, ui_utils.get_system_error(e))


def construct_single_supplier_row(object, keys):
    """
    Args:
        object: The  object which is a dict
        keys a list of keys

    Returns: a dict containing final data

    """
    function = construct_single_supplier_row.__name__
    try:
        result = {key: object.get(key) for key in keys}
        return result
    except Exception as e:
        raise Exception(function, ui_utils.get_system_error(e))


def make_export_final_response(result, data, inventory_summary_map, supplier_inventory_pricing_map, stats):
    """
    This function populates the result with 'objects' per supplier_type_codes.
    Args:
        result: result dict where objects will be stored
        data: the entire request.data
        inventory_summary_map: { 'RS' :  { 'S1' --->  inventory_summary_instance, ... } }
        supplier_inventory_pricing_map: { 'S1' ---> { 'PO': { 'allowed': True, 'price': 5000 }, 'ST': { 'allowed': True, 'price': 2122 } }, 's2' --> ... }
    Returns:
    """
    function = make_export_final_response.__name__
    try:
        for center in data:

            # obtain the dict containing centre information
            center_info_dict = construct_single_supplier_row(center['center'], v0_constants.center_keys).copy()

            for code, supplier_object_list in center['suppliers'].items():

                # calculate unique inventory codes available in the suppliers_meta dict for this supplier_type
                if center.get('suppliers_meta') and center.get('suppliers_meta').get(code):
                    try:
                        inventory_codes = center['suppliers_meta'][code]['inventory_type_selected']
                    except KeyError:
                        # happens when 'inventory_type_selected' code is missing
                        inventory_codes = []
                else:
                    center_id = center['center']['id']
                    center_error = 'This center with name {0} does not has any inventory selected. Hence No pricing data will be calculated'.format(
                        center_id)
                    inventory_codes = []
                    stats['center_error'].append(center_error)

                # get union of inventory_codes. POST will become [PO, ST] etc.
                unique_inv_codes = get_unique_inventory_codes(inventory_codes)
                for index, supplier_object in enumerate(supplier_object_list):

                    supplier_id = supplier_object['supplier_id']
                    # check weather inv summary instance exists for this supplier or not. we do not continue if it doesn't exist
                    try:
                        inventory_summary_map[code][supplier_id]
                    except KeyError:
                        stats['inventory_summary_no_instance_error'].add(supplier_id)

                    if inventory_summary_map.get(code) and inventory_summary_map[code].get(supplier_id):
                        if supplier_id in supplier_inventory_pricing_map:
                            # this module inserts a few keys in supplier_object such as 'is_allowed' and 'pricing' keys for each inventory.
                            is_error, detail = set_supplier_inventory_keys(supplier_object,
                                                                           inventory_summary_map[code][supplier_id],
                                                                           unique_inv_codes,
                                                                           supplier_inventory_pricing_map[supplier_id])
                            supplier_object = detail.copy()

                    # obtain the dict containing non-center information
                    supplier_info_dict = construct_single_supplier_row(supplier_object, result[code]['data_keys'])

                    # merge the two dicts
                    # final object ( dict ). it contains center information as well as suppliers specific information
                    final_supplier_dict = merge_two_dicts(center_info_dict, supplier_info_dict)

                    # add _price_per_flat information to the final_dict received.
                    response_data = get_union_inventory_price_per_flat(final_supplier_dict, unique_inv_codes, index)
                    # append it to the result
                    result[code]['objects'].append(response_data)

        return result, stats
    except Exception as e:
        raise Exception(function, ui_utils.get_system_error(e))


def set_supplier_inventory_keys(supplier_dict, inv_summary_instance, unique_inventory_codes,
                                supplier_inventory_pricing_map):
    """
     sets 'is_allowed' and '_price' for each inventory that is allowed in that supplier.
    :param supplier_dict:
    :param inv_summary_instance:
    :param unique_inventory_codes:
    :param supplier_inventory_pricing_map: { 'PO': { 'allowed': 'True', 'price': {..contents of pmd } }, ... }
    :return: sets a key in supplier_dict for each inventory in unique_inventory_codes which tells weather that
     inventory is allowed or not.

    """
    function = set_supplier_inventory_keys.__name__
    try:
        allowed_inventory_codes = get_inventories_allowed(inv_summary_instance)
        for code in unique_inventory_codes:
            inventory_key = '_'.join(v0_constants.inventory_code_to_name[code].lower().split())
            inventory_allowed_data_key = inventory_key + '_allowed'
            inventory_pricing_key = inventory_key + '_price'
            inventory_pricing_available_key = inventory_key + '_price_available'

            supplier_dict[inventory_key + '_price_type'] = join_with_underscores(
                ' '.join(v0_constants.inventory_type_duration_dict_list[code])).lower()

            if code not in allowed_inventory_codes:
                # this inventory is not allowed
                supplier_dict[inventory_allowed_data_key] = 0
            else:
                supplier_dict[inventory_allowed_data_key] = 1
                # set pricing only when it is allowed.
                price = supplier_inventory_pricing_map[code]['price']['actual_supplier_price']
                if not price:
                    # mark this price as not available, so that they don't eat my head 'why pricing is not available ?'.
                    supplier_dict[inventory_pricing_available_key] = 0
                    # return True, (supplier_dict['supplier_id'], code)
                else:
                    supplier_dict[inventory_pricing_available_key] = 1
                    supplier_dict[inventory_pricing_key] = price
        return False, supplier_dict
    except Exception as e:
        raise Exception(function, ui_utils.get_system_error(e))


def handle_common_filters(common_filters, supplier_type_code, proposal):
    """
    This function handles only common filters.
    Args:
        common_filters: { 'latitude': 10, 'longitude': 12, 'radius': 1, 'quality': ['UH', 'H'],
        'quantity': ['VL', 'L'] }

    Returns: a Q object based on above filters

    """
    function = handle_common_filters.__name__
    try:
        if not common_filters:
            return ui_utils.handle_response(function, data=Q(), success=True)

        # we will store field and values in this dict and later on use it to construct Q object
        query = {}

        # fetch the lat long radius
        latitude = float(common_filters['latitude'])
        longitude = float(common_filters['longitude'])
        radius = float(common_filters['radius'])

        response = get_coordinates(radius, latitude, longitude)
        if not response.data['status']:
            return response
        # get the coordinates
        coordinates = response.data['data']

        max_latitude = coordinates['max_lat']
        max_longitude = coordinates['max_long']
        min_latitude = coordinates['min_lat']
        min_longitude = coordinates['min_long']

        # start build the query
        if supplier_type_code == 'RS':
            query['society_latitude__lt'] = max_latitude
            query['society_latitude__gt'] = min_latitude
            query['society_longitude__lt'] = max_longitude
            query['society_longitude__gt'] = min_longitude

        else:
            query['address_supplier__latitude__lt'] = max_latitude
            query['address_supplier__latitude__gt'] = min_latitude
            query['address_supplier__longitude__lt'] = max_longitude
            query['address_supplier__longitude__gt'] = min_longitude

        # query['representative'] = proposal.principal_vendor

        # the keys like 'locality', 'quantity', 'quality' we receive from front end are already defined in constants
        predefined_common_filter_keys = list(v0_constants.query_dict[supplier_type_code].keys())
        # we may receive a subset of already defined keys. obtain that subset
        received_common_filter_keys = list(common_filters.keys())
        # iterate over each predefined key and check if it is what we have received.
        for filter_term in predefined_common_filter_keys:
            if filter_term in received_common_filter_keys:
                # if received, obtain the query term for that filter key. query term looks like 'locality_rating__in'
                query_term = v0_constants.query_dict[supplier_type_code][filter_term]['query']
                # fetch the query dict associated. it contains the code-value mapping of the filters codes
                query_dict = v0_constants.query_dict[supplier_type_code][filter_term]['dict']
                # fetch the filter codes received like ['HH', 'UH']
                filter_codes = common_filters.get(filter_term)
                # get the actual values against these codes for example, 'Ulta High' for 'UH'. These values are defined
                # in the query_dict fetched above.
                filter_term_value_list = [query_dict.get(code) for code in filter_codes]
                # set the query dict to list of values calculated against the db field term
                query[query_term] = filter_term_value_list

        # remove those fields that have False value or Empty values.
        response = remove_empty_fields(query)
        if not response.data['status']:
            return response
        query = response.data['data']

        # build the actual Q object once the fields are fixed
        common_filter_query = Q(**query)

        # return the query
        return ui_utils.handle_response(function, data=common_filter_query, success=True)

    except KeyError as e:
        return ui_utils.handle_response(function, data='Key Error occurred', exception_object=e)
    except Exception as e:
        return ui_utils.handle_response(function, exception_object=e)


def remove_empty_fields(q_object_dict):
    """
    Args:
        q_object_dict: This function takes a dict called q_object_dict, and removes all those keys that have None value
        against it. This is because we do not want to query against a NULL value as it produces inconsistent
        results.

    Returns: q dict with keys whose value is None removed.

    """
    function = remove_empty_fields.__name__
    try:
        q_object_dict_copy = q_object_dict.copy()
        for key, value in q_object_dict.items():
            if not value:
                del q_object_dict_copy[key]
        return ui_utils.handle_response(function, data=q_object_dict_copy, success=True)
    except Exception as e:
        return ui_utils.handle_response(function, exception_object=e)


def handle_inventory_filters(inventory_list):
    """
    Args:
        inventory_list: ['PO', 'POST', 'ST' ]

    Returns: a Q object after handling each inventory code
    query = PO | ST | (POST) | CD | (STFL)
    for single codes like 'PO', 'ST', etc the query is just there corresponding db field.
    for complex codes like 'POST', each individual code is extracted and individual database fields are joined by 'and'.
    """
    function = handle_inventory_filters.__name__
    try:
        if not inventory_list:
            return ui_utils.handle_response(function, data=Q(), success=True)

        # final Q object to be returned
        inventory_query = Q()
        # atomic inventories means 'PO', 'ST' etc.
        valid_atomic_inventories = list(v0_constants.inventory_dict.keys())
        # iterate through all the inventory list
        for inventory in inventory_list:
            # if it is atomic, that means you only need to fetch it's db field and set it to Q object
            if inventory in valid_atomic_inventories:
                # the policy is to  OR the atomic inventories
                if inventory_query:
                    inventory_query |= Q(**{v0_constants.inventory_dict[inventory]: True})
                else:
                    inventory_query = Q(**{v0_constants.inventory_dict[inventory]: True})
                continue
            # come here only it it's non atomic inventory code.
            query = {}
            step = 2
            # split the non atomic inventory code into size of 2 letters.
            individual_codes = [inventory[i:i + step] for i in range(0, len(inventory), step)]
            # for each code
            for code in individual_codes:
                # set the query
                query[v0_constants.inventory_dict[code]] = True

            # make a new Q object with query and OR it with previously calculated inventory_query.
            if inventory_query:
                inventory_query |= Q(**query)
            else:
                inventory_query = Q(**query)

        return ui_utils.handle_response(function, data=inventory_query, success=True)
    except Exception as e:
        return ui_utils.handle_response(function, exception_object=e)


def handle_specific_filters(specific_filters, supplier_type_code):
    """
    This function is called from construct query which handles filters specific to supplier
    Args:
        specific_filters: { 'real_estate_allowed': True, 'total_employee_count': 120, 'flat_type': [ '1RK', '2BHK' ]
          }
        supplier_type_code: 'RS', 'CP'
    Returns: a Q object based on above filters

    """
    function = handle_specific_filters.__name__
    try:

        if not specific_filters:
            return ui_utils.handle_response(function, data=Q(), success=True)

        # get the predefined dict of specific filters for this supplier
        master_specific_filters = v0_constants.supplier_filters[supplier_type_code]

        specific_filters_query = Q()
        # the following loop stores those fields in received filters which can be mapped directly to db columns.
        for received_filter, filter_value in specific_filters.items():

            database_field = master_specific_filters.get(received_filter)

            if database_field:
                # do things only when you get a db field
                if not specific_filters_query:
                    specific_filters_query = Q(**{database_field: filter_value})
                else:
                    specific_filters_query &= Q(**{database_field: filter_value})

        if supplier_type_code == v0_constants.corporate:
            # well, we can receive a multiple dicts for employee counts. each describing min and max employee counts.
            # todo: does not make sense
            if specific_filters.get('employees_count'):
                for employees_count_range in specific_filters.get('employees_count'):
                    max = employees_count_range['max']
                    min = employees_count_range['min']
                    if specific_filters_query:
                        specific_filters_query |= Q(totalemployees_count__gte=min, totalemployees_count__lte=max)
                    else:
                        specific_filters_query = Q(totalemployees_count__gte=min, totalemployees_count__lte=max)
        # return it
        return ui_utils.handle_response(function, data=specific_filters_query, success=True)

    except Exception as e:
        return ui_utils.handle_response(function, exception_object=e)


def handle_priority_index_filters(supplier_type_code, pi_filters_map, final_suppliers_id_list):
    """

    Returns:

    """
    function = handle_priority_index_filters.__name__
    try:
        if not supplier_type_code or not pi_filters_map or not final_suppliers_id_list:
            return {}

        content_type = ui_utils.fetch_content_type(supplier_type_code)
        supplier_model = apps.get_model(settings.APP_NAME, content_type.model)

        pi_index_map = {}
        # give space for the two keys initially
        for supplier_id in final_suppliers_id_list:
            pi_index_map[supplier_id] = {
                'total_priority_index': 0,
                'detail': {}
            }
            for filter_name, filter_value in pi_filters_map.items():
                pi_index_map[supplier_id]['detail'][filter_name] = {
                    'query': filter_value,
                    'explanation': {},
                    'assigned_pi': 0
                }

        # do if else check on supplier type code to include things particular to that supplier. Things which
        # cannot be mapped to a particular supplier and vary from supplier to supplier
        # handle filters of type min_max here. min max type filters are always there for PI.
        queryset = supplier_model.objects.filter(supplier_id__in=final_suppliers_id_list)
        for filter_name, db_value in v0_constants.pi_range_filters[supplier_type_code].items():
            if pi_filters_map.get(filter_name):
                min_value = float(pi_filters_map[filter_name]['min'])
                max_value = float(pi_filters_map[filter_name]['max'])

                # didn't issue the query of __gte and  __lte because i want to capture details of all suppliers who
                # didn't qualify for this query. this goes as part of explanation of PI
                suppliers_db_map = {item['supplier_id']: item[db_value] for item in
                                    queryset.values('supplier_id', db_value)}

                for supplier_id in final_suppliers_id_list:
                    db_value_result = suppliers_db_map.get(supplier_id)
                    if not db_value_result:
                        db_value_result = -1

                    # increment PI only when this supplier qualifies for the query.
                    if min_value <= float(db_value_result) <= max_value:
                        pi_index_map[supplier_id]['detail'][filter_name]['assigned_pi'] = 1

                    # need to record what the query was, what was the output for all the suppliers
                    pi_index_map[supplier_id]['total_priority_index'] += \
                    pi_index_map[supplier_id]['detail'][filter_name]['assigned_pi']
                    pi_index_map[supplier_id]['detail'][filter_name]['explanation'] = {filter_name: db_value_result}

        # handle amenities here
        if pi_filters_map.get('amenities'):
            filter_name = 'amenities'

            queryset = SupplierAmenitiesMap.objects.filter(content_type=content_type,
                                                           object_id__in=final_suppliers_id_list).values(
                'object_id').annotate(count=Count('amenity'))
            queryset_map = {item['object_id']: item['count'] for item in queryset}

            # if amenity query is issued from front end, all suppliers should have 'amenities' as filter name in 'detail' dict
            for supplier_id in final_suppliers_id_list:

                amenity_count = queryset_map.get(
                    supplier_id)  # set zero to amenity_count if we do not find amenity for this current supplier
                if not amenity_count:
                    amenity_count = 0
                # if else check to assign PI of 1 in case the supplier passes the amenity threshold
                if amenity_count >= v0_constants.amenity_count_threshold:
                    pi_index_map[supplier_id]['detail'][filter_name]['assigned_pi'] = 1

                pi_index_map[supplier_id]['detail'][filter_name]['explanation'] = {
                    'supplier_amenity_count': amenity_count, 'amenity_threshold': v0_constants.amenity_count_threshold}
                pi_index_map[supplier_id]['total_priority_index'] += pi_index_map[supplier_id]['detail'][filter_name][
                    'assigned_pi']

        if supplier_type_code == v0_constants.society:

            # check for standalone societies
            if pi_filters_map.get('is_standalone_society'):
                filter_name = 'is_standalone_society'
                query = Q(supplier_id__in=final_suppliers_id_list)
                standalone_supplier_detail = get_standalone_societies(query)
                for supplier_id in final_suppliers_id_list:
                    detail = standalone_supplier_detail.get(supplier_id)
                    if not detail:
                        detail = {}
                    if is_society_standalone(detail):
                        pi_index_map[supplier_id]['detail'][filter_name]['assigned_pi'] = 1
                    pi_index_map[supplier_id]['detail'][filter_name]['explanation'] = detail
                    pi_index_map[supplier_id]['total_priority_index'] += \
                    pi_index_map[supplier_id]['detail'][filter_name]['assigned_pi']

            if pi_filters_map.get('flat_type'):
                filter_name = 'flat_type'
                queryset = FlatType.objects.filter(content_type=content_type, object_id__in=final_suppliers_id_list)
                flat_result = {}
                for item in queryset:
                    if not flat_result.get(item.object_id):
                        flat_result[item.object_id] = {}

                    flat_result[item.object_id][item.flat_type] = {
                        'flat_count': item.flat_count,
                        'flat_size': item.size_builtup_area
                    }

                for supplier_id in final_suppliers_id_list:
                    flat_details = flat_result.get(supplier_id)
                    if flat_details:
                        for flat_code, detail in pi_filters_map['flat_type'].items():
                            flat_code_value = v0_constants.flat_type_dict[flat_code]
                            single_flat_detail = flat_details.get(flat_code_value)
                            if flat_code_value in list(flat_details.keys()):
                                pi_index_map[supplier_id]['detail'][filter_name]['assigned_pi'] += 1
                            if single_flat_detail and detail.get('count') and int(detail['count']['min']) <= \
                                    single_flat_detail['flat_count'] <= int(detail['count']['max']):
                                pi_index_map[supplier_id]['detail'][filter_name]['assigned_pi'] += 1
                            if single_flat_detail and detail.get('size') and float(detail['size']['min']) <= \
                                    single_flat_detail['flat_size'] <= float(detail['size']['max']):
                                pi_index_map[supplier_id]['detail'][filter_name]['assigned_pi'] += 1
                    pi_index_map[supplier_id]['detail'][filter_name][
                        'explanation'] = flat_details if flat_details else {}
                    pi_index_map[supplier_id]['total_priority_index'] += \
                    pi_index_map[supplier_id]['detail'][filter_name]['assigned_pi']

            if pi_filters_map.get('ratio_of_tenants_to_flats'):
                # check for society ratio of tenants to flats
                filter_name = 'ratio_of_tenants_to_flats'
                ratio_of_tenants_to_flats = pi_filters_map['ratio_of_tenants_to_flats']
                supplier_ratio_details = SupplierTypeSociety.objects.filter(
                    supplier_id__in=final_suppliers_id_list).values('supplier_id').annotate(ratio=ExpressionWrapper(
                    F('total_tenant_flat_count') / F('flat_count'), output_field=FloatField()))
                supplier_ratio_details_map = {item['supplier_id']: item['ratio'] for item in supplier_ratio_details}
                for supplier_id in final_suppliers_id_list:
                    ratio = supplier_ratio_details_map.get(supplier_id)
                    if not ratio:
                        ratio = 0.0
                    if float(ratio_of_tenants_to_flats['min']) <= ratio <= float(ratio_of_tenants_to_flats['max']):
                        pi_index_map[supplier_id]['detail'][filter_name]['assigned_pi'] = 1

                    pi_index_map[supplier_id]['detail'][filter_name]['explanation'] = ratio
                    pi_index_map[supplier_id]['total_priority_index'] += \
                    pi_index_map[supplier_id]['detail'][filter_name]['assigned_pi']

        return pi_index_map
    except Exception as e:
        raise Exception(function, ui_utils.get_system_error(e))


def initialize_area_subarea(result, index):
    """
    Args:
        result: a list
        index: index of the list to initialize with a dict containing 'state', 'city' , 'area', 'subarea'.

    Returns:

    """
    function = initialize_area_subarea.__name__
    try:
        # empty dict
        area_subarea_object = {}
        # set a state dict
        area_subarea_object['state'] = {}
        # set a city dict
        area_subarea_object['city'] = {}
        # set a area dict
        area_subarea_object['area'] = {}
        # set a subarea dict
        area_subarea_object['subarea'] = {}
        # set the final dict to this index
        result[index] = area_subarea_object

        return ui_utils.handle_response(function, data=result, success=True)
    except Exception as e:
        return ui_utils.handle_response(function, exception_object=e)


def handle_states(result, index, row):
    """
    Args:
        result: The result list
        index: index of the list for which state object is to be modified
        row:  a single row of the sheet

    Returns:

    """
    function = handle_states.__name__
    try:
        state_object = result[index]['state']
        state_object['state_code'] = row['state_code']
        state_object['state_name'] = row['state_name']
        result[index]['state'] = state_object
        return ui_utils.handle_response(function, data=result, success=True)
    except Exception as e:
        return ui_utils.handle_response(function, exception_object=e)


def handle_city(result, index, row):
    """
    Args:
        result: The result list
        index: index of the list for which city object is to be modified
        row:  a single row of the sheet

    Returns:

    """
    function = handle_city.__name__
    try:
        city_object = result[index]['city']
        city_object['city_code'] = row['city_code']
        city_object['city_name'] = row['city_name']
        result[index]['city'] = city_object
        return ui_utils.handle_response(function, data=result, success=True)
    except Exception as e:
        return ui_utils.handle_response(function, exception_object=e)


def handle_area(result, index, row):
    """
    Args:
        result: The result list
        index: index of the list for which area object is to be modified
        row:  a single row of the sheet

    Returns:

    """
    function = handle_area.__name__
    try:
        area_object = result[index]['area']
        area_object['area_code'] = row['area_code']
        area_object['label'] = row['area_name']
        result[index]['area'] = area_object
        return ui_utils.handle_response(function, data=result, success=True)
    except Exception as e:
        return ui_utils.handle_response(function, exception_object=e)


def handle_subarea(result, index, row):
    """
    Args:
        result: The result list
        index: index of the list for which subarea object is to be modified
        row:  a single row of the sheet

    Returns:

    """
    function = handle_subarea.__name__
    try:
        subarea_object = result[index]['subarea']
        subarea_object['subarea_code'] = row['subarea_code']
        subarea_object['subarea_name'] = row['subarea_name']
        subarea_object['locality_rating'] = row.get('locality_rating')
        result[index]['subarea'] = subarea_object
        return ui_utils.handle_response(function, data=result, success=True)
    except Exception as e:
        return ui_utils.handle_response(function, exception_object=e)


def get_inventory_pricing(supplier_type_code, inventory_names, inventory_types, inventory_durations):
    """
    Args:
        inventory_code: fetches price details for one inventory_code, 'PO', 'ST'.
        inventory_type: 'Canopy', 'Small', 'Large'
        inventory_duration: 'Campaign Weekly', 'Campaign Monthly'.

    Returns: price for that inventory from PriceMappingDefault table
    """
    function = get_inventory_pricing.__name__
    try:
        supplier_content_type = ui_utils.get_content_type(supplier_type_code)
        adinventory_type_objects = AdInventoryType.objects.filter(adinventory_name__in=inventory_names,
                                                                  adinventory_type__in=inventory_types)
        duration_type_objects = DurationType.objects.filter(duration_name__in=inventory_durations)
        inventory_prices = PriceMappingDefault.objects.prefetch_related('content_object').filter(
            content_type=supplier_content_type).values('object_id', 'actual_supplier_price', 'adinventory_type__name',
                                                       'adinventory_type__type', 'duration__type__name').filter(
            adinventory_type__in=adinventory_type_objects, duration_type__in=duration_type_objects)

        return ui_utils.handle_response(function, data=inventory_prices, success=True)
    except Exception as e:
        return ui_utils.handle_response(function, exception_object=e)


def set_supplier_extra_attributes(suppliers, supplier_type_code, inventory_codes):
    """
    Args:
        suppliers: a list containg dict for each supplier

    Returns: modified each supplier by adding some attributes

    """
    function = set_supplier_extra_attributes.__name__
    try:
        inventory_names = []
        inventory_types = []
        inventory_durations = []
        for inventory_code in inventory_codes:
            inventory_duration_dict = v0_constants.inventory_duration_dict[inventory_code]
            inventory_types.extend(type_dur_dict['type'] for type_dur_dict in inventory_duration_dict['type_duration'])
            inventory_durations.extend(
                type_dur_dict['duration'] for type_dur_dict in inventory_duration_dict['type_duration'])
            inventory_names.extend(inventory_duration_dict['name'])

        response = get_inventory_pricing(supplier_type_code, inventory_names, inventory_types, inventory_durations)
        if not response.data['status']:
            return response
        inventory_pricing = response.data['data']
        return ui_utils.handle_response(function, data=suppliers, success=True)
    except Exception as e:
        return ui_utils.handle_response(function, exception_object=e)


def save_area_subarea(result):
    """
    The result looks like this
     [
        {
          "city": {
            "city_name": "Mumabi",
            "city_code": "BOM"
          },
          "state": {
            "state_code": "MH",
            "state_name": "Maharashtra"
          },
          "subarea": {
            "locality_rating": "high",
            "subarea_code": "AE",
            "subarea_name": "Andheri(E)"
          },
          "area": {
            "area_code": "AE",
            "label": "Andheri(E)"
          }
        },
        {..}
     ]
    Args:
        result: The result

    Returns: success if the data in result is saved, else failure

    """
    function = save_area_subarea.__name__
    try:
        total_new_objects_created = 0

        state_objects = []
        city_objects = []
        area_objects = []
        subarea_objects = []

        with transaction.atomic():
            for data in result:
                # make state
                state_object, is_created = State.objects.get_or_create(state_code=data['state']['state_code'])
                if is_created:
                    total_new_objects_created += 1

                state_object.state_name = data['state']['state_name']
                state_objects.append(state_object)

                # make city
                city_object, is_created = City.objects.get_or_create(city_code=data['city']['city_code'],
                                                                     state_code=state_object)
                if is_created:
                    total_new_objects_created += 1

                city_object.city_name = data['city']['city_name']
                city_objects.append(city_object)

                # make area
                area, is_created = CityArea.objects.get_or_create(area_code=data['area']['area_code'],
                                                                  city_code=city_object)
                if is_created:
                    total_new_objects_created += 1
                area.label = data['area']['label']

                area_objects.append(area)

                # make subarea
                subarea, is_created = CitySubArea.objects.get_or_create(subarea_code=data['subarea']['subarea_code'],
                                                                        area_code=area)
                if is_created:
                    total_new_objects_created += 1
                subarea.subarea_name = data['subarea']['subarea_name']

                subarea_objects.append(subarea)

            bulk_update(state_objects)
            bulk_update(city_objects)
            bulk_update(area_objects)
            bulk_update(subarea_objects)

            return ui_utils.handle_response(function, data=total_new_objects_created, success=True)
    except Exception as e:
        return ui_utils.handle_response(function, exception_object=e)


def get_inventory_count(supplier_ids, content_type, inventory_summary_objects=None):
    """
    Args:
        supplier_ids:
        content_type:
        inventory_summary_objects
    Returns: a dict of all inventory counts allowed for each supplier

    """
    function = get_inventory_count.__name__
    try:
        # if we get inventory summary objects as param, we need to return total counts which are stored in here. will be removed in future.
        if inventory_summary_objects:
            return inventory_summary_objects.aggregate(posters=Sum('total_poster_count'),
                                                       standees=Sum('total_standee_count'),
                                                       stalls=Sum('total_stall_count'), fliers=Sum('flier_frequency'))
        # else we need to calculate the counts per supplier by counting each row in each inventory tables
        all_inventory_codes = v0_constants.inventories_with_object_id_fields
        inventory_models = {}
        for code in all_inventory_codes:
            inventory_content_type = ui_utils.fetch_content_type(code)
            model = apps.get_model(settings.APP_NAME, inventory_content_type.model)
            inventory_models[code] = model
        result = {}
        count_query_result = {}
        for inventory_code, model in inventory_models.items():
            count_query_result[inventory_code] = {detail['object_id']: detail['count'] for detail in
                                                  model.objects.filter(object_id__in=supplier_ids,
                                                                       content_type=content_type).values(
                                                      'object_id').annotate(count=Count('adinventory_id'))}

        for supplier_id in supplier_ids:
            if not result.get(supplier_id):
                result[supplier_id] = {}
            for code in all_inventory_codes:
                try:
                    result[supplier_id][code] = count_query_result[code][supplier_id]
                except KeyError:
                    result[supplier_id][code] = v0_constants.default_inventory_count
        return result
    except Exception as e:
        raise Exception(function, ui_utils.get_system_error(e))


def set_pricing_temproray(suppliers, supplier_ids, supplier_type_code, coordinates, priority_index_map):
    """
    Args:
        suppliers: a list of supplier dicts
        supplier_ids: a list of supplier id's
        supplier_type_code: CP, RS
        coordinates: a dict containing radius, lat, long information.
        priority_index_map: a map of ids to PI.

    Returns: list of suppliers with pricing set, count of inventories
    """

    function = set_pricing_temproray.__name__
    try:
        # todo: remove this function . redudant code. directly call add_inventory_summary_details
        # fetch all inventory_summary objects related to each one of suppliers
        content_type = ui_utils.fetch_content_type(supplier_type_code)
        inventory_summary_objects = InventorySummary.objects.filter(object_id__in=supplier_ids,
                                                                    content_type=content_type)
        # generate a mapping from object_id to inv_summ_object in a dict so that right object can be fetched up
        inventory_summary_objects_mapping = {inv_summary_object.object_id: inv_summary_object for inv_summary_object in
                                             inventory_summary_objects}
        suppliers_inventory_count = get_inventory_count(supplier_ids, content_type, inventory_summary_objects)
        # add the PI indexes
        for supplier in suppliers:
            supplier['priority_index'] = priority_index_map.get(supplier['supplier_id'])
        result = add_inventory_summary_details(suppliers, inventory_summary_objects_mapping, supplier_type_code,
                                               status=True)
        return result, suppliers_inventory_count
    except Exception as e:
        raise Exception(function, ui_utils.get_system_error(e))


def get_dict_value(my_dict, possible_keys):
    """

    Args:
        my_dict:
        possible_keys:

    Returns:

    """
    function = get_dict_value.__name__
    try:
        value = None
        for key in possible_keys:
            try:
                value = my_dict[key]
            except KeyError:
                pass
        return value
    except Exception as e:
        raise Exception(function, ui_utils.get_system_error(e))


def get_suppliers_within_circle(suppliers, coordinates, supplier_type_code):
    """
    list of supplier dicts
    Args:
        suppliers:
        coordinates: radius, latitude, longitude
        supplier_type_code:

    Returns:

    """
    function = get_suppliers_within_circle.__name__
    try:
        radius = float(coordinates['radius'])
        latitude = float(coordinates['latitude'])
        longitude = float(coordinates['longitude'])

        # container to hold final suppliers
        result = []

        for supplier in suppliers:

            if supplier_type_code != "RS":
                address_supplier = supplier.get('address_supplier')
                if address_supplier:
                    supplier['latitude'] = address_supplier.get('latitude')
                    supplier['longitude'] = address_supplier.get('longitude')

            # include only those suppliers that lie within the circle of radius given
            supplier_latitude = get_dict_value(supplier, ['society_latitude', 'latitude'])
            supplier_longitude = get_dict_value(supplier, ['society_longitude', 'longitude'])
           

            if space_on_circle(latitude, longitude, radius, supplier_latitude, supplier_longitude):
                result.append(supplier)
        return result
    except Exception as e:
        raise Exception(function, ui_utils.get_system_error(e))


def handle_inventory_pricing(inv_type, dur_type, supplier_id, supplier_type_code, actual_supplier_price=0):
    """
    Args:
        inv_type: type of inventory
        dur_type: duration
        supplier_id: The supplier_id for which price needs to be returned
        supplier_type_code: The supplier_type_code.
        actual_supplier_price: The price if given is set to newly created price_mapping_default object.

    Returns: price for the inventory, for this inventory type and this duration
    """
    function = handle_inventory_pricing.__name__
    try:
        response = ui_utils.get_content_type(supplier_type_code)
        if not response.data['data']:
            return response
        content_type = response.data['data']
        adinventory_type_dict = ui_utils.adinventory_func()
        duration_type_dict = ui_utils.duration_type_func()
        price_mappings = PriceMappingDefault.objects.filter(adinventory_type=adinventory_type_dict[inv_type],
                                                            duration_type=duration_type_dict[dur_type],
                                                            object_id=supplier_id, content_type=content_type)
        if not price_mappings:
            return ui_utils.handle_response(function, data=0, success=True)
        price_mapping = price_mappings[0]
        price_mapping.actual_supplier_price = actual_supplier_price
        price_mapping.save()
        return ui_utils.handle_response(function, data=price_mapping.supplier_price, success=True)
    except Exception as e:
        return ui_utils.handle_response(function, exception_object=e)


def get_file_name(user, proposal_id, is_exported=True):
    """
    Args:
        user: The user name
        proposal_id: proposal_id
        is_exported: determines weather this file was exported or imported.

    Returns: a string that wil be used as file name.

    """
    function = get_file_name.__name__
    try:
        format = v0_constants.datetime_format
        now_time = datetime.datetime.now()
        datetime_stamp = now_time.strftime(format)
        proposal = ProposalInfo.objects.get(proposal_id=proposal_id)
        account = proposal.account
        organisation = account.organisation
        if user.is_anonymous:
            user_string = 'Anonymous'
            user = None
        else:
            user_string = user.get_username()
        # file_name = user_string + '_' + organisation.name.lower() + '_' + account.name.lower() + '_' + proposal_id + '_' + datetime_stamp + '.xlsx'
        file_name = organisation.name.lower() + '_' + proposal_id + '_' + datetime_stamp + '.xlsx'

        # save this file in db
        data = {
            'user': user,
            'organisation': organisation,
            'account': account,
            'proposal': proposal,
            # 'date': now_time,
            'file_name': file_name,
            'is_exported': is_exported
        }
        GenericExportFileName.objects.get_or_create(**data)
        return file_name
    except Exception as e:
        raise Exception(function, ui_utils.get_system_error(e))


def add_metric_sheet(workbook):
    """
    Args:
        workbook: a workbook object
    Returns:  reads a sheet from empty_proposal_cost_data.xlsx, add it to export sheet for metrics.
    """
    function = add_metric_sheet.__name__
    try:
        my_file = open(v0_constants.metric_file_path, 'rb')
        wb = openpyxl.load_workbook(my_file)
        # copy first sheet from saved workbook
        first_sheet = wb.get_sheet_by_name(v0_constants.metric_sheet_name)
        # create a target sheet where data will be copied
        target_sheet = workbook.create_sheet(index=0, title=v0_constants.metric_sheet_name)
        # for each row, copy it to new sheet
        for row in first_sheet.iter_rows():
            target_row_list = [cell.value for cell in row]
            target_sheet.append(target_row_list)

        return workbook
    except Exception as e:
        raise Exception(function, ui_utils.get_system_error(e))


def send_excel_file(file_name):
    """
    converts the file in binary before returning  it. and sends the required mail
    """
    function = send_excel_file.__name__
    try:
        if os.path.exists(file_name):

            excel = open(file_name, "rb")
            file_content = excel.read()
            output = StringIO(str(file_content))
            out_content = output.getvalue()
            output.close()
            excel.close()
        else:
            # return response
            raise Exception(function, 'File does not exist on disk')
        return out_content
    except Exception as e:
        raise Exception(function, ui_utils.get_system_error(e))


def save_price_mapping_default(supplier_id, supplier_type_code, row):
    """
    saves data in price_mapping_default  for all kinds of inventories
    Args:
        supplier_id:  The supplier_id
        supplier_type_code:  RS, CP
        row: a variable representing one row of the sheet.

    Returns: success in case price entry is success.

    """
    function = save_price_mapping_default.__name__
    try:
        with transaction.atomic():
            # for each predefined inventories. each 'inventory' is actually a tuple. look in constants for more details.
            for inventory in v0_constants.current_inventories:

                inventory_type = inventory[0]  # at index 0 we have inventory_type
                inventory_duration = inventory[1]  # at index 1 we have inventory_duration
                if not row[inventory[2]]:
                    price = 0
                else:
                    price = int(row[inventory[2]])  # at index 2 we have inventory_pricing index.
                # todo: one db hit in each loop. improve if code slows down in future
                response = handle_inventory_pricing(inventory_type, inventory_duration, supplier_id, supplier_type_code,
                                                    actual_supplier_price=price)
                if not response.data['status']:
                    return response
            # done. return success.
            return ui_utils.handle_response(function, data='success', success=True)
    except Exception as e:
        return ui_utils.handle_response(function, exception_object=e)


def delete_create_final_proposal_data(proposal_id):
    """"
    This function deletes all the data related to this proposal_id whenever create-final-proposal api
    is called because we do not want extra or duplicate data

    The data that is deleted is from two tables - shortlisted_spaces, filters.
    """
    function = delete_create_final_proposal_data.__name__
    try:
        # delete all the shortlisted_spaces rows for this proposal
        ShortlistedSpaces.objects.filter(proposal_id=proposal_id).delete()
        # delete all Filter table rows for this proposal
        Filters.objects.filter(proposal_id=proposal_id).delete()
        return ui_utils.handle_response(function, data='success', success=True)
    except Exception as e:
        return ui_utils.handle_response(function, exception_object=e)


def process_template(target_string, mapping):
    """
    Args:
        target_string: The string  on which templating is to be performed.
        mapping: a dict having keys as placeholders, and values which will be in place of placeholders.

    Returns: a string with all placeholders replaced by there respective values
    """
    function = process_template.__name__
    try:
        template_string = Template(target_string)
        result_string = template_string.substitute(mapping)
        return result_string
    except Exception as e:
        raise Exception(function, ui_utils.get_system_error(e))


def proposal_centers(proposal_id):
    """
    This function basically collects centers associated with proposal

    Args:
        contains proposal_id
    Returns: a dictionary of centers and suppliers in centers
    """
    function = proposal_centers.__name__
    try:
        data = {}
        centers = ProposalCenterMapping.objects.filter(proposal_id=proposal_id).values()
        suppliers = ProposalCenterSuppliers.objects.filter(proposal_id=proposal_id).values()
        for center in centers:
            center['supplier_codes'] = []
            for supplier in suppliers:
                if supplier['center_id'] == center['id']:
                    center['supplier_codes'].append(supplier['supplier_type_code'])
        return ui_utils.handle_response(function, data=centers, success=True)
    except Exception as e:
        return ui_utils.handle_response(function, exception_object=e)


def add_shortlisted_suppliers_get_spaces(proposal_id, user, data):
    """
    Args:
        proposal_id: The proposal_id for which we need to add ShortlistedSpaces
        data:  The data is the response of get_spaces() api.
    Returns: add's status field in each of the suppliers . The status field comes form ShortlistedSpaces.
    All those suppliers which are not in ShortlistedSpaces and present in data  have a special status called 'X'.
    This is required from front end side as the suppliers marked with 'X' are shown with a different colour there.
    """
    function = add_shortlisted_suppliers_get_spaces.__name__
    try:
        # get all suppliers from ShortlistedSpaces for this proposal_id and user
        response = get_shortlisted_suppliers(proposal_id, user)
        if not response.data['status']:
            return response
        result = response.data['data']

        # replace the status of each supplier under each center to status already saved.
        # Here the union takes place between two supplier sets. one from response of get_spaces() and one from response
        # from ShortlistedSpaces, because we need to send those suppliers which were saved last time.
        for supplier_dict in data['suppliers']:
            for code in supplier_dict['center']['codes']:
                center_id = supplier_dict['center']['id']

                if result.get(center_id) and result[center_id].get(code):
                    # set it to union of two sets.
                    supplier_dict['suppliers'][code] = union_suppliers(supplier_dict['suppliers'][code],
                                                                       result[center_id][code]).values()

        return ui_utils.handle_response(function, data=data, success=True)
    except Exception as e:
        return ui_utils.handle_response(function, exception_object=e)


def union_suppliers(first_supplier_list, second_supplier_list):
    """

    Args:
        first_supplier_list: first supplier list of dicts
        second_supplier_list: second supplier list of dicts

    Returns: Union of two. sets status for all those suppliers which are in first_list  but not in second set to 'X'.

    """
    function = union_suppliers.__name__
    try:
        first_supplier_list_ids = set()
        second_supplier_list_ids = set()
        first_supplier_mapping = {}
        second_supplier_mapping = {}

        if not second_supplier_list and (not first_supplier_list):
            return {}

        if second_supplier_list:
            for supplier in second_supplier_list:
                supplier_id = supplier['supplier_id'] if supplier.get('supplier_id') else supplier['object_id']
                second_supplier_list_ids.add(supplier_id)
                second_supplier_mapping[supplier_id] = supplier

        if first_supplier_list:
            for supplier in first_supplier_list:
                supplier_id = supplier['supplier_id'] if supplier.get('supplier_id') else supplier.get('object_id')
                first_supplier_list_ids.add(supplier_id)
                first_supplier_mapping[supplier_id] = supplier
                supplier['status'] = v0_constants.status

        total_supplier_ids = first_supplier_list_ids.union(second_supplier_list_ids)
        suppliers_not_in_second_set = first_supplier_list_ids.difference(second_supplier_list_ids)

        result = {}
        for supplier_id in total_supplier_ids:
            if second_supplier_mapping.get(supplier_id) and not result.get(supplier_id):
                result[supplier_id] = second_supplier_mapping[supplier_id]

            elif first_supplier_mapping.get(supplier_id):
                result[supplier_id] = first_supplier_mapping[supplier_id]

            if supplier_id in suppliers_not_in_second_set:
                result[supplier_id]['status'] = v0_constants.status
            result[supplier_id]['supplier_id'] = supplier_id
        return result
    except Exception as e:
        logger.exception(e)
        raise Exception(function, ui_utils.get_system_error(e))


def get_shortlisted_suppliers_map(proposal_id, content_type, center_id):
    """

    Args:
        proposal_id:
        content_type:

    Returns:

    """
    function = get_shortlisted_suppliers_map.__name__
    try:
        # fetch the shortlisted supplier instances ( object_id, status only )
        shortlisted_suppliers = ShortlistedSpaces.objects.filter(proposal_id=proposal_id, content_type=content_type,
                                                                 center_id=center_id).values('object_id', 'status')
        shortlisted_ids = []
        shortlisted_suppliers_map = {}
        # prepare a map from id --> ss because later we will need to fetch it
        for instance_dict in shortlisted_suppliers:
            supplier_id = instance_dict['object_id']
            shortlisted_suppliers_map[supplier_id] = instance_dict
            shortlisted_ids.append(supplier_id)
        
        if content_type.model == "suppliertypesociety":
            instances = SupplierTypeSociety.objects.filter(supplier_id__in=shortlisted_ids).values()
        else:
            supplier_master = SupplierMaster.objects.filter(supplier_id__in=shortlisted_ids)
            supplier_master_serializer = SupplierMasterSerializer(supplier_master, many=True).data
            instances = manipulate_master_to_rs(supplier_master_serializer)

        result = {}
        for instance in instances:
            supplier_id = instance['supplier_id']
            result[supplier_id] = merge_two_dicts(instance, shortlisted_suppliers_map[supplier_id])
        return result
    except Exception as e:
        raise Exception(function, ui_utils.get_system_error(e))


def get_shortlisted_suppliers(proposal_id, user):
    """
    Args:
        proposal_id: The proposal_id
        user: The User instance.

    Returns: a dict having shortlisted suppliers against each code and each center
    {
    "45":  {
              "RS" : [ {}, {}, {} ],
              "CP": [ {}]
             },
    "44" : {
              "RS" : [ {} ]
            }
    }
    """
    function = get_shortlisted_suppliers.__name__
    try:

        # fetch the right shortlisted space instances
        shortlisted_spaces = ShortlistedSpaces.objects.filter(proposal_id=proposal_id).values()

        # to store supplier id against each supplier_type_code
        shortlisted_spaces_content_type_wise = {}
        # to store supplier objects against each center and each supplier_type_code.
        shortlisted_suppliers_center_content_type_wise = {}
        # to store unique supplier_type_codes
        supplier_type_codes = set()

        # collect all supplier_ids against each supplier_type_code.
        for item in shortlisted_spaces:
            supplier_id = item['object_id']
            supplier_type_code = item['supplier_code']
            supplier_type_codes.add(supplier_type_code)

            if not shortlisted_spaces_content_type_wise.get(supplier_type_code):
                shortlisted_spaces_content_type_wise[supplier_type_code] = set()

            shortlisted_spaces_content_type_wise[supplier_type_code].add(supplier_id)

        #  get actual supplier objects for each supplier_id.
        for code in supplier_type_codes:

            supplier_ids = shortlisted_spaces_content_type_wise[code]
            query = Q(**{'supplier_id__in': supplier_ids})
            response = get_suppliers(query, code, {})
            if not response.data['status']:
                return response
            supplier_objects = {supplier_object['supplier_id']: supplier_object for supplier_object in
                                response.data['data']}
            shortlisted_spaces_content_type_wise[code] = supplier_objects

        # store each supplier object for each center and  each supplier_type_code.
        for item in shortlisted_spaces:
            supplier_id = item['object_id']
            center_id = item['center_id']
            supplier_type_code = item['supplier_code']
            status = item['status']

            if not shortlisted_suppliers_center_content_type_wise.get(center_id):
                shortlisted_suppliers_center_content_type_wise[center_id] = {}

            if not shortlisted_suppliers_center_content_type_wise[center_id].get(supplier_type_code):
                shortlisted_suppliers_center_content_type_wise[center_id][supplier_type_code] = []

            
            if supplier_id and shortlisted_spaces_content_type_wise.get(supplier_type_code) and shortlisted_spaces_content_type_wise[supplier_type_code].get(supplier_id) :
                supplier_object = shortlisted_spaces_content_type_wise[supplier_type_code][supplier_id]
                if supplier_object:
                    supplier_object['status'] = status
                    shortlisted_suppliers_center_content_type_wise[center_id][supplier_type_code].append(supplier_object)


        return ui_utils.handle_response(function, data=shortlisted_suppliers_center_content_type_wise, success=True)
    except KeyError as e:
        return ui_utils.handle_response(function, data='key error', exception_object=e)
    except Exception as e:
        return ui_utils.handle_response(function, exception_object=e)


def manipulate_object_key_values(suppliers, supplier_type_code=v0_constants.society, **kwargs):
    """
    Args:
        suppliers: list of all suppliers
        supplier_type_code: by default 'RS'.
        kwargs: key,value pairs meant to set in each supplier.
    Returns:
        return list of suppliers by changing some keys in supplier object
    """
    function = manipulate_object_key_values.__name__
    try:
        for supplier in suppliers:

            if supplier.get('address_supplier'):
                address_supplier = supplier.get('address_supplier')
                supplier['address1'] = address_supplier['address1'] if address_supplier['address1'] else ''
                supplier['address2'] = address_supplier['address2']
                supplier['area'] = address_supplier['area']
                supplier['subarea'] = address_supplier['subarea']
                supplier['city'] = address_supplier['city']
                supplier['state'] = address_supplier['state']
                supplier['zipcode'] = address_supplier['zipcode']
                supplier['latitude'] = address_supplier['latitude']
                supplier['longitude'] = address_supplier['longitude']
                #del supplier['address_supplier']

            if supplier.get("supplier_name"):
                supplier['name'] = supplier['supplier_name']

            # replace all society specific keys with common supplier keys
            if supplier_type_code == v0_constants.society:
                for society_key, actual_key in v0_constants.society_common_keys.items():
                    if society_key in list(supplier.keys()):
                        value = supplier[society_key]
                        del supplier[society_key]
                        supplier[actual_key] = value

            if kwargs:
                # set extra key, value sent in kwargs
                for key, item in kwargs.items():
                    supplier[key] = item
        return suppliers
    except Exception as e:
        raise Exception(function, ui_utils.get_system_error(e))

def manipulate_object_key_values_generic(suppliers, supplier_type_code=v0_constants.society, **kwargs):
    """
    Args:
        suppliers: list of all suppliers
        supplier_type_code: by default 'RS'.
        kwargs: key,value pairs meant to set in each supplier.
    Returns:
        return list of suppliers by changing some keys in supplier object
    """
    function = manipulate_object_key_values_generic.__name__
    try:
        for supplier in suppliers:
            if supplier_type_code != 'RS' and supplier.get('address_supplier'):      
                address_supplier = supplier.get('address_supplier')
                supplier['address1'] = address_supplier['address1'] if address_supplier['address1'] else ''
                supplier['address2'] = address_supplier['address2']
                supplier['area'] = address_supplier['area']
                supplier['subarea'] = address_supplier['subarea']
                supplier['city'] = address_supplier['city']
                supplier['state'] = address_supplier['state']
                supplier['zipcode'] = address_supplier['zipcode']
                supplier['latitude'] = address_supplier['latitude']
                supplier['longitude'] = address_supplier['longitude']
                supplier['name'] = supplier['supplier_name']
                #del supplier.get('address_supplier')

            # replace all society specific keys with common supplier keys
            if supplier_type_code == v0_constants.society:
                for society_key, actual_key in v0_constants.society_common_keys.items():
                    if society_key in list(supplier.keys()):
                        value = supplier[society_key]
                        del supplier[society_key]
                        supplier[actual_key] = value

            if kwargs:
                # set extra key, value sent in kwargs
                for key, item in kwargs.items():
                    supplier[key] = item
        return suppliers
    except Exception as e:
        raise Exception(function, ui_utils.get_system_error(e))
    
def manipulate_master_to_rs(suppliers):
    """
    Args:
        suppliers: list of all suppliers
        supplier_type_code: by default 'RS'.
        kwargs: key,value pairs meant to set in each supplier.
    Returns:
        return list of suppliers by changing some keys in supplier object
    """
    function = manipulate_master_to_rs.__name__
    try:
        for supplier in suppliers:
            if supplier.get('address_supplier'):      
                address_supplier = supplier.get('address_supplier')
                supplier['address1'] = address_supplier['address1'] if address_supplier['address1'] else ''
                supplier['address2'] = address_supplier['address2']
                supplier['area'] = address_supplier['area']
                supplier['subarea'] = address_supplier['subarea']
                supplier['city'] = address_supplier['city']
                supplier['state'] = address_supplier['state']
                supplier['zipcode'] = address_supplier['zipcode']
                supplier['latitude'] = address_supplier['latitude']
                supplier['longitude'] = address_supplier['longitude']
                #del supplier.get('address_supplier')

            supplier['name'] = supplier.get("supplier_name")
            supplier['supplier_code'] = supplier.get("supplier_type")

            if not supplier.get("flat_count"):
                supplier["flat_count"] = supplier["unit_primary_count"] if supplier.get("unit_primary_count") else 0
            if not supplier.get("tower_count"):
                supplier["tower_count"] = supplier["unit_secondary_count"] if supplier.get("unit_secondary_count") else 0

        return suppliers
    except Exception as e:
        raise Exception(function, ui_utils.get_system_error(e))

def setup_generic_export(data, user, proposal_id):
    """
    data: Request.data
    user: User instance
    proposal_id: The proposal id

    Returns the sheet.
    """
    function = setup_generic_export.__name__
    try:
        workbook = Workbook()
        total_suppliers_map = {}
        # center_error occur when user has not selected any inventory for a center. This means for all suppliers
        # in that center, no pricing data will be fetched and displayed.
        # supplier_no_pricing_error happens when a particular inventory is allowed in a supplier but does not has pricing data available.
        # inventory_summary_no_instance_error: This occur when there is no inv_summ_instance instance for a supplier. For this supplier, hence we cannot is_allowed parts.
        stats = {'center_error': [], 'supplier_no_pricing_error': {}, 'inventory_summary_no_instance_error': set()}

        # get the supplier type codes available in the request
        unique_supplier_codes = unique_supplier_type_codes(data)

        result = {}

        # initialize the result = {} dict which will be used in inserting into sheet
        result = initialize_export_final_response(unique_supplier_codes, result)

        # collect all the extra header and database keys for all the supplier type codes and all inv codes in them
        result = extra_header_database_keys(unique_supplier_codes, data, result)

        # collect unique supplier ids first
        distinct_supplier_ids = set()
        for center in data:
            for supplier_code, supplier_object_list in center['suppliers'].items():
                if not total_suppliers_map.get(supplier_code):
                    total_suppliers_map[supplier_code] = []
                for dict_object in supplier_object_list:
                    total_suppliers_map[supplier_code].append(dict_object['supplier_id'])
                    distinct_supplier_ids.add(dict_object['supplier_id'])

        # store inv summary instances supplier_code wise.
        total_inventory_summary_instances = InventorySummary.objects.filter(object_id__in=distinct_supplier_ids)

        suppliers_without_inv_instances = []
        # check weather we have inventory summary instances for all suppliers or not. if not we are not proceeding further.
        inventory_summary_map = {instance.object_id: instance for instance in total_inventory_summary_instances}
        for supplier_id in distinct_supplier_ids:
            try:
                inventory_summary_map[supplier_id]
            except KeyError:
                suppliers_without_inv_instances.append(supplier_id)
        stats['inventory_summary_no_instance_error'] = set(suppliers_without_inv_instances)

        # now rebuild the map code wise.
        inventory_summary_map = {}
        for instance in total_inventory_summary_instances:
            # taking advantage of the fact that a supplier id contains it's code in it. 'RS' is embedded in two characters [7:8]
            supplier_code = 'RS'
            supplier_id = instance.object_id

            if not inventory_summary_map.get(supplier_code):
                inventory_summary_map[supplier_code] = {}

            if not inventory_summary_map[supplier_code].get(supplier_id):
                inventory_summary_map[supplier_code][supplier_id] = instance

        # fetch  pricing information directly from database
        supplier_pricing_map = {}
        for supplier_code, detail in inventory_summary_map.items():
            # detail is inventory_summary mapping.
            supplier_pricing_map = {}
            if len(total_suppliers_map.keys()) > 0 and supplier_code in total_suppliers_map :
                supplier_pricing_map = merge_two_dicts(
                    set_inventory_pricing(total_suppliers_map[supplier_code], supplier_code, detail, stats),
                    supplier_pricing_map)

        # make the call to generate data in the result. center_error and supplier_no_pricing_error is logged in this function
        result, stats = make_export_final_response(result, data, inventory_summary_map, supplier_pricing_map, stats)

        # print result
        workbook = insert_supplier_sheet(workbook, result)

        # make a file name for this file
        file_name = get_file_name(user, proposal_id)

        workbook.save(file_name)
        file_content = send_excel_file(file_name)
        content_type = v0_constants.mime['xlsx']

        # in order to provide custom headers in response in angular js, we need to set this header
        # first
        # headers = {
        #     'Access-Control-Expose-Headers': "file_name, Content-Disposition"
        #

        return {
            'local_path': os.path.join(settings.BASE_DIR, file_name),
            'name': file_name,
            'pricing_defaults': v0_constants.inventory_type_duration_dict_list,
            'stats': stats
        }

    except Exception as e:
        raise Exception(function, ui_utils.get_system_error(e))


def setup_create_final_proposal_post(data, proposal_id, delete_and_save_filter_data=True, delete_and_save_spaces=True,
                                     exclude_shortlisted_space=True):
    """
    Args:
        data: Request data
        proposal_id: Proposal_id
        delete_and_save_filter_data: weather to delete and recreate the filter or not
        delete_and_save_spaces: weather to delete and create spaces.
        exclude_shortlisted_space: parameter to determine weather to save or exclude suppliers with status as 'S'.

    Returns: Success in case success returns.
    """
    function = setup_create_final_proposal_post.__name__
    try:
        # get the supplier type codes available in the request
        unique_supplier_codes = unique_supplier_type_codes(data)

        with transaction.atomic():

            # containers to store shortlisted suppliers and filter information
            total_shortlisted_suppliers_list = []
            filter_data = []

            for proposal_data in data:
                proposal_data['proposal_id'] = proposal_id
                shortlisted_suppliers, filter_data = fetch_final_proposal_data(proposal_data, unique_supplier_codes)
                total_shortlisted_suppliers_list.extend(shortlisted_suppliers)
                filter_data.extend(filter_data)

            now_time = timezone.now()

            if delete_and_save_spaces:
                # by default it will not save those suppliers which have status set as 'S'.
                # delete previous  shortlisted suppliers and save new
                if not exclude_shortlisted_space:
                    final_shortlisted_list = total_shortlisted_suppliers_list
                else:
                    final_shortlisted_list = [instance for instance in total_shortlisted_suppliers_list if
                                              instance.status != v0_constants.shortlisted]
                ShortlistedSpaces.objects.filter(proposal_id=proposal_id).delete()
                ShortlistedSpaces.objects.bulk_create(final_shortlisted_list)
                ShortlistedSpaces.objects.filter(proposal_id=proposal_id).update(created_at=now_time,
                                                                                 updated_at=now_time)
            # making this conditional because we delete filters and save new filters whenever request proposal is hit, but we should not delete filters when the sheet is imported.
            if delete_and_save_filter_data:
                # delete previous and save new selected filters and update date
                Filters.objects.filter(proposal_id=proposal_id).delete()
                Filters.objects.bulk_create(filter_data)
                Filters.objects.filter(proposal_id=proposal_id).update(created_at=now_time, updated_at=now_time)

            return True
    except Exception as e:
        raise Exception(function, ui_utils.get_system_error(e))


def print_bucket_content(bucket_name):
    """
    prints all bucket contents
    Args:
        bucket_name:
    Returns: Nothing
    """
    function = print_bucket_content.__name__
    try:
        conn = boto.connect_s3(settings.AWS_ACCESS_KEY_ID, settings.AWS_SECRET_ACCESS_KEY)
        bucket = conn.get_bucket(bucket_name)
        log_file_path = settings.BASE_DIR + '/files/bucket_content.txt'
        output_file = open(log_file_path, 'w')
        for key in bucket.list():
            output_file.write(str(key) + "\n")
        output_file.close()
    except Exception as e:
        raise Exception(function, ui_utils.get_system_error(e))


def is_campaign(proposal):
    """
    The function which tells weather a proposal is a campaign or not
    Args:
        proposal: The proposal object

    Returns: True if campaign else False.
    """
    function = is_campaign.__name__
    try:

        if not proposal.invoice_number:
            return ui_utils.handle_response(function, data=errors.CAMPAIGN_NO_INVOICE_ERROR)

        if proposal.campaign_state == v0_constants.proposal_not_converted_to_campaign:
            return ui_utils.handle_response(function, data=errors.CAMPAIGN_NOT_APPROVED_ERROR)
        # todo: removing check for on-hold as of now. any proposal must have been accepted before on holding it. A proposal on
        # on hold means, a campaign was running but due to some reason, it's on hold. Still, that proposal_id was a campaign.
        #
        # if proposal.campaign_state == v0_constants.proposal_on_hold:
        #     return ui_utils.handle_response(function, data=errors.CAMPAIGN_ON_HOLD_ERROR)

        if not proposal.tentative_start_date or not proposal.tentative_end_date:
            return ui_utils.handle_response(function, data=errors.CAMPAIGN_NO_START_OR_END_DATE_ERROR.format(
                proposal.proposal_id))

        if (proposal.campaign_state != v0_constants.proposal_converted_to_campaign) and (
                proposal.campaign_state != v0_constants.proposal_on_hold):
            return ui_utils.handle_response(function,
                                            data=errors.CAMPAIGN_INVALID_STATE_ERROR.format(proposal.campaign_state,
                                                                                            v0_constants.proposal_converted_to_campaign,
                                                                                            v0_constants.proposal_on_hold))

        return ui_utils.handle_response(function, data='success', success=True)
    except Exception as e:
        return ui_utils.handle_response(function, exception_object=e)


def prepare_shortlisted_spaces_and_inventories(proposal_id, page, user, assigned, search, start_date, end_date, supplier_type_code=None, booking_status_code=None, phase_id=None, space_status=None):
    """

    Args:
        proposal_id: The proposal_id

    Returns: The data in required form.

    """

    function = prepare_shortlisted_spaces_and_inventories.__name__
    try:
        result = {}

        proposal = ProposalInfo.objects.get(proposal_id=proposal_id)

        filter_query = Q()

        filter_query &= Q(proposal_id=proposal_id)


        if supplier_type_code:
            filter_query &= Q(supplier_code=supplier_type_code)

        if booking_status_code:
            filter_query &= Q(booking_status=booking_status_code)

        if phase_id:
            filter_query &= Q(phase_no=phase_id)

        if assigned:
            assigned_suppliers_list = SupplierAssignment.objects.filter(campaign=proposal_id,assigned_to=assigned). \
                values_list('supplier_id')
            filter_query &= Q(object_id__in=assigned_suppliers_list)
            # shortlisted_spaces = ShortlistedSpaces.objects.filter(proposal_id=proposal_id, object_id__in=assigned_suppliers_list). \
            #     order_by('id')

        if search:
            filter_query &= Q(object_id=search)
            # shortlisted_spaces = ShortlistedSpaces.objects.filter(proposal_id=proposal_id, object_id=search).order_by('id')
        if start_date and end_date:
            filter_query &= Q(next_action_date__gte=start_date)
            filter_query &= Q(next_action_date__lte=end_date)
            filter_query &= Q(last_call_date__gte=start_date)
            filter_query &= Q(last_call_date__lte=end_date)
        
        if space_status:
            filter_query &= Q(status=space_status)
        
        shortlisted_spaces = ShortlistedSpaces.objects.filter(filter_query)

        if proposal.type_of_end_customer and proposal.type_of_end_customer.formatted_name in ["b_to_b_r_g", "b_to_b_l_d"]:
            shortlisted_spaces = shortlisted_spaces.order_by(F('color_code').asc(nulls_last=True), '-requirement_given_date')
        else:
            shortlisted_spaces = shortlisted_spaces.order_by('-id')

        if page:
            entries = 10
        else:
            entries = 500
            page = 1
        paginator = Paginator(shortlisted_spaces, entries)
        try:
            spaces = paginator.page(int(page))
        except PageNotAnInteger:
            spaces = paginator.page(1)
        except EmptyPage:
            spaces = paginator.page(paginator.num_pages)

        # set the campaign data
        proposal_serializer = ProposalInfoSerializer(proposal)
        result['campaign'] = proposal_serializer.data
        # set the shortlisted spaces data. it maps various supplier ids to their respective content_types
        response = get_objects_per_content_type_by_instance(spaces.object_list)

        if not response.data['status']:
            return response
        content_type_supplier_id_map, content_type_id_set, supplier_id_set = response.data['data']

        # converts the ids store in previous step to actual objects and adds additional information which is
        # supplier specific  like area, name, subarea etc.
        response = map_objects_ids_to_objects(content_type_supplier_id_map)
        if not response.data['status']:
            return response

        # the returned response is a dict in which key is (content_type, supplier_id) and value is a dict of extra
        # information for that supplier
        supplier_specific_info = response.data['data']
        response = get_contact_information(content_type_id_set, supplier_id_set)
        if not response.data['status']:
            return response

        contact_object_per_content_type_per_supplier = response.data['data']
        response = get_supplier_price_information(content_type_id_set, supplier_id_set)
        if not response.data['status']:
            return response
        supplier_price_per_content_type_per_supplier = response.data['data']

        shortlisted_suppliers_serializer = ShortlistedSpacesSerializerReadOnly(spaces, many=True)
        shortlisted_suppliers_list = shortlisted_suppliers_serializer.data
        result['shortlisted_suppliers'] = shortlisted_suppliers_list
        result['total_count'] = shortlisted_spaces.count()

        # put the extra supplier specific info like name, area, subarea in the final result.
        for supplier in shortlisted_suppliers_list:
            supplier['supplierCode'] = supplier['supplier_code']
            supplier['brand_organisation_id'] = supplier['brand_organisation_id']
            
            supplier_content_type_id = supplier['content_type']
            supplier_id = supplier['object_id']
            supplier['freebies'] = supplier['freebies'].split(',') if supplier['freebies'] else None
            supplier['stall_locations'] = supplier['stall_locations'].split(',') if supplier['stall_locations'] else None
            supplier_tuple = (supplier_content_type_id, supplier_id)
            try:
                for key, value in supplier_specific_info[supplier_tuple].items():
                    supplier[key] = value
            except:
                pass   
            # no good way to check if the tuple exist in the dict. Hence resorted to KeyError checking.
            try:
                retail_shop_instance = ContactDetails.objects.filter(object_id=supplier['object_id'])
                contact_serializer = ContactDetailsSerializer(retail_shop_instance, many=True)
                supplier['contacts'] = contact_serializer.data
            except KeyError:
                supplier['contacts'] = []
            try:
                brand_organisation_instance = Organisation.objects.filter(pk=supplier['brand_organisation_id']).first()
                brand_organisation_instance_serializer = OrganisationSerializer(brand_organisation_instance, many=False)
                supplier['brand_organisation_data'] = brand_organisation_instance_serializer.data
            except KeyError:
                supplier['brand_organisation_data'] = []
                
            try:
                pmd_objects_per_supplier = supplier_price_per_content_type_per_supplier[supplier_tuple]
            except KeyError:
                pmd_objects_per_supplier = []
                
            shortlisted_inventories = supplier['shortlisted_inventories']

            response = add_total_price_per_inventory_per_supplier(pmd_objects_per_supplier, shortlisted_inventories)
            if not response.data['status']:
                return response
            supplier['shortlisted_inventories'], total_inventory_supplier_price = response.data['data']
            supplier['total_inventory_supplier_price'] = total_inventory_supplier_price

        try:
            cache.set(str(proposal_id), result, timeout=60 * 100)
        except:
            pass
        
        return ui_utils.handle_response(function, data=result, success=True)
    except Exception as e:
        # print("e3",e)
        return ui_utils.handle_response(function, exception_object=e)


def handle_update_campaign_inventories(user, data):
    """
    Update campaign and inventories
    Args:
        data:

    Returns:

    """
    function = handle_update_campaign_inventories.__name__
    try:
        if not data:
            return ui_utils.handle_response(function, data='success', success=True)

        # collect shortlisted spaces specific details
        shortlisted_spaces = {}
        shortlisted_inventory_details = {}
        old_inventory_activity_assignments = {}
        new_inventory_activity_assignments = []
        inventory_activity_ids = set()

        for supplier in data:
            ss_global_id = supplier['id']
            if not shortlisted_spaces.get(ss_global_id):
                shortlisted_spaces[ss_global_id] = {}

            # response = check_valid_codes(supplier['payment_status'], supplier['payment_method'],
            #                              supplier['booking_status'])
            # if not response.data['status']:
            #     return response
            supplier_freebies = supplier['freebies'] if 'freebies' in supplier else None
            if supplier_freebies and isinstance(supplier_freebies,list):
                supplier_freebies = ','.join(supplier_freebies)
            supplier_stall_locations = supplier['stall_locations'] if 'stall_locations' in supplier else None
            if supplier_stall_locations and isinstance(supplier_stall_locations,list):
                supplier_stall_locations = ','.join(supplier_stall_locations)
            supplier_sunboard_locations = supplier['sunboard_location'] if 'sunboard_location' in supplier else None
            if supplier_sunboard_locations and isinstance(supplier_sunboard_locations, list):
                supplier_sunboard_locations = ','.join(supplier_sunboard_locations)
            if 'next_action_date' in supplier and supplier['next_action_date']:
                supplier['next_action_date'] = supplier['next_action_date']
            shortlisted_spaces[ss_global_id] = {
                'phase': supplier['phase'],
                'phase_no': supplier['phase_no'],
                'payment_status': supplier['payment_status'],
                'payment_method': supplier['payment_method'],
                'ifsc_code': supplier['ifsc_code'],
                'beneficiary_name': supplier['beneficiary_name'],
                'account_number': supplier['account_number'],
                'payment_message': supplier['payment_message'] if 'payment_message' in supplier else None,
                'total_negotiated_price': supplier['total_negotiated_price'],
                'booking_status': supplier['booking_status'],
                'booking_sub_status': supplier['booking_sub_status'],
                'transaction_or_check_number': supplier['transaction_or_check_number'],
                'freebies': supplier_freebies,
                'stall_locations': supplier_stall_locations,
                'is_completed' : supplier['is_completed'],
                'cost_per_flat' : supplier['cost_per_flat'],
                'booking_priority': supplier['booking_priority'],
                'sunboard_location': supplier['sunboard_location'] if 'sunboard_location' in supplier else None,
                'next_action_date': supplier['next_action_date'] if 'next_action_date' in supplier else None,
                'last_call_date': supplier['last_call_date'] if 'last_call_date' in supplier else None,
                'requirement_given': supplier['requirement_given'] if 'requirement_given' in supplier else 'no',
            }

            shortlisted_inventories = supplier['shortlisted_inventories']
            for inventory, inventory_detail in shortlisted_inventories.items():
                if inventory != 'NA' or inventory_detail != 'NA':
                    for inv in inventory_detail['detail']:
                        for inventory_activity in inv['inventory_activities']:
                            inventory_activity_id = inventory_activity['id']

                            inventory_activity_ids.add(inventory_activity_id)

                            for inventory_activity_assignment in inventory_activity['inventory_activity_assignment']:
                                if not inventory_activity_assignment.get('id'):
                                    new_inventory_activity_assignments.append(
                                        {
                                            'inventory_activity_id': inventory_activity_id,
                                            'activity_date': ui_utils.get_aware_datetime_from_string(
                                                inventory_activity_assignment['activity_date'])
                                        }
                                    )
                                else:
                                    old_inventory_activity_assignments[inventory_activity_assignment['id']] = {

                                        'inventory_activity_id': inventory_activity_id,
                                        'activity_date': ui_utils.get_aware_datetime_from_string
                                        (inventory_activity_assignment['activity_date'])
                                    }

                        sid_global_id = inv['id']
                        if not shortlisted_inventory_details.get(sid_global_id):
                            shortlisted_inventory_details[sid_global_id] = {}
                        shortlisted_inventory_details[sid_global_id] = {
                            'comment': inv['comment']
                        }
                        if inv['inventory_number_of_days']:
                            shortlisted_inventory_details[sid_global_id]['inventory_number_of_days'] = inv['inventory_number_of_days']
        data = {
            'shortlisted_spaces': shortlisted_spaces,
            'shortlisted_inventory_details': shortlisted_inventory_details,
            'old_inventory_activity_assignments': old_inventory_activity_assignments,
            'new_inventory_activity_assignments': new_inventory_activity_assignments,
            'inventory_activity_ids': inventory_activity_ids
        }
        response = update_campaign_inventories(data)
        if not response.data['status']:
            return response
        return ui_utils.handle_response(function, data='success', success=True)
    except Exception as e:
        return ui_utils.handle_response(function, exception_object=e)


def update_campaign_inventories(data):
    """
    updates the data in three tables: SS , SID, and Audit dates.
    Args:
        data: a dict of  individual dicts.

    Returns: success or failure
    """
    function = update_campaign_inventories.__name__
    try:
        shortlisted_spaces = data['shortlisted_spaces']
        shortlisted_inventory_details = data['shortlisted_inventory_details']
        new_inventory_activity_assignments = data['new_inventory_activity_assignments']  # user can issue new dates
        old_inventory_activity_assignments = data['old_inventory_activity_assignments']  # user can edit existing dates
        inventory_activity_ids = data['inventory_activity_ids']

        # mapping from id to object instance. used later to fetch objects directly from id
        inventory_activity_object_map = InventoryActivity.objects.in_bulk(inventory_activity_ids)

        shortlisted_spaces_ids = list(shortlisted_spaces.keys())
        ss_objects = ShortlistedSpaces.objects.filter(id__in=shortlisted_spaces_ids)

        for obj in ss_objects:
            ss_global_id = obj.id
            obj.phase_no = SupplierPhase.objects.get(id=shortlisted_spaces[ss_global_id]['phase_no']) if shortlisted_spaces[ss_global_id]['phase_no'] else None
            obj.phase = shortlisted_spaces[ss_global_id]['phase']
            obj.payment_status = shortlisted_spaces[ss_global_id]['payment_status']
            obj.payment_method = shortlisted_spaces[ss_global_id]['payment_method']
            obj.ifsc_code = shortlisted_spaces[ss_global_id]['ifsc_code']
            obj.beneficiary_name = shortlisted_spaces[ss_global_id]['beneficiary_name']
            obj.account_number = shortlisted_spaces[ss_global_id]['account_number']
            obj.payment_message = shortlisted_spaces[ss_global_id]['payment_message']
            obj.total_negotiated_price = shortlisted_spaces[ss_global_id]['total_negotiated_price']
            obj.booking_status = shortlisted_spaces[ss_global_id]['booking_status']
            obj.booking_sub_status = shortlisted_spaces[ss_global_id]['booking_sub_status']
            obj.transaction_or_check_number = shortlisted_spaces[ss_global_id]['transaction_or_check_number']
            obj.freebies = shortlisted_spaces[ss_global_id]['freebies']
            obj.stall_locations = shortlisted_spaces[ss_global_id]['stall_locations']
            obj.is_completed = shortlisted_spaces[ss_global_id]['is_completed']
            obj.cost_per_flat = shortlisted_spaces[ss_global_id]['cost_per_flat']
            obj.booking_priority = shortlisted_spaces[ss_global_id]['booking_priority']
            obj.sunboard_location = shortlisted_spaces[ss_global_id]['sunboard_location']
            obj.next_action_date = shortlisted_spaces[ss_global_id]['next_action_date']
            obj.last_call_date = shortlisted_spaces[ss_global_id]['last_call_date']
            
            if not obj.requirement_given == shortlisted_spaces[ss_global_id]['requirement_given']:
                obj.requirement_given = shortlisted_spaces[ss_global_id]['requirement_given']
                obj.requirement_given_date = datetime.datetime.now()

        sid_ids = list(shortlisted_inventory_details.keys())
        sid_objects = ShortlistedInventoryPricingDetails.objects.filter(id__in=sid_ids)

        for obj in sid_objects:
            sid_global_id = obj.id
            obj.comment = shortlisted_inventory_details[sid_global_id]['comment']
            if 'inventory_number_of_days' in shortlisted_inventory_details[sid_global_id]:
                obj.inventory_number_of_days = shortlisted_inventory_details[sid_global_id]['inventory_number_of_days']

        new_inventory_activity_assignment_objects = []
        for obj_dict in new_inventory_activity_assignments:
            obj_dict['inventory_activity'] = inventory_activity_object_map[obj_dict['inventory_activity_id']]
            del obj_dict['inventory_activity_id']
            new_inventory_activity_assignment_objects.append(InventoryActivityAssignment(**obj_dict))

        old_inventory_assignment_ids = list(old_inventory_activity_assignments.keys())
        old_inventory_activity_assignment_objects = InventoryActivityAssignment.objects.filter(
            id__in=old_inventory_assignment_ids)

        for instance in old_inventory_activity_assignment_objects:
            inventory_activity_assignment_id = instance.id
            instance.activity_date = old_inventory_activity_assignments[inventory_activity_assignment_id][
                'activity_date']

        bulk_update(ss_objects)
        bulk_update(sid_objects)
        bulk_update(old_inventory_activity_assignment_objects)
        InventoryActivityAssignment.objects.bulk_create(new_inventory_activity_assignment_objects)

        # todo: .save() won't be called because we are using bulk update to update the tables. as a consequence dates
        # todo: won't be updated. Find a solution later.

        return ui_utils.handle_response(function, data='success', success=True)
    except Exception as e:
        return ui_utils.handle_response(function, exception_object=e)


def check_valid_codes(payment_status, payment_method, booking_status):
    """
    checks weather these codes are valid or not
    Args:
        payment_status: payment status
        payment_method: payment method
        booking_status: booking status

    Returns:
    """
    function = check_valid_codes.__name__
    try:
        if payment_method and payment_method not in v0_constants.payment_method.values():
            return ui_utils.handle_response(function, data=errors.INVALID_PAYMENT_METHOD_CODE.format(payment_method))
        if payment_status and payment_status not in v0_constants.payment_status.values():
            return ui_utils.handle_response(function, data=errors.INVALID_PAYMENT_STATUS_CODE.format(payment_status))
        if booking_status and booking_status not in v0_constants.booking_status.values():
            return ui_utils.handle_response(function, data=errors.INVALID_BOOKING_STATUS_CODE.format(booking_status))

        return ui_utils.handle_response(function, data='success', success=True)
    except Exception as e:
        return ui_utils.handle_response(function, exception_object=e)


def add_total_price_per_inventory_per_supplier(price_mapping_default_inventories, shortlisted_inventories):
    """
    Args:
        price_mapping_default_inventories: a list of PMD entries for a supplier
        shortlisted_inventories: List of SI.

    Returns: a dict with total supplier price keyed by inventory name.

    """
    function = add_total_price_per_inventory_per_supplier.__name__
    try:
        total_price = 0
        supplier_price_arr = {}

        pmd_inv_to_supplier_price_map = {}
        pmd_inventory_names = set()

        for pmd in price_mapping_default_inventories:
            supplier_price_arr[pmd['inventory_duration']['duration_name']] = pmd
            inventory_name = pmd['inventory_type']['adinventory_name']
            inventory_supplier_price = pmd['actual_supplier_price'] if pmd['actual_supplier_price'] else 0
            if not pmd_inv_to_supplier_price_map.get(inventory_name):
                pmd_inventory_names.add(inventory_name)
                pmd_inv_to_supplier_price_map[inventory_name] = inventory_supplier_price

        result = {}
        for inventory in shortlisted_inventories:

            inventory_name = inventory['inventory_type']['adinventory_name']
           
            if not result.get(inventory_name):
                result[inventory_name] = {}
                result[inventory_name]['inventory_duration_name'] = inventory['inventory_duration']['duration_name']
                
                if supplier_price_arr.get(inventory['inventory_duration']['duration_name']) and supplier_price_arr[inventory['inventory_duration']['duration_name']].get('actual_supplier_price'):
                    result[inventory_name]['actual_supplier_price'] = supplier_price_arr[inventory['inventory_duration']['duration_name']]['actual_supplier_price']
                else:
                    result[inventory_name]['actual_supplier_price'] = 0
                
                #result[inventory_name]['actual_supplier_price'] = pmd_inv_to_supplier_price_map[
                    #inventory_name] if pmd_inv_to_supplier_price_map.get(inventory_name) else 0
                result[inventory_name]['detail'] = []
                result[inventory_name]['total_count'] = 0

                total_price += float(result[inventory_name]['actual_supplier_price'])

            result[inventory_name]['detail'].append(inventory)
            result[inventory_name]['total_count'] += 1            
            
        return ui_utils.handle_response(function, data=(result, total_price), success=True)
    except Exception as e:
        return ui_utils.handle_response(function, exception_object=e)


def get_supplier_price_information(content_type_id_set, supplier_id_set):
    """

    """
    function = get_supplier_price_information.__name__
    try:
        price_objects = PriceMappingDefault.objects.filter(content_type__id__in=content_type_id_set,
                                                           object_id__in=supplier_id_set).select_related(
            'adinventory_type', 'duration_type').values(
            'id',
            'adinventory_type__id',
            'adinventory_type__adinventory_name',
            'adinventory_type__adinventory_type',
            'duration_type__id',
            'duration_type__duration_name',
            'duration_type__days_count',
            'suggested_supplier_price',
            'actual_supplier_price',
            'content_type',
            'object_id',
        )
        price_objects_per_content_type_per_supplier = {}
        for price in price_objects:
            object_id = price['object_id']
            content_type_id = price['content_type']
            try:
                reference = price_objects_per_content_type_per_supplier[content_type_id, object_id]
            except KeyError:
                price_objects_per_content_type_per_supplier[content_type_id, object_id] = []
                reference = price_objects_per_content_type_per_supplier[content_type_id, object_id]

            data = {
                'inventory_type': {'id': price['adinventory_type__id'],
                                   'adinventory_name': price['adinventory_type__adinventory_name'],
                                   'adinventory_type': price['adinventory_type__adinventory_type']},
                'inventory_duration': {'id': price['duration_type__id'],
                                       'duration_name': price['duration_type__duration_name'],
                                       'days_count': price['duration_type__days_count']}
            }

            del price['adinventory_type__id']
            del price['adinventory_type__adinventory_name']
            del price['adinventory_type__adinventory_type']
            del price['duration_type__id']
            del price['duration_type__duration_name']
            del price['duration_type__days_count']

            price = merge_two_dicts(price, data)
            reference.append(price)
        return ui_utils.handle_response(function, data=price_objects_per_content_type_per_supplier, success=True)
    except Exception as e:
        return ui_utils.handle_response(function, exception_object=e)


def get_contact_information(content_type_id_set, supplier_id_set):
    """
    """
    function = get_contact_information.__name__
    try:
        contacts = ContactDetails.objects.filter(content_type__id__in=content_type_id_set,
                                                 object_id__in=supplier_id_set)
        contact_object_per_content_type_per_supplier = {}
        for contact in contacts:
            object_id = contact.object_id
            content_type_object = contact.content_type_id
            if (content_type_object, object_id) not in list(contact_object_per_content_type_per_supplier.keys()):
                contact_object_per_content_type_per_supplier[content_type_object, object_id] = []
            contact_object_per_content_type_per_supplier[content_type_object, object_id].append(model_to_dict(contact))
        return ui_utils.handle_response(function, data=contact_object_per_content_type_per_supplier, success=True)
    except Exception as e:
        return ui_utils.handle_response(function, exception_object=e)


def map_objects_ids_to_objects(mapping):
    """
    Purpose is to map individual object id's to actual objects.
    Args:
        mapping:  a dict of form { 45: [ s1, s2, s3 ], 46: [ 34, 56, 67 ]  }

    Returns: individual ids within each content_type is replaced by actual objects.

    """
    function = map_objects_ids_to_objects.__name__
    try:
        result = {}
        # fetch all content_type_ids
        content_type_ids = ContentType.objects.filter(id__in=list(mapping.keys()))
        # prepare a mapping like { id: content_type_object } for each  of the ids  involved.
        content_type_ids = [content_type.id for content_type in content_type_ids]
        content_type_object_mapping = ContentType.objects.in_bulk(content_type_ids)
        # iterate over mapping.
        for content_type_id, object_ids in mapping.items():
            # fetch all content_type_object
            content_type_object = content_type_object_mapping[content_type_id]
            # fetch the model class.
            model_class = SupplierTypeSociety
            if content_type_object.model != 'suppliertypesociety':
                model_class = SupplierMaster
            
            # fetch all objects.
            my_objects = model_class.objects.filter(supplier_id__in=object_ids)
            if content_type_object.model != 'suppliertypesociety':
                my_objects = SupplierMasterSerializer(my_objects, many=True).data
            else:
                my_objects = SupplierTypeSocietySerializer(my_objects, many=True).data
            # set the new mapping
            result[content_type_id] = my_objects

        output = {}
        for content_type_id, supplier_objects in result.items():
            content_type_object = content_type_object_mapping[content_type_id]
            model_name = content_type_object.model
            # we need to change the keys when we encounter a society
            # if model_name == v0_constants.society_model_name:
            supplier_objects = manipulate_object_key_values(supplier_objects)
            
            # map the extra supplier_specific attributes to content_type, supplier_id
            for supplier in supplier_objects:
                extra_data = {'area': supplier['area'], 'name': supplier['name'], 'subarea': supplier['subarea']}
                output[content_type_id, supplier['supplier_id']] = merge_two_dicts(supplier, extra_data)
        return ui_utils.handle_response(function, data=output, success=True)
    except Exception as e:
        print("e1", e)
        return ui_utils.handle_response(function, exception_object=e)


def get_objects_per_content_type_by_instance(objects):
    """

    Args:
        objects: a list of objects

    Returns: returns a dict having key as content_type and value as list of ids.

    """
    function = get_objects_per_content_type.__name__
    try:
        result = {}
        content_type_set = set()
        supplier_id_set = set()

        for my_object in objects:

            #  key can be both. one from serializer and one directly hitting .values()
            try:
                content_type_id = my_object.content_type.id
            except KeyError:
                content_type_id = my_object.content_type_id
            try:
                object_id = my_object.object_id
            except KeyError:
                object_id = my_object.supplier_id

            if not result.get(content_type_id):
                result[content_type_id] = []
                content_type_set.add(content_type_id)

            result[content_type_id].append(object_id)
            supplier_id_set.add(object_id)

        return ui_utils.handle_response(function, data=(result, content_type_set, supplier_id_set), success=True)
    except Exception as e:
        return ui_utils.handle_response(function, exception_object=e)

def get_objects_per_content_type(objects):
    """
    Args:
        objects: a list of objects
    Returns: returns a dict having key as content_type and value as list of ids.
    """
    function = get_objects_per_content_type.__name__
    try:
        result = {}
        content_type_set = set()
        supplier_id_set = set()

        for my_object in objects:

            #  key can be both. one from serializer and one directly hitting .values()
            try:
                content_type_id = my_object['content_type']
            except KeyError:
                content_type_id = my_object['content_type_id']
            try:
                object_id = my_object['object_id']
            except KeyError:
                object_id = my_object['supplier_id']

            if not result.get(content_type_id):
                result[content_type_id] = []
                content_type_set.add(content_type_id)

            result[content_type_id].append(object_id)
            supplier_id_set.add(object_id)

        return ui_utils.handle_response(function, data=(result, content_type_set, supplier_id_set), success=True)
    except Exception as e:
        return ui_utils.handle_response(function, exception_object=e)

def can_inventory_be_assigned(proposal_release_date, proposal_closure_date, dates):
    """
    returns True or False depending on weather the inventory is assigned or not
    Args:
        proposal_release_date:
        proposal_closure_date:
        dates: a list of sorted dates tuple.
    Returns:

    """
    function = can_inventory_be_assigned.__name__
    try:
        # fetch first date
        first_date_tuple = dates[0]

        # fetch last date
        left_date = first_date_tuple[0]

        # if proposal finishes before the first date of first tuple, it can be assigned that inv.
        if proposal_closure_date < left_date:
            return ui_utils.handle_response(function, data=True, success=True)

        # now we look for gaps between different time intervals. the time gap that satisfy the requirements
        # means the inv_can be assigned. left edge of the time gap.
        x = first_date_tuple[1]

        for date_tuple in dates[1:]:

            # right edge of the time gap
            y = date_tuple[0]

            if proposal_release_date > x and proposal_closure_date < y:
                return ui_utils.handle_response(function, data=True, success=True)

            # the left edge of time gap will be equal to right date of the time gap in next iteration
            x = date_tuple[1]

        if proposal_release_date > x:
            return ui_utils.handle_response(function, data=True, success=True)

        return ui_utils.handle_response(function, data=False, success=True)

    except Exception as e:
        return ui_utils.handle_response(function, exception_object=e)


def get_shortlisted_supplier_mapping(proposal_id):
    """
    maps (content_type, supplier_id) --> ss object
    Args:
        proposal_id:

    Returns: a mapping !

    """
    function = get_shortlisted_supplier_mapping.__name__
    try:
        shortlisted_spaces = ShortlistedSpaces.objects.filter(proposal_id=proposal_id)

        # make shortlisted_spaces mapping of (content_type, object_id) ----> ss object
        shortlisted_spaces_mapping = {}
        for ss in shortlisted_spaces:
            shortlisted_spaces_mapping[ss.content_type, ss.object_id, ss.center_id] = ss

        return ui_utils.handle_response(function, data=shortlisted_spaces_mapping, success=True)

    except Exception as e:
        return ui_utils.handle_response(function, exception_object=e)


def get_shortlisted_inventory_mapping(proposal_id):
    """
    maps inventory_content_type ---->  inventory_ids for this proposal
    Args:
        proposal_id:
    Returns:

    """
    function = get_shortlisted_inventory_mapping.__name__
    try:
        shortlisted_inventories = ShortlistedInventoryPricingDetails.objects.select_related(
            'shortlisted_spaces').filter(shortlisted_spaces__proposal_id=proposal_id)
        inventory_map = {}

        for inventory in shortlisted_inventories:
            content_type = inventory.inventory_content_type
            inventory_id = inventory.inventory_id
            if not inventory_map.get(content_type):
                inventory_map[content_type] = []
            inventory_map[content_type].append(inventory_id)

        return ui_utils.handle_response(function, data=inventory_map, success=True)
    except Exception as e:
        return ui_utils.handle_response(function, exception_object=e)


def prepare_bucket(inventory_name, master_sorted_list_inventories, supplier_ids):
    """
    Prepares the bucket for this inventory
    Args:
        inventory_name: Inventory name
        master_sorted_list_inventories: a list of inventory ids sorted by their assigned proposal count
        supplier_ids: list of supplier ids
    Returns:
    """
    function = prepare_bucket.__name__
    try:
        bucket = {}

        # we receive content type ids. but we want objects. hence storing in set to avoid multiple calls to
        # ContentType table.

        # get inventory_content_type.
        response = ui_utils.get_content_type(inventory_name)
        if not response.data['status']:
            return response
        inventory_content_type = response.data['data']

        model_name = inventory_content_type.model
        model_class = apps.get_model(settings.APP_NAME, model_name)
        inventories = model_class.objects.filter(object_id__in=supplier_ids)

        # fetch only the ids here.
        inventory_ids = [inv.adinventory_id for inv in inventories]

        # create a mapping from inv_id to inv object. can't use in_bulk() because ad inventory_id is not PK.
        inventory_id_to_object_map = {inventory.adinventory_id: inventory for inventory in inventories}

        # get the sorted list
        response = sort_inventory_ids_on_proposal_count(master_sorted_list_inventories, inventory_ids)
        if not response.data['status']:
            return response

        # sorted inventory ids.
        sorted_inventory_ids = response.data['data']

        # get the general object for this inventory
        response = get_inventory_general_data(inventory_name, inventory_content_type)
        if not response.data['status']:
            return response
        inventory_general_data = response.data['data']

        # stores inventory count per bucket id per supplier. Useful in making buckets and assigning frequencies.
        inventory_count_per_bucket_per_supplier = {}
        # inventory_id to tower map.
        inventory_ids_to_tower_id_map = {}

        for inv in inventories:

            content_type = inv.content_type
            object_id = inv.object_id
            response = get_tower_id(inv)

            if not response.data['status']:
                return response
            tower_id = response.data['data']

            if not inventory_ids_to_tower_id_map.get(inv.adinventory_id):
                inventory_ids_to_tower_id_map[inv.adinventory_id] = ''

            try:
                frequency_bucket = inventory_count_per_bucket_per_supplier[content_type, object_id]
            except KeyError:
                inventory_count_per_bucket_per_supplier[content_type, object_id] = {}
                frequency_bucket = inventory_count_per_bucket_per_supplier[content_type, object_id]

            try:
                frequency_bucket[tower_id] += 1
                # increment count of inventories if we encounter more inventories for this bucket
            except KeyError:
                # set inventory count against this bucket to 1 because we just encountered first inventory for this
                # bucket
                frequency_bucket[tower_id] = 1

            inventory_ids_to_tower_id_map[inv.adinventory_id] = tower_id

        for inv_id in sorted_inventory_ids:
            # get the inv object
            inv = inventory_id_to_object_map[inv_id]
            bucket_key = (inv.content_type, inv.object_id, inventory_name)
            bucket_number = inventory_ids_to_tower_id_map[inv_id]

            if bucket_key not in list(bucket.keys()):
                # this is a list of bucket_ids or tower ids for this supplier
                list_of_bucket_ids = list(inventory_count_per_bucket_per_supplier[inv.content_type, inv.object_id].keys())
                bucket_id_to_max_frequency = inventory_count_per_bucket_per_supplier[inv.content_type, inv.object_id]
                # this prepares the bucket based on the above mentioned list
                response = prepare_bucket_per_inventory(inventory_content_type, inventory_name, list_of_bucket_ids,
                                                        bucket_id_to_max_frequency)
                if not response.data['status']:
                    return response
                # a mapping like { 10: { } , 12: { }, 13: { } } is received for each bucket_id.
                buckets_per_supplier_per_inventory = response.data['data']
                bucket[bucket_key] = buckets_per_supplier_per_inventory

            # put the inventory id to the right bucket. one inventory_id can only belong to one bucket only.
            bucket[bucket_key][bucket_number]['inventory_ids'].append(inv.adinventory_id)

        return ui_utils.handle_response(function, data=(bucket, inventory_general_data), success=True)

    except Exception as e:
        return ui_utils.handle_response(function, exception_object=e)


def get_inventory_general_data(inventory_name, inventory_content_type):
    """
    returns inventory_general_data like ad inventory_type, content_type etc.
    Args:
        inventory_name:  The name of inventory in caps.
        inventory_content_type: The inventory_content_type
    Returns:

    """
    function = get_inventory_general_data.__name__
    try:
        inventory_general_data = {}

        inventory_code = v0_constants.inventory_name_to_code[inventory_name]

        if inventory_name == v0_constants.standee_name:
            # general inventory data
            inventory_general_data = {
                'ad_inventory_type': AdInventoryType.objects.get(adinventory_name=v0_constants.standee_name,
                                                                 adinventory_type=v0_constants.default_standee_type),
                'ad_inventory_duration': DurationType.objects.get(
                    duration_name=v0_constants.default_standee_duration_type),
                'inventory_content_type': inventory_content_type,
            }

        elif inventory_name == v0_constants.stall:
            # general inventory data
            inventory_general_data = {
                'ad_inventory_type': AdInventoryType.objects.get(adinventory_name=v0_constants.stall,
                                                                 adinventory_type=v0_constants.default_stall_type),
                'ad_inventory_duration': DurationType.objects.get(
                    duration_name=v0_constants.default_stall_duration_type),
                'inventory_content_type': inventory_content_type,
            }
        elif inventory_name == v0_constants.flier:
            inventory_general_data = {
                'ad_inventory_type': AdInventoryType.objects.get(adinventory_name=v0_constants.flier,
                                                                 adinventory_type=v0_constants.default_flier_type),
                'ad_inventory_duration': DurationType.objects.get(
                    duration_name=v0_constants.default_flier_duration_type),
                'inventory_content_type': inventory_content_type,
            }
        elif inventory_name == v0_constants.poster:
            inventory_general_data = {
                'ad_inventory_type': AdInventoryType.objects.get(adinventory_name=v0_constants.poster,
                                                                 adinventory_type=v0_constants.default_poster_type),
                'ad_inventory_duration': DurationType.objects.get(
                    duration_name=v0_constants.default_poster_duration_type),
                'inventory_content_type': inventory_content_type,
            }
        elif inventory_name == v0_constants.car_display:
            inventory_general_data = {
                'ad_inventory_type': AdInventoryType.objects.get(adinventory_name=v0_constants.car_display,
                                                                 adinventory_type=
                                                                 v0_constants.inventory_type_duration_dict_list[
                                                                     inventory_code][1]),
                'ad_inventory_duration': DurationType.objects.get(
                    duration_name=v0_constants.inventory_type_duration_dict_list[inventory_code][2]),
                'inventory_content_type': inventory_content_type,
            }

        return ui_utils.handle_response(function, data=inventory_general_data, success=True)
    except Exception as e:
        return ui_utils.handle_response(function, exception_object=e)


def get_tower_id(inventory_object):
    """
    returns tower_id of this object. The reason this is a function  because stall inventory for a society  is not
    associated with any tower.
    neither do the gyms and saloons have concept of towers.  Hence they all have a concept of Zero tower. Those inventories
    which have a concept of tower with them are treated differently than those who doesn't.

    Args:
        inventory_object:

    Returns:
    """
    function = get_tower_id.__name__
    try:
        class_name = inventory_object.__class__.__name__
        if class_name == v0_constants.stall_class_name or class_name == v0_constants.flier_class_name:
            return ui_utils.handle_response(function, data=0, success=True)
        elif class_name == v0_constants.standee_class_name or class_name == v0_constants.poster_class_name:
            if not inventory_object.tower:
                return ui_utils.handle_response(function, data=errors.NO_TOWER_ASSIGNED_ERROR.format(class_name,
                                                                                                     inventory_object.adinventory_id))
            return ui_utils.handle_response(function, data=inventory_object.tower_id, success=True)

    except Exception as e:
        return ui_utils.handle_response(function, exception_object=e)


def prepare_bucket_per_inventory(inventory_content_type, inventory_name, list_of_bucket_ids,
                                 bucket_id_to_max_frequency):
    """
    The function that prepares buckets per inventory_content_type
    Args:
        inventory_content_type: The inventory content type.
        inventory_name: The name of the inventory.
        list_of_bucket_ids: List of bucket ids.
        bucket_id_to_max_frequency: { 8: 10, 12: 14 .. } is a map from bucket ids to total number of inventories in that bucket
    Returns: { 0: [ { } ],

    """
    function = prepare_bucket_per_inventory.__name__
    try:
        # default assignment frequency of each bucket.
        assignment_frequency = 1
        if inventory_name == v0_constants.stall:
            assignment_frequency = v0_constants.default_stall_assignment_frequency
        if inventory_name == v0_constants.poster:
            assignment_frequency = v0_constants.default_poster_assignment_frequency
        elif inventory_name == v0_constants.standee_name:
            assignment_frequency = v0_constants.default_standee_assignment_frequency
        elif inventory_name == v0_constants.flier:
            # because all inventories for a supplier must be assigned to a proposal. We know that flier can only have
            # a bucket number as zero.
            assignment_frequency = v0_constants.default_flier_assignment_frequency

        buckets = {}
        inventory_name = inventory_content_type.model
        for bucket_id in list_of_bucket_ids:
            bucket = {
                'bucket_name': inventory_name,
                'assignment_frequency': assignment_frequency,
                'inventory_ids': []
            }
            buckets[bucket_id] = bucket

        return ui_utils.handle_response(function, data=buckets, success=True)
    except Exception as e:
        return ui_utils.handle_response(function, exception_object=e)


def sort_inventory_ids_on_proposal_count(master_inventory_ids, inventory_ids):
    """
    This function changes the sequence of inventories in 'inventory_ids' on basis of proposal count
    Args:
        master_inventory_ids: The sorted list
        inventory_ids:  The sequence of inventory ids.

    Returns: sorted inv_ids based on proposal_count.
    """
    function = sort_inventory_ids_on_proposal_count.__name__
    try:
        result = []
        for inv_id in master_inventory_ids:
            # if inv_id not in inventory_ids, you should continue.
            if inv_id not in inventory_ids:
                continue

            # append the  inv_id to result
            result.append(inv_id)

        # now see those inv_ids which have not been assigned to any proposal. They should sit at front of the result
        for inv_id in inventory_ids:
            if inv_id not in result:
                result.insert(0, inv_id)

        return ui_utils.handle_response(function, data=result, success=True)
    except Exception as e:
        return ui_utils.handle_response(function, exception_object=e)


def get_valid_suppliers(inventory_content_type):
    """

    Args:
        inventory_content_type: The content_type object  of the inventory

    Returns: a list of content_type, supplier_id for which this inventory is allowed.

    """
    function = get_valid_suppliers.__name__
    try:

        model_name = inventory_content_type.model
        model_class = apps.get_model(settings.APP_NAME, model_name)
        all_objects = model_class.objects.all()
        valid_suppliers = set()

        for inv_object in all_objects:
            content_type = inv_object.content_type
            supplier_object = inv_object.object_id
            valid_suppliers.add((content_type, supplier_object))

        return ui_utils.handle_response(function, data=valid_suppliers, success=True)
    except Exception as e:
        return ui_utils.handle_response(function, exception_object=e)


def assign_inventories(bucket_list_per_supplier_per_inventory, shortlisted_supplier, already_shortlisted_inventories,
                       inventory_general_data):
    """

    Args:
        bucket_list_per_supplier_per_inventory: The bucket list for this supplier. Inventories are assigned from this bucket
        shortlisted_supplier: The map of shortlisted object for this proposal
        already_shortlisted_inventories:The list of already shortlisted inventories for this proposal if any.
        inventory_general_data: dict of general data for inv like it's content_type etc.

    Returns: collect objects and returns.

    Assign a particular number of inventories from each bucket to the proposal. This number is dependent on type of
    inventory in question. for standees, it's the length of the bucket. for stall, it's 1. for poster it's 1.
    This number defines how many inventories from that bucket can be assigned to the proposal at a time.
    """
    function = assign_inventories.__name__
    try:
        final_objects = []
        for bucket in bucket_list_per_supplier_per_inventory:

            frequency = bucket['assignment_frequency']
            count = 0

            for inventory_id in bucket['inventory_ids']:
                # this inventory_id can be assigned to this proposal only if it has not been assigned before
                if already_shortlisted_inventories and inventory_id in already_shortlisted_inventories:
                    return ui_utils.handle_response(function, data=[], success=True)

                # we are only allowed to assign 'frequency' number of inventory_ids per bucket to this proposal
                if count < frequency:
                    data = {
                        'inventory_id': inventory_id,
                        'inventory_content_type': inventory_general_data['inventory_content_type'],
                        'ad_inventory_type': inventory_general_data['ad_inventory_type'],
                        'ad_inventory_duration': inventory_general_data['ad_inventory_duration'],
                        'shortlisted_spaces': shortlisted_supplier,
                    }
                    final_objects.append(ShortlistedInventoryPricingDetails(**data))
                    count += 1

        return ui_utils.handle_response(function, data=final_objects, success=True)
    except Exception as e:
        return ui_utils.handle_response(function, exception_object=e)


def make_inventory_assignments(proposal_id, sheet_data, supplier_type_codes):
    """

    Args:
        proposal_id:
        sheet_data: The data we collect when the sheet is imported. determines what suppliers are there in each of the
        inventory_type.
        supplier_type_codes: The set() contains unique supplier types.

    Returns: assigns the inventories to suppliers in sheet.
    """
    function = make_inventory_assignments.__name__
    try:
        inventory_names = set()
        supplier_ids = set()

        # fetch all supplier_type_codes at one place
        response = ui_utils.get_content_types(supplier_type_codes)
        if not response.data['status']:
            return response
        content_types = response.data['data']

        # we should be able to query right dict object by supplying supplier_id, content_type, and inventory_name.
        # this is required to assign fields such as factor, and price from this dict to final object
        suppliers_per_inventory_map = {}

        # we need unique inventory names and supplier_type_codes. unique inventory names because we want to collect each
        # all suppliers belonging to one inventory at one place. This will help in assigning inventory_ids
        # to those suppliers  we need set of supplier type codes so that we can fetch all content types once.

        for inventory in sheet_data:

            # do not process for status = 's'
            if inventory['status'] == v0_constants.shortlisted:
                continue

            inventory_name = inventory['inventory_name']
            supplier_type_code = inventory['supplier_type_code']
            supplier_id = inventory['supplier_id']
            content_type = content_types[supplier_type_code]
            center_id = inventory['center']

            inventory_names.add(inventory['inventory_name'])
            supplier_type_codes.add(inventory['supplier_type_code'])
            supplier_ids.add(inventory['supplier_id'])

            if not suppliers_per_inventory_map.get(inventory_name):
                suppliers_per_inventory_map[inventory_name] = []

            suppliers_per_inventory_map[inventory_name].append((content_type, supplier_id, center_id))

        # this has to be an atomic transaction
        with transaction.atomic():
            output = []
            response = get_shortlisted_supplier_mapping(proposal_id)
            if not response.data['status']:
                return response
            # this is a mapping of content_type, supplier_id, center_id --> ss object. Useful when SID object is created.
            shortlisted_suppliers_mapping = response.data['data']

            response = get_shortlisted_inventory_mapping(proposal_id)
            if not response.data['status']:
                return response
            # this is a list of content_type --> inventory_id mapping of those inventories which already have been
            # assigned to this proposal before. This is used in order to check not to assign the same inventory again
            # to that proposal
            already_shortlisted_inventory_mapping = response.data['data']

            sorted_shortlisted_inv_pricing_objects = ShortlistedInventoryPricingDetails.objects.values('inventory_id',
                                                                                                       'inventory_content_type_id').annotate(
                pcount=Count('shortlisted_spaces__proposal_id')).order_by('pcount')
            master_sorted_list_inventories = [inventory['inventory_id'] for inventory in
                                              sorted_shortlisted_inv_pricing_objects]

            # process each inventory one by one and assign the ids.
            for inventory_name in inventory_names:

                if inventory_name in v0_constants.inv_not_implemented:
                    continue

                # get inventory_content_type.
                response = ui_utils.get_content_type(inventory_name)
                if not response.data['status']:
                    return response
                inventory_content_type = response.data['data']

                # prepare the bucket for this inventory
                response = prepare_bucket(inventory_name, master_sorted_list_inventories, supplier_ids)
                if not response.data['status']:
                    return response

                # this is a mapping (content_type, supplier_id) --> [ { } , { } ] where each { } is a bucket.
                # A bucket contains inventory_ids which you want to assign to a proposal. Only a fixed number of
                # inventory can be assigned to a proposal. bucket also stores extra info like ad_inv_id, duration
                # object etc.
                bucket_list_per_supplier_per_inventory, inventory_general_data = response.data['data']

                # this is mapping of content_type, supplier_id --> [ id1, id2 ] which is inventory_id from master
                # inventory table. These are the suppliers for this this inventory is allowed.
                response = get_valid_suppliers(inventory_content_type)
                if not response.data['status']:
                    return response
                valid_suppliers = response.data['data']

                # if the supplier has that inventory mapped and it's also been in the sheet list, we try to assign
                # the inventory to that supplier.
                for supplier_tuple in suppliers_per_inventory_map[inventory_name]:

                    if (supplier_tuple[0], supplier_tuple[1]) in valid_suppliers:

                        if supplier_tuple not in list(shortlisted_suppliers_mapping.keys()):
                            return ui_utils.handle_response(function,
                                                            data='This supplier is not shortlisted yet {0}'.format(
                                                                supplier_tuple[1]))

                        response = assign_inventories(bucket_list_per_supplier_per_inventory[
                                                          supplier_tuple[0], supplier_tuple[
                                                              1], inventory_name].values(),
                                                      shortlisted_suppliers_mapping[supplier_tuple],
                                                      already_shortlisted_inventory_mapping.get(inventory_content_type),
                                                      inventory_general_data)
                        if not response.data['status']:
                            return response
                        if response.data['data']:
                            output.extend(response.data['data'])

            message = 'success'
            # if not inventories assigned, means a problem is there.
            if not len(output):
                message = 'No inventories being assigned for this proposal. Please check the error logs'
            # issue a single insert statements. be aware of disadvantages of .bulk_create usage.
            ShortlistedInventoryPricingDetails.objects.bulk_create(output)
            return message
    except Exception as e:
        return ui_utils.handle_response(function, exception_object=e)


def get_campaign_status(proposal_start_date, proposal_end_date):
    """
    Args:
        proposal_start_date: The start date
        proposal_end_date: The end date

    Returns:     returns 'completed', 'running', and 'upcoming' based on date
    """
    function = get_campaign_status.__name__
    try:
        if not proposal_end_date or not proposal_start_date:
            return ui_utils.handle_response(function, data=v0_constants.unknown, success=True)

        current_date = timezone.now()

        if proposal_end_date < current_date:
            return ui_utils.handle_response(function, data=v0_constants.completed, success=True)
        if proposal_start_date > current_date:
            return ui_utils.handle_response(function, data=v0_constants.upcoming, success=True)
        return ui_utils.handle_response(function, data=v0_constants.running, success=True)

    except Exception as e:
        return ui_utils.handle_response(function, exception_object=e)


def do_dates_overlap(base_start_date_1, base_end_date_1, target_start_date_1, target_end_date_2):
    """

    Args:
        base_start_date_1:
        base_end_date_1:
        target_start_date_1:
        target_end_date_2:

    Returns:

    """
    function = do_dates_overlap.__name__
    try:
        # check for valid dates.
        if base_end_date_1 < base_start_date_1:
            return ui_utils.handle_response(function, data=errors.INVALID_START_END_DATE.format(base_start_date_1,
                                                                                                base_end_date_1))

        if target_end_date_2 < target_start_date_1:
            return ui_utils.handle_response(function, data=errors.INVALID_START_END_DATE.format(target_start_date_1,
                                                                                                target_end_date_2))

        # now test for overlap
        if base_end_date_1 < target_start_date_1:
            return ui_utils.handle_response(function, data=False, success=True)

        if base_start_date_1 > target_end_date_2:
            return ui_utils.handle_response(function, data=False, success=True)

        return ui_utils.handle_response(function, data=True, success=True)

    except Exception as e:
        return ui_utils.handle_response(function, exception_object=e)


def get_overlapping_campaigns(proposal):
    """
    """
    function = get_overlapping_campaigns.__name__
    try:
        start_date = proposal.tentative_start_date
        end_date = proposal.tentative_end_date

        overlapping_campaigns = []
        campaigns = ProposalInfo.objects.filter(campaign_state=v0_constants.proposal_converted_to_campaign,
                                                invoice_number__isnull=False).exclude(pk=proposal.pk)

        for campaign in campaigns:
            target_start_date = campaign.tentative_start_date
            target_end_date = campaign.tentative_end_date
            # if the campaign does not have dates, it's an error. a campaign must have dates
            if not target_start_date or not target_end_date:
                return ui_utils.handle_response(function, data=errors.NO_DATES_ERROR.format(campaign.proposal_id))

            # check for overlapping
            response = do_dates_overlap(start_date, end_date, target_start_date, target_end_date)
            if not response.data['status']:
                return response
            verdict = response.data['data']
            if verdict:
                overlapping_campaigns.append(campaign)

        return ui_utils.handle_response(function, data=overlapping_campaigns, success=True)
    except Exception as e:
        return ui_utils.handle_response(function, exception_object=e)


def book_inventories(current_inventories_map, already_inventories_map):
    """

    Args:
        current_inventories_map:
        already_inventories_map:

    Returns: a list of SID objects if no object overlaps with already booked inventories.

    """
    function = book_inventories.__name__
    try:
        booked_inventories = []
        inv_errors = []
        inventory_release_closure_list = []
        for inv_tuple, inv_date_tuple in current_inventories_map.items():

            release_date = inv_date_tuple[0]
            closure_date = inv_date_tuple[1]
            inv_obj = inv_date_tuple[2]

            inv_obj.release_date = release_date
            inv_obj.closure_date = closure_date
            # todo: removing date checking logic for now. even if dates overlap we book it anyway. Insert the logic when required in future
            # if this inventory is not found in already_booked_inventories map, book this inventory.
            try:
                already_book_inv_map_reference = already_inventories_map[inv_tuple]
                is_overlapped = False

                for already_booked_inv_tuple in already_book_inv_map_reference:

                    target_release_date = already_booked_inv_tuple[0]
                    target_closure_date = already_booked_inv_tuple[1]
                    target_proposal_id = already_booked_inv_tuple[2]

                    response = do_dates_overlap(release_date, closure_date, target_release_date, target_closure_date)
                    if not response.data['status']:
                        return response
                    verdict = response.data['data']
                    # if the dates overlap we don't book this inventory
                    if verdict:
                        is_overlapped = True
                        inv_errors.append(
                            errors.DATES_OVERLAP_ERROR.format(inv_obj.inventory_id, release_date, closure_date,
                                                              target_release_date, target_closure_date,
                                                              target_proposal_id))

                is_overlapped = False  # remove this line when you need to book by dates
                # if the inv does not overlaps with any of the pre booked versions of it, we book it.
                if not is_overlapped:
                    # we collect this info as we have to insert RD and CD for this inventory later
                    inventory_release_closure_list.append((inv_obj, release_date, closure_date))
                    # book it
                    booked_inventories.append(inv_obj)
            except KeyError:
                # we collect this info as we have to insert RD and CD for this inventory later
                inventory_release_closure_list.append((inv_obj, release_date, closure_date))
                # means this inventory is already not booked. hence we book it.
                booked_inventories.append(inv_obj)

        return ui_utils.handle_response(function, data=(booked_inventories, inventory_release_closure_list, inv_errors),
                                        success=True)
    except Exception as e:
        return ui_utils.handle_response(function, exception_object=e)


def get_amenities_suppliers(supplier_type_code, amenities):
    """
    returns a list of supplier_ids which have one of the amenities in the list of amenities.
    Args:
        supplier_type_code: 'RS', 'CP'
        amenities: [ .. ]

    Returns: a list  of suppliers
    """
    function = get_amenities_suppliers.__name__
    try:
        for amenity in amenities:
            if amenity not in list(v0_constants.valid_amenities.keys()):
                return ui_utils.handle_response(function, data=errors.INVALID_AMENITY_CODE_ERROR.format(amenity))

        response = ui_utils.get_content_type(supplier_type_code)
        if not response.data['status']:
            return response
        content_type = response.data['data']

        supplier_amenity_count_list = SupplierAmenitiesMap.objects.filter(content_type=content_type,
                                                                          amenity__code__in=amenities).values(
            'object_id').annotate(amenity_count=Count('amenity'))
        return supplier_amenity_count_list

    except Exception as e:
        raise Exception(function, ui_utils.get_system_error(e))


def save_amenities_for_supplier(supplier_type_code, supplier_id, amenities):
    """
    save amenities for supplier
    Args:
        amenities: [ .. ]

    Returns:
    """
    function = save_amenities_for_supplier.__name__
    try:
        response = ui_utils.get_content_type(supplier_type_code)
        if not response.data['status']:
            return response
        content_type = response.data['data']
        # container to store amenities for supplier
        total_amenities = []
        amenity_ids = [amenity['id'] for amenity in amenities]
        amenity_objects_map = Amenity.objects.in_bulk(amenity_ids)

        for amenity in amenities:
            amenity_id = amenity['id']
            data = {
                'object_id': supplier_id,
                'amenity': amenity_objects_map[amenity_id],
                'content_type': content_type,
                'comments': amenity.get('comments')
            }
            total_amenities.append(SupplierAmenitiesMap(**data))

        # delete previous  shortlisted suppliers and save new
        SupplierAmenitiesMap.objects.filter(object_id=supplier_id, content_type=content_type).delete()
        SupplierAmenitiesMap.objects.bulk_create(total_amenities)

        return ui_utils.handle_response(function, data='success', success=True)
    except Exception as e:
        return ui_utils.handle_response(function, exception_object=e)


def get_inventory_activity_image_objects(shortlisted_inventory_to_audit_count_map, image_activity_to_audit_count_map):
    """
    prepares  inventory activity image objects
    Args:

    Returns:

    """
    function = get_inventory_activity_image_objects.__name__
    try:
        inv_act_objects = []
        for inventory_global_id, audit_count in shortlisted_inventory_to_audit_count_map.items():

            if not image_activity_to_audit_count_map.get(inventory_global_id):
                inv_act_objects.append(InventoryActivityImage(shortlisted_inventory_details_id=inventory_global_id,
                                                              activity_type=v0_constants.activity_type['RELEASE']))
                inv_act_objects.append(InventoryActivityImage(shortlisted_inventory_details_id=inventory_global_id,
                                                              activity_type=v0_constants.activity_type['CLOSURE']))
                image_activity_to_audit_count_map[inventory_global_id] = 0

            actual_audit_count = audit_count - image_activity_to_audit_count_map[inventory_global_id]
            for count in actual_audit_count:
                inv_act_objects.append(InventoryActivityImage(shortlisted_inventory_details_id=inventory_global_id,
                                                              activity_type=v0_constants.activity_type['AUDIT']))

        return ui_utils.handle_response(function, data=inv_act_objects, success=True)
    except Exception as e:
        return ui_utils.handle_response(function, exception_object=e)


def get_possible_activity_count(proposal_id):
    """
    returns possible counts of activities over inventories for a given proposal. The number of releases under a
    proposal is same as inventory count under that proposal. same thing for number of closures. Number of audits
    is number of rows in audits for all inventories under that proposal in audt date table.

    Args:
        proposal_id: The proposal id

    Returns: a dict in which key is act name and value is respective count of that act possible according to system.
    """
    function = get_possible_activity_count.__name__
    try:

        sql = " select a.activity_type, count(b.activity_date) as count from inventory_activity as a  " \
              " INNER JOIN inventory_activity_assignment as b on a.id = b.inventory_activity_id " \
              " INNER JOIN shortlisted_inventory_pricing_details as c on c.id = a.shortlisted_inventory_details_id" \
              " INNER JOIN shortlisted_spaces as d on d.id = c.shortlisted_spaces_id WHERE d.proposal_id = %s " \
              " group by a.activity_type"

        with connection.cursor() as cursor:

            cursor.execute(sql, [proposal_id])
            act_data = dict_fetch_all(cursor)

        # act_data will be an array of dicts. converting to just a dict with key as act name
        data = {}
        for item in act_data:
            data[item['activity_type']] = item['count']

        return ui_utils.handle_response(function, data=data, success=True)
    except Exception as e:
        return ui_utils.handle_response(function, exception_object=e)


def get_actual_activity_count(proposal_id):
    """
    returns actual activity count. This data lives in inventory image activity table where only then entry
    is made when user actually takes the image and uploads it to the server.
    Args:
        proposal_id:

    Returns: a dict with keys as act names and values as counts of actual acts performed

    """
    function = get_actual_activity_count.__name__
    try:

        sql = " select a.activity_type, count(b.activity_date) as count from inventory_activity as a " \
              " INNER JOIN inventory_activity_assignment as b on a.id = b.inventory_activity_id " \
              " INNER JOIN shortlisted_inventory_pricing_details as c on c.id = a.shortlisted_inventory_details_id " \
              " INNER JOIN shortlisted_spaces as d on d.id = c.shortlisted_spaces_id " \
              " INNER JOIN inventory_activity_image as e  on e.inventory_activity_assignment_id = b.id " \
              " WHERE d.proposal_id = %s group by a.activity_type "

        with connection.cursor() as cursor:
            cursor.execute(sql, [proposal_id])
            act_data = dict_fetch_all(cursor)

        # act_data will be an array of dicts. converting to just a dict with key as act name
        data = {}
        for item in act_data:
            data[item['activity_type']] = item['count']

        return ui_utils.handle_response(function, data=data, success=True)
    except Exception as e:
        return ui_utils.handle_response(function, exception_object=e)


def dict_fetch_all(cursor):
    """
    Args:
        cursor: database cursor

    Returns:     "Return all rows from a cursor as a dict"

    """
    function = dict_fetch_all.__name__
    try:
        columns = [col[0] for col in cursor.description]
        return [
            dict(zip(columns, row))
            for row in cursor.fetchall()
        ]
    except Exception as e:
        raise Exception(function, ui_utils.get_system_error(e))


def construct_date_range_query(database_field):
    """
    Args:
        database_field: The db field to query upon

    Returns: Q object
    """
    function = construct_date_range_query.__name__
    try:
        # The database field must be within delta days of the current date
        current_date = timezone.now().date()
        previous_date = current_date - timezone.timedelta(days=7)
        first_query = Q(**{database_field + '__gte': previous_date})
        second_query = Q(**{database_field + '__lte': current_date + timezone.timedelta(days=v0_constants.delta_days)})
        return first_query & second_query
    except Exception as e:
        raise Exception(function, ui_utils.get_system_error(e))


def insert_release_closure_dates(inventory_release_closure_list):
    """
    inserts release date and closure dates for all inventories in the shortlisted_inventories list
    Args:
        inventory_release_closure_list: a list of tuples where each tuple describes an inventory instance, R.D and
        C.D. This is so because in theory, each inventory can have different RD and CD.

    Returns: success in case insert happens successfully
    """
    function = insert_release_closure_dates.__name__
    try:
        # todo: using get_or_create in a loop can be a bottleneck as it hits db each time. this is simplest solution
        # need to improve if api response takes too long because of this
        inventory_activity_instance_list = []
        for shortlisted_inv_instance, release_date, closure_date in inventory_release_closure_list:
            # RELEASE
            inv_activity_instance, is_created = InventoryActivity.objects.get_or_create(
                shortlisted_inventory_details=shortlisted_inv_instance,
                activity_type=INVENTORY_ACTIVITY_TYPES[0][0]
            )

            inventory_activity_instance_list.append((inv_activity_instance, release_date.date()))

            # CLOSURE
            inv_activity_instance, is_created = InventoryActivity.objects.get_or_create(
                shortlisted_inventory_details=shortlisted_inv_instance,
                activity_type=INVENTORY_ACTIVITY_TYPES[1][0]
            )

            inventory_activity_instance_list.append((inv_activity_instance, closure_date.date()))

            # AUDIT. No dates are put against audit. it can contain many dates, that's why
            inv_activity_instance, is_created = InventoryActivity.objects.get_or_create(
                shortlisted_inventory_details=shortlisted_inv_instance,
                activity_type=INVENTORY_ACTIVITY_TYPES[2][0]
            )

        for inv_activity_instance, activity_date in inventory_activity_instance_list:
            InventoryActivityAssignment.objects. \
                get_or_create(inventory_activity=inv_activity_instance, activity_date=activity_date)

        return ui_utils.handle_response(function, data='success', success=True)
    except Exception as e:
        return ui_utils.handle_response(function, exception_object=e)


def get_societies_within_tenants_flat_ratio(min_ratio, max_ratio, query=Q()):
    """
    filters  societies by calculating a ratio of tenant to flat and ensuring the ratio within a range of
    min max.
    Args:
        min_ratio: min value
        max_ratio: max_value
        query: a custom query if any

    Returns: filtered supplier ids

    """
    function = get_societies_within_tenants_flat_ratio.__name__
    try:
        supplier_ids = SupplierTypeSociety.objects.filter(query).annotate(ratio=ExpressionWrapper(
            F('total_tenant_flat_count') / F('flat_count'), output_field=FloatField())).filter(ratio__gte=min_ratio,
                                                                                               ratio__lte=max_ratio).values_list(
            'supplier_id', flat=True)
        return supplier_ids
    except Exception as e:
        raise Exception(function, ui_utils.get_system_error(e))


def get_standalone_societies(query=Q()):
    """
    Returns: returns a list of societies that are standalone

    """
    function = get_standalone_societies.__name__
    try:
        response = ui_utils.get_content_type(v0_constants.society)
        if not response.data['status']:
            return response
        content_type = response.data['data']
        amenities = SupplierAmenitiesMap.objects.filter(content_type=content_type).values('object_id').annotate(
            amenity_count=Count('id'))
        amenities_map = {amenity['object_id']: amenity['amenity_count'] for amenity in amenities}
        societies = SupplierTypeSociety.objects.filter(query).values('supplier_id', 'tower_count', 'flat_count')
        societies_map = {
            society['supplier_id']:
                {
                    'tower_count': society['tower_count'],
                    'flat_count': society['flat_count'],
                    'amenity_count': amenities_map.get(society['supplier_id'])

                } for society in societies}
        return societies_map
    except Exception as e:
        raise Exception(function, ui_utils.get_system_error(e))


def is_society_standalone(instance_detail):
    """
    checks if society instance is standalone or not as per definition of standalone society
        Args:
        instance_detail:

    Returns: True or False
    """
    function = is_society_standalone.__name__
    try:
        if not instance_detail:
            return False

        if instance_detail['tower_count'] <= v0_constants.standalone_society_config['tower_count'] and instance_detail[
            'flat_count'] <= v0_constants.standalone_society_config['flat_count'] and instance_detail[
            'amenity_count'] <= v0_constants.standalone_society_config['amenity_count']:
            return True
        return False
    except Exception as e:
        raise Exception(function, ui_utils.get_system_error(e))


def start_download_from_amazon(proposal_id, image_map):
    """
    function responsible for downloading images from amazon and storing in a folder under files directory.
    Args:
        proposal_id: The proposal_id
        image_map: The image_map in which images are stored against each supplier_id

    Returns: task id
    """
    function = start_download_from_amazon.__name__
    # try:
    image_map = json.loads(image_map)
    # create this path before calling util to download
    path_to_master_dir = settings.BASE_DIR + '/files/downloaded_images/' + proposal_id

    if os.path.exists(path_to_master_dir):
        # shutil.rmtree(path_to_master_dir)
        os.system("rm -rf "+path_to_master_dir)

    os.system("mkdir "+path_to_master_dir)
    # os.makedirs(path_to_master_dir)

    sub_tasks = []
    for supplier_key, image_name_list in image_map.items():
        supplier_id = supplier_key.split('_')[
            0]  # supplier_id and content_type_id are joined by '_'. splitting and getting first value
        # call util function to download from amazon
        sub_tasks.append(subtask(tasks.bulk_download_from_amazon_per_supplier,
                                    args=[path_to_master_dir + '/' + supplier_id, image_name_list]))

    job = group(sub_tasks)
    result = job.apply_async()
    result.save()
    return result.id
    # except Exception as e:
    #     raise Exception(function, ui_utils.get_system_error(e))

def get_random_pattern(size=v0_constants.pattern_length, chars=string.ascii_uppercase + string.digits):
    function = get_random_pattern.__name__
    try:
        return ''.join(random.choice(chars) for _ in range(size))
    except Exception as e:
        raise Exception(function, ui_utils.get_system_error(e))


def expand_supplier_id(supplier_id):
    """
    expands supplier id into it's constituents
    Args:
        supplier_id:

    Returns:

    """
    function = expand_supplier_id.__name__
    try:
        # MUM AE BN RS KML
        data = {
            'city_code': supplier_id[0:3],
            'area_code': supplier_id[3:5],
            'subarea_code': supplier_id[5:7],
            'supplier_type_code': supplier_id[7:9],
            'supplier_code': supplier_id[9:12]
        }
        return data
    except Exception as e:
        raise Exception(function, ui_utils.get_system_error(e))


def generate_supplier_basic_sheet_mail(data):
    """

    Args:
        data:

    Returns:

    """
    function = generate_supplier_basic_sheet_mail.__name__
    try:

        workbook = Workbook()
        sheet_name = data['sheet_name']
        headers = data['headers']

        # create a new sheet for each supplier type
        ws = workbook.create_sheet(index=0, title=sheet_name)

        # set the heading
        ws.append(headers)

        for supplier_object in data['suppliers']:
            ws.append([supplier_object.get(key) for key in data['data_keys']])
        file_name = os.path.join(settings.BASE_DIR, v0_constants.all_supplier_data_file_name)
        workbook.save(file_name)
        return file_name

    except Exception as e:
        raise Exception(function, ui_utils.get_system_error(e))


def validate_supplier_headers(supplier_type_code, row, data_import_type):
    """
    Returns: True means the headers defined in constants exactly match the headers defined in sheet.
    """
    function = validate_supplier_headers.__name__
    try:
        lowercase_sheet_header_list_with_underscores = ['_'.join(field.value.lower().split(' ')) for field in row if
                                                        field.value]

        supplier_headers_per_import_type = v0_constants.supplier_headers[data_import_type]
        basic_headers = supplier_headers_per_import_type['basic_data']

        supplier_specific_headers = supplier_headers_per_import_type['supplier_specific'].get(supplier_type_code)
        amenity_headers = supplier_headers_per_import_type['amenities']
        event_headers = supplier_headers_per_import_type['events']
        flat_headers = supplier_headers_per_import_type['flats']

        all_header_list = basic_headers + amenity_headers + event_headers + flat_headers

        if supplier_specific_headers:
            all_header_list += supplier_specific_headers

        # convert each one to standard underscore format
        all_header_list = ['_'.join(current_header.lower().split(' ')) for current_header in all_header_list]
        # check what we have from sheet
        for current_header in lowercase_sheet_header_list_with_underscores:
            # check weather the header in the sheet is with us or not
            if current_header not in all_header_list:
                raise Exception(function, errors.HEADER_NOT_PRESENT_IN_SYSTEM.format(current_header))
        return True
    except Exception as e:
        raise Exception(function, ui_utils.get_system_error(e))


def collect_supplier_common_data(result, supplier_type_code, supplier_id, row_dict, data_import_type):
    """

    Args:
        result:
        supplier_type_code:
        row_dict:
        supplier_id:
        data_import_type

    Returns: popuates 'common_data' key of result dict

    """
    function = collect_supplier_common_data.__name__
    try:
        common_data_key = 'common_data'
        basic_headers = v0_constants.supplier_headers[data_import_type]['basic_data']
        supplier_specific_headers = v0_constants.supplier_headers[data_import_type]['supplier_specific'].get(
            supplier_type_code)
        if supplier_specific_headers:
            basic_headers += supplier_specific_headers
        common_headers = ['_'.join(field.lower().split(' ')) for field in basic_headers]
        for key in common_headers:
            result[supplier_id][common_data_key][key] = row_dict.get(key)
        return result
    except Exception as e:
        raise Exception(function, ui_utils.get_system_error(e))


def collect_amenity_data(result, supplier_id, row_dict):
    """

    Args:
        result:
        supplier_id:
        row_dict:

    Returns:

    """
    function = collect_amenity_data.__name__
    try:
        valid_amenities = list(v0_constants.valid_amenities.keys())
        positive_amenities_list = []
        negative_amenity_list = []
        for amenity in valid_amenities:

            key = 'amenity_' + '_'.join(amenity.lower().split(' ')) + '_present'
            if not row_dict[key]:
                continue
            if row_dict[key].lower() in [item.lower() for item in v0_constants.positive]:
                positive_amenities_list.append(amenity)
            elif row_dict[key].lower() in [item.lower() for item in v0_constants.negative]:
                negative_amenity_list.append(amenity)
            else:
                pass

        result[supplier_id]['amenities']['positive']['names'] = positive_amenities_list
        result[supplier_id]['amenities']['negative']['names'] = negative_amenity_list

        return result
    except Exception as e:
        raise Exception(function, ui_utils.get_system_error(e))


def collect_events_data(result, supplier_id, row_dict):
    """

    Args:
        result:
        supplier_id:
        row_dict:

    Returns:

    """
    function = collect_events_data.__name__
    try:
        valid_events = v0_constants.valid_events
        positive_events = []
        negative_events = []
        for event in valid_events:
            key = 'event_' + '_'.join(event.lower().split(' '))
            if not row_dict[key]:
                continue
            if row_dict[key].lower() in [item.lower() for item in v0_constants.positive]:
                positive_events.append(event)
            elif row_dict[key].lower() in [item.lower() for item in v0_constants.negative]:
                negative_events.append(event)
            else:
                pass

        result[supplier_id]['events']['positive']['names'] = positive_events
        result[supplier_id]['events']['negative']['names'] = negative_events

        return result
    except Exception as e:
        raise Exception(function, ui_utils.get_system_error(e))


def collect_flat_data(result, supplier_id, row_dict):
    """

    Args:
        result:
        supplier_id:
        row_dict:

    Returns:

    """
    function = collect_flat_data.__name__
    try:
        all_flat_types = v0_constants.flat_type_dict.values()
        for flat_type in all_flat_types:
            flat_key = '_'.join(flat_type.lower().split(' '))
            key = 'flat_' + flat_key + '_present'
            if not row_dict[key]:
                continue
            if row_dict[key].lower() in [item.lower() for item in v0_constants.positive]:
                count_key = flat_key + '_count'
                size_key = flat_key + '_size'
                rent_key = flat_key + '_rent'
                # if not row_dict[count_key] or not row_dict[size_key] or not row_dict[rent_key]:
                #    raise Exception(function, errors.COUNT_SIZE_RENT_VALUE_NOT_PRESENT.format(count_key, size_key, rent_key, supplier_id))
                result[supplier_id]['flats']['positive'][flat_type] = {
                    'count': row_dict[count_key],
                    'size': row_dict[size_key],
                    'rent': row_dict[rent_key]
                }
            elif row_dict[key].lower() in [item.lower() for item in v0_constants.negative]:
                result[supplier_id]['flats']['negative'][flat_type] = {}
            else:
                pass

        return result
    except Exception as e:
        raise Exception(function, ui_utils.get_system_error(e))


def handle_supplier_data_from_sheet(result, supplier_instance_map, content_type, supplier_type_code):
    """
    updates, creates or deletes required instances as per the case for each supplier.
    Args:
        result: a dict having all the data that is to be handled into database.
        supplier_instance_map: a map of supplier_id --> instance.
        content_type: The content type of this supplier.
        supplier_type_code:
        cases handled:
           amenities, Flats, events.
    Returns: a dict showing summary of counts of each distinct type of object created, updated or deleted.
    """
    function = handle_supplier_data_from_sheet.__name__
    try:
        # define required variables.
        supplier_instance_list = []
        supplier_amenity_instance_list = []
        positive_events = []
        negative_events = []
        positive_created_flats = []
        negative_flats = []
        negative_amenity_instances = []
        positive_updated_flats = []
        tower_instance_list = []

        event_instance_map = {}
        supplier_amenity_instances_map = {}
        flat_instance_map = {}

        amenities = Amenity.objects.all()
        amenity_map = {amenity.name: amenity for amenity in amenities}
        supplier_ids = list(result.keys())

        # prepare a map of key --> instance. This is required to fetch the instance once we know the key. key is different for
        # different types of instance.
        supplier_amenity_instances = SupplierAmenitiesMap.objects.filter(object_id__in=supplier_ids,
                                                                         content_type=content_type)
        for instance in supplier_amenity_instances:
            key = (instance.object_id, instance.content_type, instance.amenity)
            try:
                ref = supplier_amenity_instances_map[key]
            except KeyError:
                supplier_amenity_instances_map[key] = instance

        event_instances = Events.objects.filter(object_id__in=supplier_ids, content_type=content_type)
        for instance in event_instances:
            key = (instance.event_name, instance.object_id, instance.content_type)
            try:
                ref = event_instance_map[key]
            except KeyError:
                event_instance_map[key] = instance

        flat_instances = FlatType.objects.filter(object_id__in=supplier_ids, content_type=content_type)
        for instance in flat_instances:
            key = (instance.flat_type, instance.object_id, instance.content_type)
            try:
                ref = flat_instance_map[key]
            except KeyError:
                flat_instance_map[key] = instance

        # collect actual tower count by looking into society tower table
        tower_counts_list = SocietyTower.objects.filter(object_id__in=supplier_ids, content_type=content_type).values(
            'object_id').annotate(count=Count('tower_id'))
        tower_count_map = {item['object_id']: item['count'] for item in tower_counts_list}

        supplier_id = ''

        # once the maps are prepared we loop over each supplier and collect the required objects on the fly.
        # we update, create or delete in bulk outside of the loop.
        count = 0
        for supplier_id, detail in result.items():
            count = count + 1
            instance = supplier_instance_map[supplier_id]
            if supplier_type_code == v0_constants.society_code:
                # get additional tower instance to be added if any first before setting new attributes
                tower_created_list = handle_society_towers(instance, detail, tower_count_map, content_type)
                if tower_created_list:
                    tower_instance_list.extend(tower_created_list)

            instance = handle_supplier_common_attributes(instance, detail, supplier_type_code)
            supplier_instance_list.append(instance)

            positive_instances, negative_instances = handle_amenities(supplier_id, result, amenity_map,
                                                                      supplier_amenity_instances_map, content_type)
            if positive_instances:
                supplier_amenity_instance_list.extend(positive_instances)
            if negative_instances:
                negative_amenity_instances.extend(negative_instances)

            positive_instances, negative_instances = handle_events(supplier_id, result, event_instance_map,
                                                                   content_type)
            if positive_instances:
                positive_events.extend(positive_instances)
            if negative_instances:
                negative_events.extend(negative_instances)

            positive_created_instances, positive_updated_instances, negative_instances = handle_flats(supplier_id,
                                                                                                      result,
                                                                                                      flat_instance_map,
                                                                                                      content_type)
            if positive_created_instances:
                positive_created_flats.extend(positive_created_instances)
            if positive_updated_instances:
                positive_updated_flats.extend(positive_updated_instances)
            if negative_instances:
                negative_flats.extend(negative_instances)

        # bulk update suppliers and flats
        bulk_update(supplier_instance_list)
        bulk_update(positive_updated_flats)

        SupplierAmenitiesMap.objects.bulk_create(supplier_amenity_instance_list)
        Events.objects.bulk_create(positive_events)
        FlatType.objects.bulk_create(positive_created_flats)
        SocietyTower.objects.bulk_create(tower_instance_list)

        for instance in negative_amenity_instances:
            instance.delete()
        for instance in negative_events:
            instance.delete()
        for instance in negative_flats:
            instance.delete()

        debug = {

            'total_suppliers_updated': len(supplier_instance_list),
            'total_supplier_amenity_instance_created': len(supplier_amenity_instance_list),
            'total_supplier_amenity_instance_deleted': len(negative_amenity_instances),
            'total_events_created': len(positive_events),
            'total_events_deleted': len(negative_events),
            'total_flats_created': len(positive_created_flats),
            'total_flats_updated': len(positive_updated_flats),
            'total_flats_deleted': len(negative_flats),
            'total_towers_created': len(tower_instance_list)
        }
        return debug
    except Exception as e:
        raise Exception(function, ui_utils.get_system_error(e))


def handle_society_towers(instance, detail, tower_count_map, content_type):
    """
    Args:
        instance:
        detail:
        content_type:
        tower_count_map:

    Returns:

    """
    function = handle_society_towers.__name__
    try:
        positive_towers = []
        given_tower_count = detail['common_data']['tower_count']

        if not given_tower_count:
            return []

        given_tower_count = int(given_tower_count)

        # if we don't find the instance pk, the current tower count is zero
        current_tower_count = tower_count_map.get(instance.pk, 0)

        if given_tower_count == current_tower_count:
            return []

        if given_tower_count > current_tower_count:
            extra_towers = given_tower_count - current_tower_count
            for i in range(extra_towers):
                positive_towers.append(SocietyTower(object_id=instance.pk, content_type=content_type))

        return positive_towers
    except Exception as e:

        raise Exception(function, ui_utils.get_system_error(e))


def handle_supplier_common_attributes(instance, detail, supplier_type_code):
    """
    instance:  the supplier instance
    detail: a dict having details to be set to this instance

    Returns: updated instance

    """
    function = handle_supplier_common_attributes.__name__
    try:
        common_data_key = 'common_data'
        if supplier_type_code == v0_constants.society_code:

            for society_db_field, input_key in v0_constants.society_db_field_to_input_field_map.items():
                if not detail[common_data_key][input_key]:
                    continue
                setattr(instance, society_db_field, detail[common_data_key][input_key])
        else:
            for basic_db_field, input_key in v0_constants.basic_db_fields_to_input_field_map.items():
                if not detail[common_data_key][input_key]:
                    continue
                setattr(instance, basic_db_field, detail[common_data_key][input_key])

        return instance

    except Exception as e:
        raise Exception(function, ui_utils.get_system_error(e))


def handle_flats(supplier_id, result, flat_map, content_type):
    """
    Args:
        supplier_id:
        flat_map:
        content_type:
        result:

    Returns: instances which are to be updated, instances which will be created and instances which will be deleted.

    """
    function = handle_flats.__name__
    try:
        positive_updated_instances = []
        positive_created_instances = []
        negative_flat_instances = []
        try:
            positive_flat_dict = result[supplier_id]['flats']['positive']
            negative_flat_dict = result[supplier_id]['flats']['negative']
        except KeyError:
            return positive_created_instances, positive_updated_instances, negative_flat_instances

        for flat_type, detail in positive_flat_dict.items():
            try:
                instance = flat_map[flat_type, supplier_id, content_type]
                instance.flat_count = detail['count']
                instance.size_builtup_area = detail['size']
                instance.flat_rent = detail['rent']
                positive_updated_instances.append(instance)
            except KeyError:
                positive_created_instances.append(
                    FlatType(flat_type=flat_type, flat_count=detail['count'], flat_rent=detail['rent'],
                             size_builtup_area=detail['size'], object_id=supplier_id, content_type=content_type))

        for flat_type in list(negative_flat_dict.keys()):
            try:
                instance = flat_map[flat_type, supplier_id, content_type]
                negative_flat_instances.append(instance)
            except KeyError:
                pass

        return positive_created_instances, positive_updated_instances, negative_flat_instances
    except Exception as e:
        raise Exception(function, ui_utils.get_system_error(e))


def handle_events(supplier_id, result, events_map, content_type):
    """

    Args:
        supplier_id:
        result:
        content_type:
        events_map:

    Returns: instances which will be created, instances which will be deleted

    """
    function = handle_events.__name__
    try:
        pos_event_list = []
        neg_events_list = []

        try:
            positive_events = result[supplier_id]['events']['positive']['names']
            negative_events = result[supplier_id]['events']['negative']['names']
        except KeyError:
            return pos_event_list, neg_events_list

        for event in positive_events:
            try:
                ref = events_map[event, supplier_id, content_type]
                # if positive event is there already, don't do anything.
            except KeyError:
                # else create the event
                pos_event_list.append(Events(object_id=supplier_id, content_type=content_type, event_name=event))
        for event in negative_events:
            try:
                # if negative event is there already, it  will be deleted.
                ref = events_map[event, supplier_id, content_type]
                neg_events_list.append(ref)
            except KeyError:
                pass
        return pos_event_list, neg_events_list
    except Exception as e:
        raise Exception(function, ui_utils.get_system_error(e))


def handle_amenities(supplier_id, result, amenity_map, supplier_amenity_instances_map, content_type):
    """

    Args:
        supplier_id:
        result:
        amenity_map:
        supplier_amenity_instances_map:
        content_type:

    Returns: instances which will be created, instances which will be deleted.

    """
    function = handle_amenities.__name__
    try:
        supplier_amenity_instances = []
        negative_amenity_instances = []

        try:
            positive_amenities = result[supplier_id]['amenities']['positive']['names']
            negative_amenities = result[supplier_id]['amenities']['negative']['names']
        except KeyError:
            return supplier_amenity_instances, negative_amenity_instances

        for amenity_name in positive_amenities + negative_amenities:
            amenity_instance = amenity_map.get(amenity_name)
            if amenity_instance:
                key = (supplier_id, content_type, amenity_instance)
                try:
                    ref = supplier_amenity_instances_map[key]
                    if amenity_name in negative_amenities:
                        negative_amenity_instances.append(ref)
                except KeyError:
                    supplier_amenity_instances.append(
                        SupplierAmenitiesMap(content_type=content_type, object_id=supplier_id,
                                             amenity=amenity_instance))
        return supplier_amenity_instances, negative_amenity_instances
    except Exception as e:
        raise Exception(function, ui_utils.get_system_error(e))


def get_inventory_summary_map(supplier_ids=None):
    """
    returns a simple inventory summary map
    :return:
    """
    function = get_inventory_summary_map.__name__
    try:
        instances = InventorySummary.objects.filter_permission(object_id__in=supplier_ids)
        inv_sum_map = {}
        for instance in instances:
            supplier_id = instance.object_id
            if not inv_sum_map.get(supplier_id):
                inv_sum_map[supplier_id] = instance
            else:
                raise Exception('duplicate inventory summary for supplier {0}'.format(supplier_id))
        return inv_sum_map
    except Exception as e:
        raise Exception(function, ui_utils.get_system_error(e))


def get_price_mapping_map(supplier_ids=None):
    """
    returns a simple pmd map

    :param supplier_ids:
    :return:
    """
    function = get_price_mapping_map.__name__
    try:
        instances = PriceMappingDefault.objects.filter(object_id__in=supplier_ids)
        pmd_map = {}
        for instance in instances:
            supplier_id = instance.object_id
            if not pmd_map.get(supplier_id):
                pmd_map[supplier_id] = []
            pmd_map[supplier_id].append(instance)
        return pmd_map
    except Exception as e:
        raise Exception(function, ui_utils.get_system_error(e))


def join_with_underscores(str, delim=' '):
    """
    breaks the string on delim and joins indiviudal parts with underscores
    """
    function = join_with_underscores.__name__
    try:
        return '_'.join(str.split(delim))
    except Exception as e:
        raise Exception(function, ui_utils.get_system_error(e))


def upload_to_amazon(file_name, file_content=None, bucket_name=settings.BUCKET_NAME):
    """
    Args:
        file_name: The file name
        file_content: optional file content
    Returns: success in case file is uploaded, failure otherwise error
    """
    function = upload_to_amazon.__name__
    try:

        if not os.path.exists(file_name) and (not file_content):
            raise Exception(function, 'The file path {0} does not exists also NO content provided.'.format(file_name))

        conn = boto.connect_s3(settings.AWS_ACCESS_KEY_ID, settings.AWS_SECRET_ACCESS_KEY)
        bucket = conn.get_bucket(bucket_name)

        k = Key(bucket)
        k.key = file_name
        if file_content:
            k.set_contents_from_string(file_content.read())
        else:
            k.set_contents_from_filename(file_name)
        k.make_public()
        return 1

    except Exception as e:
        raise Exception(function, ui_utils.get_system_error(e))


def get_generic_id(items):
    """
    pulls first three characters from each of the item in items and adds a random 4 digit at the end to make a general id for any object where custom primary key is required
    :param items:
    :return:
    """
    function = get_generic_id.__name__
    try:
        object_id = ''
        for item in items:
            my_item = ''.join(item.split(' '))
            assert len(my_item) >= 3
            object_id += my_item[:3].upper()
        object_id += str(uuid.uuid4())[-4:].upper()
        return object_id
    except Exception as e:
        return Exception(function, ui_utils.get_system_error(e))


def create_entry_in_role_hierarchy(role):
    """
    create role entry in role_hierarchy table
    :return:
    """
    function = create_entry_in_role_hierarchy.__name__
    try:
        instance = Role.objects.get(codename=role['codename'], organisation=role['organisation'])
        data = {
            'parent': instance.id,
            'child': instance.id
        }
        serializer = RoleHierarchySerializer(data=data)
        if serializer.is_valid():
            serializer.save()
            return 1
        return ui_utils.handle_response(function, data=serializer.errors)
    except Exception as e:
        return Exception(function, ui_utils.get_system_error(e))


def get_campaigns_with_status(category, user, vendor, request):
    """
    return campaigns list by arranging in ongoing, upcoming and completed keys

    :param data:
    :return:
    """
    function = get_campaigns_with_status.__name__
    try:
        current_date = timezone.now()
        campaign_data = {
            'ongoing_campaigns': [],
            'upcoming_campaigns': [],
            'completed_campaigns': [],
            'onhold_campaigns': []
        }
        campaign_query = Q()
        vendor_query = Q()
        supplier_code_query = Q()
        if vendor:
            vendor_query = Q(campaign__principal_vendor=vendor)
        if not user.is_superuser:
            campaign_query = get_query_by_organisation_category(category,
                                                                v0_constants.category_query_status['campaign_query'],
                                                                user)
        if request.query_params.get('supplier_code') == 'mix':
            supplier_code_query = Q(campaign__is_mix=True)
        elif request.query_params.get('supplier_code') and request.query_params.get('supplier_code') != 'mix' and request.query_params.get('supplier_code') != 'all':
            shortlisted_spaces_ids = ShortlistedSpaces.objects.filter(supplier_code=request.query_params.get('supplier_code')).values_list('proposal', flat=True)
            supplier_code_query = Q(campaign_id__in=shortlisted_spaces_ids)

        campaign_data['completed_campaigns'] = CampaignAssignment.objects. \
            filter(campaign_query, vendor_query, supplier_code_query, campaign__tentative_end_date__lt=current_date, campaign__campaign_state='PTC'). \
            annotate(name=F('campaign__name'), principal_vendor=F('campaign__principal_vendor__name'),
                     organisation=F('campaign__account__organisation__name'), vendor_id=F('campaign__principal_vendor')). \
            values('campaign', 'name', 'principal_vendor', 'organisation','vendor_id').distinct()
        campaign_data['upcoming_campaigns'] = CampaignAssignment.objects. \
            filter(campaign_query, vendor_query, supplier_code_query, campaign__tentative_start_date__gt=current_date, campaign__campaign_state='PTC'). \
            annotate(name=F('campaign__name'), principal_vendor=F('campaign__principal_vendor__name'),
                     organisation=F('campaign__account__organisation__name'), vendor_id=F('campaign__principal_vendor')). \
            values('campaign', 'name', 'principal_vendor', 'organisation', 'vendor_id').distinct()
        campaign_data['ongoing_campaigns'] = CampaignAssignment.objects. \
            filter(campaign_query, vendor_query, supplier_code_query, Q(campaign__tentative_start_date__lte=current_date) & Q(
            campaign__tentative_end_date__gte=current_date), campaign__campaign_state='PTC'). \
            annotate(name=F('campaign__name'), principal_vendor=F('campaign__principal_vendor__name'),
                     organisation=F('campaign__account__organisation__name'), vendor_id=F('campaign__principal_vendor')). \
            values('campaign', 'name', 'principal_vendor', 'organisation', 'vendor_id').distinct()
        campaign_data['onhold_campaigns'] = CampaignAssignment.objects. \
            filter(campaign_query, vendor_query, supplier_code_query, campaign__campaign_state='POH'). \
            annotate(name=F('campaign__name'), principal_vendor=F('campaign__principal_vendor__name'),
                     organisation=F('campaign__account__organisation__name'), vendor_id=F('campaign__principal_vendor')). \
            values('campaign', 'name', 'principal_vendor', 'organisation', 'vendor_id').distinct()
        return campaign_data
    except Exception as e:
        return Exception(function, ui_utils.get_system_error(e))


def organise_supplier_inv_images_data(inv_act_assignment_objects, user_map, format):
    """

    :param inv_act_assignment_objects:
    :return:
    """


    function = organise_supplier_inv_images_data.__name__
    try:
        result = {}
        shortlisted_supplier_id_set = set()
        total_shortlisted_spaces_list = []  # this is required to fetch supplier details later
        inv_act_assignment_ids = set()  # this is required to fetch images data later
        # the idea of this loop is to separate different table data in different keys.
        for content in inv_act_assignment_objects:
            if not result.get('shortlisted_suppliers'):
                result['shortlisted_suppliers'] = {}

            # fetch data fro shortlisted_suppliers key
            shortlisted_space_id = content['inventory_activity__shortlisted_inventory_details__shortlisted_spaces']
            proposal_id = content['inventory_activity__shortlisted_inventory_details__shortlisted_spaces__proposal_id']
            proposal_name = content[
                'inventory_activity__shortlisted_inventory_details__shortlisted_spaces__proposal__name']
            supplier_id = content['inventory_activity__shortlisted_inventory_details__shortlisted_spaces__object_id']
            supplier_content_type_id = content[
                'inventory_activity__shortlisted_inventory_details__shortlisted_spaces__content_type']
            assigned_to = content['assigned_to']

            result['shortlisted_suppliers'][shortlisted_space_id] = {
                'shortlisted_space_id': shortlisted_space_id,
                'proposal_id': proposal_id,
                'proposal_name': proposal_name,
                'supplier_id': supplier_id,
                'content_type_id': supplier_content_type_id
            }

            total_shortlisted_spaces_list.append(
                {
                    'content_type_id': supplier_content_type_id,
                    'supplier_id': supplier_id
                }
            )

            if not result.get('shortlisted_inventories'):
                result['shortlisted_inventories'] = {}
            # fetch data for shortlisted_inventories key
            shortlisted_inventory_id = content['inventory_activity__shortlisted_inventory_details']
            inventory_id = content['inventory_activity__shortlisted_inventory_details__inventory_id']
            inventory_content_type_id = content[
                'inventory_activity__shortlisted_inventory_details__inventory_content_type']
            comment = content['inventory_activity__shortlisted_inventory_details__comment']
            inventory_name = content[
                'inventory_activity__shortlisted_inventory_details__ad_inventory_type__adinventory_name']
            inventory_duration = content[
                'inventory_activity__shortlisted_inventory_details__ad_inventory_duration__duration_name']

            result['shortlisted_inventories'][shortlisted_inventory_id] = {
                'shortlisted_spaces_id': shortlisted_space_id,
                'inventory_id': inventory_id,
                'inventory_content_type_id': inventory_content_type_id,
                'comment': comment,
                'inventory_name': inventory_name,
                'inventory_duration': inventory_duration
            }

            if not result.get('inventory_activities'):
                result['inventory_activities'] = {}

            # fetch data for inventory activity key
            inventory_activity_id = content['inventory_activity']
            activity_type = content['inventory_activity__activity_type']

            result['inventory_activities'][inventory_activity_id] = {
                'shortlisted_inventory_id': shortlisted_inventory_id,
                'activity_type': activity_type
            }

            if not result.get('inventory_activity_assignment'):
                result['inventory_activity_assignment'] = {}
            # fetch data for inventory activity assignment
            inventory_activity_assignment_id = content['id']
            activity_date = content['activity_date']
            reassigned_activity_date = content['reassigned_activity_date']
            inv_act_assignment_ids.add(inventory_activity_assignment_id)

            result['inventory_activity_assignment'][inventory_activity_assignment_id] = {
                'activity_date': activity_date.date() if activity_date else None,
                'reassigned_activity_date': reassigned_activity_date.date() if reassigned_activity_date else None,
                'inventory_activity_id': inventory_activity_id,
                'assigned_to': user_map[assigned_to] if assigned_to else v0_constants.default_assigned_to_string
            }

        # after the result is prepared, here we collect images data
        inventory_activity_images = InventoryActivityImage.objects.filter(
            inventory_activity_assignment_id__in=inv_act_assignment_ids)
        images = {}
        for inv_act_image in inventory_activity_images:
            image_id = inv_act_image.id

            if not images.get(image_id):
                images[image_id] = {}

            images[image_id] = {
                'image_path': inv_act_image.image_path,
                'comment': inv_act_image.comment,
                'actual_activity_date': inv_act_image.actual_activity_date.date() if inv_act_image.actual_activity_date else None,
                'inventory_activity_assignment_id': inv_act_image.inventory_activity_assignment_id
            }

        # # set the shortlisted spaces data. it maps various supplier ids to their respective content_types
        response = get_objects_per_content_type(total_shortlisted_spaces_list)
        if not response.data['status']:
            return response
        content_type_supplier_id_map, content_type_set, supplier_id_set = response.data['data']

        # converts the ids store in previous step to actual objects and adds additional information which is
        #  supplier specific  like area, name, subarea etc.
        response = map_objects_ids_to_objects(content_type_supplier_id_map)
        if not response.data['status']:
            return response

        # the returned response is a dict in which key is (content_type, supplier_id) and value is a dict of extra
        # information for that supplier
        supplier_detail = response.data['data']


        response = get_contact_information(content_type_set, supplier_id_set)
        if not response.data['status']:
            return response
        contact_object_per_content_type_per_supplier = response.data['data']

        # add the key 'supplier_detail' which holds all sorts of information for that supplier to final result.
        if result:
            for shortlisted_space_id, detail in result['shortlisted_suppliers'].items():
                key = (detail['content_type_id'], detail['supplier_id'])
                try:
                    raw_supplier_data = supplier_detail[key]
                    detail['supplier_detail'] = raw_supplier_data
                    # detail['supplier_detail'] = {
                    #     'name': raw_supplier_data['name'],
                    #     'address1': raw_supplier_data['address1'],
                    #     'address2': raw_supplier_data['address2'],
                    #     'area': raw_supplier_data['area'],
                    #     'subarea': raw_supplier_data['subarea'],
                    #     'city': raw_supplier_data['city'],
                    #     'state': raw_supplier_data['state'],
                    #     'zipcode': raw_supplier_data['zipcode'],
                    #     'latitude': raw_supplier_data['latitude'],
                    #     'longitude': raw_supplier_data['longitude'],
                    #     'flat_count': raw_supplier_data['flat_count'],
                    #     'user_id': raw_supplier_data['user_id'],
                    #     'created_by_id': raw_supplier_data['created_by_id'],
                    # }
                except KeyError:
                    # ideally every supplier in ss table must also be in the corresponding supplier table. But
                    # because current data is corrupt as i have manually added suppliers, i have to set this to
                    # empty when KeyError occurres. #todo change this later.
                    detail['supplier_detail'] = {}
                    # set images data to final result

                # add 'contact' key to each supplier object
                try:
                    contact_object = contact_object_per_content_type_per_supplier[key]
                    detail['supplier_detail']['contacts'] = contact_object
                except KeyError:
                    detail['supplier_detail']['contacts'] = []
            result['images'] = images
        if result == {}:
            return result
        if format == "new":
            result = restructure_supplier_inv_images_data(result)
        return result
    except Exception as e:
        return Exception(function, ui_utils.get_system_error(e))


def restructure_supplier_inv_images_data(prev_dict):
    all_assignment_data = prev_dict['inventory_activity_assignment']
    all_activity_data = prev_dict['inventory_activities']
    all_inventory_data = prev_dict['shortlisted_inventories']
    all_supplier_data = prev_dict['shortlisted_suppliers']
    all_image_data = prev_dict['images']
    new_dict = {"shortlisted_suppliers":[]}
    for curr_assignment_id in all_assignment_data:
        assignment_data = all_assignment_data[curr_assignment_id]
        curr_activity_id = assignment_data['inventory_activity_id']
        activity_data = all_activity_data[curr_activity_id]
        shortlisted_inventory_id = activity_data['shortlisted_inventory_id']
        inventory_data = all_inventory_data[shortlisted_inventory_id]
        shortlisted_spaces_id = inventory_data['shortlisted_spaces_id']
        supplier_data = all_supplier_data[shortlisted_spaces_id]
        due_date = assignment_data['activity_date']
        if assignment_data['reassigned_activity_date']:
            due_date = assignment_data['reassigned_activity_date']
        if 'activities' not in supplier_data:
            supplier_data['activities'] = {}
        supplier_data['activities'][curr_activity_id] = {
            'inventory_activity_assignment_id': curr_assignment_id,
            'shortlisted_inventory_details_id': shortlisted_inventory_id,
            'activity_id': curr_activity_id,
            'activity_type':activity_data['activity_type'],
            'inventory_id': inventory_data['inventory_id'],
            'comment': inventory_data['comment'],
            'inventory_name': inventory_data['inventory_name'],
            'inventory_duration': inventory_data['inventory_duration'],
            'activity_date': assignment_data['activity_date'],
            'reassigned_activity_date': assignment_data['reassigned_activity_date'],
            'actual_activity_date': None,
            'due_date': due_date,
            'status': 'pending',
            'images': []
        }

    for curr_image_id in all_image_data:
        image_data = all_image_data[curr_image_id]
        assignment_id = image_data['inventory_activity_assignment_id']
        activity_id = all_assignment_data[assignment_id]['inventory_activity_id']
        shortlisted_inventory_id = all_activity_data[activity_id]['shortlisted_inventory_id']
        shortlisted_spaces_id = all_inventory_data[shortlisted_inventory_id]['shortlisted_spaces_id']
        supplier_data = all_supplier_data[shortlisted_spaces_id]
        if image_data['actual_activity_date']:
            supplier_data['activities'][activity_id]['actual_activity_date'] = image_data['actual_activity_date']
            supplier_data['activities'][activity_id]['status'] = 'complete'
            supplier_data['activities'][activity_id]['images'].append({
                'image_path': image_data["image_path"],
                'comment': image_data["comment"],
            })
    for ss in prev_dict['shortlisted_suppliers']:
        prev_dict['shortlisted_suppliers'][ss]['activities'] = list(prev_dict['shortlisted_suppliers'][ss]['activities'].values())
    new_dict['shortlisted_suppliers'] = list(prev_dict['shortlisted_suppliers'].values())

    return new_dict

def save_filters(center, supplier_code, proposal_data, proposal):
    """

    :param center:
    :param supplier_code:
    :param proposal_data:
    :return:
    """
    function_name = save_filters.__name__
    try:
        content_type = ui_utils.fetch_content_type(supplier_code)
        selected_filters_list = []
        for filter_code in proposal_data['center_data'][supplier_code]['filter_codes']:
            data = {
                'center': center,
                'proposal': proposal,
                'supplier_type': content_type,
                'supplier_type_code': supplier_code,
                'filter_name': 'inventory_type_selected',
                'filter_code': filter_code['id'],
                'is_checked': True,
            }
            filter_object = Filters(**data)
            selected_filters_list.append(filter_object)
        now_time = timezone.now()
        Filters.objects.filter(proposal_id=proposal.proposal_id).delete()
        Filters.objects.bulk_create(selected_filters_list)
        Filters.objects.filter(proposal_id=proposal.proposal_id).update(created_at=now_time, updated_at=now_time)

        return ui_utils.handle_response(function_name, data={}, success=True)
    except Exception as e:
        return Exception(function_name, ui_utils.get_system_error(e))


def save_shortlisted_suppliers_data(center, supplier_code, proposal_data, proposal, old_shortlisted_suppliers_map={}):
    """

    :param center:
    :param supplier_code:
    :param proposal_data:
    :return:
    """
    function_name = save_shortlisted_suppliers_data.__name__
    try:
        content_type = ui_utils.fetch_content_type(supplier_code)

        shortlisted_suppliers = []
        for supplier in proposal_data['center_data'][supplier_code]['supplier_data']:
            if supplier['id'] not in old_shortlisted_suppliers_map:
                data = {
                    'content_type': content_type,
                    'object_id': supplier['id'],
                    'center': center,
                    'proposal': proposal,
                    'supplier_code': supplier_code,
                    'status': supplier['status'],
                    # 'booking_status' : 'BK',
                    # 'phase' : 1,
                    'total_negotiated_price': supplier[
                        'total_negotiated_price'] if 'total_negotiated_price' in supplier else None
                }
                shortlisted_suppliers.append(ShortlistedSpaces(**data))

        now_time = timezone.now()

        # ShortlistedSpaces.objects.filter(proposal_id=proposal.proposal_id).delete()
        ShortlistedSpaces.objects.bulk_create(shortlisted_suppliers)
        ShortlistedSpaces.objects.filter(proposal_id=proposal.proposal_id).update(created_at=now_time,
                                                                                  updated_at=now_time)

        return ui_utils.handle_response(function_name, data={}, success=True)
    except Exception as e:
        return Exception(function_name, ui_utils.get_system_error(e))


def save_shortlisted_inventory_pricing_details_data(center, supplier_code, proposal_data, proposal,
                                                    create_inv_act_data=False, old_shortlisted_suppliers_map={}):
    """

    :param center:
    :param supplier_code:
    :param proposal_data:
    :param proposal:
    :param create_inv_act_data:
    :return:
    """

    function_name = save_shortlisted_inventory_pricing_details_data.__name__
    try:
        is_import_sheet = proposal_data['is_import_sheet']
        supplier_ids = [id['id'] for id in proposal_data['center_data'][supplier_code]['supplier_data']]
        supplier_objects = SupplierTypeSociety.objects.filter(supplier_id__in=supplier_ids)
        supplier_objects_mapping = {sup_obj.supplier_id: sup_obj for sup_obj in supplier_objects}
        inventory_summary_objects = InventorySummary.objects.filter(object_id__in=supplier_ids)

        inventory_summary_objects_mapping = {inv_sum_object.object_id: inv_sum_object for inv_sum_object in
                                             inventory_summary_objects}
        shortlisted_suppliers = ShortlistedSpaces.objects.filter(object_id__in=supplier_ids,
                                                                 proposal=proposal.proposal_id)
        shortlisted_suppliers_mapping = {sup_obj.object_id: sup_obj for sup_obj in shortlisted_suppliers}
        shortlisted_inv_objects = []
        index = 0
        for supplier_id in supplier_ids:
            if supplier_id not in inventory_summary_objects_mapping:
                create_inventory_summary_data_for_supplier()
            for filter_code in proposal_data['center_data'][supplier_code]['filter_codes']:
                if is_import_sheet:
                    supplier_inv_count_mapping = {sup_obj['id'] : sup_obj for sup_obj in proposal_data['center_data'][supplier_code]['supplier_data']}
                    inventory_objects = create_inventory_ids(supplier_objects_mapping[supplier_id], filter_code, is_import_sheet,supplier_inv_count_mapping)
                else:
                    inventory_objects = create_inventory_ids(None, filter_code)

                response = make_final_list(filter_code, inventory_objects, shortlisted_suppliers_mapping[supplier_id])
                if response.data['status'] and response.data['data']:
                    shortlisted_inv_objects.extend(response.data['data'])
                    
        ShortlistedInventoryPricingDetails.objects.bulk_create(shortlisted_inv_objects)
        if create_inv_act_data:
            shortlisted_supplier_ids = {space_obj.id for space_obj in shortlisted_suppliers}
            shortlisted_inventory_objects = ShortlistedInventoryPricingDetails.objects.filter(
                shortlisted_spaces__in=shortlisted_supplier_ids,
                shortlisted_spaces__proposal=proposal.proposal_id)

            # response = create_inventory_activity_data(shortlisted_inventory_objects,proposal_data)
            # if not response:
            #     return response


        return ui_utils.handle_response(function_name, data={}, success=True)
    except Exception as e:
        return Exception(function_name, ui_utils.get_system_error(e))


def create_inventory_summary_data_for_supplier():
    """

    :return:
    """
    function_name = create_inventory_summary_data_for_supplier.__name__
    try:
        return 1
    except Exception as e:
        return Exception(function_name, ui_utils.get_system_error(e))


def create_inventory_ids(supplier_object, filter_code, is_import_sheet=False, supplier_inv_mapping={}):
    """

    :param supplier_object:
    :param filter_code:
    :return:
    """
    function_name = create_inventory_ids.__name__
    try:
        tower_count = 1
        inventory_ids = []
        Struct = namedtuple('Struct', 'adinventory_id')
        if is_import_sheet:
            tower_count = supplier_inv_mapping[supplier_object.supplier_id][filter_code['id']]
            if tower_count is None:
                tower_count = 1

        for count in range(int(tower_count)):
            data = Struct(adinventory_id='TESTINVID' + str(filter_code['id']) + '00' + str(count + 1))
            inventory_ids.append(data)
            
        return inventory_ids
    except Exception as e:
        print(e)
        return Exception(function_name, ui_utils.get_system_error(e))


def make_final_list(filter_code, inventory_objects, space_id):
    """

    :return:
    """
    function_name = make_final_list.__name__
    try:
        ad_inventory = v0_constants.inventory_type_duration_dict_list[filter_code['id']]
        ad_inventory_type_id = AdInventoryType.objects.get(adinventory_name=ad_inventory[0],
                                                           adinventory_type=ad_inventory[1])
        duration_type_id = DurationType.objects.get(duration_name=ad_inventory[2])
        content_type = ui_utils.fetch_content_type(filter_code['id'])
        now_time = timezone.now()
        shortlisted_suppliers = []
        for inventory in inventory_objects:
            data = {
                'ad_inventory_type': ad_inventory_type_id,
                'ad_inventory_duration': duration_type_id,
                'inventory_id': inventory.adinventory_id,
                'shortlisted_spaces': space_id,
                'created_at': now_time,
                'updated_at': now_time,
                'inventory_content_type_id': content_type.id
            }

            shortlisted_suppliers.append(ShortlistedInventoryPricingDetails(**data))
        return ui_utils.handle_response(function_name, data=shortlisted_suppliers, success=True)

    except Exception as e:
        return Exception(function_name, ui_utils.get_system_error(e))


def update_proposal_invoice_and_state(proposal_data, proposal):
    """

    :param invoice_number:
    :param proposal:
    :return:
    """
    function_name = update_proposal_invoice_and_state
    try:
        proposal.invoice_number = proposal_data['invoice_number']
        proposal.campaign_state = v0_constants.proposal_finalized
        proposal.tentative_start_date = proposal_data['tentative_start_date']
        proposal.tentative_end_date = proposal_data['tentative_end_date']
        proposal.save()
        return ui_utils.handle_response(function_name, data={}, success=True)
    except Exception as e:
        return Exception(function_name, ui_utils.get_system_error(e))


def create_generic_export_file_data(proposal):
    """

    :param proposal:
    :return:
    """
    function_name = create_generic_export_file_data
    try:
        data = {}
        data['proposal'] = proposal.proposal_id
        data['file_name'] = v0_constants.exported_file_name_default
        data['is_exported'] = False
        serializer = GenericExportFileSerializer(data=data)
        if serializer.is_valid():
            serializer.save()
            return ui_utils.handle_response(function_name, data={}, success=True)
        return ui_utils.handle_response(function_name, data=serializer.errors)
    except Exception as e:
        return Exception(function_name, ui_utils.get_system_error(e))


def create_inventory_activity_data(shortlisted_inventory_objects,proposal_data):
    """

    :param shortlisted_inventory_objects:
    :return:
    """
    function_name = create_inventory_activity_data.__name__
    try:
        inventory_ativity_objects = []
        for shortlisted_inv_instance in shortlisted_inventory_objects:
            for inv_activity_type in INVENTORY_ACTIVITY_TYPES:
                data = {
                    'shortlisted_inventory_details': shortlisted_inv_instance,
                    'activity_type': inv_activity_type[0]
                }
                inventory_ativity_objects.append(InventoryActivity(**data))
        # InventoryActivity.objects.filter(shortlisted_inventory_details__shortlisted_spaces__proposal_id=proposal_data['proposal_id'])
        InventoryActivity.objects.bulk_create(inventory_ativity_objects)
        return ui_utils.handle_response(function_name, data={}, success=True)
    except Exception as e:
        return Exception(function_name, ui_utils.get_system_error(e))


def get_total_activity_data(activity, campaign_id, content_type_id):
    """

    :param activity:
    :param campaign_id:
    :return:
    """
    function_name = get_total_activity_data.__name__
    try:
        result = InventoryActivityAssignment.objects.select_related('inventory_activity',
                                                                    'inventory_activity__shortlisted_inventory_details'). \
            filter(inventory_activity__shortlisted_inventory_details__shortlisted_spaces__proposal=campaign_id,
                   inventory_activity__activity_type=activity,
                   inventory_activity__shortlisted_inventory_details__inventory_content_type_id=content_type_id)
        serializer = InventoryActivityAssignmentSerializer(result, many=True)
        return serializer.data;

    except Exception as e:
        return Exception(function_name, ui_utils.get_system_error(e))


def get_actual_activity_data(activity, campaign_id, content_type_id):
    """

    :param activity:
    :param campaign_id:
    :return:
    """
    function_name = get_total_activity_data.__name__
    try:
        result = InventoryActivityImage.objects.select_related('inventory_activity_assignment',
                                                               'inventory_activity_assignment__inventory_activity',
                                                               'inventory_activity_assignment__inventory_activity__shortlisted_inventory_details'). \
            filter(
            inventory_activity_assignment__inventory_activity__shortlisted_inventory_details__shortlisted_spaces__proposal=campaign_id,
            inventory_activity_assignment__inventory_activity__activity_type=activity,
            inventory_activity_assignment__inventory_activity__shortlisted_inventory_details__inventory_content_type_id=content_type_id). \
            annotate(activity_date=F('inventory_activity_assignment__inventory_activity__activity_date'),
                     assignment_id=F('inventory_activity_assignemnt__id')). \
            values('activity_date', 'created_at', 'image_path', 'assignment_id')
        # serializer = InventoryActivityImageSerializer(result, many=True)
        return result;

    except Exception as e:
        return Exception(function_name, ui_utils.get_system_error(e))


def get_activity_data_by_values(campaign_id, content_type_id):
    """

    :param activity:
    :param campaign_id:
    :return:
    """
    function_name = get_activity_data_by_values.__name__
    try:
        result = list(InventoryActivityImage.objects.select_related('inventory_activity_assignment',
                                                                    'inventory_activity_assignment__inventory_activity',
                                                                    'inventory_activity_assignment__inventory_activity__shortlisted_inventory_details'). \
                      filter(
            inventory_activity_assignment__inventory_activity__shortlisted_inventory_details__shortlisted_spaces__proposal=campaign_id,
            inventory_activity_assignment__inventory_activity__shortlisted_inventory_details__inventory_content_type_id=content_type_id). \
                      annotate(object_id=F(
            'inventory_activity_assignment__inventory_activity__shortlisted_inventory_details__shortlisted_spaces__object_id'),
                               activity=F('inventory_activity_assignment__inventory_activity__activity_type')). \
                      values('object_id', 'latitude', 'longitude', 'inventory_activity_assignment_id', 'activity'))

        return result

    except Exception as e:
        return Exception(function_name, ui_utils.get_system_error(e))


def get_filters_by_campaign(campaign_id):
    """

    :param campaign_id:
    :return:
    """
    function_name = get_filters_by_campaign.__name__
    try:
        filters = Filters.objects.filter(proposal__proposal_id=campaign_id).values('filter_code').distinct()
        # serializer = FiltersSerializer(filters, many=True)
        return filters
    except Exception as e:
        return Exception(function_name, ui_utils.get_system_error(e))


def get_campaign_inventory_activity_data(campaign_id):
    """

    :param campaign_id:
    :return:
    """
    function_name = get_campaign_inventory_activity_data.__name__
    try:
        result = InventoryActivityImage.objects.select_related('inventory_activity_assignment',
                                                               'inventory_activity_assignment__inventory_activity',
                                                               'inventory_activity_assignment__inventory_activity__shortlisted_inventory_details',
                                                               'inventory_activity_assignment__inventory_activity__shortlisted_inventory_details__shortlisted_spaces',
                                                               'inventory_activity_assignment__inventory_activity__shortlisted_inventory_details__shortlisted_spaces__proposal'). \
            filter(
            inventory_activity_assignment__inventory_activity__shortlisted_inventory_details__shortlisted_spaces__proposal=campaign_id). \
            annotate(activity_type=F('inventory_activity_assignment__inventory_activity__activity_type'), inventory=F(
            'inventory_activity_assignment__inventory_activity__shortlisted_inventory_details__ad_inventory_type__adinventory_name')). \
            values('activity_type', 'inventory'). \
            annotate(total=Count('inventory_activity_assignment', distinct=True))
        data = {}
        for object in result:
            if object['inventory'] not in data:
                data[object['inventory']] = {}
            if object['activity_type'] not in data[object['inventory']]:
                data[object['inventory']][object['activity_type']] = {}
            data[object['inventory']][object['activity_type']] = object['total']
        return data

    except Exception as e:
        return Exception(function_name, ui_utils.get_system_error(e))


def calculate_location_difference_between_inventory_and_supplier(data, suppliers):
    """

    :param data:
    :param suppliers:
    :return:
    """
    function_name = calculate_location_difference_between_inventory_and_supplier.__name__
    try:
        supplier_objects_id_map = {supplier['supplier_id']: supplier for supplier in suppliers}

        for item in data:
            lat1 = item['latitude']
            lon1 = item['longitude']
            # need to be changed for other suppliers i.e society_latitude
            if item['object_id'] in supplier_objects_id_map:
                if 'society_latitude' in supplier_objects_id_map[item['object_id']]:
                    lat2 = supplier_objects_id_map[item['object_id']]['society_latitude'] if \
                        supplier_objects_id_map[item['object_id']]['society_latitude'] else None
                else :
                    lat2 = supplier_objects_id_map[item['object_id']]['latitude'] if \
                            supplier_objects_id_map[item['object_id']]['latitude'] else None
                if 'society_longitude' in supplier_objects_id_map[item['object_id']]:
                    lon2 = supplier_objects_id_map[item['object_id']]['society_longitude'] if \
                                supplier_objects_id_map[item['object_id']]['society_longitude'] else None
                else :
                    lon2 = supplier_objects_id_map[item['object_id']]['longitude'] if \
                        supplier_objects_id_map[item['object_id']]['longitude'] else None

                if lat1 and lon1 and lat2 and lon2:
                    distance = gpxpy.geo.haversine_distance(lat1, lon1, lat2, lon2)
                    item['distance'] = distance
        return data
    except Exception as e:
        return Exception(function_name, ui_utils.get_system_error(e))


def create_contact_details(account_id, contacts):
    """
    this function creates multiple contacts of one account in single hit
    :param contacts:
    :return:
    """
    function_name = create_contact_details.__name__
    try:
        for contact in contacts:
            contact['object_id'] = account_id
            serializer = ContactDetailsSerializer(data=contact)
            if serializer.is_valid():
                serializer.save()
            else:
                return ui_utils.handle_response(function_name, data=serializer.errors)
        return ui_utils.handle_response(function_name, data={}, success=True)
    except Exception as e:
        return Exception(function_name, ui_utils.get_system_error(e))


def update_contact_details(account_id, contacts):
    """
    this function will update contacts
    :param account_id:
    :param contacts:
    :return:
    """
    function_name = update_contact_details.__name__
    try:
        for contact in contacts:
            id = None
            if 'id' in contact:
                id = contact['id']
            instance, is_created = ContactDetails.objects.get_or_create(id=id)
            if is_created:
                instance.object_id = account_id
                instance.content_type = ContentType.objects.get_for_model(AccountInfo)
            serializer = ContactDetailsSerializer(data=contact, instance=instance)
            if serializer.is_valid():
                serializer.save()
            else:
                return ui_utils.handle_response(function_name, data=serializer.errors)
        return ui_utils.handle_response(function_name, data={}, success=True)
    except Exception as e:
        return Exception(function_name, ui_utils.get_system_error(e))


def get_query_by_organisation_category(category, campaign_query, user):
    """
    THis fun will return the query based on category provided
    :param category:
    :return:
    """
    function_name = get_query_by_organisation_category.__name__
    try:
        query = Q()
        # organisation_id = user.profile.organisation.organisation_id
        # if category.upper() == v0_constants.category['business']:
        #     query = Q(**{v0_constants.business_category_campaign_query[campaign_query]: organisation_id})
        # if category.upper() == v0_constants.category['business_agency']:
        #     query = Q(**{v0_constants.bus_agency_campaign_query[campaign_query]: user})
        # if category.upper() == v0_constants.category['supplier_agency'] or category.upper() == v0_constants.category['machadalo']:
        query = Q(**{v0_constants.sup_agency_campaign_query[campaign_query]: user})
        return query
    except Exception as e:
        return Exception(function_name, ui_utils.get_system_error(e))


def get_performance_metrics_data_for_inventory(campaign_id, request):
    """
    This fun will return total assignment vs actual activity completed data for inv, ontime and onlocation
    :param campaign_id:
    :param request:
    :return:
    """
    function_name = get_performance_metrics_data_for_inventory.__name__
    try:
        inv_code = request.query_params.get('inv_code', None)
        content_type = ui_utils.fetch_content_type(inv_code)
        content_type_id = content_type.id
        perf_param = request.query_params.get('perf_param', None)
        result = {}
        if perf_param == v0_constants.perf_metrics_param['inv'] or perf_param == v0_constants.perf_metrics_param[
            'on_time']:
            result['total'] = {}
            result['total']['release'] = get_total_assigned_inv_act_data(campaign_id, content_type_id,
                                                                         v0_constants.activity_type['RELEASE'])
            result['total']['audit'] = get_total_assigned_inv_act_data(campaign_id, content_type_id,
                                                                       v0_constants.activity_type['AUDIT'])
            result['total']['closure'] = get_total_assigned_inv_act_data(campaign_id, content_type_id,
                                                                         v0_constants.activity_type['CLOSURE'])
            result['actual'] = {}
            result['actual']['release'] = get_assigned_inv_performance_data(campaign_id, content_type_id,
                                                                            v0_constants.activity_type['RELEASE'])
            result['actual']['audit'] = get_assigned_inv_performance_data(campaign_id, content_type_id,
                                                                          v0_constants.activity_type['AUDIT'])
            result['actual']['closure'] = get_assigned_inv_performance_data(campaign_id, content_type_id,
                                                                            v0_constants.activity_type['CLOSURE'])

        if perf_param == v0_constants.perf_metrics_param['on_location']:
            result['actual'] = {}
            result['actual']['release'] = get_assigned_inv_location_data(campaign_id, content_type_id,
                                                                         v0_constants.activity_type['RELEASE'])
            result['actual']['audit'] = get_assigned_inv_location_data(campaign_id, content_type_id,
                                                                       v0_constants.activity_type['AUDIT'])
            result['actual']['closure'] = get_assigned_inv_location_data(campaign_id, content_type_id,
                                                                         v0_constants.activity_type['CLOSURE'])

        return result
    except Exception as e:
        return Exception(function_name, ui_utils.get_system_error(e))


def get_assigned_inv_performance_data(campaign_id, content_type_id, act_type):
    """

    :param campaign_id:
    :param inv:
    :return:
    """
    function_name = get_assigned_inv_performance_data.__name__
    try:
        image_data = InventoryActivityImage.objects. \
            filter(
            inventory_activity_assignment__inventory_activity__shortlisted_inventory_details__shortlisted_spaces__proposal__proposal_id=campaign_id,
            inventory_activity_assignment__inventory_activity__shortlisted_inventory_details__inventory_content_type_id=content_type_id,
            inventory_activity_assignment__inventory_activity__activity_type=act_type). \
            annotate(activity_date=F('inventory_activity_assignment__activity_date'),
                     inventory_id=F(
                         'inventory_activity_assignment__inventory_activity__shortlisted_inventory_details__id')). \
            values('activity_date', 'inventory_id', 'image_path', 'created_at')
        result = arrange_assignment_data_by_inventory_id(image_data)

        return result
    except Exception as e:
        return Exception(function_name, ui_utils.get_system_error(e))


def arrange_assignment_data_by_inventory_id(data):
    """

    :param data:
    :return:
    """
    function_name = arrange_assignment_data_by_inventory_id.__name__
    try:
        result = {}
        for id in data:
            if id['inventory_id'] not in result:
                result[id['inventory_id']] = []
            result[id['inventory_id']].append(id)
        return result

    except Exception as e:
        return Exception(function_name, ui_utils.get_system_error(e))


def get_assigned_inv_location_data(campaign_id, content_type_id, act_type):
    """

    :param campaign_id:
    :param content_type_id:
    :return:
    """
    function_name = get_assigned_inv_location_data.__name__
    try:
        inv_act_image_objects = InventoryActivityImage.objects. \
            filter(
            inventory_activity_assignment__inventory_activity__shortlisted_inventory_details__shortlisted_spaces__proposal__proposal_id=campaign_id,
            inventory_activity_assignment__inventory_activity__shortlisted_inventory_details__inventory_content_type_id=content_type_id,
            inventory_activity_assignment__inventory_activity__activity_type=act_type). \
            annotate(
            inventory_id=F('inventory_activity_assignment__inventory_activity__shortlisted_inventory_details__id'),
            object_id=F(
                'inventory_activity_assignment__inventory_activity__shortlisted_inventory_details__shortlisted_spaces__object_id')). \
            values('inventory_id', 'object_id', 'image_path', 'latitude', 'longitude')

        supplier_id_list = [object['object_id'] for object in inv_act_image_objects]
        supplier_objects = SupplierTypeSociety.objects.filter(supplier_id__in=supplier_id_list)
        serializer = SupplierTypeSocietySerializer(supplier_objects, many=True)
        suppliers = serializer.data
        inv_act_image_objects_with_distance = calculate_location_difference_between_inventory_and_supplier(
            inv_act_image_objects, suppliers)
        result = arrange_assignment_data_by_inventory_id(inv_act_image_objects_with_distance)
        return result

    except Exception as e:
        return Exception(function_name, ui_utils.get_system_error(e))


def get_total_assigned_inv_act_data(campaign_id, content_type_id, act_type):
    """

    :param campaign_id:
    :param content_type_id:
    :param act_type:
    :return:
    """
    function_name = get_total_assigned_inv_act_data.__name__
    try:
        result = InventoryActivityAssignment.objects. \
            filter(
            inventory_activity__shortlisted_inventory_details__shortlisted_spaces__proposal__proposal_id=campaign_id,
            inventory_activity__shortlisted_inventory_details__inventory_content_type_id=content_type_id,
            inventory_activity__activity_type=act_type). \
            values('inventory_activity__shortlisted_inventory_details__id').distinct().count()
        return result
    except Exception as e:
        return Exception(function_name, ui_utils.get_system_error(e))


def get_campaign_inv_data(campaign_id):
    """
    This function will return inv and inv count data
    :param campaign_id:
    :return:
    """
    function_name = get_campaign_inv_data.__name__
    try:

        result = {}
        inv_act_image_data = InventoryActivityImage.objects.select_related('inventory_activity_assignment',
                                                                                  'inventory_activity_assignment__inventory_activity',
                                                                                  'inventory_activity_assignment__inventory_activity__shortlisted_inventory_details',
                                                                                  'inventory_activity_assignment__inventory_activity__shortlisted_inventory_details__shortlisted_spaces',
                                                                                  'inventory_activity_assignment__inventory_activity__shortlisted_inventory_details__shortlisted_spaces__proposal'). \
            filter(
            inventory_activity_assignment__inventory_activity__shortlisted_inventory_details__shortlisted_spaces__proposal=campaign_id). \
            annotate(activity_type=F('inventory_activity_assignment__inventory_activity__activity_type'), inventory=F(
            'inventory_activity_assignment__inventory_activity__shortlisted_inventory_details__ad_inventory_type__adinventory_name'),
            object_id=F('inventory_activity_assignment__inventory_activity__shortlisted_inventory_details__shortlisted_spaces__object_id')). \
            values('activity_type', 'inventory','object_id'). \
            annotate(total=Count('inventory_activity_assignment', distinct=True))
        inv_act_image_data_map = organise_data_by_activity_and_inventory_type(inv_act_image_data)
        # inv_act_image_data_map = {supplier['object_id']:supplier for supplier in inv_act_image_data}
        total_inv_act_data = ShortlistedInventoryPricingDetails.objects. \
            filter(shortlisted_spaces__proposal=campaign_id). \
            annotate(object_id=F('shortlisted_spaces__object_id'), inventory=F('ad_inventory_type__adinventory_name')). \
            values('object_id', 'inventory'). \
            annotate(total=Count('id'))
        total_inv_act_data_map = {}

        for supplier in total_inv_act_data:
            if supplier['object_id'] not in total_inv_act_data_map:
                total_inv_act_data_map[supplier['object_id']] = {}
            if supplier['inventory'] not in total_inv_act_data_map[supplier['object_id']]:
                total_inv_act_data_map[supplier['object_id']][supplier['inventory']] = {}
            total_inv_act_data_map[supplier['object_id']][supplier['inventory']] = supplier
        inv_act_assigned_data = InventoryActivityAssignment.objects. \
            filter(inventory_activity__shortlisted_inventory_details__shortlisted_spaces__proposal=campaign_id). \
            annotate(activity_type=F('inventory_activity__activity_type'), inventory=F(
            'inventory_activity__shortlisted_inventory_details__ad_inventory_type__adinventory_name'),
                object_id=F('inventory_activity__shortlisted_inventory_details__shortlisted_spaces__object_id')). \
            values('activity_type', 'inventory','object_id'). \
            annotate(total=Count('id', distinct=True))
        inv_act_assigned_data_map = organise_data_by_activity_and_inventory_type(inv_act_assigned_data)

        for key,supplier in total_inv_act_data_map.items():
            if key not in result:
                result[key] = {}
            for inv_key,inv_data in supplier.items():
                if inv_key not in result[key]:
                    result[key][inv_key] = {}
                    result[key][inv_key]['total'] = inv_data
                if key in inv_act_assigned_data_map and inv_key in inv_act_assigned_data_map[key]:
                    result[key][inv_key]['assigned'] = inv_act_assigned_data_map[key][inv_key]
                if key in inv_act_image_data_map and inv_key in inv_act_image_data_map[key]:
                    result[key][inv_key]['completed'] = inv_act_image_data_map[key][inv_key]
        return result
    except Exception as e:
        return Exception(function_name, ui_utils.get_system_error(e))


def get_past_campaigns_data(supplier_id, campaign_id):
    """
    This function returns past campaigns data like last campaign price, last 5 campaigns etc
    :param supplier_id:
    :return:
    """
    function_name = get_past_campaigns_data.__name__
    try:
        last_campaign_data = ShortlistedSpaces.objects.filter(~Q(proposal=campaign_id), object_id=supplier_id,
                                                              total_negotiated_price__isnull=False). \
            values('total_negotiated_price', 'proposal__name').order_by('-proposal__tentative_start_date')
        if last_campaign_data:
            last_campaign_data = last_campaign_data[0]
        past_campaigns_count = ShortlistedSpaces.objects.filter(~Q(proposal=campaign_id),
                                                                object_id=supplier_id).distinct().count()
        campaign_list = ShortlistedSpaces.objects.filter(~Q(proposal=campaign_id), object_id=supplier_id). \
                            annotate(name=F('proposal__name'),
                                     organisation_name=F('proposal__account__organisation__name')). \
                            order_by('-proposal__tentative_start_date').values('name', 'organisation_name').distinct()[
                        :5]
        result = {
            'last_campaign_price': last_campaign_data,
            'past_campaigns': past_campaigns_count,
            'campaigns': campaign_list
        }
        return result
    except Exception as e:
        return Exception(function_name, ui_utils.get_system_error(e))

def organise_data_by_activity_and_inventory_type(data):
    """
    THis function takes data which should contain activity_type, inventory_type and object_id which will return data by orgaised way
    :param data:
    :return:
    """
    function_name = organise_data_by_activity_and_inventory_type.__name__
    try:
        result = {}
        for supplier in data:
            if supplier['object_id'] not in result:
                result[supplier['object_id']] = {}
            if supplier['inventory'] not in result[supplier['object_id']]:
                result[supplier['object_id']][supplier['inventory']] = {}
            if supplier['activity_type'] not in result[supplier['object_id']][supplier['inventory']]:
                result[supplier['object_id']][supplier['inventory']][supplier['activity_type']] = supplier
        return result
    except Exception as e:
        return Exception(function_name, ui_utils.get_system_error(e))


def get_address_from_lat_long(lat, long):
    geolocator = GoogleV3(api_key="AIzaSyCy_uR_SVnzgxCQTw1TS6CYbBTQEbf6jOY")
    location = geolocator.reverse("{0}, {1}".format(lat, long),  exactly_one=True)
    return location.address


def add_string_to_image(image,message):
    filename = str(image)
    with default_storage.open("./" + filename, 'wb+') as destination:
        for chunk in image.chunks():
            destination.write(chunk)
        im = Image.open(destination)
        draw = ImageDraw.Draw(im)
        font = ImageFont.truetype(os.path.join(fonts_path, "Roboto-Bold.ttf"), 40)
        (x, y) = (5, 5)
        color = 'rgb(0, 0, 0)'
        draw.text((x, y), message, fill=color, font=font)
        im.save(str(destination))
        return str(destination)

def import_proposal_cost_data(file,proposal_id):
    function = import_proposal_cost_data.__name__
    # load the workbook
    wb = openpyxl.load_workbook(file)
    # read the sheet
    ws = wb.get_sheet_by_name(v0_constants.metric_sheet_name)

    # before inserting delete all previous data as we don't want to duplicate things.
    response = delete_proposal_cost_data(proposal_id)
    if not response.data['status']:
        return response

    with transaction.atomic():
        try:
            count = 0
            master_data = {}
            # DATA COLLECTION  in order to  collect data in master_data, initialize with proper data structures
            master_data = initialize_master_data(master_data)
            for index, row in enumerate(ws.iter_rows()):

                # ignore empty rows
                if is_empty_row(row):
                    continue
                # send one row for processing
                response = handle_offline_pricing_row(row, master_data)
                if not response.data['status']:
                    return response
                # update master_data with response
                master_data = response.data['data']
                count += 1

            # DATA INSERTION time to save the data
            master_data['proposal_master_cost']['proposal'] = proposal_id
            response = save_master_data(master_data)
            if not response.data['status']:
                return response
            
            return True
        except Exception as e:
            return ui_utils.handle_response(function, exception_object=e)