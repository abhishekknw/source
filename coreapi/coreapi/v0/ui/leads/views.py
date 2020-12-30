from __future__ import print_function
from __future__ import absolute_import

import difflib

from rest_framework.views import APIView
from rest_framework.response import Response
from openpyxl import load_workbook, Workbook
from .models import (get_leads_summary, LeadsPermissions, ExcelDownloadHash, CampaignExcelDownloadHash)
from v0.ui.analytics.views import (get_data_analytics, get_details_by_higher_level,
                                   get_details_by_higher_level_geographical, geographical_parent_details)
from v0.ui.supplier.models import SupplierTypeSociety, SupplierTypeRetailShop, SupplierMaster
from v0.ui.supplier.serializers import SupplierTypeSocietySerializer, SupplierMasterSerializer
from v0.ui.finances.models import ShortlistedInventoryPricingDetails
from v0.ui.proposal.models import ShortlistedSpaces, SupplierAssignment
from v0.ui.inventory.models import (InventoryActivityAssignment, InventoryActivity)
import operator
from v0.ui.utils import handle_response, get_user_organisation_id, create_validation_msg
import boto3
import os
import datetime
from bulk_update.helper import bulk_update
from v0.ui.common.models import BaseUser
from v0.ui.campaign.models import CampaignAssignment, CampaignComments
from v0.constants import (campaign_status, proposal_on_hold, booking_code_to_status,
                          payment_code_to_status, booking_priority_code_to_status, booking_substatus_code_to_status)
from django.http import HttpResponse
from celery import shared_task
from django.conf import settings
from v0.ui.common.models import mongo_client, mongo_test
import pprint
from random import randint
import random, string
from v0.ui.website.utils import prepare_shortlisted_spaces_and_inventories, manipulate_object_key_values, return_price, manipulate_master_to_rs

from v0.ui.proposal.views import convert_date_format

pp = pprint.PrettyPrinter(depth=6)
import hashlib
from bson.objectid import ObjectId
from v0.ui.proposal.models import ProposalInfo, ProposalCenterSuppliers, BookingSubstatus
from v0.ui.account.models import Profile
from v0.ui.b2b.models import Requirement
from v0.ui.dynamic_suppliers.utils import get_dynamic_single_supplier_data
from .utils import convert_xldate_as_datetime, add_society_to_campaign

import logging
logger = logging.getLogger(__name__)

def is_user_permitted(permission_type, user, **kwargs):
    is_permitted = True
    validation_msg_dict = {'msg': None}
    return is_permitted, validation_msg_dict
    leads_form_id = kwargs['leads_form_id'] if 'leads_form_id' in kwargs else None
    campaign_id = kwargs['campaign_id'] if 'campaign_id' in kwargs else None
    permission_list = list(LeadsPermissions.objects.raw({'profile_id': user.profile_id}))
    if len(permission_list) == 0:
        is_permitted = True
        validation_msg_dict['msg'] = 'no_permission_document'
        return is_permitted, validation_msg_dict
    else:
        permission_obj = permission_list[0]
        leads_permissions = permission_obj.leads_permissions
        if leads_form_id:
            all_campaign_ids = list(mongo_client.leads_forms.find({"leads_form_id": int(leads_form_id)}))
            campaign_id = all_campaign_ids[0]['campaign_id']
            leads_level_permissions = permission_obj.leads_permissions['leads_forms']
            if str(leads_form_id) not in leads_level_permissions:
                champaign_level_permissions = permission_obj.leads_permissions['campaigns']
                if campaign_id not in champaign_level_permissions:
                    is_permitted = False
                    validation_msg_dict['msg'] = 'not_permitted'
                    return is_permitted, validation_msg_dict
                else:
                    if permission_type not in champaign_level_permissions[campaign_id]:
                        is_permitted = False
                        validation_msg_dict['msg'] = 'not_permitted'
                        return is_permitted, validation_msg_dict
            else:
                if permission_type not in leads_level_permissions[str(leads_form_id)]:
                    is_permitted = False
                    validation_msg_dict['msg'] = 'not_permitted'
                    return is_permitted, validation_msg_dict
        elif campaign_id:
            if 'campaigns' not in permission_obj.leads_permissions:
                validation_msg_dict['msg'] = 'not_permitted'
                return False, validation_msg_dict
            champaign_level_permissions = permission_obj.leads_permissions['campaigns']
            if campaign_id not in champaign_level_permissions:
                is_permitted = False
                validation_msg_dict['msg'] = 'not_permitted'
                return is_permitted, validation_msg_dict
            else:
                if permission_type not in champaign_level_permissions[campaign_id]:
                    is_permitted = False
                    validation_msg_dict['msg'] = 'not_permitted'
                    return is_permitted, validation_msg_dict
    return is_permitted, validation_msg_dict


def enter_lead_to_mongo(lead_data, supplier_id, campaign_id, lead_form, entry_id):
    entry_id = entry_id + 1
    all_form_items_dict = lead_form['data']
    timestamp = datetime.datetime.utcnow()
    lead_dict = {"data": [], "is_hot": False, "created_at": timestamp, "supplier_id": supplier_id, "campaign_id": campaign_id,
                 "leads_form_id": lead_form['leads_form_id'], "entry_id": entry_id, "status": "active"}
    lead_for_hash = {"data": []}
    for lead_item_data in lead_data:
        if "value" not in lead_item_data:
            lead_item_data["value"] = None
        item_dict = {}
        item_id = lead_item_data["item_id"]
        key_name = all_form_items_dict[str(item_id)]["key_name"]
        key_type = all_form_items_dict[str(item_id)]["key_type"]
        value = lead_item_data["value"]
        item_dict = {
            'key_name': key_name,
            'value': value,
            'item_id': item_id,
            'key_type': key_type
        }
        lead_dict["data"].append(item_dict)
        
        try:
            if not all_form_items_dict[str(item_id)]["isHotLead"]:
                lead_for_hash["data"].append(item_dict)
        except:
            lead_for_hash["data"].append(item_dict)
    
    lead_for_hash["leads_form_id"] = lead_form['leads_form_id']
    lead_sha_256 = create_lead_hash(lead_for_hash)
    lead_dict["lead_sha_256"] = lead_sha_256
    lead_dict["is_hot"], lead_dict["multi_level_is_hot"], lead_dict["hotness_level"] = calculate_is_hot(lead_dict, lead_form['global_hot_lead_criteria'])
    lead_already_exist = True if len(list(mongo_client.leads.find({"lead_sha_256": lead_sha_256}))) > 0 else False
    existing_lead = {}
    hot_lead_count=0
    leads_count = 0
    if not lead_already_exist:
        mongo_client.leads.insert_one(lead_dict)
        mongo_client.leads_forms.update_one({"leads_form_id": lead_form['leads_form_id']}, {"$set": {"last_entry_id": entry_id}})
        leads_count = 1

        if lead_dict["is_hot"]:
            hot_lead_count=1
    else:
        existing_lead_obj = dict(mongo_client.leads.find_one({"lead_sha_256": lead_sha_256}))
        existing_lead = existing_lead_obj["multi_level_is_hot"]
        if lead_dict["is_hot"] and not existing_lead_obj["is_hot"]:
            hot_lead_count=1
        mongo_client.leads.update_one({"lead_sha_256": lead_sha_256},{"$set":lead_dict})

    lead_summary = mongo_client.leads_summary.find_one({"campaign_id": campaign_id,"supplier_id": supplier_id})

    if lead_summary:
        hot_leads = lead_summary.get("hot_leads",{})

        for i in range(0,lead_dict["hotness_level"]):
            key = "is_hot_level_"+str(i+1)
            if not existing_lead.get(key):
                hot_leads[key] = hot_leads.get(key,0)+1
        
        total_leads_count=lead_summary['total_leads_count']+leads_count
        hot_lead_count=lead_summary['total_hot_leads_count']+hot_lead_count
        mongo_client.leads_summary.update_one({"_id": ObjectId(lead_summary['_id'])},
                                                {"$set": {'total_leads_count': total_leads_count,'total_hot_leads_count': hot_lead_count,'hot_leads': hot_leads}})
    else:
        hot_leads = {}
        for i in range(0,lead_dict["hotness_level"]):
            key = "is_hot_level_"+str(i+1)
            hot_leads[key] = 1

        mongo_client.leads_summary.insert_one({
            "campaign_id": campaign_id,
            "supplier_id": supplier_id,
            "lead_date": None,
            "total_leads_count": 1,
            "total_hot_leads_count":hot_lead_count,
            "total_booking_confirmed": 0,
            "total_orders_punched": 0,
            "hot_leads": hot_leads
        })
    return


def convertToNumber(str):
    try:
        return int(str)
    except:
        return str

def get_data_by_supplier_type_code(lead_form, supplier_id):
    proposal_id = lead_form['campaign_id']
    supplier_type_code = ProposalCenterSuppliers.objects.filter(proposal_id=proposal_id)
    if len(supplier_type_code) > 0:
        try:
            code = supplier_type_code[0].supplier_type_code
            if code == 'RS':
                supplier_data = SupplierTypeSociety.objects.get(supplier_id=supplier_id)
                return supplier_data, supplier_data.society_name
            else:
                supplier_data = SupplierMaster.objects.get(supplier_id=supplier_id)
                return supplier_data, supplier_data.supplier_name
        except Exception as e:
            supplier_data = {}
            supplier_data = get_dynamic_single_supplier_data(supplier_id)
            if supplier_data:
                return supplier_data, supplier_data['name']
    return

def get_supplier_all_leads_entries(leads_form_id, supplier_id, page_number=0, **kwargs):
    leads_per_page = 25
    leads_forms = mongo_client.leads_forms.find_one({"leads_form_id": int(leads_form_id)}, {"_id":0, "data":1, "campaign_id": 1})
    if not leads_forms:
        return None
    leads_forms_items = leads_forms["data"]
    if supplier_id == 'All':
        leads_data = mongo_client.leads.find({"$and": [{"leads_form_id": int(leads_form_id)}, {"status": {"$ne": "inactive"}}]},
                                             {"_id": 0})
        leads_data_list = list(leads_data)
        suppliers_list = []
        for lead_data in leads_data_list:
            suppliers_list.append(lead_data['supplier_id'])
        suppliers_list = list(set(suppliers_list))
        suppliers_names = SupplierTypeSociety.objects.filter(supplier_id__in=suppliers_list).values_list(
            'supplier_id','society_name')

        master_suppliers_names = SupplierMaster.objects.filter(supplier_id__in=suppliers_list).values_list(
             'supplier_id','supplier_name')

        all_supplier_names = list(suppliers_names) + list(master_suppliers_names)

        supplier_id_names = dict((x, y) for x, y in all_supplier_names)
    else:
        leads_data = mongo_client.leads.find({"$and": [{"leads_form_id": int(leads_form_id)}, {"supplier_id": supplier_id},
                                                       {"status": {"$ne": "inactive"}}]}, {"_id": 0})
        leads_data_list = list(leads_data)
        supplier_data, supplier_name = get_data_by_supplier_type_code(leads_forms, supplier_id)


    hot_leads = [x['entry_id'] for x in leads_data_list if x['is_hot'] == True]
    headers = []
    all_order_ids = []
    for form_item in leads_forms_items:
        try:
            curr_item = leads_forms_items[form_item]
        except:
            curr_item = form_item
        headers.append({
            "order_id": curr_item["order_id"] if "order_id" in curr_item else None,
            "key_name": curr_item["key_name"],
        })
        if curr_item["order_id"] not in all_order_ids:
            all_order_ids.append(curr_item["order_id"])
    headers.extend((
        {
            "order_id": 0,
            "key_name": "Entry Id"
        },
        {
            "order_id": 0,
            "key_name": "Supplier Name"
        },
        {
            "order_id": 0,
            "key_name": "Lead Date"
        }))
    headers = sorted(headers,key=operator.itemgetter('order_id'))

    values = []

    if 'start_date' in kwargs and kwargs['start_date']:
        leads_data_start = [x for x in leads_data_list if x['created_at'].date() >= kwargs['start_date']]
        leads_data_list = leads_data_start
    if 'end_date' in kwargs and kwargs['end_date']:
        leads_data_start_end = [x for x in leads_data_start if x['created_at'].date() <= kwargs['end_date']]
        leads_data_list = leads_data_start_end


    leads_data_values = []
    missing_suppliers= []

    for entry in leads_data_list:
        curr_entry = entry['data']
        entry_date = entry['created_at']
        entry_id = entry['entry_id']
        if supplier_id == 'All':
            curr_supplier_id = entry['supplier_id']
            if curr_supplier_id not in supplier_id_names:
                missing_suppliers.append(curr_supplier_id)
                continue
            else:
                curr_supplier_name = supplier_id_names[curr_supplier_id]
        else:
            curr_supplier_name = supplier_name
        new_entry = [
            {
                "entry_id": 0,
                "value": entry_id
            },
            {
                "order_id": 0,
                "value": curr_supplier_name
            },
            {
                "order_id": 0,
                "value": entry_date
            }
        ]
        for item in curr_entry:
            if item["item_id"] not in all_order_ids:
                continue
            value = None
            if item["value"]:
                if isinstance(item["value"], (str,bytes)):
                    value = item["value"].strip()
                value = convertToNumber(item["value"])  # if possible
            
            key_name = item.get("key_name","")
            key_type = item.get("key_type","")
           
            new_entry.append({"order_id": item.get("item_id"), "value": value, "key_name":key_name, "key_type":key_type})
        leads_data_values.append(new_entry)

    final_data = {"hot_leads": hot_leads, "headers": headers, "values": leads_data_values, "missing_suppliers": missing_suppliers}

    return final_data

class GetLeadsEntries(APIView):
    @staticmethod
    def get(request, leads_form_id):
        try:
            supplier_id = str(request.query_params.get('supplier_id')) if request.query_params.get('supplier_id'
                                                                                                ) is not None else 'All'

            page_number = int(request.query_params.get('page_number',0))
            supplier_all_lead_entries = get_supplier_all_leads_entries(leads_form_id, supplier_id, page_number)
            return handle_response({}, data=supplier_all_lead_entries, success=True)
        except Exception as e:
            logger.exception(e)
            return handle_response({}, data="No Leads Found", success=False)

class GetLeadsEntriesBySupplier(APIView):
    @staticmethod
    def get(request, leads_form_id, supplier_id):
        page_number = int(request.query_params.get('page_number', 0))
        supplier_all_lead_entries = get_supplier_all_leads_entries(leads_form_id, supplier_id,page_number)
        return handle_response({}, data=supplier_all_lead_entries, success=True)


class GetLeadsEntriesBySupplierCount(APIView):
    @staticmethod
    def get(request, leads_form_id, supplier_id):
        page_number = int(request.query_params.get('page_number', 0))
        supplier_all_lead_entries = get_supplier_all_leads_entries(leads_form_id, supplier_id,page_number)
        return handle_response({}, data=len(supplier_all_lead_entries['values']), success=True)


class GetLeadsEntriesByCampaignId(APIView):
    # it is assumed that a form belongs to a campaign
    @staticmethod
    def get(request, campaign_id, supplier_id='All'):
        page_number = int(request.query_params.get('page_number', 0))
        first_leads_form_id = mongo_client.leads_forms.find_one({"campaign_id":campaign_id})

        if first_leads_form_id is not None:
            leads_form_id = first_leads_form_id['leads_form_id']
            supplier_all_lead_entries = get_supplier_all_leads_entries(leads_form_id, supplier_id, page_number)
            return handle_response({}, data=supplier_all_lead_entries, success=True)
        else:
            return handle_response({}, data="No leads found", success=False)


class CreateLeadsForm(APIView):
    @staticmethod
    def post(request, campaign_id):
        leads_form_name = request.data['leads_form_name']
        leads_form_items = request.data['leads_form_items']
        global_hot_lead_criteria = request.data['global_hot_lead_criteria'] if 'global_hot_lead_criteria' in request.data else None
        hotness_mapping = request.data['hotness_mapping'] if 'hotness_mapping' in request.data else None
        item_id = 0
        max_id_data = mongo_client.leads_forms.find_one(sort=[('leads_form_id', -1)])
        max_id = max_id_data['leads_form_id'] if max_id_data is not None else 0
        new_leads_form_id = max_id+1
        mongo_dict = {
            'leads_form_id': new_leads_form_id,
            'campaign_id': campaign_id,
            'leads_form_name': leads_form_name,
            'data': {},
            'status': 'active'
        }
        for item in leads_form_items:
            item_id = item_id + 1
            key_options = item["key_options"] if 'key_options' in item else None
            if key_options and not isinstance(key_options, list):
                key_options = key_options.split(',')
            item["item_id"] = item_id
            mongo_dict['data'][str(item_id)] = item
        if global_hot_lead_criteria:
            mongo_dict['global_hot_lead_criteria'] = global_hot_lead_criteria
        else:
            mongo_dict['global_hot_lead_criteria'] = create_global_hot_lead_criteria(mongo_dict)
        if hotness_mapping:
            mongo_dict['hotness_mapping'] = hotness_mapping
        mongo_client.leads_forms.insert_one(mongo_dict)
        add_single_leads_permission(request.user.profile_id, new_leads_form_id, ["EDIT", "VIEW", "DELETE", "FILL", "FREEZE", "UNFREEZE"])
        return handle_response({}, data='success', success=True)


class GetLeadsForm(APIView):
    @staticmethod
    def get(request, campaign_id):
        campaign_lead_form = mongo_client.leads_forms.find(
            {"$and": [{"campaign_id": campaign_id}, {"status": {"$ne": "inactive"}}]}, {"_id": 0})
        lead_form_dict = {}
        for lead_form in campaign_lead_form:
            lead_form_dict[lead_form["leads_form_id"]] = {
                "leads_form_name": lead_form['leads_form_name'],
                "leads_form_id": lead_form['leads_form_id'],
                "leads_form_items": lead_form['data'],
                "global_hot_lead_criteria": lead_form['global_hot_lead_criteria'],
                "hotness_mapping": lead_form['hotness_mapping'] if 'hotness_mapping' in lead_form else None
            }
        return handle_response({}, data=lead_form_dict, success=True)

class GetLeadsFormById(APIView):
    @staticmethod
    def get(request, leads_form_id):
        lead_form = list(mongo_client.leads_forms.find(
            {"$and": [{"leads_form_id": int(leads_form_id)}, {"status": {"$ne": "inactive"}}]}))
        lead_form_dict = {}
        if len(lead_form) > 0:
            lead_form = lead_form[0]
            lead_form_dict = {
                    "leads_form_name": lead_form['leads_form_name'],
                    "leads_form_id": lead_form['leads_form_id'],
                    "leads_form_items": lead_form['data']
                }
        return handle_response({}, data=lead_form_dict, success=True)

def get_supplier_data_by_type(name):
    suppliers = SupplierTypeSociety.objects.filter(society_name=name).values('supplier_id',
                                                                                     'society_name').all()
    if len(suppliers) > 0:
        return suppliers
    else:
        suppliers = SupplierMaster.objects.filter(supplier_name=name).values('supplier_id', 'supplier_name').all()
        if len(suppliers) > 0:
            return suppliers
    return []

class LeadsFormBulkEntry(APIView):
    @staticmethod
    def post(request, leads_form_id):
        try:
            is_permitted, validation_msg_dict = is_user_permitted("FILL", request.user, leads_form_id=leads_form_id)
            if not is_permitted:
                return handle_response('', data=validation_msg_dict, success=False)
            source_file = request.data['file']
            wb = load_workbook(source_file)
            ws = wb.get_sheet_by_name(wb.get_sheet_names()[0])
            lead_form = mongo_client.leads_forms.find_one({"leads_form_id": int(leads_form_id)})
            fields = len(lead_form['data'])
            campaign_id = lead_form['campaign_id']
            global_hot_lead_criteria = lead_form['global_hot_lead_criteria']
            entry_id = lead_form['last_entry_id'] if 'last_entry_id' in lead_form else 0

            missing_societies = []
            inv_activity_assignment_missing_societies = []
            inv_activity_assignment_activity_date_missing_societies = []
            inv_activity_missing_societies = []
            not_present_in_shortlisted_societies = []
            more_than_ones_same_shortlisted_society = []
            unresolved_societies = []

            leads_dict = []
            all_sha256 = list(mongo_client.leads.find({"leads_form_id": int(leads_form_id)},{"lead_sha_256": 1, "_id": 0}))
            all_sha256_list = [str(element['lead_sha_256']) for element in all_sha256]
            apartment_index = None
            phone_number_index = None
            club_name_index = None
            phone_number = None
            lead_entry_date = None
            lead_entry_date_index = None
            for index, row in enumerate(ws.iter_rows()):
                if index == 0:
                    for idx, i in enumerate(row):
                        if i.value and 'apartment' in i.value.lower():
                            apartment_index = idx
                        if i.value and i.value.lower() == 'phone_number':
                            phone_number_index = idx
                        if i.value and i.value.lower() == 'lead_entry_date':
                            lead_entry_date_index = idx
                    if not apartment_index:
                        for idx, i in enumerate(row):
                            if i.value and 'club name' in i.value.strip().lower():
                                club_name_index = idx
                                break
                    if not phone_number_index:
                        return handle_response('', data='phone_number header is missing')
                    if not lead_entry_date_index:
                        return handle_response('', data='lead_entry_date header is missing')
                if apartment_index is None and club_name_index is None:
                    return handle_response('', data="neither apartment nor club found in the sheet", success=False)
                entity_index = apartment_index if apartment_index else club_name_index

                if index > 0:
                    if not (row[entity_index].value is None):
                        society_name = row[entity_index].value

                    suppliers = get_supplier_data_by_type(society_name)

                    if not (row[phone_number_index].value is None):
                        phone_number = row[phone_number_index].value

                    if not (row[lead_entry_date_index].value is None):
                        lead_entry_date = row[lead_entry_date_index].value
                        try:
                            lead_entry_date = convert_xldate_as_datetime(lead_entry_date)
                            lead_entry_date = lead_entry_date.isoformat()
                        except Exception as e:
                            return handle_response('', data="Pass lead_entry_date as dd/mm/yy format",
                                                   success=False)

                    if len(suppliers) == 0:
                        if society_name not in missing_societies and society_name is not None:
                            missing_societies.append(society_name)
                        continue
                    else:
                        if len(suppliers) == 1:
                            found_supplier_id = suppliers[0]['supplier_id']
                        else:
                            supplier_ids = []
                            for s in suppliers:
                                supplier_ids.append(s['supplier_id'])
                            shortlisted_spaces = ShortlistedSpaces.objects.filter(proposal_id=campaign_id,
                                                                                  object_id__in=supplier_ids).values(
                                'object_id', 'id').all()
                            # if len(shortlisted_spaces) > 1:
                            #     more_than_ones_same_shortlisted_society.append(society_name)
                            #     continue
                            if len(shortlisted_spaces) == 0:
                                not_present_in_shortlisted_societies.append(society_name)
                                add_society_to_campaign(campaign_id, supplier_id)
                            else:
                                found_supplier_id = shortlisted_spaces[0]['object_id']

                    shortlisted_spaces = ShortlistedSpaces.objects.filter(object_id=found_supplier_id).filter(
                        proposal_id=campaign_id).all()
                    if len(shortlisted_spaces) == 0:
                        not_present_in_shortlisted_societies.append(society_name)
                        add_society_to_campaign(campaign_id, found_supplier_id)

                    if len(shortlisted_spaces) > 0:
                        inventory_list = ShortlistedInventoryPricingDetails.objects.filter(
                            shortlisted_spaces_id=shortlisted_spaces[0].id).all()
                        stall = None

                        for inventory in inventory_list:
                            if inventory.ad_inventory_type_id >= 8 and inventory.ad_inventory_type_id <= 11:
                                stall = inventory
                                break
                        if not stall:
                            continue
                        shortlisted_inventory_details_id = stall.id
                        inventory_list = InventoryActivity.objects.filter(
                            shortlisted_inventory_details_id=shortlisted_inventory_details_id, activity_type='RELEASE').all()

                        if len(inventory_list) == 0:
                            inv_activity_missing_societies.append(society_name)
                            continue
                        inventory_activity_id = inventory_list[0].id
                        inventory_activity_list = InventoryActivityAssignment.objects.filter(
                            inventory_activity_id=inventory_activity_id).all()
                        if len(inventory_activity_list) == 0:
                            inv_activity_assignment_missing_societies.append(society_name)
                            continue

                        created_at = inventory_activity_list[0].activity_date if inventory_activity_list[0].activity_date else None
                        if not created_at:
                            inv_activity_assignment_activity_date_missing_societies.append(society_name)
                    else:
                        created_at = datetime.datetime.now()
                    lead_dict = {"data": [], "is_hot": False, "created_at": created_at, "supplier_id": found_supplier_id,
                                 "campaign_id": campaign_id, "leads_form_id": int(leads_form_id), "entry_id": entry_id,
                                 "phone_number": phone_number, 'lead_entry_date': lead_entry_date}
                    lead_for_hash = {"data": [], "phone_number": phone_number}
                    for item_id in range(0, fields):
                        curr_item_id = item_id + 1
                        curr_form_item_dict = lead_form['data'][str(curr_item_id)]
                        key_name = curr_form_item_dict['key_name']
                        value = row[item_id].value if row[item_id].value else None
                        if isinstance(value, datetime.datetime) or isinstance(value, datetime.time):
                            value = str(value)
                        item_dict = {
                            'key_name': key_name,
                            'value': value,
                            'item_id': curr_item_id
                        }
                        lead_dict["data"].append(item_dict)
                        
                        try:
                            if not curr_form_item_dict["isHotLead"]:
                                lead_for_hash["data"].append(item_dict)
                        except:
                            lead_for_hash["data"].append(item_dict)
                    
                    lead_for_hash["leads_form_id"] = int(leads_form_id)
                    lead_sha_256 = create_lead_hash(lead_for_hash)
                    lead_dict["lead_sha_256"] = lead_sha_256
                    lead_dict["is_hot"], lead_dict["multi_level_is_hot"], lead_dict["hotness_level"] = calculate_is_hot(lead_dict, global_hot_lead_criteria)
                    lead_already_exist = True if lead_sha_256 in all_sha256_list else False
                    existing_lead = {}
                    hot_lead_count=0
                    leads_count = 0
                    if not lead_already_exist:
                        entry_id = entry_id + 1
                        lead_dict["entry_id"] = entry_id
                        mongo_client.leads.insert_one(lead_dict)
                        mongo_client.leads_forms.update_one({"leads_form_id": leads_form_id}, {"$set": {"last_entry_id": entry_id}})
                        leads_count = 1

                        if lead_dict["is_hot"]:
                            hot_lead_count=1
                    else:
                        existing_lead_obj = dict(mongo_client.leads.find_one({"lead_sha_256": lead_sha_256}))
                        existing_lead = existing_lead_obj["multi_level_is_hot"]
                        if lead_dict["is_hot"] and not existing_lead_obj["is_hot"]:
                            hot_lead_count=1
                        mongo_client.leads.update_one({"lead_sha_256": lead_sha_256},{"$set":lead_dict})

                    lead_summary = mongo_client.leads_summary.find_one({"campaign_id": campaign_id,"supplier_id": found_supplier_id})
            
                    if lead_summary:
                        hot_leads = lead_summary.get("hot_leads",{})

                        for i in range(0,lead_dict["hotness_level"]):
                            key = "is_hot_level_"+str(i+1)
                            if not existing_lead.get(key):
                                hot_leads[key] = hot_leads.get(key,0)+1
                        
                        total_leads_count=lead_summary['total_leads_count']+leads_count
                        hot_lead_count=lead_summary['total_hot_leads_count']+hot_lead_count
                        mongo_client.leads_summary.update_one({"_id": ObjectId(lead_summary['_id'])},
                                                                {"$set": {'total_leads_count': total_leads_count,'total_hot_leads_count': hot_lead_count,'hot_leads': hot_leads}})
                    else:
                        hot_leads = {}
                        for i in range(0,lead_dict["hotness_level"]):
                            key = "is_hot_level_"+str(i+1)
                            hot_leads[key] = 1

                        mongo_client.leads_summary.insert_one({
                            "campaign_id": campaign_id,
                            "supplier_id": found_supplier_id,
                            "lead_date": None,
                            "total_leads_count": 1,
                            "total_hot_leads_count":hot_lead_count,
                            "total_booking_confirmed": 0,
                            "total_orders_punched": 0,
                            "hot_leads": hot_leads
                        })

            missing_societies.sort()

            missing_dict = {
                "missing_societies": missing_societies,
                "unresolved_societies": list(set(unresolved_societies)),
                "more_than_ones_same_shortlisted_society": list(set(more_than_ones_same_shortlisted_society)),
                "inv_activity_assignment_missing_societies": list(set(inv_activity_assignment_missing_societies)),
                "inv_activity_assignment_activity_date_missing_societies": list(
                    set(inv_activity_assignment_activity_date_missing_societies)),
                "inv_activit_missing_societies": list(set(inv_activity_missing_societies)),
                "not_present_in_shortlisted_societies": list(set(not_present_in_shortlisted_societies))
            }
            return handle_response({}, data=missing_dict, success=True)
        except Exception as ex:
            print(ex)
            return handle_response({}, data="failed", success=False)


class LeadsFormEntry(APIView):
    @staticmethod
    def post(request, leads_form_id):
        is_permitted, validation_msg_dict = is_user_permitted("FILL", request.user, leads_form_id=leads_form_id)
        if not is_permitted:
            return handle_response('', data=validation_msg_dict, success=False)
        supplier_id = request.data['supplier_id']
        lead_form = mongo_client.leads_forms.find_one({"leads_form_id": int(leads_form_id)})
        entry_id = lead_form['last_entry_id'] if ('last_entry_id' in lead_form and lead_form['last_entry_id']) else 0
        campaign_id = lead_form['campaign_id']
        lead_data = request.data["leads_form_entries"]
        enter_lead_to_mongo(lead_data, supplier_id, campaign_id, lead_form, entry_id)

        return handle_response({}, data='success', success=True)


@shared_task()
def sanitize_leads_data():
    all_leads_forms = list(mongo_client.leads_forms.find({}))
    campaign_list = list(set([leads_form['campaign_id'] for leads_form in all_leads_forms]))
    for campaign_id in campaign_list:
        all_leads_data = mongo_client.leads.find({"campaign_id":campaign_id}).sort("entry_id",-1)
        last_entry_id = all_leads_data[0]['entry_id'] if all_leads_data.count() > 0 else None
        lead_form = mongo_client.leads_forms.find({"campaign_id":campaign_id})
        new_entry_id = last_entry_id + 1 if last_entry_id else 1
        all_leads_items = lead_form[0]['data']
        scholarship_item_id = other_child_class_item_id = counseling_item_id = first_child_class_item_id = first_child_name_item_id = other_child_item_id = None
        for idx,item in all_leads_items.items():
            if "scholarship" in item['key_name'].lower():
                if "test" in item['key_name'].lower():
                    scholarship_item_id = item['item_id']
            if "name of other child" in item['key_name'].lower():
                other_child_item_id = item['item_id']
            if "counseling" in item['key_name'].lower() or "counselling" in item['key_name'].lower() or "counceling" in item['key_name'].lower():
                counseling_item_id = item['item_id']
            if "class of other child" in item['key_name'].lower() or "class of second child" in item['key_name'].lower():
                other_child_class_item_id = item['item_id']
            if "name of first child" in item['key_name'].lower() or "name of child" in item['key_name'].lower():
                first_child_name_item_id = item['item_id']
            if "class of first child" in item['key_name'].lower():
                first_child_class_item_id = item['item_id']
        for lead_data in all_leads_data:
            leads_data_dict = {single_item_data['item_id']:single_item_data for single_item_data in lead_data['data']}
            create_other_child_lead = False
            if leads_data_dict[scholarship_item_id]['value'] == 'NA':
                if not leads_data_dict[counseling_item_id]['value']:
                    leads_data_dict[counseling_item_id]['value'] = "CounselingScheduled"
            if leads_data_dict[other_child_class_item_id]['value'] and len(
                    leads_data_dict[other_child_class_item_id]['value']) > 2:
                create_other_child_lead = True
                other_child_name = leads_data_dict[other_child_item_id]['value']
                other_child_class = leads_data_dict[other_child_class_item_id]['value']
                leads_data_dict[other_child_item_id]['value'] = None
                leads_data_dict[other_child_class_item_id]['value'] = None
            new_data = [leads_data_dict[single_item_id] for single_item_id in leads_data_dict]
            is_hot, multi_level_is_hot, hotness_level = calculate_is_hot({'data':new_data}, lead_form[0]['global_hot_lead_criteria'])
            mongo_client.leads.update_one({"leads_form_id": int(lead_data['leads_form_id']), "entry_id": int(lead_data['entry_id'])},
                                          {"$set": {"data": new_data, "is_hot": is_hot}})
            if create_other_child_lead:
                new_lead_dict = {"is_hot": False, "multi_level_is_hot": multi_level_is_hot, "hotness_level": hotness_level,
                                 "created_at": lead_data['created_at'],
                                 "supplier_id": lead_data['supplier_id'],
                                 "campaign_id": lead_data['campaign_id'],
                                 "leads_form_id": lead_data['leads_form_id'], "entry_id": new_entry_id, "status": "active",
                                 "data": []}
                leads_data_dict[first_child_name_item_id]["value"] = other_child_name
                leads_data_dict[first_child_class_item_id]["value"] = other_child_class
                leads_data_dict[other_child_item_id]["value"] = None
                leads_data_dict[other_child_class_item_id]["value"] = None
                new_lead_dict["data"] = [leads_data_dict[single_item_id] for single_item_id in leads_data_dict]
                new_lead_dict["is_hot"], new_lead_dict["multi_level_is_hot"], new_lead_dict["hotness_level"] = calculate_is_hot(new_lead_dict,
                                                                                                lead_form[0]['global_hot_lead_criteria'])
                lead_sha_256 = create_lead_hash(new_lead_dict)
                new_lead_dict["lead_sha_256"] = lead_sha_256
                lead_already_exist = True if len(
                    list(mongo_client.leads.find({"lead_sha_256": lead_sha_256}))) > 0 else False
                if not lead_already_exist:
                    mongo_client.leads.insert_one(new_lead_dict)
                    new_entry_id = new_entry_id + 1
    return


class SanitizeLeadsData(APIView):
    def put(self, request):
        class_name = self.__class__.__name__
        sanitize_leads_data.delay()
        return handle_response(class_name, data='success', success=True)


def write_keys_to_file(keys_list):
    book = Workbook()
    sheet = book.active
    sheet.append(keys_list)
    now = datetime.datetime.now()
    current_date = now.strftime("%d%m%Y_%H%M")
    cwd = os.path.dirname(os.path.realpath(__file__))
    filename = 'leads_form_' + current_date + '.xlsx'
    filepath = cwd + '/' + filename
    book.save(filepath)

    s3 = boto3.client(
        's3',
        aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
        aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY
    )
    with open(filepath, 'rb') as f:
        try:
            s3.put_object(Body=f, Bucket='leads-forms-templates', Key=filename)
            os.unlink(filepath)
        except Exception as ex:
            print(ex)
    return filename

class GenerateLeadForm(APIView):
    @staticmethod
    def get(request, leads_form_id):
        leads_form_data_mongo = mongo_client.leads_forms.find_one({"leads_form_id": int(leads_form_id)},{"data":1, "_id":0})
        leads_form_data = leads_form_data_mongo["data"]
        keys_list = ['supplier_id', 'lead_entry_date (format: dd/mm/yyyy)']
        for lead in leads_form_data:
            keys_list.append(leads_form_data[lead]["key_name"])
        filename = write_keys_to_file(keys_list)
        return handle_response({}, data={
            'filepath': 'https://s3.ap-south-1.amazonaws.com/leads-forms-templates/' + filename}, success=True)


def get_leads_excel_sheet(leads_form_id, supplier_id,**kwargs):
    start_date = kwargs['start_date'] if 'start_date' in kwargs else None
    end_date = kwargs['end_date'] if 'end_date' in kwargs else None
    all_leads = get_supplier_all_leads_entries(leads_form_id, supplier_id, start_date=start_date, end_date=end_date)
    keys_list = []
    for item in all_leads['headers']:
        keys_list.append(item['key_name'])

    book = Workbook()
    sheet = book.active
    sheet.append(keys_list)
    total_leads_count = 0
    for lead in all_leads["values"]:
        value_list = []
        for item_dict in lead:
            value_list.append(item_dict["value"])
        sheet.append(value_list)
        if value_list != []:
            total_leads_count += 1
    return (book, total_leads_count)


class GenerateLeadDataExcel(APIView):
    @staticmethod
    def get(request, leads_form_id):
        supplier_id = request.GET.get('supplier_id', 'ALL')
        excel_download_hash_dict = {"leads_form_id": int(leads_form_id), "supplier_id": supplier_id}
        now = datetime.datetime.now()
        excel_download_hash_dict["created_at"] = now
        one_time_hash = hashlib.sha256(str(now).encode('utf-8')).hexdigest()
        excel_download_hash_dict["one_time_hash"] = one_time_hash
        ExcelDownloadHash(**excel_download_hash_dict).save()
        return handle_response({}, data={"one_time_hash": one_time_hash}, success=True)


class DownloadLeadDataExcel(APIView):
    permission_classes = ()
    authentication_classes = ()
    @staticmethod
    def get(request, one_time_hash):
        start_date = None
        end_date = None
        st_date = request.query_params.get('start_date', None)
        e_date = request.query_params.get('end_date', None)
        if st_date:
            start_date = st_date[:10]
            start_date = datetime.datetime.strptime(str(start_date), '%Y-%m-%d').date()
        if e_date:
            end_date = e_date[:10]
            end_date = datetime.datetime.strptime(str(end_date), '%Y-%m-%d').date()
        excel_download_hash = list(ExcelDownloadHash.objects.raw({"one_time_hash": one_time_hash}))
        if len(excel_download_hash) > 0:
            supplier_id = excel_download_hash[0].supplier_id
            leads_form_id = excel_download_hash[0].leads_form_id
            (excel_book, total_leads_count) = get_leads_excel_sheet(leads_form_id, 'All', start_date=start_date,
                                                              end_date=end_date)
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=mydata.xlsx'
            excel_book.save(response)
            ExcelDownloadHash.objects.raw({'one_time_hash': one_time_hash}).delete()
            return response
        return handle_response({}, data='success', success=True)


class DeleteLeadForm(APIView):
    # Entire form is deactivated
    @staticmethod
    def put(request, form_id):
        is_permitted, validation_msg_dict = is_user_permitted("DELETE", request.user, leads_form_id=form_id)
        if not is_permitted:
            return handle_response('', data=validation_msg_dict, success=False)
        result = mongo_client.leads_forms.update_one({"leads_form_id": int(form_id)},
                                     {"$set": {"status": "inactive"}})
        return handle_response({}, data='success', success=True)


class DeleteLeadEntry(APIView):
    @staticmethod
    def put(request, form_id, entry_id):
        is_permitted, validation_msg_dict = is_user_permitted("DELETE", request.user, leads_form_id=form_id)
        if not is_permitted:
            return handle_response('', data=validation_msg_dict, success=False)
        result = mongo_client.leads.update_one({"leads_form_id": int(form_id), "entry_id": int(entry_id)},
                                     {"$set": {"status": "inactive"}})
        return handle_response(result, data='success', success=True)

    @staticmethod
    def delete(request):
        request_body = request.data
        supplier_ids = request_body.get("supplier_ids",None)
        campaign_id = request_body.get("campaign_id",None)
        for supplier_id in supplier_ids:
            mongo_client.leads.remove({"campaign_id": campaign_id, "supplier_id": supplier_id})
        return handle_response('', data={"success": True}, success=True)


class DeleteLeadItem(APIView):
    @staticmethod
    def put(request, form_id, item_id):
        is_permitted, validation_msg_dict = is_user_permitted("EDIT", request.user, leads_form_id=form_id)
        if not is_permitted:
            return handle_response('', data=validation_msg_dict, success=False)
        old_form = mongo_client.leads_forms.find_one({"leads_form_id": int(form_id)})
        if not old_form:
            return handle_response({}, data={"status": "No for Found with this id"}, success=True)
        old_form_items = old_form['data']
        new_form_items = {}
        for form_item in old_form_items:
            if str(old_form_items[form_item]['item_id']) != str(item_id):
                new_form_items[str(old_form_items[form_item]['item_id'])] = old_form_items[form_item]
        mongo_client.leads_forms.update_one({"leads_form_id": int(form_id)}, {"$set": {"data": new_form_items}})
        return handle_response({}, data='success', success=True)


class AddLeadFormItems(APIView):
    # this function is used to add
    def put(self, request, form_id):
        is_permitted, validation_msg_dict = is_user_permitted("EDIT", request.user, leads_form_id=form_id)
        if not is_permitted:
            return handle_response('', data=validation_msg_dict, success=False)
        class_name = self.__class__.__name__
        items_dict_list = request.data
        for items_dict in items_dict_list:
            old_form = mongo_client.leads_forms.find_one({"leads_form_id": int(form_id)})
            if not old_form:
                return handle_response(class_name, data={"status": "No form Found with this id"}, success=True)
            old_form_items = old_form['data']
            max_item_id = 0
            max_order_id = 0
            for item in old_form_items:
                if int(item) > max_item_id:
                    max_item_id = int(item)
                if 'order_id' in old_form_items[item] and old_form_items[item]['order_id'] > max_order_id:
                   max_order_id = int(old_form_items[item]['order_id'])
            new_form_item = {
                "key_name": items_dict['key_name'],
                "campaign_id": old_form['campaign_id'],
                "hot_lead_criteria": items_dict["hot_lead_criteria"] if "hot_lead_criteria" in items_dict else None,
                "isHotLead": items_dict["isHotLead"] if "isHotLead" in items_dict else None,
                "key_type": items_dict["key_type"],
                "key_options": items_dict["key_options"] if "key_options" in items_dict else None,
                "leads_form_id": form_id,
                "is_required": items_dict["is_required"] if "is_required" in items_dict else None,
                "item_id": max_item_id + 1,
                "order_id": max_order_id + 1,
            }
            old_form_items[str(max_item_id + 1)] = new_form_item
            global_hot_lead_criteria = create_global_hot_lead_criteria({'data':old_form_items})
            mongo_client.leads_forms.update_one({"leads_form_id": int(form_id)}, {"$set": {"data": old_form_items, 'global_hot_lead_criteria':global_hot_lead_criteria}})
        return handle_response(class_name, data='success', success=True)


class EditLeadsForm(APIView):
    @staticmethod
    def put(request, form_id):
        is_permitted, validation_msg_dict = is_user_permitted("EDIT", request.user, leads_form_id=form_id)
        if not is_permitted:
            return handle_response('', data=validation_msg_dict, success=False)
        leads_form_items = request.data['leads_form_items'] if 'leads_form_items' in request.data.keys() else None
        global_hot_lead_criteria = request.data[
            'global_hot_lead_criteria'] if 'global_hot_lead_criteria' in request.data.keys() else None
        hotness_mapping = request.data[
            'hotness_mapping'] if 'hotness_mapping' in request.data.keys() else None
        name = request.data['leads_form_name'] if 'leads_form_name' in request.data.keys() else None #changed 'name' to 'leads_form_name' so that edit form works properly.
        set_dict = {}
        if name:
            set_dict["leads_form_name"] = name
        if leads_form_items:
            set_dict["data"] = leads_form_items
        if global_hot_lead_criteria:
            set_dict["global_hot_lead_criteria"] = global_hot_lead_criteria
        if hotness_mapping:
            set_dict["hotness_mapping"] = hotness_mapping
        mongo_client.leads_forms.update_one({"leads_form_id": int(form_id)}, {"$set": set_dict})
        return handle_response({}, data='success', success=True)


class InsertExtraLeads(APIView):
    @staticmethod
    def post(request, form_id):
        is_permitted, validation_msg_dict = is_user_permitted("FILL", request.user, leads_form_id=form_id)
        if not is_permitted:
            return handle_response('', data=validation_msg_dict, success=False)
        supplier_id = request.data['supplier_id'] if 'supplier_id' in request.data.keys() else None
        campaign_id = request.data['campaign_id'] if 'campaign_id' in request.data.keys() else None
        if not supplier_id:
            return handle_response({}, data='failed', success=False)
        if not campaign_id:
            return handle_response({}, data='failed', success=False)

        extra_leads = request.data['extra_leads'] if 'extra_leads' in request.data.keys() else None
        extra_hot_leads = request.data['extra_hot_leads'] if 'extra_hot_leads' in request.data.keys() else None
        set_dict = {}
        # if extra_leads:
        set_dict["extra_leads"] = extra_leads
        # if extra_hot_leads:
        set_dict["extra_hot_leads"] = extra_hot_leads

        if set_dict != {}:
            set_dict["supplier_id"] = supplier_id
            set_dict["campaign_id"] = campaign_id
            set_dict["leads_form_id"] = int(form_id)
            set_dict["created_at"] = datetime.datetime.now()

            mongo_client.leads_extras.insert_one(set_dict)
        return handle_response({}, data='success', success=True)


class GenerateDemoData(APIView):
    def put(self, request):
        #leads_form_test_data = mongo_test.leads_forms.find_one({"leads_form_id": 13})
        leads_data_all = mongo_client.leads.find({})
        # copy leads form data
        leads_form_object = mongo_client.leads_forms.find({})
        leads_form_all = list(leads_form_object)
        allchar = string.ascii_letters.lower()
        for curr_form in leads_form_all:
            curr_form['leads_form_name'] = "".join(random.choice(allchar) for x in range(randint(6, 10)))
            curr_form['leads_form_name'] = curr_form['leads_form_name'].title() + " Lead Form"
            total_data_items = int(len(curr_form['data'].keys()))
            if total_data_items > 0:
                next_data_object = {**curr_form['data'][str(total_data_items)]}
                next_data_object["key_name"] = "Is Meeting Fixed"
                next_data_object["item_id"] = total_data_items + 1
                curr_form['data'][str(total_data_items + 1)] = next_data_object
                next_data_object2 = {**curr_form['data'][str(total_data_items)]}
                next_data_object2["key_name"] = "Is Meeting Completed"
                next_data_object2["item_id"] = total_data_items + 2
                curr_form['data'][str(total_data_items + 2)] = next_data_object2
                next_data_object3 = {**curr_form['data'][str(total_data_items)]}
                next_data_object3["key_name"] = "Is Lead Converted"
                next_data_object3["item_id"] = total_data_items + 3
                curr_form['data'][str(total_data_items + 3)] = next_data_object3
                curr_form['global_hot_lead_criteria']['is_hot_level_2'] = {
                    'or': {str(total_data_items + 1): ['Y', 'y', 'Yes', 'yes', 'YES']}
                }
                curr_form['global_hot_lead_criteria']['is_hot_level_3'] = {
                    'or': {str(total_data_items + 2): ['Y', 'y', 'Yes', 'yes', 'YES']}
                }
                curr_form['global_hot_lead_criteria']['is_hot_level_4'] = {
                    'or': {str(total_data_items + 3): ['Y', 'y', 'Yes', 'yes', 'YES']}
                }
            mongo_test.leads_forms.insert_one(curr_form)
        for curr_lead in leads_data_all:
            leads_form_id = curr_lead['leads_form_id']
            curr_leads_form = [x for x in leads_form_all if x['leads_form_id'] == leads_form_id]
            if curr_lead['is_hot']:
                if random.random() < .5:
                    curr_lead['multi_level_is_hot']['is_hot_level_2'] = True
                    curr_lead['hotness_level'] = 2
                    if random.random() < .4:
                        curr_lead['multi_level_is_hot']['is_hot_level_3'] = True
                        curr_lead['hotness_level'] = 3
                    if random.random() < .3:
                        curr_lead['multi_level_is_hot']['is_hot_level_4'] = True
                        curr_lead['hotness_level'] = 4
            curr_data = curr_lead['data']
            new_data = []
            for curr_data_item in curr_data:
                item_id = curr_data_item['item_id']
                if len(curr_leads_form) == 0:
                    continue
                curr_leads_form_items = curr_leads_form[0]['data']
                curr_leads_form_item = curr_leads_form_items[str(item_id)]
                key_name = curr_data_item['key_name']
                key_type = curr_leads_form_item['key_type']
                if 'number' in key_name.lower() or 'phone' in key_name.lower():
                    curr_data_item['value'] = randint(9000000000,9999999999)
                if 'email' in key_type.lower():
                    r1 = "".join(random.choice(allchar) for x in range(randint(4, 6)))
                    r2 = "".join(random.choice(allchar) for x in range(randint(4, 6)))
                    curr_data_item['value'] = r1+'@'+r2 + '.com'
                if 'name' in key_name.lower():
                    if 'apartment' not in key_name.lower():
                        curr_data_item['value'] = "".join(random.choice(allchar) for x in range(randint(6, 10)))
                curr_data_item['key_type'] = key_type
                new_data.append(curr_data_item)
            original_data_len = len(new_data)
            new_data.append({
                'item_id': original_data_len + 1,
                'order_id': original_data_len + 1,
                'key_name': "Is Meeting Fixed",
                'value': 'Y' if curr_lead['hotness_level'] >= 2 else None
            })
            new_data.append({
                'item_id': original_data_len + 2,
                'order_id': original_data_len + 2,
                'key_name': "Is Meeting Completed",
                'value': 'Y' if curr_lead['hotness_level'] >= 3 else None
            })
            new_data.append({
                'item_id': original_data_len + 3,
                'order_id': original_data_len + 3,
                'key_name': "Is Lead Converted",
                'value': 'Y' if curr_lead['hotness_level'] >= 4 else None
            })
            curr_lead['data'] = new_data
            mongo_test.leads.insert_one(curr_lead)
        return handle_response({}, data='success', success=True)


class LeadsSummary(APIView):

    def get(self, request):
        class_name = self.__class__.__name__
        try:
            user_id = request.user.id
            vendor = request.query_params.get('vendor',None)
            if vendor:
                campaign_list = CampaignAssignment.objects.filter(assigned_to_id=user_id,
                                                                campaign__principal_vendor=vendor).values_list(
                    'campaign_id', flat=True).distinct()
            else:
                campaign_list = CampaignAssignment.objects.filter(assigned_to_id=user_id,
                                                                ).values_list('campaign_id', flat=True).distinct()
            
            if request.query_params.get('supplier_code') == "mix":
                campaign_list = ProposalInfo.objects.filter(proposal_id__in=campaign_list,is_mix=True).values_list('proposal_id', flat=True)

            if request.query_params.get('supplier_code') and request.query_params.get('supplier_code') != "mix" and request.query_params.get('supplier_code') != "all":
                campaign_list = ShortlistedSpaces.objects.filter(proposal_id__in=campaign_list,supplier_code=request.query_params.get('supplier_code')).values_list('proposal_id', flat=True).distinct()
            
            campaign_list = [campaign_id for campaign_id in campaign_list]

            all_shortlisted_supplier = ShortlistedSpaces.objects.filter(proposal_id__in=campaign_list).\
                values('proposal_id', 'object_id', 'phase_no_id', 'is_completed', 'proposal__name', 'proposal__tentative_start_date',
                    'proposal__tentative_end_date', 'proposal__campaign_state', 'supplier_code')

            all_campaign_dict = {}
            all_shortlisted_supplier_id = [supplier['object_id'] for supplier in all_shortlisted_supplier if supplier['supplier_code'] == 'RS']
            all_supplier_society = SupplierTypeSociety.objects.filter(supplier_id__in=all_shortlisted_supplier_id).values('supplier_id', 'flat_count')

            all_supplier_id = [supplier['object_id'] for supplier in all_shortlisted_supplier if supplier['supplier_code'] != 'RS']
            all_supplier_master = SupplierMaster.objects.filter(supplier_id__in=all_supplier_id).values('supplier_id', 'unit_primary_count')

            all_supplier_society_dict = {}
            current_date = datetime.datetime.now().date()
            for supplier in all_supplier_society:
                all_supplier_society_dict[supplier['supplier_id']] = {'flat_count': supplier['flat_count']}

            for supplier in all_supplier_master:
                all_supplier_society_dict[supplier['supplier_id']] = {'flat_count': supplier['unit_primary_count']}

            for shortlisted_supplier in all_shortlisted_supplier:
                if shortlisted_supplier['proposal_id'] not in all_campaign_dict:
                    all_campaign_dict[shortlisted_supplier['proposal_id']] = {
                    'all_supplier_ids': [], 'all_phase_ids': [], 'total_flat_counts': 0, 'total_leads':0, 'hot_leads':0}
                if shortlisted_supplier['object_id'] not in all_campaign_dict[shortlisted_supplier['proposal_id']]['all_supplier_ids']:
                    all_campaign_dict[shortlisted_supplier['proposal_id']]['all_supplier_ids'].append(shortlisted_supplier['object_id'])
                    if shortlisted_supplier['object_id'] in all_supplier_society_dict and all_supplier_society_dict[shortlisted_supplier['object_id']]['flat_count']:
                        all_campaign_dict[shortlisted_supplier['proposal_id']]['total_flat_counts'] += all_supplier_society_dict[shortlisted_supplier['object_id']]['flat_count']
                if shortlisted_supplier['phase_no_id'] and shortlisted_supplier['phase_no_id'] not in all_campaign_dict[shortlisted_supplier['proposal_id']]['all_phase_ids']:
                    if shortlisted_supplier['proposal__tentative_end_date'].date() < current_date:
                        all_campaign_dict[shortlisted_supplier['proposal_id']]['all_phase_ids'].append(
                            shortlisted_supplier['phase_no_id'])
                all_campaign_dict[shortlisted_supplier['proposal_id']]['name'] = shortlisted_supplier['proposal__name']
                all_campaign_dict[shortlisted_supplier['proposal_id']]['start_date'] = shortlisted_supplier['proposal__tentative_start_date']
                all_campaign_dict[shortlisted_supplier['proposal_id']]['end_date'] = shortlisted_supplier['proposal__tentative_end_date']
                all_campaign_dict[shortlisted_supplier['proposal_id']]['campaign_status'] = shortlisted_supplier['proposal__campaign_state']
            all_campaign_summary = get_leads_summary(campaign_list)
            all_leads_summary = []
            for campaign_summary in all_campaign_summary:
                if campaign_summary['campaign_id'] not in all_campaign_dict:
                    continue
                all_campaign_dict[campaign_summary['campaign_id']]['hot_leads'] += campaign_summary['hot_leads_count']
                all_campaign_dict[campaign_summary['campaign_id']]['total_leads'] += campaign_summary['total_leads_count']
            for campaign_id in all_campaign_dict:
                this_campaign_status = None
                if not all_campaign_dict[campaign_id]['campaign_status'] == proposal_on_hold:
                    if all_campaign_dict[campaign_id]['start_date'].date() > current_date:
                        this_campaign_status = campaign_status['upcoming_campaigns']
                    elif all_campaign_dict[campaign_id]['end_date'].date() >= current_date:
                        this_campaign_status = campaign_status['ongoing_campaigns']
                    elif all_campaign_dict[campaign_id]['end_date'].date() < current_date:
                        this_campaign_status = campaign_status['completed_campaigns']
                else:
                    this_campaign_status = "on_hold"
                all_leads_summary.append({
                    "campaign_id": campaign_id,
                    "name": all_campaign_dict[campaign_id]['name'],
                    "start_date": all_campaign_dict[campaign_id]['start_date'],
                    "end_date": all_campaign_dict[campaign_id]['end_date'],
                    "phase_complete": len(all_campaign_dict[campaign_id]['all_phase_ids']),
                    "supplier_count": len(all_campaign_dict[campaign_id]['all_supplier_ids']),
                    "flat_count": all_campaign_dict[campaign_id]['total_flat_counts'],
                    "total_leads": all_campaign_dict[campaign_id]['total_leads'],
                    "hot_leads": all_campaign_dict[campaign_id]['hot_leads'],
                    "campaign_status": this_campaign_status
                })
            return handle_response(class_name, data=all_leads_summary, success=True)
        except Exception as e:
            logger.exception(e)
            return handle_response(class_name, exception_object=e, request=request)

# class SmsContact(APIView):
#
#     def post(self, request, form_id):
#         class_name = self.__class__.__name__
#         contact_details = request.data
#         contact_mobile = contact_details['mobile']
#         data = (LeadsFormContacts(**{
#             'contact_name': contact_details['name'],
#             'contact_mobile': contact_details['mobile'],
#             'form_id': form_id
#         }))
#         data.save()
#         return handle_response(class_name, data='success', success=True)
#
#     def get(self, request, form_id):
#         class_name = self.__class__.__name__
#         contacts_data_object = LeadsFormContacts.objects.filter(form_id=form_id).values('contact_name',
#                                                                                         'contact_mobile')
#         contacts_data = []
#         for data in contacts_data_object:
#             contacts_data.append(data)
#         return handle_response(class_name, data=contacts_data, success=True)


def create_lead_hash(lead_dict):
    lead_hash_string = ''
    lead_hash_string += str(lead_dict['leads_form_id'])

    for item in lead_dict['data']:
        if item['value']:
            if isinstance(item["value"], (str,bytes)):
                lead_hash_string += str(item['value'].strip())
            else:
                lead_hash_string += str(item['value'])
    return hashlib.sha256(lead_hash_string.encode('utf-8')).hexdigest()


class UpdateLeadsDataSHA256(APIView):
    def put(self, request):
        refresh_all = request.data['refresh_all'] if 'refresh_all' in request.data else False
        find_param = {"lead_sha_256": {"$exists": False}}
        if refresh_all:
            find_param = {}
        leads_data_all = mongo_client.leads.find(find_param, no_cursor_timeout=True)
        bulk = mongo_client.leads.initialize_unordered_bulk_op()
        counter = 0
        for curr_lead in leads_data_all:
            entry_id = curr_lead['entry_id']
            lead_sha_256 = create_lead_hash(curr_lead)
            bulk.find({"entry_id": int(entry_id)}).update({"$set": {"lead_sha_256": lead_sha_256}})
            counter += 1
            if counter % 500 == 0:
                bulk.execute()
                bulk = mongo_client.leads.initialize_unordered_bulk_op()
        if counter > 0:
            bulk.execute()
        return handle_response({}, data='success', success=True)


def create_global_hot_lead_criteria(curr_lead_form):
    global_hot_lead_criteria = {
        'is_hot_level_1': {
            'or': {},
        }
    }
    items_dict = curr_lead_form['data']
    for item in items_dict:
        key_name = items_dict[item]['key_name'].lower()
        if 'hot_lead_criteria' in items_dict[item] and items_dict[item]['hot_lead_criteria']:
            global_hot_lead_criteria['is_hot_level_1']['or'][item] = [items_dict[item]['hot_lead_criteria']]
            if items_dict[item]['hot_lead_criteria'] == 'Y':
                global_hot_lead_criteria['is_hot_level_1']['or'][item] += ['y','Yes', 'yes', 'YES']
        if "counseling" in key_name or "counselling" in key_name or "counceling" in key_name:
            global_hot_lead_criteria['is_hot_level_1']['or'][item] = ['AnyValue']
    return global_hot_lead_criteria


class UpdateGlobalHotLeadCriteria(APIView):
    def put(self, request):
        refresh_all = request.data['refresh_all'] if 'refresh_all' in request.data else False
        find_param = {"global_hot_lead_criteria": {"$exists": False}}
        if refresh_all:
            find_param = {}
        leads_form_all = mongo_client.leads_forms.find(find_param, no_cursor_timeout=True)
        bulk = mongo_client.leads_forms.initialize_unordered_bulk_op()
        counter = 0
        for curr_lead_form in leads_form_all:
            leads_form_id = curr_lead_form['leads_form_id']
            global_hot_lead_criteria = create_global_hot_lead_criteria(curr_lead_form)
            bulk.find({"leads_form_id": int(leads_form_id)}).update({"$set": {"global_hot_lead_criteria": global_hot_lead_criteria}})
            counter += 1
            if counter % 500 == 0:
                bulk.execute()
                bulk = mongo_client.leads_forms.initialize_unordered_bulk_op()
        if counter > 0:
            bulk.execute()
        return handle_response({}, data='success', success=True)


def calculate_hotness_level(multi_level_is_hot):
    hotness_level = 0
    for is_hot_level in multi_level_is_hot:
        if multi_level_is_hot[is_hot_level]:
            if int(is_hot_level[-1]) > hotness_level:
                hotness_level = int(is_hot_level[-1])
    return hotness_level


def calculate_is_hot(curr_lead, global_hot_lead_criteria):
    # checking 'or' global_hot_lead_criteria
    any_is_hot = False
    multi_level_is_hot = {is_hot_level: False for is_hot_level in global_hot_lead_criteria}
    curr_lead_data_dict = {str(item['item_id']): item for item in curr_lead['data']}
    for is_hot_level in global_hot_lead_criteria:
        for item_id in global_hot_lead_criteria[is_hot_level]['or']:
            if item_id in curr_lead_data_dict and curr_lead_data_dict[item_id]['value'] is not None:
                if str(curr_lead_data_dict[item_id]['value']) in global_hot_lead_criteria[is_hot_level]['or'][item_id]:
                    multi_level_is_hot[is_hot_level] = True
                    any_is_hot = True
                if "AnyValue" in global_hot_lead_criteria[is_hot_level]['or'][item_id] and str(
                        curr_lead_data_dict[item_id]['value']).lower() != 'na':
                    multi_level_is_hot[is_hot_level] = True
                    any_is_hot = True
    hotness_level = calculate_hotness_level(multi_level_is_hot)
    return any_is_hot, multi_level_is_hot, hotness_level


@shared_task()
def update_leads_data_is_hot():
    leads_form_all = mongo_client.leads_forms.find({}, no_cursor_timeout=True)
    for leads_form_curr in leads_form_all:
        leads_form_id = leads_form_curr['leads_form_id']
        global_hot_lead_criteria = leads_form_curr['global_hot_lead_criteria']
        leads_data_all = mongo_client.leads.find({"leads_form_id": leads_form_id}, no_cursor_timeout=True)
        bulk = mongo_client.leads.initialize_unordered_bulk_op()
        counter = 0
        for curr_lead in leads_data_all:
            entry_id = curr_lead['entry_id']
            is_hot, multi_level_is_hot, hotness_level = calculate_is_hot(curr_lead, global_hot_lead_criteria)
            bulk.find({"entry_id": int(entry_id)}).update({"$set": {"is_hot": is_hot,
                                                                    "multi_level_is_hot": multi_level_is_hot,
                                                                    "hotness_level": hotness_level}})
            counter += 1
            if counter % 500 == 0:
                bulk.execute()
                bulk = mongo_client.leads.initialize_unordered_bulk_op()
        if counter > 0:
            bulk.execute()


class UpdateLeadsDataIsHot(APIView):
    def put(self, request):
        update_leads_data_is_hot()
        return handle_response({}, data='success', success=True)


class GetAllLeadFormsByCampaigns(APIView):
    # used for getting a list of all checklists of a campaign
    def get(self, request):
        class_name = self.__class__.__name__
        campaign_list = CampaignAssignment.objects.filter(assigned_to_id=request.user.id).values_list('campaign_id', flat=True) \
            .distinct()
        campaign_list = [campaign_id for campaign_id in campaign_list]
        all_campaign_name = ProposalInfo.objects.filter(proposal_id__in=campaign_list).values('proposal_id',
                                                                                                     'name')
        all_campaign_name_dict = {campaign['proposal_id']: campaign['name'] for campaign in all_campaign_name}
        all_campaign_leads_forms = list(mongo_client.leads_forms.find({"$and": [{"campaign_id": {"$in": campaign_list}},
                                                                         {"status": {"$ne": "inactive"}}]}))
        all_leads_forms_names = {leads_forms["leads_form_id"]: leads_forms["leads_form_name"] for leads_forms in all_campaign_leads_forms}
        all_campaign_leads_forms_dict = {}
        for single_object in all_campaign_leads_forms:
            if single_object['campaign_id'] not in all_campaign_leads_forms_dict:
                all_campaign_leads_forms_dict[single_object['campaign_id']] = []
            all_campaign_leads_forms_dict[single_object['campaign_id']].append(single_object['leads_form_id'])
        for campaign_id in campaign_list:
            if campaign_id not in all_campaign_leads_forms_dict:
                all_campaign_leads_forms_dict[campaign_id] = []
        all_campaign_leads_forms_list = []
        for campaign_id in all_campaign_leads_forms_dict:
            campaign_checklist_dict = {
                "campaign_id": campaign_id,
                "campaign_name": all_campaign_name_dict[campaign_id],
                "leads_forms": []
            }

            for checklist_id in all_campaign_leads_forms_dict[campaign_id]:
                campaign_checklist_dict["leads_forms"].append({
                    "lead_form_name": all_leads_forms_names[checklist_id],
                    "lead_form_id": checklist_id
                })
            all_campaign_leads_forms_list.append(campaign_checklist_dict)
        return handle_response(class_name, data=all_campaign_leads_forms_list, success=True)


def add_single_leads_permission(profile_id, leads_form_id, new_permissions):
    existing_leads_form_permissions_dict = list(LeadsPermissions.objects.raw({"profile_id": profile_id}))
    if len(existing_leads_form_permissions_dict) == 0:
        return handle_response({}, data='something_is_wrong', success=False)
    existing_leads_form_permissions_dict = existing_leads_form_permissions_dict[0]
    permissions_id = str(existing_leads_form_permissions_dict._id)
    old_leads_form_permissions = existing_leads_form_permissions_dict.leads_permissions
    if "leads_forms" not in old_leads_form_permissions:
        old_leads_form_permissions["leads_forms"] = {}
    old_leads_form_permissions["leads_forms"][str(leads_form_id)] = new_permissions
    dict_of_req_attributes = {
        "profile_id": profile_id,
        "leads_permissions": old_leads_form_permissions,
        "updated_at": datetime.datetime.now()
    }
    LeadsPermissions.objects.raw({'_id': ObjectId(permissions_id)}).update({"$set": dict_of_req_attributes})


class LeadsPermissionsAPI(APIView):
    @staticmethod
    def post(request):
        for single_obj in request.data:
            leads_permissions = single_obj['leads_permissions']
            profile_id = single_obj['profile_id']
            organisation_id = get_user_organisation_id(request.user)
            dict_of_req_attributes = {"leads_permissions": leads_permissions,
                                      "organisation_id": organisation_id, "profile_id": profile_id}
            (is_valid, validation_msg_dict) = create_validation_msg(dict_of_req_attributes)
            if not is_valid:
                return handle_response('', data=validation_msg_dict, success=False)
            leads_permissions_dict = dict_of_req_attributes
            leads_permissions_dict["created_by"] = request.user.id
            leads_permissions_dict["created_at"] = datetime.datetime.now()
            LeadsPermissions(**leads_permissions_dict).save()
        return handle_response('', data={"success": True}, success=True)

    @staticmethod
    def get(request):
        organisation_id = get_user_organisation_id(request.user)
        if organisation_id is not None:
            leads_permissions_all = LeadsPermissions.objects.raw({'organisation_id': organisation_id})
            all_user_id_list = []
            all_profile_id_list = []
            for permission in leads_permissions_all:
                if permission.profile_id not in all_profile_id_list:
                    all_profile_id_list.append(permission.profile_id)
                if permission.created_by not in all_user_id_list:
                    all_user_id_list.append(permission.created_by)
            all_user_objects = BaseUser.objects.filter(id__in=all_user_id_list).all()
            all_user_dict = {user.id: {"first_name": user.first_name,
                                       "last_name": user.last_name,
                                       "username": user.username,
                                       "email": user.email,
                                       "id": user.id
                                       } for user in all_user_objects}
            all_profile_objects = Profile.objects.filter(id__in=all_profile_id_list).all()
            all_profile_dict = {profile.id: {
                    "id": profile.id,
                    "name": profile.name,
                    "is_standard": profile.is_standard,
                    "organisation_id": profile.organisation_id,
                } for profile in all_profile_objects}
            data = []
            for permission in leads_permissions_all:
                permission_data = {
                    "id": str(permission._id),
                    "profile_id": all_profile_dict[int(permission.profile_id)] if int(permission.profile_id) in all_profile_dict else None,
                    "organisation_id": permission.organisation_id,
                    "leads_permissions": permission.leads_permissions,
                    "created_by": all_user_dict[int(permission.created_by)] if int(
                        permission.created_by) in all_user_dict else None
                }
                data.append(permission_data)
            return handle_response('', data=data, success=True)
        else:
            return handle_response('', data='organisation not found', success=True)

    @staticmethod
    def delete(request):
        permission_id = request.query_params.get("permission_id", None)
        if not permission_id:
            return handle_response('', data="Permission Id Not Provided", success=False)
        exist_permission_query = LeadsPermissions.objects.raw({'_id': ObjectId(permission_id)})[0]
        exist_permission_query.delete()
        return handle_response('', data="success", success=True)

    @staticmethod
    def put(request):
        permissions = request.data
        for permission in permissions:
            curr_permission = permission.copy()
            curr_permission.pop('id')
            curr_permission['updated_at'] = datetime.datetime.now()
            LeadsPermissions.objects.raw({'_id': ObjectId(permission['id'])}).update({"$set": curr_permission})
        return handle_response('', data={"success": True}, success=True)


class LeadsPermissionsByProfileIdAPI(APIView):
    @staticmethod
    def get(request, profile_id):
        organisation_id = get_user_organisation_id(request.user)
        leads_permissions = list(LeadsPermissions.objects.raw(
            {"profile_id": int(profile_id), "organisation_id": organisation_id}))
        all_user_id_list = []
        all_profile_id_list = []
        for permission in leads_permissions:
            if permission.profile_id not in all_profile_id_list:
                all_profile_id_list.append(permission.profile_id)
            if permission.created_by not in all_user_id_list:
                all_user_id_list.append(permission.created_by)

        all_user_objects = BaseUser.objects.filter(id__in=all_user_id_list).all()
        all_user_dict = {user.id: {"first_name": user.first_name,
                                   "last_name": user.last_name,
                                   "username": user.username,
                                   "email": user.email,
                                   "id": user.id
                                   } for user in all_user_objects}
        all_profile_objects = Profile.objects.filter(id__in=all_profile_id_list).all()
        all_profile_dict = {profile.id: {"name": profile.name,
                                         "id": profile.id,
                                         "is_standard": profile.is_standard,
                                         "organisation_id": profile.organisation_id,
                                         } for profile in all_profile_objects}
        if len(leads_permissions) == 0:
            return handle_response('', data="no_permission_exists", success=True)

        leads_permissions = leads_permissions[0]
        permission_data = {
            "id": str(leads_permissions._id),
            "profile_id": all_profile_dict[int(leads_permissions.profile_id)] if int(leads_permissions.profile_id) in all_profile_dict else None,
            "organisation_id": leads_permissions.organisation_id,
            "leads_permissions": leads_permissions.leads_permissions,
            "created_by": all_user_dict[int(leads_permissions.created_by)] if int(leads_permissions.created_by) in all_user_dict else None
        }
        return handle_response('', data=permission_data, success=True)


class LeadsPermissionsSelfAPI(APIView):
    @staticmethod
    def get(request):
        organisation_id = get_user_organisation_id(request.user)
        leads_permissions = list(LeadsPermissions.objects.raw(
            {"profile_id": int(request.user.profile_id), "organisation_id": organisation_id}))
        all_user_id_list = []
        all_profile_id_list = []
        for permission in leads_permissions:
            if permission.created_by not in all_user_id_list:
                all_user_id_list.append(permission.created_by)
            if permission.profile_id not in all_profile_id_list:
                all_profile_id_list.append(permission.profile_id)

        all_user_objects = BaseUser.objects.filter(id__in=all_user_id_list).all()
        all_user_dict = {user.id: {"first_name": user.first_name,
                                   "last_name": user.last_name,
                                   "username": user.username,
                                   "email": user.email,
                                   "id": user.id
                                   } for user in all_user_objects}
        all_profile_objects = Profile.objects.filter(id__in=all_profile_id_list).all()
        all_profile_dict = {profile.id: {"name": profile.name,
                                         "is_standard": profile.is_standard,
                                         "organisation_id": profile.organisation_id,
                                         } for profile in all_profile_objects}
        if len(leads_permissions) == 0:
            return handle_response('', data="no_permission_exists", success=True)

        leads_permissions = leads_permissions[0]
        permission_data = {
            "id": str(leads_permissions._id),
            "profile_id": all_profile_dict[int(leads_permissions.profile_id)] if int(leads_permissions.profile_id) in all_profile_dict else None,
            "organisation_id": leads_permissions.organisation_id,
            "leads_permissions": leads_permissions.leads_permissions,
            "created_by": all_user_dict[int(leads_permissions.created_by)] if int(leads_permissions.created_by) in all_user_dict else None
        }
        return handle_response('', data=permission_data, success=True)



class DeleteExtraLeadEntry(APIView):
    @staticmethod
    def delete(request, id):
        mongo_client.leads_extras.remove({"_id":ObjectId(id)})
        return handle_response('', data={"success": True}, success=True)


class GetListsCounts(APIView):
    @staticmethod
    def put(request):
        data = request.data
        highest_level = data['highest_level']
        lowest_level = data['lowest_level']
        highest_level_list = data['highest_level_values']
        default_value_type = data['default_value_type'] if 'default_value_type' in data else None
        final_output = get_details_by_higher_level(highest_level,lowest_level,highest_level_list, default_value_type)
        return handle_response('', data=final_output, success=True)


class GeographicalLevelsTest(APIView):

    @staticmethod
    def put(request):
        highest_level = request.data['highest_level']
        lowest_level = request.data['lowest_level'] if 'lowest_level' in request.data else 'supplier'
        highest_level_list = request.data['highest_level_list']
        results_by_lowest_level = request.data['results_by_lowest_level'] \
            if 'results_by_lowest_level' in request.data else 0
        result_dict = get_details_by_higher_level_geographical(
            highest_level, highest_level_list, lowest_level, results_by_lowest_level)
        return handle_response('',data={'final_dict':result_dict['final_dict'],
                                        'single_list':result_dict['single_list']},success=True)

class GetLeadsEntry(APIView):
    @staticmethod
    def get(request, form_id, supplier_id, entry_id):
        data = mongo_client.leads.find({'leads_form_id': int(form_id), 'supplier_id': supplier_id, 'entry_id': int(entry_id)})

        lead_form = list(mongo_client.leads_forms.find(
            {"$and": [{"leads_form_id": int(form_id)}, {"status": {"$ne": "inactive"}}]}))
        lead_form_dict = {}
        if len(lead_form) > 0:
            lead_form = lead_form[0]
            lead_form_dict = {
                "leads_form_name": lead_form['leads_form_name'],
                "leads_form_id": lead_form['leads_form_id'],
                "leads_form_items": lead_form['data']
            }
        # leads_form_items_map_by_item_id = {item['item_id']: item for item in lead_form_dict['leads_form_items']}
        lead_entry_map_by_item_id = {item['item_id']:item for item in data[0]['data']}

        for key,value in lead_form_dict['leads_form_items'].items():
            value['value'] = lead_entry_map_by_item_id[value['item_id']]['value']

        return handle_response('', data=lead_form_dict, success=True)


class UpdateLeadsEntry(APIView):
    @staticmethod
    def put(request, form_id, supplier_id, entry_id):
        data = request.data
        lead_dict = list(mongo_client.leads.find({'leads_form_id': int(form_id), 'supplier_id': supplier_id, 'entry_id': int(entry_id)}))
        lead_form = mongo_client.leads_forms.find({"leads_form_id": int(form_id)})[0]
        if len(lead_dict) == 0:
            return handle_response('', data={"error_msg": "lead_not_present"}, success=False)
        lead_dict = lead_dict[0]
        lead_entry_map_by_item_id = {item['item_id']: item for k, item in data.items()}
        for lead_item in lead_dict['data']:
            lead_item['value'] = lead_entry_map_by_item_id[int(lead_item['item_id'])]['value']

        lead_dict["is_hot"], lead_dict["multi_level_is_hot"], lead_dict["hotness_level"] = calculate_is_hot(lead_dict, lead_form['global_hot_lead_criteria'])
        lead_sha_256 = create_lead_hash(lead_dict)
        lead_dict["lead_sha_256"] = lead_sha_256
        lead_already_exist = True if len(list(mongo_client.leads.find({"lead_sha_256": lead_sha_256}))) > 0 else False
        if not lead_already_exist:
            mongo_client.leads.update_one(
                {"leads_form_id": int(form_id), "entry_id": int(entry_id), "supplier_id": supplier_id},
                {"$set": {"data": lead_dict["data"], "lead_sha_256": lead_sha_256, "is_hot": lead_dict["is_hot"],
                          "multi_level_is_hot": lead_dict["multi_level_is_hot"],
                          "hotness_level": lead_dict["hotness_level"]}})
        return handle_response('', data={"success": True}, success=True)

def prepare_campaign_specific_data_in_excel(data, comment_list):
    
    inventory_list = []
    for supplier in data['shortlisted_suppliers']:
        if supplier['shortlisted_inventories']:
            for key,value in supplier['shortlisted_inventories'].items():
                if key not in inventory_list:
                    inventory_list.append(key)

    header_list = [
        'Index', 'Proposal Id', 'Supplier Id', 'Supplier Name', 'Supplier Type' , 'Subarea', 'Area', 'City', 'Address',
        'Landmark', 'PinCode', 'Unit Primary Count / Flat Count', 'Unit Secondary Count / Tower Count', 'Average Household Points',
        'Cost Per Unit', 'Booking Priority', 'Booking Status', 'Booking Sub-status', 'Next Action Date', 'Date of Last Call',
        'Payment Method', 'Payment Status', 'Completion Status', 'Total Price',
        'Internal Comment', 'External Comment', 'Rating', 'Assigned To',
        # 'Poster Allowed', 'Poster Count', 'Poster Price',
        # 'Standee Allowed', 'Standee Count', 'Standee Price',
        # 'Stall Allowed', 'Stall Count', 'Stall Price',
        # 'Flier Allowed', 'Flier Count', 'Flier Price',
        # 'Banner Allowed', 'Banner Count', 'Banner Price',
    ]

    if data["campaign"].get("type_of_end_customer_formatted_name") == "b_to_b_r_g":
        header_list.append("Requirement Given")
        header_list.append("Requirement Given Date")

    for inventory in inventory_list:
        header_list.append(inventory+" Allowed")
        header_list.append(inventory+" Count")
        header_list.append(inventory+" Price")

    book = Workbook()
    sheet = book.active
    sheet.append(header_list)
    index = 0
    for supplier in data['shortlisted_suppliers']:
        index = index + 1
        supplier_data = []

        supplier_data.append(index)

        supplier_data.append(supplier['proposal'])
        supplier_data.append(supplier['object_id'])
        supplier_data.append(supplier['name'])
        supplier_data.append(supplier['supplier_type'])
        supplier_data.append(supplier['subarea'])
        supplier_data.append(supplier['area'])
        supplier_data.append(supplier['city'])
        supplier_data.append(str(supplier['address1']) + ' '+ str(supplier['address2']))

        supplier_data.append(supplier['landmark'])
        supplier_data.append(supplier['zipcode'])

        primary_count = supplier.get('flat_count')
        secondary_count = supplier.get('tower_count')
        if supplier.get('unit_primary_count'):
            primary_count = supplier.get('unit_primary_count')

        if supplier.get('unit_secondary_count'):
            secondary_count = supplier.get('unit_secondary_count')

        supplier_data.append(primary_count)
        supplier_data.append(secondary_count)
        avg_household_occupants = None
        if supplier.get('avg_household_occupants'):
            avg_household_occupants = supplier.get('avg_household_occupants')
        supplier_data.append(avg_household_occupants)

        supplier_data.append(supplier['cost_per_flat'])
        supplier_data.append(booking_priority_code_to_status[supplier['booking_priority']] if supplier['booking_priority'] else None)
        supplier_data.append(booking_code_to_status[supplier['booking_status']] if supplier['booking_status'] else None)
        
        booking_substatus = BookingSubstatus.objects.filter(code=supplier['booking_sub_status']).first()
        supplier_data.append(booking_substatus.name if booking_substatus else None)

        # supplier_data.append(supplier['next_action_date'])
        next_action_date = None
        if supplier['next_action_date']:
            next_action_date = datetime.datetime.strptime(supplier['next_action_date'], '%Y-%m-%dT%H:%M:%SZ').strftime("%d/%m/%Y")
        supplier_data.append(next_action_date)

        last_call_date = None
        if supplier['last_call_date']:
            last_call_date = datetime.datetime.strptime(supplier['last_call_date'], '%Y-%m-%dT%H:%M:%SZ').strftime("%d/%m/%Y")
        supplier_data.append(last_call_date)

        supplier_data.append(supplier['payment_method'])
        supplier_data.append(payment_code_to_status[supplier['payment_status']] if supplier['payment_status'] else None)
        supplier_data.append('Yes' if supplier['is_completed'] else 'No')
        supplier_data.append(supplier['total_negotiated_price'])

        internal_comment = ""
        external_comment = ""
        
        if comment_list["INTERNAL"].get(supplier["id"]):
            internal_comment = "\n".join(comment_list["INTERNAL"][supplier["id"]])
        
        if comment_list["EXTERNAL"].get(supplier["id"]):
            external_comment = "\n".join(comment_list["EXTERNAL"][supplier["id"]])

        supplier_data.append(internal_comment)
        supplier_data.append(external_comment)

        supplier_data.append(supplier["quality_rating"])

        assigned_user = SupplierAssignment.objects.filter(campaign_id=supplier['proposal'], supplier_id=supplier['object_id']).first()
        assigned_to = None
        if assigned_user:
            assigned_to =  assigned_user.assigned_to.username
            supplier_data.append(assigned_to)
        else :
            supplier_data.append(assigned_to)

        if data["campaign"].get("type_of_end_customer_formatted_name") == "b_to_b_r_g":
            supplier_data.append(supplier['requirement_given'])
            requirement_given_date = None
            if supplier['requirement_given_date']:
                requirement_given_date = datetime.datetime.strptime(supplier['requirement_given_date'], '%Y-%m-%dT%H:%M:%S.%fZ').strftime("%d/%m/%Y")

            supplier_data.append(requirement_given_date)

        for row in inventory_list:
            supplier_data.append('Yes' if row in supplier['shortlisted_inventories'] else 'No')
            supplier_data.append(
                supplier['shortlisted_inventories'][row]['total_count'] if row in supplier[
                    'shortlisted_inventories'] else None)
            supplier_data.append(
                supplier['shortlisted_inventories'][row]['actual_supplier_price'] if row in supplier[
                    'shortlisted_inventories'] else None)

        sheet.append(supplier_data)

    return book

class CampaignDataInExcelSheet(APIView):
    permission_classes = ()
    authentication_classes = ()
    @staticmethod
    def get(request, one_time_hash):
        excel_download_hash = list(CampaignExcelDownloadHash.objects.raw({"one_time_hash": one_time_hash}))
        data = {}
        data["shortlisted_suppliers"] = []
        if len(excel_download_hash) > 0:
            campaign_id = excel_download_hash[0].campaign_id
            response = prepare_shortlisted_spaces_and_inventories(campaign_id, None, request.user, 0, None, None, None, space_status='F')
            if response.data['status']:
                data = response.data['data']

        comments = CampaignComments.objects.filter(campaign_id=campaign_id)
        comment_list = {
            "INTERNAL":{},
            "EXTERNAL":{}
        }
        for row in comments:
            if row.related_to in comment_list:
                if not comment_list[row.related_to].get(row.shortlisted_spaces_id):
                    comment_list[row.related_to][row.shortlisted_spaces_id] = []

                comment_list[row.related_to][row.shortlisted_spaces_id].append(row.comment)
        
        if data["campaign"].get("type_of_end_customer_formatted_name") == "b_to_b_l_d":
            excel_book = prepare_campaign_leads_data_in_excel(data, comment_list)
        else:    
            excel_book = prepare_campaign_specific_data_in_excel(data, comment_list)
        resp = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = 'attachment; filename=mydata.xlsx'
        excel_book.save(resp)
        return resp

class GenerateCampaignExcelDownloadHash(APIView):
    @staticmethod
    def get(request, campaign_id):
        excel_download_hash_dict = {"campaign_id": campaign_id}
        now = datetime.datetime.now()
        excel_download_hash_dict["created_at"] = now
        one_time_hash = hashlib.sha256(str(now).encode('utf-8')).hexdigest()
        excel_download_hash_dict["one_time_hash"] = one_time_hash
        CampaignExcelDownloadHash(**excel_download_hash_dict).save()
        return handle_response({}, data={"one_time_hash": one_time_hash}, success=True)


class AddHotnessLevelsToLeadForm(APIView):
    @staticmethod
    def get(request):
        lead_forms = mongo_client.leads_forms.find()
        for lead_form in lead_forms:
            index = len(lead_form['data'].keys())
            index = index + 1

            lead_form['data'][
                str(index)] = {
                    "order_id": None,
                    "key_name": "Is Meeting Fixed",
                    "key_type": "STRING",
                    "item_id": index,
                    "is_required": False
                }

            lead_form["global_hot_lead_criteria"][
                "is_hot_level_2"] = {
                    "or": {
                        str(index) : [
                            "Y",
                            "y",
                            "Yes",
                            "yes",
                            "YES"
                        ]
                    }
                }
            index = index + 1
            lead_form['data'][
                str(index)] = {
                    "order_id": None,
                    "key_name": "Is Meeting Completed",
                    "key_type": "STRING",
                    "item_id": index,
                    "is_required": False
                }
            lead_form["global_hot_lead_criteria"][
            "is_hot_level_3" ] = {
                "or": {
                    str(index): [
                        "Y",
                        "y",
                        "Yes",
                        "yes",
                        "YES"
                    ]
                }
            }
            index = index + 1
            lead_form['data'][
                str(index)] = {
                    "order_id": None,
                    "key_name": "Is Lead Converted",
                    "key_type": "STRING",
                    "item_id": index,
                    "is_required": False
                }
            lead_form["global_hot_lead_criteria"][
            "is_hot_level_4"] = {
                "or": {
                    str(index): [
                        "Y",
                        "y",
                        "Yes",
                        "yes",
                        "YES"
                    ]
                }
            }
            index = index + 1
            lead_form['data'][
                str(index)] = {
                "order_id": None,
                "key_name": "Order Punched Date",
                "key_type": "DATE",
                "item_id": index,
                "is_required": False
            }
            mongo_client.leads_forms.update_one({"leads_form_id": lead_form['leads_form_id']},
                                            {"$set": {'data': lead_form['data'],
                                                    'global_hot_lead_criteria': lead_form['global_hot_lead_criteria']}})
        leads = mongo_client.leads.find({})
        fields_array = ["Is Meeting Fixed", "Is Meeting Completed", "Is Lead Converted"]

        bulk = mongo_client.leads.initialize_unordered_bulk_op()
        counter = 0
        for lead in leads:
            count = len(lead['data'])
            for field in fields_array:
                count = count + 1
                lead['data'].append({
                    "key_name": field,
                    "key_type": "STRING",
                    "item_id": count,
                    "value": None
                })
            count = count + 1
            lead['data'].append({
                "key_name": "Order Punched Date",
                "key_type": "DATE",
                "item_id": count,
                "value": None
            })

            lead_sha_256 = create_lead_hash(lead)
            lead["lead_sha_256"] = lead_sha_256

            bulk.find({"_id": ObjectId(lead['_id'])}).update({"$set":
                                                      {"data": lead["data"], "lead_sha_256": lead_sha_256}})
            counter += 1
            if counter % 1000 == 0:
                bulk.execute()
                bulk = mongo_client.leads.initialize_unordered_bulk_op()
        if counter > 0:
            bulk.execute()

                # mongo_client.leads.update_one(
                #     {"leads_form_id": int(lead['leads_form_id']), "entry_id": int(lead['entry_id']),
                #      "supplier_id": lead['supplier_id']},
                #     {"$set": {"data": lead["data"], "lead_sha_256": lead_sha_256}})

        return handle_response('', data={"success": True}, success=True)

class UpdateConvertedLeadsFromSheet(APIView):
    @staticmethod
    def post(request):
        source_file = request.data['file']
        wb = load_workbook(source_file)
        ws = wb.get_sheet_by_name(wb.get_sheet_names()[1])
        supplier_names_list = set()
        campaign_names_list = set()
        supplier_campaign_name_map = {}
        for index, row in enumerate(ws.iter_rows()):
            if index == 0:
                continue
            if row[4].value:
                supplier_names_list.add(row[4].value.lower())
                campaign_names_list.add(row[8].value.lower())
                supplier_campaign_name_map[row[4].value.lower()] = row[8].value.lower()
        suppliers = SupplierTypeSociety.objects.filter(society_name__in=list(supplier_names_list))

        campaigns = ProposalInfo.objects.filter(name__in=list(campaign_names_list))
        suppliers_map_with_id = { supplier.society_name.lower(): supplier.supplier_id for supplier in suppliers }
        suppliers_map_with_name = {supplier.supplier_id: supplier.society_name.lower() for supplier in suppliers}
        campaigns_map_with_id = { campaign.name.lower(): campaign.proposal_id for campaign in campaigns }

        ws = wb.get_sheet_by_name(wb.get_sheet_names()[4])
        purchase_order_objects = []
        suppliers_in_purchase_order = set()
        for index, row in enumerate(ws.iter_rows()):
            if index == 0:
                continue
            purchase_order_objects.append({
                "name": row[0].value.lower(),
                "supplier_name": row[1].value.lower(),
                "date": row[2].value
            })
            suppliers_in_purchase_order.add(row[1].value.lower())

        supplier_leads = {}
        mismatch_suppliers_in_punched_order = []
        for supplier in suppliers_in_purchase_order:
            supplier_leads[supplier] = {}
            if supplier in suppliers_map_with_id:
                supplier_leads[supplier]['leads'] = mongo_client.leads.find({'supplier_id': suppliers_map_with_id[supplier],
                                                         'campaign_id': campaigns_map_with_id[supplier_campaign_name_map[supplier]]})
                if supplier_leads[supplier]['leads'].count() > 0:
                    supplier_leads[supplier]['name_list'] = []
                    supplier_leads[supplier]['names_objects_map'] = {}
                    for lead in supplier_leads[supplier]['leads']:
                        if 'data' in lead:
                            name = lead['data'][0]['value'].lower()
                            if name:
                                supplier_leads[supplier]['name_list'].append(name)
                                supplier_leads[supplier]['names_objects_map'][name] = lead
            else:
                mismatch_suppliers_in_punched_order.append(supplier)
        for item in purchase_order_objects:
            if 'name_list' in supplier_leads[item['supplier_name']]:
                matched_name = difflib.get_close_matches(item['name'], supplier_leads[item['supplier_name']]['name_list'])
                if len(matched_name) > 0:
                    name = matched_name[0]
                    lead = supplier_leads[item['supplier_name']]['names_objects_map'][name]
                    count = len(lead['data'])
                    if lead['data'][count-1]['key_name'] == 'Order Punched Date':
                        lead['data'][count-1]['value'] = item['date']
                    if lead['data'][count-2]['key_name'] == 'Is Lead Converted':
                        lead['data'][count-2]['value'] = 'Yes'
                    lead['multi_level_is_hot']['is_hot_level_2'] = True
                    lead['multi_level_is_hot']['is_hot_level_3'] = True
                    lead['multi_level_is_hot']['is_hot_level_4'] = True

                    lead_sha_256 = create_lead_hash(lead)
                    lead["lead_sha_256"] = lead_sha_256

                    mongo_client.leads.update_one(
                        {"leads_form_id": int(lead['leads_form_id']), "entry_id": int(lead['entry_id']), "supplier_id": lead['supplier_id']},
                        {"$set": {"data": lead["data"], "lead_sha_256": lead_sha_256,
                                  "hotness_level": 4, "multi_level_is_hot": lead['multi_level_is_hot']}})
        # To update leads summary table
        ws = wb.get_sheet_by_name(wb.get_sheet_names()[1])
        for index, row in enumerate(ws.iter_rows()):
            if index == 0:
                continue
            if row[4].value:
                lead = list(mongo_client.leads_summary.find({"supplier_id": suppliers_map_with_id[row[4].value.lower()],
                                                 "campaign_id": campaigns_map_with_id[supplier_campaign_name_map[row[4].value.lower()]]}))

                if len(lead) > 0:
                    total_booking_confirmed = row[12].value if row[12].value else None
                    total_orders_punched = row[13].value if row[13].value else None

                    mongo_client.leads_summary.update({"_id": ObjectId(lead[0]['_id'])},
                                                      {"$set": {'total_booking_confirmed': total_booking_confirmed,
                                                                'total_orders_punched': total_orders_punched}})
        data = {
            "mismatch_suppliers_in_punched_order": mismatch_suppliers_in_punched_order,
        }
        return handle_response('', data=data, success=True)

class UpdateLeadSummary(APIView):
    @staticmethod
    def get(request):
        campaign_list = ProposalInfo.objects.filter().values_list('proposal_id', flat=True)
        data = get_leads_summary(list(campaign_list))
        for item in data:
            mongo_client.leads_summary.insert_one({
                "campaign_id": item['campaign_id'],
                "supplier_id": item['supplier_id'],
                "lead_date": item['created_at'] if 'created_at' in item else None,
                "total_leads_count": item['total_leads_count'],
                "total_hot_leads_count": item['hot_leads_count'],
                "total_booking_confirmed": 0,
                "total_orders_punched": 0
            })
        return handle_response('', data={"success": True}, success=True)

class UpdateOrderId(APIView):
    @staticmethod
    def get(request):
        lead_forms = mongo_client.leads_forms.find({})
        for lead_form in lead_forms:
            if len(lead_form['data']) > 0:
                for item in lead_form['data'].values():
                    if not item['order_id']:
                        item['order_id'] = item['item_id']
                mongo_client.leads_forms.update({"_id": ObjectId(lead_form['_id'])},
                                                  {"$set": {'data': lead_form['data'] }})


        return handle_response('', data={"success": True}, success=True)

class MigrateLeadsFormsByAlias(APIView):
    @staticmethod
    def get(request):
        leads_forms = mongo_client.leads_forms.find({})
        alias_dict = {
            'is_hot_level_1': "Hot Leads",
            'is_hot_level_2': "Is Meeting Fixed",
            'is_hot_level_3': "Total Bookings Confirmed",
            'is_hot_level_4': "Total Orders Punched",
        }
        bulk = mongo_client.leads_forms.initialize_unordered_bulk_op()
        counter = 0
        for lead_form in leads_forms:
            data = {}
            if 'global_hot_lead_criteria' in lead_form:
                for key in lead_form['global_hot_lead_criteria']:
                    print(key)
                    if key in alias_dict:
                        data[key] = alias_dict[key]
                bulk.find({"id": ObjectId(lead_form['_id'])}).update({"$set": {"hotness_mapping": data}})
                counter += 1
                if counter % 500 == 0:
                    bulk.execute()
                    bulk = mongo_client.leads.initialize_unordered_bulk_op()
        if counter > 0:
            bulk.execute()
        return handle_response('', data={"success": True}, success=True)

class LeadsKeys(APIView):
    def put(self, request):
        leads_forms = mongo_client.leads_forms.find({})
        campaign_ids = request.data.get("campaign_ids")
        forms = list(mongo_client.leads_forms.find({"campaign_id":{"$in":campaign_ids}}))

        hotness_mapping = {}

        for row in forms:
            if row.get("hotness_mapping"):
                for key, value in row['hotness_mapping'].items():
                    hotness_mapping[key] = True
        
        return handle_response('', data=hotness_mapping, success=True)

def prepare_campaign_leads_data_in_excel(data, comment_list):
    
    header_list = [
        'Index', 'Supplier Name', 'Supplier Type' , 'Current Partner', 'FeedBack', 'Preferred Partner', 'Preferred Partner Other',
        'L1 Answer 1 ', 'L1 Answer 1', 'L2 Answer 1', 'L2 Answer 2', 'Implementation Time', 'Meeting Time', 'Lead Given by', 'Subarea',
        'Area', 'City', 'Address', 'Landmark', 'PinCode', 'Unit Primary Count / Flat Count',
        'Unit Secondary Count / Tower Count', 'Average Household Points',
    ]

    book = Workbook()
    sheet = book.active
    sheet.append(header_list)
    index = 0
    for supplier in data['shortlisted_suppliers']:
        index = index + 1
        supplier_data = []

        supplier_data.append(index)

        supplier_data.append(supplier['name'])
        supplier_data.append(supplier['supplier_type'])

        requirement = Requirement.objects.filter(company_campaign_id=supplier['proposal'], company_shortlisted_spaces_id=supplier['id']).first()
        if requirement:
            supplier_data.append(requirement.current_company.name if requirement.current_company else None)
            supplier_data.append(requirement.current_patner_feedback)
            preferred_company = None
            if requirement.preferred_company:
                preferred_company = ", ".join(requirement.preferred_company)
            supplier_data.append(preferred_company)
            supplier_data.append(requirement.preferred_company_other)
            supplier_data.append(requirement.l1_answers)
            supplier_data.append(requirement.l1_answer_2)
            supplier_data.append(requirement.l2_answers)
            supplier_data.append(requirement.l2_answer_2)
            supplier_data.append(requirement.impl_timeline)
            supplier_data.append(requirement.meating_timeline)
            supplier_data.append(requirement.lead_by.name)

        supplier_data.append(supplier['subarea'])
        supplier_data.append(supplier['area'])
        supplier_data.append(supplier['city'])
        supplier_data.append(str(supplier['address1']) + ' '+ str(supplier['address2']))

        supplier_data.append(supplier['landmark'])
        supplier_data.append(supplier['zipcode'])

        primary_count = supplier.get('flat_count')
        secondary_count = supplier.get('tower_count')
        if supplier.get('unit_primary_count'):
            primary_count = supplier.get('unit_primary_count')

        if supplier.get('unit_secondary_count'):
            secondary_count = supplier.get('unit_secondary_count')

        supplier_data.append(primary_count)
        supplier_data.append(secondary_count)
        avg_household_occupants = None
        if supplier.get('avg_household_occupants'):
            avg_household_occupants = supplier.get('avg_household_occupants')
        supplier_data.append(avg_household_occupants)

        sheet.append(supplier_data)

    return book